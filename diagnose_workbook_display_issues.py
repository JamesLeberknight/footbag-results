#!/usr/bin/env python3
"""
diagnose_workbook_display_issues.py

Scans Footbag_Results_Community_FINAL_v13.xlsx for visible display problems.

Focus:
- lowercase-first-name issues
- location-prefixed names (e.g. 'Urbana Jeff Cruz')
- suspicious result lines
- bad division labels
- year-only Date metadata in year sheets

Edit WORKBOOK below if needed, then run:
    python diagnose_workbook_display_issues.py
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path.home() / "projects" / "FOOTBAG_DATA" / "Footbag_Results_Community_FINAL_v13.xlsx"

YEAR_SHEET_RE = re.compile(r"^\d{4}$")

# Common location tokens seen leaking into player display
LOCATION_PREFIXES = {
    "urbana", "sherman", "chicago", "milan", "charleston", "selma",
    "wichita", "austin", "hebron", "mcpherson", "vancouver", "pittsburgh",
    "portland", "golden", "memphis", "oregon", "san", "boulder"
}

DIVISION_TYPO_PATTERNS = [
    re.compile(r"\bIntrmediate\b", re.I),
    re.compile(r"\bWomens\b"),
]

LOWERCASE_FIRSTNAME_RE = re.compile(r"\b[a-z][a-z]+ [A-Z][a-z]+(?:[-'][A-Z][a-z]+)?\b")
YEAR_ONLY_RE = re.compile(r"^\d{4}$")

# result rows often start with medal emoji or numeric place
RESULT_LINE_RE = re.compile(r"^(🥇|🥈|🥉)?\s*\d+\s+(.+)$")

# simple team split
TEAM_SPLIT_RE = re.compile(r"\s/\s")

def is_year_sheet(name: str) -> bool:
    return bool(YEAR_SHEET_RE.fullmatch(str(name).strip()))

def extract_result_text(cell_text: str) -> str | None:
    m = RESULT_LINE_RE.match(cell_text.strip())
    if not m:
        return None
    return m.group(2).strip()

def looks_like_location_prefixed_name(name: str) -> bool:
    """
    Flags patterns like:
      Urbana Jeff Cruz
      Sherman Josh Vorvel
    Conservative heuristic: first token is known location prefix,
    remaining text looks like 2+ capitalized name tokens.
    """
    toks = name.strip().split()
    if len(toks) < 3:
        return False
    if toks[0].lower() not in LOCATION_PREFIXES:
        return False
    # remaining tokens look like name parts
    remain = toks[1:]
    caps = sum(1 for t in remain if re.fullmatch(r"[A-Z][a-z]+(?:[-'][A-Z][a-z]+)?", t))
    return caps >= 2

def analyze_workbook(path: Path):
    wb = load_workbook(path, data_only=False)

    year_only_dates = []
    division_typos = []
    lowercase_names = []
    location_prefixed = []
    suspicious_team_lines = []
    same_place_duplicates = []

    # Track result texts by sheet to find conflicting variants nearby
    result_name_index = defaultdict(list)

    for ws in wb.worksheets:
        sheet = ws.title

        # Date row on year sheets is usually row 4 in your layout
        if is_year_sheet(sheet):
            for cell in ws[4]:
                v = cell.value
                if isinstance(v, str) and YEAR_ONLY_RE.fullmatch(v.strip()):
                    year_only_dates.append((sheet, cell.coordinate, v.strip()))

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                text = v.strip()
                if not text:
                    continue

                # division typos
                for pat in DIVISION_TYPO_PATTERNS:
                    if pat.search(text):
                        division_typos.append((sheet, cell.coordinate, text))
                        break

                # lowercase-first-name anywhere
                if LOWERCASE_FIRSTNAME_RE.search(text):
                    lowercase_names.append((sheet, cell.coordinate, text))

                # result-line-specific diagnostics
                result_text = extract_result_text(text)
                if result_text:
                    result_name_index[sheet].append((cell.coordinate, result_text))

                    # team separator oddities
                    if " + " in result_text:
                        suspicious_team_lines.append((sheet, cell.coordinate, text, "plus_separator"))

                    # location-prefixed player names
                    if TEAM_SPLIT_RE.search(result_text):
                        parts = TEAM_SPLIT_RE.split(result_text)
                        for p in parts:
                            if looks_like_location_prefixed_name(p):
                                location_prefixed.append((sheet, cell.coordinate, text, p))
                    else:
                        if looks_like_location_prefixed_name(result_text):
                            location_prefixed.append((sheet, cell.coordinate, text, result_text))

        # duplicate-ish display rows in same sheet:
        # e.g. Jeff Cruz and Urbana Jeff Cruz both appearing
        names = [t for _, t in result_name_index[sheet]]
        for i, a in enumerate(names):
            for b in names[i + 1:]:
                if a == b:
                    continue
                # one string contains the other with one extra leading token
                a_toks = a.split()
                b_toks = b.split()
                if len(a_toks) + 1 == len(b_toks) and b.endswith(a):
                    same_place_duplicates.append((sheet, a, b))
                elif len(b_toks) + 1 == len(a_toks) and a.endswith(b):
                    same_place_duplicates.append((sheet, b, a))

    return {
        "year_only_dates": year_only_dates,
        "division_typos": division_typos,
        "lowercase_names": lowercase_names,
        "location_prefixed": location_prefixed,
        "suspicious_team_lines": suspicious_team_lines,
        "same_place_duplicates": same_place_duplicates,
    }

def print_section(title: str, rows, limit: int = 50):
    print(f"\n{title}")
    print("-" * len(title))
    print(f"Count: {len(rows)}")
    for item in rows[:limit]:
        print(item)
    if len(rows) > limit:
        print(f"... and {len(rows) - limit} more")

def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    findings = analyze_workbook(WORKBOOK)

    print(f"Workbook: {WORKBOOK}")

    print_section("YEAR-ONLY DATE CELLS", findings["year_only_dates"])
    print_section("DIVISION TYPO HITS", findings["division_typos"])
    print_section("LOWERCASE NAME HITS", findings["lowercase_names"])
    print_section("LOCATION-PREFIXED NAME HITS", findings["location_prefixed"])
    print_section("SUSPICIOUS TEAM-LINE HITS", findings["suspicious_team_lines"])
    print_section("POSSIBLE DISPLAY DUPLICATE / PREFIX PAIRS", findings["same_place_duplicates"])

if __name__ == "__main__":
    main()
