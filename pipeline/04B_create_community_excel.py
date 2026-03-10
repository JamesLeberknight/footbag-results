#!/usr/bin/env python3
"""
04B_create_community_excel.py

Produces Footbag_Results_Community.xlsx — a reader-friendly Footbag Historical
Almanac targeted at the footbag community.

Read-only transformation only.  No identity changes, no canonical mutations.

Inputs  (from out/):
    stage2_canonical_events.csv  — event metadata + division source order
    Placements_Flat.csv          — identity-resolved placements
    Placements_ByPerson.csv      — for leaderboard computation
    Persons_Truth.csv            — for honours matching

Inputs  (from inputs/):
    bap_data.csv                 — Big Add Posse inductees
    fbhof_data.csv               — Footbag Hall of Fame inductees

Output:
    Footbag_Results_Community.xlsx
"""

import csv
import json
import re
import sys
import unicodedata
from collections import defaultdict, OrderedDict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

csv.field_size_limit(10_000_000)

REPO      = Path(__file__).resolve().parent.parent
OUT_DIR   = REPO / "out"
INPUT_DIR = REPO / "inputs"
XLSX      = REPO / "Footbag_Results_Community.xlsx"
KNOWN_ISSUES_CSV = REPO / "overrides" / "known_issues.csv"


# ── Date sorting ──────────────────────────────────────────────────────────────

_MONTH_NUM = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

def _date_sort_key(date_str: str, eid: str) -> tuple:
    """Return (month, day, event_id) for sorting within a year.

    Parses text dates like 'September 29, 2001' or 'November 9-11, 2001'.
    Events with missing or unparseable dates sort last (month=13).
    """
    m = re.match(r"([A-Za-z]+)\s+(\d+)", (date_str or "").strip())
    if m:
        month = _MONTH_NUM.get(m.group(1).lower(), 13)
        day   = int(m.group(2))
    else:
        month, day = 13, 0
    return (month, day, eid)


# ── Palette & styles ──────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

def _font(**kw) -> Font:
    return Font(**kw)

def _border_top(color="BBBBBB") -> Border:
    return Border(top=Side(style="thin", color=color))

def _border_bottom(color="BBBBBB") -> Border:
    return Border(bottom=Side(style="thin", color=color))

FILL_BANNER   = _fill("1F4E79")   # dark navy  — event banner
FILL_BANNER_WORLDS = _fill("7D4607")  # deep amber — Worlds event banner
FILL_META     = _fill("EBF3FB")   # pale blue  — location / host / date
FILL_PLAYERS  = _fill("F5F5F5")   # near-white — players count
FILL_DIV      = _fill("E2E2E2")   # light grey — division header
FILL_GOLD     = _fill("FFF3CC")   # soft gold  — 1st place
FILL_SILVER   = _fill("F0F0F0")   # near-white — 2nd place
FILL_BRONZE   = _fill("FDEBD0")   # pale orange— 3rd place
FILL_WHITE    = _fill("FFFFFF")
FILL_ALT      = _fill("F7FAFD")   # very pale blue — alternate data rows
FILL_HDR      = _fill("1F4E79")   # sheet header row
FILL_HON_BAP  = _fill("FFF8E1")   # BAP honour row tint
FILL_HON_FBHOF= _fill("E8F5E9")   # FBHOF honour row tint
FILL_HON_BOTH = _fill("F3E5F5")   # both honours row tint
FILL_ROW_LABEL= _fill("F0F4F8")   # column-A row labels in year sheets

FONT_BANNER   = Font(bold=True,   size=12, color="FFFFFF")
FONT_META     = Font(             size=9,  color="1F4E79")
FONT_HOST     = Font(italic=True, size=9,  color="444444")
FONT_PLAYERS  = Font(             size=9,  color="888888")
FONT_DIV      = Font(bold=True,   size=9)
FONT_PODIUM   = Font(bold=True,   size=9)
FONT_PLACE    = Font(             size=9)
FONT_TITLE    = Font(bold=True,   size=16)
FONT_SECTION  = Font(bold=True,   size=12)
FONT_SUBHEAD  = Font(bold=True,   size=10)
FONT_NORMAL   = Font(             size=10)
FONT_SMALL    = Font(             size=9,  color="555555")
FONT_HDR      = Font(bold=True,   size=10, color="FFFFFF")
FONT_LINK     = Font(             size=10, color="0563C1", underline="single")
FONT_ITALIC   = Font(italic=True, size=10)
FONT_ROW_LBL  = Font(             size=8,  color="888888")

ALIGN_WRAP    = Alignment(wrap_text=True, vertical="top")
ALIGN_TOP     = Alignment(vertical="top")
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT   = Alignment(horizontal="right",  vertical="top")

MEDALS = {1: "🥇", 2: "🥈", 3: "🥉"}

COL_W_MIN     = 24   # minimum column width for event columns
COL_W_LABEL   = 12   # column A (row-label) width in year sheets

# Honour symbols
SYM_BAP   = "★"
SYM_FBHOF = "☆"

# Known name-variant pairs between BAP and FBHOF files
_BAP_TO_FBHOF_ALIASES = {
    "Ken Shults":      "Kenny Shults",
    "Vasek Klouda":    "Václav Klouda",
    "Sebastien Lefay": "Sébastien Lefay",
}
_FBHOF_TO_BAP_ALIASES = {v: k for k, v in _BAP_TO_FBHOF_ALIASES.items()}

# Honours display names → person_canon in Persons_Truth
# (for cases where normalisation alone doesn't find the match)
_HONOURS_TO_PT = {
    # BAP file variants (name as it appears in BAP file → person_canon in Persons_Truth)
    "Sebastien Duschesne":      "Sebastien Duchesne",
    "Arek Dzudzinski":          "Arek Dudzinski",
    "Rene Ruhr":                "Rene Ruehr",
    "Nick Polini":              "Nick Pollini",
    "Ken Shults":               "Kenneth Shults",
    "Kenny Shults":             "Kenneth Shults",
    "Dave Holton":              "David Holton",
    "Gordon Scott Bevier":      "Scott Bevier",
    "Bryan Fournier":           "Brian Fournier",
    "Johnny Murphy":            "Jonathan Murphy",
    "Phillip Morrison":         "Philip Morrison",
    "Tina Aberli":              "Tina Aeberli",
    "Vasek Klouda":             "Vaclav Klouda",
    "Vaclav (Vasek) Klouda":    "Vaclav Klouda",
    "Václav Klouda":            "Vaclav Klouda",
    "Olav Piwowar":             "Olaf Piwowar",
    "Jindra Smola":             "Jindrich Smola",
    "Rafał Kaleta":             "Rafal Kaleta",
    "Jani Markkanen":           "Jani Sakari Markkanen",
    "Jakob Wagner Revstein":    "Jakob Wagner",
    # FBHOF file variants
    "Becca English":            "Becca English Ross",
    "Chris Siebert":            "Christopher Michael Siebert",
    "David Leberknight":        "Dave Leberknight",
    "Jim Caveney":              "Jimmy Caveney",
    "Lon Skyler Smith":         "Skyler Lon Smith",
    "Maude Landreville":        "Maude Laudreville",
    "PT Lovern":                "Paul Lovern",
    "Sam Conlon":               "Samantha Conlon",
    "Vasek Klouda":             "Vaclav Klouda",
}


# ── Name helpers ──────────────────────────────────────────────────────────────

_TRANSLIT = str.maketrans("łŁøØðÐđĐ", "lloodddd")

def _display_name(s: str) -> str:
    """Return s with title-casing applied if stored in ALL-CAPS (Latin American event data)."""
    s = (s or "").strip()
    if not s:
        return s
    alpha = [c for c in s if c.isalpha()]
    if alpha and all(c.isupper() for c in alpha):
        return s.title()
    return s


def _norm_name(s: str) -> str:
    """Lowercase + strip diacritics for fuzzy matching.
    Handles Polish ł, Norwegian ø, Icelandic ð, etc. that don't NFD-decompose."""
    s = s.translate(_TRANSLIT)
    nfd = unicodedata.normalize("NFD", s.lower().strip())
    return "".join(c for c in nfd if unicodedata.category(c) != "Mn")


def _to_int(v) -> int:
    try:
        return int(float(v or 0))
    except (ValueError, TypeError):
        return 0


def _split_location(loc: str):
    """'City, State, Country' → (city_region, country)"""
    parts = [p.strip() for p in loc.split(",")]
    if len(parts) >= 2:
        return ", ".join(parts[:-1]), parts[-1]
    return loc, ""


# ── Data loading ──────────────────────────────────────────────────────────────

def _load_event_field_overrides() -> dict:
    """Load location/host_club overrides from events_overrides.jsonl."""
    path = REPO / "overrides" / "events_overrides.jsonl"
    overrides: dict[str, dict] = {}
    if not path.exists():
        return overrides
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                continue
            eid = str(obj.get("event_id", "")).strip()
            if not eid:
                continue
            entry = {}
            if "location" in obj:
                entry["location"] = obj["location"]
            if "host_club" in obj:
                entry["host_club"] = obj["host_club"]
            if "event_type" in obj:
                entry["event_type"] = obj["event_type"]
            if entry:
                overrides.setdefault(eid, {}).update(entry)
    return overrides


def load_stage2_events() -> dict:
    """
    Load stage2_canonical_events.csv.
    Returns dict event_id → {year, event_name, date, location, host_club,
                              div_order: [division_canon ...]}
    """
    path = OUT_DIR / "stage2_canonical_events.csv"
    field_overrides = _load_event_field_overrides()
    events = {}
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.DictReader(fh):
            eid = row["event_id"].strip()
            try:
                placements = json.loads(row.get("placements_json") or "[]")
            except (json.JSONDecodeError, TypeError):
                placements = []

            seen, div_order = set(), []
            for p in placements:
                dc = (p.get("division_canon") or "").strip()
                if dc and dc not in seen:
                    div_order.append(dc)
                    seen.add(dc)

            fo = field_overrides.get(eid, {})
            loc = fo.get("location") or (row.get("location") or "").strip()
            city, country = _split_location(loc)
            events[eid] = {
                "event_id":   eid,
                "year":       _to_int(row.get("year")),
                "event_name": (row.get("event_name") or "").strip(),
                "date":       (row.get("date") or "").strip(),
                "location":   loc,
                "city":       city,
                "country":    country,
                "host_club":  fo.get("host_club") or (row.get("host_club") or "").strip(),
                "event_type": fo.get("event_type") or (row.get("event_type") or "").strip(),
                "div_order":  div_order,
            }
    return events


def load_placements_flat() -> pd.DataFrame:
    return pd.read_csv(
        OUT_DIR / "Placements_Flat.csv", dtype=str, encoding="utf-8",
    ).fillna("")


def load_placements_by_person() -> pd.DataFrame:
    return pd.read_csv(
        OUT_DIR / "Placements_ByPerson.csv", dtype=str, encoding="utf-8",
    ).fillna("")


def load_persons_truth() -> pd.DataFrame:
    df = pd.read_csv(
        OUT_DIR / "Persons_Truth.csv", dtype=str, encoding="utf-8",
    ).fillna("")
    internal = {"effective_person_id", "player_ids_seen", "player_names_seen",
                "alias_statuses", "norm_key", "last_token",
                "person_canon_clean", "person_canon_clean_reason"}
    return df.drop(columns=[c for c in df.columns if c in internal])


def load_honours(pt_df: pd.DataFrame) -> dict:
    """
    Load BAP and FBHOF CSVs, unify name variants, match against person_canon.

    Returns dict:
        person_canon → {
            'bap':        bool,
            'fbhof':      bool,
            'nickname':   str,
            'bap_year':   int,
            'fbhof_year': int,
            'symbol':     str,   # e.g. '★', '☆', '★☆'
        }

    Also returns (bap_rows, fbhof_rows) as raw lists for the Honours sheet.
    """
    # Build normalised-name → person_canon reverse lookup
    canon_list = pt_df["person_canon"].dropna().tolist() if not pt_df.empty else []
    norm_to_canon: dict = {}
    for pc in canon_list:
        norm_to_canon[_norm_name(pc)] = pc

    def _match_canon(name: str) -> str:
        """Return matched person_canon or '' if not found."""
        import re as _re
        # Hard-coded alias table
        if name in _HONOURS_TO_PT:
            return _HONOURS_TO_PT[name]
        # Exact
        if name in norm_to_canon.values():
            return name
        # Normalised
        n = _norm_name(name)
        if n in norm_to_canon:
            return norm_to_canon[n]
        # Strip parenthetical variants: "Vaclav (Vasek) Klouda" → "Vaclav Klouda"
        stripped = _re.sub(r"\s*\([^)]*\)", "", name).strip()
        if stripped != name:
            if stripped in _HONOURS_TO_PT:
                return _HONOURS_TO_PT[stripped]
            n2 = _norm_name(stripped)
            if n2 in norm_to_canon:
                return norm_to_canon[n2]
        return ""

    honours: dict = {}   # person_canon → info dict

    def _ensure(canon: str) -> dict:
        if canon not in honours:
            honours[canon] = {
                "bap": False, "fbhof": False,
                "nickname": "", "bap_year": 0, "fbhof_year": 0, "symbol": "",
            }
        return honours[canon]

    bap_rows  = []
    fbhof_rows = []

    # ── Load BAP ──────────────────────────────────────────────────────────────
    try:
        bap_df = pd.read_csv(INPUT_DIR / "bap_data_updated.csv", dtype=str).fillna("")
        for _, row in bap_df.iterrows():
            name = row["name"].strip()
            if not name:
                continue
            nick  = row.get("nickname", "").strip()
            year  = _to_int(row.get("year_inducted", ""))
            canon = _match_canon(name)
            bap_rows.append({"name": name, "nickname": nick, "year": year, "canon": canon})
            if canon:
                h = _ensure(canon)
                h["bap"]      = True
                h["nickname"] = h["nickname"] or nick
                h["bap_year"] = year
    except Exception as exc:
        print(f"  WARN: could not load bap_data.csv: {exc}", file=sys.stderr)

    # ── Load FBHOF ────────────────────────────────────────────────────────────
    try:
        fbhof_df = pd.read_csv(INPUT_DIR / "fbhof_data_updated.csv", dtype=str).fillna("")
        for _, row in fbhof_df.iterrows():
            name = row["name"].strip()
            if not name:
                continue
            raw_year = row.get("year_inducted", "")
            year = 0 if str(raw_year).strip().lower() == "unknown" else _to_int(raw_year)
            # Map FBHOF name to BAP canonical name if known
            bap_name = _FBHOF_TO_BAP_ALIASES.get(name, name)
            canon    = _match_canon(bap_name) or _match_canon(name)
            fbhof_rows.append({"name": name, "year": year, "canon": canon})
            if canon:
                h = _ensure(canon)
                h["fbhof"]      = True
                h["fbhof_year"] = year
    except Exception as exc:
        print(f"  WARN: could not load fbhof_data.csv: {exc}", file=sys.stderr)

    # Compute symbol strings
    for h in honours.values():
        h["symbol"] = (SYM_BAP if h["bap"] else "") + (SYM_FBHOF if h["fbhof"] else "")

    matched   = sum(1 for r in bap_rows if r["canon"])
    unmatched = [r["name"] for r in bap_rows if not r["canon"]]
    print(f"  Honours: {matched}/{len(bap_rows)} BAP names matched to person_canon")
    if unmatched:
        print(f"  Honours unmatched: {unmatched}", file=sys.stderr)

    return honours, bap_rows, fbhof_rows


# ── Placement data for year sheets ────────────────────────────────────────────

_FLAG_RANK = {"complete": 0, "mostly_complete": 1, "partial": 2, "sparse": 3}

def compute_event_coverage(pf: pd.DataFrame) -> dict:
    """Return dict event_id → worst coverage_flag string for that event."""
    result = {}
    for eid, grp in pf.groupby("event_id"):
        worst = max(grp["coverage_flag"], key=lambda f: _FLAG_RANK.get(f.lower(), 0))
        result[str(eid)] = worst.lower()
    return result


def build_event_placements(pf: pd.DataFrame, events: dict) -> dict:
    """
    Returns dict: event_id → OrderedDict{division_canon: [(place_int, display, cat)]}.
    Divisions are in source order (from events[eid]['div_order']).
    Doubles are deduplicated by team_person_key; team_display_name used.
    __NON_PERSON__ and unresolved rows are excluded.
    """
    result = {}

    for eid, edf in pf.groupby("event_id"):
        if eid not in events:
            continue

        div_order     = events[eid]["div_order"]
        div_placements: dict = {}

        for div_canon, ddf in edf.groupby("division_canon"):
            div_canon = div_canon.rstrip(":").strip()
            ddf = ddf.copy()
            ddf["_place"] = pd.to_numeric(ddf["place"], errors="coerce")
            ddf = ddf.sort_values(["_place", "team_person_key", "person_canon"],
                                  na_position="last")

            entries    = []
            seen_teams: set = set()
            # Determine if this is a true team division using only VISIBLE rows
            # (exclude __NON_PERSON__ / unresolved which may be mis-parsed teams).
            _visible = ddf[
                ~ddf["person_canon"].str.strip().isin(["", "__NON_PERSON__"]) &
                ~ddf["person_unresolved"].str.lower().isin(("true", "1"))
            ]
            is_team_division = (
                not _visible.empty and
                (_visible["competitor_type"].str.lower() == "team").mean() > 0.5
            )

            for _, row in ddf.iterrows():
                person       = (row.get("person_canon") or "").strip()
                comp         = (row.get("competitor_type") or "player").lower()
                tpk          = (row.get("team_person_key") or "").strip()
                cat          = (row.get("division_category") or "").strip()
                team_display = (row.get("team_display_name") or "").strip()

                if not person or person == "__NON_PERSON__":
                    # Allow team entries whose display name is fully populated
                    # (both partners named, e.g. 1997 Worlds where neither player
                    # reached PT threshold). Exclude "solo-in-doubles" artifacts
                    # that end with "/ ?" (one partner genuinely unknown).
                    if not (comp == "team" and tpk and team_display
                            and not team_display.rstrip().endswith("/ ?")):
                        continue
                if (row.get("person_unresolved") or "").lower() in ("true", "1"):
                    continue

                try:
                    place_int = int(float(row["place"]))
                except (ValueError, TypeError):
                    continue

                if comp == "team" and tpk:
                    if tpk in seen_teams:
                        continue
                    seen_teams.add(tpk)
                    display = team_display
                    if not display:
                        members = ddf[ddf["team_person_key"] == tpk]["person_canon"].tolist()
                        display = " / ".join(_display_name(m) for m in members if m)
                elif comp == "team" and team_display:
                    # team_person_key missing (built before partners reached PT threshold);
                    # use stored team_display_name directly; dedup on display string.
                    if team_display in seen_teams:
                        continue
                    seen_teams.add(team_display)
                    display = team_display
                else:
                    display = _display_name(person)
                    # Solo entry in a true doubles division — partner not recorded
                    if is_team_division:
                        display = f"{display} / ?"

                entries.append((place_int, display, cat))

            if entries:
                div_placements[div_canon] = entries

        # Reorder: source order first, then any unseen divisions
        ordered: OrderedDict = OrderedDict()
        for dc in div_order:
            if dc in div_placements:
                ordered[dc] = div_placements[dc]
        for dc, dp in div_placements.items():
            if dc not in ordered:
                ordered[dc] = dp

        result[eid] = ordered

    return result


# ── Leaderboard computation ───────────────────────────────────────────────────

def compute_leaderboards(pbp: pd.DataFrame) -> pd.DataFrame:
    """Compute wins / podiums / placements / events / career_span per person."""
    df = pbp.copy()
    df = df[~df["person_unresolved"].str.lower().isin(("true", "1"))]
    df = df[df["person_canon"].str.strip() != ""]
    df = df[df["person_canon"].str.strip() != "__NON_PERSON__"]
    df["_place"] = pd.to_numeric(df["place"], errors="coerce")
    df["_year"]  = pd.to_numeric(df["year"],  errors="coerce")

    wins     = df[df["_place"] == 1].groupby("person_canon").size().rename("wins")
    podiums  = df[df["_place"] <= 3].groupby("person_canon").size().rename("podiums")
    total    = df.groupby("person_canon").size().rename("placements")
    events   = df.groupby("person_canon")["event_id"].nunique().rename("events")
    first_yr = df.groupby("person_canon")["_year"].min().rename("first_year")
    last_yr  = df.groupby("person_canon")["_year"].max().rename("last_year")

    stats = pd.concat([wins, podiums, total, events, first_yr, last_yr], axis=1).fillna(0)
    stats["wins"]        = stats["wins"].astype(int)
    stats["podiums"]     = stats["podiums"].astype(int)
    stats["placements"]  = stats["placements"].astype(int)
    stats["events"]      = stats["events"].astype(int)
    stats["first_year"]  = stats["first_year"].astype(int)
    stats["last_year"]   = stats["last_year"].astype(int)
    stats["career_span"] = stats["last_year"] - stats["first_year"]
    return stats.reset_index()


def compute_leaderboards_by_cat(pbp: pd.DataFrame) -> dict:
    """Wins and podiums per person per division_category."""
    df = pbp.copy()
    df = df[~df["person_unresolved"].str.lower().isin(("true", "1"))]
    df = df[~df["person_canon"].str.strip().isin(["", "__NON_PERSON__"])]
    df["_place"] = pd.to_numeric(df["place"], errors="coerce")
    by_cat = {}
    for cat, cdf in df.groupby("division_category"):
        wins    = cdf[cdf["_place"] == 1].groupby("person_canon").size().rename("wins")
        podiums = cdf[cdf["_place"] <= 3].groupby("person_canon").size().rename("podiums")
        merged  = pd.concat([wins, podiums], axis=1).fillna(0).astype(int)
        merged  = merged.sort_values("wins", ascending=False).reset_index()
        by_cat[cat] = merged
    return by_cat


# ── Cell helper ───────────────────────────────────────────────────────────────

def _c(ws, row: int, col: int, value=None, *,
       font=None, fill=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    return cell


def _get_symbol(display: str, honours: dict) -> str:
    """Return honour symbol string for a display name (person or team)."""
    if display in honours:
        return honours[display]["symbol"]
    if "/" in display:
        symbols = set()
        for m in display.split("/"):
            m = m.strip()
            if m in honours:
                if honours[m]["bap"]:   symbols.add(SYM_BAP)
                if honours[m]["fbhof"]: symbols.add(SYM_FBHOF)
        if symbols:
            return (SYM_BAP if SYM_BAP in symbols else "") + \
                   (SYM_FBHOF if SYM_FBHOF in symbols else "")
    return ""


# ── ReadMe sheet ──────────────────────────────────────────────────────────────

def build_readme(wb: Workbook, events: dict, pf: pd.DataFrame):
    ws = wb.create_sheet("ReadMe")
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22

    n_events     = len(events)
    n_placements = len(pf[~pf["person_canon"].isin(["", "__NON_PERSON__"])])
    years        = sorted({ev["year"] for ev in events.values() if ev["year"]})
    yr_range     = f"{years[0]}–{years[-1]}" if years else "?"

    _c(ws, 1, 1, "Footbag Historical Results Archive", font=FONT_TITLE)

    sections = [
        (3,  "Coverage",   yr_range),
        (5,  "Events",     f"{n_events:,}"),
        (7,  "Placements", f"{n_placements:,}"),
    ]
    for row, label, value in sections:
        _c(ws, row, 1, label, font=FONT_SUBHEAD)
        _c(ws, row, 2, value, font=FONT_NORMAL)

    _c(ws, 9,  1, "Note", font=FONT_SUBHEAD)
    _c(ws, 10, 1, "Results data is incomplete for early years (pre-1997).", font=FONT_NORMAL)
    _c(ws, 11, 1, "Coverage improves significantly from 1997 onward.", font=FONT_SMALL)

    _c(ws, 13, 1, "Sources", font=FONT_SUBHEAD)
    _c(ws, 14, 1, "footbag.org archive",          font=FONT_NORMAL)
    _c(ws, 15, 1, "historical tournament records", font=FONT_NORMAL)

    _c(ws, 17, 1, "Compiled by",        font=FONT_SUBHEAD)
    _c(ws, 18, 1, "James Leberknight",  font=FONT_NORMAL)
    _c(ws, 19, 1, str(datetime.now().year), font=FONT_SMALL)



# ── Honours sheet ─────────────────────────────────────────────────────────────

def build_honours_sheet(wb: Workbook, honours: dict,
                        bap_rows: list, fbhof_rows: list,
                        stats: pd.DataFrame):
    ws = wb.create_sheet("Honours")
    ws.freeze_panes = "A3"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 8

    # Title
    ws.merge_cells("A1:H1")
    _c(ws, 1, 1, "Footbag Honours — Big Add Posse & Hall of Fame",
       font=FONT_TITLE, align=ALIGN_CENTER)

    # Sub-header
    ws.merge_cells("A2:H2")
    _c(ws, 2, 1,
       f"{SYM_BAP} Big Add Posse  |  {SYM_FBHOF} Footbag Hall of Fame",
       font=FONT_SMALL, align=ALIGN_CENTER)

    # Column headers
    hdrs = ["Year", "Name", SYM_BAP, SYM_FBHOF, "Nickname", "Wins", "Podiums", "Events"]
    for c, h in enumerate(hdrs, start=1):
        _c(ws, 3, c, h, font=FONT_HDR, fill=FILL_HDR, align=ALIGN_CENTER)

    # Build unified inductee list, keyed by BAP canonical name
    # (FBHOF-only entries use FBHOF name if no BAP match)
    bap_set  = {r["name"] for r in bap_rows}
    fbhof_set = {r["name"] for r in fbhof_rows}

    inductees = []  # list of dicts: {year, name, bap, fbhof, nickname, canon}

    # Collect all BAP entries first
    for r in bap_rows:
        fbhof_year = 0
        # Check if this person is also in FBHOF (via alias or exact)
        fbhof_name = _BAP_TO_FBHOF_ALIASES.get(r["name"], r["name"])
        for fr in fbhof_rows:
            if fr["name"] == fbhof_name or fr["name"] == r["name"]:
                fbhof_year = fr["year"]
                break
        inductees.append({
            "year":       r["year"],
            "name":       r["name"],
            "bap":        True,
            "fbhof":      fbhof_year > 0,
            "fbhof_year": fbhof_year,
            "nickname":   r["nickname"],
            "canon":      r["canon"],
        })

    # FBHOF-only entries (not in BAP)
    bap_canonical_names = {r["name"] for r in bap_rows}
    bap_aliases         = set(_BAP_TO_FBHOF_ALIASES.values())
    for fr in fbhof_rows:
        if fr["name"] not in bap_canonical_names and fr["name"] not in bap_aliases:
            inductees.append({
                "year":       fr["year"],
                "name":       fr["name"],
                "bap":        False,
                "fbhof":      True,
                "fbhof_year": fr["year"],
                "nickname":   "",
                "canon":      fr["canon"],
            })

    # Sort by induction year
    inductees.sort(key=lambda x: x["year"])

    # Build stats lookup  canon → plain dict (avoid Series truth-value issues)
    stats_map: dict = {}
    if not stats.empty:
        for _, row in stats.iterrows():
            stats_map[row["person_canon"]] = {
                "wins":    int(row.get("wins", 0)    or 0),
                "podiums": int(row.get("podiums", 0) or 0),
                "events":  int(row.get("events", 0)  or 0),
            }

    row_idx = 4
    for ind in inductees:
        canon  = ind["canon"]
        s      = stats_map.get(canon)
        wins   = s["wins"]    if s else 0
        pods   = s["podiums"] if s else 0
        evts   = s["events"]  if s else 0
        symbol = (SYM_BAP if ind["bap"] else "") + (SYM_FBHOF if ind["fbhof"] else "")

        if ind["bap"] and ind["fbhof"]:
            fill = FILL_HON_BOTH
        elif ind["bap"]:
            fill = FILL_HON_BAP
        else:
            fill = FILL_HON_FBHOF

        ws.cell(row=row_idx, column=1, value=ind["year"] or None)
        ws.cell(row=row_idx, column=2, value=_display_name(ind["name"]))
        ws.cell(row=row_idx, column=3, value=SYM_BAP   if ind["bap"]   else "")
        ws.cell(row=row_idx, column=4, value=SYM_FBHOF if ind["fbhof"] else "")
        ws.cell(row=row_idx, column=5, value=ind["nickname"] or "")
        ws.cell(row=row_idx, column=6, value=wins  or None)
        ws.cell(row=row_idx, column=7, value=pods  or None)
        ws.cell(row=row_idx, column=8, value=evts  or None)

        for c in range(1, 9):
            ws.cell(row=row_idx, column=c).fill = fill
            ws.cell(row=row_idx, column=c).alignment = ALIGN_CENTER \
                if c in (1, 3, 4, 6, 7, 8) else ALIGN_TOP

        row_idx += 1

    # Legend below table
    row_idx += 1
    _c(ws, row_idx, 1, "Legend", font=FONT_SUBHEAD)
    _c(ws, row_idx + 1, 1, f"{SYM_BAP} = Big Add Posse inductee",          font=FONT_SMALL)
    _c(ws, row_idx + 2, 1, f"{SYM_FBHOF} = Footbag Hall of Fame inductee", font=FONT_SMALL)
    _c(ws, row_idx + 3, 1, "BAP data includes nickname column where available.", font=FONT_SMALL)
    fill_legend = [FILL_HON_BAP, FILL_HON_FBHOF, FILL_HON_BOTH]
    legend_labels = ["BAP only", "FBHOF only", "Both"]
    for i, (f, lbl) in enumerate(zip(fill_legend, legend_labels)):
        _c(ws, row_idx + 5 + i, 1, lbl, font=FONT_SMALL, fill=f)


# ── Summary sheet ─────────────────────────────────────────────────────────────

def build_summary(wb: Workbook, events: dict, event_placements: dict,
                  stats: pd.DataFrame, pbp: pd.DataFrame):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width =  8
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width =  7
    ws.column_dimensions["E"].width =  7
    ws.column_dimensions["F"].width =  8

    N_COLS = 6
    def _merge_title(row, text, font):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        c = ws.cell(row=row, column=1, value=text)
        c.font = font
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22 if font.size >= 14 else 16

    # Title
    _merge_title(1, "Footbag Historical Results Archive", FONT_TITLE)

    # ── About ─────────────────────────────────────────────────────────────────
    _merge_title(3, "About", FONT_SECTION)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=N_COLS)
    about = ws.cell(row=4, column=1,
        value="A historical archive of competitive footbag results reconstructed from "
              "the Footbag.org website and curated identity records. Coverage is "
              "comprehensive from 1997 onward; pre-1997 data is partial. All player "
              "identities are human-verified.")
    about.font = FONT_SMALL
    about.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 36

    # ── Dataset overview ──────────────────────────────────────────────────────
    _merge_title(6, "Dataset Overview", FONT_SECTION)

    n_events     = len(events)
    n_placements = sum(len(v) for ep in event_placements.values() for v in ep.values())
    n_players    = len(stats) if not stats.empty else "?"
    years        = sorted({ev["year"] for ev in events.values() if ev["year"]})
    yr_range     = f"{years[0]}–{years[-1]}" if years else "?"

    for row_i, (label, value) in enumerate([
        ("Events",         f"{n_events:,}"),
        ("Years covered",  yr_range),
        ("Placements",     f"{n_placements:,}"),
        ("Unique players", f"{n_players:,}" if isinstance(n_players, int) else n_players),
    ], start=7):
        _c(ws, row_i, 1, label, font=FONT_SUBHEAD)
        _c(ws, row_i, 2, value, font=FONT_NORMAL)

    # ── Largest events ────────────────────────────────────────────────────────
    _merge_title(12, "Largest Events", FONT_SECTION)
    _c(ws, 13, 1, "Event",    font=Font(bold=True, size=9, color="FFFFFF"), fill=FILL_HDR)
    _c(ws, 13, 2, "Year",     font=Font(bold=True, size=9, color="FFFFFF"), fill=FILL_HDR)
    _c(ws, 13, 3, "Location", font=Font(bold=True, size=9, color="FFFFFF"), fill=FILL_HDR)
    _c(ws, 13, 4, "Players",  font=Font(bold=True, size=9, color="FFFFFF"), fill=FILL_HDR)
    ws.row_dimensions[13].height = 13

    event_sizes = []
    for eid, ep in event_placements.items():
        n = sum(len(v) for v in ep.values())
        ev = events.get(eid, {})
        event_sizes.append((ev.get("event_name", eid), ev.get("year", 0),
                            ev.get("location", ""), n))
    top_events = sorted(event_sizes, key=lambda x: x[3], reverse=True)[:15]

    for i, (name, year, loc, n) in enumerate(top_events):
        fill = FILL_ALT if i % 2 else FILL_WHITE
        r = 14 + i
        for ci, val in [(1, name), (2, year or ""), (3, loc), (4, n)]:
            c = ws.cell(row=r, column=ci, value=val)
            c.font = Font(size=9)
            c.fill = fill
        ws.row_dimensions[r].height = 13

    # ── Navigate this workbook ────────────────────────────────────────────────
    guide_start = 30
    _merge_title(guide_start, "Navigate this Workbook", FONT_SECTION)
    _c(ws, guide_start + 1, 1, "Sheet",       font=Font(bold=True, size=9, color="FFFFFF"), fill=FILL_HDR)
    _c(ws, guide_start + 1, 2, "Description", font=Font(bold=True, size=9, color="FFFFFF"), fill=FILL_HDR)
    ws.merge_cells(start_row=guide_start + 1, start_column=2,
                   end_row=guide_start + 1,   end_column=N_COLS)
    ws.row_dimensions[guide_start + 1].height = 13

    sheet_guide = [
        ("Summary",        "This page — dataset overview, largest events, and navigation guide"),
        ("Records",        "All-time and category leaderboards (top 15 players)"),
        ("Consecutives",   "Consecutive kicks records, world records, and milestone firsts"),
        ("Index",          "Full list of all events with dates, locations, coverage flags, and data quality notes"),
        ("Player Stats",   "Career statistics for each player"),
        ("Player Results", "Full placement history searchable by player name"),
        ("Year sheets",    "All events and division results for each year (1980–2026)"),
    ]
    for i, (sheet, desc) in enumerate(sheet_guide):
        fill = FILL_ALT if i % 2 else FILL_WHITE
        r = guide_start + 2 + i
        name_cell = ws.cell(row=r, column=1, value=sheet)
        name_cell.font = Font(bold=True, size=9)
        name_cell.fill = fill
        desc_cell = ws.cell(row=r, column=2, value=desc)
        desc_cell.font = Font(size=9)
        desc_cell.fill = fill
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=N_COLS)
        ws.row_dimensions[r].height = 13


# ── Records sheet ─────────────────────────────────────────────────────────────

def build_records(wb: Workbook, stats: pd.DataFrame, cat_stats: dict,
                  honours: dict):
    ws = wb.create_sheet("Records")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width =  4
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width =  4

    _c(ws, 1, 1, "Records", font=FONT_TITLE)

    def _write_lb(row, col, title, df_in, val_col, val_label, n=15):
        _c(ws, row, col,   title,     font=FONT_SUBHEAD, border=_border_top())
        _c(ws, row, col+1, val_label, font=FONT_SUBHEAD, border=_border_top())
        r = row + 1
        try:
            top = (df_in[["person_canon", val_col]].copy()
                   .assign(**{val_col: pd.to_numeric(df_in[val_col], errors="coerce")})
                   .dropna(subset=[val_col])
                   .nlargest(n, val_col))
            for i, (_, row_data) in enumerate(top.iterrows()):
                fill = FILL_ALT if i % 2 else FILL_WHITE
                name_val = _display_name(str(row_data["person_canon"]))
                num_val  = int(row_data[val_col])
                nc = ws.cell(row=r, column=col,   value=name_val)
                vc = ws.cell(row=r, column=col+1, value=num_val)
                nc.font = Font(size=9); nc.fill = fill
                vc.font = Font(size=9); vc.fill = fill
                vc.alignment = Alignment(horizontal="right")
                r += 1
        except Exception:
            ws.cell(row=r, column=col, value="(unavailable)")
        return r + 1

    # ── Column A/B: all-time records ──────────────────────────────────────────
    r = 3
    if not stats.empty:
        r = _write_lb(r, 1, "Most Wins — All Time",    stats, "wins",        "Wins")
        r = _write_lb(r, 1, "Most Podium Finishes",    stats, "podiums",     "Podiums")
        r = _write_lb(r, 1, "Longest Careers (years)", stats, "career_span", "Years")

    # ── Column D/E: by category — wins then podiums ───────────────────────────
    cat_order = [
        ("freestyle", "Freestyle Wins",    "wins"),
        ("freestyle", "Freestyle Podiums", "podiums"),
        ("net",       "Net Wins",          "wins"),
        ("net",       "Net Podiums",       "podiums"),
        ("golf",      "Golf Wins",         "wins"),
    ]
    rc = 3
    for cat, label, col_key in cat_order:
        if cat in cat_stats and not cat_stats[cat].empty and col_key in cat_stats[cat].columns:
            rc = _write_lb(rc, 4, label, cat_stats[cat], col_key, col_key.capitalize())



# ── Consecutives Records sheet ────────────────────────────────────────────────

def _load_consecutives_records(path: Path) -> list[dict]:
    """Load out/consecutives_combined.csv; return list of row dicts sorted by sort_order."""
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    for r in rows:
        try:
            r["_sort"] = int(r.get("sort_order") or 0)
        except ValueError:
            r["_sort"] = 0
    rows.sort(key=lambda r: r["_sort"])
    return rows


def build_consecutives_records(wb: Workbook, records: list[dict]):
    if not records:
        return

    ws = wb.create_sheet("Consecutives")

    # Column widths (A–J)
    col_widths = [6, 5, 26, 22, 22, 10, 36, 13, 36, 30]
    col_labels = ["Year", "Rank", "Division", "Person / Team",
                  "Partner", "Score", "Note", "Event Date", "Event Name", "Location"]
    n_cols = len(col_widths)
    last_col_ltr = get_column_letter(n_cols)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    FILL_SUBSEC = _fill("D6E4F0")   # light blue for subsection headers
    FILL_COLHDR = _fill("EBF3FB")   # pale blue for column header row

    def _banner(row, text, fill, font):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font  = font
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1, wrap_text=False)
        ws.row_dimensions[row].height = 18
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n_cols)

    def _col_headers(row):
        for ci, label in enumerate(col_labels, 1):
            c = ws.cell(row=row, column=ci, value=label)
            c.font      = Font(bold=True, size=8, color="FFFFFF")
            c.fill      = FILL_HDR
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 13

    # Title row
    ws.row_dimensions[1].height = 22
    cell = ws.cell(row=1, column=1, value="Consecutive Kicks Records")
    cell.font      = FONT_TITLE
    cell.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    r = 3

    # Group by section, then subsection (order preserved from sort_order)
    from itertools import groupby
    for section, sec_rows in groupby(records, key=lambda x: x["section"]):
        sec_rows = list(sec_rows)
        _banner(r, section, FILL_BANNER, FONT_BANNER)
        r += 1

        for subsection, sub_rows in groupby(sec_rows, key=lambda x: x["subsection"]):
            sub_rows = list(sub_rows)
            _banner(r, subsection, FILL_SUBSEC,
                    Font(bold=True, size=10, color="1F4E79"))
            r += 1
            _col_headers(r)
            r += 1

            for i, row_data in enumerate(sub_rows):
                fill = FILL_ALT if i % 2 else FILL_WHITE

                year    = row_data.get("year", "")
                rank    = row_data.get("rank", "")
                div     = row_data.get("division", "")
                person  = row_data.get("person_or_team", "")
                partner = row_data.get("partner", "")
                raw_score = row_data.get("score", "")
                note    = row_data.get("note", "")
                edate   = row_data.get("event_date", "")
                ename   = row_data.get("event_name", "")
                loc     = row_data.get("location", "")

                try:
                    score_val = int(raw_score) if raw_score else None
                except ValueError:
                    score_val = raw_score or None

                values = [year, rank, div, person, partner,
                          score_val, note, edate, ename, loc]

                for ci, val in enumerate(values, 1):
                    c = ws.cell(row=r, column=ci, value=val or None)
                    c.fill = fill
                    c.font = Font(size=9)
                    c.alignment = Alignment(vertical="top", wrap_text=(ci in (7, 9, 10)))
                    if ci == 6 and isinstance(val, int):  # Score: right-align, bold
                        c.font      = Font(size=9, bold=True)
                        c.alignment = Alignment(horizontal="right", vertical="top")

                ws.row_dimensions[r].height = 14
                r += 1

        r += 1   # blank row between sections

    ws.freeze_panes = "A3"


# ── Known issues ──────────────────────────────────────────────────────────────

def load_known_issues() -> dict[str, dict]:
    """Return dict event_id → {severity, note} from overrides/known_issues.csv."""
    result = {}
    if not KNOWN_ISSUES_CSV.exists():
        return result
    with open(KNOWN_ISSUES_CSV, newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            result[row["event_id"]] = {
                "severity": row.get("severity", "minor"),
                "note":     row["note"],
            }
    return result

# Row background colours for Index sheet (traffic-light)
_ROW_FILL = {
    None:       ("F0FFF0", "E8F5E9"),  # no issue  — green tones (even/odd)
    "minor":    ("FFFDE7", "FFF9C4"),  # minor     — pale yellow
    "moderate": ("FFF3E0", "FFE0B2"),  # moderate  — pale orange
    "severe":   ("FFEBEE", "FFCDD2"),  # severe    — pale red
}
# Note text colours
_NOTE_COLOR = {
    "minor":    "997700",
    "moderate": "CC5500",
    "severe":   "CC0000",
}


# ── Index sheet ───────────────────────────────────────────────────────────────

def build_index_real(wb: Workbook, events: dict, event_placements: dict,
                     event_col_map: dict, insert_at: int,
                     event_coverage: dict = None,
                     known_issues: dict = None):
    """Build the Index sheet with hyperlinks and insert at the correct position."""
    ws = wb.create_sheet("Index")
    wb.move_sheet("Index", offset=-(len(wb.sheetnames) - 1 - insert_at))

    ws.freeze_panes = "A2"

    hdrs   = ["Year", "Event", "City / Region", "Country", "Host Club",
              "Divisions", "Players", "Coverage", "Data Notes"]
    widths = [7, 48, 32, 12, 30, 11, 10, 12, 30]
    for c, (h, w) in enumerate(zip(hdrs, widths), start=1):
        _c(ws, 1, c, h, font=FONT_HDR, fill=FILL_HDR, align=ALIGN_CENTER)
        ws.column_dimensions[get_column_letter(c)].width = w

    all_eids = sorted(
        events.keys(),
        key=lambda eid: (events[eid]["year"],
                         _date_sort_key(events[eid].get("date", ""), eid)),
    )

    for row_idx, eid in enumerate(all_eids, start=2):
        ev  = events[eid]
        ep  = event_placements.get(eid, {})
        n_p = sum(len(v) for v in ep.values())
        n_d = len(ep)

        ws.cell(row=row_idx, column=1, value=ev["year"] or "?")
        ws.cell(row=row_idx, column=3, value=ev.get("city", ev.get("location", "")))
        ws.cell(row=row_idx, column=4, value=ev.get("country", ""))
        ws.cell(row=row_idx, column=5, value=ev.get("host_club", ""))
        ws.cell(row=row_idx, column=6, value=n_d)
        ws.cell(row=row_idx, column=7, value=n_p)

        flag = (event_coverage or {}).get(str(eid), "complete")
        if flag in ("partial", "sparse"):
            cov_cell = ws.cell(row=row_idx, column=8, value=flag)
            cov_cell.font = Font(italic=True, size=9, color="CC0000")

        issue   = (known_issues or {}).get(str(eid))
        severity = issue["severity"] if issue else None
        if issue:
            note_cell = ws.cell(row=row_idx, column=9, value=issue["note"])
            note_cell.font = Font(italic=True, size=9,
                                  color=_NOTE_COLOR.get(severity, "996600"))

        cell = ws.cell(row=row_idx, column=2, value=ev["event_name"])
        if eid in event_col_map:
            sheet_name, col_letter = event_col_map[eid]
            safe = sheet_name.replace("'", "''")
            cell.hyperlink = f"#'{safe}'!{col_letter}1"
            cell.font = FONT_LINK
        else:
            cell.font = FONT_NORMAL

        # Row background: traffic-light by severity (even/odd variants)
        row_fills = _ROW_FILL[severity]
        fill_hex  = row_fills[row_idx % 2]
        for c in range(1, 10):
            cell_obj = ws.cell(row=row_idx, column=c)
            if cell_obj.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                cell_obj.fill = _fill(fill_hex)


# ── Player Stats sheet ────────────────────────────────────────────────────────

def build_player_stats(wb: Workbook, stats: pd.DataFrame, honours: dict,
                       persons_df: pd.DataFrame | None = None):
    """
    One row per resolved player showing career statistics.
    Filterable so any player can look themselves up.
    """
    ws = wb.create_sheet("Player Stats")
    ws.freeze_panes = "A2"

    hdrs   = ["Player", "Nickname", "Wins", "Podiums", "Placements", "Events",
              "Legacy ID"]
    widths = [32, 18, 8, 8, 12, 8, 10]

    for c, (h, w) in enumerate(zip(hdrs, widths), start=1):
        _c(ws, 1, c, h, font=FONT_HDR, fill=FILL_HDR)
        ws.column_dimensions[get_column_letter(c)].width = w

    if stats.empty:
        return

    # Build legacyid lookup from Persons_Truth
    legacyid_map: dict = {}
    if persons_df is not None and "legacyid" in persons_df.columns:
        for _, pr in persons_df.iterrows():
            lid = pr.get("legacyid", "")
            if lid:
                legacyid_map[pr["person_canon"]] = lid

    df = stats.sort_values("person_canon").reset_index(drop=True)

    for r_idx, row in df.iterrows():
        pc        = row["person_canon"]
        excel_row = r_idx + 2

        ws.cell(row=excel_row, column=1, value=_display_name(pc))
        nick = honours.get(pc, {}).get("nickname", "") if honours.get(pc, {}).get("bap") else ""
        ws.cell(row=excel_row, column=2, value=nick or None)
        ws.cell(row=excel_row, column=3, value=int(row["wins"]))
        ws.cell(row=excel_row, column=4, value=int(row["podiums"]))
        ws.cell(row=excel_row, column=5, value=int(row["placements"]))
        ws.cell(row=excel_row, column=6, value=int(row["events"]))
        lid = legacyid_map.get(pc, "")
        ws.cell(row=excel_row, column=7, value=int(lid) if lid else None)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{len(df) + 1}"


# ── Player Results sheet ──────────────────────────────────────────────────────

def build_player_results(wb: Workbook, pf: pd.DataFrame, events: dict):
    ws = wb.create_sheet("Player Results")
    ws.freeze_panes = "A2"

    hdrs   = ["Year", "Event", "Location", "Division", "Category",
              "Place", "Player", "Partner"]
    widths = [7, 48, 32, 28, 12, 7, 28, 28]

    for c, (h, w) in enumerate(zip(hdrs, widths), start=1):
        _c(ws, 1, c, h, font=FONT_HDR, fill=FILL_HDR)
        ws.column_dimensions[get_column_letter(c)].width = w

    df = pf.copy()
    df = df[~df["person_unresolved"].str.lower().isin(("true", "1"))]
    df = df[~df["person_canon"].isin(["", "__NON_PERSON__"])]

    df["_place"] = pd.to_numeric(df["place"], errors="coerce")
    df["_year"]  = pd.to_numeric(df["year"],  errors="coerce")
    df = df.sort_values(["_year", "event_id", "division_canon", "_place",
                          "team_person_key", "person_canon"],
                        na_position="last")

    # Pre-build partner lookup for doubles
    seen_teams: dict = {}
    for _, row in df[df["competitor_type"] == "team"].iterrows():
        tpk = (row.get("team_person_key") or "").strip()
        if not tpk:
            continue
        grp = df[
            (df["event_id"]       == row["event_id"]) &
            (df["division_canon"] == row["division_canon"]) &
            (df["place"]          == row["place"]) &
            (df["team_person_key"]== tpk) &
            (df["person_canon"]   != row["person_canon"])
        ]["person_canon"].tolist()
        seen_teams[(row["event_id"], row["division_canon"], row["place"],
                    tpk, row["person_canon"])] = " / ".join(grp) if grp else ""

    row_idx = 2
    for _, row in df.iterrows():
        eid    = row["event_id"]
        ev     = events.get(eid, {})
        person = (row.get("person_canon") or "").strip()
        tpk    = (row.get("team_person_key") or "").strip()

        partner = ""
        if (row.get("competitor_type") or "").lower() == "team" and tpk:
            partner = seen_teams.get(
                (eid, row["division_canon"], row["place"], tpk, person), ""
            )
            if not partner:
                td = (row.get("team_display_name") or "").strip()
                if td and person in td:
                    partner = td.replace(person, "").strip(" /")
                elif td:
                    partner = td

        try:
            place_val = int(float(row["place"]))
        except (ValueError, TypeError):
            place_val = row["place"]

        ws.cell(row=row_idx, column=1, value=ev.get("year") or _to_int(row.get("year")))
        ws.cell(row=row_idx, column=2, value=ev.get("event_name") or eid)
        ws.cell(row=row_idx, column=3, value=ev.get("location", ""))
        ws.cell(row=row_idx, column=4, value=row.get("division_canon", "").rstrip(":").strip())
        ws.cell(row=row_idx, column=5, value=row.get("division_category", ""))
        ws.cell(row=row_idx, column=6, value=place_val)
        ws.cell(row=row_idx, column=7, value=_display_name(person))
        ws.cell(row=row_idx, column=8, value=partner)
        row_idx += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{row_idx - 1}"


# ── Year sheets ───────────────────────────────────────────────────────────────

# Fixed row positions for the event header block
# (columns B onward; column A holds row labels)
_R_NAME    = 1   # Event name
_R_LOC     = 2   # Location
_R_HOST    = 3   # Host club  (italic per spec)
_R_DATE    = 4   # Date
_R_PLAYERS = 5   # Players count
_R_EVTYPE  = 6   # Event type
_R_EID     = 7   # Legacy event ID
_R_BLANK   = 8   # Spacer
_R_DATA    = 9   # First division / placement row

_ROW_LABELS = {
    _R_NAME:    "Event",
    _R_LOC:     "Location",
    _R_HOST:    "Host Club",
    _R_DATE:    "Date",
    _R_PLAYERS: "Players",
    _R_EVTYPE:  "Event Type",
    _R_EID:     "Event ID",
}


# Category display constants for grouped division headers in year sheets
CAT_ORDER  = ["net", "freestyle", "golf", "sideline", "unknown"]
CAT_LABELS = {"net": "NET", "freestyle": "FREESTYLE", "golf": "GOLF",
              "sideline": "SIDELINE", "unknown": "OTHER"}
FILL_CAT   = _fill("D0D0D0")
FONT_CAT   = Font(bold=True, size=8, color="444444")


def _write_event_col(ws, col: int, ev: dict, placements: OrderedDict,
                     honours: dict, coverage_flag: str = "complete") -> tuple:
    """
    Write one event into column `col` (1-based, already offset for label col).
    Returns (last_row_written, max_content_length).
    Divisions are grouped by category (NET / FREESTYLE / GOLF / SIDELINE / OTHER).
    """
    n_players   = sum(len(v) for v in placements.values())
    max_content = max(len(ev.get("event_name", "")), 24)

    def _write(r, val, font, fill, align=ALIGN_TOP):
        nonlocal max_content
        if val:
            max_content = max(max_content, len(str(val)))
        _c(ws, r, col, val, font=font, fill=fill, align=align)

    is_worlds   = ev.get("event_type", "") == "worlds"
    banner_fill = FILL_BANNER_WORLDS if is_worlds else FILL_BANNER
    _write(_R_NAME,    ev["event_name"],                FONT_BANNER,  banner_fill, ALIGN_WRAP)
    _write(_R_LOC,     ev["location"] or "—",           FONT_META,    FILL_META)
    _write(_R_HOST,    ev["host_club"] or "",            FONT_HOST,    FILL_META)
    _write(_R_DATE,    ev["date"] or "",                 FONT_META,    FILL_META)
    _write(_R_PLAYERS, f"Players: {n_players}",         FONT_PLAYERS, FILL_PLAYERS)
    _write(_R_EVTYPE,  ev.get("event_type") or "",      FONT_META,    FILL_META)
    _write(_R_EID,     ev.get("event_id")  or "",      FONT_ROW_LBL, FILL_META)

    row = _R_DATA

    # Group divisions by category while preserving source order within each category
    cat_to_divs: dict = {}
    for div_name, entries in placements.items():
        if not entries:
            continue
        cat = (entries[0][2] or "unknown")
        if cat not in cat_to_divs:
            cat_to_divs[cat] = []
        cat_to_divs[cat].append((div_name, entries))

    for cat in CAT_ORDER:
        if cat not in cat_to_divs:
            continue
        # Category header row (only if this category has ≥1 division)
        cat_label = CAT_LABELS.get(cat, "OTHER")
        _c(ws, row, col, cat_label, font=FONT_CAT, fill=FILL_CAT, align=ALIGN_TOP)
        max_content = max(max_content, len(cat_label) + 2)
        row += 1

        for div_name, entries in cat_to_divs[cat]:
            # Division header: bold, light-grey, top border
            _c(ws, row, col, div_name,
               font=FONT_DIV, fill=FILL_DIV, border=_border_top(), align=ALIGN_TOP)
            max_content = max(max_content, len(div_name) + 2)
            row += 1

            for place_int, display, _ in entries:
                medal = MEDALS.get(place_int, "")

                # Build display text:  🥇 1  Name
                parts = []
                if medal:
                    parts.append(medal)
                parts.append(f"{place_int:>3} ")
                parts.append(display)
                text = " ".join(parts) if medal else "".join(parts)

                if place_int == 1:
                    fill, font = FILL_GOLD,   FONT_PODIUM
                elif place_int == 2:
                    fill, font = FILL_SILVER, FONT_PODIUM
                elif place_int == 3:
                    fill, font = FILL_BRONZE, FONT_PODIUM
                else:
                    fill, font = FILL_WHITE,  FONT_PLACE

                _c(ws, row, col, text, font=font, fill=fill, align=ALIGN_TOP)
                max_content = max(max_content, len(text) + 2)
                row += 1

            row += 1   # blank row between divisions

    return row - 1, max_content


def build_year_sheet(wb: Workbook, year: int, eids: list,
                     events: dict, event_placements: dict,
                     honours: dict, event_coverage: dict = None) -> dict:
    """
    Build one year sheet with:
    - Column A: row labels (Event, Location, Host Club, Date, Players)
    - Columns B onward: one event per column
    - freeze_panes = "B1"  (column A stays visible when scrolling right)
    - Auto column width (min COL_W_MIN)

    Returns dict event_id → column_letter (B, C, D, …)
    """
    ws = wb.create_sheet(title=str(year))

    # ── Column A: row-label column ────────────────────────────────────────────
    ws.column_dimensions["A"].width = COL_W_LABEL
    for row_num, label in _ROW_LABELS.items():
        _c(ws, row_num, 1, label,
           font=FONT_ROW_LBL, fill=FILL_ROW_LABEL, align=ALIGN_RIGHT)

    sorted_eids = sorted(
        eids,
        key=lambda eid: _date_sort_key(events[eid].get("date", ""), eid),
    )

    event_col_map: dict = {}
    col_max_widths: dict = {}

    for col_offset, eid in enumerate(sorted_eids, start=2):   # B=2, C=3, …
        ev         = events[eid]
        placements = event_placements.get(eid, OrderedDict())
        flag       = (event_coverage or {}).get(str(eid), "complete")
        last_row, max_w = _write_event_col(ws, col_offset, ev, placements, honours,
                                           coverage_flag=flag)
        event_col_map[eid]       = get_column_letter(col_offset)
        col_max_widths[col_offset] = max_w

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[_R_NAME].height    = 36
    ws.row_dimensions[_R_LOC].height     = 15
    ws.row_dimensions[_R_HOST].height    = 15
    ws.row_dimensions[_R_DATE].height    = 15
    ws.row_dimensions[_R_PLAYERS].height = 15
    ws.row_dimensions[_R_EVTYPE].height  = 15
    ws.row_dimensions[_R_EID].height    = 13

    # ── Auto-width per event column (min COL_W_MIN, cap at 60) ───────────────
    for col_idx, max_w in col_max_widths.items():
        ltr = get_column_letter(col_idx)
        ws.column_dimensions[ltr].width = max(min(max_w + 4, 60), COL_W_MIN)

    # ── Freeze: column A always visible when scrolling right ─────────────────
    ws.freeze_panes = "B1"

    return event_col_map


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading data…")
    s2_events  = load_stage2_events()
    pf         = load_placements_flat()
    pbp        = load_placements_by_person()
    persons_df = load_persons_truth()

    print("Loading honours (BAP / FBHOF)…")
    honours, bap_rows, fbhof_rows = load_honours(persons_df)
    print(f"  {sum(1 for h in honours.values() if h['bap'])} BAP  "
          f"| {sum(1 for h in honours.values() if h['fbhof'])} FBHOF  "
          f"| {sum(1 for h in honours.values() if h['bap'] and h['fbhof'])} both")

    print("Building event placements…")
    event_placements = build_event_placements(pf, s2_events)
    event_coverage   = compute_event_coverage(pf)
    known_issues     = load_known_issues()

    print("Computing leaderboards…")
    stats     = compute_leaderboards(pbp)
    cat_stats = compute_leaderboards_by_cat(pbp)

    # Group events by year (only years with placements get a sheet)
    year_to_eids: dict = defaultdict(list)
    for eid in s2_events:
        yr = s2_events[eid]["year"]
        if yr and eid in event_placements and event_placements[eid]:
            year_to_eids[yr].append(eid)

    print("Creating workbook…")
    wb = Workbook()
    wb.remove(wb.active)

    cons_records = _load_consecutives_records(OUT_DIR / "consecutives_combined.csv")

    # Sheet order: Summary, Records, Consecutives, [Index placeholder],
    #              Player Stats, Player Results, year sheets
    build_summary(wb, s2_events, event_placements, stats, pbp)
    build_records(wb, stats, cat_stats, honours)
    build_consecutives_records(wb, cons_records)

    # Index placeholder — correct content added after year sheets are built
    idx_placeholder = wb.create_sheet("Index")

    build_player_stats(wb, stats, honours, persons_df=persons_df)
    build_player_results(wb, pf, s2_events)

    # ── Year sheets ───────────────────────────────────────────────────────────
    all_event_col_map: dict = {}   # event_id → (sheet_title, col_letter)

    sorted_years = sorted(year_to_eids.keys())
    print(f"Building {len(sorted_years)} year sheets…")
    for year in sorted_years:
        col_map = build_year_sheet(
            wb, year, year_to_eids[year], s2_events, event_placements, honours,
            event_coverage=event_coverage,
        )
        for eid, col_letter in col_map.items():
            all_event_col_map[eid] = (str(year), col_letter)

    # ── Rebuild Index now that year positions are known ───────────────────────
    # Sheet order after year sheets: [..., Index placeholder at position 2]
    # We remove placeholder and insert real Index at position 2
    # (Summary=0, Records=1, Index=2, …)
    wb.remove(idx_placeholder)
    build_index_real(wb, s2_events, event_placements, all_event_col_map,
                     insert_at=3, event_coverage=event_coverage,
                     known_issues=known_issues)

    # ── Save ──────────────────────────────────────────────────────────────────
    print(f"Saving {XLSX}…")
    wb.save(XLSX)
    n_placements = sum(len(v) for ep in event_placements.values() for v in ep.values())
    print(f"Done.  Events: {len(s2_events)}, Placements: {n_placements:,}, "
          f"Year sheets: {len(sorted_years)}")


if __name__ == "__main__":
    main()
