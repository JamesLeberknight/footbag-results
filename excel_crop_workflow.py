#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import shutil
import tempfile
from pathlib import Path
from typing import Optional

from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

VALID_EXTS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=THIN_GRAY)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def iter_images(in_dir: Path) -> list[Path]:
    return sorted(
        p for p in in_dir.iterdir()
        if p.is_file() and p.suffix.lower() in VALID_EXTS
    )


def pixels_to_row_height(px: int) -> float:
    # Approx Excel row height conversion.
    return max(18.0, px * 0.75)


def build_workbook(input_dir: Path, workbook_path: Path, thumb_max_w: int, thumb_max_h: int) -> None:
    files = iter_images(input_dir)
    if not files:
        raise SystemExit(f"No images found in {input_dir}")

    wb = Workbook()
    ws = wb.active
    ws.title = "CropReview"
    ws.freeze_panes = "A2"

    headers = [
        "filename",
        "status",
        "crop_x1",
        "crop_y1",
        "crop_x2",
        "crop_y2",
        "orig_w",
        "orig_h",
        "notes",
        "image",
    ]
    ws.append(headers)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    widths = {
        "A": 42,
        "B": 12,
        "C": 10,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 24,
        "J": 52,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    note_text = (
        "Enter crop_x1, crop_y1, crop_x2, crop_y2 in original-image pixels. "
        "Leave blank to skip. Status can be crop / skip / review."
    )
    ws["L1"] = "Instructions"
    ws["L1"].font = Font(bold=True)
    ws["L2"] = note_text
    ws["L2"].alignment = WRAP
    ws.column_dimensions["L"].width = 48

    tmp_dir = Path(tempfile.mkdtemp(prefix="excel_crop_thumbs_"))
    try:
        row = 2
        for img_path in files:
            with Image.open(img_path) as im:
                orig_w, orig_h = im.size
                thumb = im.copy()
                thumb.thumbnail((thumb_max_w, thumb_max_h))
                thumb_path = tmp_dir / f"{img_path.stem}.png"
                thumb.save(thumb_path, format="PNG")
                thumb_w, thumb_h = thumb.size

            ws.cell(row, 1, img_path.name)
            ws.cell(row, 2, "review")
            ws.cell(row, 7, orig_w)
            ws.cell(row, 8, orig_h)
            ws.cell(row, 9, "")
            for c in range(1, 10):
                ws.cell(row, c).alignment = WRAP
                ws.cell(row, c).border = BORDER

            img = XLImage(str(thumb_path))
            img.anchor = f"J{row}"
            ws.add_image(img)
            ws.row_dimensions[row].height = pixels_to_row_height(thumb_h + 8)
            row += 1

        ws.auto_filter.ref = f"A1:J{row-1}"
        wb.save(workbook_path)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def parse_coord(value) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(round(float(value)))
    except Exception:
        return None


def apply_crops(workbook_path: Path, input_dir: Path, output_dir: Path, status_filter: set[str]) -> tuple[int, int]:
    wb = load_workbook(workbook_path, data_only=True)
    if "CropReview" not in wb.sheetnames:
        raise SystemExit("Workbook missing sheet 'CropReview'")
    ws = wb["CropReview"]

    output_dir.mkdir(parents=True, exist_ok=True)
    written = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        filename, status, x1, y1, x2, y2, orig_w, orig_h, notes, image = row[:10]
        if not filename:
            continue
        status_norm = str(status or "review").strip().lower()
        if status_filter and status_norm not in status_filter:
            skipped += 1
            continue

        x1 = parse_coord(x1)
        y1 = parse_coord(y1)
        x2 = parse_coord(x2)
        y2 = parse_coord(y2)
        if None in (x1, y1, x2, y2):
            skipped += 1
            continue

        src = input_dir / str(filename)
        if not src.exists():
            skipped += 1
            continue

        with Image.open(src) as im:
            w, h = im.size
            x1c = max(0, min(w, x1))
            x2c = max(0, min(w, x2))
            y1c = max(0, min(h, y1))
            y2c = max(0, min(h, y2))
            if x2c <= x1c or y2c <= y1c:
                skipped += 1
                continue
            cropped = im.crop((x1c, y1c, x2c, y2c))
            out_path = output_dir / f"{Path(filename).stem}_crop.png"
            cropped.save(out_path, format="PNG")
            written += 1

    return written, skipped


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create an Excel workbook for manual crop review, or apply crop coordinates from that workbook."
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_make = sub.add_parser("make", help="Create crop review workbook")
    p_make.add_argument("--input", default="out/preprocessed", help="Directory of preprocessed images")
    p_make.add_argument("--workbook", default="out/crop_review.xlsx", help="Workbook to create")
    p_make.add_argument("--thumb-max-w", type=int, default=320, help="Thumbnail max width")
    p_make.add_argument("--thumb-max-h", type=int, default=420, help="Thumbnail max height")

    p_apply = sub.add_parser("apply", help="Apply crop coordinates from workbook")
    p_apply.add_argument("--input", default="out/preprocessed", help="Directory of source images")
    p_apply.add_argument("--workbook", default="out/crop_review.xlsx", help="Workbook to read")
    p_apply.add_argument("--output", default="out/manual_crops", help="Directory for cropped outputs")
    p_apply.add_argument(
        "--status",
        default="crop",
        help="Comma-separated statuses to process, e.g. crop or crop,review",
    )

    args = parser.parse_args()

    if args.command == "make":
        build_workbook(
            input_dir=Path(args.input),
            workbook_path=Path(args.workbook),
            thumb_max_w=args.thumb_max_w,
            thumb_max_h=args.thumb_max_h,
        )
        print(f"Created {args.workbook}")
        return 0

    if args.command == "apply":
        statuses = {s.strip().lower() for s in args.status.split(",") if s.strip()}
        written, skipped = apply_crops(
            workbook_path=Path(args.workbook),
            input_dir=Path(args.input),
            output_dir=Path(args.output),
            status_filter=statuses,
        )
        print(f"Wrote {written} crops; skipped {skipped} rows")
        return 0

    return 1


if __name__ == "__main__":
    raise SystemExit(main())
