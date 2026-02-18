#!/usr/bin/env python3
"""
03_build_excel.py — Stage 3: Build final Excel workbook

This script:
- Reads out/stage2_canonical_events.csv
- Generates Excel workbook with one sheet per year
- Outputs: Footbag_Results_Canonical.xlsx

Input: out/stage2_canonical_events.csv
Output: Footbag_Results_Canonical.xlsx
"""

from __future__ import annotations

import csv
import sys
import json
import re
import string
import hashlib
import unicodedata
from copy import copy
from pathlib import Path
from typing import Optional
from collections import defaultdict

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Import master QC orchestrator
try:
    from qc_master import run_qc_for_stage, print_qc_summary
    from qc_slop_detection import run_slop_detection_checks_stage3_excel
    USE_MASTER_QC = True
except ImportError:
    run_slop_detection_checks_stage3_excel = None
    print("Warning: Could not import qc_master, Stage 3 QC will not run")
    USE_MASTER_QC = False


# Excel/openpyxl rejects control chars: 0x00-0x08, 0x0B-0x0C, 0x0E-0x1F
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def load_alias_map(path):
    """
    Read-only alias map.
    Returns: dict[player_id] -> alias_group_id

    No guessing.
    No transitive closure.
    Blank means 'unknown / not yet reviewed'.
    """
    import csv
    from pathlib import Path

    alias_map = {}
    p = Path(path)

    if not p.exists():
        return alias_map

    with p.open("r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            pid = row.get("player_id", "").strip()
            agid = row.get("alias_group_id", "").strip()

            if pid and agid:
                alias_map[pid] = agid

    return alias_map


ALIAS_MAP = load_alias_map("out/person_alias_map_bootstrap.csv")


def sanitize_excel_strings(df: pd.DataFrame) -> pd.DataFrame:
    """Required to write .xlsx safely (not semantic cleaning)."""
    if df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_string_dtype(out[col]) or out[col].dtype == object:
            out[col] = out[col].apply(
                lambda v: _ILLEGAL_XLSX_RE.sub("", v) if isinstance(v, str) else v
            )
    return out


def sanitize_string(s: str) -> str:
    """Sanitize a single string for Excel."""
    if not isinstance(s, str):
        return s
    return _ILLEGAL_XLSX_RE.sub("", s)


def _strip_diacritics(s: str) -> str:
    if not isinstance(s, str):
        return ""
    # NFKD splits accents; we drop combining marks
    return "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))


def normalize_person_key(name: str) -> str:
    """
    Conservative, presentation-only normalization key for alias-candidate grouping.
    Lowercase, strip diacritics, remove punctuation, collapse whitespace.
    """
    if not isinstance(name, str) or not name.strip():
        return ""
    t = _strip_diacritics(name).lower()
    t = re.sub(r"[^a-z0-9\s]", " ", t)   # remove punctuation/symbols
    t = re.sub(r"\s+", " ", t).strip()
    # Optional: remove single-letter middle initials (keeps first + last)
    t = re.sub(r"\b([a-z])\b", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _stable_group_id(prefix: str, key: str) -> str:
    h = hashlib.sha1(f"{prefix}:{key}".encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{h}"


def _best_display_name(names: list[str]) -> str:
    cleaned = [n.strip() for n in names if isinstance(n, str) and n.strip()]
    if not cleaned:
        return ""
    # Prefer shortest (usually most canonical-looking), then alpha
    cleaned.sort(key=lambda x: (len(x), x.lower()))
    return cleaned[0]


def _alias_confidence(alias_names: list[str], key: str) -> str:
    """
    Simple heuristic:
    - high: all aliases normalize to same key
    - med: >1 alias and share same last token after normalization
    - low: otherwise
    """
    normed = [normalize_person_key(a) for a in alias_names if isinstance(a, str)]
    normed = [n for n in normed if n]
    if normed and all(n == key for n in normed):
        return "high"
    # last-token check
    toks = [n.split() for n in normed if n.split()]
    lasts = [t[-1] for t in toks if t]
    if len(set(lasts)) == 1 and len(lasts) >= 2:
        return "med"
    return "low"


def normalize_team_key(p1: str, p2: str) -> str:
    """
    Presentation-only team grouping key based on *member names* (not IDs),
    order-invariant: (A,B) == (B,A).
    """
    a = normalize_person_key(p1)
    b = normalize_person_key(p2)
    if not a or not b:
        return ""
    left, right = sorted([a, b])
    return f"{left} // {right}"


def _best_team_display(alias_pairs: list[str]) -> str:
    """
    Choose a stable best display label for team: shortest alias pair string.
    """
    cleaned = [s.strip() for s in alias_pairs if isinstance(s, str) and s.strip()]
    if not cleaned:
        return ""
    cleaned.sort(key=lambda x: (len(x), x.lower()))
    return cleaned[0]


def team_display_name(n1: str, n2: str) -> str:
    a = (n1 or "").strip()
    b = (n2 or "").strip()
    if not a or not b:
        return a or b
    return " / ".join(sorted([a, b], key=lambda x: x.lower()))


def _collapse_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def normalize_division_key(s: str) -> str:
    """
    Presentation-only key to group near-identical divisions for the Divisions_Normalized sheet.
    Conservative: remove punctuation + collapse whitespace + lowercase.
    Also normalizes trailing 'competition/comp' to reduce obvious redundancy
    (e.g., 'Freestyle Circle' vs 'Freestyle Circle Competition').
    """
    if not isinstance(s, str) or not s.strip():
        return ""
    t = s.lower().strip()
    # Replace punctuation with spaces (keeps words separated)
    trans = str.maketrans({ch: " " for ch in string.punctuation})
    t = t.translate(trans)
    t = _collapse_ws(t)
    # Common harmless suffix noise
    t = re.sub(r"\bcompetition\b$", "", t).strip()
    t = re.sub(r"\bcomp\b$", "", t).strip()
    t = _collapse_ws(t)
    return t


def pick_best_division_display(aliases: list[str]) -> str:
    """
    Choose a human-friendly label from observed aliases.
    Heuristic: prefer shorter, non-empty; preserve original capitalization.
    """
    cleaned = [a.strip() for a in (aliases or []) if isinstance(a, str) and a.strip()]
    if not cleaned:
        return ""
    # Prefer the shortest variant (often the canonical-looking one)
    cleaned.sort(key=lambda x: (len(x), x.lower()))
    return cleaned[0]


def year_to_sheet_name(y) -> str:
    if y is None:
        return "Unknown"
    try:
        # Handles "2001.0", 2001.0, "2001"
        yi = int(float(str(y).strip()))
        return str(yi)
    except Exception:
        return str(y)


def display_date(date_str: str, year) -> str:
    """
    Presentation-only: if date has no explicit YYYY and we *know* the sheet year,
    append ', YYYY' to improve consistency across the final workbook.
    """
    s = sanitize_string(date_str or "")
    if not s:
        return ""
    if year is None:
        return s
    # If a 4-digit year is already present, leave it alone
    if re.search(r"\b(19|20)\d{2}\b", s):
        return s
    try:
        y = int(year)
    except Exception:
        return s
    return f"{s}, {y}"


def is_team_division(division_name: str) -> bool:
    """
    Determine if a division is a team division based on division name.
    
    Returns True if division contains team indicators (doubles, pairs, team),
    False if it contains "singles" or is empty/None.
    """
    if not division_name:
        return False

    name = division_name.lower()

    # Explicit singles exclusions
    if "singles" in name:
        return False

    # Explicit team indicators
    if any(k in name for k in ["doubles", "pairs", "team"]):
        return True

    return False


# ------------------------------------------------------------
# Results formatting from placements
# ------------------------------------------------------------

# Category display order and labels
CATEGORY_ORDER = ["net", "freestyle", "golf", "sideline", "unknown"]
CATEGORY_LABELS = {
    "net": "NET",
    "freestyle": "FREESTYLE",
    "golf": "GOLF",
    "sideline": "OTHER",
    "unknown": "OTHER",
}


def format_results_from_placements(placements: list[dict], players_by_id: Optional[dict] = None) -> Optional[str]:
    """
    Build a deterministic, consistent results blob from canonical placements.
    Groups results by category (NET, FREESTYLE, GOLF, OTHER) with clear headers.

    Format:
      === NET ===
      OPEN SINGLES NET
      1. Name
      2. Name / Name

      === FREESTYLE ===
      SHRED 30
      1. Name

    We do NOT invent missing facts. If no placements exist -> None.
    """
    if not placements:
        return None

    # Group by category, then by division
    by_category = {}
    for p in placements:
        cat = p.get("division_category", "unknown") or "unknown"
        div = p.get("division_canon") or p.get("division_raw") or "Unknown"

        if cat not in by_category:
            by_category[cat] = {}
        if div not in by_category[cat]:
            by_category[cat][div] = []
        by_category[cat][div].append(p)

    out_lines = []

    # Output categories in defined order
    for cat in CATEGORY_ORDER:
        if cat not in by_category:
            continue

        divisions = by_category[cat]
        if not divisions:
            continue

        # Add category header
        label = CATEGORY_LABELS.get(cat, cat.upper())
        out_lines.append(f"<<< {label} >>>")
        out_lines.append("")

        # Sort divisions alphabetically within category
        for div in sorted(divisions.keys(), key=str.casefold):
            entries = divisions[div]

            # Sort entries by place, then by player name
            def sort_key(p):
                place = p.get("place", 999)
                try:
                    place = int(place)
                except (ValueError, TypeError):
                    place = 999
                name = _build_name_line(p, players_by_id)
                return (place, name.lower() if name else "")

            entries.sort(key=sort_key)

            out_lines.append(f"--- {div.upper()} ---")

            # Deduplicate: skip (place, name) combos already output (source data can have dupes)
            seen_line_key = set()
            for p in entries:
                place = p.get("place")
                try:
                    place_int = int(place)
                    place_txt = f"{place_int}."
                except (ValueError, TypeError):
                    place_txt = f"{place}." if place is not None else ""

                name = _build_name_line(p, players_by_id)
                line_key = (place_txt, (name or "").lower().strip())
                if line_key in seen_line_key:
                    continue
                seen_line_key.add(line_key)

                if place_txt:
                    out_lines.append(f"{place_txt} {name}".rstrip())
                else:
                    out_lines.append(name)

            out_lines.append("")  # blank line between divisions

    # Remove trailing blank lines
    while out_lines and out_lines[-1] == "":
        out_lines.pop()

    return "\n".join(out_lines) if out_lines else None


def _build_name_line(placement: dict, players_by_id: Optional[dict] = None) -> str:
    """Build display name from placement dict, preferring Stage 2.5 cleaned names."""

    def _lookup_clean(which: str) -> str:
        import re
        if which == "player1":
            pid = placement.get("player1_id") or placement.get("player_id") or placement.get("player1_player_id") or ""
            raw = (placement.get("player1_name") or "").strip()
        else:
            pid = placement.get("player2_id") or placement.get("player2_player_id") or ""
            raw = (placement.get("player2_name") or "").strip()
        if players_by_id and pid and pid in players_by_id:
            clean = (players_by_id[pid].get("player_name_clean") or "").strip()
            return clean or raw
        # Fallback (display-only): strip common slop when player_id is missing
        s = re.sub(r'^\s*(?:\*\-|\*|&)\s*', '', raw)
        s = re.sub(r'\s*-\s*scratch\b.*$', '', s, flags=re.IGNORECASE)
        s = re.sub(r'\s{2,}', ' ', s).strip(" ,.-")
        return s

    p1 = _lookup_clean("player1")
    p2 = _lookup_clean("player2")
    if p2:
        return f"{p1} / {p2}".strip()
    return p1


# ------------------------------------------------------------
# CSV reading
# ------------------------------------------------------------
def read_stage2_csv(csv_path: Path) -> list[dict]:
    """Read stage2 CSV and return list of event records."""
    # Increase CSV field size limit to handle large JSON fields
    csv.field_size_limit(min(2**31 - 1, 10 * 1024 * 1024))  # 10MB limit
    records = []
    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Convert year to int if present
            if row.get("year"):
                try:
                    row["year"] = int(row["year"])
                except ValueError:
                    row["year"] = None
            else:
                row["year"] = None

            # Parse placements JSON
            placements_json = row.get("placements_json", "[]")
            try:
                row["placements"] = json.loads(placements_json)
            except json.JSONDecodeError:
                row["placements"] = []

            records.append(row)
    return records


def build_players_by_id(players_df: Optional[pd.DataFrame]) -> dict:
    """Build lookup: player_id -> {player_name_clean, country_clean, name_status}."""
    if players_df is None or len(players_df) == 0:
        return {}
    dfp = players_df.copy()
    # Filter out junk rows so they cannot leak into Excel outputs
    if 'name_status' in dfp.columns:
        dfp = dfp[dfp['name_status'].isin(['ok','suspicious','needs_review'])].copy()
    out = {}
    for _, r in dfp.iterrows():
        pid = str(r.get('player_id') or '').strip()
        if not pid:
            continue
        out[pid] = {
            'player_name_clean': str(r.get('player_name_clean') or r.get('player_name_raw') or '').strip(),
            'country_clean': str(r.get('country_clean') or r.get('country_observed') or '').strip(),
            'name_status': str(r.get('name_status') or '').strip(),
        }
    return out


_WS = re.compile(r"\s+")


def _one_line(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    # collapse any \r \n \t etc into single spaces
    return _WS.sub(" ", s).strip()


def build_placements_flat_df(records: list[dict], players_by_id: dict) -> pd.DataFrame:
    """
    Flatten placements into a single truth-preserving table for analysis/QC.

    Output columns (stable):
      event_id, year,
      division_canon, division_raw, division_category,
      competitor_type,
      place,
      player1_id, player1_name, player2_id, player2_name,
      team_display_name
    """
    rows = []
    for rec in records:
        eid = str(rec.get("event_id") or "").strip()
        year = rec.get("year")
        placements = rec.get("placements", []) or []
        for p in placements:
            div_canon = (p.get("division_canon") or "").strip()
            div_raw = (p.get("division_raw") or "").strip()
            div_cat = (p.get("division_category") or "unknown") or "unknown"
            competitor_type = (p.get("competitor_type") or "").strip()

            # Place (keep original if non-numeric)
            place = p.get("place", "")

            # Prefer cleaned player display if we can
            def _display(pid_key: str, name_key: str) -> str:
                pid = str(p.get(pid_key) or "").strip()
                raw = str(p.get(name_key) or "").strip()
                if pid and pid in players_by_id:
                    clean = str(players_by_id[pid].get("player_name_clean") or "").strip()
                    return clean or raw
                return raw

            player1_id = str(p.get("player1_id") or p.get("player_id") or p.get("player1_player_id") or "").strip()
            player2_id = str(p.get("player2_id") or p.get("player2_player_id") or "").strip()

            player1_name = _display("player1_id", "player1_name") if p.get("player1_name") is not None else _display("player_id", "player1_name")
            # If original data uses player_name instead of player1_name, fall back:
            if not player1_name:
                player1_name = str(p.get("player_name") or "").strip()
                if player1_id and player1_id in players_by_id:
                    player1_name = (players_by_id[player1_id].get("player_name_clean") or "").strip() or player1_name

            player2_name = _display("player2_id", "player2_name")

            team_disp = team_display_name(player1_name, player2_name) if player2_name else player1_name

            # Sanitize: collapse any embedded newlines/tabs to single spaces (one line per field)
            player1_name = _one_line(player1_name)
            player2_name = _one_line(player2_name)
            team_disp = _one_line(team_disp)
            div_canon = _one_line(div_canon)
            div_raw = _one_line(div_raw)
            player1_id = _one_line(player1_id)
            player2_id = _one_line(player2_id)

            rows.append({
                "event_id": eid,
                "year": year if year is not None else "",
                "division_canon": div_canon,
                "division_raw": div_raw,
                "division_category": div_cat,
                "competitor_type": competitor_type,
                "place": place,
                "player1_id": player1_id,
                "player1_name": player1_name,
                "player2_id": player2_id,
                "player2_name": player2_name,
                "team_display_name": team_disp,
            })

    df = pd.DataFrame(rows)
    # Stable ordering: by year, event_id, division, place
    if not df.empty:
        def _place_sort(x):
            try:
                return int(x)
            except Exception:
                return 999999

        df["_place_sort"] = df["place"].apply(_place_sort)
        df.sort_values(
            by=["year", "event_id", "division_canon", "division_raw", "_place_sort", "team_display_name"],
            ascending=[True, True, True, True, True, True],
            inplace=True,
        )
        df.drop(columns=["_place_sort"], inplace=True)

    return df


def build_persons_truth(df_pf: pd.DataFrame) -> pd.DataFrame:
    """
    Build Persons_Truth table:
    - one row per effective_person_id
    - NO guessing
    - person_id from overrides OR fallback to player_id
    """
    if df_pf is None or (isinstance(df_pf, pd.DataFrame) and df_pf.empty):
        return pd.DataFrame()

    rows = []

    for _, r in df_pf.iterrows():
        pid = str(r.get("person_id") or "").strip()
        player_id = str(r.get("player_id") or r.get("player1_id") or "").strip()

        if pid:
            effective_person_id = pid
            source = "override"
        elif player_id:
            effective_person_id = player_id
            source = "fallback_player_id"
        else:
            continue  # impossible case

        rows.append({
            "effective_person_id": effective_person_id,
            "source": source,
            "player_id": player_id,
            "player_name_clean": str(r.get("player_name_clean") or "").strip(),
        })

    df = pd.DataFrame(rows)

    if df.empty:
        return df

    # Aggregate deterministically
    out = []
    for ep, g in df.groupby("effective_person_id"):
        names = sorted({n for n in g["player_name_clean"] if n})
        player_ids = sorted({p for p in g["player_id"] if p})

        out.append({
            "person_id": ep,
            "identity_source": g["source"].iloc[0],
            "player_ids_seen": " | ".join(player_ids),
            "player_names_seen": " | ".join(names),
            "player_id_count": len(player_ids),
            "name_variant_count": len(names),
        })

    return pd.DataFrame(out).sort_values(
        by=["identity_source", "name_variant_count"],
        ascending=[True, False],
    )


# ------------------------------------------------------------
# Excel writer
# ------------------------------------------------------------
def write_excel(
    out_xlsx: Path,
    records: list[dict],
    players_df: Optional[pd.DataFrame] = None,
    placements_flat_df: Optional[pd.DataFrame] = None,
) -> None:
    """
    Archive workbook writer (matches Footbag_Results_Canonical.xlsx layout):
    - One sheet per year named YYYY.0
    - Columns are event_id
    - Rows are fixed labels (Tournament Name, Date, Location, ...)
    - Results are generated from placements (canonical), not copied raw
    """
    players_by_id = build_players_by_id(players_df)

    # Build results map from placements (use cleaned player names when available)
    results_map = {}
    for rec in records:
        eid = rec.get("event_id")
        if eid:
            placements = rec.get("placements", [])
            results_map[str(eid)] = format_results_from_placements(placements, players_by_id)

    # Fixed row labels (index) to match the example workbook
    row_labels = [
        "Tournament Name",
        "Date",
        "Location",
        "Event Type",
        "Host Club",
        "Results",
    ]

    # Sort key for event IDs
    def _eid_sort_key(x: str):
        try:
            return int(re.sub(r"\D+", "", x) or "0")
        except Exception:
            return 0

    # Group records by year
    by_year = {}
    unknown_year = []
    for rec in records:
        year = rec.get("year")
        if year is not None:
            if year not in by_year:
                by_year[year] = []
            by_year[year].append(rec)
        else:
            unknown_year.append(rec)

    event_locator = {}  # event_id(str) -> (sheet_name(str), col_idx(int))

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xw:
        # Build one sheet per year
        for y in sorted(by_year.keys()):
            year_records = by_year[y]
            eids = sorted([str(r.get("event_id", "")) for r in year_records], key=_eid_sort_key)

            # Excel columns: A=1 is row labels, so first event column is B=2
            sheet_name = year_to_sheet_name(y)
            for j, eid in enumerate(eids, start=2):
                event_locator[str(eid)] = (sheet_name, j)

            data = {}
            for eid in eids:
                rec = next((r for r in year_records if str(r.get("event_id")) == eid), None)
                if not rec:
                    continue

                # Use integer event_id as column header to avoid Excel apostrophe prefix
                col_key = int(eid) if eid.isdigit() else eid
                data[col_key] = [
                    sanitize_string(rec.get("event_name") or ""),
                    display_date(rec.get("date") or "", y),
                    sanitize_string(rec.get("location") or ""),
                    sanitize_string(rec.get("event_type") or ""),
                    sanitize_string(rec.get("host_club") or ""),
                    sanitize_string(results_map.get(eid) or ""),
                ]

            df_year = pd.DataFrame(data, index=row_labels)
            df_year.index.name = "event_id"  # puts "event_id" in A1 like the example

            df_year = sanitize_excel_strings(df_year)
            df_year.to_excel(xw, sheet_name=sheet_name)

            # Apply wrap_text formatting to Results row (row 7)
            worksheet = xw.sheets[sheet_name]
            for col_idx in range(2, len(eids) + 2):  # Start from column B (2)
                cell = worksheet.cell(row=7, column=col_idx)
                a = copy(cell.alignment)
                a.wrap_text = True
                cell.alignment = a

        # Unknown-year sheet
        if unknown_year:
            eids = sorted([str(r.get("event_id", "")) for r in unknown_year], key=_eid_sort_key)
            # Excel columns: A=1 is row labels, so first event column is B=2
            for j, eid in enumerate(eids, start=2):
                event_locator[str(eid)] = ("unknown_year", j)
            data = {}
            for eid in eids:
                rec = next((r for r in unknown_year if str(r.get("event_id")) == eid), None)
                if not rec:
                    continue

                # Use integer event_id as column header to avoid Excel apostrophe prefix
                col_key = int(eid) if eid.isdigit() else eid
                data[col_key] = [
                    sanitize_string(rec.get("event_name") or ""),
                    display_date(rec.get("date") or "", None),
                    sanitize_string(rec.get("location") or ""),
                    sanitize_string(rec.get("event_type") or ""),
                    sanitize_string(rec.get("host_club") or ""),
                    sanitize_string(results_map.get(eid) or ""),
                ]

            df_unk = pd.DataFrame(data, index=row_labels)
            df_unk.index.name = "event_id"
            df_unk = sanitize_excel_strings(df_unk)
            df_unk.to_excel(xw, sheet_name="unknown_year")

            # Apply wrap_text formatting to Results row
            worksheet = xw.sheets["unknown_year"]
            for col_idx in range(2, len(eids) + 2):
                cell = worksheet.cell(row=7, column=col_idx)
                a = copy(cell.alignment)
                a.wrap_text = True
                cell.alignment = a

        # Build Index sheet: one row per event with hyperlinks
        index_data = []
        for rec in records:
            eid = str(rec.get("event_id", ""))
            if not eid:
                continue
            
            year = rec.get("year")
            placements = rec.get("placements", [])
            results_text = results_map.get(eid, "")
            results_lines = len(results_text.splitlines()) if results_text else 0
            
            index_data.append({
                "event_id": eid,
                "year": year if year is not None else "",
                "Tournament Name": sanitize_string(rec.get("event_name") or ""),
                "Date": sanitize_string(rec.get("date") or ""),
                "Location": sanitize_string(rec.get("location") or ""),
                "Event Type": sanitize_string(rec.get("event_type") or ""),
                "Host Club": sanitize_string(rec.get("host_club") or ""),
                "placements_count": len(placements),
                "results_lines": results_lines,
            })
        
        # Sort by year, then event_id
        index_data.sort(key=lambda x: (x["year"] if x["year"] != "" else 9999, _eid_sort_key(x["event_id"])))
        
        df_index = pd.DataFrame(index_data)
        df_index = sanitize_excel_strings(df_index)
        df_index.to_excel(xw, sheet_name="Index", index=False)
        
        # Apply formatting to Index sheet: hyperlinks, filters, freeze panes
        index_ws = xw.sheets["Index"]
        
        # Add hyperlinks to event_id column (column A, starting at row 2)
        hyperlink_font = Font(color="0563C1", underline="single")  # Blue, underlined
        for idx, row_data in enumerate(index_data, start=2):
            eid = row_data["event_id"]
            if eid in event_locator:
                sheet_name, col_idx = event_locator[eid]
                col_letter = get_column_letter(col_idx)
                # Hyperlink format: #SheetName!ColumnLetter1 (e.g., #1999!B1)
                hyperlink = f"#{sheet_name}!{col_letter}1"
                cell = index_ws.cell(row=idx, column=1)  # Column A
                cell.hyperlink = hyperlink
                cell.font = hyperlink_font
        
        # Freeze first row (header)
        index_ws.freeze_panes = "A2"
        
        # Add auto filter to header row
        index_ws.auto_filter.ref = index_ws.dimensions

        # Build Summary sheet: rollups and health metrics
        total_events = len(records)
        total_placements = sum(len(rec.get("placements", [])) for rec in records)
        
        years_with_events = [rec.get("year") for rec in records if rec.get("year") is not None]
        year_min = min(years_with_events) if years_with_events else None
        year_max = max(years_with_events) if years_with_events else None
        
        # Year table: year, events, placements
        year_stats = defaultdict(lambda: {"events": 0, "placements": 0})
        for rec in records:
            year = rec.get("year")
            if year is not None:
                year_stats[year]["events"] += 1
                year_stats[year]["placements"] += len(rec.get("placements", []))
        
        summary_data = [
            {"Metric": "Total Events", "Value": total_events},
            {"Metric": "Total Placements", "Value": total_placements},
            {"Metric": "Year Min", "Value": year_min if year_min else ""},
            {"Metric": "Year Max", "Value": year_max if year_max else ""},
        ]
        
        df_summary_metrics = pd.DataFrame(summary_data)
        df_summary_metrics = sanitize_excel_strings(df_summary_metrics)
        
        year_table_data = [
            {"year": year, "events": stats["events"], "placements": stats["placements"]}
            for year, stats in sorted(year_stats.items())
        ]
        df_year_table = pd.DataFrame(year_table_data)
        df_year_table = sanitize_excel_strings(df_year_table)
        
        # Write Summary sheet with metrics and year table
        df_summary_metrics.to_excel(xw, sheet_name="Summary", index=False, startrow=0)
        df_year_table.to_excel(xw, sheet_name="Summary", index=False, startrow=len(summary_data) + 3)
        
        summary_ws = xw.sheets["Summary"]
        summary_ws.freeze_panes = None
        summary_ws.auto_filter.ref = None

        # ------------------------------------------------------------
        # Players sheets (prefer Stage 2.5 cleaned player tokens)
        # ------------------------------------------------------------
        if players_df is not None and len(players_df) > 0:
            dfp = players_df.copy()

            # Ensure required columns exist (defensive)
            for col in [
                'player_id','player_name_raw','country_observed','player_name_clean',
                'name_status','junk_reason','country_clean','usage_count','source_hint','name_key'
            ]:
                if col not in dfp.columns:
                    dfp[col] = ''

            # Main Players sheet: one row per player_id (truth-preserving, duplicates allowed)
            df_players = dfp[dfp['name_status'].isin(['ok','suspicious','needs_review'])].copy()
            df_players = df_players[['player_id','player_name_clean','country_clean','name_status']]
            df_players = sanitize_excel_strings(df_players)
            df_players.to_excel(xw, sheet_name='Players', index=False)

            # Players_Clean: ok + suspicious (plus audit columns)
            df_clean = dfp[dfp['name_status'].isin(['ok','suspicious'])].copy()
            df_clean = df_clean[['player_id','player_name_clean','country_clean','name_status',
                                 'player_name_raw','country_observed','usage_count','source_hint','name_key']]
            df_clean = sanitize_excel_strings(df_clean)
            df_clean.to_excel(xw, sheet_name='Players_Clean', index=False)

            # Players_Junk: junk only (audit)
            df_junk = dfp[dfp['name_status'] == 'junk'].copy()
            df_junk = df_junk[['player_id','player_name_raw','junk_reason','usage_count','source_hint','name_key']]
            df_junk = sanitize_excel_strings(df_junk)
            df_junk.to_excel(xw, sheet_name='Players_Junk', index=False)

            # ------------------------------------------------------------
            # Players_Alias_Candidates (presentation-only; NO merges)
            # Groups name variants that currently map to different player_id.
            # ------------------------------------------------------------
            df_nonjunk = dfp[dfp['name_status'].isin(['ok','suspicious','needs_review'])].copy()

            # Build grouping key:
            # - prefer provided name_key (Stage 2.5)
            # - fallback to normalized player_name_clean or raw
            def _row_group_key(r):
                nk = r.get('name_key')
                if isinstance(nk, str) and nk.strip():
                    return nk.strip()
                nm = r.get('player_name_clean')
                if not (isinstance(nm, str) and nm.strip()):
                    nm = r.get('player_name_raw')
                return normalize_person_key(nm)

            df_nonjunk['alias_group_key'] = df_nonjunk.apply(_row_group_key, axis=1)
            df_nonjunk['alias_group_key'] = df_nonjunk['alias_group_key'].fillna('').astype(str)

            # Only keep meaningful keys
            df_nonjunk = df_nonjunk[df_nonjunk['alias_group_key'].str.len() > 0]

            candidates_rows = []
            for gkey, g in df_nonjunk.groupby('alias_group_key'):
                # Only interesting if multiple distinct player_ids (your stated problem)
                player_ids = sorted(set([str(x) for x in g['player_id'].dropna().astype(str).tolist() if str(x).strip()]))
                if len(player_ids) < 2:
                    continue

                clean_names = g['player_name_clean'].dropna().astype(str).tolist()
                raw_names = g['player_name_raw'].dropna().astype(str).tolist()
                # Prefer clean names for display; keep raw names as additional aliases if different
                aliases = []
                for n in clean_names:
                    if isinstance(n, str) and n.strip():
                        aliases.append(n.strip())
                for n in raw_names:
                    if isinstance(n, str) and n.strip():
                        aliases.append(n.strip())
                # unique while preserving stable order
                seen = set()
                aliases_u = []
                for a in aliases:
                    if a.lower() not in seen:
                        seen.add(a.lower())
                        aliases_u.append(a)

                display_best = _best_display_name([a for a in aliases_u if a])
                conf = _alias_confidence(aliases_u, normalize_person_key(display_best) or gkey)

                usage = 0
                if 'usage_count' in g.columns:
                    try:
                        usage = int(pd.to_numeric(g['usage_count'], errors='coerce').fillna(0).sum())
                    except Exception:
                        usage = 0

                # country rollup (optional)
                countries = []
                if 'country_clean' in g.columns:
                    countries = sorted(set([str(x).strip() for x in g['country_clean'].dropna().astype(str).tolist() if str(x).strip()]))

                candidates_rows.append({
                    "candidate_group_id": _stable_group_id("p", gkey),
                    "name_key": gkey,
                    "display_name_best": display_best,
                    "aliases": " | ".join(aliases_u),
                    "player_ids": " | ".join(player_ids),
                    "countries_seen": " | ".join(countries),
                    "usage_count_total": usage,
                    "confidence": conf,
                    "review_priority": (
                        "high" if conf == "high" and usage >= 10
                        else "med" if conf in ("high", "med")
                        else "low"
                    ),
                    "merge_decision": "",   # merge / not_merge / unsure
                    "notes": "",
                })

            if candidates_rows:
                # High usage first, then confidence, then name
                conf_rank = {"high": 0, "med": 1, "low": 2}
                candidates_rows.sort(key=lambda r: (-int(r.get("usage_count_total") or 0),
                                                   conf_rank.get(r.get("confidence"), 9),
                                                   (r.get("display_name_best") or "").lower()))
                df_cand = pd.DataFrame(candidates_rows)
                df_cand = sanitize_excel_strings(df_cand)
                df_cand.to_excel(xw, sheet_name="Players_Alias_Candidates", index=False)

                ws = xw.sheets["Players_Alias_Candidates"]
                for col_name in ("player_ids", "name_key"):
                    if col_name in df_cand.columns:
                        idx = df_cand.columns.get_loc(col_name) + 1
                        ws.column_dimensions[get_column_letter(idx)].hidden = True
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

        else:
            # Fallback (legacy): derive Players from placements (may include slop)
            # Deterministic improvement: if Placements_Flat has player_name_clean, prefer it for display.
            clean_name_by_player_id: dict[str, str] = {}
            if placements_flat_df is not None and isinstance(placements_flat_df, pd.DataFrame) and not placements_flat_df.empty:
                # Support common id/name column spellings; do NOT invent columns if missing.
                id_col = None
                name_col = None
                for c in ("player_id", "player1_id"):
                    if c in placements_flat_df.columns:
                        id_col = c
                        break
                for c in ("player_name_clean", "player1_name_clean"):
                    if c in placements_flat_df.columns:
                        name_col = c
                        break
                if id_col and name_col:
                    for _pid, _nm in zip(placements_flat_df[id_col].astype(str), placements_flat_df[name_col].astype(str)):
                        _pid = (_pid or "").strip()
                        _nm = (_nm or "").strip()
                        if _pid and _nm and _pid not in clean_name_by_player_id:
                            clean_name_by_player_id[_pid] = _nm

            players_map = {}  # player_id or name -> {name, country}
            for rec in records:
                placements = rec.get('placements', [])
                for p in placements:
                    player_id = (p.get('player1_id') or p.get('player_id') or p.get('player1_player_id') or '')
                    player_name = (p.get('player1_name') or '').strip()
                    if not player_name:
                        continue
                    country = (p.get('home_country') or p.get('country') or p.get('nation') or p.get('player1_country') or p.get('player1_home_country') or '')
                    country = country.strip() if country else ''
                    key = player_id if player_id else player_name.lower()
                    if key not in players_map:
                        players_map[key] = {'player_id': player_id, 'player_name': player_name, 'home_country': country}

            # Prefer clean names for display when available
            for key, val in players_map.items():
                pid = val.get("player_id") or ""
                if pid and pid in clean_name_by_player_id:
                    val["player_name"] = clean_name_by_player_id[pid]

            players_data = list(players_map.values())
            players_data.sort(key=lambda x: (x.get('player_name') or '').lower())
            df_players = pd.DataFrame(players_data)
            df_players = sanitize_excel_strings(df_players)
            df_players.to_excel(xw, sheet_name='Players', index=False)

        # Format Players-related sheets
        for sheet_name in ['Players','Players_Clean','Players_Junk']:
            if sheet_name in xw.sheets:
                ws = xw.sheets[sheet_name]
                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = ws.dimensions

        # Build Divisions sheet: one row per division_canon
        divisions_map = defaultdict(lambda: {"placements": 0, "events": set()})
        
        for rec in records:
            eid = rec.get("event_id")
            placements = rec.get("placements", [])
            for p in placements:
                div_canon = p.get("division_canon", "")
                div_category = p.get("division_category", "unknown")
                if div_canon:
                    divisions_map[div_canon]["placements"] += 1
                    divisions_map[div_canon]["events"].add(eid)
                    # Store category (should be consistent per division, but take first seen)
                    if "category" not in divisions_map[div_canon]:
                        divisions_map[div_canon]["category"] = div_category
        
        divisions_data = [
            {
                "division_canon": div,
                "division_category": divisions_map[div].get("category", "unknown"),
                "count_placements": divisions_map[div]["placements"],
                "count_events": len(divisions_map[div]["events"]),
            }
            for div in sorted(divisions_map.keys())
        ]
        
        df_divisions = pd.DataFrame(divisions_data)
        df_divisions = sanitize_excel_strings(df_divisions)
        df_divisions.to_excel(xw, sheet_name="Divisions", index=False)
        
        divisions_ws = xw.sheets["Divisions"]
        divisions_ws.freeze_panes = "A2"
        divisions_ws.auto_filter.ref = divisions_ws.dimensions

        # ------------------------------------------------------------
        # Placements_Flat sheet (truth-preserving analysis table)
        # ------------------------------------------------------------
        if placements_flat_df is not None and len(placements_flat_df) > 0:
            df_pf = sanitize_excel_strings(placements_flat_df.copy())
            df_pf.to_excel(xw, sheet_name="Placements_Flat", index=False)
            ws = xw.sheets["Placements_Flat"]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # ------------------------------------------------------------
        # IMPORTANT:
        # Stage 04 owns the final strict Option-A Persons_Truth (presentation/pivot-ready).
        # To avoid confusion/overwrites, Stage 03 writes its source/tracing version under
        # a distinct sheet name (no data changes).
        # ------------------------------------------------------------
        persons_truth_df = build_persons_truth(placements_flat_df)
        if persons_truth_df is not None and not persons_truth_df.empty:
            persons_truth_df = sanitize_excel_strings(persons_truth_df)
            persons_truth_df.to_excel(xw, sheet_name="Persons_Truth_Source", index=False)

            ws = xw.sheets["Persons_Truth_Source"]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # ------------------------------------------------------------
        # Divisions_Normalized sheet (presentation-only grouping)
        # ------------------------------------------------------------
        norm_map = defaultdict(lambda: {
            "placements": 0,
            "events": set(),
            "categories": set(),
            "aliases": set(),
        })

        for rec in records:
            eid = rec.get("event_id")
            placements = rec.get("placements", [])
            for p in placements:
                div_canon = (p.get("division_canon") or "").strip()
                div_raw = (p.get("division_raw") or "").strip()
                div_cat = (p.get("division_category") or "unknown").strip() or "unknown"

                # Use canon if present; also record raw as alias
                base = div_canon or div_raw
                if not base:
                    continue

                key = normalize_division_key(base)
                if not key:
                    continue

                nm = norm_map[key]
                nm["placements"] += 1
                if eid:
                    nm["events"].add(eid)
                if div_cat:
                    nm["categories"].add(div_cat)

                if div_canon:
                    nm["aliases"].add(div_canon)
                if div_raw:
                    nm["aliases"].add(div_raw)

        divisions_norm_data = []
        for key, nm in norm_map.items():
            aliases = sorted(nm["aliases"], key=lambda x: (len(x), x.lower()))
            # If multiple categories seen, keep 'mixed' to be honest
            cats = sorted([c for c in nm["categories"] if c])
            cat = cats[0] if len(cats) == 1 else ("mixed" if cats else "unknown")

            years_seen = sorted(
                {rec.get("year") for rec in records if rec.get("event_id") in nm["events"] and rec.get("year") is not None}
            )

            divisions_norm_data.append({
                "division_key": key,
                "division_display": pick_best_division_display(aliases),
                "preferred_display_name": "",  # manual override
                "division_category": cat,
                "count_placements": nm["placements"],
                "count_events": len(nm["events"]),
                "first_year_seen": years_seen[0] if years_seen else "",
                "last_year_seen": years_seen[-1] if years_seen else "",
                "aliases": " | ".join(aliases),
            })

        if divisions_norm_data:
            # Highest frequency first
            divisions_norm_data.sort(key=lambda r: (-int(r["count_placements"]), r["division_display"].lower()))
            df_div_norm = pd.DataFrame(divisions_norm_data)
            df_div_norm = sanitize_excel_strings(df_div_norm)
            df_div_norm.to_excel(xw, sheet_name="Divisions_Normalized", index=False)

            ws = xw.sheets["Divisions_Normalized"]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # Build Teams sheet: team placements only
        teams_data = []
        for rec in records:
            eid = rec.get("event_id")
            year = rec.get("year")
            placements = rec.get("placements", [])
            for p in placements:
                if p.get("competitor_type") == "team":
                    division_name = p.get("division_canon", "") or p.get("division_raw", "")
                    if not is_team_division(division_name):
                        continue  # skip singles
                    p1 = sanitize_string(p.get("player1_name", ""))
                    p2 = sanitize_string(p.get("player2_name", ""))

                    teams_data.append({
                        "event_id": eid,
                        "year": year if year is not None else "",
                        "division_canon": sanitize_string(p.get("division_canon", "")),
                        "place": p.get("place", ""),
                        "team_display_name": team_display_name(p1, p2),
                        "player1_name": p1,
                        "player2_name": p2,
                        "player1_id": sanitize_string(p.get("player1_id") or p.get("player_id") or ""),
                        "player2_id": sanitize_string(p.get("player2_id") or ""),
                    })
        
        if teams_data:
            df_teams = pd.DataFrame(teams_data)
            df_teams = sanitize_excel_strings(df_teams)
            df_teams.to_excel(xw, sheet_name="Teams", index=False)
            
            teams_ws = xw.sheets["Teams"]
            teams_ws.freeze_panes = "A2"
            teams_ws.auto_filter.ref = teams_ws.dimensions

            # ------------------------------------------------------------
            # Teams_Alias_Candidates (presentation-only; NO merges)
            # Group by normalized member-name key (order-invariant).
            # ------------------------------------------------------------
            df_t = pd.DataFrame(teams_data).copy()
            for col in ["player1_name","player2_name","player1_id","player2_id","division_canon","event_id","year","place"]:
                if col not in df_t.columns:
                    df_t[col] = ""

            df_t["team_key"] = df_t.apply(lambda r: normalize_team_key(r.get("player1_name",""), r.get("player2_name","")), axis=1)
            df_t = df_t[df_t["team_key"].astype(str).str.len() > 0]

            cand_rows = []
            for tkey, g in df_t.groupby("team_key"):
                # Build alias strings like "Name1 / Name2" as seen in data
                alias_pairs = []
                for _, r in g.iterrows():
                    n1 = str(r.get("player1_name","") or "").strip()
                    n2 = str(r.get("player2_name","") or "").strip()
                    if n1 and n2:
                        alias_pairs.append(f"{n1} / {n2}")

                # Unique aliases (case-insensitive)
                seen = set()
                alias_u = []
                for a in alias_pairs:
                    k = a.lower().strip()
                    if k and k not in seen:
                        seen.add(k)
                        alias_u.append(a.strip())

                # ID pairs observed (order-invariant)
                id_pairs = set()
                for _, r in g.iterrows():
                    i1 = str(r.get("player1_id","") or "").strip()
                    i2 = str(r.get("player2_id","") or "").strip()
                    if i1 and i2:
                        left, right = sorted([i1, i2])
                        id_pairs.add(f"{left} | {right}")
                id_pairs = sorted(id_pairs)

                # Only interesting if there is some variation:
                # - more than 1 alias spelling/order OR
                # - more than 1 distinct id-pair (signals inconsistent player IDs)
                if len(alias_u) < 2 and len(id_pairs) < 2:
                    continue

                divisions = sorted(set([str(x).strip() for x in g["division_canon"].dropna().astype(str).tolist() if str(x).strip()]))
                years = sorted(set([str(x).strip() for x in g["year"].dropna().astype(str).tolist() if str(x).strip()]))

                cand_rows.append({
                    "candidate_group_id": _stable_group_id("t", tkey),
                    "team_key": tkey,
                    "team_display_best": _best_team_display(alias_u),
                    "aliases": " | ".join(alias_u),
                    "player_id_pairs": " || ".join(id_pairs),   # keep for traceability (can hide in Excel)
                    "divisions_seen": " | ".join(divisions),
                    "years_seen": " | ".join(years),
                    "count_placements": int(len(g)),
                    "confidence": "med" if len(alias_u) >= 2 else "low",
                    "decision": "",  # merge / not_merge / unsure
                    "notes": "",
                })

            if cand_rows:
                cand_rows.sort(key=lambda r: (-int(r.get("count_placements") or 0),
                                             (r.get("team_display_best") or "").lower()))
                df_tc = pd.DataFrame(cand_rows)
                df_tc = sanitize_excel_strings(df_tc)
                df_tc.to_excel(xw, sheet_name="Teams_Alias_Candidates", index=False)

                ws = xw.sheets["Teams_Alias_Candidates"]
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

        # Build QC_TopIssues sheet: from QC output files
        qc_summary_path = Path(__file__).resolve().parent / "out" / "stage3_qc_summary.json"
        qc_issues_path = Path(__file__).resolve().parent / "out" / "stage3_qc_issues.jsonl"
        
        if qc_summary_path.exists() and qc_issues_path.exists():
            try:
                import json
                
                # Read summary
                with open(qc_summary_path, "r", encoding="utf-8") as f:
                    qc_summary = json.load(f)
                
                # Build counts by check_id table
                counts_by_check = qc_summary.get("counts_by_check", {})
                qc_counts_data = []
                for check_id, counts in counts_by_check.items():
                    qc_counts_data.append({
                        "check_id": check_id,
                        "ERROR": counts.get("ERROR", 0),
                        "WARN": counts.get("WARN", 0),
                        "INFO": counts.get("INFO", 0),
                        "Total": counts.get("ERROR", 0) + counts.get("WARN", 0) + counts.get("INFO", 0),
                    })
                qc_counts_data.sort(key=lambda x: x["Total"], reverse=True)
                
                # Read sample issues
                qc_issues_data = []
                with open(qc_issues_path, "r", encoding="utf-8") as f:
                    for line_num, line in enumerate(f, 1):
                        if line_num > 100:  # Limit to first 100 issues
                            break
                        try:
                            issue = json.loads(line.strip())
                            qc_issues_data.append({
                                "event_id": issue.get("event_id", ""),
                                "check_id": issue.get("check_id", ""),
                                "severity": issue.get("severity", ""),
                                "field": issue.get("field", ""),
                                "message": sanitize_string(issue.get("message", ""))[:100],
                                "example_value": sanitize_string(str(issue.get("example_value", "")))[:50],
                                "context": sanitize_string(str(issue.get("context", "")))[:100],
                            })
                        except json.JSONDecodeError:
                            continue
                
                # Write QC sheet
                if qc_counts_data:
                    df_qc_counts = pd.DataFrame(qc_counts_data)
                    df_qc_counts = sanitize_excel_strings(df_qc_counts)
                    df_qc_counts.to_excel(xw, sheet_name="QC_TopIssues", index=False, startrow=0)
                
                if qc_issues_data:
                    df_qc_issues = pd.DataFrame(qc_issues_data)
                    df_qc_issues = sanitize_excel_strings(df_qc_issues)
                    start_row = len(qc_counts_data) + 3 if qc_counts_data else 0
                    df_qc_issues.to_excel(xw, sheet_name="QC_TopIssues", index=False, startrow=start_row)
                
                qc_ws = xw.sheets.get("QC_TopIssues")
                if qc_ws:
                    qc_ws.freeze_panes = "A2"
                    qc_ws.auto_filter.ref = qc_ws.dimensions
            except Exception as e:
                # Silently skip QC sheet if files don't exist or are malformed
                pass

        # ------------------------------------------------------------
        # README sheet (one-page explanation of workbook structure)
        # ------------------------------------------------------------
        readme_rows = [
            {
                "Section": "Purpose",
                "Description": (
                    "This workbook is an archive-quality canonical dataset derived from "
                    "historical footbag event results. It preserves original data while "
                    "providing normalized and review-friendly views for analysis and curation."
                ),
            },
            {
                "Section": "Canonical Sheets",
                "Description": (
                    "These sheets represent the canonical truth tables produced by the pipeline. "
                    "They preserve provenance and should not be edited manually:\n"
                    "- Events / Placements / Players / Teams\n"
                    "- Year-specific result sheets"
                ),
            },
            {
                "Section": "Normalized Sheets",
                "Description": (
                    "These sheets group near-identical labels for readability and analysis, "
                    "without changing canonical truth:\n"
                    "- Divisions_Normalized"
                ),
            },
            {
                "Section": "Alias Candidate Sheets",
                "Description": (
                    "These sheets identify potential duplicates caused by spelling variants, "
                    "encoding differences, or historical inconsistencies. "
                    "No merges are applied automatically:\n"
                    "- Players_Alias_Candidates\n"
                    "- Teams_Alias_Candidates\n\n"
                    "Use the confidence and usage columns to prioritize review."
                ),
            },
            {
                "Section": "Important Notes",
                "Description": (
                    "- No identity merges are performed automatically.\n"
                    "- Internal IDs may be present but hidden for traceability.\n"
                    "- Presentation sheets may change over time; canonical sheets should remain stable."
                ),
            },
        ]

        df_readme = pd.DataFrame(readme_rows)
        df_readme = sanitize_excel_strings(df_readme)
        df_readme.to_excel(xw, sheet_name="README", index=False)

        ws = xw.sheets["README"]
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 120
        ws.auto_filter.ref = ws.dimensions

    # Save event_locator map for downstream hyperlinks (Stage 04)
    locator_path = Path(out_xlsx).parent / "out" / "event_locator.json"
    locator_path.parent.mkdir(exist_ok=True)
    with open(locator_path, "w", encoding="utf-8") as f:
        json.dump(event_locator, f)


def run_stage3_qc(
    records: list[dict], results_map: dict, out_dir: Path, players_by_id: dict = None
) -> None:
    """Run Stage 3 QC checks on Excel workbook data and write outputs (fallback when qc_master unavailable)."""
    if run_slop_detection_checks_stage3_excel is None:
        return
    print("\n" + "="*60)
    print("Running Stage 3 QC: Excel Cell Scanning")
    print("="*60)

    issues = run_slop_detection_checks_stage3_excel(
        records, results_map, players_by_id=players_by_id
    )

    # Build summary
    counts_by_check = defaultdict(lambda: {"ERROR": 0, "WARN": 0, "INFO": 0})
    for issue in issues:
        issue_dict = issue.to_dict() if hasattr(issue, 'to_dict') else issue
        counts_by_check[issue_dict["check_id"]][issue_dict["severity"]] += 1

    total_errors = sum(1 for i in issues if (i.to_dict() if hasattr(i, 'to_dict') else i)["severity"] == "ERROR")
    total_warnings = sum(1 for i in issues if (i.to_dict() if hasattr(i, 'to_dict') else i)["severity"] == "WARN")
    total_info = sum(1 for i in issues if (i.to_dict() if hasattr(i, 'to_dict') else i)["severity"] == "INFO")

    summary = {
        "stage": "stage3",
        "total_events": len(records),
        "total_errors": total_errors,
        "total_warnings": total_warnings,
        "total_info": total_info,
        "counts_by_check": dict(counts_by_check),
    }

    # Write Stage 3 QC outputs
    summary_path = out_dir / "stage3_qc_summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    print(f"Wrote: {summary_path}")

    issues_path = out_dir / "stage3_qc_issues.jsonl"
    with open(issues_path, "w", encoding="utf-8") as f:
        for issue in issues:
            issue_dict = issue.to_dict() if hasattr(issue, 'to_dict') else issue
            f.write(json.dumps(issue_dict, ensure_ascii=False) + "\n")
    print(f"Wrote: {issues_path} ({len(issues)} issues)")

    # Print summary
    print(f"\nStage 3 QC Results:")
    print(f"  Total issues: {len(issues)}")
    print(f"  Errors: {total_errors}")
    print(f"  Warnings: {total_warnings}")
    print(f"  Info: {total_info}")

    if counts_by_check:
        print(f"\nIssues by check:")
        for check_id in sorted(counts_by_check.keys()):
            counts = counts_by_check[check_id]
            err = counts.get("ERROR", 0)
            warn = counts.get("WARN", 0)
            info = counts.get("INFO", 0)
            print(f"  {check_id}: {err} errors, {warn} warnings, {info} info")

    print("="*60)


def print_verification_stats(records: list[dict], out_xlsx: Path) -> None:
    """Print verification gate statistics."""
    total = len(records)
    print(f"\n{'='*60}")
    print("VERIFICATION GATE: Stage 3 (Excel Output)")
    print(f"{'='*60}")
    print(f"Total events in output: {total}")

    if total == 0:
        return

    # Count by year
    by_year = {}
    unknown = 0
    for rec in records:
        year = rec.get("year")
        if year is not None:
            by_year[year] = by_year.get(year, 0) + 1
        else:
            unknown += 1

    years = sorted(by_year.keys())
    print(f"\nSheet count: {len(years)} year sheets" + (", 1 unknown_year sheet" if unknown else ""))

    if years:
        print(f"Year range: {min(years)} - {max(years)}")

    print("\nEvents per sheet (first 10):")
    for y in years[:10]:
        print(f"  {int(y)}.0: {by_year[y]} events")
    if len(years) > 10:
        print(f"  ... and {len(years) - 10} more year sheets")
    if unknown:
        print(f"  unknown_year: {unknown} events")

    # Spot check 10 events
    print("\nSpot check (10 sample events):")
    import random
    sample = random.sample(records, min(10, len(records)))
    for rec in sample:
        eid = rec.get("event_id")
        year = rec.get("year")
        name = str(rec.get("event_name", ""))[:30]
        placements = len(rec.get("placements", []))
        print(f"  {eid:6s} | {year or '????'} | {name:30s} | {placements} placements")

    print(f"\nOutput file: {out_xlsx}")
    print(f"{'='*60}\n")


def main():
    """
    Read stage2 CSV and output final Excel workbook.
    """
    repo_dir = Path(__file__).resolve().parent
    out_dir = repo_dir / "out"
    in_csv = out_dir / "stage2_canonical_events.csv"
    out_xlsx = repo_dir / "Footbag_Results_Canonical.xlsx"

    players_csv = out_dir / "stage2p5_players_clean.csv"
    players_df = None
    if players_csv.exists():
        players_df = pd.read_csv(players_csv)
    else:
        print(f"Warning: Stage 2.5 players file not found: {players_csv} (falling back to placement-derived Players)")

    if not in_csv.exists():
        print(f"ERROR: Input file not found: {in_csv}")
        print("Run 02_canonicalize_results.py first.")
        return

    print(f"Reading: {in_csv}")
    records = read_stage2_csv(in_csv)

    # Build results_map for Stage 3 QC (use cleaned player names when available)
    players_by_id = build_players_by_id(players_df)
    results_map = {}
    for rec in records:
        eid = rec.get("event_id")
        if eid:
            placements = rec.get("placements", [])
            results_map[str(eid)] = format_results_from_placements(placements, players_by_id)

    # Placements_Flat must exist (person_id-enriched from 02p5); do not overwrite
    placements_flat_csv = out_dir / "Placements_Flat.csv"
    if not placements_flat_csv.exists():
        print(f"ERROR: Missing required {placements_flat_csv}. Run 02p5 to generate it.", file=sys.stderr)
        return

    df_placements_flat = pd.read_csv(placements_flat_csv)
    print(f"Loaded {placements_flat_csv} ({len(df_placements_flat)} rows)")

    print(f"Writing Excel with {len(records)} events...")
    write_excel(out_xlsx, records, players_df=players_df, placements_flat_df=df_placements_flat)

    # Run Stage 3 QC on Excel workbook data
    if USE_MASTER_QC:
        qc_summary, qc_issues = run_qc_for_stage(
            "stage3", records, results_map=results_map, players_by_id=players_by_id, out_dir=out_dir
        )
        print_qc_summary(qc_summary, "stage3")
    else:
        print("Skipping Stage 3 QC (qc_master not available)")

    print_verification_stats(records, out_xlsx)
    print(f"Wrote: {out_xlsx}")


if __name__ == "__main__":
    main()
