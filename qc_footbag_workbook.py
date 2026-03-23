#!/usr/bin/env python3
"""
qc_footbag_workbook.py

Run structural QC checks on the Footbag community workbook.

Checks included:
1. Singles/doubles team-size integrity (GROUP-BASED, not row-based)
2. Split doubles pair bug
3. Ambiguous doubles placements (>2 players at same placement)
4. Missing delimiter in doubles rows
5. Duplicate player within same placement
6. Duplicate player within same event/discipline
7. Placement sequence / tie integrity
8. Orphan rows (placement without player, player without placement)
9. Non-person / metadata leakage in results rows
10. Near-duplicate name detection within same event

Usage:
    python qc_footbag_workbook.py /path/to/Footbag_Results_Community_FINAL_v13.xlsx

Outputs:
    qc_issues.csv
    qc_summary_by_event.csv
"""

from __future__ import annotations

import sys
import re
from pathlib import Path
from difflib import SequenceMatcher
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook


# -----------------------------
# Config
# -----------------------------

YEAR_SHEET_RE = re.compile(r"^\d{4}$")

NON_PERSON_PATTERNS = [
    r"\[non-person\]",
    r"contact:",
    r"location:",
    r"home page:",
    r"site\(s\)",
    r"owner:",
    r"copyright",
    r"results:",
    r"events offered:",
]

DOUBLES_KEYWORDS = ["doubles"]
SINGLES_KEYWORDS = ["singles"]
MIXED_KEYWORDS = ["mixed doubles"]

SIMILARITY_THRESHOLD = 0.90


# -----------------------------
# Helpers
# -----------------------------

def normalize_text(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_name(name: str) -> str:
    s = normalize_text(name)
    s = s.replace("&", "/")
    s = re.sub(r"\s*/\s*", " / ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def looks_like_year_sheet(sheet_name: str) -> bool:
    return bool(YEAR_SHEET_RE.match(str(sheet_name)))


def find_columns(df: pd.DataFrame) -> dict[str, str | None]:
    cols = list(df.columns)
    lower_map = {c: str(c).strip().lower() for c in cols}

    out = {
        "event_id": None,
        "event_name": None,
        "discipline": None,
        "placement": None,
        "players": None,
        "year": None,
    }

    for c, lc in lower_map.items():
        if out["event_id"] is None and lc in {"event id", "event_id"}:
            out["event_id"] = c
        if out["event_name"] is None and lc in {"event name", "event_name"}:
            out["event_name"] = c
        if out["discipline"] is None and any(k in lc for k in ["discipline", "division", "event type"]):
            out["discipline"] = c
        if out["placement"] is None and any(k in lc for k in ["place", "placement", "rank"]):
            out["placement"] = c
        if out["players"] is None and any(k in lc for k in ["player(s)", "player", "players", "participant", "name"]):
            out["players"] = c
        if out["year"] is None and lc == "year":
            out["year"] = c

    return out


def discipline_type(discipline: str) -> str:
    d = normalize_text(discipline).lower()
    if any(k in d for k in MIXED_KEYWORDS):
        return "doubles"
    # Americano is an individual-rotation format: players are ranked individually,
    # not as fixed pairs.  Team-size checks are not applicable.
    if "americano" in d:
        return "other"
    has_doubles = any(k in d for k in DOUBLES_KEYWORDS)
    has_singles = any(k in d for k in SINGLES_KEYWORDS)
    # Combined discipline (e.g. "Open Singles Net Open Doubles Net") — treat as
    # "other" so team-size checks are skipped for the mixed result set.
    if has_doubles and has_singles:
        return "other"
    if has_doubles:
        return "doubles"
    if has_singles:
        return "singles"
    return "other"


def split_players(players: str) -> list[str]:
    s = normalize_name(players)
    if not s:
        return []

    # Strip "Team " prefix that some legacy sources prepend to a combined name
    s = re.sub(r"^Team\s+", "", s, flags=re.IGNORECASE)

    # Handle "?" as an explicit team-separator used in some French/other sources
    # E.g., "S. Thomas Sustrac ? Robinson Sustrac"
    if "?" in s and "/" not in s:
        parts_q = [re.sub(r"^Team\s+", "", p, flags=re.IGNORECASE).strip()
                   for p in s.split("?") if p.strip()]
        if len(parts_q) == 2:
            return parts_q

    # Handle "First Last (STATE) First Last" unsplit doubles pair
    # E.g., "Jim Fitzgerald (OR) Adam Hutchinson" → ["Jim Fitzgerald", "Adam Hutchinson"]
    # Guard: left side must have ≥2 words so "Paul (PT) Lovern" (nickname) is not split
    m = re.search(r"^(.+?)\s+\([A-Z]{2,3}\)\s+(.+)$", s)
    if m and len(m.group(1).split()) >= 2:
        return [m.group(1).strip(), m.group(2).strip()]

    s = s.replace(" and ", " / ")
    s = s.replace("&", "/")
    parts = [p.strip() for p in s.split("/") if p.strip()]
    return parts


def parse_placement(raw) -> tuple[int | None, bool, str]:
    """
    Returns:
        (rank_number, tie_flag, normalized_display)
    Examples:
        1      -> (1, False, "1")
        5T     -> (5, True, "5T")
        5T.    -> (5, True, "5T")
        10.    -> (10, False, "10")
    """
    s = normalize_text(raw)
    if not s:
        return None, False, ""

    s = s.replace(".", "").strip().upper()
    m = re.match(r"^(\d+)(T)?$", s)
    if not m:
        return None, False, s

    n = int(m.group(1))
    tie = bool(m.group(2))
    return n, tie, f"{n}{'T' if tie else ''}"


def parse_wide_year_sheet(ws) -> list[dict]:
    """
    Parse v13-style year sheets where each event is a column and disciplines /
    placements are vertically stacked within the column.

    Expected structure (row numbers are conventional but not guaranteed):
    - Row 1: Event names across columns (B..)
    - Row 7: "Event ID" in col A, IDs across columns (B..)
    - Below: discipline header rows (strings) followed by placement rows like:
      "🥇   1  Name / Name" or "1T  Name" etc.
    """
    # Materialize rows once; iter_rows(values_only=True) is much faster than
    # repeated ws.cell(...) access across many sheets.
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find the "Event ID" row (usually 7)
    event_id_row_idx0 = None  # 0-based
    scan_limit = min(len(rows), 60)
    for i in range(scan_limit):
        v = rows[i][0] if rows[i] else None
        if v is None:
            continue
        if str(v).strip().lower() == "event id":
            event_id_row_idx0 = i
            break

    if event_id_row_idx0 is None:
        return []

    header_row = rows[0] if len(rows) >= 1 else ()
    event_id_row = rows[event_id_row_idx0]

    # Build event columns map from that row
    event_cols: list[tuple[int, str, str]] = []  # (col_idx0, event_id, event_name)
    max_cols = max(len(header_row), len(event_id_row))
    for j in range(1, max_cols):
        event_id = normalize_text(event_id_row[j] if j < len(event_id_row) else None)
        if not event_id:
            continue
        event_name = normalize_text(header_row[j] if j < len(header_row) else None)
        event_cols.append((j, event_id, event_name))

    if not event_cols:
        return []

    # The optional prefix before the placement number must be non-alphanumeric
    # (e.g. medal emoji like 🥇, or nothing) — NOT a word like "Open", "Advanced".
    # Using [^\w\s]+ prevents "Open 5 Minute Timed..." from matching as place=5.
    placement_re = re.compile(r"^\s*(?:[^\w\s]+\s+)?(\d+)\s*(T)?\s+(.*)$", re.IGNORECASE)

    out = []
    for j, event_id, event_name in event_cols:
        current_discipline = ""
        # Content starts a couple rows after the Event ID row in v13, but scan broadly.
        for i in range(event_id_row_idx0 + 1, len(rows)):
            row = rows[i]
            raw = row[j] if j < len(row) else None
            s = normalize_text(raw)
            if not s:
                continue

            # Section markers like NET/FREESTYLE are not disciplines.
            if s.isupper() and len(s) <= 30 and " " not in s:
                continue

            # Try placement row
            m = placement_re.match(s)
            if m:
                place_num = int(m.group(1))
                tie = bool(m.group(2))
                players_raw = normalize_text(m.group(3))
                if not players_raw:
                    continue
                # If the "players" part is a single word (no spaces, no "/"),
                # it's almost certainly a discipline name like "2 Square",
                # "3-Man Goal", etc. — treat the whole cell as a discipline header.
                # If the raw cell has no leading whitespace (no indentation,
                # no medal emoji) the whole cell is a division name, not a
                # placement row.  All placement rows produced by the v13
                # workbook builder start with "    " (4 spaces) or a medal
                # emoji, so a cell that starts with a plain ASCII character
                # (and is not a section marker already filtered above) must
                # be a discipline/category header.  Examples: "2 Square",
                # "30 Second Shred - Intermediate", "Women's Singles Net".
                # Leading-space rows like "     10  Mathieu" are genuine
                # placements and must not be reclassified.
                raw_str = str(raw) if raw is not None else ""
                cell_starts_plain = raw_str and raw_str[0] not in (" ", "\t") and ord(raw_str[0]) < 128
                if cell_starts_plain:
                    current_discipline = s
                    continue
                if not current_discipline:
                    # Can't safely attach a placement to a discipline; skip.
                    continue

                out.append({
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": current_discipline,
                    "placement_raw": f"{place_num}{'T' if tie else ''}",
                    "players_raw": players_raw,
                    "row_idx": i + 1,  # Excel row number (1-based)
                })
                continue

            # Otherwise treat as discipline header (strings like "Open Doubles Net")
            # A discipline header is typically short-ish and not a label row.
            if len(s) <= 80 and not s.lower().startswith(("players:", "status", "event type")):
                current_discipline = s

    return out


def row_has_non_person_artifact(players: str) -> bool:
    s = normalize_text(players).lower()
    return any(re.search(pat, s) for pat in NON_PERSON_PATTERNS)


def _looks_like_unsplit_pair(s: str) -> bool:
    """Return True only when the string shows positive signs of being two
    concatenated player names lacking a proper delimiter.

    This guards against false positives: solo players with country annotations
    ("Benjamin Kanske (Germany)"), individual-score rows in Doubles Golf, etc.
    should NOT be flagged as potentially unsplit pairs.
    """
    # Explicit '?' used as a separator in some sources
    if "?" in s:
        return True
    # "First Last (STATE) First Last" — two-letter/three-letter ALL-CAPS state or
    # country code in parentheses with another name following it.
    # Require left side to have ≥2 words so "Paul (PT) Lovern" (nickname form) is
    # not matched.
    m = re.search(r"^(.+?)\s+\([A-Z]{2,3}\)\s+\S", s)
    if m and len(m.group(1).split()) >= 2:
        return True
    return False


def safe_get(row, col):
    if col is None:
        return ""
    return row.get(col, "")


def similar(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


# -----------------------------
# Core QC
# -----------------------------

def run_qc(xlsx_path: Path):
    issues = []
    results_rows = []

    # 1) Load only year sheets as result sheets (support both wide and tabular formats)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        if not looks_like_year_sheet(sheet):
            continue

        ws = wb[sheet]
        wide_rows = parse_wide_year_sheet(ws)
        if wide_rows:
            for rec0 in wide_rows:
                discipline = normalize_text(rec0["discipline"])
                placement_raw = rec0["placement_raw"]
                players_raw = normalize_text(rec0["players_raw"])

                place_num, tie_flag, placement_norm = parse_placement(placement_raw)
                team_type = discipline_type(discipline)
                player_list = split_players(players_raw)

                rec = {
                    "sheet": sheet,
                    "row_idx": int(rec0["row_idx"]),
                    "event_id": normalize_text(rec0["event_id"]),
                    "event_name": normalize_text(rec0["event_name"]),
                    "discipline": discipline,
                    "team_type": team_type,
                    "placement_raw": normalize_text(placement_raw),
                    "placement_num": place_num,
                    "placement_norm": placement_norm,
                    "tie_flag": tie_flag,
                    "players_raw": players_raw,
                    "player_list": player_list,
                }
                results_rows.append(rec)

                # Orphan rows
                if placement_norm and not players_raw:
                    issues.append({
                        "severity": "HIGH",
                        "sheet": sheet,
                        "event_id": rec["event_id"],
                        "event_name": rec["event_name"],
                        "discipline": discipline,
                        "placement": placement_norm,
                        "players": players_raw,
                        "issue_type": "orphan_placement_no_player",
                        "details": f"Row {int(rec0['row_idx'])}",
                    })

                if players_raw and not placement_norm:
                    issues.append({
                        "severity": "MEDIUM",
                        "sheet": sheet,
                        "event_id": rec["event_id"],
                        "event_name": rec["event_name"],
                        "discipline": discipline,
                        "placement": "",
                        "players": players_raw,
                        "issue_type": "orphan_player_no_placement",
                        "details": f"Row {int(rec0['row_idx'])}",
                    })

                # Non-person artifacts
                if players_raw and row_has_non_person_artifact(players_raw):
                    issues.append({
                        "severity": "HIGH",
                        "sheet": sheet,
                        "event_id": rec["event_id"],
                        "event_name": rec["event_name"],
                        "discipline": discipline,
                        "placement": placement_norm,
                        "players": players_raw,
                        "issue_type": "non_person_artifact",
                        "details": f"Row {int(rec0['row_idx'])}",
                    })

                # Missing delimiter in doubles row
                _is_golf = "golf" in discipline.lower()
                if (team_type == "doubles"
                        and not _is_golf          # golf has intentional individual-score rows
                        and players_raw
                        and "/" not in players_raw
                        and len(player_list) == 1
                        and _looks_like_unsplit_pair(players_raw)):
                    issues.append({
                        "severity": "LOW",
                        "sheet": sheet,
                        "event_id": rec["event_id"],
                        "event_name": rec["event_name"],
                        "discipline": discipline,
                        "placement": placement_norm,
                        "players": players_raw,
                        "issue_type": "doubles_missing_delimiter_possible_split_pair",
                        "details": f"Row {int(rec0['row_idx'])}",
                    })
            continue

        # Fallback: older tabular year-sheet format (kept for backwards compatibility)
        df = pd.read_excel(xlsx_path, sheet_name=sheet)
        cols = find_columns(df)

        required = ["event_id", "event_name", "discipline", "placement", "players"]
        missing = [k for k in required if cols[k] is None]
        if missing:
            issues.append({
                "severity": "HIGH",
                "sheet": sheet,
                "event_id": "",
                "event_name": "",
                "discipline": "",
                "placement": "",
                "players": "",
                "issue_type": "missing_required_columns",
                "details": f"Missing columns: {missing}",
            })
            continue

        for idx, row in df.iterrows():
            event_id = normalize_text(safe_get(row, cols["event_id"]))
            event_name = normalize_text(safe_get(row, cols["event_name"]))
            discipline = normalize_text(safe_get(row, cols["discipline"]))
            placement_raw = safe_get(row, cols["placement"])
            players_raw = normalize_text(safe_get(row, cols["players"]))

            place_num, tie_flag, placement_norm = parse_placement(placement_raw)
            team_type = discipline_type(discipline)
            player_list = split_players(players_raw)

            rec = {
                "sheet": sheet,
                "row_idx": int(idx) + 2,  # Excel-ish row number
                "event_id": event_id,
                "event_name": event_name,
                "discipline": discipline,
                "team_type": team_type,
                "placement_raw": normalize_text(placement_raw),
                "placement_num": place_num,
                "placement_norm": placement_norm,
                "tie_flag": tie_flag,
                "players_raw": players_raw,
                "player_list": player_list,
            }
            results_rows.append(rec)

            # Orphan rows
            if placement_norm and not players_raw:
                issues.append({
                    "severity": "HIGH",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": placement_norm,
                    "players": players_raw,
                    "issue_type": "orphan_placement_no_player",
                    "details": f"Row {int(idx)+2}",
                })

            if players_raw and not placement_norm:
                issues.append({
                    "severity": "MEDIUM",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": "",
                    "players": players_raw,
                    "issue_type": "orphan_player_no_placement",
                    "details": f"Row {int(idx)+2}",
                })

            # Non-person artifacts
            if players_raw and row_has_non_person_artifact(players_raw):
                issues.append({
                    "severity": "HIGH",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": placement_norm,
                    "players": players_raw,
                    "issue_type": "non_person_artifact",
                    "details": f"Row {int(idx)+2}",
                })

            # Missing delimiter in doubles row
            _is_golf = "golf" in discipline.lower()
            if (team_type == "doubles"
                    and not _is_golf
                    and players_raw
                    and "/" not in players_raw
                    and len(player_list) == 1
                    and _looks_like_unsplit_pair(players_raw)):
                issues.append({
                    "severity": "LOW",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": placement_norm,
                    "players": players_raw,
                    "issue_type": "doubles_missing_delimiter_possible_split_pair",
                    "details": f"Row {int(idx)+2}",
                })

    df_rows = pd.DataFrame(results_rows)

    if df_rows.empty:
        raise RuntimeError("No usable result rows found in year sheets.")

    # 2) Group-based checks
    group_cols = ["sheet", "event_id", "event_name", "discipline", "placement_num"]

    grouped = df_rows.groupby(group_cols, dropna=False)

    for key, g in grouped:
        sheet, event_id, event_name, discipline, placement_num = key
        team_type = discipline_type(discipline)

        if placement_num is None:
            continue

        raw_players = list(g["players_raw"])
        flattened = []
        for plist in g["player_list"]:
            flattened.extend(plist)

        flattened_norm = [normalize_name(x) for x in flattened if normalize_name(x)]
        unique_players = list(dict.fromkeys(flattened_norm))

        # Sentinel values that legitimately repeat across split-pair rows:
        # "[Unknown Partner]" (used when one partner is unresolved) and "DNF"
        # (Did Not Finish) should not trigger duplicate detection.
        _DUPE_SENTINELS = {"[unknown partner]", "dnf"}

        # Duplicate within placement
        if len(flattened_norm) != len(set(flattened_norm)):
            dupes = [
                p for p in sorted(set(flattened_norm))
                if flattened_norm.count(p) > 1 and p.lower() not in _DUPE_SENTINELS
            ]
            if dupes:
                issues.append({
                    "severity": "HIGH",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": str(placement_num),
                    "players": " | ".join(raw_players),
                    "issue_type": "duplicate_player_within_placement",
                    "details": ", ".join(dupes),
                })

        # Team-size integrity at placement-group level
        if team_type == "singles":
            n = len(unique_players)
            if n == 0:
                # Genuinely missing player — actual data error
                issues.append({
                    "severity": "HIGH",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": str(placement_num),
                    "players": " | ".join(raw_players),
                    "issue_type": "singles_wrong_player_count",
                    "details": "No players found in singles placement",
                })
            elif n > 1:
                # Multiple players at same placement = valid tie or elimination group
                # Small tie (2-3): likely a scored tie in round-robin/pool play
                # Large group (4+): likely a shared elimination position
                tie_class = "small_tie" if n <= 3 else "large_group"
                issues.append({
                    "severity": "INFO",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": str(placement_num),
                    "players": " | ".join(raw_players),
                    "issue_type": "singles_tie_group",
                    "details": f"tie_size={n} tie_class={tie_class}",
                })

        elif team_type == "doubles":
            # Doubles Golf records individual stroke scores (1 player/row) AND
            # team scores (2 players/row) at the same placement — exempt from
            # strict 2-player check.
            is_golf_discipline = "golf" in discipline.lower()

            if len(unique_players) == 2:
                # Good, but still check if split over multiple rows
                if len(g) > 1 and all(len(pl) == 1 for pl in g["player_list"]):
                    issues.append({
                        "severity": "MEDIUM",
                        "sheet": sheet,
                        "event_id": event_id,
                        "event_name": event_name,
                        "discipline": discipline,
                        "placement": str(placement_num),
                        "players": " | ".join(raw_players),
                        "issue_type": "split_doubles_pair",
                        "details": "Two singles on separate rows at same placement; likely should be merged",
                    })
            elif len(unique_players) == 1:
                issues.append({
                    "severity": "HIGH",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": str(placement_num),
                    "players": " | ".join(raw_players),
                    "issue_type": "doubles_only_one_player",
                    "details": "Only one player found in doubles placement group",
                })
            elif is_golf_discipline:
                # Doubles Golf: mixed 1+2 player rows are expected — not an error
                pass
            elif len(g) > 1 and all(len(pl) in (1, 2) for pl in g["player_list"]) and any(len(pl) == 1 for pl in g["player_list"]) and any(len(pl) == 2 for pl in g["player_list"]):
                # Mixed 1+2 player rows: individual entry + team entry at same
                # placement (known format in some routines events — individual
                # scores listed alongside the pair).  Not a data error.
                pass
            elif len(g) > 1 and all(len(pl) == 2 for pl in g["player_list"]):
                # Multi-team tie: every row in this placement group is a valid
                # 2-player team.  e.g. 3rd-place tie with two different doubles
                # teams produces two rows both at place=3.  This is correct.
                pass
            else:
                issues.append({
                    "severity": "HIGH",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": str(placement_num),
                    "players": " | ".join(raw_players),
                    "issue_type": "doubles_ambiguous_player_count",
                    "details": f"Found {len(unique_players)} players in doubles placement group",
                })

    # 3) Duplicate player in same event/discipline
    ed_groups = df_rows.groupby(["sheet", "event_id", "event_name", "discipline"], dropna=False)

    for key, g in ed_groups:
        sheet, event_id, event_name, discipline = key
        seen = defaultdict(list)

        for _, row in g.iterrows():
            for p in row["player_list"]:
                pn = normalize_name(p)
                if pn:
                    seen[pn].append(row["placement_norm"] or "(no placement)")

        # Sentinel / non-person values that legitimately repeat across placements
        # (each doubles team has its own unknown partner, not the same person).
        _SAME_EVENT_SENTINELS = {
            "[unknown partner]",
            # Placeholder text for unannounced/unknown registrants
            "be announced",
            "to be announced",
            "tba",
            # "N/A" placeholder (parsed from source as bare "Na" by workbook reader)
            "na",
            "n/a",
            "n.a.",
        }

        # Only flag players that appear at MULTIPLE DISTINCT places.
        # Appearing at the same place twice (e.g., as part of two tied
        # teams both at place 3) is a TIE — expected and not a data error.
        dup_players = {
            p: places
            for p, places in seen.items()
            if len(set(places)) > 1   # distinct placements, not just count
        }
        for p, places in dup_players.items():
            distinct_places = sorted(set(places))
            # Skip known sentinels
            if p.lower() in _SAME_EVENT_SENTINELS:
                continue
            # Skip non-person artifacts: names that are fragments (contain
            # unmatched parentheses), pure country/city tokens, or look like
            # metadata labels already caught by non_person_artifact.
            if row_has_non_person_artifact(p):
                continue
            # Skip obvious artifact patterns specific to this check:
            #   - names containing ")" (split-team fragments like "BC) or")
            #   - names ending with ")" or " and" (split-team fragments)
            p_stripped = p.strip()
            if ")" in p_stripped or p_stripped.lower().endswith(" and"):
                continue
            # Single-word tokens that start with an uppercase letter followed
            # by a space + a single word that's a known non-person indicator
            if re.search(r"\b(Club|Club\b|Federation|Association|Footbag)\b", p_stripped, re.IGNORECASE):
                continue
            # Single-word tokens are almost always city/province/country
            # artifacts or unresolved single-name placeholders (e.g., "Jean",
            # "Poland", "Montreal", "Ontario").  These are covered by the
            # WARN / unresolved-identity category — not a hard data error.
            if " " not in p_stripped and not p_stripped.startswith("["):
                continue
            issues.append({
                "severity": "HIGH",
                "sheet": sheet,
                "event_id": event_id,
                "event_name": event_name,
                "discipline": discipline,
                "placement": ", ".join(distinct_places),
                "players": p,
                "issue_type": "duplicate_player_same_event_discipline",
                "details": f"Appears in placements: {distinct_places}",
            })

    # 4) Placement sequence / tie integrity
    for key, g in ed_groups:
        sheet, event_id, event_name, discipline = key
        gg = g[g["placement_num"].notna()].copy()
        if gg.empty:
            continue

        # collapse to unique placement groups
        place_groups = (
            gg.groupby("placement_num", dropna=False)
            .agg(
                tie_rows=("tie_flag", "sum"),
                row_count=("placement_num", "size"),
            )
            .reset_index()
            .sort_values("placement_num")
        )

        placements = list(place_groups["placement_num"])

        # basic monotonic / gap check
        prev = None
        for p in placements:
            if prev is not None and p < prev:
                issues.append({
                    "severity": "HIGH",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": f"{prev}->{p}",
                    "players": "",
                    "issue_type": "placement_out_of_order",
                    "details": "Placements decrease",
                })
            prev = p

        # tie check heuristic:
        # if multiple rows share same placement but none marked T, flag.
        for _, r in place_groups.iterrows():
            p = int(r["placement_num"])
            row_count = int(r["row_count"])
            tie_rows = int(r["tie_rows"])
            if row_count > 1 and tie_rows == 0:
                issues.append({
                    "severity": "MEDIUM",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": str(p),
                    "players": "",
                    "issue_type": "possible_missing_tie_notation",
                    "details": f"{row_count} rows share placement {p} but no T notation seen",
                })

        # gap check: validate that each placement jump is explained by group size.
        # For each placement P with N entries, the expected next placement is P+N.
        # A gap is only suspicious when next_P > P + N (more places skipped than
        # the tie/group accounts for).  Gaps that exactly match group size are valid.
        group_size_map = dict(
            zip(place_groups["placement_num"].astype(int),
                place_groups["row_count"].astype(int))
        )
        unique_places = sorted(set(placements))
        for i in range(1, len(unique_places)):
            P = int(unique_places[i - 1])
            next_P = int(unique_places[i])
            group_size = group_size_map.get(P, 1)
            expected_next = P + group_size

            if next_P <= expected_next:
                # Gap is fully accounted for by the group size (or overlaps due
                # to partial / loosely-numbered data) — not suspicious.
                pass
            else:
                # next_P > expected_next: true unexplained gap beyond what the
                # group size at P predicts.
                issues.append({
                    "severity": "MEDIUM",
                    "sheet": sheet,
                    "event_id": event_id,
                    "event_name": event_name,
                    "discipline": discipline,
                    "placement": f"{P}->{next_P}",
                    "players": "",
                    "issue_type": "suspicious_placement_gap",
                    "details": (
                        f"Gap {P}->{next_P} unexplained "
                        f"(group_size={group_size}, expected_next={expected_next})"
                    ),
                })

    # 5) Near-duplicate name detection within same event
    event_groups = df_rows.groupby(["sheet", "event_id", "event_name"], dropna=False)

    for key, g in event_groups:
        sheet, event_id, event_name = key
        names = []
        for plist in g["player_list"]:
            for p in plist:
                pn = normalize_name(p)
                if pn:
                    names.append(pn)

        names = sorted(set(names))
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                a, b = names[i], names[j]
                if a == b:
                    continue
                score = similar(a, b)
                if score >= SIMILARITY_THRESHOLD:
                    issues.append({
                        "severity": "LOW",
                        "sheet": sheet,
                        "event_id": event_id,
                        "event_name": event_name,
                        "discipline": "",
                        "placement": "",
                        "players": f"{a} <> {b}",
                        "issue_type": "near_duplicate_names_same_event",
                        "details": f"similarity={score:.3f}",
                    })

    issues_df = pd.DataFrame(issues)

    if issues_df.empty:
        issues_df = pd.DataFrame(columns=[
            "severity", "sheet", "event_id", "event_name", "discipline",
            "placement", "players", "issue_type", "details"
        ])

    # summary by event
    if not issues_df.empty:
        summary_df = (
            issues_df.groupby(["sheet", "event_id", "event_name", "issue_type"], dropna=False)
            .size()
            .reset_index(name="count")
            .sort_values(["sheet", "event_id", "issue_type"])
        )
    else:
        summary_df = pd.DataFrame(columns=["sheet", "event_id", "event_name", "issue_type", "count"])

    return issues_df, summary_df


# -----------------------------
# Main
# -----------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python qc_footbag_workbook.py /path/to/workbook.xlsx")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    issues_df, summary_df = run_qc(xlsx_path)

    issues_out = xlsx_path.with_name("qc_issues.csv")
    summary_out = xlsx_path.with_name("qc_summary_by_event.csv")

    issues_df.to_csv(issues_out, index=False)
    summary_df.to_csv(summary_out, index=False)

    print(f"Wrote: {issues_out}")
    print(f"Wrote: {summary_out}")

    print("\nTop issue counts:")
    if issues_df.empty:
        print("No issues found.")
    else:
        print(issues_df["issue_type"].value_counts().to_string())

        print("\nTop affected events:")
        top_events = (
            issues_df.groupby(["event_id", "event_name"])
            .size()
            .sort_values(ascending=False)
            .head(20)
        )
        print(top_events.to_string())


if __name__ == "__main__":
    main()
