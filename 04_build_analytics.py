#!/usr/bin/env python3
"""
04_build_analytics.py — Stage 4: Add analytics sheets to the canonical workbook.

Reads:
  - out/Placements_Flat.csv  (must exist; produced by 02p5)
  - Footbag_Results_Canonical.xlsx (produced by 03)
Optionally reads:
  - overrides/person_aliases.csv   (human-verified)

Writes (adds/replaces sheets in the workbook):
  - Person_Stats
  - Player_Stats
  - Division_Stats
  - Person_Aliases (optional)

No guessing: if person_id missing, falls back to player_id.
"""

from __future__ import annotations

import csv
import re
import sys
import uuid
import unicodedata
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
import json

from qc_common import PERSONS, PLACEMENTS

# GATE 2 COVERAGE THRESHOLDS — do not change casually; changing these alters
# the meaning of all historical coverage annotations in downstream artifacts.
_G2_COMPLETE        = 0.95
_G2_MOSTLY_COMPLETE = 0.75
_G2_PARTIAL         = 0.40
# Below _G2_PARTIAL → "sparse"

_COVERAGE_FILLS = {
    "complete":        PatternFill(fill_type="solid", fgColor="92D050"),  # green
    "mostly_complete": PatternFill(fill_type="solid", fgColor="FFEB9C"),  # yellow
    "partial":         PatternFill(fill_type="solid", fgColor="FFC000"),  # orange
    "sparse":          PatternFill(fill_type="solid", fgColor="FFC7CE"),  # light red
}


def _coverage_flag(ratio) -> str:
    if pd.isna(ratio):
        return ""
    if ratio >= _G2_COMPLETE:
        return "complete"
    if ratio >= _G2_MOSTLY_COMPLETE:
        return "mostly_complete"
    if ratio >= _G2_PARTIAL:
        return "partial"
    return "sparse"


def read_csv_optional(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path, dtype=str, encoding="utf-8").fillna("")
    except UnicodeDecodeError:
        return pd.read_csv(path, dtype=str, encoding="cp1252").fillna("")


# Regex-level "presentability"
RE_BAD_SEPARATORS = re.compile(r"[+/\\=]|(\b(and|or)\b)", re.IGNORECASE)
RE_HAS_DIGIT = re.compile(r"\d")
RE_BAD_WORDS = re.compile(r"\b(results?|final|place|pts?|points?|scratch|victory)\b", re.IGNORECASE)


def is_presentable_person(s: str) -> bool:
    s = (s or "").strip()
    if not s:
        return False
    if RE_BAD_SEPARATORS.search(s):  # multi-person / operator
        return False
    if RE_HAS_DIGIT.search(s):       # ranks, scores, ages
        return False
    if RE_BAD_WORDS.search(s):       # non-name artifacts
        return False
    toks = s.split()
    if len(toks) < 2:                # require at least First Last
        return False
    if len(toks) > 3:                # reject likely "two people" / junk strings
        return False
    if not any(ch.isalpha() for ch in s):
        return False
    return True


def _normalize_status(s: str) -> str:
    return (s or "").strip().upper()


def _canon_key(s: str) -> str:
    return (s or "").strip().casefold()


# --- Triage for Persons_Unresolved (presentation-only heuristics) ---
_TRIAGE_HARD_REJECT_PHRASES = {
    "pdl","sky","flip bags","flipbags","kc blender","blender",
    "bag","bags","footbag","freestyle","net","crew","posse","team","club",
    "jam","demo","workshop","tournament","results","final"
}
_TRIAGE_HARD_REJECT_ISSUE_TYPES = {
    "not_person_like","club_or_group","heading_or_meta","equipment_or_brand",
    "trick_notation","location_only","event_title","instructional_text"
}
_TRIAGE_PERSON_LIKE_ISSUE_TYPES = {"encoding_corruption","diacritics","misspelling","needs_alias","variant"}

_TRIAGE_REJECT_RX = [
    ("trick_symbols", re.compile(r"[><=~^*+/\\]|::|->")),
    ("many_digits", re.compile(r"\d{2,}")),
    ("url_like", re.compile(r"(https?://|www\.)", re.I)),
    ("email_like", re.compile(r"\b\S+@\S+\.\S+\b")),
]
_TRIAGE_PERSON_RX = [
    ("two_words_cap", re.compile(r"^[A-Z][a-z]+ [A-Z][a-z]+$")),
    ("three_words_cap", re.compile(r"^[A-Z][a-z]+ [A-Z][a-z]+ [A-Z][a-z]+$")),
    ("initial_last", re.compile(r"^[A-Z]\. [A-Z][a-z]+$")),
    ("last_first", re.compile(r"^[A-Z][a-z]+, [A-Z][a-z]+$")),
    ("non_ascii", re.compile(r"[^\x00-\x7F]")),
]

def _triage_persons_unresolved(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add heuristic triage columns to Persons_Unresolved-style table.

    Expected columns (if present):
      person_canon, name_raw, issue_type, appearances

    Adds:
      likelihood_score (int)
      resolution_likelihood (HIGH|MEDIUM|LOW|REJECT)
      triage_reasons (semicolon list)

    Deterministic. No merges. Safe for presentation only.
    """
    if df is None or df.empty:
        return df

    def norm(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip())

    def toks(s: str):
        s = re.sub(r"[^a-z0-9 ]+", " ", (s or "").lower())
        return [t for t in s.split() if t]

    def to_int(x) -> int:
        try:
            return int(str(x).strip() or "0")
        except Exception:
            return 0

    def score_row(name: str, issue_type: str, appearances: int):
        n = norm(name)
        nlow = n.lower()
        token_set = set(toks(n))
        reasons = []
        score = 0

        it = (issue_type or "").strip().lower()
        if it in _TRIAGE_HARD_REJECT_ISSUE_TYPES:
            reasons.append(f"hard_reject_issue_type:{it}")
            score -= 120
        if it in _TRIAGE_PERSON_LIKE_ISSUE_TYPES:
            reasons.append(f"person_like_issue_type:{it}")
            score += 20

        if n.isupper() and 2 <= len(n) <= 8 and re.fullmatch(r"[A-Z0-9]+", n):
            reasons.append("all_caps_acronym")
            score -= 60
        if len(n) <= 2:
            reasons.append("too_short")
            score -= 50

        for ph in _TRIAGE_HARD_REJECT_PHRASES:
            if " " in ph and ph in nlow:
                reasons.append(f"hard_reject_phrase:{ph}")
                score -= 80
        for t in token_set:
            if t in _TRIAGE_HARD_REJECT_PHRASES:
                reasons.append(f"hard_reject_token:{t}")
                score -= 70

        for label, rx in _TRIAGE_REJECT_RX:
            if rx.search(n):
                reasons.append(label)
                score -= 15

        for label, rx in _TRIAGE_PERSON_RX:
            if rx.search(n):
                reasons.append(label)
                score += 40

        if appearances >= 20:
            score += 20; reasons.append("appearances>=20")
        elif appearances >= 10:
            score += 12; reasons.append("appearances>=10")
        elif appearances >= 3:
            score += 6; reasons.append("appearances>=3")
        elif appearances == 1:
            score -= 3; reasons.append("appearances==1")

        score = max(-200, min(200, score))

        if score <= -60: bucket = "REJECT"
        elif score <= -10: bucket = "LOW"
        elif score < 50: bucket = "MEDIUM"
        else: bucket = "HIGH"

        return score, bucket, ";".join(sorted(set(reasons)))

    out = df.copy()
    # prefer person_canon, fallback to name_raw
    name_series = out.get("person_canon", pd.Series([""] * len(out))).astype(str)
    if "name_raw" in out.columns:
        name_series = name_series.where(name_series.str.strip() != "", out["name_raw"].astype(str))

    issue_series = out.get("issue_type", pd.Series([""] * len(out))).astype(str)
    app_series = out.get("appearances", pd.Series(["0"] * len(out))).apply(to_int)

    scores, buckets, reasons = [], [], []
    for n, it, app in zip(name_series.tolist(), issue_series.tolist(), app_series.tolist()):
        sc, b, rs = score_row(n, it, app)
        scores.append(sc); buckets.append(b); reasons.append(rs)

    out["likelihood_score"] = scores
    out["resolution_likelihood"] = buckets
    out["triage_reasons"] = reasons

    # stable sort
    order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "REJECT": 3}
    out["_rank"] = out["resolution_likelihood"].map(order).fillna(9).astype(int)
    if "appearances" in out.columns:
        out["_app_int"] = app_series
        out = out.sort_values(by=["_rank", "likelihood_score", "_app_int", "person_canon"],
                              ascending=[True, False, False, True]).drop(columns=["_rank", "_app_int"])
    else:
        out = out.sort_values(by=["_rank", "likelihood_score", "person_canon"],
                              ascending=[True, False, True]).drop(columns=["_rank"])

    return out


def build_aliases_presentable_from_overrides(person_aliases_df: pd.DataFrame) -> dict[str, str]:
    """
    Return person_id -> 'alias1 | alias2 | ...' (VERIFIED only, presentable only, deterministic order).
    """
    if person_aliases_df is None or len(person_aliases_df) == 0:
        return {}
    df = person_aliases_df.copy()
    for c in ("alias", "person_id", "status"):
        if c not in df.columns:
            return {}
    df["alias"] = df["alias"].astype(str).fillna("").str.strip()
    df["person_id"] = df["person_id"].astype(str).fillna("").str.strip()
    df["status"] = df["status"].astype(str).fillna("").map(_normalize_status)
    df = df[(df["person_id"] != "") & (df["alias"] != "") & (df["status"] == "VERIFIED")]
    if df.empty:
        return {}
    df = df[df["alias"].map(is_presentable_person)]
    out: dict[str, list[str]] = {}
    for pid, g in df.groupby("person_id", sort=True):
        aliases = sorted(set(g["alias"].tolist()), key=lambda x: (_canon_key(x), x))
        out[str(pid)] = aliases
    return {pid: " | ".join(v) for pid, v in out.items()}


def quarantine_duplicate_display_names(
    persons_df: pd.DataFrame, name_col: str, id_col: str
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Presentation rule: no duplicate display names across different person IDs.
    Deterministic + safe: do not auto-merge; quarantine instead.
    Returns: (kept_df, quarantined_df)
    """
    if persons_df is None or len(persons_df) == 0:
        return (pd.DataFrame() if persons_df is None else persons_df.copy(), pd.DataFrame())
    df = persons_df.copy()
    df["_canon_key"] = df[name_col].astype(str).fillna("").str.strip().str.casefold()
    df["_id_key"] = df[id_col].astype(str).fillna("").str.strip()
    dup_mask = (df["_canon_key"] != "") & df.duplicated("_canon_key", keep=False)
    quarantined = df.loc[dup_mask].copy()
    kept = df.loc[~dup_mask].copy()
    kept = kept.drop(columns=["_canon_key", "_id_key"], errors="ignore")
    quarantined = quarantined.drop(columns=["_canon_key", "_id_key"], errors="ignore")
    return kept, quarantined


def hide_columns_by_header(ws, headers_to_hide: set[str]) -> None:
    """Hide columns by header name. Assumes headers in row 1."""
    for col_idx, cell in enumerate(ws[1], start=1):
        h = str(cell.value or "").strip()
        if h in headers_to_hide:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True


def hide_columns_by_prefix(ws, prefixes: tuple[str, ...]) -> None:
    """Hide columns whose header starts with any of the given prefixes. Assumes headers in row 1."""
    for col_idx, cell in enumerate(ws[1], start=1):
        h = str(cell.value or "").strip()
        if any(h.startswith(p) for p in prefixes):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True


# Year sheet classifier
RE_YEAR_SHEET = re.compile(r"^(?:Y)?(19\d{2}|20\d{2})$")


def is_year_sheet(name: str) -> bool:
    return bool(RE_YEAR_SHEET.match((name or "").strip()))


def is_qc_sheet(name: str) -> bool:
    """QC/diagnostic sheets should remain fully visible (no hiding)."""
    n = (name or "").strip().lower()
    if n.startswith("qc"):
        return True
    if "qc" in n:
        return True
    if n.endswith(("_full", "_excluded", "_quarantine", "_duplicates")):
        return True
    if n in {"persons_unresolved", "placements_unresolved",
             "players_alias_candidates", "teams_alias_candidates", "divisions_normalized"}:
        return True
    return False


# ----------------------------
# Presentability helpers for aliases (Option A)
# ----------------------------
# Presentability is a stronger constraint than correctness for any value that is displayed.
# Correctness is evaluated only on presentable values.
def add_or_replace_readme_sheet(wb, readme_df: pd.DataFrame | None = None, title: str = "README") -> None:
    """
    Ensure README sheet is at index 0.

    Behavior:
    - If readme_df is provided and non-empty: replace README content with full table and pin as first sheet.
    - Else if README already exists: DO NOT overwrite; just move it to first sheet.
    - Else: create a minimal fallback README as first sheet.
    """

    # Helper: move an existing sheet to index 0 without rewriting content
    def _move_to_front(ws):
        try:
            wb._sheets.remove(ws)
            wb._sheets.insert(0, ws)
        except Exception:
            # If internal API changes, fallback to remove+create (but try to preserve content)
            pass

    has_readme = title in wb.sheetnames
    df_ok = readme_df is not None and len(readme_df.columns) >= 1 and len(readme_df) > 0

    # Case A: No external README provided → preserve Stage 03 README (if present)
    if not df_ok and has_readme:
        _move_to_front(wb[title])
        return

    # Case B: External README provided → replace sheet content (full table)
    if has_readme:
        ws_old = wb[title]
        wb.remove(ws_old)

    ws = wb.create_sheet(title, 0)

    if df_ok:
        df = readme_df.fillna("").astype(str)

        # Header row
        for c_idx, col_name in enumerate(df.columns.tolist(), start=1):
            ws.cell(row=1, column=c_idx, value=str(col_name))

        # Data rows
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, v in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=str(v) if v is not None else "")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Deterministic column widths (sample first 200 rows)
        for c_idx, col_name in enumerate(df.columns.tolist(), start=1):
            letter = get_column_letter(c_idx)
            best = len(str(col_name))
            for v in df.iloc[:200, c_idx - 1].tolist():
                best = max(best, len(str(v)))
            ws.column_dimensions[letter].width = max(10, min(best + 2, 80))

        return

    # Case C: No README at all → create fallback minimal README
    lines = [
        "Footbag Results — Canonical Archive",
        "",
        "Generated by pipeline stages 03 (build Excel) + 04 (post-process + analytics).",
        "Do not hand-edit data tables; use overrides/*.csv instead.",
        "",
        "Option A sheets are pivot/presentation ready (clean names, IDs hidden).",
        "Option B sheets are QC/diagnostics (may include raw/noise).",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 110


def reorder_sheets(wb) -> None:
    """
    Deterministic sheet order:

      1) README
      2) Year sheets (ascending)
      3) Pivot/presentation sheets (Option A)
      4) Analytics sheets
      5) QC / diagnostics sheets (Option B)

    Any unknown sheets are appended at the end in existing relative order.
    """
    names = list(wb.sheetnames)

    readme = [n for n in names if n == "README"]
    years = sorted(
        [n for n in names if is_year_sheet(n)],
        key=lambda n: int(RE_YEAR_SHEET.match(n).group(1))
    )

    # Preferred (presence-safe) order after year sheets
    preferred = [
        "Index",
        "Summary",
        "Persons_Truth",          # promoted: first analytics sheet
        "Analytics_Safe_Surface", # new: default pivot source
        "Divisions",
        "Divisions_Normalized",
        "Division_Stats",
        "Person_Stats",
        "PersonStats_ByDivCat",
        "Placements_ByPerson",
        "Persons_Unresolved",
        "Placements_Unresolved",
        "Coverage_ByEventDiv",
        "Data_Integrity",         # new
        "Players_Clean",
        "Placements_Flat",
    ]
    preferred_present = [n for n in preferred if n in names]

    used = set(readme + years + preferred_present)
    rest = [n for n in names if n not in used]

    target_order = readme + years + preferred_present + rest
    for i, name in enumerate(target_order):
        sheet = wb[name]
        current_idx = wb.worksheets.index(sheet)
        if current_idx != i:
            wb.move_sheet(sheet, offset=i - current_idx)


def _hide_id_columns_sheet(ws) -> None:
    """Hide ID-like columns in a single sheet (generic rule)."""
    id_header_re = re.compile(r"(^id$|.*(_id|_ids)$|.*(uuid|guid|hash)$|.*(person_id|player_id)$)", re.IGNORECASE)
    for col_idx, cell in enumerate(ws[1], start=1):
        v = cell.value
        if not isinstance(v, str):
            continue
        h = v.strip()
        if id_header_re.match(h) or h in {"effective_person_id", "player_ids_seen"}:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True


def _apply_sheet_hiding(wb) -> None:
    """Apply per-sheet column hiding rules."""
    _prefixes_placements = (
        "player1_name_raw", "player1_name_clean", "player1_name_noise",
        "player2_name_raw", "player2_name_clean", "player2_name_noise",
        "player1_person_canon", "player2_person_canon",
    )
    _prefixes_year = _prefixes_placements + (
        "player1_person_id", "player2_person_id",
    )
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1:
            continue
        if sheet_name == "Persons_Truth":
            hide_columns_by_header(ws, {"effective_person_id"})
            _hide_id_columns_sheet(ws)
        elif sheet_name == "Placements_ByPerson":
            hide_columns_by_header(ws, {"player1_id", "player2_id", "team_person_key"})
            hide_columns_by_prefix(ws, _prefixes_placements)
            _hide_id_columns_sheet(ws)
        elif is_year_sheet(sheet_name):
            hide_columns_by_header(ws, {"player1_id", "player2_id"})
            hide_columns_by_prefix(ws, _prefixes_year)
            _hide_id_columns_sheet(ws)
        else:
            # Keep QC/diagnostic sheets fully visible; presentation sheets hide ID-like columns.
            if not is_qc_sheet(sheet_name):
                _hide_id_columns_sheet(ws)


def _apply_coverage_colors(wb) -> None:
    """Color-code coverage_flag column cells in all sheets that have it."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        cov_col = None
        for col_idx in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=col_idx).value or "").strip() == "coverage_flag":
                cov_col = col_idx
                break
        if cov_col is None:
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=cov_col)
            val = str(cell.value or "").strip()
            fill = _COVERAGE_FILLS.get(val)
            if fill:
                cell.fill = fill


def _as_int_place(x) -> Optional[int]:
    try:
        if pd.isna(x):
            return None
        s = str(x).strip()
        if not s:
            return None
        return int(float(s))
    except Exception:
        return None


def _norm(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    if not isinstance(s, str):
        s = str(s)
    return " ".join(s.strip().split())


_UUID_NS_PERSON = uuid.UUID("3b5d5c7e-7c4b-4d21-8b44-3c39d1a0f4d6")  # any fixed UUID you choose once

_RE_TRAIL_AND = re.compile(r"\band\s+([A-Z][A-Za-z'’-]+(?:\s+[A-Z][A-Za-z'’-]+){1,3})\s*$", re.IGNORECASE)


def clean_person_label_no_guess(s: str) -> tuple[str, str]:
    """
    Returns (clean_label, reason)
    reason is "" if unchanged.
    """
    if not isinstance(s, str):
        return ("", "non_string")
    t = s.strip()
    if not t:
        return ("", "blank")

    # strip quoted nicknames: Kenneth "Kenny" Shults -> Kenneth  Shults
    t = re.sub(r'"[^"]*"', " ", t)
    t = re.sub(r"\u201C[^\u201D]*\u201D", " ", t)
    t = re.sub(r"\s+", " ", t).strip()

    # hard reject explicit multi-person separators
    if "\\" in t or "/" in t:
        return ("", "multi_person_sep")

    # 1) "..., USA) and Rick Reese" -> "Rick Reese"
    m = _RE_TRAIL_AND.search(t)
    if m:
        return (m.group(1).strip(), "trail_and")

    # 2) "Rick Reese- Ft. Collins" -> "Rick Reese"
    # Only treat dash as suffix separator if dash is followed by a space.
    if re.search(r"-\s", t):
        left = t.split("-", 1)[0].strip()
        if left:
            return (left, "dash_suffix")

    # 3) strip parenthetical notes: "Aleksi (FIN) ?" -> "Aleksi ?"
    t2 = re.sub(r"\([^)]*\)", " ", t)
    t2 = re.sub(r"[?]+", " ", t2)
    t2 = re.sub(r"\s+", " ", t2).strip()

    # reject digits in final label
    if any(ch.isdigit() for ch in t2):
        return ("", "has_digits")

    # require 2-4 tokens for results world
    parts = t2.split()
    if not (2 <= len(parts) <= 4):
        return ("", "bad_token_count")

    # reject headings/notes
    if re.search(r"\b(results?|partners|place|points?|victory|scratch)\b", t2, re.IGNORECASE):
        return ("", "heading_or_note")

    return (t2, "strip_parens_punct" if t2 != t else "")


def is_person_like(name: str) -> bool:
    if not name:
        return False
    t = name.strip().lower()

    # obvious junk tokens
    if t in {"()", "na", "nd", "rd", "th"}:
        return False

    # ordinal / result text
    if "position match" in t:
        return False

    # club / org keywords
    if any(k in t for k in ["fc ", "club", "team"]):
        return False

    # locations (already detected by 02p5)
    if t in {"helsinki", "california", "arizona", "quebec"}:
        return False

    return True


# allowed: letters, spaces, hyphens, apostrophes
_RE_ALLOWED_CHARS = re.compile(r"^[A-Za-zÀ-ÖØ-öø-ÿ'’ -]+$")

# disallowed tokens (metadata / junk)
_RE_BAD_TOKENS = re.compile(
    r"""
    \b(
        usa|canada|germany|de|ger|fin|cz|
        victory|points?|scratch|results?|open|
        place|position|playoff|rank|
        pixie|ducking|paradox|swirl|torque|
        and|with|plus
    )\b
    """,
    re.IGNORECASE | re.VERBOSE,
)

# hard separators
_RE_SEPARATORS = re.compile(r"[+/\\=]|🇩🇪|🇫🇮|🇨🇦|🇺🇸")


def is_presentable_person_canon(s: str) -> bool:
    if not isinstance(s, str):
        return False

    t = unicodedata.normalize("NFKC", s).strip()
    if not t:
        return False

    # hard rejects
    if _RE_SEPARATORS.search(t):
        return False
    if any(ch.isdigit() for ch in t):
        return False
    if not _RE_ALLOWED_CHARS.match(t):
        return False
    if _RE_BAD_TOKENS.search(t):
        return False

    parts = t.split()
    if not (2 <= len(parts) <= 4):
        return False

    # Each token must look like a name
    for p in parts:
        if len(p) == 1:
            return False   # single-letter initials not allowed
        if p.isupper() and len(p) <= 3:
            return False   # country codes
    return True


def _tokenize_simple(name: str) -> list[str]:
    t = unicodedata.normalize("NFKC", (name or "")).strip()
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"[\.,;:]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return [x for x in t.split(" ") if x]


_UUID_RE = re.compile(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", re.I)


def _is_uuid(x: str) -> bool:
    return bool(_UUID_RE.match((x or "").strip()))


def _uuid_rate(s: pd.Series) -> float:
    return s.fillna("").astype(str).str.strip().str.match(_UUID_RE).mean()


def _uuid5_person(label: str) -> str:
    return str(uuid.uuid5(_UUID_NS_PERSON, label))


def _uuid5_person_from_canon(canon: str) -> str:
    return _uuid5_person(canon)


def detect_two_people_in_one_slot(
    persons_truth_full: pd.DataFrame,
    pf: pd.DataFrame,
    aliases_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Option A quarantine detector.

    Output: subset of persons_truth_full with added columns:
      - quarantine_reason
      - quarantine_evidence

    Heuristics (deterministic, NO guessing):
      1) Two known person names concatenated into one string (token split check).
      2) Doubles/team competitor_type but missing player2 AND player1 name looks like two names.
    """
    if persons_truth_full.empty:
        return persons_truth_full.copy()

    known: set[str] = set()
    if "person_canon" in persons_truth_full.columns:
        for v in persons_truth_full["person_canon"].fillna("").astype(str):
            vv = v.strip()
            if vv and is_presentable_person_canon(vv):
                known.add(vv)
    if aliases_df is not None and not aliases_df.empty and "person_canon" in aliases_df.columns:
        for v in aliases_df["person_canon"].fillna("").astype(str):
            vv = v.strip()
            if vv and is_presentable_person_canon(vv):
                known.add(vv)

    known_by_toklen: dict[int, set[str]] = {}
    for nm in known:
        toks = tuple(_tokenize_simple(nm))
        if len(toks) >= 2:
            known_by_toklen.setdefault(len(toks), set()).add(" ".join(toks))

    def _split_two_known(name: str) -> tuple[bool, str]:
        toks = _tokenize_simple(name)
        if len(toks) < 4:
            return (False, "")
        for i in range(2, len(toks) - 1):
            left = " ".join(toks[:i])
            right = " ".join(toks[i:])
            if left in known and right in known:
                return (True, f"split_known:{left} || {right}")
            if left in known_by_toklen.get(len(left.split()), set()) and right in known_by_toklen.get(len(right.split()), set()):
                return (True, f"split_toknorm:{left} || {right}")
        return (False, "")

    quarantined_ids: dict[str, tuple[str, str]] = {}

    for r in persons_truth_full.itertuples(index=False):
        eff = str(getattr(r, "effective_person_id", "") or "").strip()
        canon = str(getattr(r, "person_canon", "") or "").strip()
        if not eff or not canon:
            continue
        ok, ev = _split_two_known(canon)
        if ok:
            quarantined_ids[eff] = ("two_people_concat", ev)

    if pf is not None and not pf.empty:
        comp = pf.get("competitor_type", pd.Series([""] * len(pf))).fillna("").astype(str).str.lower()
        is_teamish = comp.str.contains("team|double|pair|doubles", regex=True)

        p2_blank = pf.get("player2_name", pd.Series([""] * len(pf))).fillna("").astype(str).str.strip().eq("")
        p1_name = pf.get("player1_name", pd.Series([""] * len(pf))).fillna("").astype(str).str.strip()

        mask = is_teamish & p2_blank & (p1_name != "")
        if mask.any():
            sub = pf.loc[mask, ["player1_name", "player1_person_id"]].copy()
            for _, row in sub.iterrows():
                name = str(row.get("player1_name", "")).strip()
                pid = str(row.get("player1_person_id", "")).strip()
                if not pid:
                    pid = _uuid5_person(name)
                ok, ev = _split_two_known(name)
                if ok and pid:
                    quarantined_ids.setdefault(pid, ("team_missing_player2_two_people", ev))

    if not quarantined_ids:
        return persons_truth_full.iloc[0:0].copy()

    q = persons_truth_full[persons_truth_full["effective_person_id"].astype(str).str.strip().isin(quarantined_ids.keys())].copy()
    if q.empty:
        return q

    q["quarantine_reason"] = q["effective_person_id"].map(lambda x: quarantined_ids.get(str(x).strip(), ("", ""))[0])
    q["quarantine_evidence"] = q["effective_person_id"].map(lambda x: quarantined_ids.get(str(x).strip(), ("", ""))[1])
    return q


def _mk_truth_row_from_canon(canon: str) -> dict:
    pid = _uuid5_person_from_canon(canon)
    return {
        "effective_person_id": pid,
        "person_canon": canon,
        "player_ids_seen": "",
        "player_names_seen": canon,
        "aliases": "",
        "alias_statuses": "",
        "notes": "Added for QC07 referential integrity (no merges).",
        "source": "coverage_closure",
        "person_canon_clean": canon,
        "person_canon_clean_reason": "coverage_closure",
    }


def load_person_aliases(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=["alias", "person_id", "person_canon", "status", "notes"])
    rows = []
    with path.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            rows.append({
                "alias": _norm(row.get("alias", "")),
                "person_id": _norm(row.get("person_id", "")),
                "person_canon": _norm(row.get("person_canon", "")),
                "status": _norm(row.get("status", "")),
                "notes": _norm(row.get("notes", "")),
            })
    return pd.DataFrame(rows)


def explode_to_people(pf: pd.DataFrame) -> pd.DataFrame:
    """
    Convert Placements_Flat (one row per placement with player1/player2) into
    one row per person appearance, so person stats are easy.

    Team rows become two rows (one for player1, one for player2).
    Individual rows become one row (player1).
    """
    pf = pf.copy()

    # Ensure person columns exist (02p5 should emit them, but be defensive)
    for col in ["player1_person_id", "player1_person_canon", "player2_person_id", "player2_person_canon"]:
        if col not in pf.columns:
            pf[col] = ""
    for col in ["player1_name_clean", "player1_name_raw", "player2_name_clean", "player2_name_raw"]:
        if col not in pf.columns:
            pf[col] = ""

    # identity_source for Persons_Truth filter: override when person_id from 02p5, else fallback
    pf["p1_identity_source"] = pf.apply(
        lambda r: "override" if _norm(r.get("player1_person_id", "")) else "fallback_player_id",
        axis=1,
    )
    pf["p2_identity_source"] = pf.apply(
        lambda r: "override" if _norm(r.get("player2_person_id", "")) else "fallback_player_id",
        axis=1,
    )

    base_cols = [
        "event_id", "year",
        "division_canon", "division_raw", "division_category",
        "competitor_type",
        "place",
    ]
    for c in base_cols:
        if c not in pf.columns:
            pf[c] = ""

    pf["place_int"] = pf["place"].apply(_as_int_place)
    base_cols_with_place_int = base_cols + ["place_int"]

    # Player 1 rows: map person_id <- player1_person_id (UUID), person_canon <- player1_person_canon (name)
    p1 = pf[base_cols_with_place_int + [
        "player1_person_id", "player1_person_canon",
        "player1_id", "player1_name", "player1_name_clean", "player1_name_raw",
        "p1_identity_source", "team_display_name"
    ]].copy()
    p1.rename(columns={
        "player1_person_id": "person_id",
        "player1_person_canon": "person_canon",
        "player1_id": "player_id",
        "player1_name_raw": "player_name_raw",
        "player1_name_clean": "player_name_clean",
        "player1_name": "player_name",
        "p1_identity_source": "identity_source",
    }, inplace=True)
    p1["member_role"] = "player1"

    # Player 2 rows (only if present): map person_id <- player2_person_id, person_canon <- player2_person_canon
    has_p2 = pf["player2_name"].fillna("").astype(str).str.strip().str.len() > 0
    p2 = pf[has_p2][base_cols_with_place_int + [
        "player2_person_id", "player2_person_canon",
        "player2_id", "player2_name", "player2_name_clean", "player2_name_raw",
        "p2_identity_source", "team_display_name"
    ]].copy()
    p2.rename(columns={
        "player2_person_id": "person_id",
        "player2_person_canon": "person_canon",
        "player2_id": "player_id",
        "player2_name_raw": "player_name_raw",
        "player2_name_clean": "player_name_clean",
        "player2_name": "player_name",
        "p2_identity_source": "identity_source",
    }, inplace=True)
    p2["member_role"] = "player2"

    out = pd.concat([p1, p2], ignore_index=True)
    # Fallback: if person_id blank use player_id; if person_canon blank use player_name
    out["person_id"] = out["person_id"].fillna("").map(_norm)
    out["person_canon"] = out["person_canon"].fillna("").map(_norm)
    out.loc[out["person_id"] == "", "person_id"] = out.loc[out["person_id"] == "", "player_id"].fillna("").map(_norm)
    out.loc[out["person_canon"] == "", "person_canon"] = out.loc[out["person_canon"] == "", "player_name"].fillna("").map(_norm)
    out["player_id"] = out["player_id"].fillna("").map(_norm)
    out["player_name"] = out["player_name"].fillna("").map(_norm)
    for c in ["player_name_clean", "player_name_raw"]:
        if c in out.columns:
            out[c] = out[c].fillna("").map(lambda x: _norm(str(x)) if isinstance(x, str) else "")
    if "identity_source" in out.columns:
        out["identity_source"] = out["identity_source"].fillna("").astype(str).str.strip()
    out["division_canon"] = out["division_canon"].fillna("").map(_norm)
    out["division_category"] = out["division_category"].fillna("").map(lambda x: _norm(str(x)) or "unknown")
    return out


def build_person_stats(per: pd.DataFrame) -> pd.DataFrame:
    per = per.copy()
    per["is_win"] = per["place_int"].apply(lambda x: 1 if x == 1 else 0)
    per["is_podium"] = per["place_int"].apply(lambda x: 1 if isinstance(x, int) and 1 <= x <= 3 else 0)
    per["has_place"] = per["place_int"].apply(lambda x: 1 if isinstance(x, int) else 0)

    # unique event participation: person_id + event_id
    grp = per.groupby(["person_id", "person_canon"], dropna=False)

    stats = grp.agg(
        events_competed=("event_id", lambda s: int(pd.Series(s).nunique())),
        placements_total=("event_id", "count"),
        placements_with_numeric_place=("has_place", "sum"),
        wins=("is_win", "sum"),
        podiums=("is_podium", "sum"),
        first_year=("year", lambda s: int(pd.to_numeric(s, errors="coerce").min()) if pd.to_numeric(s, errors="coerce").notna().any() else ""),
        last_year=("year", lambda s: int(pd.to_numeric(s, errors="coerce").max()) if pd.to_numeric(s, errors="coerce").notna().any() else ""),
    ).reset_index()

    # Derived columns
    stats["win_rate"] = (
        stats["wins"] / stats["placements_with_numeric_place"].replace(0, float("nan"))
    ).round(3)

    stats["podium_rate"] = (
        stats["podiums"] / stats["placements_with_numeric_place"].replace(0, float("nan"))
    ).round(3)

    stats["years_active"] = stats.apply(
        lambda r: (int(r["last_year"]) - int(r["first_year"]) + 1)
        if (str(r.get("first_year", "")).strip() and str(r.get("last_year", "")).strip()
            and pd.notna(r.get("first_year")) and pd.notna(r.get("last_year")))
        else 0,
        axis=1,
    )

    # Sort: wins desc, events desc, then name
    stats.sort_values(
        by=["wins", "podiums", "events_competed", "placements_total", "person_canon"],
        ascending=[False, False, False, False, True],
        inplace=True
    )
    return stats


def build_player_stats(per: pd.DataFrame) -> pd.DataFrame:
    per = per.copy()
    per["is_win"] = per["place_int"].apply(lambda x: 1 if x == 1 else 0)
    per["is_podium"] = per["place_int"].apply(lambda x: 1 if isinstance(x, int) and 1 <= x <= 3 else 0)
    grp = per.groupby(["player_id", "player_name"], dropna=False)
    stats = grp.agg(
        events_competed=("event_id", lambda s: int(pd.Series(s).nunique())),
        placements_total=("event_id", "count"),
        wins=("is_win", "sum"),
        podiums=("is_podium", "sum"),
    ).reset_index()

    stats.sort_values(
        by=["wins", "podiums", "events_competed", "placements_total", "player_name"],
        ascending=[False, False, False, False, True],
        inplace=True
    )
    return stats


def build_division_stats(pf: pd.DataFrame, out_dir: Path) -> pd.DataFrame:
    pf = pf.copy()
    pf["place_int"] = pf["place"].apply(_as_int_place)
    pf["is_win"] = pf["place_int"].apply(lambda x: 1 if x == 1 else 0)
    grp = pf.groupby(["division_category", "division_canon"], dropna=False)
    stats = grp.agg(
        events_with_division=("event_id", lambda s: int(pd.Series(s).nunique())),
        placements_total=("event_id", "count"),
        wins_total=("is_win", "sum"),
    ).reset_index()
    stats.sort_values(
        by=["placements_total", "events_with_division", "division_category", "division_canon"],
        ascending=[False, False, True, True],
        inplace=True
    )

    # Attach worst coverage flag per division (lowest ratio seen across all events)
    cov_path = out_dir / "Coverage_ByEventDivision.csv"
    if cov_path.exists():
        cov = pd.read_csv(cov_path, dtype=str).fillna("")
        cov["coverage_ratio"] = pd.to_numeric(cov["coverage_ratio"], errors="coerce")
        cov_agg = (
            cov.groupby(["division_category", "division_canon"], as_index=False)
               ["coverage_ratio"].min()
               .rename(columns={"coverage_ratio": "_min_cov"})
        )
        cov_agg["worst_coverage_flag"] = cov_agg["_min_cov"].map(_coverage_flag)
        cov_agg = cov_agg.drop(columns=["_min_cov"])
        stats = stats.merge(cov_agg, on=["division_category", "division_canon"], how="left")
        stats["worst_coverage_flag"] = stats["worst_coverage_flag"].fillna("")
    else:
        stats["worst_coverage_flag"] = ""

    return stats


def build_person_stats_by_div_category(per: pd.DataFrame) -> pd.DataFrame:
    per = per.copy()
    per["is_win"] = per["place_int"].apply(lambda x: 1 if x == 1 else 0)
    per["is_podium"] = per["place_int"].apply(lambda x: 1 if isinstance(x, int) and 1 <= x <= 3 else 0)

    grp = per.groupby(["person_id", "person_canon", "division_category"], dropna=False)

    stats = grp.agg(
        events_competed=("event_id", lambda s: int(pd.Series(s).nunique())),
        placements_total=("event_id", "count"),
        wins=("is_win", "sum"),
        podiums=("is_podium", "sum"),
        first_year=("year", lambda s: int(pd.to_numeric(s, errors="coerce").min()) if pd.to_numeric(s, errors="coerce").notna().any() else ""),
        last_year=("year", lambda s: int(pd.to_numeric(s, errors="coerce").max()) if pd.to_numeric(s, errors="coerce").notna().any() else ""),
    ).reset_index()

    stats.sort_values(
        by=["wins", "podiums", "events_competed", "placements_total", "person_canon", "division_category"],
        ascending=[False, False, False, False, True, True],
        inplace=True
    )
    return stats


def build_persons_truth(per: pd.DataFrame, aliases_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build Persons_Truth: one row per effective_person_id from placements + overrides.
    NO guessing: person_id comes from 02p5 (or fallback to player_id in explode_to_people).
    Normalizes ID/canon so effective_person_id is always the UUID (or stable player_id), not the name.
    """
    base_cols = ["effective_person_id", "person_canon", "player_ids_seen", "player_names_seen",
                 "aliases", "alias_statuses", "notes", "source", "person_canon_clean", "person_canon_clean_reason"]
    empty_pt = pd.DataFrame(columns=base_cols)

    per = per.copy()
    pid_raw = per["person_id"].fillna("").astype(str).str.strip()
    canon_raw = per["person_canon"].fillna("").astype(str).str.strip()

    def _norm_row_id_canon(pid: str, canon: str) -> tuple[str, str]:
        """Return (effective_id, canon_name). effective_id is always UUID or uuid5(name), never a name."""
        if _is_uuid(pid):
            return pid, canon
        if _is_uuid(canon):
            return canon, pid
        name = canon if canon else pid
        return (_uuid5_person(name), name) if name else ("", "")

    eff_canon = pd.DataFrame(
        [_norm_row_id_canon(p, c) for p, c in zip(pid_raw, canon_raw)],
        index=per.index,
        columns=["_eff_id", "_canon"],
    )
    per["_eff_id"] = eff_canon["_eff_id"]
    per["_canon"] = eff_canon["_canon"]
    per = per[per["_eff_id"].str.len() > 0]
    if per.empty:
        pt = empty_pt.copy()
    else:
        rows = []
        for pid, g in per.groupby("_eff_id", dropna=False):
            pid = str(pid).strip()
            if not pid:
                continue
            canons = g["_canon"].fillna("").astype(str).str.strip()
            canons = canons[canons != ""]
            person_canon = canons.mode().iloc[0] if len(canons) else ""
            if not person_canon:
                pn = g["player_name"].fillna("").astype(str).str.strip()
                pn = pn[pn != ""]
                person_canon = pn.mode().iloc[0] if len(pn) else pid
            player_ids = sorted({str(x).strip() for x in g["player_id"] if str(x).strip()})
            names = sorted({str(x).strip() for x in g["_canon"] if str(x).strip()})
            if not names:
                names = sorted({str(x).strip() for x in g["player_name"] if str(x).strip()})
            identity_source = g["identity_source"].iloc[0] if "identity_source" in g.columns else "fallback_player_id"
            source = "overrides+data" if (str(identity_source).strip() == "override") else "data_only"
            rows.append({
                "effective_person_id": pid,
                "person_canon": person_canon,
                "player_ids_seen": " | ".join(player_ids),
                "player_names_seen": " | ".join(names) if names else person_canon,
                "aliases": "",
                "alias_statuses": "",
                "notes": "",
                "source": source,
                "person_canon_clean": person_canon,
                "person_canon_clean_reason": "",
            })
        pt = pd.DataFrame(rows)

    # Add override-only persons from aliases (person_id in aliases but not in per)
    if not aliases_df.empty and "person_id" in aliases_df.columns:
        existing_ids = set(pt["effective_person_id"].astype(str).str.strip())
        for _, r in aliases_df.iterrows():
            aid = str(r.get("person_id", "")).strip()
            if not aid or aid in existing_ids:
                continue
            acanon = str(r.get("person_canon", "")).strip()
            existing_ids.add(aid)
            pt = pd.concat([
                pt,
                pd.DataFrame([{
                    "effective_person_id": aid,
                    "person_canon": acanon or aid,
                    "player_ids_seen": "",
                    "player_names_seen": acanon or "",
                    "aliases": str(r.get("alias", "")).strip(),
                    "alias_statuses": str(r.get("status", "")).strip(),
                    "notes": str(r.get("notes", "")).strip(),
                    "source": "overrides_only",
                    "person_canon_clean": acanon or aid,
                    "person_canon_clean_reason": "",
                }]),
            ], ignore_index=True)

    # Defensive: ensure effective_person_id is never name-like (fix swapped id/canon rows)
    if len(pt) > 0:
        eff = pt["effective_person_id"].fillna("").astype(str).str.strip()
        canon = pt["person_canon"].fillna("").astype(str).str.strip()
        name_like = eff.map(_looks_like_person)
        if name_like.any():
            sub_canon = canon.loc[name_like]
            fix_vals = sub_canon.where(sub_canon.map(_is_uuid)).fillna(
                sub_canon.map(lambda x: _uuid5_person(x) if x else "")
            )
            pt = pt.copy()
            pt.loc[name_like, "effective_person_id"] = fix_vals
        # Deduplicate by effective_person_id (fix can create duplicates); prefer non-UUID for canon
        def _best_canon(s):
            vals = [str(x).strip() for x in s if x and str(x).strip()]
            non_uuid = [v for v in vals if not _is_uuid(v)]
            return (non_uuid[0] if non_uuid else (vals[0] if vals else ""))

        pt = pt.groupby("effective_person_id", as_index=False).agg(
            person_canon=("person_canon", _best_canon),
            player_ids_seen=("player_ids_seen", lambda s: " | ".join(sorted({x for v in s if v for x in str(v).split(" | ")}))),
            player_names_seen=("player_names_seen", lambda s: " | ".join(sorted({x for v in s if v for x in str(v).split(" | ")}))),
            aliases=("aliases", lambda s: " | ".join({x for v in s if v for x in str(v).split(" | ")})),
            alias_statuses=("alias_statuses", lambda s: s.iloc[0] if len(s) else ""),
            notes=("notes", lambda s: s.iloc[0] if len(s) else ""),
            source=("source", lambda s: "overrides+data" if (s == "overrides+data").any() else s.iloc[0]),
            person_canon_clean=("person_canon_clean", _best_canon),
            person_canon_clean_reason=("person_canon_clean_reason", lambda s: s.iloc[0] if len(s) else ""),
        )
        # Where person_canon is still UUID (no name in group), use first name from player_names_seen or placeholder
        canon = pt["person_canon"].fillna("").astype(str).str.strip()
        uuid_canon = canon.map(_is_uuid)
        if uuid_canon.any():
            idx = pt.index[uuid_canon].tolist()
            names_seen = pt.loc[idx, "player_names_seen"].fillna("").astype(str).str.strip()

            def first_non_uuid(s: str) -> str:
                parts = [p.strip() for p in str(s).split(" | ") if p.strip()]
                for p in parts:
                    if not _is_uuid(p):
                        return p
                return "Unknown"

            vals = [first_non_uuid(n) for n in names_seen]
            pt.loc[idx, "person_canon"] = vals
            pt.loc[idx, "person_canon_clean"] = vals

    return pt


def qc_persons_truth(pt: pd.DataFrame) -> None:
    """
    HARD QC gate for Persons_Truth.

    Guarantees (definitive, no guessing):
      - effective_person_id is present, non-empty, unique
      - effective_person_id is NOT name-like (should be UUID or legacy player_id)
      - person_canon is present and NOT UUID-like
      - source (if present) is within expected enum
    """
    required = {"effective_person_id", "person_canon"}
    missing = sorted(required - set(pt.columns))
    if missing:
        raise ValueError(f"Persons_Truth missing required columns: {missing}")

    # Non-empty IDs
    eff = pt["effective_person_id"].fillna("").astype(str).str.strip()
    if (eff == "").any():
        raise ValueError(
            f"Persons_Truth has {(eff == '').sum()} blank effective_person_id values"
        )

    # Uniqueness
    dup = eff.duplicated(keep=False)
    if dup.any():
        sample = pt.loc[dup, ["effective_person_id", "person_canon"]].head(20)
        raise ValueError(
            "Persons_Truth has duplicate effective_person_id values:\n"
            + sample.to_string(index=False)
        )

    # IDs must NOT look like person names
    name_like_ids = eff.map(_looks_like_person)
    if name_like_ids.any():
        sample = pt.loc[name_like_ids, ["effective_person_id", "person_canon"]].head(30)
        raise ValueError(
            f"Persons_Truth has {int(name_like_ids.sum())} name-like IDs "
            "(IDs must be UUIDs or stable player_ids).\n"
            f"Sample:\n{sample.to_string(index=False)}"
        )

    # Canon names must NOT be UUIDs (catch swapped columns)
    canon = pt["person_canon"].fillna("").astype(str).str.strip()
    uuidish_canon = canon.map(_is_uuid)
    if uuidish_canon.any():
        sample = pt.loc[uuidish_canon, ["effective_person_id", "person_canon"]].head(30)
        raise ValueError(
            f"Persons_Truth has {int(uuidish_canon.sum())} UUID-like person_canon values "
            "(likely swapped columns).\n"
            f"Sample:\n{sample.to_string(index=False)}"
        )

    # Source enum check (if present)
    if "source" in pt.columns:
        allowed = {"data_only", "overrides_only", "overrides+data", "coverage_closure"}
        bad = ~pt["source"].fillna("").astype(str).isin(allowed)
        if bad.any():
            sample = pt.loc[bad, ["effective_person_id", "person_canon", "source"]].head(30)
            raise ValueError(
                f"Persons_Truth has unexpected source values. "
                f"Allowed={sorted(allowed)}.\n"
                f"Sample:\n{sample.to_string(index=False)}"
            )


def _looks_like_person(name: str) -> bool:
    """
    Heuristic for Excel diagnostics only.
    NO GUESSING: this does not merge identities; it only helps prioritize review.
    """
    s = _norm(name)
    if not s:
        return False
    # must have at least 2 tokens to look like "First Last"
    if len(s.split()) < 2:
        return False
    low = s.lower()
    # common non-person tokens seen in Top_Unmapped_Names
    if low in {"na", "dnf", "()", "nd", "th"}:
        return False
    # obvious non-person phrases
    bad_sub = ["club", "footbag", "position", "match", "results", "team", "canada", "usa"]
    if any(b in low for b in bad_sub):
        return False
    return True


def build_top_unmapped_names(pf: pd.DataFrame, limit: int = 200) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = []

    for side in ["player1", "player2"]:
        name_col = f"{side}_name"
        pid_col = f"{side}_person_id"

        if name_col not in pf.columns or pid_col not in pf.columns:
            continue

        sub = pf[
            (pf[name_col].fillna("").str.strip() != "") &
            (pf[pid_col].fillna("").str.strip() == "")
        ]

        counts = sub[name_col].value_counts()

        for name, cnt in counts.items():
            rows.append({
                "name": name,
                "appearances": cnt,
                f"as_{side}": cnt,
            })

    if not rows:
        empty = pd.DataFrame(columns=["name", "appearances", "as_player1", "as_player2"])
        return empty, empty.copy()

    df = pd.DataFrame(rows)

    df = (
        df
        .groupby("name", as_index=False)
        .agg(
            appearances=("appearances", "sum"),
            as_player1=("as_player1", "sum"),
            as_player2=("as_player2", "sum"),
        )
        .sort_values(by=["appearances", "name"], ascending=[False, True])
        .reset_index(drop=True)
    )

    df["personlike"] = df["name"].map(_looks_like_person)
    personlike = df[df["personlike"]].drop(columns=["personlike"]).head(limit).reset_index(drop=True)
    noise = df[~df["personlike"]].drop(columns=["personlike"]).head(limit).reset_index(drop=True)
    return personlike, noise


def build_coverage_by_event_division(
    pf: pd.DataFrame,          # Gate-1-filtered Placements_Flat (in memory)
    out_dir: Path,
    quarantine_path: Path | None = None,
) -> pd.DataFrame:
    """
    Coverage definition:
      For each (event_id, year, division_canon), compute:
        placements_present = count of distinct place values present
        min_place, max_place (numeric)
        expected_span = max_place - min_place + 1
        missing_places = expected_span - placements_present
        coverage_ratio = placements_present / expected_span

    IMPORTANT:
      - Uses only "clean" placements by default.
      - If quarantine_path is provided and exists, removes those rows from the surface
        using an (event_id, division_canon, place, player/team key) join.
      - No guessing, no inference of missing results.
    """

    df = pf.copy()
    # Standardize key columns (defensive) — ensure string dtype
    for c in ["event_id", "year", "division_canon", "division_category", "place",
              "competitor_type", "player1_name", "player2_name", "team_display_name"]:
        if c not in df.columns:
            df[c] = ""
        else:
            df[c] = df[c].fillna("").astype(str)

    # Parse place as int where possible (ignore non-numeric places)
    df["place_num"] = pd.to_numeric(df["place"], errors="coerce")

    # Keep only rows with a numeric place (coverage is defined on ordinal places)
    df = df[df["place_num"].notna()].copy()
    df["place_num"] = df["place_num"].astype(int)

    # Optionally exclude quarantined rows if a quarantine file exists.
    # This makes coverage reflect the analytics surface, not the diagnostic set.
    if quarantine_path is not None and Path(quarantine_path).exists():
        q = pd.read_csv(quarantine_path, dtype=str).fillna("")
        for c in ["event_id", "division_canon", "division_category", "place",
                  "competitor_type", "player1_name", "player2_name", "team_display_name"]:
            if c not in q.columns:
                q[c] = ""
        q["place_num"] = pd.to_numeric(q["place"], errors="coerce")
        q = q[q["place_num"].notna()].copy()
        q["place_num"] = q["place_num"].astype(int)

        # Build a conservative row identity key. We do NOT use IDs (since presentation-clean).
        # This aims to remove only the exact quarantined rows, not "similar" ones.
        def row_key(d: pd.DataFrame) -> pd.Series:
            # Prefer team_display_name when competitor_type is team; else use player1|player2
            teamish = (d["competitor_type"].str.lower() == "team")
            key = d["player1_name"].str.strip() + " | " + d["player2_name"].str.strip()
            key = key.where(~teamish, d["team_display_name"].str.strip())
            return key.str.strip()

        df["_rk"] = row_key(df)
        q["_rk"] = row_key(q)

        q_key = q[["event_id", "division_canon", "place_num", "_rk"]].copy()
        q_key["_is_quarantined"] = 1

        # Left join to mark quarantined
        merged = df.merge(
            q_key.drop_duplicates(),
            on=["event_id", "division_canon", "place_num", "_rk"],
            how="left",
        )
        merged["_is_quarantined"] = merged["_is_quarantined"].fillna(0).astype(int)
        df = merged[merged["_is_quarantined"] == 0].copy()
        df.drop(columns=["_rk", "_is_quarantined"], inplace=True, errors="ignore")
    else:
        df.drop(columns=["_rk"], inplace=True, errors="ignore")

    # Aggregate coverage by (event_id, year, division_canon)
    grp_cols = ["event_id", "year", "division_canon", "division_category"]
    cov = (
        df.groupby(grp_cols, dropna=False)
          .agg(
              placements_present=("place_num", lambda s: int(pd.Series(s).nunique())),
              min_place=("place_num", "min"),
              max_place=("place_num", "max"),
          )
          .reset_index()
    )

    cov["expected_span"] = (cov["max_place"] - cov["min_place"] + 1).astype(int)
    cov["missing_places"] = (cov["expected_span"] - cov["placements_present"]).astype(int)

    # Avoid division by zero (shouldn't happen, but keep deterministic)
    cov["coverage_ratio"] = cov.apply(
        lambda r: (r["placements_present"] / r["expected_span"]) if r["expected_span"] > 0 else 0.0,
        axis=1
    )

    # Sort for readability
    cov = cov.sort_values(["year", "event_id", "division_category", "division_canon"], kind="mergesort")

    # Add coverage flag (self-contained: consumers don't need to re-implement thresholds)
    cov["coverage_flag"] = cov["coverage_ratio"].map(_coverage_flag)

    # Write CSV output
    out_path = out_dir / "Coverage_ByEventDivision.csv"
    cov.to_csv(out_path, index=False)
    print(f"Wrote: {out_path} ({len(cov)} rows)")

    return cov


def build_placements_by_person_clean(
    pf: pd.DataFrame,      # Gate-1-filtered Placements_Flat (with person_id columns)
    cov_df: pd.DataFrame,  # from build_coverage_by_event_division
    out_dir: Path,
) -> pd.DataFrame:
    """
    Build clean Placements_ByPerson following canonical sequence:
    Step 2: person_id already applied (02p5 did the LEFT JOIN)
    Step 3: canonicalize competitor identity; build team_person_key
    Step 4: collapse duplicate player-token rows per competitor identity
    Step 5: apply coverage_flag
    Step 6: output clean schema
    """
    df = pf.copy()

    # Defensive: ensure person_id columns exist and are clean strings
    for col in ["player1_person_id", "player1_person_canon",
                "player2_person_id", "player2_person_canon"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str).str.strip()

    # Ensure base columns exist
    for col in ["event_id", "year", "division_canon", "division_category",
                "place", "competitor_type", "team_display_name"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str).str.strip()

    is_team = df["competitor_type"].str.lower() == "team"

    # --- Step 3: Canonicalize competitor identity ---

    # Singles: identity = player1_person_id
    df.loc[~is_team, "person_id"] = df.loc[~is_team, "player1_person_id"]
    df.loc[~is_team, "person_canon"] = df.loc[~is_team, "player1_person_canon"]
    df.loc[~is_team, "team_person_key"] = ""
    df.loc[~is_team, "team_display_name"] = ""

    # Teams: stable key = sorted(p1_id, p2_id) joined by "|"
    def _team_key(row):
        parts = sorted(x for x in [row["player1_person_id"], row["player2_person_id"]] if x)
        return "|".join(parts) if parts else ""

    df.loc[is_team, "team_person_key"] = df[is_team].apply(_team_key, axis=1)
    df.loc[is_team, "person_id"] = ""

    # Team display: use existing team_display_name if present, else build from canons
    def _team_display(row):
        existing = str(row.get("team_display_name", "")).strip()
        if existing:
            return existing
        p1 = str(row["player1_person_canon"]).strip()
        p2 = str(row["player2_person_canon"]).strip()
        return f"{p1} / {p2}" if p2 else p1

    df.loc[is_team, "team_display_name"] = df[is_team].apply(_team_display, axis=1)
    df.loc[is_team, "person_canon"] = ""

    # --- Step 4: Deduplicate by competitor identity ---
    # Group key: person_id for singles, team_person_key for teams
    df["_group_key"] = df.apply(
        lambda r: r["team_person_key"] if str(r.get("competitor_type", "")).lower() == "team"
                  else r["person_id"],
        axis=1,
    )

    group_cols = ["event_id", "division_canon", "place", "_group_key"]
    df["_place_int"] = df["place"].apply(_as_int_place)
    df = df.sort_values(["event_id", "division_canon", "_place_int", "_group_key"])
    df = df.drop_duplicates(subset=group_cols, keep="first")
    df = df.drop(columns=["_group_key", "_place_int"], errors="ignore")

    # --- Step 5: Apply coverage_flag ---
    if not cov_df.empty:
        _cov = cov_df[["event_id", "division_canon", "coverage_flag"]].drop_duplicates()
        df = df.drop(columns=["coverage_flag"], errors="ignore")
        df = df.merge(_cov, on=["event_id", "division_canon"], how="left")
        df["coverage_flag"] = df["coverage_flag"].fillna("")
    else:
        df["coverage_flag"] = ""

    # person_unresolved flag
    df["person_unresolved"] = ""
    _is_team2 = df["competitor_type"].fillna("").str.strip().str.lower() == "team"
    df.loc[~_is_team2 & (df["person_id"].fillna("") == ""), "person_unresolved"] = "true"
    df.loc[_is_team2 & (df["team_person_key"].fillna("") == ""), "person_unresolved"] = "true"

    # --- Step 6: Output clean schema ---
    OUTPUT_COLS = [
        "event_id", "year", "division_canon", "division_category",
        "place", "competitor_type",
        "person_id", "team_person_key", "person_canon", "team_display_name",
        "coverage_flag", "person_unresolved",
    ]
    for c in OUTPUT_COLS:
        if c not in df.columns:
            df[c] = ""
    out = df[OUTPUT_COLS].copy()
    out_path = out_dir / "Placements_ByPerson.csv"
    out.to_csv(out_path, index=False)
    n_unresolved = out["person_unresolved"].eq("true").sum()
    print(f"Wrote: {out_path} ({len(out)} rows, {n_unresolved} unresolved)")
    return out


def build_persons_unresolved(
    pf: pd.DataFrame,
    per_all: pd.DataFrame,
    out_dir: Path,
) -> pd.DataFrame:
    """
    Build Persons_Unresolved: identity punch list surfacing unmapped players,
    excluded persons, and multi-person collisions needing human review.

    Output columns:
      player_id | name_raw | name_clean | person_id | person_canon |
      issue_type | appearances | evidence | suggested_action
    """
    COLS = ["player_id", "name_raw", "name_clean", "person_id", "person_canon",
            "issue_type", "appearances", "evidence", "suggested_action"]
    parts = []

    # 1. Unmapped player tokens: rows where player1_person_id is blank
    p1_pid = pf.get("player1_person_id", pd.Series([""] * len(pf))).fillna("").astype(str).str.strip()
    unmapped = pf[p1_pid == ""].copy()
    if not unmapped.empty:
        for col in ["player1_id", "player1_name_raw", "player1_name_clean"]:
            if col not in unmapped.columns:
                unmapped[col] = ""
            unmapped[col] = unmapped[col].fillna("").astype(str).str.strip()
        grp = (
            unmapped
            .groupby(["player1_id", "player1_name_raw", "player1_name_clean"], dropna=False)
            .agg(appearances=("event_id", "count"))
            .reset_index()
            .rename(columns={
                "player1_id": "player_id",
                "player1_name_raw": "name_raw",
                "player1_name_clean": "name_clean",
            })
        )
        grp["person_id"] = ""
        grp["person_canon"] = ""
        grp["issue_type"] = "unmapped_player_id"
        grp["evidence"] = "no alias in person_aliases.csv"
        grp["suggested_action"] = grp["appearances"].map(
            lambda n: "add alias" if n > 1 else "accept as unresolved (single appearance)"
        )
        parts.append(grp)

    # 2. Excluded persons from Persons_Truth_Excluded.csv
    excluded_path = out_dir / "Persons_Truth_Excluded.csv"
    if excluded_path.exists():
        excl = pd.read_csv(excluded_path, dtype=str).fillna("")
        if not excl.empty and "effective_person_id" in excl.columns:
            exclude_reason_map = {
                "synthetic_sparse_single":   ("sparse_coverage_only",  "ignore (non-analytic)"),
                "two_people_quarantine":     ("multi_person_collision", "split or merge person_ids"),
                "synthetic_quarantine_only": ("quarantine_only",        "accept as unresolved"),
                "not_presentable_strict":    ("not_presentable",        "fix name or add alias"),
                "duplicate_person_canon":    ("duplicate_display_name", "resolve in person_aliases.csv"),
            }
            app_map: dict[str, int] = {}
            if not per_all.empty and "person_id" in per_all.columns:
                _app = (
                    per_all.groupby("person_id", dropna=False)
                    .size()
                    .reset_index(name="n")
                )
                app_map = dict(zip(_app["person_id"].astype(str).str.strip(), _app["n"]))
            rows = []
            for _, r in excl.iterrows():
                pid = str(r.get("effective_person_id", "")).strip()
                canon = str(r.get("person_canon", "")).strip()
                reason = str(r.get("exclude_reason", "")).strip()
                issue_type, suggested_action = exclude_reason_map.get(reason, (reason, "review manually"))
                rows.append({
                    "player_id": "",
                    "name_raw": "",
                    "name_clean": "",
                    "person_id": pid,
                    "person_canon": canon,
                    "issue_type": issue_type,
                    "appearances": app_map.get(pid, 0),
                    "evidence": reason,
                    "suggested_action": suggested_action,
                })
            if rows:
                parts.append(pd.DataFrame(rows))

    # 3. Multi-person quarantine from Persons_Truth_Quarantine_TwoPeople.csv
    qua_path = out_dir / "Persons_Truth_Quarantine_TwoPeople.csv"
    if qua_path.exists():
        qua = pd.read_csv(qua_path, dtype=str).fillna("")
        if not qua.empty and "effective_person_id" in qua.columns:
            app_map2: dict[str, int] = {}
            if not per_all.empty and "person_id" in per_all.columns:
                _app2 = (
                    per_all.groupby("person_id", dropna=False)
                    .size()
                    .reset_index(name="n")
                )
                app_map2 = dict(zip(_app2["person_id"].astype(str).str.strip(), _app2["n"]))
            rows2 = []
            for _, r in qua.iterrows():
                pid = str(r.get("effective_person_id", "")).strip()
                canon = str(r.get("person_canon", "")).strip()
                evidence = str(r.get("quarantine_evidence", "")).strip()
                rows2.append({
                    "player_id": "",
                    "name_raw": "",
                    "name_clean": "",
                    "person_id": pid,
                    "person_canon": canon,
                    "issue_type": "multi_person_collision",
                    "appearances": app_map2.get(pid, 0),
                    "evidence": evidence,
                    "suggested_action": "split person_ids manually",
                })
            if rows2:
                parts.append(pd.DataFrame(rows2))

    if not parts:
        out = pd.DataFrame(columns=COLS)
    else:
        out = pd.concat(parts, ignore_index=True)
        for c in COLS:
            if c not in out.columns:
                out[c] = ""
        out = out[COLS]
        out["appearances"] = pd.to_numeric(out["appearances"], errors="coerce").fillna(0).astype(int)
        out = out.sort_values(["issue_type", "appearances"], ascending=[True, False])

    out_path = out_dir / "Persons_Unresolved.csv"
    out.to_csv(out_path, index=False)
    print(f"Wrote: {out_path} ({len(out)} rows)")
    return out


def build_placements_unresolved(
    placements_by_person_df: pd.DataFrame,
    out_dir: Path,
) -> pd.DataFrame:
    """
    Build Placements_Unresolved: excluded/unresolved placements with recovery scope.

    Output columns:
      event_id | year | division_canon | place | competitor_type |
      name_display | reason_excluded | recovery_candidate
    """
    base_cols = ["event_id", "year", "division_canon", "place", "competitor_type",
                 "name_display", "reason_excluded", "recovery_candidate"]
    parts = []

    # 1. In-analytics but identity-missing (person_unresolved == "true")
    if not placements_by_person_df.empty and "person_unresolved" in placements_by_person_df.columns:
        unres = placements_by_person_df[
            placements_by_person_df["person_unresolved"].fillna("").str.strip() == "true"
        ].copy()
        if not unres.empty:
            def _name_unres(r):
                name = str(r.get("person_canon", "") or "").strip()
                if not name:
                    name = str(r.get("team_display_name", "") or "").strip()
                return name
            unres["name_display"] = unres.apply(_name_unres, axis=1)
            unres["reason_excluded"] = "unmapped_identity"
            unres["recovery_candidate"] = unres["place"].apply(
                lambda p: "yes" if _as_int_place(p) is not None else "no"
            )
            for c in ["event_id", "year", "division_canon", "place", "competitor_type"]:
                if c not in unres.columns:
                    unres[c] = ""
            parts.append(unres[[c for c in base_cols if c in unres.columns or True]].reindex(columns=base_cols, fill_value=""))

    # 2. Rejected placements
    rej_path = out_dir / "Placements_ByPerson_Rejected.csv"
    if rej_path.exists():
        rej = pd.read_csv(rej_path, dtype=str).fillna("")
        if not rej.empty:
            def _name_rej(r):
                name = str(r.get("player1_name_clean", "") or "").strip()
                if not name:
                    name = str(r.get("player1_name_raw", "") or "").strip()
                return name
            rej["name_display"] = rej.apply(_name_rej, axis=1)
            rej["reason_excluded"] = "rejected_missing_id"
            rej["recovery_candidate"] = rej.apply(
                lambda r: "yes" if (_as_int_place(r.get("place", "")) is not None
                                    and str(r.get("name_display", "")).strip()) else "no",
                axis=1,
            )
            for c in ["event_id", "year", "division_canon", "place", "competitor_type"]:
                if c not in rej.columns:
                    rej[c] = ""
            parts.append(rej.reindex(columns=base_cols, fill_value=""))

    # 3. Unpresentable placements
    exc_path = out_dir / "qc" / "excluded_results_rows_unpresentable.csv"
    if exc_path.exists():
        exc = pd.read_csv(exc_path, dtype=str).fillna("")
        if not exc.empty:
            exc["name_display"] = exc.get("player1_name_raw", pd.Series([""] * len(exc))).fillna("").astype(str).str.strip()
            exc["reason_excluded"] = "unpresentable"
            exc["recovery_candidate"] = "no"
            for c in ["event_id", "year", "division_canon", "place", "competitor_type"]:
                if c not in exc.columns:
                    exc[c] = ""
            parts.append(exc.reindex(columns=base_cols, fill_value=""))

    if not parts:
        out = pd.DataFrame(columns=base_cols)
    else:
        out = pd.concat(parts, ignore_index=True)
        for c in base_cols:
            if c not in out.columns:
                out[c] = ""
        out = out[base_cols]
        out = out.drop_duplicates(subset=["event_id", "division_canon", "place", "name_display"])
        _year_int = pd.to_numeric(out["year"], errors="coerce").fillna(0).astype(int)
        _place_int = out["place"].apply(lambda p: _as_int_place(p) if _as_int_place(p) is not None else 9999)
        out = out.iloc[(-_year_int).argsort(kind="stable")]
        out = out.reset_index(drop=True)

    out_path = out_dir / "Placements_Unresolved.csv"
    out.to_csv(out_path, index=False)
    print(f"Wrote: {out_path} ({len(out)} rows)")
    return out


def build_analytics_safe_surface(
    placements_by_person_df: pd.DataFrame,
    out_dir: Path,
) -> pd.DataFrame:
    """
    Pre-filtered pivot-ready surface: coverage complete/mostly_complete,
    identity-locked (person_unresolved != true).
    This is the recommended default pivot source for external analysis.
    """
    _complete_flags = {"complete", "mostly_complete"}
    df = placements_by_person_df[
        placements_by_person_df["coverage_flag"].isin(_complete_flags) &
        (placements_by_person_df["person_unresolved"].fillna("").str.strip() != "true")
    ].copy()

    OUTPUT_COLS = [
        "year", "division_category", "division_canon",
        "place", "competitor_type",
        "person_canon", "team_display_name",
        "coverage_flag",
    ]
    for c in OUTPUT_COLS:
        if c not in df.columns:
            df[c] = ""
    df = df[OUTPUT_COLS].copy()

    df["_year_int"] = pd.to_numeric(df["year"], errors="coerce").fillna(0).astype(int)
    df["_place_int"] = df["place"].apply(_as_int_place).apply(lambda x: x if x is not None else 9999)
    df = df.sort_values(["_year_int", "division_category", "division_canon", "_place_int"],
                        ascending=[False, True, True, True])
    df = df.drop(columns=["_year_int", "_place_int"])

    out_path = out_dir / "Analytics_Safe_Surface.csv"
    df.to_csv(out_path, index=False)
    print(f"Wrote: {out_path} ({len(df)} rows, coverage-filtered + identity-locked)")
    return df


def build_data_integrity(
    pf_raw_count: int,
    pf: pd.DataFrame,
    placements_by_person_df: pd.DataFrame,
    analytics_safe_df: pd.DataFrame,
    cov_df: pd.DataFrame,
    out_dir: Path,
) -> pd.DataFrame:
    """
    Build Data_Integrity metrics registry sheet.
    Reads Persons_Truth_Excluded.csv from out_dir for Gate 3 drop breakdown.
    """
    rows = []

    def _row(category, metric, value, notes=""):
        return {"category": category, "metric": metric, "value": value, "notes": notes}

    # Placements
    rows.append(_row("Placements", "Total in source (raw)", pf_raw_count, "Before Gate 1"))
    rows.append(_row("Placements", "Surviving Gate 1", len(pf), "Rejected + unpresentable removed"))
    n_unresolved = placements_by_person_df["person_unresolved"].eq("true").sum()
    rows.append(_row("Placements", "In Analytics_Safe_Surface", len(analytics_safe_df),
                     "Coverage-filtered + identity-locked"))
    rows.append(_row("Placements", "Unresolved identity", n_unresolved,
                     "person_unresolved == true"))

    # Persons
    excl_path = out_dir / "Persons_Truth_Excluded.csv"
    pt_path = out_dir / "Persons_Truth.csv"
    n_persons = 0
    if pt_path.exists():
        _pt = pd.read_csv(pt_path, dtype=str).fillna("")
        n_persons = len(_pt)
    rows.append(_row("Persons", "Total (Gate 3)", n_persons,
                     "Presentable, non-synthetic, non-duplicate"))
    if excl_path.exists():
        excl = pd.read_csv(excl_path, dtype=str).fillna("")
        rows.append(_row("Persons", "Excluded in Gate 3", len(excl), "All exclusion reasons"))
        by_reason = excl.groupby("exclude_reason").size().reset_index(name="count")
        for _, r in by_reason.iterrows():
            rows.append(_row("Persons (excluded)", r["exclude_reason"], int(r["count"]), ""))

    # Coverage
    if not cov_df.empty and "coverage_flag" in cov_df.columns:
        n_total = len(cov_df)
        for flag in ["complete", "mostly_complete", "partial", "sparse"]:
            n = (cov_df["coverage_flag"] == flag).sum()
            rows.append(_row("Coverage", flag, n,
                             f"{n/n_total*100:.1f}% of event/division combinations"))
        n_ok = ((cov_df["coverage_flag"] == "complete") |
                (cov_df["coverage_flag"] == "mostly_complete")).sum()
        rows.append(_row("Coverage", "Analytic-safe %",
                         f"{n_ok/n_total*100:.1f}%",
                         "complete + mostly_complete"))

    out = pd.DataFrame(rows)
    out_path = out_dir / "Data_Integrity.csv"
    out.to_csv(out_path, index=False)
    print(f"Wrote: {out_path} ({len(out)} rows)")
    return out


def build_coverage_gap_priority(
    cov_df: pd.DataFrame,
    out_dir: Path,
) -> pd.DataFrame:
    """
    Classify coverage gaps by recoverability using quarantine/rejected evidence.

    Gap classes:
      recoverable      — ratio 0.4–0.8, missing>=2, quarantine evidence exists
      possibly_recoverable — ratio 0.2–0.4, little/no quarantine evidence
      not_recoverable  — ratio <=0.2, no quarantine, consistently sparse
      document_only    — ratio >=0.8 but <1.0, or missing_places==1

    Practical rule: work upstream only when ratio>=0.4, missing>=2, evidence exists.
    """
    # Filter to gaps only
    gaps = cov_df[cov_df["coverage_ratio"] < 1.0].copy()
    if gaps.empty:
        empty = pd.DataFrame(columns=[
            "event_id", "year", "division_canon", "division_category",
            "placements_present", "expected_span", "missing_places", "coverage_ratio",
            "quarantine_rows", "rejected_rows", "excluded_rows",
            "gap_class", "priority_score",
        ])
        out_path = out_dir / "Coverage_GapPriority.csv"
        empty.to_csv(out_path, index=False)
        print(f"Wrote: {out_path} (0 rows — no gaps)")
        return empty

    gaps["event_id"] = gaps["event_id"].astype(str).str.strip()
    gaps["division_canon"] = gaps["division_canon"].astype(str).str.strip()

    # Load evidence sources
    def _load_evidence(path: Path) -> pd.DataFrame:
        if not path.exists():
            return pd.DataFrame(columns=["event_id", "division_canon"])
        df = pd.read_csv(path, dtype=str).fillna("")
        for c in ["event_id", "division_canon"]:
            if c not in df.columns:
                df[c] = ""
            df[c] = df[c].astype(str).str.strip()
        return df

    rejected = _load_evidence(out_dir / "Placements_ByPerson_Rejected.csv")
    quarantine = _load_evidence(out_dir / "Placements_ByPerson_SinglesQuarantine.csv")
    excluded = _load_evidence(out_dir / "qc" / "excluded_results_rows_unpresentable.csv")

    # Count evidence per (event_id, division_canon)
    def _count_by_key(df: pd.DataFrame) -> dict:
        if df.empty:
            return {}
        counts = df.groupby(["event_id", "division_canon"]).size()
        return {(str(eid), str(div)): int(cnt) for (eid, div), cnt in counts.items()}

    rej_counts = _count_by_key(rejected)
    qua_counts = _count_by_key(quarantine)
    exc_counts = _count_by_key(excluded)

    def _lookup(counts, eid, div):
        return counts.get((eid, div), 0)

    gaps["rejected_rows"] = gaps.apply(lambda r: _lookup(rej_counts, r["event_id"], r["division_canon"]), axis=1)
    gaps["quarantine_rows"] = gaps.apply(lambda r: _lookup(qua_counts, r["event_id"], r["division_canon"]), axis=1)
    gaps["excluded_rows"] = gaps.apply(lambda r: _lookup(exc_counts, r["event_id"], r["division_canon"]), axis=1)
    gaps["evidence_total"] = gaps["rejected_rows"] + gaps["quarantine_rows"] + gaps["excluded_rows"]

    # Classify
    def _classify(r):
        ratio = r["coverage_ratio"]
        missing = r["missing_places"]
        evidence = r["evidence_total"]

        if ratio >= 0.4 and missing >= 2 and evidence > 0:
            return "recoverable"
        if 0.2 <= ratio < 0.4:
            return "possibly_recoverable"
        if ratio < 0.2:
            return "not_recoverable"
        # ratio >= 0.4 but no evidence, or missing < 2
        return "document_only"

    gaps["gap_class"] = gaps.apply(_classify, axis=1)

    # Priority score: higher = more worth fixing
    # Factors: missing_places (volume), coverage_ratio (inversely), evidence_total
    gaps["priority_score"] = (
        gaps["missing_places"] * (1 - gaps["coverage_ratio"]) * (1 + gaps["evidence_total"].clip(upper=10))
    ).round(1)

    # Sort: recoverable first, then by priority_score descending
    class_order = {"recoverable": 0, "possibly_recoverable": 1, "document_only": 2, "not_recoverable": 3}
    gaps["_class_order"] = gaps["gap_class"].map(class_order)
    gaps.sort_values(["_class_order", "priority_score"], ascending=[True, False], inplace=True)
    gaps.drop(columns=["_class_order"], inplace=True)

    out_path = out_dir / "Coverage_GapPriority.csv"
    gaps.to_csv(out_path, index=False)

    # Summary
    for cls in ["recoverable", "possibly_recoverable", "document_only", "not_recoverable"]:
        subset = gaps[gaps["gap_class"] == cls]
        if len(subset) > 0:
            print(f"  {cls}: {len(subset)} gaps, {int(subset['missing_places'].sum())} missing places")

    print(f"Wrote: {out_path} ({len(gaps)} rows)")
    return gaps


SHEET_RENAMES = {"Person_Stats_ByDivisionCategory": "PersonStats_ByDivCat"}


def write_sheets_append(xlsx_path: Path, sheets: list[Tuple[str, pd.DataFrame]], readme_df: pd.DataFrame | None = None) -> None:
    # Append mode with replace semantics for these sheets
    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
        for name, df in sheets:
            sheet_name = SHEET_RENAMES.get(name, name)
            df.to_excel(xw, sheet_name=sheet_name, index=False)

            # ---- Excel usability formatting ----
            ws = xw.book[sheet_name]

            # Freeze header row
            ws.freeze_panes = "A2"

            # AutoFilter over the written range
            ws.auto_filter.ref = ws.dimensions

            # Autosize columns based on header + first N rows
            max_rows_scan = min(len(df), 200)
            for col_idx, col_name in enumerate(df.columns, start=1):
                letter = get_column_letter(col_idx)

                # measure header + sample rows
                best = len(str(col_name))
                if max_rows_scan > 0:
                    series = df[col_name].head(max_rows_scan)
                    for v in series:
                        s = "" if v is None else str(v)
                        if len(s) > best:
                            best = len(s)

                # set width with caps
                width = max(10, min(best + 2, 60))
                ws.column_dimensions[letter].width = width

            # Wrap text for very long narrative columns (keeps width sane)
            wrap_cols = {"examples", "divisions_seen", "divisions_top", "player_names"}
            for col_idx, col_name in enumerate(df.columns, start=1):
                if str(col_name) in wrap_cols:
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Make header row slightly nicer
            for cell in ws[1]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Gate 3: Persons_Truth banner row + sheet protection (applied unconditionally)
            if sheet_name == "Persons_Truth":
                ws.insert_rows(2)
                banner_text = "IDENTITY LOCKED — One row per real person. All analytics derive identity from this sheet only."
                ws.merge_cells(f"A2:{get_column_letter(ws.max_column)}2")
                ws["A2"] = banner_text
                ws["A2"].font = Font(bold=True, italic=True)
                ws["A2"].alignment = Alignment(horizontal="center")
                ws.freeze_panes = "A3"   # shift freeze below banner row
                ws.protection.sheet = True
                ws.protection.enable()
            elif sheet_name in {"Placements_ByPerson", "Persons_Unresolved", "Placements_Unresolved"}:
                ws.protection.sheet = True
                ws.protection.enable()

        # Post-process workbook in-memory (before ExcelWriter saves): README, reorder, hide
        wb = xw.book
        add_or_replace_readme_sheet(wb, readme_df=readme_df, title="README")
        reorder_sheets(wb)
        _apply_sheet_hiding(wb)
        _apply_coverage_colors(wb)


def main() -> int:
    import argparse
    parser = argparse.ArgumentParser(description="Stage 4: Build analytics sheets for footbag data.")
    parser.add_argument("--force-identity", action="store_true",
                        help="Overwrite Persons_Truth even if persons_truth.lock exists")
    args = parser.parse_args()

    repo = Path(__file__).resolve().parent
    out_dir = repo / "out"
    overrides_dir = repo / "overrides"

    # Gate 3: sentinel file check — lock is created manually after Gate 3 is verified
    lock_path = out_dir / "persons_truth.lock"
    persons_truth_csv = out_dir / "Persons_Truth.csv"
    skip_identity_overwrite = False
    if lock_path.exists() and not args.force_identity:
        if not persons_truth_csv.exists():
            print(f"ERROR: Persons_Truth is locked ({lock_path}) but {persons_truth_csv} is missing.")
            print("       Either restore Persons_Truth.csv or run with --force-identity to rebuild.")
            raise SystemExit(1)
        print(f"INFO: Persons_Truth is locked ({lock_path}).")
        print("      Will NOT overwrite Persons_Truth. Continuing with existing identity outputs.")
        skip_identity_overwrite = True

    # README template (user-provided) — accept either naming convention
    readme_df = read_csv_optional(repo / "readme-excel.csv")
    if readme_df.empty:
        readme_df = read_csv_optional(repo / "readme_excel.csv")

    # Overrides (displayable)
    person_aliases_overrides_df = read_csv_optional(overrides_dir / "person_aliases.csv")

    pf_csv = out_dir / "Placements_Flat.csv"
    xlsx = repo / "Footbag_Results_Canonical.xlsx"

    if not pf_csv.exists():
        print(f"ERROR: missing {pf_csv} (run 02p5 first)", file=sys.stderr)
        return 2
    if not xlsx.exists():
        print(f"ERROR: missing {xlsx} (run 03 first)", file=sys.stderr)
        return 2

    pf = pd.read_csv(pf_csv)
    pf_raw_count = len(pf)

    # --- Gate 1 enforcement: exclude rejected / unpresentable rows from analytics ---
    _rej_path = out_dir / "Placements_ByPerson_Rejected.csv"
    _exc_path = out_dir / "qc" / "excluded_results_rows_unpresentable.csv"
    _exclude_parts = []
    for _p in [_rej_path, _exc_path]:
        if _p.exists():
            _df = pd.read_csv(_p, dtype=str).fillna("")
            if not _df.empty:
                _exclude_parts.append(_df)
    if _exclude_parts:
        _join_cols = ["event_id", "division_raw", "place", "player1_name_raw"]
        _excl = pd.concat(_exclude_parts, ignore_index=True)
        for c in _join_cols:
            _excl[c] = _excl[c].astype(str).str.strip()
            pf[c] = pf[c].astype(str).str.strip()
        _excl_keys = _excl[_join_cols].drop_duplicates()
        _excl_keys["_gate1_exclude"] = True
        pf = pf.merge(_excl_keys, on=_join_cols, how="left")
        _n_excl = pf["_gate1_exclude"].notna().sum()
        print(f"[Gate1] Excluding {_n_excl} rows from analytics (rejected + unpresentable). Remaining: {(~pf['_gate1_exclude'].notna()).sum()}")
        pf = pf[pf["_gate1_exclude"].isna()].drop(columns=["_gate1_exclude"])

    # --- Coverage metric by event/division ---
    cov_df = build_coverage_by_event_division(
        pf=pf,
        out_dir=out_dir,
        quarantine_path=out_dir / "Placements_ByPerson_SinglesQuarantine.csv",
    )

    # --- Gap priority analysis ---
    gap_df = build_coverage_gap_priority(cov_df, out_dir)

    per_all = explode_to_people(pf)

    # --- Repair + QC: detect and fix inverted person_id / person_canon rows ---
    inv = (~per_all["person_id"].map(_is_uuid)) & (per_all["person_canon"].map(_is_uuid))
    print(f"[QC] Placements_ByPerson inversion rows: {inv.sum()} / {len(per_all)} ({inv.mean():.3%})")
    if inv.any():
        tmp = per_all.loc[inv, "person_id"].copy()
        per_all.loc[inv, "person_id"] = per_all.loc[inv, "person_canon"]
        per_all.loc[inv, "person_canon"] = tmp
    # --- Extra guard: if name_clean got UUID, replace with person_canon (name) ---
    if "player_name_clean" in per_all.columns:
        bad_name_clean = per_all["player_name_clean"].map(_is_uuid) & (~per_all["person_canon"].map(_is_uuid))
        if bad_name_clean.any():
            per_all.loc[bad_name_clean, "player_name_clean"] = per_all.loc[bad_name_clean, "person_canon"]

    # STEP 2: drop non-person-like rows (presentation / analytics only)
    def is_person_row(r):
        name_clean = (r.get("person_canon") or "").strip()
        name_raw = (r.get("player_name") or "").strip()
        comp_type = (r.get("competitor_type") or "").strip().lower()

        # must have something name-like
        if not is_person_like(name_clean):
            return False

        # reject obvious junk in RAW
        raw_l = name_raw.lower()
        if any(k in raw_l for k in [
            "result", "position", "partner", "tournament",
            "did not", "playoff", "annual"
        ]):
            return False

        # reject team concatenations
        if any(sym in name_raw for sym in ["+", "/", " ? ", " and "]):
            return False

        # reject locations / clubs (raw check)
        if raw_l in {"helsinki", "california", "arizona", "quebec"}:
            return False

        # reject numeric-heavy blobs
        if sum(c.isdigit() for c in name_raw) >= 3:
            return False

        return True

    per = per_all[per_all.apply(is_person_row, axis=1)].copy()

    _complete_flags = {"complete", "mostly_complete"}
    _cov_keys = (cov_df[cov_df["coverage_flag"].isin(_complete_flags)]
                 [["event_id", "division_canon"]].drop_duplicates())
    per_covered = per.merge(_cov_keys, on=["event_id", "division_canon"], how="inner")
    person_stats = build_person_stats(per_covered)
    player_stats = build_player_stats(per)
    division_stats = build_division_stats(pf, out_dir)
    person_by_cat = build_person_stats_by_div_category(per_covered)
    top_unmapped_people, top_unmapped_noise = build_top_unmapped_names(pf)

    aliases_csv = repo / "overrides" / "person_aliases.csv"
    aliases_df = load_person_aliases(aliases_csv)

    if not skip_identity_overwrite:
        # NO-GUESSING person dimension (one row per effective_person_id)
        persons_truth_full = build_persons_truth(per_all, aliases_df)
        qc_persons_truth(persons_truth_full)

        # Try to derive a presentable canon for any row whose current canon is not presentable.
        cleaned_all = persons_truth_full["person_canon"].map(clean_person_label_no_guess)
        persons_truth_full["person_canon_clean"] = cleaned_all.map(lambda t: t[0])
        persons_truth_full["person_canon_clean_reason"] = cleaned_all.map(lambda t: t[1])

        orig_ok = persons_truth_full["person_canon"].map(is_presentable_person_canon)
        clean_ok = persons_truth_full["person_canon_clean"].map(is_presentable_person_canon)

        # Only adopt cleaned canon when original is NOT presentable but cleaned IS presentable.
        use_clean = (~orig_ok) & clean_ok & persons_truth_full["person_canon_clean"].fillna("").ne("")
        persons_truth_full.loc[use_clean, "person_canon"] = persons_truth_full.loc[use_clean, "person_canon_clean"]

        # ---- Option A strict gate + quarantine ----
        mask_presentable = persons_truth_full["person_canon"].map(is_presentable_person_canon)

        not_presentable = persons_truth_full.loc[~mask_presentable].copy()
        not_presentable["exclude_reason"] = "not_presentable_strict"

        quarantine = detect_two_people_in_one_slot(persons_truth_full, pf=pf, aliases_df=aliases_df)
        quarantine_ids = set(quarantine["effective_person_id"].astype(str).str.strip()) if not quarantine.empty else set()

        excluded = not_presentable.copy()
        if quarantine_ids:
            q2 = persons_truth_full.loc[persons_truth_full["effective_person_id"].astype(str).str.strip().isin(quarantine_ids)].copy()
            q2["exclude_reason"] = "two_people_quarantine"
            if "quarantine_reason" in quarantine.columns and "quarantine_evidence" in quarantine.columns:
                q2 = q2.merge(
                    quarantine[["effective_person_id", "quarantine_reason", "quarantine_evidence"]],
                    on="effective_person_id",
                    how="left",
                )
            excluded = pd.concat([excluded, q2], ignore_index=True)

        persons_truth = persons_truth_full.loc[mask_presentable].copy()
        if quarantine_ids:
            persons_truth = persons_truth.loc[~persons_truth["effective_person_id"].astype(str).str.strip().isin(quarantine_ids)].copy()

        # Gate 3 Step 3.3: Exclude synthetic persons (all placements in sparse divisions + total == 1)
        _cov_flags = cov_df[["event_id", "division_canon", "coverage_flag"]].drop_duplicates()
        _per_cov = per_all[["person_id", "event_id", "division_canon"]].merge(
            _cov_flags, on=["event_id", "division_canon"], how="left"
        )
        _per_cov["coverage_flag"] = _per_cov["coverage_flag"].fillna("")
        _syn_agg = (
            _per_cov.groupby("person_id")
            .agg(
                total=("event_id", "count"),
                sparse=("coverage_flag", lambda s: (s == "sparse").sum()),
            )
            .reset_index()
        )
        _syn_agg["is_synthetic"] = (_syn_agg["sparse"] == _syn_agg["total"]) & (_syn_agg["total"] == 1)
        _synthetic_ids = set(_syn_agg.loc[_syn_agg["is_synthetic"], "person_id"].astype(str))

        # Exclude persons whose only placements are quarantined (logic present even when quarantine is empty)
        _qua_path = out_dir / "Placements_ByPerson_SinglesQuarantine.csv"
        if _qua_path.exists():
            _qua_df = pd.read_csv(_qua_path, dtype=str).fillna("")
            _qua_pids: set[str] = set()
            for _side in ["player1_person_id", "player2_person_id"]:
                if _side in _qua_df.columns:
                    _qua_pids |= set(_qua_df[_side].astype(str).str.strip())
            _qua_pids.discard("")
            if _qua_pids:
                _all_pids = set(per_all["person_id"].astype(str).str.strip())
                _qua_only = _qua_pids - _all_pids
                if _qua_only:
                    _qua_only_rows = persons_truth[
                        persons_truth["effective_person_id"].astype(str).str.strip().isin(_qua_only)
                    ].copy()
                    _qua_only_rows["exclude_reason"] = "synthetic_quarantine_only"
                    excluded = pd.concat([excluded, _qua_only_rows], ignore_index=True)
                    persons_truth = persons_truth[
                        ~persons_truth["effective_person_id"].astype(str).str.strip().isin(_qua_only)
                    ].copy()
                    print(f"[Gate3] Excluded {len(_qua_only_rows)} persons whose only placements are quarantined")

        if _synthetic_ids:
            _synthetic_rows = persons_truth[
                persons_truth["effective_person_id"].astype(str).str.strip().isin(_synthetic_ids)
            ].copy()
            if not _synthetic_rows.empty:
                _synthetic_rows["exclude_reason"] = "synthetic_sparse_single"
                excluded = pd.concat([excluded, _synthetic_rows], ignore_index=True)
                print(f"[Gate3] Excluded {len(_synthetic_rows)} synthetic persons (sparse+single-appearance)")
            persons_truth = persons_truth[
                ~persons_truth["effective_person_id"].astype(str).str.strip().isin(_synthetic_ids)
            ].copy()

        # --- coverage closure (only on strict, presentable set) ---
        # Collect all person canons referenced in Placements_Flat and ensure each appears
        # in Persons_Truth (prevents orphan canons from slipping through presentation).
        used_canons = set()
        for col in ["player1_person_canon", "player2_person_canon"]:
            if col in pf.columns:
                for v in pf[col].astype(str):
                    vv = str(v or "").strip()
                    if not vv:
                        continue
                    # normalize using the exact same no-guess cleaner used in QC07
                    cleaned, _reason = clean_person_label_no_guess(vv)
                    key = (cleaned or vv).strip()
                    if key and is_presentable_person_canon(key):
                        used_canons.add(key)
        existing = set(persons_truth["person_canon"].astype(str).str.strip())
        existing.discard("")
        missing = sorted(c for c in used_canons if c not in existing)
        if missing:
            add_rows = [_mk_truth_row_from_canon(c) for c in missing]
            persons_truth = pd.concat([persons_truth, pd.DataFrame(add_rows)], ignore_index=True)

        # Persons views (presentation surface): aliases_presentable from overrides (VERIFIED only)
        alias_map = build_aliases_presentable_from_overrides(person_aliases_overrides_df)
        if "effective_person_id" in persons_truth.columns:
            persons_truth["aliases_presentable"] = persons_truth["effective_person_id"].astype(str).str.strip().map(alias_map).fillna("")
        else:
            persons_truth["aliases_presentable"] = ""

        def _drop_self_alias(row: pd.Series) -> str:
            canon = (row.get("person_canon") or "").strip()
            aliases = (row.get("aliases_presentable") or "").strip()
            if not canon or not aliases:
                return aliases
            parts = [p.strip() for p in aliases.split(" | ") if p.strip()]
            parts = [p for p in parts if _canon_key(p) != _canon_key(canon)]
            return " | ".join(parts)

        persons_truth["aliases_presentable"] = persons_truth.apply(_drop_self_alias, axis=1)

        # Presentation rule: no duplicate display names across different IDs
        persons_truth, persons_truth_dupe_quarantine = quarantine_duplicate_display_names(
            persons_truth, name_col="person_canon", id_col="effective_person_id"
        )
        if len(persons_truth_dupe_quarantine) > 0:
            persons_truth_dupe_quarantine = persons_truth_dupe_quarantine.copy()
            persons_truth_dupe_quarantine["exclude_reason"] = "duplicate_person_canon"
            out_dir = repo / "out"
            out_dir.mkdir(exist_ok=True)
            persons_truth_dupe_quarantine.to_csv(out_dir / "Persons_DuplicateDisplay.csv", index=False)

        # Option A display sheet: slim, pivot-ready, one row per effective_person_id
        persons_truth_display_cols = ["person_canon", "aliases_presentable", "source", "notes", "effective_person_id"]
        persons_truth_display_cols = [c for c in persons_truth_display_cols if c in persons_truth.columns]
        persons_truth_display = persons_truth[persons_truth_display_cols].copy()

        persons_truth_full_out = persons_truth_full.copy()

        # ---- Persist definitive CSV artifacts (deterministic) ----
        out_dir = repo / "out"
        out_dir.mkdir(exist_ok=True)

        persons_truth.to_csv(PERSONS, index=False)

        persons_truth_full_out.to_csv(out_dir / "Persons_Truth_Full.csv", index=False)
        excluded.to_csv(out_dir / "Persons_Truth_Excluded.csv", index=False)
        if not quarantine.empty:
            quarantine.to_csv(out_dir / "Persons_Truth_Quarantine_TwoPeople.csv", index=False)

        # Persons_Public: canonical name + aliases only (no source/notes/IDs)
        persons_public = persons_truth_display[["person_canon", "aliases_presentable"]].copy()
    else:
        # Lock active: use existing Persons_Truth.csv, do not overwrite
        persons_truth = pd.read_csv(persons_truth_csv, dtype=str).fillna("")
        persons_truth_display_cols = ["person_canon", "aliases_presentable", "source", "notes", "effective_person_id"]
        persons_truth_display_cols = [c for c in persons_truth_display_cols if c in persons_truth.columns]
        persons_truth_display = persons_truth[persons_truth_display_cols].copy()
        persons_public = persons_truth_display[["person_canon", "aliases_presentable"]].copy()

    # Build clean competitor-centric Placements_ByPerson (new schema)
    placements_by_person_df = build_placements_by_person_clean(pf, cov_df, out_dir)
    persons_unresolved_df = build_persons_unresolved(pf, per_all, out_dir)
    # Workbook uses triaged version for Persons_Unresolved sheet when available
    df_unresolved = read_csv_optional(out_dir / "Persons_Unresolved_Triage.csv")
    if df_unresolved.empty:
        df_unresolved = read_csv_optional(out_dir / "Persons_Unresolved.csv")
    if not df_unresolved.empty:
        persons_unresolved_df = df_unresolved
    placements_unresolved_df = build_placements_unresolved(placements_by_person_df, out_dir)
    analytics_safe_df = build_analytics_safe_surface(placements_by_person_df, out_dir)
    data_integrity_df = build_data_integrity(pf_raw_count, pf, placements_by_person_df,
                                             analytics_safe_df, cov_df, out_dir)

    # Add career columns to Persons_Truth (coverage-filtered, so stats match Person_Stats)
    _career = (
        per_covered.groupby("person_id", dropna=False)
        .agg(
            total_placements_gate3=("event_id", "count"),
            first_year_active=("year", lambda s: int(pd.to_numeric(s, errors="coerce").min())
                               if pd.to_numeric(s, errors="coerce").notna().any() else ""),
            last_year_active=("year", lambda s: int(pd.to_numeric(s, errors="coerce").max())
                              if pd.to_numeric(s, errors="coerce").notna().any() else ""),
        )
        .reset_index()
        .rename(columns={"person_id": "effective_person_id"})
    )
    persons_truth_display = persons_truth_display.merge(_career, on="effective_person_id", how="left")
    persons_truth_display["total_placements_gate3"] = (
        persons_truth_display["total_placements_gate3"].fillna(0).astype(int)
    )
    persons_truth_display["years_active_count"] = persons_truth_display.apply(
        lambda r: (int(r["last_year_active"]) - int(r["first_year_active"]) + 1)
        if (pd.notna(r.get("first_year_active")) and r.get("first_year_active") != "" and
            pd.notna(r.get("last_year_active")) and r.get("last_year_active") != "")
        else 0,
        axis=1,
    )

    # ---- Workbook sheets (presentation only — diagnostics go to Review workbook) ----
    sheets = []
    sheets.append(("Placements_ByPerson", placements_by_person_df))
    sheets.append(("Persons_Truth", persons_truth_display))
    sheets.append(("Analytics_Safe_Surface", analytics_safe_df))
    # Put likelihood fields near the front for Persons_Unresolved if present
    if not persons_unresolved_df.empty:
        preferred = [
            "person_canon", "issue_type", "appearances",
            "resolution_likelihood", "likelihood_score", "triage_reasons",
            "evidence", "suggested_action", "player_id", "person_id", "name_raw", "name_clean"
        ]
        cols = [c for c in preferred if c in persons_unresolved_df.columns] + [
            c for c in persons_unresolved_df.columns if c not in preferred
        ]
        persons_unresolved_df = persons_unresolved_df[cols]
    sheets.append(("Persons_Unresolved", persons_unresolved_df))
    sheets.append(("Placements_Unresolved", placements_unresolved_df))
    if len(person_stats) > 0:
        sheets.append(("Person_Stats", person_stats))
    if len(person_by_cat) > 0:
        sheets.append(("PersonStats_ByDivCat", person_by_cat))
    sheets.append(("Division_Stats", division_stats))
    if len(cov_df) > 0:
        sheets.append(("Coverage_ByEventDiv", cov_df))
    sheets.append(("Data_Integrity", data_integrity_df))

    write_sheets_append(xlsx, sheets, readme_df=readme_df)

    # ---- Remove diagnostic/obsolete sheets from Stage 03 ----
    wb = openpyxl.load_workbook(xlsx)
    sheets_to_remove = [
        "Players", "Players_Junk", "Players_Alias_Candidates",
        "Persons_Truth_Source",
        "Teams", "Teams_Alias_Candidates", "QC_TopIssues",
    ]
    for name in sheets_to_remove:
        if name in wb.sheetnames:
            del wb[name]

    # ---- Add hyperlinks from Placements_ByPerson event_id → year sheets ----
    locator_path = repo / "out" / "event_locator.json"
    if locator_path.exists() and "Placements_ByPerson" in wb.sheetnames:
        with open(locator_path, encoding="utf-8") as f:
            event_locator = json.load(f)
        ws = wb["Placements_ByPerson"]
        # Find event_id column index
        eid_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == "event_id":
                eid_col = col_idx
                break
        if eid_col:
            hyperlink_font = Font(color="0563C1", underline="single")
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=eid_col)
                eid = str(cell.value or "").strip()
                if eid in event_locator:
                    sheet_name, col_idx = event_locator[eid]
                    col_letter = get_column_letter(col_idx)
                    cell.hyperlink = f"#{sheet_name}!{col_letter}1"
                    cell.font = hyperlink_font

    # ---- Add coverage_ratio + coverage_flag rows to year sheets ----
    if len(cov_df) > 0:
        # Build per-event aggregate: min coverage_ratio across divisions
        cov_by_event = (
            cov_df.groupby("event_id", dropna=False)
            .agg(coverage_ratio=("coverage_ratio", "min"))
            .reset_index()
        )
        cov_by_event["event_id"] = cov_by_event["event_id"].astype(str).str.strip()

        cov_by_event["coverage_flag"] = cov_by_event["coverage_ratio"].map(_coverage_flag)
        cov_lookup = dict(zip(cov_by_event["event_id"], zip(cov_by_event["coverage_ratio"], cov_by_event["coverage_flag"])))

        for sheet_name in wb.sheetnames:
            if not is_year_sheet(sheet_name):
                continue
            ws = wb[sheet_name]
            if ws.max_column < 2:
                continue

            # Year sheets: row 1 = header (event_ids as column headers), rows 2-7 = data
            # Add two new rows after the last data row
            next_row = ws.max_row + 1
            ratio_row = next_row
            flag_row = next_row + 1

            ws.cell(row=ratio_row, column=1, value="Coverage Ratio")
            ws.cell(row=flag_row, column=1, value="Coverage Flag")

            for col_idx in range(2, ws.max_column + 1):
                eid = str(ws.cell(row=1, column=col_idx).value or "").strip()
                if eid in cov_lookup:
                    ratio, flag = cov_lookup[eid]
                    ws.cell(row=ratio_row, column=col_idx, value=round(ratio, 3))
                    ws.cell(row=flag_row, column=col_idx, value=flag)

    wb.save(xlsx)
    wb.close()

    # Gate 3 completion check: COUNT(person_id) == COUNT(DISTINCT person_canon)
    n_ids = persons_truth["effective_person_id"].nunique()
    n_canons = persons_truth["person_canon"].nunique()
    if n_ids == n_canons:
        print(f"[Gate3] PASS: COUNT(person_id) == COUNT(person_canon) = {n_ids}")
    else:
        print(f"[Gate3] FAIL: person_id count ({n_ids}) != person_canon count ({n_canons})")
        print(f"        Run: python qc02_canon_multiple_person_ids.py")

    print(f"OK: updated {xlsx} with: {', '.join([n for n, _ in sheets])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
