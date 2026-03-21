#!/usr/bin/env python3
"""
update_workbook_event_ids.py

Replace legacy event IDs in the community workbook with canonical event_key slugs.

No command-line arguments required.
Edit the configuration block below if you want different paths or behavior.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


# ============================================================================
# CONFIG
# ============================================================================

ROOT = Path.home() / "projects" / "FOOTBAG_DATA"

EVENTS_CSV = ROOT / "out" / "canonical" / "events.csv"
INPUT_XLSX = ROOT / "Footbag_Results_Community_FINAL_v13.xlsx"
OUTPUT_XLSX = ROOT / "Footbag_Results_Community_FINAL_v13_eventkeys.xlsx"

# If True, overwrite INPUT_XLSX instead of writing OUTPUT_XLSX
IN_PLACE = True

# If True, print every replacement
VERBOSE = False


# ============================================================================
# HELPERS
# ============================================================================

def build_event_map(events_csv: Path) -> dict[str, str]:
    df = pd.read_csv(events_csv, dtype=str).fillna("")

    id_col = None
    for candidate in ("legacy_event_id", "event_id"):
        if candidate in df.columns:
            id_col = candidate
            break

    if id_col is None:
        raise ValueError(
            f"Could not find legacy event ID column in {events_csv}. "
            "Expected 'legacy_event_id' or 'event_id'."
        )

    if "event_key" not in df.columns:
        raise ValueError(f"Could not find 'event_key' column in {events_csv}.")

    event_map: dict[str, str] = {}
    dupes: dict[str, set[str]] = defaultdict(set)

    for _, row in df.iterrows():
        legacy_id = str(row[id_col]).strip()
        event_key = str(row["event_key"]).strip()

        if not legacy_id or not event_key:
            continue

        if legacy_id in event_map and event_map[legacy_id] != event_key:
            dupes[legacy_id].update({event_map[legacy_id], event_key})

        event_map[legacy_id] = event_key

    if dupes:
        msg = "\n".join(
            f"  {eid}: {sorted(list(keys))}" for eid, keys in sorted(dupes.items())
        )
        raise ValueError(
            "Found conflicting mappings for some legacy event IDs:\n" + msg
        )

    return event_map


def normalize_cell_value(value: Any) -> str | None:
    """
    Normalize a cell value to a candidate legacy event ID string.

    Handles:
    - strings like "1035277529"
    - ints like 1035277529
    - floats like 1035277529.0

    Rejects:
    - formulas
    - non-integer floats
    - strings with extra text
    """
    if value is None:
        return None

    if isinstance(value, str):
        s = value.strip()
        if not s or s.startswith("="):
            return None
        if s.isdigit():
            return s
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return None

    return None


def replace_ids_in_workbook(
    in_xlsx: Path,
    out_xlsx: Path,
    event_map: dict[str, str],
    verbose: bool = False,
) -> tuple[int, dict[str, int]]:
    wb = load_workbook(in_xlsx)
    replacements_by_sheet: dict[str, int] = defaultdict(int)
    total = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                legacy_id = normalize_cell_value(cell.value)
                if legacy_id is None:
                    continue
                if legacy_id in event_map:
                    old = cell.value
                    new = event_map[legacy_id]
                    cell.value = new
                    replacements_by_sheet[ws.title] += 1
                    total += 1
                    if verbose:
                        print(f"{ws.title}!{cell.coordinate}: {old!r} -> {new!r}")

    wb.save(out_xlsx)
    return total, dict(replacements_by_sheet)


# ============================================================================
# MAIN
# ============================================================================

def main() -> None:
    if not EVENTS_CSV.exists():
        raise FileNotFoundError(f"events.csv not found: {EVENTS_CSV}")
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"Workbook not found: {INPUT_XLSX}")

    out_xlsx = INPUT_XLSX if IN_PLACE else OUTPUT_XLSX

    print(f"Loading event map from: {EVENTS_CSV}")
    event_map = build_event_map(EVENTS_CSV)
    print(f"Loaded {len(event_map):,} legacy_event_id -> event_key mappings")

    print(f"Updating workbook: {INPUT_XLSX}")
    total, by_sheet = replace_ids_in_workbook(
        in_xlsx=INPUT_XLSX,
        out_xlsx=out_xlsx,
        event_map=event_map,
        verbose=VERBOSE,
    )

    print(f"\nWrote: {out_xlsx}")
    print(f"Total replacements: {total:,}")

    if by_sheet:
        print("Replacements by sheet:")
        for sheet, count in sorted(by_sheet.items()):
            print(f"  {sheet}: {count:,}")
    else:
        print("No matching event IDs found in workbook.")


if __name__ == "__main__":
    main()
