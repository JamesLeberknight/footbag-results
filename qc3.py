#!/usr/bin/env python3
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

DEFAULT_XLSX = Path("Footbag_Results_Community_FINAL_v13.xlsx")
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX

wb = load_workbook(XLSX, data_only=False)

legacy_id_hits = []
event_prefix_hits = []
plus_separator_hits = []
date_year_only_hits = []
typo_hits = []

legacy_id_re = re.compile(r"^\d{6,}$")
event_prefix_re = re.compile(r"^event_\d{4}_")
year_only_re = re.compile(r"^\d{4}$")

typo_patterns = [
    re.compile(r"\bIntrmediate\b", re.I),
    re.compile(r"\bnd=\b", re.I),
    re.compile(r"\bdavid Butcher\b"),
]

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()

            if legacy_id_re.fullmatch(s):
                legacy_id_hits.append((ws.title, cell.coordinate, s))

            if event_prefix_re.match(s):
                event_prefix_hits.append((ws.title, cell.coordinate, s))

            if " + " in s:
                plus_separator_hits.append((ws.title, cell.coordinate, s))

            if year_only_re.fullmatch(s):
                if cell.row <= 10:
                    date_year_only_hits.append((ws.title, cell.coordinate, s))

            for pat in typo_patterns:
                if pat.search(s):
                    typo_hits.append((ws.title, cell.coordinate, s))
                    break

print("\nLEGACY NUMERIC IDS")
for x in legacy_id_hits[:200]:
    print(x)

print("\nEVENT_ PREFIX IDS")
for x in event_prefix_hits[:200]:
    print(x)

print("\nPLUS SEPARATOR HITS")
for x in plus_separator_hits[:200]:
    print(x)

print("\nYEAR-ONLY METADATA CELLS")
for x in date_year_only_hits[:200]:
    print(x)

print("\nTYPO / ODD STRING HITS")
for x in typo_hits[:200]:
    print(x)
