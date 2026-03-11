"""
tools/53_all_events_viewer.py
─────────────────────────────
Comprehensive all-events side-by-side viewer.

Produces two artefacts in out/review/:
  all_events_overview.xlsx  — sortable, filterable spreadsheet of all 774 events
  all_events_inspector.html — searchable HTML, one accordion section per event

Both include:
  • Event metadata (year, name, country, event_type, host_club)
  • Stage2 division + placement counts
  • PBP division + placement counts (identity-locked)
  • Coverage flags summary (complete / partial / sparse)
  • Known-issue flag + severity
  • Quarantine flag + reason
  • Review heat label + disposition (from MEDIUM decisions if available)
  • All divisions listed with place range and top-3 names (HTML only)

Usage:
    .venv/bin/python tools/53_all_events_viewer.py
"""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

csv.field_size_limit(sys.maxsize)

ROOT   = Path(__file__).resolve().parent.parent
OUT    = ROOT / "out"
REVIEW = OUT / "review"
REVIEW.mkdir(parents=True, exist_ok=True)

# ── inputs ───────────────────────────────────────────────────────────────────
STAGE2_CSV    = OUT / "stage2_canonical_events.csv"
PBP_CSV       = OUT / "Placements_ByPerson.csv"           # filtered (no unresolved)
PBP_FULL_CSV  = ROOT / "inputs/identity_lock/Placements_ByPerson_v60.csv"
COVERAGE_CSV  = OUT / "Coverage_ByEventDivision.csv"
KNOWN_CSV     = ROOT / "overrides/known_issues.csv"
QUARANTINE_CSV= ROOT / "inputs/review_quarantine_events.csv"
DECISIONS_CSV = REVIEW / "medium_review_decisions.csv"
REVIEW_PKT    = REVIEW / "Footbag_Event_Review_Packet.xlsx"

# ── load helpers ─────────────────────────────────────────────────────────────

def _load_csv(path: Path) -> list[dict]:
    if not path.exists():
        print(f"  WARN: {path.name} not found — skipping")
        return []
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _load_review_heat() -> dict[str, dict]:
    """Load heat label + review_status from the review packet xlsx."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(REVIEW_PKT), read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            r = dict(zip(headers, row))
            eid = str(r.get("event_id", ""))
            result[eid] = {
                "heat_label":    r.get("review_heat_label", ""),
                "heat_score":    r.get("review_heat_score", 0),
                "diff_summary":  r.get("diff_summary", ""),
                "review_status": r.get("review_status", ""),
            }
        wb.close()
        return result
    except Exception as e:
        print(f"  WARN: could not load review packet: {e}")
        return {}


# ── main data assembly ────────────────────────────────────────────────────────

def build_event_table() -> list[dict]:
    print("Loading stage2 events…")
    stage2_rows = _load_csv(STAGE2_CSV)

    print("Loading PBP (full v60)…")
    pbp_full = _load_csv(PBP_FULL_CSV)

    print("Loading coverage flags…")
    coverage_rows = _load_csv(COVERAGE_CSV)

    print("Loading known issues…")
    known_issues = {r["event_id"]: r for r in _load_csv(KNOWN_CSV)}

    print("Loading quarantine…")
    quarantine    = {r["event_id"]: r for r in _load_csv(QUARANTINE_CSV)}

    print("Loading review decisions…")
    decisions     = {r["event_id"]: r for r in _load_csv(DECISIONS_CSV)}

    print("Loading review heat labels…")
    heat_map      = _load_review_heat()

    # ── index PBP by event_id ─────────────────────────────────────────────
    pbp_by_event: dict[str, list[dict]] = defaultdict(list)
    for row in pbp_full:
        pbp_by_event[row["event_id"]].append(row)

    # ── index coverage by event_id ────────────────────────────────────────
    cov_by_event: dict[str, list[dict]] = defaultdict(list)
    for row in coverage_rows:
        cov_by_event[str(row.get("event_id", ""))].append(row)

    # ── build table ───────────────────────────────────────────────────────
    table = []
    for ev in stage2_rows:
        eid  = str(ev["event_id"])
        year = ev.get("year", "")

        # stage2 placements
        try:
            placements = json.loads(ev.get("placements_json", "[]") or "[]")
        except Exception:
            placements = []

        s2_divs  = sorted({p.get("division_canon", "") for p in placements if p.get("division_canon")})
        s2_place_count = len(placements)
        s2_div_count   = len(s2_divs)

        # PBP
        pbp_rows  = pbp_by_event.get(eid, [])
        pbp_divs  = sorted({r.get("division_canon", "") for r in pbp_rows if r.get("division_canon")})
        pbp_place_count  = len(pbp_rows)
        pbp_div_count    = len(pbp_divs)
        pbp_unresolved   = sum(1 for r in pbp_rows if r.get("person_unresolved") == "1")

        # coverage flags
        cov_rows  = cov_by_event.get(eid, [])
        cov_flags = sorted({r.get("coverage_flag", "") for r in cov_rows if r.get("coverage_flag")})
        cov_summary = "/".join(cov_flags) if cov_flags else ""
        has_partial = any(f in ("partial", "sparse") for f in cov_flags)

        # known issues
        ki = known_issues.get(eid)
        ki_flag     = 1 if ki else 0
        ki_severity = ki["severity"] if ki else ""
        ki_note     = ki["note"]     if ki else ""

        # quarantine
        qt = quarantine.get(eid)
        qt_flag   = 1 if qt else 0
        qt_reason = qt["reason"] if qt else ""

        # review heat
        hm = heat_map.get(eid, {})
        heat_label   = "QUARANTINED" if qt_flag else hm.get("heat_label", "")
        heat_score   = 0 if qt_flag else (hm.get("heat_score") or 0)
        diff_summary = hm.get("diff_summary", "")

        # MEDIUM decision if available
        dec = decisions.get(eid, {})
        disposition  = dec.get("recommended_disposition", "")
        dec_note     = dec.get("short_note", "")
        pattern_fam  = dec.get("pattern_family", "")

        # division details for HTML
        div_details = []
        by_div: dict[str, list] = defaultdict(list)
        for p in placements:
            dc = p.get("division_canon", "Unknown")
            by_div[dc].append(p)

        for dc in s2_divs:
            rows_d = sorted(by_div.get(dc, []), key=lambda x: x.get("place", 999))
            places  = [r.get("place", 0) for r in rows_d]
            min_p   = min(places) if places else 0
            max_p   = max(places) if places else 0
            cat     = rows_d[0].get("division_category", "") if rows_d else ""
            top3    = []
            for r in rows_d[:3]:
                name = r.get("player1_name", "")
                if r.get("player2_name"):
                    name += " / " + r["player2_name"]
                top3.append(f"#{r.get('place','')} {name}")
            div_details.append({
                "division": dc,
                "category": cat,
                "count": len(rows_d),
                "min_place": min_p,
                "max_place": max_p,
                "top3": top3,
            })

        table.append({
            "event_id":         eid,
            "year":             year,
            "event_name":       ev.get("event_name", ""),
            "country":          ev.get("location", ""),
            "event_type":       ev.get("event_type", ""),
            "host_club":        ev.get("host_club", ""),
            "date":             ev.get("date", ""),
            # stage2
            "s2_div_count":     s2_div_count,
            "s2_place_count":   s2_place_count,
            "s2_divs":          ", ".join(s2_divs),
            # pbp
            "pbp_div_count":    pbp_div_count,
            "pbp_place_count":  pbp_place_count,
            "pbp_unresolved":   pbp_unresolved,
            "pbp_divs":         ", ".join(pbp_divs),
            # coverage
            "coverage_flags":   cov_summary,
            "has_partial":      has_partial,
            # known issues
            "ki_flag":          ki_flag,
            "ki_severity":      ki_severity,
            "ki_note":          ki_note,
            # quarantine
            "qt_flag":          qt_flag,
            "qt_reason":        qt_reason,
            # review
            "heat_label":       heat_label,
            "heat_score":       heat_score,
            "diff_summary":     diff_summary,
            "disposition":      disposition,
            "pattern_family":   pattern_fam,
            "dec_note":         dec_note,
            # for HTML only
            "_div_details":     div_details,
        })

    table.sort(key=lambda r: (int(r["year"]) if r["year"].isdigit() else 0, r["event_name"]))
    print(f"  {len(table)} events assembled")
    return table


# ── XLSX output ───────────────────────────────────────────────────────────────

XLSX_COLS = [
    ("event_id",        "Event ID",           12),
    ("year",            "Year",                6),
    ("event_name",      "Event Name",         45),
    ("country",         "Location",           22),
    ("event_type",      "Type",               10),
    ("host_club",       "Host Club",          18),
    ("date",            "Date",               12),
    ("s2_div_count",    "S2 Divs",             8),
    ("s2_place_count",  "S2 Places",           9),
    ("pbp_div_count",   "PBP Divs",            9),
    ("pbp_place_count", "PBP Places",         10),
    ("pbp_unresolved",  "Unresolved",          9),
    ("coverage_flags",  "Coverage",           16),
    ("ki_flag",         "Known Issue",         9),
    ("ki_severity",     "KI Severity",        10),
    ("ki_note",         "KI Note",            40),
    ("qt_flag",         "Quarantine",          9),
    ("qt_reason",       "QT Reason",          20),
    ("heat_label",      "Heat",               12),
    ("heat_score",      "Score",               6),
    ("diff_summary",    "Diff",               20),
    ("pattern_family",  "Pattern Family",     18),
    ("disposition",     "Disposition",        14),
    ("dec_note",        "Decision Note",      45),
    ("s2_divs",         "Stage2 Divisions",   60),
]


def write_xlsx(table: list[dict], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  SKIP xlsx: openpyxl not available")
        return

    print(f"Writing {out_path.name}…")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Events"

    # header
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, (key, label, width) in enumerate(XLSX_COLS, 1):
        cell = ws.cell(1, ci, label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = "A2"

    # row fills
    FILL = {
        "HIGH":        PatternFill("solid", fgColor="FFC7CE"),
        "QUARANTINED": PatternFill("solid", fgColor="CCCCCC"),
        "MEDIUM":      PatternFill("solid", fgColor="FFEB9C"),
        "LOW":         PatternFill("solid", fgColor="EBF1DE"),
        "CLEAN":       PatternFill("solid", fgColor="FFFFFF"),
        "":            PatternFill("solid", fgColor="FFFFFF"),
    }
    DISP_FILL = {
        "ACCEPT":         PatternFill("solid", fgColor="C6EFCE"),
        "SOURCE_PARTIAL": PatternFill("solid", fgColor="FFEB9C"),
        "DEFER":          PatternFill("solid", fgColor="FFC7CE"),
    }
    small_font = Font(size=9)
    heat_col_idx = next(i for i, (k, *_) in enumerate(XLSX_COLS, 1) if k == "heat_label")
    disp_col_idx = next(i for i, (k, *_) in enumerate(XLSX_COLS, 1) if k == "disposition")

    for ri, row in enumerate(table, 2):
        hl = row.get("heat_label", "")
        row_fill = FILL.get(hl, FILL[""])
        for ci, (key, label, width) in enumerate(XLSX_COLS, 1):
            val = row.get(key, "")
            cell = ws.cell(ri, ci, val)
            cell.fill = row_fill
            cell.font = small_font
            cell.alignment = Alignment(wrap_text=False, vertical="top")
        # heat cell colour override
        ws.cell(ri, heat_col_idx).fill = FILL.get(hl, FILL[""])
        # disposition cell colour override
        disp = row.get("disposition", "")
        if disp in DISP_FILL:
            ws.cell(ri, disp_col_idx).fill = DISP_FILL[disp]

    ws.auto_filter.ref = f"A1:{get_column_letter(len(XLSX_COLS))}1"
    ws.row_dimensions[1].height = 36

    wb.save(str(out_path))
    print(f"  Saved: {out_path} ({len(table)} rows)")


# ── HTML output ───────────────────────────────────────────────────────────────

HEAT_COLOUR = {
    "HIGH":        "#FFC7CE",
    "QUARANTINED": "#CCCCCC",
    "MEDIUM":      "#FFEB9C",
    "LOW":         "#EBF1DE",
    "CLEAN":       "#F7F7F7",
    "":            "#FFFFFF",
}
DISP_COLOUR = {
    "ACCEPT":         "#C6EFCE",
    "SOURCE_PARTIAL": "#FFEB9C",
    "DEFER":          "#FFC7CE",
}
KI_COLOUR = {
    "severe":   "#FFC7CE",
    "moderate": "#FFEB9C",
    "minor":    "#EBF1DE",
    "":         "transparent",
}


def _esc(s) -> str:
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def write_html(table: list[dict], out_path: Path) -> None:
    print(f"Writing {out_path.name}…")

    # build year groups
    years: dict[str, list[dict]] = defaultdict(list)
    for row in table:
        years[str(row["year"])].append(row)
    sorted_years = sorted(years.keys(), key=lambda y: int(y) if y.isdigit() else 0, reverse=True)

    # stats for header
    total = len(table)
    high   = sum(1 for r in table if r["heat_label"] == "HIGH")
    med    = sum(1 for r in table if r["heat_label"] == "MEDIUM")
    low    = sum(1 for r in table if r["heat_label"] == "LOW")
    clean  = sum(1 for r in table if r["heat_label"] == "CLEAN")
    quar   = sum(1 for r in table if r["heat_label"] == "QUARANTINED")
    ki_cnt = sum(1 for r in table if r["ki_flag"])

    html_parts: list[str] = []
    html_parts.append(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Footbag Archive — All Events Viewer</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;font-size:13px;background:#f0f2f5;color:#222}}
#header{{background:#1F3864;color:#fff;padding:16px 24px}}
#header h1{{font-size:20px;margin-bottom:4px}}
#header .stats{{display:flex;gap:16px;flex-wrap:wrap;margin-top:8px}}
.stat{{background:rgba(255,255,255,.15);border-radius:6px;padding:4px 10px;font-size:12px}}
.stat b{{font-size:14px}}
#controls{{background:#fff;padding:10px 24px;border-bottom:1px solid #ddd;display:flex;gap:12px;flex-wrap:wrap;align-items:center;position:sticky;top:0;z-index:100;box-shadow:0 2px 6px rgba(0,0,0,.1)}}
#controls input,#controls select{{padding:6px 10px;border:1px solid #ccc;border-radius:4px;font-size:12px}}
#controls input{{width:280px}}
#controls label{{font-size:12px;color:#555}}
#year-nav{{background:#fff;padding:8px 24px;border-bottom:1px solid #ddd;display:flex;flex-wrap:wrap;gap:6px}}
#year-nav a{{font-size:11px;color:#1F3864;text-decoration:none;padding:2px 8px;border:1px solid #1F3864;border-radius:3px}}
#year-nav a:hover{{background:#1F3864;color:#fff}}
#main{{padding:12px 24px;max-width:1600px}}
.year-section{{margin-bottom:28px}}
.year-header{{font-size:16px;font-weight:bold;color:#1F3864;margin-bottom:8px;padding-bottom:4px;border-bottom:2px solid #1F3864}}
.event-card{{background:#fff;border:1px solid #ddd;border-radius:6px;margin-bottom:6px;overflow:hidden}}
.event-header{{display:flex;align-items:stretch;cursor:pointer;user-select:none}}
.event-header:hover .ev-name{{text-decoration:underline}}
.heat-bar{{width:6px;flex-shrink:0}}
.ev-main{{padding:8px 12px;flex:1;min-width:0}}
.ev-name{{font-weight:600;font-size:13px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.ev-meta{{font-size:11px;color:#666;margin-top:2px}}
.ev-badges{{display:flex;gap:5px;flex-wrap:wrap;margin-top:4px}}
.badge{{font-size:10px;padding:2px 7px;border-radius:10px;font-weight:600;white-space:nowrap}}
.ev-counts{{padding:8px 12px;display:flex;flex-direction:column;justify-content:center;gap:2px;min-width:160px;border-left:1px solid #eee;font-size:11px;color:#444}}
.ev-counts .cnt-row{{display:flex;justify-content:space-between;gap:8px}}
.ev-counts .cnt-label{{color:#888}}
.toggle-arrow{{padding:8px 12px;display:flex;align-items:center;font-size:14px;color:#888}}
.event-body{{display:none;border-top:1px solid #eee;padding:12px;background:#fafafa}}
.event-body.open{{display:block}}
.div-table{{width:100%;border-collapse:collapse;font-size:11px;margin-top:6px}}
.div-table th{{background:#1F3864;color:#fff;padding:4px 8px;text-align:left;font-weight:600}}
.div-table td{{padding:4px 8px;border-bottom:1px solid #eee;vertical-align:top}}
.div-table tr:hover td{{background:#f0f4ff}}
.cat-net{{color:#1565C0;font-weight:600}}
.cat-freestyle{{color:#6A1B9A;font-weight:600}}
.cat-golf{{color:#2E7D32;font-weight:600}}
.cat-other{{color:#555}}
.section-label{{font-size:11px;font-weight:bold;color:#555;margin:8px 0 4px}}
.meta-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:6px;font-size:11px}}
.meta-item{{background:#f0f0f0;border-radius:4px;padding:4px 8px}}
.meta-item b{{color:#444}}
.dec-box{{margin-top:8px;padding:6px 10px;border-radius:4px;font-size:11px;border:1px solid #ddd}}
.hidden{{display:none!important}}
</style>
</head>
<body>
<div id="header">
  <h1>Footbag Archive — All Events Viewer</h1>
  <div class="stats">
    <div class="stat"><b>{total}</b> events</div>
    <div class="stat" style="background:#FFC7CE;color:#900">HIGH <b>{high}</b></div>
    <div class="stat" style="background:#FFEB9C;color:#7A6000">MEDIUM <b>{med}</b></div>
    <div class="stat" style="background:#EBF1DE;color:#3A5200">LOW <b>{low}</b></div>
    <div class="stat" style="background:#E8E8E8;color:#444">CLEAN <b>{clean}</b></div>
    <div class="stat" style="background:#CCCCCC;color:#444">QUAR <b>{quar}</b></div>
    <div class="stat" style="background:#E3F2FD;color:#0D47A1">Known Issues <b>{ki_cnt}</b></div>
  </div>
</div>
<div id="controls">
  <input type="text" id="search" placeholder="Search event name, location, club…" oninput="filterEvents()">
  <select id="filter-heat" onchange="filterEvents()">
    <option value="">All heat levels</option>
    <option>HIGH</option><option>MEDIUM</option><option>LOW</option>
    <option>CLEAN</option><option>QUARANTINED</option>
  </select>
  <select id="filter-year" onchange="filterEvents()">
    <option value="">All years</option>
""")
    for yr in sorted_years:
        html_parts.append(f'    <option>{_esc(yr)}</option>\n')
    html_parts.append("""  </select>
  <select id="filter-ki" onchange="filterEvents()">
    <option value="">Any known-issue</option>
    <option value="1">Has known issue</option>
    <option value="0">No known issue</option>
  </select>
  <select id="filter-type" onchange="filterEvents()">
    <option value="">All event types</option>
    <option>open</option><option>regional</option><option>national</option>
    <option>worlds</option><option>continental</option><option>invitational</option>
  </select>
  <button onclick="expandAll()" style="padding:5px 12px;border:1px solid #ccc;border-radius:4px;cursor:pointer;font-size:12px">Expand all</button>
  <button onclick="collapseAll()" style="padding:5px 12px;border:1px solid #ccc;border-radius:4px;cursor:pointer;font-size:12px">Collapse all</button>
  <span id="visible-count" style="font-size:11px;color:#888"></span>
</div>
<div id="year-nav">
""")
    for yr in sorted_years:
        html_parts.append(f'  <a href="#year-{_esc(yr)}">{_esc(yr)}</a>\n')
    html_parts.append("</div>\n<div id=\"main\">\n")

    for yr in sorted_years:
        yr_events = years[yr]
        html_parts.append(f'<div class="year-section" id="year-{_esc(yr)}">\n')
        html_parts.append(f'  <div class="year-header">{_esc(yr)} — {len(yr_events)} event{"s" if len(yr_events)!=1 else ""}</div>\n')

        for row in yr_events:
            eid  = row["event_id"]
            hl   = row["heat_label"]
            heat_bg = HEAT_COLOUR.get(hl, "#fff")
            disp = row["disposition"]
            ki_sev = row["ki_severity"]

            # badges
            badges = []
            if row["qt_flag"]:
                badges.append(('<span class="badge" style="background:#CCCCCC;color:#444">'
                               f'QUARANTINED: {_esc(row["qt_reason"])}</span>'))
            if row["ki_flag"]:
                ki_bg = KI_COLOUR.get(ki_sev, "transparent")
                badges.append(f'<span class="badge" style="background:{ki_bg};color:#333">'
                               f'KNOWN ISSUE ({_esc(ki_sev)})</span>')
            if disp:
                d_bg = DISP_COLOUR.get(disp, "#eee")
                badges.append(f'<span class="badge" style="background:{d_bg};color:#333">{_esc(disp)}</span>')
            if row["pattern_family"]:
                badges.append(f'<span class="badge" style="background:#E3F2FD;color:#0D47A1">{_esc(row["pattern_family"])}</span>')
            if row["has_partial"]:
                badges.append('<span class="badge" style="background:#FFF3E0;color:#E65100">partial coverage</span>')
            if row["pbp_unresolved"]:
                badges.append(f'<span class="badge" style="background:#F3E5F5;color:#4A148C">'
                               f'{_esc(row["pbp_unresolved"])} unresolved</span>')

            badges_html = "".join(badges)

            # data attributes for filtering
            da = (f'data-heat="{_esc(hl)}" '
                  f'data-year="{_esc(yr)}" '
                  f'data-ki="{row["ki_flag"]}" '
                  f'data-type="{_esc(row["event_type"])}" '
                  f'data-search="{_esc((row["event_name"]+" "+row["country"]+" "+row["host_club"]).lower())}"')

            html_parts.append(f'<div class="event-card" {da}>\n')
            html_parts.append(f'  <div class="event-header" onclick="toggleBody(this)">\n')
            html_parts.append(f'    <div class="heat-bar" style="background:{heat_bg}"></div>\n')
            html_parts.append(f'    <div class="ev-main">\n')
            html_parts.append(f'      <div class="ev-name">{_esc(row["event_name"])}</div>\n')
            meta_parts = []
            if row["country"]:  meta_parts.append(_esc(row["country"]))
            if row["event_type"]: meta_parts.append(_esc(row["event_type"]))
            if row["host_club"]:  meta_parts.append(_esc(row["host_club"]))
            if row["date"]:       meta_parts.append(_esc(row["date"]))
            html_parts.append(f'      <div class="ev-meta">{" · ".join(meta_parts)}</div>\n')
            if badges_html:
                html_parts.append(f'      <div class="ev-badges">{badges_html}</div>\n')
            html_parts.append(f'    </div>\n')

            # counts column
            s2_p = row["s2_place_count"]
            pb_p = row["pbp_place_count"]
            s2_d = row["s2_div_count"]
            pb_d = row["pbp_div_count"]
            cov  = row["coverage_flags"] or "—"
            diff_val = abs(s2_p - pb_p)
            diff_str = f'+{diff_val}' if s2_p > pb_p else (f'-{diff_val}' if s2_p < pb_p else '=')
            html_parts.append(f'''    <div class="ev-counts">
      <div class="cnt-row"><span class="cnt-label">S2 places</span><b>{s2_p}</b></div>
      <div class="cnt-row"><span class="cnt-label">PBP places</span><b>{pb_p}</b></div>
      <div class="cnt-row"><span class="cnt-label">Divs (s2/pbp)</span><b>{s2_d}/{pb_d}</b></div>
      <div class="cnt-row"><span class="cnt-label">Diff</span><b>{diff_str}</b></div>
      <div class="cnt-row"><span class="cnt-label">Coverage</span><b>{_esc(cov)}</b></div>
    </div>\n''')
            html_parts.append(f'    <div class="toggle-arrow">▶</div>\n')
            html_parts.append(f'  </div>\n')  # event-header

            # ── body ─────────────────────────────────────────────────────
            html_parts.append(f'  <div class="event-body">\n')

            # meta grid
            html_parts.append('    <div class="section-label">Event metadata</div>\n')
            html_parts.append('    <div class="meta-grid">\n')
            for label, key in [("Event ID", "event_id"), ("Year", "year"),
                                 ("Type", "event_type"), ("Location", "country"),
                                 ("Host club", "host_club"), ("Date", "date")]:
                val = row.get(key, "") or "—"
                html_parts.append(f'      <div class="meta-item"><b>{label}:</b> {_esc(val)}</div>\n')
            html_parts.append('    </div>\n')

            # review box
            if hl or row["diff_summary"] or disp:
                dec_bg = DISP_COLOUR.get(disp, "#f5f5f5")
                html_parts.append(f'    <div class="dec-box" style="background:{dec_bg}">\n')
                if hl:
                    html_parts.append(f'      <b>Heat:</b> {_esc(hl)} (score {row["heat_score"]}) &nbsp;')
                if row["diff_summary"]:
                    html_parts.append(f'<b>Diff:</b> {_esc(row["diff_summary"])} &nbsp;')
                if disp:
                    html_parts.append(f'<b>Disposition:</b> {_esc(disp)} &nbsp;')
                if row["pattern_family"]:
                    html_parts.append(f'<b>Pattern:</b> {_esc(row["pattern_family"])}<br>\n')
                if row["dec_note"]:
                    html_parts.append(f'      <i>{_esc(row["dec_note"])}</i>\n')
                if row["ki_note"]:
                    html_parts.append(f'      <br><b>Known issue:</b> {_esc(row["ki_note"])}\n')
                if row["qt_reason"]:
                    html_parts.append(f'      <br><b>Quarantine reason:</b> {_esc(row["qt_reason"])}\n')
                html_parts.append('    </div>\n')

            # divisions table
            div_details = row["_div_details"]
            if div_details:
                html_parts.append('    <div class="section-label">Stage2 divisions</div>\n')
                html_parts.append('    <table class="div-table">\n')
                html_parts.append('      <thead><tr><th>#</th><th>Division</th><th>Cat</th>'
                                   '<th>Count</th><th>Places</th><th>Top 3</th></tr></thead>\n')
                html_parts.append('      <tbody>\n')
                for i, dd in enumerate(div_details, 1):
                    cat  = dd["category"]
                    cat_cls = f"cat-{cat}" if cat in ("net","freestyle","golf") else "cat-other"
                    place_range = f"{dd['min_place']}–{dd['max_place']}" if dd["min_place"] != dd["max_place"] else str(dd["min_place"])
                    top3_str = " | ".join(_esc(n) for n in dd["top3"])
                    html_parts.append(
                        f'        <tr><td>{i}</td>'
                        f'<td>{_esc(dd["division"])}</td>'
                        f'<td class="{cat_cls}">{_esc(cat)}</td>'
                        f'<td>{dd["count"]}</td>'
                        f'<td>{place_range}</td>'
                        f'<td>{top3_str}</td></tr>\n'
                    )
                html_parts.append('      </tbody>\n    </table>\n')

            html_parts.append('  </div>\n')  # event-body
            html_parts.append('</div>\n')   # event-card

        html_parts.append('</div>\n')  # year-section

    html_parts.append("""</div>
<script>
function toggleBody(header) {
  const body = header.parentElement.querySelector('.event-body');
  const arrow = header.querySelector('.toggle-arrow');
  if (body.classList.toggle('open')) {
    arrow.textContent = '▼';
  } else {
    arrow.textContent = '▶';
  }
}
function filterEvents() {
  const q     = document.getElementById('search').value.toLowerCase();
  const heat  = document.getElementById('filter-heat').value;
  const year  = document.getElementById('filter-year').value;
  const ki    = document.getElementById('filter-ki').value;
  const type  = document.getElementById('filter-type').value;
  let visible = 0;
  document.querySelectorAll('.event-card').forEach(card => {
    const matchSearch = !q || card.dataset.search.includes(q);
    const matchHeat   = !heat || card.dataset.heat === heat;
    const matchYear   = !year || card.dataset.year === year;
    const matchKi     = !ki   || card.dataset.ki   === ki;
    const matchType   = !type || card.dataset.type.includes(type);
    const show = matchSearch && matchHeat && matchYear && matchKi && matchType;
    card.classList.toggle('hidden', !show);
    if (show) visible++;
  });
  // hide empty year sections
  document.querySelectorAll('.year-section').forEach(sec => {
    const hasVisible = Array.from(sec.querySelectorAll('.event-card'))
                            .some(c => !c.classList.contains('hidden'));
    sec.classList.toggle('hidden', !hasVisible);
  });
  document.getElementById('visible-count').textContent = visible + ' events shown';
}
function expandAll() {
  document.querySelectorAll('.event-card:not(.hidden) .event-body').forEach(b => {
    b.classList.add('open');
    b.parentElement.querySelector('.toggle-arrow').textContent = '▼';
  });
}
function collapseAll() {
  document.querySelectorAll('.event-body').forEach(b => {
    b.classList.remove('open');
    const arrow = b.parentElement.querySelector('.toggle-arrow');
    if (arrow) arrow.textContent = '▶';
  });
}
// init count
window.addEventListener('load', () => {
  document.getElementById('visible-count').textContent =
    document.querySelectorAll('.event-card').length + ' events shown';
});
</script>
</body>
</html>""")

    out_path.write_text("".join(html_parts), encoding="utf-8")
    size_kb = out_path.stat().st_size / 1024
    print(f"  Saved: {out_path} ({size_kb:.1f} KB, {len(table)} events)")


# ── entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    table = build_event_table()

    write_xlsx(table, REVIEW / "all_events_overview.xlsx")
    write_html(table, REVIEW / "all_events_inspector.html")

    print("\n=== DONE ===")
    print(f"  Spreadsheet: out/review/all_events_overview.xlsx")
    print(f"  HTML:        out/review/all_events_inspector.html")


if __name__ == "__main__":
    main()
