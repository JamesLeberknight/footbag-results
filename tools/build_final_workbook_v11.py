#!/usr/bin/env python3
"""
build_final_workbook_v11.py
Part 9: Build Final Community Workbook v11

Changes from v10:
- STATISTICS sheet rebuilt from authoritative canonical data with correct values:
    * Events documented = 774 (total in registry, all statuses incl. quarantined)
    * Events with results = 724 (non-quarantined + completed status)
    * Canonical players (registry) = 3,441 (Persons_Truth row count)
    * Players in results = 2,841 (distinct non-__NON_PERSON__ person_ids in PF)
    * Countries = 26 (real countries, excluding 'Global')
    * Cities = 181 (accent-normalised distinct city/country pairs)
    * Montreal / Montréal merged via accent normalisation (single row showing 48)
    * "Global" moved to bottom as "Multi-country / Online" row
    * Notes column added to DATASET OVERVIEW table
    * Year counts sourced from events_normalized.csv (authoritative)
    * Discipline counts sourced from Placements_Flat (non-quarantined)
    * freeze_panes = "A2" added
    * Column C widened to 30

All other sheets (README, DATA NOTES, EVENT INDEX, PLAYER SUMMARY,
CONSECUTIVE RECORDS) and all year sheets are copied unchanged from v10.
"""

import copy
import csv
import os
import sys
import unicodedata
from collections import Counter, defaultdict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

csv.field_size_limit(sys.maxsize)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SOURCE_V10_PATH = os.path.join(BASE_DIR, "Footbag_Results_Community_FINAL_v10.xlsx")
OUTPUT_PATH     = os.path.join(BASE_DIR, "Footbag_Results_Community_FINAL_v11.xlsx")
EVENTS_CSV      = os.path.join(BASE_DIR, "out", "canonical", "events_normalized.csv")
QUARANTINE_CSV  = os.path.join(BASE_DIR, "inputs", "review_quarantine_events.csv")
PF_CSV          = os.path.join(BASE_DIR, "out", "Placements_Flat.csv")
PT_CSV          = os.path.join(BASE_DIR, "out", "Persons_Truth.csv")


# ── Styles ────────────────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


FILL_NONE   = PatternFill(fill_type=None)
FILL_HEADER = _fill("D9D9D9")

FONT_SECTION = Font(bold=True, size=12)
FONT_SUBSECT = Font(bold=True, size=11)
FONT_HEADER  = Font(bold=True, size=11)
FONT_DATA    = Font(size=11)

ALIGN_LEFT   = Alignment(horizontal="left",  vertical="top")
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="top")


# ── Data computation ──────────────────────────────────────────────────────────

def normalize_city(c: str) -> str:
    return unicodedata.normalize("NFD", c).encode("ascii", "ignore").decode().lower().strip()


def compute_statistics() -> dict:
    with open(QUARANTINE_CSV, encoding="utf-8") as f:
        quarantine = {str(r["event_id"]) for r in csv.DictReader(f)}

    with open(EVENTS_CSV, encoding="utf-8") as f:
        all_events = list(csv.DictReader(f))

    with open(PF_CSV, encoding="utf-8") as f:
        pf_rows = list(csv.DictReader(f))

    with open(PT_CSV, encoding="utf-8") as f:
        pt_rows = list(csv.DictReader(f))

    # Dataset overview
    total_events      = len(all_events)
    quarantined_count = sum(1 for r in all_events if str(r.get("legacy_event_id", "")) in quarantine)
    non_q             = [r for r in all_events if str(r.get("legacy_event_id", "")) not in quarantine]
    events_with_results = sum(1 for r in non_q if r.get("status", "") == "completed")

    years = sorted(set(int(r["year"]) for r in all_events if r.get("year", "").strip()))
    year_range = f"{years[0]}\u2013{years[-1]}"  # en-dash

    real_countries = sorted(
        set(
            r.get("country", "").strip()
            for r in non_q
            if r.get("country", "").strip() and r.get("country", "").strip() != "Global"
        )
    )
    country_count = len(real_countries)

    city_keys = set()
    for r in non_q:
        city = r.get("city", "").strip()
        country = r.get("country", "").strip()
        if city:
            city_keys.add((normalize_city(city), country))
    city_count = len(city_keys)

    total_placements  = len(pf_rows)
    canonical_players = len(pt_rows)
    players_in_results = len(
        set(
            r.get("person_id", "")
            for r in pf_rows
            if r.get("person_id", "") not in ("", "__NON_PERSON__")
        )
    )

    # Events by year (all events incl. quarantined)
    year_counts = Counter(int(r["year"]) for r in all_events if r.get("year", "").strip())
    events_by_year = [(y, year_counts[y]) for y in sorted(year_counts)]

    # Discipline history (non-quarantined events)
    eid_to_year = {
        r.get("legacy_event_id", ""): int(r["year"])
        for r in all_events
        if r.get("year", "").strip() and r.get("legacy_event_id", "")
    }

    disc_events     = defaultdict(set)
    disc_first_year = defaultdict(lambda: 9999)

    for row in pf_rows:
        eid = row.get("event_id", "")
        if eid in quarantine:
            continue
        cat = row.get("division_category", "").strip()
        if not cat:
            continue
        year = eid_to_year.get(eid)
        if year:
            disc_events[cat].add(eid)
            if year < disc_first_year[cat]:
                disc_first_year[cat] = year

    DISC_LABELS = {
        "net":       "Net (Rallye / Side-Out)",
        "freestyle": "Freestyle / Shred / Circles",
        "golf":      "Footbag Golf",
        "sideline":  "Consecutive / Sideline",
        "unknown":   "Unknown / Unclassified",
    }
    disciplines = [
        (DISC_LABELS.get(cat, cat.title()), fy, len(disc_events[cat]))
        for cat, fy in sorted(disc_first_year.items(), key=lambda x: x[1])
    ]

    # Events by country
    country_counts = Counter(
        r.get("country", "").strip()
        for r in non_q
        if r.get("country", "").strip()
    )
    global_count = country_counts.pop("Global", 0)
    ranked_countries = [(c, n) for c, n in country_counts.most_common()]

    # Top host cities (accent-normalised, pick canonical spelling)
    city_counter    = Counter()
    city_canonical  = defaultdict(Counter)
    for r in non_q:
        city    = r.get("city", "").strip()
        country = r.get("country", "").strip()
        if city:
            norm = normalize_city(city)
            key  = (norm, country)
            city_counter[key] += 1
            city_canonical[key][city] += 1

    top_cities = [
        (city_canonical[k].most_common(1)[0][0], k[1], count)
        for k, count in city_counter.most_common(20)
    ]

    return {
        "total_events":        total_events,
        "quarantined_count":   quarantined_count,
        "events_with_results": events_with_results,
        "year_range":          year_range,
        "country_count":       country_count,
        "city_count":          city_count,
        "total_placements":    total_placements,
        "canonical_players":   canonical_players,
        "players_in_results":  players_in_results,
        "events_by_year":      events_by_year,
        "disciplines":         disciplines,
        "ranked_countries":    ranked_countries,
        "global_count":        global_count,
        "top_cities":          top_cities,
    }


# ── Sheet helpers ──────────────────────────────────────────────────────────────

def _w(ws, row: int, col: int, value=None, *,
       font=None, fill=None, align=None):
    """Write a cell with optional formatting."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font  is not None: cell.font      = font
    if fill  is not None: cell.fill      = fill
    if align is not None: cell.alignment = align
    return cell


def _section_title(ws, row: int, text: str) -> int:
    """Write a bold section title in col A and return next row."""
    _w(ws, row, 1, text, font=FONT_SECTION, align=ALIGN_LEFT)
    return row + 1


def _sub_header(ws, row: int, text: str) -> int:
    """Write a bold sub-header in col A and return next row."""
    _w(ws, row, 1, text, font=FONT_SUBSECT, align=ALIGN_LEFT)
    return row + 1


def _table_headers(ws, row: int, *headers) -> int:
    """Write bold grey header row and return next row."""
    for col, h in enumerate(headers, 1):
        _w(ws, row, col, h, font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT)
    return row + 1


def _data_row(ws, row: int, *values) -> int:
    """Write plain data row; right-align numeric values."""
    for col, v in enumerate(values, 1):
        align = ALIGN_RIGHT if isinstance(v, (int, float)) else ALIGN_LEFT
        _w(ws, row, col, v, font=FONT_DATA, align=align)
    return row + 1


def build_statistics_sheet(wb: Workbook, stats: dict) -> None:
    """Create (or replace) the STATISTICS sheet with accurate, consistent data."""
    # Remove existing sheet if present
    if "STATISTICS" in wb.sheetnames:
        idx = wb.sheetnames.index("STATISTICS")
        del wb["STATISTICS"]
        ws = wb.create_sheet("STATISTICS", idx)
    else:
        ws = wb.create_sheet("STATISTICS")

    row = 1

    # ── DATASET OVERVIEW ────────────────────────────────────────────────────────
    row = _section_title(ws, row, "DATASET OVERVIEW")
    row = _table_headers(ws, row, "Metric", "Value", "Notes")

    overview_rows = [
        ("Events documented",           stats["total_events"],
         f"Includes {stats['quarantined_count']} quarantined events"),
        ("Events with results",          stats["events_with_results"],
         "Excludes quarantined events"),
        ("Years covered",                stats["year_range"],
         ""),
        ("Countries represented",        stats["country_count"],
         "Excludes multi-country / online events"),
        ("Cities represented",           stats["city_count"],
         "Accent-normalised"),
        ("Total placements",             stats["total_placements"],
         "All rows in Placements_Flat"),
        ("Canonical players (registry)", stats["canonical_players"],
         "All persons in Persons_Truth"),
        ("Players appearing in results", stats["players_in_results"],
         "Distinct persons with at least one placement"),
    ]
    for metric, value, note in overview_rows:
        row = _data_row(ws, row, metric, value, note or None)

    row += 1  # blank

    # ── EVENTS BY YEAR ──────────────────────────────────────────────────────────
    row = _section_title(ws, row, "EVENTS BY YEAR")
    row = _table_headers(ws, row, "Year", "Events")
    for year, count in stats["events_by_year"]:
        row = _data_row(ws, row, year, count)

    row += 1  # blank

    # ── DISCIPLINE HISTORY ──────────────────────────────────────────────────────
    row = _section_title(ws, row, "DISCIPLINE HISTORY")
    row = _table_headers(ws, row, "Discipline", "First Year", "Events")
    for disc_label, first_year, n_events in stats["disciplines"]:
        row = _data_row(ws, row, disc_label, first_year, n_events)

    row += 1  # blank

    # ── GEOGRAPHIC DISTRIBUTION ─────────────────────────────────────────────────
    row = _section_title(ws, row, "GEOGRAPHIC DISTRIBUTION")

    # Events by country
    row = _sub_header(ws, row, "Events by Country")
    row = _table_headers(ws, row, "Country", "Events")
    for country, count in stats["ranked_countries"]:
        row = _data_row(ws, row, country, count)
    # Global / Online at the bottom
    if stats["global_count"]:
        row += 1  # blank separator
        row = _data_row(ws, row, "Multi-country / Online", stats["global_count"])

    row += 1  # blank

    # Top host cities
    row = _sub_header(ws, row, "Top Host Cities")
    row = _table_headers(ws, row, "City", "Country", "Events")
    for city, country, count in stats["top_cities"]:
        row = _data_row(ws, row, city, country, count)

    # ── Column widths ────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30

    # ── Freeze pane ──────────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    print(f"  STATISTICS sheet written: {row - 1} rows")


# ── Sheet copy utilities ───────────────────────────────────────────────────────

def copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        for attr in ("font", "fill", "alignment", "border", "number_format"):
            try:
                val = getattr(src_cell, attr)
                setattr(dst_cell, attr,
                        copy.copy(val) if attr != "number_format" else val)
            except Exception:
                pass
    if src_cell.hyperlink:
        try:
            dst_cell.hyperlink = copy.copy(src_cell.hyperlink)
        except Exception:
            pass


def copy_sheet_to(src_ws, dst_ws):
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width  = col_dim.width or 8.43
        dst_ws.column_dimensions[col_letter].hidden = col_dim.hidden
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height or 15
    for row in src_ws.iter_rows():
        for cell in row:
            copy_cell(cell, dst_ws.cell(row=cell.row, column=cell.column))
    for merge_range in src_ws.merged_cells.ranges:
        try:
            dst_ws.merge_cells(str(merge_range))
        except Exception:
            pass
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Computing statistics from canonical data sources...")
    stats = compute_statistics()

    print("\nOpening source workbook (v10)...")
    src_wb = openpyxl.load_workbook(SOURCE_V10_PATH)
    print(f"  Sheets in v10: {src_wb.sheetnames[:8]} ...")

    print("\nCreating output workbook...")
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    # Copy all sheets from v10 in order, replacing STATISTICS with rebuilt version
    for sheet_name in src_wb.sheetnames:
        if sheet_name == "STATISTICS":
            print(f"  Rebuilding STATISTICS sheet...")
            dst_ws = out_wb.create_sheet(sheet_name)
            # build_statistics_sheet will replace it via index logic
            build_statistics_sheet(out_wb, stats)
        else:
            print(f"  Copying {sheet_name}...", end="", flush=True)
            dst_ws = out_wb.create_sheet(sheet_name)
            copy_sheet_to(src_wb[sheet_name], dst_ws)
            print(" done")

    # Verify sheet order
    print(f"\nSheet order: {out_wb.sheetnames[:8]} ...")
    front_expected = ["README", "DATA NOTES", "STATISTICS", "EVENT INDEX",
                      "PLAYER SUMMARY", "CONSECUTIVE RECORDS"]
    actual_front = list(out_wb.sheetnames[:6])
    if actual_front == front_expected:
        print("  Front sheet order: CORRECT")
    else:
        print(f"  WARNING: Expected {front_expected}, got {actual_front}")

    year_sheets = [s for s in out_wb.sheetnames if s.isdigit()]
    print(f"  Year sheets: {year_sheets[0]}–{year_sheets[-1]} ({len(year_sheets)} sheets)")

    # Save
    print(f"\nSaving to: {OUTPUT_PATH}")
    out_wb.save(OUTPUT_PATH)
    size_bytes = os.path.getsize(OUTPUT_PATH)
    size_mb    = size_bytes / (1024 * 1024)
    print(f"Saved: {size_mb:.1f} MB ({size_bytes:,} bytes)")
    print("\nDone.")


if __name__ == "__main__":
    main()
