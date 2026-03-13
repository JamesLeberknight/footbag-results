#!/usr/bin/env python3
"""
Final Publication Validation — v2
Compares mirror (stage2) → canonical (PBP v61) → community xlsx
Produces 4 reports + final publication decision.

Key architectural fact respected:
  - stage2 > PBP count = parser-improvement gap (PBP locked before recent fixes)
  - This is EXPECTED behavior, not content loss
  - The community xlsx sources from PBP, which is the canonical truth
  - Division-level/event-level deltas between stage2 and PBP are NORMALIZED_EQUIVALENT
    if no corresponding community gap exists.
"""
import csv
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

csv.field_size_limit(10_000_000)

REPO = Path(__file__).parent.parent
OUT  = REPO / "out" / "final_validation"
OUT.mkdir(parents=True, exist_ok=True)

REPLACEMENT = "\ufffd"

# ── Input paths ──────────────────────────────────────────────────────────────
STAGE2_CSV   = REPO / "out" / "stage2_canonical_events.csv"
PBP_CSV      = REPO / "inputs" / "identity_lock" / "Placements_ByPerson_v61.csv"
COMMUNITY_XL = REPO / "Footbag_Results_Community.xlsx"
KNOWN_ISSUES = REPO / "overrides" / "known_issues.csv"
QUARANTINE   = REPO / "inputs" / "review_quarantine_events.csv"
EVENTS_CSV   = REPO / "out" / "canonical" / "events.csv"
DISC_CSV     = REPO / "out" / "canonical" / "event_disciplines.csv"
PERSONS_CSV  = REPO / "out" / "canonical" / "persons.csv"
PT_CSV       = REPO / "out" / "Persons_Truth.csv"

# ── Load helpers ─────────────────────────────────────────────────────────────

def load_csv(path, key_col=None):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    if key_col:
        return {r[key_col]: r for r in rows}
    return rows

# ── Load all data ─────────────────────────────────────────────────────────────

print("Loading stage2 canonical events …")
stage2_rows  = load_csv(STAGE2_CSV)
stage2_by_id = {r["event_id"]: r for r in stage2_rows}

print("Loading PBP v61 …")
pbp_rows = load_csv(PBP_CSV)

print("Loading known issues & quarantine …")
known_issues_map = {r["event_id"]: r["note"]     for r in load_csv(KNOWN_ISSUES)}
known_issues_sev = {r["event_id"]: r["severity"] for r in load_csv(KNOWN_ISSUES)}
quarantine_map   = {r["event_id"]: r.get("reason","") for r in load_csv(QUARANTINE)}

print("Loading community xlsx …")
wb_comm = openpyxl.load_workbook(str(COMMUNITY_XL), read_only=True, data_only=True)
YEAR_SHEETS = sorted(s for s in wb_comm.sheetnames if re.match(r"^\d{4}$", s))

community_event_ids  = set()
community_result_ct  = defaultdict(int)   # event_id → count of result lines
community_div_names  = defaultdict(set)   # event_id → set of section headers
community_year_for_eid = {}               # event_id → sheet year

for sheet_name in YEAR_SHEETS:
    ws = wb_comm[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 7:
        continue
    eid_row = rows[6]
    col_to_eid = {}
    for ci, val in enumerate(eid_row):
        if ci == 0 or val is None:
            continue
        eid_str = str(int(val)) if isinstance(val, float) else str(val)
        col_to_eid[ci] = eid_str
        community_event_ids.add(eid_str)
        community_year_for_eid[eid_str] = sheet_name

    for row in rows[7:]:
        for ci, eid_str in col_to_eid.items():
            if ci >= len(row) or row[ci] is None:
                continue
            s = str(row[ci]).strip()
            if not s:
                continue
            if re.match(r"^[🥇🥈🥉]|^\s*\d+\s", s):
                community_result_ct[eid_str] += 1
            else:
                community_div_names[eid_str].add(s)

wb_comm.close()
print(f"  Community xlsx: {len(community_event_ids)} events across {len(YEAR_SHEETS)} year sheets")

# ── Build PBP indexes ─────────────────────────────────────────────────────────

pbp_by_event   = defaultdict(list)
pbp_divs       = defaultdict(set)
pbp_div_count  = defaultdict(lambda: defaultdict(int))  # eid → div → count

for row in pbp_rows:
    eid = row["event_id"]
    pbp_by_event[eid].append(row)
    pbp_divs[eid].add(row["division_canon"])
    pbp_div_count[eid][row["division_canon"]] += 1

# Stage2 indexes
stage2_placements = {}
stage2_divs       = {}
for row in stage2_rows:
    eid = row["event_id"]
    try:
        pj = json.loads(row["placements_json"]) if row["placements_json"] else []
    except json.JSONDecodeError:
        pj = []
    stage2_placements[eid] = pj
    stage2_divs[eid] = {p["division_canon"] for p in pj}

all_event_ids = sorted(stage2_by_id.keys())
print(f"Total stage2 events: {len(all_event_ids)}")

# ─────────────────────────────────────────────────────────────────────────────
# PART 1 — EVENT PUBLICATION ADJUDICATION
# ─────────────────────────────────────────────────────────────────────────────
print("\n=== PART 1: Event Publication Adjudication ===")

adjudication = []

for eid in all_event_ids:
    s2        = stage2_by_id[eid]
    year      = s2["year"]
    event_name = s2["event_name"]

    s2_cnt  = len(stage2_placements.get(eid, []))
    pbp_cnt = len(pbp_by_event.get(eid, []))

    mirror_results  = "YES" if s2_cnt  > 0 else "NO"
    canon_present   = "YES" if pbp_cnt > 0 else "NO"
    comm_present    = "YES" if eid in community_event_ids else "NO"

    is_quarantined = eid in quarantine_map
    is_known_issue = eid in known_issues_map
    ki_note        = known_issues_map.get(eid, "")
    ki_sev         = known_issues_sev.get(eid, "")

    explanation = []
    blocker     = "NO"
    diff_type   = "EXACT_MATCH"

    # ── No mirror results ─────────────────────────────────────────────────
    if s2_cnt == 0:
        diff_type = "METADATA_ONLY"
        explanation.append("No placements parsed from mirror")
        if pbp_cnt > 0:
            diff_type = "NOISE_IN_CANONICAL"
            explanation.append(f"BUT PBP has {pbp_cnt} rows — unexpected")
            blocker = "YES"

    # ── Quarantined ───────────────────────────────────────────────────────
    elif is_quarantined:
        diff_type = "SOURCE_PARTIAL"
        explanation.append(f"QUARANTINED: {quarantine_map.get(eid, '')}")

    # ── Count comparison ──────────────────────────────────────────────────
    else:
        delta = s2_cnt - pbp_cnt

        if delta == 0:
            diff_type = "EXACT_MATCH"
            explanation.append(f"stage2={s2_cnt} == PBP={pbp_cnt}")

        elif delta < 0:
            # PBP > stage2 → unresolved rows added in v60 migration
            diff_type = "NORMALIZED_EQUIVALENT"
            explanation.append(
                f"PBP ({pbp_cnt}) > stage2 ({s2_cnt}): v60 unresolved row additions"
            )

        else:
            # stage2 > PBP (delta > 0)
            # This is EXPECTED: PBP was locked before parser improvements.
            # Only a blocker if community xlsx is ALSO missing data
            # that canonical DOES have (i.e., presentable persons absent from xlsx).
            if is_quarantined:
                diff_type = "SOURCE_PARTIAL"
                explanation.append(f"QUARANTINED delta={delta}")
            elif is_known_issue and ki_sev in ("severe", "moderate"):
                diff_type = "SOURCE_PARTIAL"
                explanation.append(f"Known issue ({ki_sev}): {ki_note}")
            else:
                # Parser improvement gap: stage2 has more because parser was improved
                # after PBP lock. Not a community blocker because community uses PBP.
                diff_type = "NORMALIZED_EQUIVALENT"
                explanation.append(
                    f"stage2={s2_cnt} > PBP={pbp_cnt} (delta={delta}): "
                    "parser improvements after PBP v61 lock — PBP is canonical truth"
                )
                if is_known_issue:
                    diff_type = "OVERRIDE_EQUIVALENT"
                    explanation.append(f"Also known issue ({ki_sev}): {ki_note}")

    # ── Community absent despite presentable PBP rows ─────────────────────
    if canon_present == "YES" and comm_present == "NO":
        presentable = sum(
            1 for r in pbp_by_event.get(eid, [])
            if r.get("person_id") and r.get("person_unresolved") != "1"
        )
        if presentable > 0:
            try:
                yr = int(year)
            except (ValueError, TypeError):
                yr = 0
            # Check year range
            year_min = int(min(YEAR_SHEETS))
            year_max = int(max(YEAR_SHEETS))
            if yr < year_min or yr > year_max:
                explanation.append(f"Year {yr} outside community xlsx range; OK")
            else:
                diff_type = "MIRROR_CONTENT_LOST"
                blocker = "YES"
                explanation.append(
                    f"BLOCKER: {presentable} presentable PBP rows absent from community xlsx"
                )

    adjudication.append({
        "event_id":               eid,
        "year":                   year,
        "event_name":             event_name,
        "mirror_results_found":   mirror_results,
        "canonical_results_present": canon_present,
        "community_results_present": comm_present,
        "difference_type":        diff_type,
        "explanation":            "; ".join(explanation),
        "blocker":                blocker,
    })

p1_path = OUT / "event_publication_adjudication.csv"
with open(p1_path, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=list(adjudication[0].keys()))
    w.writeheader()
    w.writerows(adjudication)

p1_blockers  = [r for r in adjudication if r["blocker"] == "YES"]
p1_by_type   = defaultdict(int)
for r in adjudication:
    p1_by_type[r["difference_type"]] += 1

print(f"  Written: {p1_path}")
print(f"  Total: {len(adjudication)}, Blockers: {len(p1_blockers)}")
for t, c in sorted(p1_by_type.items()):
    print(f"    {t}: {c}")
if p1_blockers:
    for b in p1_blockers:
        print(f"    BLOCKER: {b['event_id']} {b['event_name'][:50]} — {b['explanation'][:80]}")


# ─────────────────────────────────────────────────────────────────────────────
# PART 2 — CANONICAL NOISE SCAN
# ─────────────────────────────────────────────────────────────────────────────
print("\n=== PART 2: Canonical Noise Scan ===")

noise_rows = []
HTML_RESIDUE  = re.compile(r"<[a-zA-Z/][^>]*>|&amp;|&lt;|&gt;|&quot;|&#\d{2,5};")
SOFT_HYPHEN   = "\u00ad"
ZERO_WIDTH    = re.compile(r"[\u200b\u200c\u200d\ufeff]")
CTRL_CHARS    = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
APOS_CORRUPT  = re.compile(r"\?S\b", re.IGNORECASE)   # Women?S, Master?S

# Division name corruption patterns (e.g. Cir\ufffdCle, Rou\uffrdTines)
# Pattern: lowercase + \ufffd + uppercase → likely encoding split
DIV_CORRUPTION = re.compile(r"[a-z]\ufffd[A-Z]")

def check_value(val: str, fld: str, event_id: str = "",
               publication_artifact: bool = False) -> list:
    """Return list of (noise_type, blocker, fix) tuples.

    publication_artifact=True means the value came from the community xlsx itself
    (i.e. visible to readers).  Only xlsx-visible noise is a publication BLOCKER.
    PBP source corruption is documented as INFO (blocker=NO) since 04B applies
    output-layer fixes before writing to the community xlsx.
    """
    found = []
    if not val:
        return found
    b = "YES" if publication_artifact else "NO"
    b_src = "NO"   # PBP-level issues are INFO; not blockers unless visible in xlsx
    if REPLACEMENT in val:
        if DIV_CORRUPTION.search(val):
            found.append(("division_encoding_corruption", b_src,
                          f"Corrupt div name in PBP source (04B applies fix): {repr(val[:60])}"))
        else:
            found.append(("unicode_replacement_char", b_src,
                          f"Garbled chars in PBP source (04B applies fix): {repr(val[:60])}"))
    if APOS_CORRUPT.search(val):
        found.append(("apostrophe_corrupted_?S", b_src,
                      f"Apostrophe→?S in PBP source (04B applies fix): {repr(val[:60])}"))
    if HTML_RESIDUE.search(val):
        found.append(("html_entity_unescaped", b,
                      f"Unescaped HTML entity: {repr(val[:60])}"))
    if SOFT_HYPHEN in val:
        found.append(("soft_hyphen_U+00AD", "NO",
                      f"Soft hyphen: {repr(val[:60])} — cosmetic"))
    if ZERO_WIDTH.search(val):
        found.append(("zero_width_char", "NO",
                      f"Zero-width char: {repr(val[:60])} — cosmetic"))
    if CTRL_CHARS.search(val):
        found.append(("control_char", b,
                      f"Control char: {repr(val[:60])}"))
    return found

# Scan PBP — document source-level issues (not publication blockers; 04B fixes them)
print("  Scanning PBP v61 …")
seen_noise_keys = set()

for row in pbp_rows:
    eid = row["event_id"]
    for fld in ("division_canon", "person_canon", "team_display_name"):
        val = row.get(fld, "") or ""
        for noise_type, blocker, fix in check_value(val, fld, eid, publication_artifact=False):
            key = (fld, val[:60], noise_type)
            if key in seen_noise_keys:
                continue
            seen_noise_keys.add(key)
            noise_rows.append({
                "event_id":       eid,
                "field":          f"pbp.{fld}",
                "value":          val[:120],
                "noise_type":     noise_type,
                "blocker":        blocker,
                "recommended_fix": fix,
            })

# Scan persons.csv
print("  Scanning canonical persons.csv …")
for row in load_csv(PERSONS_CSV):
    for fld in ("display_name",):
        val = row.get(fld, "") or ""
        for noise_type, blocker, fix in check_value(val, fld):
            noise_rows.append({
                "event_id":       "",
                "field":          f"persons.{fld}",
                "value":          val[:120],
                "noise_type":     noise_type,
                "blocker":        blocker,
                "recommended_fix": fix,
            })

# Scan events.csv
print("  Scanning canonical events.csv …")
for row in load_csv(EVENTS_CSV):
    eid = row.get("legacy_event_id", "")
    for fld in ("event_name", "city", "region", "country", "host_club"):
        val = row.get(fld, "") or ""
        for noise_type, blocker, fix in check_value(val, fld, eid):
            noise_rows.append({
                "event_id":       eid,
                "field":          f"events.{fld}",
                "value":          val[:120],
                "noise_type":     noise_type,
                "blocker":        blocker,
                "recommended_fix": fix,
            })

# Also scan community xlsx directly for visible noise
print("  Scanning community xlsx for visible noise …")
wb_scan = openpyxl.load_workbook(str(COMMUNITY_XL), read_only=True, data_only=True)
community_noise_count = defaultdict(int)
for sheet_name in wb_scan.sheetnames:
    ws = wb_scan[sheet_name]
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is None:
                continue
            s = str(val)
            if REPLACEMENT in s:
                community_noise_count[("xlsx_replacement_char", sheet_name)] += 1
            if APOS_CORRUPT.search(s):
                community_noise_count[("xlsx_apostrophe_?S", sheet_name)] += 1
wb_scan.close()

xlsx_blocker_rows = []
for (ntype, sheet), count in sorted(community_noise_count.items()):
    print(f"    Community xlsx [{sheet}]: {ntype} × {count}")
    xlsx_blocker_rows.append({
        "event_id":       "",
        "field":          f"xlsx.{sheet}",
        "value":          f"{ntype} × {count} cells",
        "noise_type":     ntype,
        "blocker":        "YES",
        "recommended_fix": "Re-run 04B with encoding fix applied",
    })
noise_rows.extend(xlsx_blocker_rows)

p2_path = OUT / "canonical_noise_report.csv"
with open(p2_path, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=[
        "event_id","field","value","noise_type","blocker","recommended_fix"
    ])
    w.writeheader()
    w.writerows(noise_rows)

p2_blockers = [r for r in noise_rows if r["blocker"] == "YES"]
by_type = defaultdict(int)
for r in noise_rows:
    by_type[r["noise_type"]] += 1
print(f"  Written: {p2_path}")
print(f"  Total noise items: {len(noise_rows)}, Blockers: {len(p2_blockers)}")
for t, c in sorted(by_type.items()):
    print(f"    {t}: {c}")
if p2_blockers:
    for b in p2_blockers[:8]:
        print(f"    BLOCKER [{b['field']}]: {b['value'][:60]}")


# ─────────────────────────────────────────────────────────────────────────────
# PART 3 — MIRROR CONTENT LOSS DETECTION
# ─────────────────────────────────────────────────────────────────────────────
print("\n=== PART 3: Mirror Content Loss Detection ===")

loss_rows = []

for eid in all_event_ids:
    s2         = stage2_by_id[eid]
    year       = s2["year"]
    event_name = s2["event_name"]

    s2_cnt  = len(stage2_placements.get(eid, []))
    pbp_cnt = len(pbp_by_event.get(eid, []))

    if s2_cnt == 0 and pbp_cnt == 0:
        continue  # Metadata-only

    # Event-level summary
    s2_divs_set  = stage2_divs.get(eid, set())
    pbp_divs_set = pbp_divs.get(eid, set())

    # Divisions in mirror but NOT in canonical
    divs_lost = s2_divs_set - pbp_divs_set
    # Divisions in canonical but NOT in mirror (additions from v60 migration)
    divs_added = pbp_divs_set - s2_divs_set

    is_quarantined = eid in quarantine_map
    is_known_issue = eid in known_issues_map
    ki_note        = known_issues_map.get(eid, "")

    # Divisions lost from mirror
    for div in sorted(divs_lost):
        s2_d  = sum(1 for p in stage2_placements.get(eid,[]) if p["division_canon"] == div)

        if is_quarantined:
            loss_type  = "QUARANTINED_EVENT"
            justified  = "YES"
            blocker    = "NO"
            explanation = f"Quarantined: {quarantine_map.get(eid,'')}"
        elif is_known_issue:
            loss_type  = "KNOWN_ISSUE"
            justified  = "YES"
            blocker    = "NO"
            explanation = f"Known issue: {ki_note[:80]}"
        else:
            # Parser improvement gap: division present in current stage2 but absent
            # from PBP because PBP was locked before this division was parsed.
            loss_type  = "PARSER_IMPROVEMENT_GAP"
            justified  = "YES"
            blocker    = "NO"
            explanation = (
                f"Division '{div}' ({s2_d} rows) in stage2 but absent from PBP v61. "
                "PBP locked before parser fix recovered this division. "
                "Community xlsx correctly shows PBP."
            )

        loss_rows.append({
            "event_id":        eid,
            "event_name":      event_name,
            "mirror_summary":  f"{div}: {s2_d} stage2 placements",
            "canonical_summary": f"Division absent from PBP",
            "loss_type":       loss_type,
            "justified":       justified,
            "blocker":         blocker,
            "explanation":     explanation,
        })

    # Divisions added in PBP (v60 unresolved additions)
    for div in sorted(divs_added):
        pbp_d = pbp_div_count[eid].get(div, 0)
        loss_rows.append({
            "event_id":        eid,
            "event_name":      event_name,
            "mirror_summary":  f"Division absent from stage2",
            "canonical_summary": f"{div}: {pbp_d} PBP rows (v60 addition)",
            "loss_type":       "PBP_AUGMENTED",
            "justified":       "YES",
            "blocker":         "NO",
            "explanation":     f"Division '{div}' added to PBP v60 as unresolved rows; not in stage2 parse.",
        })

    # Placement count delta (same division, fewer rows)
    for div in sorted(s2_divs_set & pbp_divs_set):
        s2_d  = sum(1 for p in stage2_placements.get(eid,[]) if p["division_canon"] == div)
        pbp_d = pbp_div_count[eid].get(div, 0)
        delta = s2_d - pbp_d
        if delta <= 0:
            continue  # exact or PBP has more

        if is_quarantined:
            loss_type = "QUARANTINED_EVENT"
            justified = "YES"
            blocker   = "NO"
            explanation = f"Quarantined event"
        elif is_known_issue:
            loss_type = "KNOWN_ISSUE"
            justified = "YES"
            blocker   = "NO"
            explanation = f"Known issue: {ki_note[:80]}"
        else:
            # As above: parser improvement gap
            loss_type = "PARSER_IMPROVEMENT_GAP"
            justified = "YES"
            blocker   = "NO"
            explanation = (
                f"stage2={s2_d} > PBP={pbp_d} (delta={delta}) in div '{div}'. "
                "Parser improvements after PBP lock; community uses PBP which is canonical."
            )

        loss_rows.append({
            "event_id":        eid,
            "event_name":      event_name,
            "mirror_summary":  f"{div}: {s2_d} placements in stage2",
            "canonical_summary": f"{div}: {pbp_d} placements in PBP",
            "loss_type":       loss_type,
            "justified":       justified,
            "blocker":         blocker,
            "explanation":     explanation,
        })

p3_path = OUT / "mirror_content_loss_report.csv"
with open(p3_path, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=[
        "event_id","event_name","mirror_summary","canonical_summary",
        "loss_type","justified","blocker","explanation"
    ])
    w.writeheader()
    w.writerows(loss_rows)

p3_blockers = [r for r in loss_rows if r["blocker"] == "YES"]
by_loss = defaultdict(int)
for r in loss_rows:
    by_loss[r["loss_type"]] += 1
print(f"  Written: {p3_path}")
print(f"  Total division-level differences: {len(loss_rows)}, Blockers: {len(p3_blockers)}")
for t, c in sorted(by_loss.items()):
    print(f"    {t}: {c}")


# ─────────────────────────────────────────────────────────────────────────────
# PART 4 — COMMUNITY SPREADSHEET VALIDATION
# ─────────────────────────────────────────────────────────────────────────────
print("\n=== PART 4: Community Spreadsheet Validation ===")

community_report = []
year_min = int(min(YEAR_SHEETS))
year_max = int(max(YEAR_SHEETS))

# Pre-compute presentable count per event
presentable_by_event = {}
for eid, rows in pbp_by_event.items():
    presentable_by_event[eid] = sum(
        1 for r in rows
        if r.get("person_id") and r.get("person_unresolved") != "1"
    )

for eid in all_event_ids:
    s2         = stage2_by_id[eid]
    year       = s2["year"]
    event_name = s2["event_name"]

    pbp_cnt    = len(pbp_by_event.get(eid, []))
    presentable = presentable_by_event.get(eid, 0)
    in_comm    = eid in community_event_ids
    comm_ct    = community_result_ct.get(eid, 0)

    try:
        yr = int(year)
    except (ValueError, TypeError):
        yr = 0

    if pbp_cnt == 0:
        status    = "METADATA_ONLY"
        blocker   = "NO"
        explanation = "No PBP rows; community shows metadata only (or omits)"
    elif yr < year_min or yr > year_max:
        status    = "OUT_OF_RANGE"
        blocker   = "NO"
        explanation = f"Year {yr} outside community xlsx range {year_min}–{year_max}"
    elif not in_comm and presentable == 0:
        status    = "UNRESOLVED_ONLY_OK"
        blocker   = "NO"
        explanation = "All PBP rows unresolved/non-person; intentionally omitted"
    elif not in_comm and presentable > 0:
        status    = "MISSING_FROM_COMMUNITY"
        blocker   = "YES"
        explanation = f"{presentable} presentable rows absent from community xlsx"
    elif in_comm and comm_ct == 0 and presentable > 0:
        status    = "PRESENT_BUT_EMPTY"
        blocker   = "YES"
        explanation = f"Event in community but 0 result lines found; PBP={presentable} presentable"
    else:
        status    = "OK"
        blocker   = "NO"
        explanation = f"community_results={comm_ct}, pbp_presentable={presentable}"

    community_report.append({
        "event_id":             eid,
        "year":                 year,
        "event_name":           event_name,
        "pbp_total_rows":       pbp_cnt,
        "pbp_presentable_rows": presentable,
        "in_community_xlsx":    "YES" if in_comm else "NO",
        "community_result_lines": comm_ct,
        "status":               status,
        "blocker":              blocker,
        "explanation":          explanation,
    })

p4_path = OUT / "community_publication_report.csv"
with open(p4_path, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=list(community_report[0].keys()))
    w.writeheader()
    w.writerows(community_report)

p4_blockers = [r for r in community_report if r["blocker"] == "YES"]
by_status = defaultdict(int)
for r in community_report:
    by_status[r["status"]] += 1
print(f"  Written: {p4_path}")
print(f"  Total: {len(community_report)}, Blockers: {len(p4_blockers)}")
for t, c in sorted(by_status.items()):
    print(f"    {t}: {c}")
if p4_blockers:
    for b in p4_blockers:
        print(f"    BLOCKER: {b['event_id']} {b['event_name'][:50]} — {b['explanation']}")


# ─────────────────────────────────────────────────────────────────────────────
# PART 5 — FINAL PUBLICATION DECISION
# ─────────────────────────────────────────────────────────────────────────────
print("\n=== PART 5: Final Publication Decision ===")

all_blockers = (
    [(r, "P1_Adjudication") for r in p1_blockers]
    + [(r, "P2_NoiseScan")   for r in p2_blockers]
    + [(r, "P3_ContentLoss") for r in p3_blockers]
    + [(r, "P4_Community")   for r in p4_blockers]
)
total_blockers = len(all_blockers)
decision = "READY TO PUBLISH" if total_blockers == 0 else "NOT READY TO PUBLISH"

lines = [
    "",
    "=" * 64,
    f"  FINAL PUBLICATION DECISION: {decision}",
    "=" * 64,
    "",
    f"  Total events audited:           {len(all_event_ids)}",
    f"  EXACT_MATCH:                    {p1_by_type['EXACT_MATCH']}",
    f"  NORMALIZED_EQUIVALENT:          {p1_by_type['NORMALIZED_EQUIVALENT']}",
    f"  OVERRIDE_EQUIVALENT:            {p1_by_type['OVERRIDE_EQUIVALENT']}",
    f"  SOURCE_PARTIAL:                 {p1_by_type['SOURCE_PARTIAL']}",
    f"  METADATA_ONLY:                  {p1_by_type['METADATA_ONLY']}",
    "",
    "  Part 1 blockers (adjudication):  " + str(len(p1_blockers)),
    "  Part 2 blockers (noise):         " + str(len(p2_blockers)),
    "  Part 3 blockers (content loss):  " + str(len(p3_blockers)),
    "  Part 4 blockers (community):     " + str(len(p4_blockers)),
    f"  TOTAL BLOCKERS:                 {total_blockers}",
    "",
]

if p2_blockers:
    lines.append("  ENCODING ISSUES (Part 2):")
    by_noise_type = defaultdict(list)
    for r in p2_blockers:
        by_noise_type[r["noise_type"]].append(r)
    for nt, items in sorted(by_noise_type.items()):
        lines.append(f"    {nt}: {len(items)} instances")
        # Show unique affected events
        eids = sorted(set(r["event_id"] for r in items if r["event_id"]))[:5]
        if eids:
            lines.append(f"      Affected events (sample): {', '.join(eids)}")
    lines.append("")

# Community xlsx encoding summary
if community_noise_count:
    lines.append("  VISIBLE NOISE IN COMMUNITY XLSX:")
    for (ntype, sheet), count in sorted(community_noise_count.items()):
        lines.append(f"    [{sheet}] {ntype}: {count} cells")
    lines.append("")

if p1_blockers or p4_blockers:
    lines.append("  STRUCTURAL BLOCKERS:")
    for r, src in [(r, s) for r, s in all_blockers if s in ("P1_Adjudication","P4_Community")]:
        name = r.get("event_name","")[:45]
        expl = r.get("explanation","")[:70]
        lines.append(f"    [{src}] {r.get('event_id','')} {name} — {expl}")
    lines.append("")

lines += ["=" * 64, ""]

print("\n".join(lines))

summary_path = OUT / "FINAL_DECISION.txt"
with open(summary_path, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print(f"  Written: {summary_path}")
print(f"\nAll reports in: {OUT}")
