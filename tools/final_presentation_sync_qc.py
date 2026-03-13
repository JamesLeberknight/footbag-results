#!/usr/bin/env python3
"""
tools/final_presentation_sync_qc.py

Final presentation-layer synchronization QC.

Verifies that Footbag_Results_Community.xlsx (Index sheet + year sheets) is
internally consistent and agrees with canonical data sources.  Read-only.

Outputs:
    out/final_validation/presentation_sync_report.csv
    out/final_validation/presentation_sync_summary.md
"""

import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

csv.field_size_limit(10_000_000)

REPO      = Path(__file__).resolve().parent.parent
OUT_DIR   = REPO / "out"
INPUT_DIR = REPO / "inputs"
XLSX      = REPO / "Footbag_Results_Community.xlsx"

VAL_DIR   = OUT_DIR / "final_validation"
VAL_DIR.mkdir(parents=True, exist_ok=True)

# ── Allowed discrepancy types ─────────────────────────────────────────────────
DISC_NONE          = "NONE"
DISC_STATUS        = "STATUS_MISMATCH"
DISC_COVERAGE      = "COVERAGE_MISMATCH"
DISC_COUNT         = "COUNT_MISMATCH"
DISC_NOTE_MISSING  = "NOTE_MISSING"
DISC_QUAR_LABEL    = "QUARANTINE_LABEL_MISSING"
DISC_QUAR_COLOR    = "QUARANTINE_COLOR_MISMATCH"
DISC_OTHER         = "OTHER"

# ── Status label prefixes written to row 8 by 04B ────────────────────────────
_STATUS_PREFIX = {
    "SOURCE_PARTIAL": "⚠ PARTIAL RESULTS",
    "KNOWN_ISSUE":    "⚠ DATA ISSUE",
    "METADATA_ONLY":  "ℹ METADATA ONLY",
    "QUARANTINED":    "⛔ QUARANTINED",
}

# ── Row indices in year sheets ────────────────────────────────────────────────
_R_EID    = 7   # event_id
_R_STATUS = 8   # status label (or blank for OK)

# ── Index fill colors ─────────────────────────────────────────────────────────
_FILL_OK          = "E8F5E9"
_FILL_QUARANTINED = "FFCDD2"
_FILL_DEFAULT     = "FFFFFF"

# Coverage flag rank (higher = worse)
_FLAG_RANK = {"complete": 0, "mostly_complete": 1, "partial": 2, "sparse": 3}


# ── Loaders ───────────────────────────────────────────────────────────────────

def load_stage2_events() -> dict:
    path = OUT_DIR / "stage2_canonical_events.csv"
    events = {}
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.DictReader(fh):
            eid = row["event_id"].strip()
            try:
                placements = json.loads(row.get("placements_json") or "[]")
            except (json.JSONDecodeError, TypeError):
                placements = []
            div_order = []
            seen = set()
            for p in placements:
                dc = (p.get("division_canon") or "").strip()
                if dc and dc not in seen:
                    div_order.append(dc)
                    seen.add(dc)
            events[eid] = {
                "event_id":   eid,
                "year":       row.get("year"),
                "event_name": (row.get("event_name") or "").strip(),
                "div_order":  div_order,
            }
    return events


def load_placements_flat() -> dict:
    """Return dict event_id → {division_canon → placements count}."""
    path = OUT_DIR / "Placements_Flat.csv"
    result: dict = defaultdict(lambda: defaultdict(int))
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.DictReader(fh):
            eid = str(row.get("event_id", "")).strip()
            div = (row.get("division_canon") or "").strip()
            if eid and div:
                result[eid][div] += 1
    return dict(result)


def load_coverage() -> dict:
    """Return dict event_id → worst coverage_flag string.

    Reads coverage_flag from Placements_Flat.csv (same source as 04B's
    compute_event_coverage), NOT from Coverage_ByEventDivision.csv,
    to guarantee identical values.
    """
    path = OUT_DIR / "Placements_Flat.csv"
    per_event: dict = defaultdict(list)
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.DictReader(fh):
            eid  = str(row.get("event_id", "")).strip()
            flag = (row.get("coverage_flag") or "complete").lower()
            if eid:
                per_event[eid].append(flag)
    result = {}
    for eid, flags in per_event.items():
        worst = max(flags, key=lambda f: _FLAG_RANK.get(f, 0))
        result[str(eid)] = worst
    return result


def load_known_issues() -> dict:
    path = REPO / "overrides" / "known_issues.csv"
    result = {}
    if not path.exists():
        return result
    with open(path, newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            result[row["event_id"]] = {
                "severity": row.get("severity", "minor"),
                "note":     row.get("note", ""),
            }
    return result


def load_quarantine_set() -> set:
    path = INPUT_DIR / "review_quarantine_events.csv"
    result = set()
    if not path.exists():
        return result
    with open(path, newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            eid = (row.get("event_id") or "").strip()
            if eid:
                result.add(eid)
    return result


# ── Derived data_status ───────────────────────────────────────────────────────

def derive_data_status(eid: str, has_placements: bool,
                       known_issues: dict, quarantine_set: set) -> str:
    if eid in quarantine_set:
        return "QUARANTINED"
    if not has_placements:
        return "METADATA_ONLY"
    issue = known_issues.get(eid)
    if issue:
        if issue["severity"] in ("moderate", "severe"):
            return "SOURCE_PARTIAL"
        return "KNOWN_ISSUE"
    return "OK"


def derive_results_coverage(eid: str, data_status: str,
                             coverage_map: dict, has_placements: bool) -> str:
    if data_status == "QUARANTINED":
        return "quarantined"
    if not has_placements:
        return "none"
    return coverage_map.get(eid, "complete").lower()


# ── Build rendered event content from PBP (mirrors 04B logic) ─────────────────

def build_rendered_counts(s2_events: dict, pf_map: dict,
                          known_issues: dict, quarantine_set: set) -> dict:
    """
    For each event, compute how many placements and divisions would actually
    be rendered in the community xlsx (matching 04B filters: exclude
    __NON_PERSON__, unresolved; keep team entries where team_display_name present).

    Returns dict event_id → {placements_count, division_count}.
    We use the same source as 04B: Placements_Flat (pf_map) as a proxy.
    For rendered counts we count PF rows that are identity-resolved
    (excluding __NON_PERSON__ and unresolved), then count distinct divisions.
    """
    path = OUT_DIR / "Placements_Flat.csv"
    result: dict = {}
    eid_divs: dict = defaultdict(set)
    eid_plc: dict  = defaultdict(set)  # (eid, div, place) tuples for dedup

    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.DictReader(fh):
            eid    = str(row.get("event_id", "")).strip()
            person = (row.get("person_canon") or "").strip()
            unres  = (row.get("person_unresolved") or "").lower()
            comp   = (row.get("competitor_type") or "player").lower()
            tpk    = (row.get("team_person_key") or "").strip()
            tdn    = (row.get("team_display_name") or "").strip()
            div    = (row.get("division_canon") or "").strip()
            place  = row.get("place", "")

            if not eid:
                continue

            # Apply same filters as build_event_placements in 04B
            if not person or person == "__NON_PERSON__":
                # Keep team entries with full display name (no "/ ?")
                if not (comp == "team" and tpk and tdn
                        and not tdn.rstrip().endswith("/ ?")):
                    continue
            if unres in ("true", "1"):
                continue

            # Count distinct (div, place, tpk-or-person) combos
            key = (eid, div, place, tpk if comp == "team" else person)
            if key not in eid_plc[eid]:
                eid_plc[eid].add(key)
            eid_divs[eid].add(div)

    for eid in s2_events:
        result[eid] = {
            "placements_count": len(eid_plc.get(eid, set())),
            "division_count":   len(eid_divs.get(eid, set())),
        }
    return result


# ── Read Index sheet from xlsx ────────────────────────────────────────────────

def read_index_sheet(wb) -> dict:
    """
    Returns dict event_id → {year, event_name, placements_count, division_count,
                               results_coverage, data_status, notes,
                               fill_color (of col 1), row_idx}.
    Expects column order: event_id(1), year(2), event_name(3), city(4), country(5),
                          start_date(6), placements(7), divisions(8),
                          results_coverage(9), data_status(10), notes(11).
    """
    ws = wb["Index"]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        eid_cell = row[0]   # col A = event_id
        eid = str(eid_cell.value or "").strip()
        if not eid:
            continue

        def _v(col_idx):
            cell = row[col_idx]
            return str(cell.value or "").strip()

        # Extract fill color from col A
        fill = eid_cell.fill
        fill_rgb = ""
        try:
            fill_rgb = fill.fgColor.rgb[-6:] if fill and fill.fgColor else ""
        except Exception:
            fill_rgb = ""

        result[eid] = {
            "year":              _v(1),
            "event_name":        _v(2),
            "city":              _v(3),
            "country":           _v(4),
            "start_date":        _v(5),
            "placements_count":  _v(6),
            "division_count":    _v(7),
            "results_coverage":  _v(8).lower(),
            "data_status":       _v(9),
            "notes":             _v(10),
            "fill_color":        fill_rgb.upper(),
        }
    return result


# ── Read year sheets from xlsx ────────────────────────────────────────────────

def read_year_sheets(wb) -> dict:
    """
    Returns dict event_id → status_label_text (or "" if OK / blank).
    Reads row _R_EID for event_id, row _R_STATUS for status label.
    """
    result = {}
    year_pattern = re.compile(r"^\d{4}$")
    for sheet_name in wb.sheetnames:
        if not year_pattern.match(sheet_name):
            continue
        ws = wb[sheet_name]
        # Determine max column
        max_col = ws.max_column
        for col_idx in range(2, max_col + 1):  # skip col A (row labels)
            eid_cell   = ws.cell(row=_R_EID,    column=col_idx)
            label_cell = ws.cell(row=_R_STATUS, column=col_idx)
            eid   = str(eid_cell.value or "").strip()
            label = str(label_cell.value or "").strip()
            if eid:
                result[eid] = label
    return result


# ── Main QC logic ─────────────────────────────────────────────────────────────

def main():
    print("Loading canonical data…")
    s2_events     = load_stage2_events()
    pf_map        = load_placements_flat()
    coverage_map  = load_coverage()
    known_issues  = load_known_issues()
    quarantine_set = load_quarantine_set()

    print("Building rendered counts (participant-level from PF)…")
    rendered = build_rendered_counts(s2_events, pf_map, known_issues, quarantine_set)

    print(f"Reading {XLSX}…")
    wb = openpyxl.load_workbook(str(XLSX), data_only=True, read_only=False)

    index_data  = read_index_sheet(wb)
    ys_data     = read_year_sheets(wb)

    print("Running sync checks…")

    report_rows = []
    disc_counter: Counter = Counter()

    for eid, ev in sorted(s2_events.items(), key=lambda t: (t[1]["year"] or 0, t[0])):
        year       = ev.get("year") or "?"
        event_name = ev.get("event_name", "")

        # ── Derive canonical values ───────────────────────────────────────────
        rend        = rendered.get(eid, {"placements_count": 0, "division_count": 0})
        rend_plc    = rend["placements_count"]
        rend_div    = rend["division_count"]
        has_plc     = rend_plc > 0

        canon_status = derive_data_status(eid, has_plc, known_issues, quarantine_set)
        canon_cov    = derive_results_coverage(eid, canon_status, coverage_map, has_plc)

        issue     = known_issues.get(eid)
        if canon_status == "QUARANTINED":
            canon_note = "Excluded — ambiguous structure prevents deterministic parsing"
        elif canon_status == "METADATA_ONLY":
            canon_note = "No competitive results available"
        elif issue:
            canon_note = issue["note"]
        else:
            canon_note = ""

        # ── Index sheet values ────────────────────────────────────────────────
        idx = index_data.get(eid)
        if not idx:
            # Event missing from Index entirely
            report_rows.append({
                "event_id":              eid,
                "year":                  year,
                "event_name":            event_name,
                "index_status":          "MISSING",
                "year_sheet_status":     ys_data.get(eid, ""),
                "index_results_coverage": "",
                "derived_results_coverage": canon_cov,
                "index_notes_present":   "NO",
                "placements_count_index":   "",
                "placements_count_rendered": rend_plc,
                "division_count_index":     "",
                "division_count_rendered":  rend_div,
                "sync_ok":               "NO",
                "discrepancy_type":      DISC_OTHER,
                "recommended_fix":       "Event missing from Index sheet",
            })
            disc_counter[DISC_OTHER] += 1
            continue

        idx_status   = idx["data_status"]
        idx_cov      = idx["results_coverage"]
        idx_notes    = idx["notes"]
        idx_plc      = idx["placements_count"]
        idx_div      = idx["division_count"]
        idx_fill     = idx["fill_color"]

        # Parse numeric counts from Index (may be empty string for 0)
        idx_plc_int  = int(idx_plc) if idx_plc.isdigit() else 0
        idx_div_int  = int(idx_div) if idx_div.isdigit() else 0

        # ── Year sheet label ──────────────────────────────────────────────────
        ys_label = ys_data.get(eid, "")  # "" means event not in any year sheet

        # Expected year-sheet label prefix
        expected_label_prefix = _STATUS_PREFIX.get(canon_status, "")
        if expected_label_prefix:
            ys_has_label = ys_label.startswith(expected_label_prefix)
        else:
            ys_has_label = (ys_label == "")  # OK → no label expected

        # ── Expected fill color ───────────────────────────────────────────────
        if canon_status == "OK":
            expected_fill = _FILL_OK
        elif canon_status == "QUARANTINED":
            expected_fill = _FILL_QUARANTINED
        else:
            expected_fill = _FILL_DEFAULT

        # Normalize fill for comparison (handle "00" prefix from openpyxl ARGB)
        idx_fill_cmp = idx_fill[-6:] if len(idx_fill) >= 6 else idx_fill

        # ── Collect discrepancies ─────────────────────────────────────────────
        discrepancies = []
        fixes         = []

        # 1. Status mismatch
        if idx_status != canon_status:
            discrepancies.append(DISC_STATUS)
            fixes.append(f"data_status: Index={idx_status} Expected={canon_status}")

        # 2. Coverage mismatch
        if idx_cov != canon_cov:
            discrepancies.append(DISC_COVERAGE)
            fixes.append(f"results_coverage: Index={idx_cov} Expected={canon_cov}")

        # 3. Count sanity (not independent recount — just logical consistency)
        # METADATA_ONLY must have 0 placements in Index; others must have > 0 when has_plc
        if canon_status == "METADATA_ONLY" and idx_plc_int > 0:
            discrepancies.append(DISC_COUNT)
            fixes.append(f"placements: METADATA_ONLY event has Index count={idx_plc_int} (expected 0)")
        elif canon_status != "METADATA_ONLY" and has_plc and idx_plc_int == 0:
            discrepancies.append(DISC_COUNT)
            fixes.append(f"placements: non-METADATA event has Index count=0 but rendered={rend_plc}")

        # 4. Missing note for non-OK events
        if canon_status != "OK" and not idx_notes:
            discrepancies.append(DISC_NOTE_MISSING)
            fixes.append(f"notes: missing for {canon_status}")

        # 5. Quarantine label in year sheet
        if canon_status in _STATUS_PREFIX and not ys_has_label:
            # Only flag if event appears in any year sheet (not metadata-only)
            if eid in ys_data or has_plc:
                if canon_status == "QUARANTINED":
                    discrepancies.append(DISC_QUAR_LABEL)
                    fixes.append(f"year-sheet row 8: expected '{expected_label_prefix}' got '{ys_label}'")
                elif ys_label != expected_label_prefix and ys_label and not ys_label.startswith(expected_label_prefix):
                    discrepancies.append(DISC_OTHER)
                    fixes.append(f"year-sheet label mismatch: expected '{expected_label_prefix}' got '{ys_label}'")

        # 6. Fill color
        if idx_fill_cmp and idx_fill_cmp.upper() != expected_fill.upper():
            # Only flag for OK and QUARANTINED (the two explicit fill rules)
            if canon_status in ("OK", "QUARANTINED"):
                discrepancies.append(DISC_QUAR_COLOR)
                fixes.append(f"fill: Index={idx_fill_cmp} Expected={expected_fill}")

        # Determine primary discrepancy type
        if not discrepancies:
            disc_type = DISC_NONE
            sync_ok   = "YES"
        else:
            # Priority order
            priority = [DISC_STATUS, DISC_QUAR_LABEL, DISC_QUAR_COLOR,
                        DISC_COVERAGE, DISC_COUNT, DISC_NOTE_MISSING, DISC_OTHER]
            disc_type = next((d for d in priority if d in discrepancies), discrepancies[0])
            sync_ok   = "NO"
            disc_counter[disc_type] += 1

        report_rows.append({
            "event_id":               eid,
            "year":                   year,
            "event_name":             event_name,
            "index_status":           idx_status,
            "year_sheet_status":      ys_label[:60] if ys_label else "OK (no label)",
            "index_results_coverage": idx_cov,
            "derived_results_coverage": canon_cov,
            "index_notes_present":    "YES" if idx_notes else "NO",
            "placements_count_index":    idx_plc_int,
            "placements_count_rendered": rend_plc,
            "division_count_index":      idx_div_int,
            "division_count_rendered":   rend_div,
            "sync_ok":                sync_ok,
            "discrepancy_type":       disc_type,
            "recommended_fix":        "; ".join(fixes) if fixes else "",
        })

    # ── Write CSV report ──────────────────────────────────────────────────────
    report_path = VAL_DIR / "presentation_sync_report.csv"
    fieldnames = [
        "event_id", "year", "event_name",
        "index_status", "year_sheet_status",
        "index_results_coverage", "derived_results_coverage",
        "index_notes_present",
        "placements_count_index", "placements_count_rendered",
        "division_count_index",   "division_count_rendered",
        "sync_ok", "discrepancy_type", "recommended_fix",
    ]
    with open(report_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(report_rows)
    print(f"  Report: {report_path}")

    # ── Compute summary stats ─────────────────────────────────────────────────
    total        = len(report_rows)
    total_ok     = sum(1 for r in report_rows if r["sync_ok"] == "YES")
    total_fail   = total - total_ok

    status_counts: Counter = Counter(r["index_status"] for r in report_rows)
    cov_counts: Counter    = Counter(r["index_results_coverage"] for r in report_rows)

    fail_rows = [r for r in report_rows if r["sync_ok"] == "NO"]
    blockers  = [r for r in fail_rows if r["discrepancy_type"] in
                 (DISC_STATUS, DISC_QUAR_LABEL, DISC_QUAR_COLOR)]

    # ── Write summary md ──────────────────────────────────────────────────────
    summary_path = VAL_DIR / "presentation_sync_summary.md"
    verdict      = "PRESENTATION_SYNC_PASS" if total_fail == 0 else "PRESENTATION_SYNC_FAIL"

    with open(summary_path, "w", encoding="utf-8") as sf:
        sf.write("# Presentation Sync QC Summary\n\n")
        sf.write(f"Total events checked: {total}  \n")
        sf.write(f"SYNC OK: {total_ok}  \n")
        sf.write(f"SYNC FAIL: {total_fail}  \n\n")

        sf.write("## Counts by data_status (Index)\n\n")
        for s in ("OK", "KNOWN_ISSUE", "SOURCE_PARTIAL", "METADATA_ONLY", "QUARANTINED", "MISSING"):
            n = status_counts.get(s, 0)
            if n:
                sf.write(f"- {s}: {n}\n")

        sf.write("\n## Counts by results_coverage (Index)\n\n")
        for c in ("complete", "mostly_complete", "partial", "sparse", "none", "quarantined", ""):
            n = cov_counts.get(c, 0)
            if n:
                label = c if c else "(blank)"
                sf.write(f"- {label}: {n}\n")

        sf.write("\n## Discrepancy breakdown\n\n")
        if disc_counter:
            for dtype, cnt in sorted(disc_counter.items(), key=lambda x: -x[1]):
                sf.write(f"- {dtype}: {cnt}\n")
        else:
            sf.write("_(none)_\n")

        if fail_rows:
            sf.write("\n## All discrepancies\n\n")
            sf.write("| event_id | year | event_name | discrepancy_type | recommended_fix |\n")
            sf.write("|---|---|---|---|---|\n")
            for r in fail_rows:
                name = (r["event_name"] or "")[:50]
                sf.write(f"| {r['event_id']} | {r['year']} | {name} | "
                         f"{r['discrepancy_type']} | {r['recommended_fix']} |\n")

        sf.write(f"\n---\n\n## Result\n\n**{verdict}**\n")

    print(f"  Summary: {summary_path}")
    print(f"\n{verdict}")
    print(f"  {total_ok}/{total} events in sync"
          + (f"  |  {total_fail} discrepancies" if total_fail else ""))
    if disc_counter:
        for dtype, cnt in sorted(disc_counter.items(), key=lambda x: -x[1]):
            print(f"    {dtype}: {cnt}")

    return 0 if total_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
