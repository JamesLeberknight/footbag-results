#!/usr/bin/env python3
"""
build_merged_workbook_v14.py

Merged canonical edition of the Community v13 workbook.

v14 changes from v13 (minimal):
  - OUTPUT_PATH → Footbag_Results_Merged_FINAL.xlsx
  - PF_CSV      → out/merged_placements_flat.csv   (pre-1997 + post-1997, overlap-cleaned)
  - EVENTS_CSV  → out/merged_events_normalized.csv (pre-1997 + post-1997, overlap-cleaned)
  - README text updated to reflect merged dataset scope
  - DATA NOTES updated to document pre-1997 reconstruction and overlap filtering

All v13 logic (STATISTICS, PLAYER SUMMARY, FREESTYLE INSIGHTS, CONSECUTIVE RECORDS,
EVENT INDEX hyperlinking, year sheets) is preserved unchanged.
"""

import csv
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

csv.field_size_limit(sys.maxsize)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── Paths ─────────────────────────────────────────────────────────────────────

OUTPUT_PATH      = os.path.join(BASE_DIR, "Footbag_Results_Merged_FINAL.xlsx")

# v14: merged feeds (pre-1997 + post-1997, overlap-cleaned)
PF_CSV           = os.path.join(BASE_DIR, "out", "merged_placements_flat.csv")
PT_CSV           = os.path.join(BASE_DIR, "out", "Persons_Truth.csv")
EVENTS_CSV       = os.path.join(BASE_DIR, "out", "merged_events_normalized.csv")
QUARANTINE_CSV   = os.path.join(BASE_DIR, "inputs", "review_quarantine_events.csv")

BAP_CSV          = os.path.join(BASE_DIR, "inputs", "bap_data_updated.csv")
FBHOF_CSV        = os.path.join(BASE_DIR, "inputs", "hof.csv")
MEMBER_IDS_CSV   = os.path.join(BASE_DIR, "out", "member_id_enrichment", "member_id_assignments.csv")

DIFFICULTY_CSV   = os.path.join(BASE_DIR, "out", "noise_aggregates", "player_difficulty_profiles.csv")
DIVERSITY_CSV    = os.path.join(BASE_DIR, "out", "noise_aggregates", "player_diversity_profiles.csv")
TRICK_FREQ_CSV   = os.path.join(BASE_DIR, "out", "noise_aggregates", "trick_frequency.csv")
TRANSITIONS_CSV  = os.path.join(BASE_DIR, "out", "noise_aggregates", "trick_transition_network.csv")
DIFF_YEAR_CSV    = os.path.join(BASE_DIR, "out", "noise_aggregates", "difficulty_by_year.csv")
INNOVATION_CSV   = os.path.join(BASE_DIR, "out", "noise_aggregates", "trick_innovation_timeline.csv")
TRICK_NODE_CSV        = os.path.join(BASE_DIR, "out", "noise_aggregates", "trick_node_metrics.csv")
SEQ_DIFF_CSV          = os.path.join(BASE_DIR, "out", "noise_aggregates", "sequence_difficulty_conservative.csv")
CHAIN_COMPLEXITY_CSV  = os.path.join(BASE_DIR, "out", "noise_aggregates", "chain_complexity_by_year.csv")


# ── Styles ────────────────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

FILL_NONE    = PatternFill(fill_type=None)
FILL_HEADER  = _fill("D9D9D9")
FILL_SECTION = _fill("E8F0FE")   # light blue for section titles
FILL_HOF     = _fill("FFF2CC")   # pale gold for FBHOF rows
FILL_BAP     = _fill("E2EFDA")   # pale green for BAP rows
FILL_BOTH    = _fill("EAD1F5")   # light purple for both

FONT_TITLE   = Font(bold=True, size=13)
FONT_SECTION = Font(bold=True, size=12)
FONT_HEADER  = Font(bold=True, size=11)
FONT_DATA    = Font(size=11)

ALIGN_LEFT   = Alignment(horizontal="left",  vertical="top", wrap_text=False)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="top")


# ── Name normalisation helpers ────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Lower-case, strip accents, collapse spaces."""
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode()
    return " ".join(s.lower().split())


# Manual overrides: honor-CSV name  →  person_canon (exact string)
# Extend this dict whenever a new mismatch is found.
_HONOR_OVERRIDES: dict[str, str] = {
    "ken shults":               "Kenny Shults",
    "kenny shults":             "Kenny Shults",
    "kenneth shults":           "Kenny Shults",
    "vasek klouda":             "Václav Klouda",
    "vaclav (vasek) klouda":    "Václav Klouda",
    "tina aberli":              "Tina Aeberli",
    "eli piltz":                "Eliot Piltz Galán",
    "eliot piltz galan":        "Eliot Piltz Galán",
    "eliott piltz galan":       "Eliot Piltz Galán",
    "eliot galan":              "Eliot Piltz Galán",
    "eliott galan":             "Eliot Piltz Galán",
    "evanne lamarch":           "Evanne LaMarche",
    "evanne lamarche":          "Evanne LaMarche",
    "evanne lemarche":          "Evanne LaMarche",
    "arek dzudzinski":          "Arkadiusz Dudzinski",
    "martin cote":              "Martin Côté",
    "sebastien duchesne":       "Sébastien Duchesne",
    "sebastien duschesne":      "Sébastien Duchesne",
    "jonathan schneider":       "Jonathan Schneider",
    "lon smith":                "Lon Skyler Smith",
    "lon skyler smith":         "Lon Skyler Smith",
    "aleksi airinen":           "Aleksi Airinen",
    "lauri airinen":            "Lauri Airinen",
    "ales zelinka":             "Aleš Zelinka",
    "jere vainikka":            "Jere Väinikkä",
    "jukka peltola":            "Jukka Peltola",
    "tuomas karki":             "Tuomas Kärki",
    "tuukka antikainen":        "Tuukka Antikainen",
    "rafal kaleta":             "Rafał Kaleta",
    "pawel nowak":              "Paweł Nowak",
    "jakub mosciszewski":       "Jakub Mościszewski",
    "dominik simku":            "Dominik Šimků",
    "honza weber":              "Jan Weber",
    "carol wedemeyer":          "Carol Wedemeyer",
    "scott-mag hughes":         "Scott-Mag Hughes",
    "cheryl aubin hughes":      "Cheryl Aubin Hughes",
    "heather squires thomas":   "Heather Squires Thomas",
    "lisa mcdaniel jones":      "Lisa McDaniel Jones",
    "lori jean conover":        "Lori Jean Conover",
    "jody badger welch":        "Jody Badger Welch",
    "genevieve bousquet":       "Geneviève Bousquet",
    "becca english":            "Becca English-Ross",
    "becca english-ross":       "Becca English-Ross",
    "pt lovern":                "P.T. Lovern",
    "p.t. lovern":              "P.T. Lovern",
    "kendall kic":              "Kendall KIC",
    "taishi ishida":            "Taishi Ishida",
    "wiktor debski":            "Wiktor Dębski",
    "wiktor d\u0119bski":       "Wiktor Dębski",
    "florian gotze":            "Florian Götze",
    "grischa tellenbach":       "Grischa Tellenbach",
    "chantelle laurent":        "Chantelle Laurent",
    # BAP name variants → PT canonical
    "gordon scott bevier":      "Gordon Bevier",
    "dave holton":              "David Holton",
    "bryan fournier":           "Brian Fournier",
    "olav piwowar":             "Olaf Piwowar",
    "jindra smola":             "Jindrich Smola",
    "rene ruhr":                "Rene Ruehr",
    "nick polini":              "Nick Pollini",
    "rafa kaleta":              "Rafal Kaleta",   # ł strips in norm()
    "rafal kaleta":             "Rafal Kaleta",
    "phillip morrison":         "Philip Morrison",
    "johnny murphy":            "Jonathan Murphy",
    "johnathon murphy":         "Jonathan Murphy",
}


def match_honor_name(raw_name: str, canon_by_norm: dict[str, str]) -> str | None:
    """
    Try to resolve an honor-CSV name to a person_canon string.
    1. Check _HONOR_OVERRIDES (normalised key)
    2. Exact normalised lookup in Persons_Truth
    Returns person_canon or None.
    """
    key = _norm(raw_name)
    if key in _HONOR_OVERRIDES:
        return _HONOR_OVERRIDES[key]
    if key in canon_by_norm:
        return canon_by_norm[key]
    return None


# ── Data loading helpers ───────────────────────────────────────────────────────

def load_csv(path: str) -> list[dict]:
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_persons_truth() -> tuple[list[dict], dict[str, str]]:
    """Returns (rows, norm→person_canon mapping)."""
    rows = load_csv(PT_CSV)
    norm_map: dict[str, str] = {}
    for r in rows:
        pc = r.get("person_canon", "").strip()
        if pc:
            norm_map[_norm(pc)] = pc
            # Also index the norm_key field if present
            nk = r.get("norm_key", "").strip()
            if nk:
                norm_map[nk] = pc
    return rows, norm_map


def is_real_person(r: dict) -> bool:
    pid = r.get("effective_person_id", "") or r.get("person_id", "")
    pc  = r.get("person_canon", "")
    if pid == "__NON_PERSON__" or pc == "__NON_PERSON__":
        return False
    excl = r.get("exclusion_reason", "") or ""
    if "non_person" in excl.lower():
        return False
    return bool(pc)


# ── Placement aggregations ────────────────────────────────────────────────────

def build_placement_stats(
    pf_rows: list[dict],
) -> dict[str, dict]:
    """
    Returns {person_id: {events, wins, podiums, placements, year_first, year_last}}
    Counts are across ALL division categories.
    """
    stats: dict[str, dict] = defaultdict(lambda: {
        "events": set(), "wins": 0, "podiums": 0,
        "placements": 0, "years": set(),
    })

    for row in pf_rows:
        pid = row.get("person_id", "")
        if not pid or pid == "__NON_PERSON__":
            continue
        eid  = row.get("event_id", "")
        year = row.get("year", "")
        try:
            place = int(row.get("place", 0) or 0)
        except ValueError:
            place = 0

        s = stats[pid]
        if eid:
            s["events"].add(eid)
        if year:
            try:
                s["years"].add(int(year))
            except ValueError:
                pass
        s["placements"] += 1
        if place == 1:
            s["wins"] += 1
        if 1 <= place <= 3:
            s["podiums"] += 1

    return {
        pid: {
            "events":     len(d["events"]),
            "wins":       d["wins"],
            "podiums":    d["podiums"],
            "placements": d["placements"],
            "year_first": min(d["years"]) if d["years"] else None,
            "year_last":  max(d["years"]) if d["years"] else None,
        }
        for pid, d in stats.items()
    }


def years_active_str(s: dict) -> str:
    yf = s.get("year_first")
    yl = s.get("year_last")
    if yf and yl:
        return str(yf) if yf == yl else f"{yf}\u2013{yl}"
    return ""


# ── Member ID loading ─────────────────────────────────────────────────────────

def load_member_ids() -> dict[str, str]:
    """
    Returns {effective_person_id: member_id} from member_id_assignments.csv.
    Falls back to empty dict if the file doesn't exist.
    """
    if not os.path.exists(MEMBER_IDS_CSV):
        print(f"  [WARN] member_id_assignments.csv not found, legacy IDs will be PT-only")
        return {}
    result = {}
    with open(MEMBER_IDS_CSV, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            pid = r.get("effective_person_id", "").strip()
            mid = r.get("member_id", "").strip()
            if pid and mid:
                result[pid] = mid
    print(f"  Member ID assignments: {len(result)} entries loaded")
    return result


# ── Honors loading ────────────────────────────────────────────────────────────

def load_bap(canon_by_norm: dict[str, str]) -> dict[str, dict]:
    """
    Returns {person_canon: {bap_number, year_inducted, nickname}}
    BAP members are numbered 1..N in chronological order by year_inducted.
    """
    rows = load_csv(BAP_CSV)
    result: dict[str, dict] = {}
    for i, row in enumerate(rows, 1):
        raw  = row.get("name", "").strip()
        year = row.get("year_inducted", "").strip()
        nick = row.get("nickname", "").strip()
        pc   = match_honor_name(raw, canon_by_norm)
        if pc:
            result[pc] = {
                "bap_number":    i,
                "year_inducted": year,
                "nickname":      nick,
                "raw_name":      raw,
            }
        else:
            print(f"  [WARN] BAP name unmatched: {raw!r}")
    return result


def load_fbhof(canon_by_norm: dict[str, str],
               pid_to_canon: dict[str, str] | None = None) -> dict[str, dict]:
    """
    Returns {person_canon: {year_inducted}}
    Reads from inputs/hof.csv (full_name, induction_year, person_id columns).
    Uses explicit person_id when available; falls back to name matching.
    """
    rows = load_csv(FBHOF_CSV)
    result: dict[str, dict] = {}
    pid_map = pid_to_canon or {}
    for row in rows:
        raw  = row.get("full_name", "").strip()
        year = row.get("induction_year", "").strip()
        pid  = row.get("person_id", "").strip()
        # Prefer direct person_id → canon lookup
        if pid and pid in pid_map:
            pc = pid_map[pid]
        else:
            pc = match_honor_name(raw, canon_by_norm)
        if pc:
            result[pc] = {"year_inducted": year, "raw_name": raw}
        else:
            print(f"  [WARN] FBHOF name unmatched: {raw!r}")
    return result


def bap_label(info: dict) -> str:
    n    = info["bap_number"]
    year = info["year_inducted"]
    return f"BAP #{n} ({year})" if year else f"BAP #{n}"


def fbhof_label(info: dict) -> str:
    year = info.get("year_inducted", "")
    return f"FBHOF {year}" if year and year != "unknown" else "FBHOF"


# ── Freestyle analytics ───────────────────────────────────────────────────────

def load_difficulty_profiles() -> dict[str, dict]:
    """Returns {person_id: row_dict}"""
    rows = load_csv(DIFFICULTY_CSV)
    return {r["person_id"]: r for r in rows if r.get("person_id")}


def load_diversity_profiles() -> dict[str, str]:
    """Returns {person_id: most_common_trick (first of top_tricks pipe-list)}"""
    rows = load_csv(DIVERSITY_CSV)
    result: dict[str, str] = {}
    for r in rows:
        pid  = r.get("person_id", "")
        tops = r.get("top_tricks", "")
        if pid and tops:
            result[pid] = tops.split(" | ")[0].strip()
    return result


# ── Sheet helpers ─────────────────────────────────────────────────────────────

def _w(ws, row: int, col: int, value=None, *,
       font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font  is not None: cell.font      = font
    if fill  is not None: cell.fill      = fill
    if align is not None: cell.alignment = align
    return cell


def _hrow(ws, row: int, *headers) -> int:
    for col, h in enumerate(headers, 1):
        _w(ws, row, col, h, font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT)
    return row + 1


def _drow(ws, row: int, *values) -> int:
    for col, v in enumerate(values, 1):
        align = ALIGN_RIGHT if isinstance(v, (int, float)) else ALIGN_LEFT
        _w(ws, row, col, v, font=FONT_DATA, align=align)
    return row + 1


def _section(ws, row: int, text: str) -> int:
    _w(ws, row, 1, text, font=FONT_SECTION, fill=FILL_SECTION, align=ALIGN_LEFT)
    return row + 1


# ── README sheet ──────────────────────────────────────────────────────────────

def build_readme(wb: Workbook, quarantine_count: int = 0,
                 event_count: int = 0, person_count: int = 0,
                 placement_count: int = 0) -> None:
    if "README" in wb.sheetnames:
        del wb["README"]
    idx = 0
    ws = wb.create_sheet("README", idx)

    row = 1
    _w(ws, row, 1, "Footbag Historical Results — Official Merged Canonical Workbook",
       font=FONT_TITLE, align=ALIGN_LEFT)
    row += 2

    _w(ws, row, 1, "About This Workbook", font=FONT_SECTION, align=ALIGN_LEFT)
    row += 1
    for line in [
        "This workbook contains historical footbag competition results spanning 1980 to the present.",
        "POST-1997 results are sourced from the Footbag.org mirror archive (authoritative, identity-locked).",
        "PRE-1997 results are reconstructed from Footbag World magazine, IFAB summaries, and OLD_RESULTS.txt.",
        "Where both a PRE-1997 reconstruction and a POST-1997 legacy stub exist for the same event, the",
        "  reconstruction takes precedence and the redundant stub is excluded from this workbook.",
        "Player identities are human-verified. Unresolved names are preserved as-is from the source.",
        f"{quarantine_count} events are quarantined due to parsing ambiguity and excluded from statistics.",
    ]:
        _w(ws, row, 1, line, font=FONT_DATA, align=ALIGN_LEFT)
        row += 1
    row += 1

    _w(ws, row, 1, "Sheet Guide", font=FONT_SECTION, align=ALIGN_LEFT)
    row += 1
    _w(ws, row, 1, "Sheet", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT)
    _w(ws, row, 2, "Contents", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT)
    row += 1
    for name, desc in [
        ("README",              "This sheet — dataset overview and sheet guide"),
        ("DATA NOTES",          "Source quality notes, known limitations, quarantined events"),
        ("STATISTICS",          "Career podiums, event wins, events competed, career spans, events by year"),
        ("EVENT INDEX",         "One row per event — year, name, location, discipline counts"),
        ("PLAYER SUMMARY",      "Competition history per player — wins, podiums, placements, events"),
        ("CONSECUTIVE RECORDS", "Documented consecutives world records"),
        ("FREESTYLE INSIGHTS",  "Trick-sequence analytics — difficulty by year, backbone tricks, transitions, innovation timeline"),
        ("1980 – 2026",         "One sheet per year — all placement results for that year (including unresolved entries)"),
    ]:
        _w(ws, row, 1, name, font=FONT_DATA, align=ALIGN_LEFT)
        _w(ws, row, 2, desc, font=FONT_DATA, align=ALIGN_LEFT)
        row += 1
    row += 1

    _w(ws, row, 1, "Coverage Notes", font=FONT_SECTION, align=ALIGN_LEFT)
    row += 1
    for note in [
        f"{event_count:,} events documented, 1980–present (merged: PRE-1997 reconstruction + POST-1997 mirror).",
        f"{person_count:,} canonically identified players.",
        f"{placement_count:,} placements (identity-locked where possible; PRE-1997 entries are partial-resolution).",
        "POST-1997 coverage is comprehensive from 1997 onward (Footbag.org mirror archive).",
        "PRE-1997 coverage (32 reconstructed events, 1980–1996):",
        "  • World Championships and major organization championships (NHSA, WFA, IFAB) where sources exist.",
        "  • 1980–1992: NHSA Nationals, WFA World Championships, and generic World Championships.",
        "  • 1993–1996: IFAB World Championships (4 events).",
        "  • 1980–1988: US Nationals, European Championships, and regional events where documented.",
        "  • Results are partial — top finishers only; year-level date precision only.",
        "Overlap filtering: 25 POST-1997 legacy stubs for years <1997 are suppressed where a PRE-1997",
        "  reconstruction exists for the same real-world event. Full overlap table: out/canonical_all_union/.",
        "FREESTYLE INSIGHTS draws from events that reported trick sequences; coverage is a subset of all events.",
    ]:
        _w(ws, row, 1, "•  " + note, font=FONT_DATA, align=ALIGN_LEFT)
        row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70
    print("  README sheet written")


# ── STATISTICS sheet ───────────────────────────────────────────────────────────

def build_statistics(wb: Workbook, pf_rows: list[dict], pt_rows: list[dict],
                     events_dict: dict | None = None) -> None:
    if "STATISTICS" in wb.sheetnames:
        del wb["STATISTICS"]
    # Insert after DATA NOTES
    if "DATA NOTES" in wb.sheetnames:
        idx = wb.sheetnames.index("DATA NOTES") + 1
    else:
        idx = 1
    ws = wb.create_sheet("STATISTICS", idx)

    # Build lookup tables
    valid_pids: set[str] = {
        r.get("effective_person_id", "") for r in pt_rows if is_real_person(r)
    }
    pid_to_canon: dict[str, str] = {
        r.get("effective_person_id", ""): r.get("person_canon", "")
        for r in pt_rows if r.get("effective_person_id")
    }

    # Aggregate from placements
    from collections import defaultdict
    podiums:       dict[str, dict] = defaultdict(lambda: {1: 0, 2: 0, 3: 0})
    events_by_pid: dict[str, set]  = defaultdict(set)
    years_by_pid:  dict[str, set]  = defaultdict(set)
    events_by_year: dict[int, set] = defaultdict(set)

    for row in pf_rows:
        pid = row.get("person_id", "")
        if not pid or pid not in valid_pids:
            continue
        eid  = row.get("event_id", "")
        year = row.get("year", "")
        try:
            place = int(row.get("place", 0) or 0)
        except (ValueError, TypeError):
            place = 0

        if eid:
            events_by_pid[pid].add(eid)
        if year:
            try:
                y = int(year)
                years_by_pid[pid].add(y)
                if eid:
                    events_by_year[y].add(eid)
            except (ValueError, TypeError):
                pass
        if 1 <= place <= 3:
            podiums[pid][place] += 1

    # ── Table helpers ─────────────────────────────────────────────────────────
    row_num = 1
    _w(ws, row_num, 1, "STATISTICS", font=FONT_TITLE, align=ALIGN_LEFT)
    row_num += 2

    TOP_N = 25

    # ── 1. Most Career Podiums ────────────────────────────────────────────────
    row_num = _section(ws, row_num, "MOST CAREER PODIUMS")
    row_num += 1
    row_num = _hrow(ws, row_num, "Player", "1st", "2nd", "3rd", "Total Podiums")
    podium_rows = [
        (pid_to_canon.get(pid, pid),
         d[1], d[2], d[3], d[1] + d[2] + d[3])
        for pid, d in podiums.items()
        if d[1] + d[2] + d[3] > 0
    ]
    podium_rows.sort(key=lambda x: (-x[4], x[0].lower()))
    for canon, p1, p2, p3, total in podium_rows[:TOP_N]:
        row_num = _drow(ws, row_num, canon, p1, p2, p3, total)
    row_num += 2

    # ── 2. Most Event Wins ────────────────────────────────────────────────────
    row_num = _section(ws, row_num, "MOST EVENT WINS")
    row_num += 1
    row_num = _hrow(ws, row_num, "Player", "Wins")
    wins_rows = [
        (pid_to_canon.get(pid, pid), d[1])
        for pid, d in podiums.items()
        if d[1] > 0
    ]
    wins_rows.sort(key=lambda x: (-x[1], x[0].lower()))
    for canon, wins in wins_rows[:TOP_N]:
        row_num = _drow(ws, row_num, canon, wins)
    row_num += 2

    # ── 3. Most Events Competed ───────────────────────────────────────────────
    row_num = _section(ws, row_num, "MOST EVENTS COMPETED")
    row_num += 1
    row_num = _hrow(ws, row_num, "Player", "Events Competed")
    events_rows = [
        (pid_to_canon.get(pid, pid), len(eids))
        for pid, eids in events_by_pid.items()
        if pid in valid_pids
    ]
    events_rows.sort(key=lambda x: (-x[1], x[0].lower()))
    for canon, count in events_rows[:TOP_N]:
        row_num = _drow(ws, row_num, canon, count)
    row_num += 2

    # ── 4. Longest Competitive Careers ───────────────────────────────────────
    row_num = _section(ws, row_num, "LONGEST COMPETITIVE CAREERS")
    row_num += 1
    row_num = _hrow(ws, row_num,
                    "Player", "First Event Year", "Last Event Year", "Career Span (Years)")
    career_rows = []
    for pid, years in years_by_pid.items():
        if pid not in valid_pids or not years:
            continue
        yf, yl = min(years), max(years)
        span = yl - yf
        career_rows.append((pid_to_canon.get(pid, pid), yf, yl, span))
    career_rows.sort(key=lambda x: (-x[3], x[0].lower()))
    for canon, yf, yl, span in career_rows[:TOP_N]:
        row_num = _drow(ws, row_num, canon, yf, yl, span)
    row_num += 2

    # ── 5. Events by Year ─────────────────────────────────────────────────────
    row_num = _section(ws, row_num, "EVENTS BY YEAR")
    row_num += 1
    row_num = _hrow(ws, row_num, "Year", "Events")
    for year in sorted(events_by_year):
        row_num = _drow(ws, row_num, year, len(events_by_year[year]))
    row_num += 2

    # ── 6. Dataset Overview ───────────────────────────────────────────────────
    if events_dict:
        row_num = _section(ws, row_num, "DATASET OVERVIEW")
        row_num += 1

        ev_list = list(events_dict.values())
        pre97_ev   = [e for e in ev_list if e.get("year","").isdigit() and int(e["year"]) < 1997]
        post97_ev  = [e for e in ev_list if e.get("year","").isdigit() and int(e["year"]) >= 1997]
        worlds_ev  = [e for e in ev_list if _is_worlds(e.get("event_type",""),
                                                        e.get("event_key",""),
                                                        e.get("year",""))]
        total_ev   = len(ev_list)

        row_num = _hrow(ws, row_num, "Metric", "Count")
        for label, cnt in [
            ("Total events documented",          total_ev),
            ("  PRE-1997 events (reconstructed)", len(pre97_ev)),
            ("  POST-1997 events (mirror)",       len(post97_ev)),
            ("World Championship events",         len(worlds_ev)),
            ("Total placements",                  len(pf_rows)),
            ("Canonically identified players",    sum(1 for r in pt_rows if is_real_person(r))),
        ]:
            row_num = _drow(ws, row_num, label, cnt)
        row_num += 2

    # ── 7. World Championships by Year ────────────────────────────────────────
    if events_dict:
        row_num = _section(ws, row_num, "WORLD CHAMPIONSHIPS BY YEAR")
        row_num += 1
        row_num = _hrow(ws, row_num, "Year", "Event Name", "Location")
        worlds_sorted = sorted(
            [e for e in events_dict.values()
             if _is_worlds(e.get("event_type",""), e.get("event_key",""), e.get("year",""))],
            key=lambda e: (e.get("year","9999"), e.get("event_name",""))
        )
        for e in worlds_sorted:
            row_num = _drow(ws, row_num,
                            int(e["year"]) if e.get("year","").isdigit() else e.get("year",""),
                            e.get("event_name",""),
                            e.get("location","") or "—")
        row_num += 2

    # ── 8. Events by Discipline Type ─────────────────────────────────────────
    div_cat_count: dict[str, int] = Counter()
    for row in pf_rows:
        cat = (row.get("division_category") or "").strip() or "unknown"
        div_cat_count[cat] += 1

    row_num = _section(ws, row_num, "PLACEMENTS BY DISCIPLINE CATEGORY")
    row_num += 1
    row_num = _hrow(ws, row_num, "Category", "Placements")
    for cat, cnt in sorted(div_cat_count.items(), key=lambda x: -x[1]):
        row_num = _drow(ws, row_num, cat.upper() if cat else "UNKNOWN", cnt)
    row_num += 2

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    ws.freeze_panes = "A2"
    print("  STATISTICS sheet written")


# ── PLAYER SUMMARY ────────────────────────────────────────────────────────────

PLAYER_SUMMARY_HEADERS = [
    "Player", "BAP Nickname",
    "Wins", "Podiums", "Placements", "Events",
    "Legacy ID",
]

PLAYER_COL_WIDTHS = [32, 20, 5, 7, 10, 6, 10]


def build_player_summary(wb: Workbook,
                         pt_rows: list[dict],
                         placement_stats: dict[str, dict],
                         bap_map: dict[str, dict],
                         member_id_map: dict[str, str] | None = None) -> None:
    if "PLAYER SUMMARY" in wb.sheetnames:
        idx = wb.sheetnames.index("PLAYER SUMMARY")
        del wb["PLAYER SUMMARY"]
        ws = wb.create_sheet("PLAYER SUMMARY", idx)
    else:
        ws = wb.create_sheet("PLAYER SUMMARY")

    row = 1
    row = _hrow(ws, row, *PLAYER_SUMMARY_HEADERS)

    # Build rows: one per real person who has at least one placement
    persons = [r for r in pt_rows if is_real_person(r)]
    # Sort: alphabetically by name
    persons.sort(key=lambda r: r.get("person_canon", "").lower())

    n_written = 0
    for r in persons:
        pid   = r.get("effective_person_id", "")
        pc    = r.get("person_canon", "")
        lid   = r.get("legacyid", "") or (member_id_map.get(pid) if member_id_map else None) or None

        stats = placement_stats.get(pid, {})
        if not stats:
            continue  # skip persons with no placements

        bap_nick    = bap_map[pc].get("nickname") if pc in bap_map else None

        values = [
            _fix_encoding(pc),
            bap_nick,
            stats.get("wins"),
            stats.get("podiums"),
            stats.get("placements"),
            stats.get("events"),
            lid,
        ]
        for col, v in enumerate(values, 1):
            align = ALIGN_RIGHT if isinstance(v, (int, float)) else ALIGN_LEFT
            _w(ws, row, col, v, font=FONT_DATA, align=align)
        row += 1
        n_written += 1

    # Column widths
    for col, width in enumerate(PLAYER_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PLAYER_SUMMARY_HEADERS))}1"

    print(f"  PLAYER SUMMARY: {n_written} rows written")


# ── HONORS sheet ──────────────────────────────────────────────────────────────

def build_honors_sheet(wb: Workbook,
                       pt_rows: list[dict],
                       placement_stats: dict[str, dict],
                       bap_map: dict[str, dict],
                       fbhof_map: dict[str, dict]) -> None:
    if "HONORS" in wb.sheetnames:
        del wb["HONORS"]

    # Insert after CONSECUTIVE RECORDS (or PLAYER SUMMARY if not found)
    insert_after = "CONSECUTIVE RECORDS"
    if insert_after in wb.sheetnames:
        idx = wb.sheetnames.index(insert_after) + 1
    else:
        idx = wb.sheetnames.index("PLAYER SUMMARY") + 1
    ws = wb.create_sheet("HONORS", idx)

    # Build lookup: person_canon → placement stats
    pc_to_pid = {
        r.get("person_canon", ""): r.get("effective_person_id", "")
        for r in pt_rows if r.get("person_canon")
    }

    def ya_for_canon(pc: str) -> str:
        pid   = pc_to_pid.get(pc, "")
        stats = placement_stats.get(pid, {})
        return years_active_str(stats)

    row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    _w(ws, row, 1, "COMMUNITY HONORS", font=FONT_TITLE, align=ALIGN_LEFT)
    row += 2

    # ── FBHOF Section ─────────────────────────────────────────────────────────
    row = _section(ws, row, "FOOTBAG HALL OF FAME")
    row += 1
    row = _hrow(ws, row, "Player", "Year Inducted", "Years Active")

    # Sort: numeric years first (ascending), "unknown" at end, then alpha
    def _fbhof_sort(item):
        pc, info = item
        y = info.get("year_inducted", "")
        try:
            return (0, int(y), pc.lower())
        except (ValueError, TypeError):
            return (1, 0, pc.lower())

    for pc, info in sorted(fbhof_map.items(), key=_fbhof_sort):
        year = info.get("year_inducted", "")
        ya   = ya_for_canon(pc)
        cell = ws.cell(row=row, column=1)
        cell.value     = pc
        cell.font      = FONT_DATA
        cell.fill      = FILL_HOF
        cell.alignment = ALIGN_LEFT
        for col, v in enumerate([year if year != "unknown" else "?", ya], 2):
            c = ws.cell(row=row, column=col)
            c.value     = v
            c.font      = FONT_DATA
            c.fill      = FILL_HOF
            c.alignment = ALIGN_LEFT
        row += 1

    row += 2

    # ── BAP Section ───────────────────────────────────────────────────────────
    row = _section(ws, row, "BIG ADD POSSE")
    row += 1
    row = _hrow(ws, row, "Player", "BAP #", "Year Inducted", "Nickname", "Years Active")

    # Sort by BAP number (chronological)
    for pc, info in sorted(bap_map.items(), key=lambda x: x[1]["bap_number"]):
        ya   = ya_for_canon(pc)
        bnum = info["bap_number"]
        year = info["year_inducted"]
        nick = info["nickname"]
        cell = ws.cell(row=row, column=1)
        cell.value     = pc
        cell.font      = FONT_DATA
        cell.fill      = FILL_BAP
        cell.alignment = ALIGN_LEFT
        for col, v in enumerate([bnum, year, nick or None, ya], 2):
            c = ws.cell(row=row, column=col)
            c.value     = v
            c.font      = FONT_DATA
            c.fill      = FILL_BAP
            c.alignment = ALIGN_RIGHT if isinstance(v, int) else ALIGN_LEFT
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12

    ws.freeze_panes = "A2"

    n_hof = len(fbhof_map)
    n_bap = len(bap_map)
    print(f"  HONORS sheet: {n_hof} FBHOF + {n_bap} BAP entries")


# ── FREESTYLE INSIGHTS sheet ─────────────────────────────────────────────────

def build_freestyle_insights(wb: Workbook) -> None:
    """Build FREESTYLE INSIGHTS sheet — vertical stacked tables, compact formatting."""
    from openpyxl.styles import Border, Side

    trick_freq  = load_csv(TRICK_FREQ_CSV)
    transitions = load_csv(TRANSITIONS_CSV)
    seq_diff    = load_csv(SEQ_DIFF_CSV)
    complexity  = load_csv(CHAIN_COMPLEXITY_CSV)
    diversity   = load_csv(DIVERSITY_CSV)
    trick_node  = load_csv(TRICK_NODE_CSV)

    data_missing = not any([trick_freq, transitions, seq_diff])

    if "FREESTYLE INSIGHTS" in wb.sheetnames:
        del wb["FREESTYLE INSIGHTS"]

    for anchor in ("CONSECUTIVE RECORDS", "PLAYER SUMMARY"):
        if anchor in wb.sheetnames:
            idx = wb.sheetnames.index(anchor) + 1
            break
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet("FREESTYLE INSIGHTS", idx)

    if data_missing:
        ws.column_dimensions["A"].width = 80
        msg = ws.cell(row=2, column=1)
        msg.value = (
            "Freestyle analytics not available. "
            "Run tools/09_compute_difficulty_analytics.py, "
            "10_compute_extended_analytics.py, and "
            "11_build_transition_network.py to generate the required data, "
            "then rebuild this workbook."
        )
        msg.font      = Font(italic=True, size=11, color="888888")
        msg.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 40
        print("  FREESTYLE INSIGHTS: placeholder written (analytics CSVs not found)")
        return

    # ── Styles ────────────────────────────────────────────────────────────────
    _thin    = Side(style="thin")
    _border  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _hfill   = _fill("D9D9D9")
    _no_wrap = Alignment(horizontal="left",  vertical="top", wrap_text=False)
    _no_wrap_r = Alignment(horizontal="right", vertical="top", wrap_text=False)

    def _c(row: int, col: int, value=None, *, header: bool = False) -> None:
        """Write a single bordered, non-wrapping cell."""
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = _border
        if header:
            cell.font  = Font(bold=True, size=11)
            cell.fill  = _hfill
            cell.alignment = _no_wrap
        else:
            cell.font  = Font(size=11)
            cell.alignment = (
                _no_wrap_r if isinstance(value, (int, float)) else _no_wrap
            )

    def _title(row: int, text: str) -> int:
        ws.cell(row=row, column=1).value = text
        ws.cell(row=row, column=1).font  = Font(bold=True, size=12)
        return row + 1

    def _hdr(row: int, *col_header_pairs) -> int:
        """Write header cells. col_header_pairs: (col, label), ..."""
        for col, h in col_header_pairs:
            _c(row, col, h, header=True)
        return row + 1

    def _narrative(row: int, text: str, *, italic: bool = False) -> int:
        """Write a full-width wrapped narrative paragraph."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1)
        cell.value = text
        cell.font = Font(italic=italic, size=10, color="333333")
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        lines = max(2, len(text) // 90 + 1)
        ws.row_dimensions[row].height = max(28, lines * 14)
        return row + 1

    def _section(row: int, text: str, *, color: str = "1F3864") -> int:
        """Write a bold section heading."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1)
        cell.value = text
        cell.font = Font(bold=True, size=12, color=color)
        cell.alignment = Alignment(horizontal="left", vertical="top")
        return row + 1

    # Column layout (shared across all stacked tables):
    #  A(1)=5   rank / era
    #  B(2)=28  trick / player / transition / era label
    #  C(3)=8   ADD / year / count / unique tricks / chains
    #  D(4)=12  mentions / connections / avg add / years active
    #  E(5)=10  players / length
    #  F(6)=50  events / sequence / notes
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 50

    _MODIFIERS = {
        "pixie", "ducking", "spinning", "atomic", "symposium",
        "stepping", "gyro", "barraging", "blazing", "tapping", "paradox",
    }

    def _addv(trick: str, adds_raw: str) -> object:
        if trick in _MODIFIERS:
            return "modifier"
        try:
            return int(float(adds_raw)) if adds_raw else None
        except (ValueError, TypeError):
            return None

    node_by_trick = {r["trick"]: r for r in trick_node}

    row = 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 1: Most Used Freestyle Tricks
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Used Freestyle Tricks")
    row = _hdr(row, (1, "#"), (2, "Trick"), (3, "ADD"),
               (4, "Mentions"), (5, "Players"), (6, "Events"))
    freq_sorted = sorted(trick_freq,
                         key=lambda r: _int(r, "total_mentions") or 0, reverse=True)
    for rank, r in enumerate(freq_sorted[:25], 1):
        trick = r.get("trick_canon", "")
        _c(row, 1, rank)
        _c(row, 2, trick)
        _c(row, 3, _addv(trick, r.get("adds", "")))
        _c(row, 4, _int(r, "total_mentions"))
        _c(row, 5, _int(r, "n_players"))
        _c(row, 6, _int(r, "n_events"))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 2: Most Influential Connector Tricks
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Influential Connector Tricks")
    row = _hdr(row, (1, "#"), (2, "Trick"), (3, "ADD"),
               (4, "Connections"), (5, "Players"), (6, "Events"))
    enriched = []
    for r in trick_freq:
        trick = r.get("trick_canon", "")
        nd  = node_by_trick.get(trick, {})
        deg = _int(nd, "degree") or 0
        enriched.append((deg, r))
    enriched.sort(key=lambda x: x[0], reverse=True)
    for rank, (deg, r) in enumerate(enriched[:15], 1):
        trick = r.get("trick_canon", "")
        _c(row, 1, rank)
        _c(row, 2, trick)
        _c(row, 3, _addv(trick, r.get("adds", "")))
        _c(row, 4, deg)
        _c(row, 5, _int(r, "n_players"))
        _c(row, 6, _int(r, "n_events"))
        row += 1
    row += 1
    row = _narrative(row,
        "From a network perspective, freestyle sequences exhibit a clear directional structure. "
        "Blurry whirl functions as the primary launch node, initiating high-difficulty sequences, "
        "while whirl serves as the dominant attractor, acting as the most common resolution point. "
        "This creates a highly asymmetric flow pattern in which sequences tend to begin with "
        "high-complexity rotational entries and resolve into more stable, clipper-based terminations.")
    row = _narrative(row,
        "The most common two-trick structure — blurry whirl \u2192 whirl — represents an optimal "
        "difficulty architecture, combining a high-ADD entry (5 ADD) with a stable resolution "
        "(3 ADD), balancing risk and control.")
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 3: Most Common Trick Transitions
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Common Trick Transitions")
    row = _hdr(row, (1, "#"), (2, "Transition"), (3, "Count"), (4, "Players"))
    trans_sorted = sorted(transitions,
                          key=lambda r: _int(r, "count") or 0, reverse=True)
    for rank, r in enumerate(trans_sorted[:20], 1):
        ta = r.get("trick_a", "")
        tb = r.get("trick_b", "")
        _c(row, 1, rank)
        _c(row, 2, f"{ta} → {tb}")
        _c(row, 3, _int(r, "count"))
        _c(row, 4, _int(r, "n_players"))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 4: Hardest Documented Sequences
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Hardest Documented Sequences")
    row = _hdr(row, (1, "#"), (2, "Player"), (3, "Year"),
               (4, "ADD"), (5, "Length"), (6, "Sequence"))
    scored_seqs = [
        r for r in seq_diff
        if r.get("sequence_add") and r.get("person_canon", "").strip()
        and r["person_canon"].strip() not in ("", "__NON_PERSON__")
    ]
    scored_seqs.sort(key=lambda r: float(r.get("sequence_add") or 0), reverse=True)
    for rank, r in enumerate(scored_seqs[:10], 1):
        _c(row, 1, rank)
        _c(row, 2, r.get("person_canon"))
        _c(row, 3, _int(r, "year"))
        _c(row, 4, _floatv(r, "sequence_add"))
        _c(row, 5, _int(r, "normalized_length"))
        _c(row, 6, r.get("tricks_normalized", "").replace(">", " > "))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 5: Most Diverse Players
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Diverse Players")
    row = _hdr(row, (1, "#"), (2, "Player"), (3, "Unique Tricks"), (4, "Years Active"))
    for rank, r in enumerate(diversity[:15], 1):
        y1 = _int(r, "year_first")
        y2 = _int(r, "year_last")
        _c(row, 1, rank)
        _c(row, 2, r.get("person_canon"))
        _c(row, 3, _int(r, "unique_tricks"))
        _c(row, 4, f"{y1}–{y2}" if y1 and y2 else "")
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 6: Evolution of Difficulty
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Evolution of Difficulty")
    row = _hdr(row, (2, "Era"), (3, "Chains"), (4, "Avg ADD"))
    complexity_by_year: dict[int, dict] = {}
    for r in complexity:
        y = _int(r, "year")
        if y:
            complexity_by_year[y] = r
    for label, y1, y2 in [
        ("2001–2003", 2001, 2003), ("2004–2006", 2004, 2006),
        ("2007–2009", 2007, 2009), ("2010–2015", 2010, 2015),
        ("2016–2025", 2016, 2025),
    ]:
        era_rows = [complexity_by_year[y] for y in range(y1, y2 + 1)
                    if y in complexity_by_year]
        if not era_rows:
            continue
        total_chains = sum(_int(r, "n_chains") or 0 for r in era_rows)
        weighted_sum = sum(
            (float(r.get("avg_avg_add") or 0)) * (_int(r, "n_chains") or 0)
            for r in era_rows
        )
        avg_add = round(weighted_sum / total_chains, 2) if total_chains else None
        _c(row, 2, label)
        _c(row, 3, total_chains)
        _c(row, 4, avg_add)
        row += 1
    row += 1
    row = _narrative(row,
        "This plateau suggests that freestyle did not continue to increase in raw technical "
        "difficulty after the mid-2000s. Instead, progress shifted toward consistency, execution "
        "quality, and the number of players capable of reaching the established ceiling, indicating "
        "a transition from technical expansion to competitive depth.")
    row = _narrative(row,
        "In this mature phase, innovation occurs primarily through recombination of existing "
        "components, rather than the introduction of fundamentally new trick structures.")
    row += 1
    row = _section(row, "European Dominance")
    row = _narrative(row,
        "The concentration of both podium finishes and high-difficulty sequence data among "
        "European players indicates that the competitive center of freestyle shifted geographically "
        "during this period. While early innovation was driven largely by North American players, "
        "the post-2005 era is characterized by European dominance in both performance and "
        "participation density.")
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 7: ADD Composition Examples
    # ═══════════════════════════════════════════════════════════════════════════
    row = _section(row, "ADD System")
    row = _narrative(row,
        "Modifiers represent additional body mechanics layered onto base tricks — including "
        "rotations (spinning, blurry), dexterities, and positional constraints (ducking, "
        "symposium, paradox, atomic). These increase not only nominal ADD value but also the "
        "timing precision, spatial coordination, and execution risk required within a single "
        "set cycle. Difficulty therefore scales not linearly, but through the interaction of "
        "multiple simultaneous constraints on body motion and control.")
    row = _narrative(row,
        "Some informal modifiers (e.g., quantum) have been proposed within the community but "
        "were never standardized within the ADD system. As such, they are excluded from this "
        "analysis to maintain consistency across the dataset.",
        italic=True)
    row += 1
    row = _title(row, "ADD Composition Examples")
    row = _hdr(row, (2, "Trick"), (3, "ADD"), (4, "Notes"))
    for trick, add, note in [
        ("whirl",         3,           "Most-connected trick in the network"),
        ("blurry whirl",  5,           "Rotational base + blurry modifier (+2)"),
        ("blurriest",     6,           "Maximum documented base ADD"),
        ("ripwalk",       4,           "High-frequency transition trick"),
        ("ducking whirl", 4,           "Modifier stack on rotational base"),
        ("pixie",         "modifier",  "Standalone difficulty modifier, no fixed ADD"),
    ]:
        _c(row, 2, trick)
        _c(row, 3, add)
        _c(row, 4, note)
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION: Limits of Freestyle Difficulty
    # ═══════════════════════════════════════════════════════════════════════════
    row = _section(row, "Limits of Freestyle Difficulty")
    row = _narrative(row,
        "Despite the theoretical openness of the ADD system, the dataset shows no sustained "
        "increase in single-trick difficulty beyond 6 ADD. This suggests a practical ceiling "
        "imposed by human biomechanics rather than scoring rules.")
    row += 1
    for bullet in [
        "finite airtime within a single set",
        "constraints on rotational speed and body positioning",
        "increasing coordination complexity with stacked modifiers",
        "the requirement for controlled stall completion",
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1)
        cell.value = "\u2022  " + bullet
        cell.font = Font(size=10, color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="top", indent=2)
        row += 1
    row += 1
    row = _narrative(row,
        "While higher ADD values (7+) may be theoretically possible, they appear to be extremely "
        "rare and not reproducible in competitive conditions. The observed plateau therefore "
        "reflects a physical boundary on achievable complexity.")
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # CONCLUSION
    # ═══════════════════════════════════════════════════════════════════════════
    row = _section(row, "Conclusion")
    row = _narrative(row,
        "Freestyle footbag evolved through two distinct phases: an early period of rapid "
        "innovation in which the core vocabulary was established, followed by a mature phase "
        "in which that vocabulary was fully exploited. The stabilization of difficulty, combined "
        "with increasing competitive depth and a geographic shift toward Europe, indicates that "
        "the sport has reached a state of structural completeness, where progress is defined not "
        "by new elements, but by the refinement and recombination of existing ones.")

    ws.freeze_panes = "A2"
    print(f"  FREESTYLE INSIGHTS sheet written ({row - 1} rows)")


def _int(r: dict, key: str) -> int | None:
    v = r.get(key)
    if v is None or v == "" or str(v).lower() == "nan":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _floatv(r: dict, key: str) -> float | None:
    v = r.get(key)
    if v is None or v == "" or str(v).lower() == "nan":
        return None
    try:
        return round(float(v), 3)
    except (ValueError, TypeError):
        return None


# ── DATA NOTES sheet ──────────────────────────────────────────────────────────

QUARANTINE_CSV_FULL = os.path.join(BASE_DIR, "inputs", "review_quarantine_events.csv")

def build_data_notes(wb: Workbook, fffd_count: int = 0, quarantine_count: int = 0) -> None:
    """Build DATA NOTES sheet natively from known_issues.csv and review_quarantine_events.csv."""
    if "DATA NOTES" in wb.sheetnames:
        idx = wb.sheetnames.index("DATA NOTES")
        del wb["DATA NOTES"]
        ws = wb.create_sheet("DATA NOTES", idx)
    else:
        ws = wb.create_sheet("DATA NOTES")

    FONT_TITLE_DN  = Font(bold=True, size=14)
    FONT_SECTION_DN = Font(bold=True, size=12)
    FONT_SUBHDR_DN  = Font(bold=True, size=11)
    FONT_BODY_DN    = Font(size=11)
    FONT_LABEL_DN   = Font(bold=True, size=11)
    FILL_SECTION_DN = _fill("D9E1F2")
    FILL_SUBSEC_DN  = _fill("EDF2FB")
    ALIGN_TOP_WRAP  = Alignment(wrap_text=True, vertical="top")
    ALIGN_TOP_L     = Alignment(vertical="top")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    row = 1

    def title(text):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = FONT_TITLE_DN
        ws.row_dimensions[row].height = 22
        row += 1

    def section(text):
        nonlocal row
        row += 1
        c = ws.cell(row=row, column=1, value=text)
        c.font = FONT_SECTION_DN
        c.fill = FILL_SECTION_DN
        ws.row_dimensions[row].height = 18
        row += 1

    def subhdr(text):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = FONT_SUBHDR_DN
        c.fill = FILL_SUBSEC_DN
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 16
        row += 1

    def body(text):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = FONT_BODY_DN
        c.alignment = ALIGN_TOP_WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 30
        row += 1

    def kv(label, value):
        nonlocal row
        a = ws.cell(row=row, column=1, value=label)
        a.font = FONT_LABEL_DN
        a.alignment = ALIGN_TOP_L
        b = ws.cell(row=row, column=2, value=value)
        b.font = FONT_BODY_DN
        b.alignment = ALIGN_TOP_WRAP
        ws.row_dimensions[row].height = max(15, min(60, len(str(value or "")) // 3))
        row += 1

    def blank():
        nonlocal row
        row += 1

    # ── Content ──────────────────────────────────────────────────────────────
    title("Data Notes — Source Limitations and Data Quality (Merged Canonical Edition)")

    section("Dataset Overview")
    body("This is the official merged canonical workbook combining the POST-1997 mirror dataset "
         "(Footbag.org, 1997–present) and the PRE-1997 historical reconstruction (Footbag World "
         "magazine, IFAB summaries, OLD_RESULTS.txt). Where both sources document the same real-world "
         "event, the PRE-1997 reconstruction takes precedence and the POST-1997 stub is excluded.")
    kv("Overlap filtering", "25 POST-1997 legacy event stubs for years <1997 are suppressed because a "
                             "richer PRE-1997 reconstruction exists for the same championship. The "
                             "full overlap analysis is at out/canonical_all_union/early_overlap_candidates.csv.")
    kv("POST-1997 early events kept", "18 POST-1997 events before 1997 have no PRE-1997 equivalent "
                                       "(regional opens, non-championship events) and are retained.")
    kv("Audit reference", "The provenance-preserving union (all events, no filtering) is at "
                           "out/canonical_all_union/. Use it for source research and gap analysis.")

    section("Source Coverage")
    body("POST-1997 coverage is from the Footbag.org mirror archive (comprehensive, 1997–present). "
         "PRE-1997 coverage is from Footbag World magazine scans, IFAB event summaries, and the "
         "OLD_RESULTS.txt compilation maintained by the footbag community.")
    kv("PRE-1997 sources", "FBW (Footbag World magazine) · IFAB (event summaries) · OLD_RESULTS (community compilation)")
    kv("1980–1986", "Partial — PRE-1997 reconstruction only. Primarily Worlds and major championships. "
                    "Standings are typically top finishers only; complete fields are rarely available.")
    kv("1987–1992", "Partial — PRE-1997 WFA World Championships + supplementary events. "
                    "1990–1991 results include East Coast Championships from the mirror.")
    kv("1993–1996", "Sparse — IFAB World Championships (4 events) + 1995 European Open. "
                    "1993 has no mirror coverage; reconstructed from IFAB summaries only.")
    kv("1997–2025", "Comprehensive — Footbag.org mirror archive. "
                    "Some events have partial standings (see Known-Issue Events below).")
    kv("2026",      "Included — season in progress at time of publication.")

    section("Data Quality Limitations")
    body("Player statistics (wins, podiums, placements) are computed from this incomplete record. "
         "Treat all career counts as lower bounds — they reflect documented results only.")
    kv("Missing dates",          "47 pre-1997 events have year-level precision only — specific month/day "
                                  "dates are not available from the source. These events are placed in "
                                  "their correct year. Events from 1997 onward have full dates.")
    kv("Host club coverage",     "Host club information is absent for many events. "
                                  "Only events where it appeared in the original source are populated.")
    kv("Location normalization", "Locations are standardised to City, Region, Country format. "
                                  "Some remote or rural events have approximate locations.")
    kv("Character encoding",     f"{fffd_count} U+FFFD replacement characters were present in the "
                                  "source data (primarily accented characters in French, Finnish, "
                                  "German, Polish, and Czech names from HTML archive encoding loss). "
                                  "Best-effort repair is applied at display time; residual cases "
                                  "cannot be recovered without the original source pages.")
    kv("Division merging",       "Some events combined Open and Intermediate divisions, or pool + "
                                  "final standings, under a single heading on Footbag.org. These are "
                                  "preserved as-is; the merged division name is documented in the "
                                  "Event Index sheet.")
    kv("Partial standings",      "For some events the source only published top-3 or top-5 finishers. "
                                  "These are marked 'partial' in the Event Index.")
    kv("Quarantined events",     f"{quarantine_count} events have structural issues (complex competition formats, "
                                  "duplicate results, or irreconcilable source data) that make "
                                  "deterministic parsing impossible. They appear in the Event Index "
                                  "highlighted in red but are excluded from all statistics.")
    kv("Player identity",        "All player identities are human-verified. 82 competitors remain "
                                  "'unresolved' — their names appear in the source but cannot be "
                                  "confidently matched to a canonical person. They are excluded from "
                                  "Player Summary statistics.")

    # ── Known-Issue Events ────────────────────────────────────────────────────
    # Load events to get names + years
    events_for_notes: dict = {}
    if os.path.exists(EVENTS_CSV):
        for r in load_csv(EVENTS_CSV):
            # v14: merged_events_normalized uses legacy_event_id as key
            eid = r.get("legacy_event_id", "") or r.get("event_id", "")
            events_for_notes[eid] = r

    ki_rows: list[dict] = []
    if os.path.exists(KNOWN_ISSUES_CSV):
        ki_rows = load_csv(KNOWN_ISSUES_CSV)

    section(f"Known-Issue Events ({len(ki_rows)} events — minor, moderate, or severe source limitations)")
    body("These events are included in the dataset but have documented data quality issues. "
         "The severity reflects impact on standings accuracy: minor = incomplete positions; "
         "moderate = merged divisions; severe = unreliable standings.")

    if ki_rows:
        hdrs = ["Year", "Event", "Severity", "Note"]
        widths_ki = [6, 46, 10, 46]
        for col_i, (h, w) in enumerate(zip(hdrs, widths_ki), start=1):
            c = ws.cell(row=row, column=col_i, value=h)
            c.font = FONT_SUBHDR_DN
            c.fill = FILL_SUBSEC_DN
            ws.column_dimensions[get_column_letter(col_i)].width = w
        row += 1
        for ki in sorted(ki_rows, key=lambda r: (events_for_notes.get(r.get("event_id",""), {}).get("year","9999"),
                                                   r.get("event_id",""))):
            eid = ki.get("event_id", "")
            ev  = events_for_notes.get(eid, {})
            yr  = ev.get("year") or ""
            nm  = ev.get("event_name") or eid
            ws.cell(row=row, column=1, value=yr).font = FONT_BODY_DN
            ws.cell(row=row, column=2, value=nm).font = FONT_BODY_DN
            ws.cell(row=row, column=3, value=ki.get("severity","")).font = FONT_BODY_DN
            c = ws.cell(row=row, column=4, value=ki.get("note",""))
            c.font = FONT_BODY_DN
            c.alignment = ALIGN_TOP_WRAP
            ws.row_dimensions[row].height = 14
            row += 1
        # Restore main col widths after table (table used all 4)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 90

    # ── Quarantined Events ────────────────────────────────────────────────────
    quar_rows: list[dict] = []
    if os.path.exists(QUARANTINE_CSV_FULL):
        quar_rows = load_csv(QUARANTINE_CSV_FULL)

    section(f"Quarantined Events ({len(quar_rows)} events — excluded from results)")

    if quar_rows:
        hdrs_q = ["Year", "Event", "Reason"]
        widths_q = [6, 56, 36]
        for col_i, (h, w) in enumerate(zip(hdrs_q, widths_q), start=1):
            c = ws.cell(row=row, column=col_i, value=h)
            c.font = FONT_SUBHDR_DN
            c.fill = FILL_SUBSEC_DN
            ws.column_dimensions[get_column_letter(col_i)].width = w
        row += 1
        for qr in sorted(quar_rows, key=lambda r: (r.get("year",""), r.get("event_name",""))):
            ws.cell(row=row, column=1, value=qr.get("year","")).font = FONT_BODY_DN
            c = ws.cell(row=row, column=2, value=qr.get("event_name",""))
            c.font = FONT_BODY_DN
            ws.cell(row=row, column=3, value=qr.get("reason","")).font = FONT_BODY_DN
            ws.row_dimensions[row].height = 14
            row += 1
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 90

    print(f"  DATA NOTES: {row - 1} rows written")


# ── CONSECUTIVE RECORDS sheet ─────────────────────────────────────────────────

CONSECUTIVES_CSV = os.path.join(BASE_DIR, "out", "consecutives_combined.csv")

def build_consecutive_records(wb: Workbook) -> None:
    """Build CONSECUTIVE RECORDS sheet from out/consecutives_combined.csv."""
    if "CONSECUTIVE RECORDS" in wb.sheetnames:
        idx = wb.sheetnames.index("CONSECUTIVE RECORDS")
        del wb["CONSECUTIVE RECORDS"]
        ws = wb.create_sheet("CONSECUTIVE RECORDS", idx)
    else:
        ws = wb.create_sheet("CONSECUTIVE RECORDS")

    if not os.path.exists(CONSECUTIVES_CSV):
        ws.cell(row=1, column=1, value="Consecutive records data not available.")
        ws.cell(row=1, column=1).font = Font(italic=True, color="888888")
        print("  CONSECUTIVE RECORDS: data file missing — placeholder written")
        return

    rows = load_csv(CONSECUTIVES_CSV)

    FONT_TITLE_CR   = Font(bold=True, size=14)
    FONT_SECTION_CR = Font(bold=True, size=12)
    FONT_SUBSEC_CR  = Font(bold=True, size=11)
    FONT_HDR_CR     = Font(bold=True, size=10)
    FONT_DATA_CR    = Font(size=10)
    FILL_SECTION_CR = _fill("D9E1F2")
    FILL_SUBSEC_CR  = _fill("EDF2FB")
    FILL_HDR_CR     = _fill("F2F2F2")
    ALIGN_L         = Alignment(horizontal="left", vertical="top")
    ALIGN_R         = Alignment(horizontal="right", vertical="top")

    COL_HEADERS = ["Year", "Rank", "Division", "Person / Team", "Partner",
                   "Score", "Note", "Event Date", "Event Name", "Location"]
    COL_KEYS    = ["year", "rank", "division", "person_or_team", "partner",
                   "score", "note", "event_date", "event_name", "location"]
    COL_WIDTHS  = [6, 6, 20, 28, 20, 10, 28, 13, 40, 28]

    for col_i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    cur_row = 1

    # Title
    c = ws.cell(row=cur_row, column=1, value="Consecutive Kicks Records")
    c.font = FONT_TITLE_CR
    ws.row_dimensions[cur_row].height = 22
    cur_row += 2

    cur_section    = None
    cur_subsection = None

    for r in sorted(rows, key=lambda x: (x.get("section",""), int(x.get("sort_order",0) or 0))):
        sec    = r.get("section", "")
        subsec = r.get("subsection", "")

        # Section header
        if sec != cur_section:
            cur_row += 1
            c = ws.cell(row=cur_row, column=1, value=sec)
            c.font = FONT_SECTION_CR
            c.fill = FILL_SECTION_CR
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(COL_HEADERS))
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1
            cur_section    = sec
            cur_subsection = None  # force subsection re-emit

        # Subsection header
        if subsec != cur_subsection:
            c = ws.cell(row=cur_row, column=1, value=subsec)
            c.font = FONT_SUBSEC_CR
            c.fill = FILL_SUBSEC_CR
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(COL_HEADERS))
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

            # Column headers
            for col_i, h in enumerate(COL_HEADERS, start=1):
                c = ws.cell(row=cur_row, column=col_i, value=h)
                c.font = FONT_HDR_CR
                c.fill = FILL_HDR_CR
            ws.row_dimensions[cur_row].height = 14
            cur_row += 1
            cur_subsection = subsec

        # Data row
        for col_i, key in enumerate(COL_KEYS, start=1):
            val = r.get(key) or None
            if val and key == "score":
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass
            align = ALIGN_R if key in ("score", "rank", "year") else ALIGN_L
            c = ws.cell(row=cur_row, column=col_i, value=val)
            c.font = FONT_DATA_CR
            c.alignment = align
        ws.row_dimensions[cur_row].height = 14
        cur_row += 1

    ws.freeze_panes = "A2"
    print(f"  CONSECUTIVE RECORDS: {len(rows)} records written")


# ── Year sheet builder ────────────────────────────────────────────────────────

CANONICAL_EVENTS_CSV = os.path.join(BASE_DIR, "out", "merged_events_normalized.csv")

# Styles
_YR_FILL_BANNER  = PatternFill("solid", fgColor="1F3864")
_YR_FILL_WORLDS  = PatternFill("solid", fgColor="2E4057")
_YR_FILL_META    = PatternFill("solid", fgColor="F2F2F2")
_YR_FILL_LABEL   = PatternFill("solid", fgColor="E4E4E4")
_YR_FILL_CAT     = PatternFill("solid", fgColor="F4F4F4")   # very light — subordinate label
_YR_FILL_DIV     = PatternFill("solid", fgColor="EBEBEB")
_YR_FILL_GOLD    = PatternFill("solid", fgColor="FFF3CC")
_YR_FILL_SILVER  = PatternFill("solid", fgColor="F0F0F0")
_YR_FILL_BRONZE  = PatternFill("solid", fgColor="FDEBD0")
_YR_FILL_PLACE   = PatternFill(fill_type=None)
_YR_FILL_QUAR    = PatternFill("solid", fgColor="FFCDD2")

_YR_FONT_BANNER  = Font(bold=True, size=11, color="FFFFFF")
_YR_FONT_META    = Font(size=9, color="555555")
_YR_FONT_LABEL   = Font(bold=True, size=9)
_YR_FONT_CAT     = Font(size=8, color="909090")   # small, muted — visually subordinate
_YR_FONT_DIV     = Font(bold=True, size=9)
_YR_FONT_PODIUM  = Font(bold=True, size=10)
_YR_FONT_PLACE   = Font(size=10)
_YR_FONT_STATUS  = Font(bold=True, size=9, color="B71C1C")

_YR_ALIGN_TOP    = Alignment(vertical="top", wrap_text=False)
_YR_ALIGN_WRAP   = Alignment(vertical="top", wrap_text=True)
_YR_ALIGN_RIGHT  = Alignment(horizontal="right", vertical="top", wrap_text=False)

# Fixed row positions in each event column
_YR_R_NAME    = 1
_YR_R_LOC     = 2
_YR_R_HOST    = 3
_YR_R_DATE    = 4
_YR_R_PLAYERS = 5
_YR_R_TYPE    = 6
_YR_R_EID     = 7
_YR_R_STATUS  = 8
_YR_R_DATA    = 9

_YR_ROW_LABELS = {
    _YR_R_NAME:    "Event",
    _YR_R_LOC:     "Location",
    _YR_R_HOST:    "Host Club",
    _YR_R_DATE:    "Date",
    _YR_R_PLAYERS: "Players",
    _YR_R_TYPE:    "Event Type",
    _YR_R_EID:     "Event ID",
    _YR_R_STATUS:  "Status",
}

_YR_CAT_ORDER  = ["net", "freestyle", "golf", "sideline", "unknown"]
_YR_CAT_LABELS = {
    "net": "NET", "freestyle": "FREESTYLE", "golf": "GOLF",
    "sideline": "SIDELINE", "unknown": "OTHER",
}
_YR_MEDALS = {1: "🥇", 2: "🥈", 3: "🥉"}

_YR_COL_W_LABEL = 10
_YR_COL_W_MIN   = 22


def load_events_for_year_sheets() -> dict:
    """
    Returns dict: legacy_event_id → event metadata dict.
    Uses out/canonical/events.csv for normalized location and all metadata.
    """
    result: dict = {}
    if not os.path.exists(CANONICAL_EVENTS_CSV):
        print(f"  WARNING: {CANONICAL_EVENTS_CSV} not found — year sheets will be empty")
        return result
    for r in load_csv(CANONICAL_EVENTS_CSV):
        eid = r.get("legacy_event_id", "").strip()
        if not eid:
            continue
        event_key = r.get("event_key", "").strip() or eid
        city    = r.get("city", "").strip()
        region  = r.get("region", "").strip()
        country = r.get("country", "").strip()
        # Location display: US/Canada → "City, State"; other → "City, Country"
        if city and country in ("United States", "Canada") and region:
            loc = f"{city}, {region}"
        elif city and country:
            loc = f"{city}, {country}"
        elif country:
            loc = country
        else:
            loc = ""
        def _iso_to_dmy(x: str) -> str:
            """
            Convert ISO date (YYYY-MM-DD) → DD/MM/YYYY.
            Leaves non-ISO inputs unchanged.
            """
            if not x:
                return x
            x = str(x).strip()
            m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", x)
            if not m:
                return x
            y, mo, d = m.groups()
            return f"{d}/{mo}/{y}"

        # Date: prefer start_date, fall back to year
        start = r.get("start_date", "").strip()
        end   = r.get("end_date", "").strip()
        if start and end and end != start:
            date_str = f"{_iso_to_dmy(start)} – {_iso_to_dmy(end)}"
        elif start:
            date_str = _iso_to_dmy(start)
        else:
            date_str = r.get("year", "").strip()
        result[eid] = {
            "event_id":   eid,
            "event_key":  event_key,
            "event_name": r.get("event_name", "").strip(),
            "year":       r.get("year", "").strip(),
            "date":       date_str,
            "location":   loc,
            "country":    country,
            "host_club":  r.get("host_club", "").strip(),
            "event_type": r.get("event_type", "").strip(),
        }
    return result


def load_quarantine_set() -> set:
    if not os.path.exists(QUARANTINE_CSV):
        return set()
    qs: set = set()
    with open(QUARANTINE_CSV, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            eid = r.get("event_id", "").strip()
            if eid:
                qs.add(eid)
    return qs


def build_placements_for_year_sheets(pf_rows: list[dict],
                                     canon_by_norm: dict[str, str] | None = None) -> dict:
    """
    Returns dict: event_id → {division_canon: [(place_int, display_name, category)]}.

    Includes resolved and unresolved persons. Skips __NON_PERSON__ singletons
    (but keeps doubles rows that have a full team_display_name).
    Deduplicates doubles by team_person_key.
    """
    def _infer_cat(div_name: str) -> str:
        """Infer discipline_category from division name when field is empty."""
        dn = div_name.lower()
        if any(k in dn for k in ("routines", "shred", "circle", "sick", "battle",
                                  "trick", "combo", "freestyle", "request", "ironman",
                                  "consecutive", "consec")):
            return "freestyle"
        if "golf" in dn:
            return "golf"
        if any(k in dn for k in ("sideline", "side-out", "side out", "rallye")):
            return "sideline"
        return "net"

    by_event: dict = defaultdict(list)
    for r in pf_rows:
        eid = str(r.get("event_id", "")).strip()
        if eid:
            by_event[eid].append(r)

    def _norm_div_abbrevs(name: str) -> str:
        """Expand pre-1997 division abbreviations so legacy CSVs don't split visually."""
        import re as _re
        name = _re.sub(r"\bdbls\b",   "Doubles", name, flags=_re.IGNORECASE)
        name = _re.sub(r"\bsgls\b",   "Singles", name, flags=_re.IGNORECASE)
        name = _re.sub(r"\bdobles\b", "Doubles", name, flags=_re.IGNORECASE)
        return name

    result: dict = {}
    for eid, rows in by_event.items():
        by_div: dict = defaultdict(list)
        for r in rows:
            dc = (r.get("division_canon") or "").strip() or "Unknown"
            dc = _norm_div_abbrevs(dc)
            by_div[dc].append(r)

        div_result: dict = {}
        for dc, div_rows in by_div.items():
            def _place_key(r):
                try:
                    return int(float(r.get("place") or 99999))
                except (ValueError, TypeError):
                    return 99999

            div_rows = sorted(div_rows, key=_place_key)
            entries: list = []
            seen_teams: set = set()

            for r in div_rows:
                person  = (r.get("person_canon") or "").strip()
                comp    = (r.get("competitor_type") or "player").lower()
                tpk     = (r.get("team_person_key") or "").strip()
                raw_cat = (r.get("division_category") or "").strip()
                cat     = raw_cat if raw_cat else _infer_cat(dc)
                team_d  = (r.get("team_display_name") or "").strip()

                # Skip __NON_PERSON__ singletons; keep doubles with a full display name
                if person == "__NON_PERSON__":
                    if not (comp == "team" and tpk and team_d
                            and not team_d.rstrip().endswith("/ ?")):
                        continue

                if not person and not team_d:
                    continue

                try:
                    place_int = int(float(r.get("place") or 0))
                except (ValueError, TypeError):
                    continue

                # Deduplicate doubles by team_person_key
                if comp == "team" and tpk:
                    if tpk in seen_teams:
                        continue
                    seen_teams.add(tpk)
                    display = team_d or person
                else:
                    display = person

                fixed = _fix_display(display)
                if canon_by_norm:
                    fixed = _canonicalize_display(fixed, canon_by_norm)
                entries.append((place_int, fixed, cat))

            if entries:
                div_result[dc] = entries

        if div_result:
            result[eid] = div_result

    return result


_EMPTY_PARTNER = re.compile(r'\s*/\s*\(\)|\(\)\s*/\s*')


_ISO2_MAP = {
    "\u00b9": "\u0161",   # ¹ → š
    "\u00b8": "\u017e",   # ¸ → ž
    "\u00a6": "\u015a",   # ¦ → Ś
    "\u00bf": "\u017c",   # ¿ → ż
    "\u00bc": "\u017a",   # ¼ → ź
    "\u00e8": "\u010d",   # è → č
    "\u00f2": "\u0142",   # ò → ł
    "\u00b6": "\u015b",   # ¶ → ś
}
_RE_FFFD_UC   = re.compile(r"\ufffd([A-Z])")
_RE_QS_APOS   = re.compile(r"\b(\w+)\?([Ss])\b")
_RE_MOJI_Q    = re.compile(r"Ï(.+?)Ó")


def _fix_encoding(s: str) -> str:
    """Best-effort repair of known encoding corruptions in display strings."""
    # ISO-8859-2 bytes misread as Latin-1
    for bad, good in _ISO2_MAP.items():
        s = s.replace(bad, good)
    # Mojibake smart-quotes: ÏwordÓ → "word"
    s = _RE_MOJI_Q.sub(lambda m: f'"{m.group(1)}"', s)
    # Women?s → Women's
    s = _RE_QS_APOS.sub(lambda m: m.group(1) + "'" + m.group(2).lower(), s)
    # U+FFFD before uppercase: FranÿCois → François (best-effort lowercase)
    s = _RE_FFFD_UC.sub(lambda m: m.group(1).lower(), s)
    # Strip remaining replacement chars
    s = s.replace("\ufffd", "")
    return s


def _fix_display(s: str) -> str:
    """
    Normalise a year-sheet display name for readability.

    1. Repair known encoding corruptions (ISO-8859-2, U+FFFD, QS_APOS, mojibake).
    2. Title-case tokens that are entirely uppercase (e.g. "ANIBAL MONTES" →
       "Anibal Montes").  Mixed-case tokens are left untouched.
    3. Strip "()" empty-partner placeholders (e.g. "Leanne Makcrow / ()" →
       "Leanne Makcrow").
    """
    # Encoding repair first
    s = _fix_encoding(s)

    # Team separator normalisation for display strings:
    #   "Kiss + Gyáni" → "Kiss / Gyáni"
    # Accept variable whitespace (and NBSP) around "+".
    s = s.replace("\u00a0", " ")

    # Handle "?" as a team separator used in some legacy French sources
    # E.g., "Team S. Thomas Sustrac ? Robinson Sustrac"
    # Strip "Team " prefix and replace " ? " with " / "
    if " ? " in s and " / " not in s:
        s = re.sub(r"^Team\s+", "", s, flags=re.IGNORECASE)
        s = s.replace(" ? ", " / ")

    # Handle "First Last (STATE) First Last" unsplit doubles pair
    # E.g., "Jim Fitzgerald (OR) Adam Hutchinson" → "Jim Fitzgerald / Adam Hutchinson"
    # Guard: left side must have ≥2 words so "Paul (PT) Lovern" (nickname) is not split.
    if " / " not in s:
        _m = re.search(r"^(.+?)\s+\([A-Z]{2,3}\)\s+(.+)$", s)
        if _m and len(_m.group(1).split()) >= 2:
            s = f"{_m.group(1).strip()} / {_m.group(2).strip()}"

    # Only convert "+" as team separator when the string does NOT already have
    # a " / " separator.  If " / " is present, the "+" is part of a player's
    # display name (e.g. "Michi+mr. Germany GER / [UNKNOWN PARTNER]") and must
    # not be split again.
    if " / " not in s:
        s = re.sub(r"(\S)\s*\+\s*(\S)", r"\1 / \2", s)

    # Strip empty-partner placeholder
    s = _EMPTY_PARTNER.sub("", s).strip().rstrip("/").strip()

    # Title-case each "/" -separated segment's words that are all-uppercase
    parts = s.split(" / ")
    fixed_parts = []
    for part in parts:
        words = part.split()
        fixed_words = []
        for w in words:
            alpha = [c for c in w if c.isalpha()]
            # Skip short all-caps tokens: country/state abbreviations like (USA), (CAN)
            if alpha and all(c.isupper() for c in alpha) and len(alpha) > 3:
                fixed_words.append(w.title())
            else:
                fixed_words.append(w)
        fixed_parts.append(" ".join(fixed_words))
    return " / ".join(fixed_parts)


def _canonicalize_display(s: str, canon_by_norm: dict[str, str]) -> str:
    """
    Resolve each "/" -separated segment of a display name against Persons_Truth.

    For each segment: normalize via _norm(), look up in canon_by_norm.
    If found, replace with the canonical PT name.  Otherwise leave unchanged.
    This fixes casing variants like "david Butcher" → "David Butcher" without
    any hardcoded names or naive title-casing.
    """
    parts = s.split(" / ")
    out = []
    for part in parts:
        key = _norm(part)
        if key and key in canon_by_norm:
            out.append(canon_by_norm[key])
        else:
            out.append(part)
    return " / ".join(out)


def _yr_c(ws, row: int, col: int, value=None, *,
          font=None, fill=None, align=None) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align


# ── Worlds classification ─────────────────────────────────────────────────────

_WORLDS_SLUG_OVERRIDES = frozenset({
    # 2021 Online Worlds: event_type=freestyle (format-specific), but IS a world championship
    "2021_worlds_online",
})


def _is_worlds(event_type: str, event_key: str, year_str: str) -> bool:
    """
    Return True if this event should be classified as a World Championship.

    PRE1997 worlds events now have event_type='worlds' in canonical_all.
    POST1997 worlds events have event_type='worlds' from the mirror pipeline.
    Slug overrides cover edge cases (e.g. online-only year).
    """
    if event_type == "worlds":
        return True
    if event_key in _WORLDS_SLUG_OVERRIDES:
        return True
    return False
    return False


def _effective_event_type(declared: str, event_key: str, year_str: str,
                           placements: dict) -> str:
    """
    Compute display event type from declared type + actual division categories.

    Rules:
    - worlds  (after _is_worlds check) → always 'worlds'
    - sideline / unknown categories excluded from inference
    - Remaining distinct categories:
        single category  → that category label ('net', 'freestyle', 'golf')
        multiple         → 'mixed'
    - No placements (or all sideline/unknown) → fall back to declared value
    """
    if _is_worlds(declared, event_key, year_str):
        return "worlds"
    cats: set[str] = set()
    for entries in placements.values():
        for _, _, cat in entries:
            if cat and cat not in ("sideline", "unknown"):
                cats.add(cat)
    if not cats:
        return declared or "unknown"
    return list(cats)[0] if len(cats) == 1 else "mixed"


def _write_year_event_col(ws, col: int, ev: dict, placements: dict,
                           is_quarantined: bool) -> tuple[int, int]:
    """
    Write one event into column `col`. Returns (last_row_written, max_content_len).
    """
    max_w = max(len(ev.get("event_name", "")), 24)

    def _w(r, val, font, fill, align=_YR_ALIGN_TOP):
        nonlocal max_w
        if val:
            max_w = max(max_w, len(str(val)))
        _yr_c(ws, r, col, val, font=font, fill=fill, align=align)

    eff_type   = _effective_event_type(ev.get("event_type", ""),
                                        ev.get("event_key", ""),
                                        ev.get("year", ""),
                                        placements)
    is_worlds  = eff_type == "worlds"
    ban_fill   = _YR_FILL_WORLDS if is_worlds else _YR_FILL_BANNER

    # Worlds: prefix event name with "WORLD CHAMPIONSHIPS —" label
    display_name = ev.get("event_name") or ""
    if is_worlds and "WORLD" not in display_name.upper():
        display_name = f"🌐 WORLD CHAMPIONSHIPS — {display_name}"
    elif is_worlds:
        display_name = f"🌐 {display_name}"

    _w(_YR_R_NAME,    display_name or "",                          _YR_FONT_BANNER, ban_fill, _YR_ALIGN_WRAP)
    _w(_YR_R_LOC,     ev.get("location")   or "—",                 _YR_FONT_META,   _YR_FILL_META)
    _w(_YR_R_HOST,    ev.get("host_club")  or "Not recorded",      _YR_FONT_META,   _YR_FILL_META)
    _w(_YR_R_DATE,    ev.get("date")       or "Not recorded",      _YR_FONT_META,   _YR_FILL_META)

    n_players = sum(len(v) for v in placements.values()) if placements else 0
    is_incomplete = (not is_quarantined) and 1 <= n_players <= 2

    _w(_YR_R_PLAYERS, f"Players: {n_players}",                     _YR_FONT_META,   _YR_FILL_META)
    _w(_YR_R_TYPE,    eff_type or "—",                             _YR_FONT_META,   _YR_FILL_META)
    # Show slug (event_key) as the canonical ID, fall back to numeric event_id
    _w(_YR_R_EID,     ev.get("event_key") or ev.get("event_id") or "",
                                                                    _YR_FONT_META,   _YR_FILL_META)

    if is_quarantined:
        label = "⛔ QUARANTINED — results may be incomplete or ambiguous"
        _yr_c(ws, _YR_R_STATUS, col, label,
              font=_YR_FONT_STATUS, fill=_YR_FILL_QUAR, align=_YR_ALIGN_TOP)
        max_w = max(max_w, len(label))
    elif is_incomplete:
        label = f"⚠ INCOMPLETE — only {n_players} placement{'s' if n_players != 1 else ''} on record"
        _yr_c(ws, _YR_R_STATUS, col, label,
              font=_YR_FONT_STATUS, fill=_YR_FILL_QUAR, align=_YR_ALIGN_TOP)
        max_w = max(max_w, len(label))

    row = _YR_R_DATA

    if not placements:
        _yr_c(ws, row, col, "No parseable results",
              font=_YR_FONT_META, fill=_YR_FILL_PLACE, align=_YR_ALIGN_TOP)
        return row, max_w

    # Group divisions by category, preserving dict insertion order within each
    cat_to_divs: dict = {}
    for div_name, entries in placements.items():
        if not entries:
            continue
        cat = entries[0][2] or "unknown"
        cat_to_divs.setdefault(cat, []).append((div_name, entries))

    for cat in _YR_CAT_ORDER:
        if cat not in cat_to_divs:
            continue
        cat_label = "  " + _YR_CAT_LABELS.get(cat, "OTHER")   # indent under event title
        _yr_c(ws, row, col, cat_label,
              font=_YR_FONT_CAT, fill=_YR_FILL_CAT, align=_YR_ALIGN_TOP)
        max_w = max(max_w, len(cat_label) + 2)
        row += 1

        for div_name, entries in cat_to_divs[cat]:
            _yr_c(ws, row, col, div_name,
                  font=_YR_FONT_DIV, fill=_YR_FILL_DIV, align=_YR_ALIGN_TOP)
            max_w = max(max_w, len(div_name) + 2)
            row += 1

            _tie_places = {p for p, cnt in Counter(e[0] for e in entries).items() if cnt > 1}

            for place_int, display, _ in entries:
                medal = _YR_MEDALS.get(place_int, "")
                _t = "T" if place_int in _tie_places else " "
                text  = f"{medal} {place_int:>3}{_t} {display}" if medal else f"    {place_int:>3}{_t} {display}"

                if place_int == 1:
                    fill, font = _YR_FILL_GOLD,   _YR_FONT_PODIUM
                elif place_int == 2:
                    fill, font = _YR_FILL_SILVER, _YR_FONT_PODIUM
                elif place_int == 3:
                    fill, font = _YR_FILL_BRONZE, _YR_FONT_PODIUM
                else:
                    fill, font = _YR_FILL_PLACE,  _YR_FONT_PLACE

                _yr_c(ws, row, col, text, font=font, fill=fill, align=_YR_ALIGN_TOP)
                max_w = max(max_w, len(text) + 2)
                row += 1

            row += 1  # blank row between divisions

    return row - 1, max_w


def build_all_year_sheets(wb: Workbook, pf_rows: list[dict],
                           events: dict, quarantine_set: set,
                           canon_by_norm: dict[str, str] | None = None) -> tuple[dict, dict]:
    """
    Build all year sheets directly from Placements_Flat.
    Quarantined events are included and marked with ⛔.
    Returns (event_col_map, placements_by_event):
      event_col_map:      {event_id: (sheet_name, col_letter)} for INDEX hyperlinks.
      placements_by_event: {event_id: {division: [(place, display, cat)]}}
    """
    print("\nBuilding year sheets from canonical data...")
    placements_by_event = build_placements_for_year_sheets(pf_rows, canon_by_norm)

    # Group events by year — include quarantined events even if no placements
    year_to_eids: dict = defaultdict(list)
    for eid, ev in events.items():
        yr = str(ev.get("year", "")).strip()
        if not yr or not yr.isdigit():
            continue
        if eid in placements_by_event or eid in quarantine_set:
            year_to_eids[yr].append(eid)

    total_events = sum(len(v) for v in year_to_eids.values())
    print(f"  {len(year_to_eids)} year sheets, {total_events} events total")

    event_col_map: dict = {}   # eid → (sheet_name, col_letter)

    for yr in sorted(year_to_eids.keys(), key=int):
        eids = sorted(
            year_to_eids[yr],
            key=lambda eid: events[eid].get("date", "") or "",
        )
        ws = wb.create_sheet(title=yr)

        # Column A: row labels
        ws.column_dimensions["A"].width = _YR_COL_W_LABEL
        for row_num, label in _YR_ROW_LABELS.items():
            _yr_c(ws, row_num, 1, label,
                  font=_YR_FONT_LABEL, fill=_YR_FILL_LABEL, align=_YR_ALIGN_RIGHT)

        col_max_widths: dict = {}
        for col_offset, eid in enumerate(eids, start=2):
            ev         = events[eid]
            placements = placements_by_event.get(eid, {})
            is_quar    = eid in quarantine_set
            _, max_w   = _write_year_event_col(ws, col_offset, ev, placements, is_quar)
            col_max_widths[col_offset] = max_w
            col_letter = get_column_letter(col_offset)
            event_col_map[eid] = (yr, col_letter)

            # Register workbook-level defined name anchor for this event
            anchor_name = f"event_{eid}"
            safe_yr     = yr.replace("'", "''")
            dn = DefinedName(
                name=anchor_name,
                attr_text=f"'{safe_yr}'!${col_letter}${_YR_R_NAME}",
            )
            wb.defined_names[anchor_name] = dn

        # Row heights
        ws.row_dimensions[_YR_R_NAME].height    = 36
        ws.row_dimensions[_YR_R_LOC].height     = 15
        ws.row_dimensions[_YR_R_HOST].height    = 15
        ws.row_dimensions[_YR_R_DATE].height    = 15
        ws.row_dimensions[_YR_R_PLAYERS].height = 15
        ws.row_dimensions[_YR_R_TYPE].height    = 15
        ws.row_dimensions[_YR_R_EID].height     = 13
        ws.row_dimensions[_YR_R_STATUS].height  = 15

        # Auto column widths
        for col_idx, max_w in col_max_widths.items():
            ltr = get_column_letter(col_idx)
            ws.column_dimensions[ltr].width = max(min(max_w + 4, 60), _YR_COL_W_MIN)

        ws.freeze_panes = "B1"

        qcount = sum(1 for e in eids if e in quarantine_set)
        qnote  = f" ({qcount} quarantined)" if qcount else ""
        print(f"  {yr}: {len(eids)} events{qnote}")

    return event_col_map, placements_by_event


# ── EVENT INDEX builder ───────────────────────────────────────────────────────

KNOWN_ISSUES_CSV = os.path.join(BASE_DIR, "overrides", "known_issues.csv")
COVERAGE_CSV     = os.path.join(BASE_DIR, "out", "Coverage_ByEventDivision.csv")

_IDX_FILL_OK       = PatternFill("solid", fgColor="E8F5E9")   # soft green
_IDX_FILL_QUAR     = PatternFill("solid", fgColor="FFCDD2")   # soft red
_IDX_FILL_SPARSE   = PatternFill("solid", fgColor="FFCDD2")   # same red — ≤2 placements = incomplete
_IDX_FILL_WORLDS   = PatternFill("solid", fgColor="FFF8E1")   # soft amber — world championship events
_IDX_FILL_HDR  = PatternFill("solid", fgColor="1F3864")
_IDX_FONT_HDR  = Font(bold=True, size=10, color="FFFFFF")
_IDX_FONT_LINK = Font(size=10, color="1155CC", underline="single")
_IDX_FONT_NORM = Font(size=10)
_IDX_FONT_NOTE = Font(italic=True, size=9, color="666666")
_IDX_ALIGN_CTR = Alignment(horizontal="center", vertical="top")
_IDX_ALIGN_L   = Alignment(horizontal="left",   vertical="top")

_FLAG_RANK = {"complete": 0, "partial": 1, "sparse": 2, "none": 3}


def load_known_issues() -> dict:
    """Returns {event_id: {severity, note}}."""
    result: dict = {}
    if not os.path.exists(KNOWN_ISSUES_CSV):
        return result
    for r in load_csv(KNOWN_ISSUES_CSV):
        eid = r.get("event_id", "").strip()
        if eid:
            result[eid] = {"severity": r.get("severity", ""), "note": r.get("note", "")}
    return result


def load_coverage_by_event() -> dict:
    """Returns {event_id: worst_coverage_flag}."""
    result: dict = {}
    if not os.path.exists(COVERAGE_CSV):
        return result
    for r in load_csv(COVERAGE_CSV):
        eid  = str(r.get("event_id", "")).strip()
        flag = r.get("coverage_flag", "complete").strip().lower()
        if eid:
            prev = result.get(eid, "complete")
            if _FLAG_RANK.get(flag, 0) > _FLAG_RANK.get(prev, 0):
                result[eid] = flag
    return result


def build_event_index(wb: Workbook,
                      events: dict,
                      placements_by_event: dict,
                      quarantine_set: set,
                      known_issues: dict,
                      coverage_by_event: dict,
                      event_col_map: dict) -> None:
    """
    Build EVENT INDEX sheet with one row per event, sorted by year then date.
    Event names are hyperlinked to their column in the corresponding year sheet.
    Quarantined rows are red; OK rows are green.
    Replaces the copy from v11 if it exists.
    """
    if "EVENT INDEX" in wb.sheetnames:
        idx_pos = wb.sheetnames.index("EVENT INDEX")
        del wb["EVENT INDEX"]
    else:
        idx_pos = 3   # after DATA NOTES, STATISTICS

    ws = wb.create_sheet("EVENT INDEX", idx_pos)
    ws.freeze_panes = "A2"

    cols   = ["Event Key", "Year", "Event Name", "City / Region", "Country",
              "Start Date", "Placements", "Divisions", "Coverage",
              "Status", "Notes"]
    widths = [38, 6, 48, 28, 18, 13, 11, 10, 12, 16, 48]

    for c, (h, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = _IDX_FONT_HDR
        cell.fill      = _IDX_FILL_HDR
        cell.alignment = _IDX_ALIGN_CTR
        ws.column_dimensions[get_column_letter(c)].width = w

    all_eids = sorted(
        events.keys(),
        key=lambda e: (events[e].get("year", "9999"),
                       events[e].get("date", "") or ""),
    )

    for row_idx, eid in enumerate(all_eids, start=2):
        ev  = events[eid]
        ep  = placements_by_event.get(eid, {})
        n_p = sum(len(v) for v in ep.values())
        n_d = len(ep)

        is_quar  = eid in quarantine_set
        issue    = known_issues.get(eid)
        cov_flag = coverage_by_event.get(eid, "complete" if n_p else "none")

        is_sparse    = (not is_quar) and (1 <= n_p <= 2)
        is_worlds_ev = _is_worlds(ev.get("event_type", ""),
                                   ev.get("event_key", eid),
                                   ev.get("year", ""))

        if is_quar:
            status = "QUARANTINED"
            notes  = "Excluded — ambiguous structure prevents deterministic parsing"
        elif n_p == 0:
            status = "NO_RESULTS"
            notes  = "No competitive results in dataset"
        elif is_sparse:
            status = "INCOMPLETE"
            notes  = f"Only {n_p} placement{'s' if n_p != 1 else ''} — likely partial record"
        elif issue:
            status = f"KNOWN_ISSUE ({issue['severity']})"
            notes  = issue["note"]
        elif is_worlds_ev:
            status = "WORLDS"
            notes  = ""
        else:
            status = "OK"
            notes  = ""

        # City display: US/Canada keep region, others just city
        city    = ev.get("location", "").split(",")[0].strip() if ev.get("location") else ""
        country = ev.get("country", "")

        values = [ev.get("event_key", eid), ev.get("year", ""), None,  # col 3 = event name (written below)
                  city, country,
                  ev.get("date", ""),
                  n_p or None, n_d or None,
                  cov_flag, status, notes or None]

        for c, val in enumerate(values, start=1):
            if c == 3:
                continue
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.font      = _IDX_FONT_NORM
            cell.alignment = _IDX_ALIGN_L
            if c == 11 and val:   # notes italic
                cell.font = _IDX_FONT_NOTE

        # Event name with hyperlink to named anchor on year sheet
        name_cell = ws.cell(row=row_idx, column=3, value=ev.get("event_name", ""))
        if eid in event_col_map:
            name_cell.hyperlink = f"#event_{eid}"
            name_cell.font      = _IDX_FONT_LINK
        else:
            name_cell.font = _IDX_FONT_NORM
        name_cell.alignment = _IDX_ALIGN_L

        # Row fill — priority: red (quarantine/incomplete) > amber (worlds) > green (OK)
        if is_quar or is_sparse:
            row_fill = _IDX_FILL_QUAR
        elif is_worlds_ev and not is_sparse:
            row_fill = _IDX_FILL_WORLDS
        elif status == "OK":
            row_fill = _IDX_FILL_OK
        else:
            row_fill = None
        if row_fill:
            for c in range(1, len(cols) + 1):
                ws.cell(row=row_idx, column=c).fill = row_fill

    print(f"  EVENT INDEX: {len(all_eids)} rows")


# ── Main ──────────────────────────────────────────────────────────────────────

FRONT_SHEETS_COPY = {
    "DATA NOTES", "EVENT INDEX", "CONSECUTIVE RECORDS",
}

# Sheets rebuilt or newly created (not copied from v11)
SHEETS_REBUILT = {"README", "STATISTICS", "FREESTYLE INSIGHTS"}

# ── Canonical-all statistics engine ───────────────────────────────────────────

CANONICAL_ALL_DIR = os.path.join(BASE_DIR, "out", "canonical_all")

# Event types that count as World Championships
_WORLDS_TYPES = frozenset({
    "worlds", "WORLD_CHAMPIONSHIPS", "WFA_WORLD_CHAMPIONSHIPS",
    "IFAB_WORLD_CHAMPIONSHIPS",
})

# Pre-1997 "nationals" events that were effectively world championships
_WORLDS_ALSO = frozenset({
    "NHSA_NATIONALS",  # 1980–1984 NHSA Nationals = world level
})


def _gender_cat(disc_name: str) -> str:
    dn = disc_name.lower()
    if any(w in dn for w in ("women", "woman", "ladies")):
        return "WOMEN"
    if "mixed" in dn:
        return "MIXED"
    return "OPEN"


def _dtype(disc_cat: str) -> str:
    dc = (disc_cat or "").lower().strip()
    if dc == "freestyle":
        return "FREESTYLE"
    if dc == "sideline":
        return "SIDELINE"
    if dc == "golf":
        return "GOLF"
    return "NET"


def _era(year: int) -> str:
    if year < 1990:
        return "1980s"
    if year < 2000:
        return "1990s"
    if year < 2010:
        return "2000s"
    if year < 2020:
        return "2010s"
    return "2020s+"


def _is_worlds_ctype(event_type: str, event_id: str) -> bool:
    if event_type in _WORLDS_TYPES:
        return True
    if event_id in _WORLDS_SLUG_OVERRIDES:
        return True
    return False


def load_canonical_stats():
    """
    Load canonical_all CSVs and build enriched per-placement records.
    Returns (records, pid_to_canon, events_list, persons_list) where each
    record is a dict with fields:
      person_id, canon, event_id, year, event_type, placement,
      disc_name, disc_cat, team_type, gender_cat, dtype, is_worlds, era
    """
    participants = load_csv(os.path.join(CANONICAL_ALL_DIR, "event_result_participants.csv"))
    disciplines  = load_csv(os.path.join(CANONICAL_ALL_DIR, "event_disciplines.csv"))
    events_list  = load_csv(os.path.join(CANONICAL_ALL_DIR, "events.csv"))
    persons_list = load_csv(os.path.join(CANONICAL_ALL_DIR, "persons.csv"))

    print(f"  canonical_all: {len(participants)} participants, {len(disciplines)} disciplines, "
          f"{len(events_list)} events, {len(persons_list)} persons")

    # Build lookup tables
    disc_meta: dict[tuple, dict] = {}
    for d in disciplines:
        disc_meta[(d["event_id"], d["discipline"])] = d

    event_meta: dict[str, dict] = {e["event_id"]: e for e in events_list}

    pid_to_canon: dict[str, str] = {
        p["person_id"]: p["person_canon"]
        for p in persons_list if p.get("person_id")
    }

    # Build enriched records
    records = []
    for p in participants:
        pid = p.get("person_id", "").strip()
        if not pid or pid == "__NON_PERSON__":
            continue
        canon = pid_to_canon.get(pid, "")
        if not canon or canon == "__NON_PERSON__":
            continue

        try:
            placement = int(p.get("placement", 0) or 0)
        except (ValueError, TypeError):
            placement = 0
        if placement <= 0:
            continue

        eid = p.get("event_id", "")
        disc_key = p.get("discipline", "")

        dm = disc_meta.get((eid, disc_key), {})
        disc_name = dm.get("discipline_name", disc_key) or disc_key
        disc_cat  = dm.get("discipline_category", "") or ""
        team_type = dm.get("team_type", "") or ""

        em = event_meta.get(eid, {})
        year_str   = em.get("year", "") or ""
        event_type = em.get("event_type", "") or ""

        try:
            year = int(year_str)
        except (ValueError, TypeError):
            year = 0

        records.append({
            "person_id":  pid,
            "canon":      canon,
            "event_id":   eid,
            "year":       year,
            "event_type": event_type,
            "event_name": em.get("event_name", ""),
            "placement":  placement,
            "disc_key":   disc_key,
            "disc_name":  disc_name,
            "disc_cat":   disc_cat,
            "team_type":  team_type,
            "gender_cat": _gender_cat(disc_name),
            "dtype":      _dtype(disc_cat),
            "is_worlds":  _is_worlds_ctype(event_type, eid),
            "era":        _era(year) if year else "unknown",
        })

    return records, pid_to_canon, events_list, persons_list


def _agg_bucket() -> dict:
    return {"p1": 0, "p2": 0, "p3": 0, "events": set(), "years": set()}


def _add_to_bucket(bucket: dict, placement: int, event_id: str, year: int) -> None:
    if placement == 1:
        bucket["p1"] += 1
    elif placement == 2:
        bucket["p2"] += 1
    elif placement == 3:
        bucket["p3"] += 1
    if event_id:
        bucket["events"].add(event_id)
    if year:
        bucket["years"].add(year)


def compute_canonical_stats(records: list[dict]) -> dict[str, dict]:
    """
    Aggregate per-person statistics from canonical_all records.
    Returns {person_id: stats_dict}.
    """
    stats: dict[str, dict] = {}

    for r in records:
        pid       = r["person_id"]
        placement = r["placement"]
        eid       = r["event_id"]
        year      = r["year"]
        gender    = r["gender_cat"]
        dtype     = r["dtype"]
        era       = r["era"]
        is_worlds = r["is_worlds"]

        if pid not in stats:
            stats[pid] = {
                "canon":    r["canon"],
                "total":    _agg_bucket(),
                "by_cat":   {g: _agg_bucket() for g in ("OPEN", "WOMEN", "MIXED")},
                "by_dtype": {t: _agg_bucket() for t in ("NET", "FREESTYLE", "SIDELINE", "GOLF")},
                "worlds":   _agg_bucket(),
                "by_era":   {},
                "all_events": set(),
                "all_years":  set(),
            }

        s = stats[pid]
        s["all_events"].add(eid)
        if year:
            s["all_years"].add(year)

        if placement <= 3:
            _add_to_bucket(s["total"], placement, eid, year)
            if gender in s["by_cat"]:
                _add_to_bucket(s["by_cat"][gender], placement, eid, year)
            if dtype in s["by_dtype"]:
                _add_to_bucket(s["by_dtype"][dtype], placement, eid, year)
            if is_worlds:
                _add_to_bucket(s["worlds"], placement, eid, year)

            if era not in ("unknown",):
                if era not in s["by_era"]:
                    s["by_era"][era] = _agg_bucket()
                _add_to_bucket(s["by_era"][era], placement, eid, year)

        # Also track event participation regardless of placement (for career stats)
        # (already done via s["all_events"])

    return stats


def _podium_total(bucket: dict) -> int:
    return bucket["p1"] + bucket["p2"] + bucket["p3"]


def _top_n(stats: dict, key_fn, n: int = 30) -> list[tuple]:
    """
    Sort persons by key_fn(stat_dict), return top-n as
    [(canon, stat_dict), ...].
    """
    ranked = [
        (s["canon"], s)
        for s in stats.values()
        if s["canon"]
    ]
    ranked.sort(key=lambda x: (-key_fn(x[1]), x[0].lower()))
    return ranked[:n]


def _lb_section_podiums(ws, row: int, title: str,
                        ranked: list[tuple], bucket_key: str,
                        top: int = 25) -> int:
    """Write a podium leaderboard section to ws, return next row."""
    row = _section(ws, row, title)
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total Podiums")
    for rank, (canon, s) in enumerate(ranked[:top], 1):
        b = s[bucket_key] if isinstance(s.get(bucket_key), dict) else s
        if _podium_total(b) == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    return row + 2


def _lb_section_wins(ws, row: int, title: str,
                     ranked: list[tuple], bucket_key: str,
                     top: int = 25) -> int:
    """Write a wins-only leaderboard section, return next row."""
    row = _section(ws, row, title)
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Wins")
    for rank, (canon, s) in enumerate(ranked[:top], 1):
        b = s[bucket_key] if isinstance(s.get(bucket_key), dict) else s
        if b["p1"] == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"])
    return row + 2


# ── LEADERBOARDS sheet ────────────────────────────────────────────────────────

def build_leaderboards(wb: Workbook, stats: dict[str, dict]) -> None:
    sheet_name = "LEADERBOARDS"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    # Insert after STATISTICS
    if "STATISTICS" in wb.sheetnames:
        idx = wb.sheetnames.index("STATISTICS") + 1
    else:
        idx = 3
    ws = wb.create_sheet(sheet_name, idx)

    _w(ws, 1, 1, "PODIUM & WIN LEADERBOARDS", font=FONT_TITLE, align=ALIGN_LEFT)
    _w(ws, 2, 1,
       "All-time career podiums and wins, by division category.  "
       "Doubles partners each receive individual credit.",
       font=FONT_DATA, align=ALIGN_LEFT)
    row = 4

    TOP = 25

    # ── OVERALL ──────────────────────────────────────────────────────────────
    row = _section(ws, row, "OVERALL — MOST CAREER PODIUMS")
    row += 1
    overall_pod = _top_n(stats, lambda s: _podium_total(s["total"]))
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
    for rank, (canon, s) in enumerate(overall_pod[:TOP], 1):
        b = s["total"]
        if _podium_total(b) == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    row += 2

    row = _section(ws, row, "OVERALL — MOST CAREER WINS")
    row += 1
    overall_wins = _top_n(stats, lambda s: s["total"]["p1"])
    row = _hrow(ws, row, "Rank", "Player", "Wins")
    for rank, (canon, s) in enumerate(overall_wins[:TOP], 1):
        if s["total"]["p1"] == 0:
            break
        row = _drow(ws, row, rank, canon, s["total"]["p1"])
    row += 2

    # ── OPEN ─────────────────────────────────────────────────────────────────
    open_pod = _top_n(stats, lambda s: _podium_total(s["by_cat"]["OPEN"]))
    row = _section(ws, row, "OPEN DIVISIONS — MOST CAREER PODIUMS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
    for rank, (canon, s) in enumerate(open_pod[:TOP], 1):
        b = s["by_cat"]["OPEN"]
        if _podium_total(b) == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    row += 2

    open_wins = _top_n(stats, lambda s: s["by_cat"]["OPEN"]["p1"])
    row = _section(ws, row, "OPEN DIVISIONS — MOST CAREER WINS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Wins")
    for rank, (canon, s) in enumerate(open_wins[:TOP], 1):
        if s["by_cat"]["OPEN"]["p1"] == 0:
            break
        row = _drow(ws, row, rank, canon, s["by_cat"]["OPEN"]["p1"])
    row += 2

    # ── WOMEN'S ───────────────────────────────────────────────────────────────
    wom_pod = _top_n(stats, lambda s: _podium_total(s["by_cat"]["WOMEN"]))
    row = _section(ws, row, "WOMEN'S DIVISIONS — MOST CAREER PODIUMS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
    for rank, (canon, s) in enumerate(wom_pod[:TOP], 1):
        b = s["by_cat"]["WOMEN"]
        if _podium_total(b) == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    row += 2

    wom_wins = _top_n(stats, lambda s: s["by_cat"]["WOMEN"]["p1"])
    row = _section(ws, row, "WOMEN'S DIVISIONS — MOST CAREER WINS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Wins")
    for rank, (canon, s) in enumerate(wom_wins[:TOP], 1):
        if s["by_cat"]["WOMEN"]["p1"] == 0:
            break
        row = _drow(ws, row, rank, canon, s["by_cat"]["WOMEN"]["p1"])
    row += 2

    # ── MIXED ─────────────────────────────────────────────────────────────────
    mix_pod = _top_n(stats, lambda s: _podium_total(s["by_cat"]["MIXED"]))
    row = _section(ws, row, "MIXED DIVISIONS — MOST CAREER PODIUMS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
    for rank, (canon, s) in enumerate(mix_pod[:TOP], 1):
        b = s["by_cat"]["MIXED"]
        if _podium_total(b) == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    row += 2

    mix_wins = _top_n(stats, lambda s: s["by_cat"]["MIXED"]["p1"])
    row = _section(ws, row, "MIXED DIVISIONS — MOST CAREER WINS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Wins")
    for rank, (canon, s) in enumerate(mix_wins[:TOP], 1):
        if s["by_cat"]["MIXED"]["p1"] == 0:
            break
        row = _drow(ws, row, rank, canon, s["by_cat"]["MIXED"]["p1"])
    row += 2

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "A4"
    print(f"  {sheet_name} sheet written")


# ── WORLDS STATS sheet ────────────────────────────────────────────────────────

def build_worlds_stats(wb: Workbook, stats: dict[str, dict],
                       records: list[dict]) -> None:
    sheet_name = "WORLDS STATS"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    if "LEADERBOARDS" in wb.sheetnames:
        idx = wb.sheetnames.index("LEADERBOARDS") + 1
    else:
        idx = 4
    ws = wb.create_sheet(sheet_name, idx)

    _w(ws, 1, 1, "WORLDS CHAMPIONSHIP STATISTICS", font=FONT_TITLE, align=ALIGN_LEFT)
    _w(ws, 2, 1,
       "Covers all World Footbag Championships: NHSA Nationals (1980–1984), "
       "WFA Worlds (1986–1992), IFAB Worlds (1993–1996), IFPA Worlds (1997–present).",
       font=FONT_DATA, align=ALIGN_LEFT)
    row = 4
    TOP = 25

    # ── Overall Worlds podiums ────────────────────────────────────────────────
    w_pod = _top_n(stats, lambda s: _podium_total(s["worlds"]))
    row = _section(ws, row, "MOST WORLDS CHAMPIONSHIP PODIUMS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
    for rank, (canon, s) in enumerate(w_pod[:TOP], 1):
        b = s["worlds"]
        if _podium_total(b) == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    row += 2

    # ── Worlds wins ───────────────────────────────────────────────────────────
    w_wins = _top_n(stats, lambda s: s["worlds"]["p1"])
    row = _section(ws, row, "MOST WORLDS CHAMPIONSHIP WINS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Wins")
    for rank, (canon, s) in enumerate(w_wins[:TOP], 1):
        if s["worlds"]["p1"] == 0:
            break
        row = _drow(ws, row, rank, canon, s["worlds"]["p1"])
    row += 2

    # ── Women's Worlds leaders ────────────────────────────────────────────────
    # Build per-person women's worlds bucket from records
    ww_stats: dict[str, dict] = {}
    for r in records:
        if not r["is_worlds"] or r["gender_cat"] != "WOMEN":
            continue
        pid = r["person_id"]
        if pid not in ww_stats:
            ww_stats[pid] = {"canon": r["canon"], "b": _agg_bucket()}
        if r["placement"] <= 3:
            _add_to_bucket(ww_stats[pid]["b"], r["placement"], r["event_id"], r["year"])

    ww_pod = sorted(
        [(v["canon"], v["b"]) for v in ww_stats.values() if _podium_total(v["b"]) > 0],
        key=lambda x: (-_podium_total(x[1]), x[0].lower())
    )
    row = _section(ws, row, "WOMEN'S WORLDS CHAMPIONSHIP LEADERS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
    for rank, (canon, b) in enumerate(ww_pod[:TOP], 1):
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    row += 2

    # ── World Champions list ──────────────────────────────────────────────────
    # One row per (year, discipline) winner at a Worlds event
    worlds_wins: list[tuple] = []
    # Track seen (event_id, disc_key) to avoid duplicate participant_orders
    seen_result: set[tuple] = set()
    for r in records:
        if not r["is_worlds"] or r["placement"] != 1:
            continue
        key = (r["event_id"], r["disc_key"])
        if key in seen_result:
            # Doubles: both partners are gold — aggregate display
            # We already added this placement, now add partner name
            for entry in worlds_wins:
                if entry[0] == r["year"] and entry[2] == r["event_id"] and entry[3] == r["disc_key"]:
                    # Append partner
                    entry_list = list(entry)
                    if r["canon"] not in entry_list[5]:
                        entry_list[5] = entry_list[5] + " / " + r["canon"]
                    worlds_wins[worlds_wins.index(entry)] = tuple(entry_list)
                    break
        else:
            seen_result.add(key)
            worlds_wins.append((
                r["year"],
                r["event_name"],
                r["event_id"],
                r["disc_key"],
                r["disc_name"],
                r["canon"],
            ))

    worlds_wins.sort(key=lambda x: (x[0], x[4].lower()))

    row = _section(ws, row, "WORLD CHAMPIONS — ALL TIME")
    row += 1
    row = _hrow(ws, row, "Year", "Event", "Division", "Champion(s)")
    for (year, ename, eid, dkey, dname, champs) in worlds_wins:
        row = _drow(ws, row, year or "?", ename, dname, champs)
    row += 2

    # Column widths chosen to work for both the podium tables and the World Champions list:
    # A=Rank/Year(6), B=Player/Event(40), C=Gold/Division(28), D=Silver/Champion(40),
    # E=Bronze(8), F=Total(10)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10

    ws.freeze_panes = "A4"
    print(f"  {sheet_name} sheet written ({len(worlds_wins)} world champion entries)")


# ── DISCIPLINE LEADERS sheet ──────────────────────────────────────────────────

def build_discipline_leaders(wb: Workbook, stats: dict[str, dict],
                             records: list[dict]) -> None:
    sheet_name = "DISCIPLINE LEADERS"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    if "WORLDS STATS" in wb.sheetnames:
        idx = wb.sheetnames.index("WORLDS STATS") + 1
    else:
        idx = 5
    ws = wb.create_sheet(sheet_name, idx)

    _w(ws, 1, 1, "DISCIPLINE LEADERS & ERA LEADERBOARDS", font=FONT_TITLE, align=ALIGN_LEFT)
    row = 3
    TOP = 25

    # ── Net discipline ────────────────────────────────────────────────────────
    net_pod = _top_n(stats, lambda s: _podium_total(s["by_dtype"]["NET"]))
    row = _section(ws, row, "NET DISCIPLINE — MOST CAREER PODIUMS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
    for rank, (canon, s) in enumerate(net_pod[:TOP], 1):
        b = s["by_dtype"]["NET"]
        if _podium_total(b) == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    row += 2

    # ── Freestyle discipline ──────────────────────────────────────────────────
    fs_pod = _top_n(stats, lambda s: _podium_total(s["by_dtype"]["FREESTYLE"]))
    row = _section(ws, row, "FREESTYLE DISCIPLINE — MOST CAREER PODIUMS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
    for rank, (canon, s) in enumerate(fs_pod[:TOP], 1):
        b = s["by_dtype"]["FREESTYLE"]
        if _podium_total(b) == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    row += 2

    # ── Sideline discipline ───────────────────────────────────────────────────
    sl_pod = _top_n(stats, lambda s: _podium_total(s["by_dtype"]["SIDELINE"]))
    row = _section(ws, row, "SIDELINE DISCIPLINE — MOST CAREER PODIUMS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
    for rank, (canon, s) in enumerate(sl_pod[:TOP], 1):
        b = s["by_dtype"]["SIDELINE"]
        if _podium_total(b) == 0:
            break
        row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
    row += 2

    # ── Era leaderboards ──────────────────────────────────────────────────────
    eras = ["1980s", "1990s", "2000s", "2010s", "2020s+"]
    for era_label in eras:
        era_ranked = sorted(
            [(s["canon"], s["by_era"].get(era_label, _agg_bucket()))
             for s in stats.values()
             if s.get("canon") and _podium_total(s["by_era"].get(era_label, _agg_bucket())) > 0],
            key=lambda x: (-_podium_total(x[1]), x[0].lower())
        )
        if not era_ranked:
            continue
        row = _section(ws, row, f"{era_label} ERA LEADERS — MOST PODIUMS")
        row += 1
        row = _hrow(ws, row, "Rank", "Player", "Gold", "Silver", "Bronze", "Total")
        for rank, (canon, b) in enumerate(era_ranked[:TOP], 1):
            row = _drow(ws, row, rank, canon, b["p1"], b["p2"], b["p3"], _podium_total(b))
        row += 2

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "A3"
    print(f"  {sheet_name} sheet written")


# ── CAREER STATS sheet ────────────────────────────────────────────────────────

def build_career_stats(wb: Workbook, stats: dict[str, dict],
                       records: list[dict]) -> None:
    sheet_name = "CAREER STATS"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    if "DISCIPLINE LEADERS" in wb.sheetnames:
        idx = wb.sheetnames.index("DISCIPLINE LEADERS") + 1
    else:
        idx = 6
    ws = wb.create_sheet(sheet_name, idx)

    _w(ws, 1, 1, "CAREER STATISTICS", font=FONT_TITLE, align=ALIGN_LEFT)
    row = 3
    TOP = 25

    # ── Longest careers ───────────────────────────────────────────────────────
    career_rows = []
    for s in stats.values():
        years = s["all_years"]
        if not years:
            continue
        yf, yl = min(years), max(years)
        span = yl - yf
        if span > 0:
            career_rows.append((s["canon"], yf, yl, span))
    career_rows.sort(key=lambda x: (-x[3], x[0].lower()))

    row = _section(ws, row, "LONGEST COMPETITIVE CAREERS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "First Year", "Last Year", "Career Span")
    for rank, (canon, yf, yl, span) in enumerate(career_rows[:TOP], 1):
        row = _drow(ws, row, rank, canon, yf, yl, span)
    row += 2

    # ── Most events competed ──────────────────────────────────────────────────
    events_rows = sorted(
        [(s["canon"], len(s["all_events"])) for s in stats.values() if s.get("canon")],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = _section(ws, row, "MOST EVENTS COMPETED")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Events")
    for rank, (canon, cnt) in enumerate(events_rows[:TOP], 1):
        if cnt == 0:
            break
        row = _drow(ws, row, rank, canon, cnt)
    row += 2

    # ── Most women's events competed ──────────────────────────────────────────
    wom_events: dict[str, set] = {}
    for r in records:
        if r["gender_cat"] != "WOMEN":
            continue
        pid = r["person_id"]
        wom_events.setdefault(pid, set()).add(r["event_id"])

    pid_to_canon = {s["canon"]: pid for pid, s in stats.items() for s in [s]}
    wom_events_rows = sorted(
        [(stats[pid]["canon"], len(eids))
         for pid, eids in wom_events.items()
         if pid in stats],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = _section(ws, row, "MOST EVENTS COMPETED — WOMEN'S DIVISIONS")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Events")
    for rank, (canon, cnt) in enumerate(wom_events_rows[:TOP], 1):
        if cnt == 0:
            break
        row = _drow(ws, row, rank, canon, cnt)
    row += 2

    # ── Most Worlds events ────────────────────────────────────────────────────
    worlds_events_rows = sorted(
        [(s["canon"], len(s["worlds"]["events"]))
         for s in stats.values()
         if s.get("canon") and s["worlds"]["events"]],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = _section(ws, row, "MOST WORLDS CHAMPIONSHIPS COMPETED")
    row += 1
    row = _hrow(ws, row, "Rank", "Player", "Worlds Events")
    for rank, (canon, cnt) in enumerate(worlds_events_rows[:TOP], 1):
        if cnt == 0:
            break
        row = _drow(ws, row, rank, canon, cnt)
    row += 2

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A3"
    print(f"  {sheet_name} sheet written")


def main():

    print("\nLoading canonical data...")
    pt_rows, canon_by_norm = load_persons_truth()
    pf_rows = load_csv(PF_CSV)
    print(f"  Persons Truth: {len(pt_rows)} rows")
    print(f"  Placements Flat: {len(pf_rows)} rows")

    print("\nLoading honors + stats data...")
    placement_stats  = build_placement_stats(pf_rows)
    bap_map          = load_bap(canon_by_norm)
    member_id_map    = load_member_ids()

    print("\nLoading year-sheet data...")
    yr_events      = load_events_for_year_sheets()
    yr_quarantine  = load_quarantine_set()
    print(f"  {len(yr_events)} events, {len(yr_quarantine)} quarantined")

    print("\nCreating output workbook...")
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    # Build front sheets in order
    # Compute dynamic stats for README / DATA NOTES
    _event_count     = len(yr_events)
    _person_count    = sum(1 for r in pt_rows if is_real_person(r))
    _placement_count = len(pf_rows)
    _fffd_count      = sum(
        v.count("\ufffd")
        for r in pf_rows
        for v in r.values()
        if isinstance(v, str)
    )

    print("\nBuilding README sheet...")
    build_readme(out_wb, quarantine_count=len(yr_quarantine),
                 event_count=_event_count, person_count=_person_count,
                 placement_count=_placement_count)

    print("\nBuilding DATA NOTES sheet...")
    build_data_notes(out_wb, fffd_count=_fffd_count, quarantine_count=len(yr_quarantine))

    print("\nLoading canonical_all for statistics...")
    canon_records, _c_pid_canon, _c_events, _c_persons = load_canonical_stats()
    canon_stats = compute_canonical_stats(canon_records)
    print(f"  {len(canon_records)} enriched records, {len(canon_stats)} persons with stats")

    print("\nBuilding STATISTICS sheet...")
    build_statistics(out_wb, pf_rows, pt_rows, events_dict=yr_events)

    print("\nBuilding LEADERBOARDS sheet...")
    build_leaderboards(out_wb, canon_stats)

    print("\nBuilding WORLDS STATS sheet...")
    build_worlds_stats(out_wb, canon_stats, canon_records)

    print("\nBuilding DISCIPLINE LEADERS sheet...")
    build_discipline_leaders(out_wb, canon_stats, canon_records)

    print("\nBuilding CAREER STATS sheet...")
    build_career_stats(out_wb, canon_stats, canon_records)

    print("\nBuilding PLAYER SUMMARY sheet...")
    build_player_summary(out_wb, pt_rows, placement_stats, bap_map, member_id_map)

    print("\nBuilding CONSECUTIVE RECORDS sheet...")
    build_consecutive_records(out_wb)

    print("\nBuilding FREESTYLE INSIGHTS sheet...")
    build_freestyle_insights(out_wb)

    # Build all year sheets directly from Placements_Flat (includes quarantined events)
    event_col_map, placements_by_event = build_all_year_sheets(
        out_wb, pf_rows, yr_events, yr_quarantine, canon_by_norm
    )

    # Build EVENT INDEX natively (must come after year sheets so hyperlinks resolve)
    print("\nBuilding EVENT INDEX sheet...")
    known_issues      = load_known_issues()
    coverage_by_event = load_coverage_by_event()
    # Pre-position: insert placeholder before first year sheet
    first_year = next((s for s in out_wb.sheetnames if s.isdigit()), None)
    if first_year:
        year_idx = out_wb.sheetnames.index(first_year)
        out_wb.create_sheet("EVENT INDEX", year_idx)
    build_event_index(
        out_wb, yr_events, placements_by_event,
        yr_quarantine, known_issues, coverage_by_event, event_col_map
    )

    # Verify every event that has a year-sheet column has a matching defined-name anchor.
    # (Events with no placements and not quarantined have no column → no anchor expected.)
    print("\nVerifying event anchor coverage...")
    defined = set(out_wb.defined_names.keys())
    missing_anchors = []
    for eid in event_col_map:
        anchor = f"event_{eid}"
        if anchor not in defined:
            missing_anchors.append(eid)
    if missing_anchors:
        print(f"  ERROR: {len(missing_anchors)} event(s) with year-sheet column but missing anchor:")
        for eid in missing_anchors[:10]:
            print(f"    {eid}")
        if len(missing_anchors) > 10:
            print(f"    ... and {len(missing_anchors) - 10} more")
    else:
        n_with_col = len(event_col_map)
        n_no_results = len(yr_events) - n_with_col
        print(f"  PASS: all {n_with_col} events with year-sheet columns have anchors "
              f"({n_no_results} events have no results and no column)")

    # Final sheet order check
    sheets = out_wb.sheetnames
    year_sheets = [s for s in sheets if s.isdigit()]
    front = [s for s in sheets if not s.isdigit()]
    print(f"\nFront sheets: {front}")
    print(f"Year sheets: {year_sheets[0]}–{year_sheets[-1]} ({len(year_sheets)} sheets)")

    print(f"\nSaving to: {OUTPUT_PATH}")
    out_wb.save(OUTPUT_PATH)
    size_mb = os.path.getsize(OUTPUT_PATH) / (1024 * 1024)
    print(f"Saved: {size_mb:.1f} MB")
    print("Done.")


def _insert_before(wb: Workbook, before_sheet: str, new_sheet_name: str) -> None:
    """Create new_sheet_name positioned immediately before before_sheet."""
    if new_sheet_name in wb.sheetnames:
        return  # already exists
    if before_sheet in wb.sheetnames:
        idx = wb.sheetnames.index(before_sheet)
        wb.create_sheet(new_sheet_name, idx)
    else:
        wb.create_sheet(new_sheet_name)


if __name__ == "__main__":
    main()
