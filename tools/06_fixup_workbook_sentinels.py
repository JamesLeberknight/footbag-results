#!/usr/bin/env python3
"""
06_fixup_workbook_sentinels.py

Post-build artifact repair step.

Ensures referential closure in the final Excel workbook by injecting
the reserved sentinel identity '__NON_PERSON__' into the Persons_Truth
sheet if it is referenced by Placements_ByPerson but not present in
Persons_Truth.

This does NOT modify the canonical identity lock.
It operates only on the workbook artifact.

Rationale:
The pipeline treats '__NON_PERSON__' as a classification bucket,
not a real identity. QC closure requires that any referenced canon
exist in Persons_Truth. This script enforces closure at the artifact level.
"""
from openpyxl import load_workbook

SENTINEL_CANON = "__NON_PERSON__"
SENTINEL_ID = "__NON_PERSON_ID__"

def find_col(headers, candidates):
    h = [str(x or "").strip().lower() for x in headers]
    for c in candidates:
        if c.lower() in h:
            return h.index(c.lower()) + 1
    return None

def main(inp, out):
    wb = load_workbook(inp)

    ws_pbp = wb["Placements_ByPerson"]
    ws_pt  = wb["Persons_Truth"]

    # PBP: locate person_canon col
    pbp_headers = [cell.value for cell in ws_pbp[1]]
    pbp_canon_col = find_col(pbp_headers, ["person_canon"])
    if not pbp_canon_col:
        raise SystemExit("Placements_ByPerson missing person_canon column")

    # Does PBP reference the sentinel?
    pbp_has = False
    for r in range(2, ws_pbp.max_row + 1):
        v = ws_pbp.cell(r, pbp_canon_col).value
        if str(v or "").strip() == SENTINEL_CANON:
            pbp_has = True
            break

    # PT: locate person_canon + person_id cols
    pt_headers = [cell.value for cell in ws_pt[1]]
    pt_canon_col = find_col(pt_headers, ["person_canon"])
    pt_id_col    = find_col(pt_headers, ["person_id", "effective_person_id"])
    if not pt_canon_col:
        raise SystemExit("Persons_Truth missing person_canon column")

    # Does PT already contain it?
    pt_has = False
    for r in range(2, ws_pt.max_row + 1):
        v = ws_pt.cell(r, pt_canon_col).value
        if str(v or "").strip() == SENTINEL_CANON:
            pt_has = True
            break

    # If PBP has it and PT doesn't, append a minimal sentinel row
    if pbp_has and not pt_has:
        new_row = [""] * ws_pt.max_column
        new_row[pt_canon_col - 1] = SENTINEL_CANON
        if pt_id_col:
            new_row[pt_id_col - 1] = SENTINEL_ID
        ws_pt.append(new_row)

    wb.save(out)

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("usage: python 06_fixup_workbook_sentinels.py IN.xlsx OUT.xlsx")
        raise SystemExit(2)
    main(sys.argv[1], sys.argv[2])
