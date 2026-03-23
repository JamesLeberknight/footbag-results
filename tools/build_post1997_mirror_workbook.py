#!/usr/bin/env python3
"""
build_post1997_mirror_workbook.py

Generates footbag_results_post1997_mirror_only.xlsx

Scope:
  - Events with year >= 1997 only
  - Source: Footbag.org HTML mirror
  - Pre-1997 and FBW-derived data explicitly excluded
  - Reads from out/canonical/*.csv

This is a NEW dataset with its own statistics, not a filtered view of the
full community workbook.  All statistics are recomputed from the subset.
"""

import csv, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

csv.field_size_limit(sys.maxsize)

ROOT         = Path(__file__).resolve().parent.parent
CANONICAL    = ROOT / "out" / "canonical"
NOISE_AGG    = ROOT / "out" / "noise_aggregates"
OUT_XLSX     = ROOT / "out" / "footbag_results_post1997_mirror_only.xlsx"
QUARANTINE_CSV = ROOT / "inputs" / "review_quarantine_events.csv"

# ── Constants ──────────────────────────────────────────────────────────────────

MIN_YEAR = 1997

# Worlds detection: conservative substring matches (uppercase comparison).
# Only these two patterns reliably identify World Championships without
# false-positives like "Worlds Warm-Up" or "World Record Attempt".
WORLDS_TRIGGERS = ("WORLD FOOTBAG", "IFPA WORLD")

# event_type values that are actually discipline categories (must be blanked)
REMOVE_EVENT_TYPES = {"net", "freestyle", "golf", "shred"}

# Discipline category display order
DISC_CAT_ORDER = {
    "net":      0,
    "freestyle": 1,
    "golf":     2,
    "sideline": 3,
    "other":    4,
    "unknown":  5,
    "":         6,
}

MEDALS = {1: "🥇", 2: "🥈", 3: "🥉"}

# ── Style palette ──────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def _align(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="top", wrap_text=wrap)

FILL_DARK_HDR      = _fill("1F3864")
FILL_WORLDS_HDR    = _fill("D4A800")   # amber – distinct but not garish
FILL_QUARANTINE    = _fill("FFCCCC")   # light red for quarantined events
FONT_QUARANTINE_HDR = _font(bold=True, color="990000", size=10)  # dark red bold
FILL_SECTION    = _fill("D9E1F2")
FILL_DISC       = _fill("EEF2F7")
FILL_CAT        = _fill("F5F5F5")
FILL_GOLD       = _fill("FFF2CC")
FILL_SILVER     = _fill("F2F2F2")
FILL_BRONZE     = _fill("FCE4D6")
FILL_STRIPE     = _fill("FAFAFA")

FONT_HDR        = _font(bold=True, color="FFFFFF", size=10)
FONT_HDR_LG     = _font(bold=True, color="FFFFFF", size=12)
FONT_WORLDS_COL = _font(bold=True, color="5C3D00", size=10)   # dark amber text on amber fill
FONT_SECT_COL   = _font(bold=True, color="1F3864", size=10)   # navy text on section fill
FONT_DISC_HDR   = _font(bold=True, color="1F3864", size=9)
FONT_CAT_HDR    = _font(bold=False, color="777777", size=8)
FONT_BOLD_10    = _font(bold=True, size=10)
FONT_NORM_10    = _font(size=10)
FONT_NORM_9     = _font(size=9)
FONT_BOLD_9     = _font(bold=True, size=9)
FONT_ITALIC_9   = _font(italic=True, size=9)
FONT_WORLDS_TAG = _font(bold=True, color="8B6914", size=9)

ALIGN_L   = _align("left",   wrap=True)
ALIGN_C   = _align("center", wrap=True)
ALIGN_L_NW = _align("left",  wrap=False)


def _w(ws, row, col, value, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:  cell.font  = font
    if fill:  cell.fill  = fill
    if align: cell.alignment = align
    return cell


def _hdr(ws, row, col, text, ncols=1):
    """Dark blue section header spanning ncols columns."""
    cell = _w(ws, row, col, text, font=FONT_HDR, fill=FILL_DARK_HDR, align=ALIGN_L)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + ncols - 1)
    return cell


# ── Data loading ───────────────────────────────────────────────────────────────

def _load(path):
    p = Path(path)
    if not p.exists():
        print(f"  WARNING: {p} not found")
        return []
    with open(p, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _anchor_name(event_key: str) -> str:
    """Return a valid Excel DefinedName for the event (must start with letter)."""
    import re
    return "ev_" + re.sub(r"[^A-Za-z0-9_]", "_", event_key)


def _int(r: dict, key: str):
    v = r.get(key)
    if v is None or v == "" or str(v).lower() == "nan":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _floatv(r: dict, key: str):
    v = r.get(key)
    if v is None or v == "" or str(v).lower() == "nan":
        return None
    try:
        return round(float(v), 3)
    except (ValueError, TypeError):
        return None


# ── Worlds detection & event_type normalisation ────────────────────────────────

def _is_worlds(event_name: str, existing_type: str) -> bool:
    if existing_type == "worlds":
        return True
    n = event_name.upper()
    return any(kw in n for kw in WORLDS_TRIGGERS)


def _norm_event_type(event_name: str, existing_type: str) -> str:
    if _is_worlds(event_name, existing_type):
        return "worlds"
    if existing_type in REMOVE_EVENT_TYPES:
        return ""
    return existing_type   # keep "mixed", "social" etc. as-is


# ── Filtering ──────────────────────────────────────────────────────────────────

def filter_events(events_raw: list) -> list:
    """Return events with year >= 1997, with normalised event_type."""
    out = []
    for r in events_raw:
        try:
            year = int(r.get("year") or 0)
        except ValueError:
            continue
        if year < MIN_YEAR:
            continue
        r = dict(r)
        r["year"] = year
        r["event_type"] = _norm_event_type(r.get("event_name", ""), r.get("event_type", ""))
        out.append(r)
    return out


# ── Lookup structures ──────────────────────────────────────────────────────────

def build_lookups(filtered_events, disciplines_raw, participants_raw):
    """
    Returns:
      events_by_key      : {event_key → event_dict}
      disciplines        : {(event_key, disc_key) → disc_dict}
      discs_per_event    : {event_key → [disc_dicts in display order]}
      placements_by_disc : {(event_key, disc_key) → [(place_int, display_str)]}
    """
    event_keys = {r["event_key"] for r in filtered_events}

    events_by_key = {r["event_key"]: r for r in filtered_events}

    # Index disciplines
    disciplines: dict = {}
    for r in disciplines_raw:
        if r["event_key"] in event_keys:
            disciplines[(r["event_key"], r["discipline_key"])] = r

    # Group disciplines per event in display order
    discs_per_event: dict = defaultdict(list)
    for (ek, dk), d in disciplines.items():
        discs_per_event[ek].append(d)
    for ek in discs_per_event:
        discs_per_event[ek].sort(key=lambda d: (
            DISC_CAT_ORDER.get(d.get("discipline_category", ""), 99),
            int(d.get("sort_order") or 0),
            d.get("discipline_name", ""),
        ))

    # Group participants by (event_key, disc_key, placement)
    raw_parts: dict = defaultdict(list)
    for r in participants_raw:
        if r["event_key"] in event_keys:
            raw_parts[(r["event_key"], r["discipline_key"], r["placement"])].append(r)

    # Build placements_by_disc with proper doubles joining
    placements_by_disc: dict = defaultdict(list)
    for (ek, dk, pl), parts in raw_parts.items():
        disc_info  = disciplines.get((ek, dk), {})
        team_type  = disc_info.get("team_type", "singles")
        place      = int(pl)

        sorted_p = sorted(parts, key=lambda x: int(x.get("participant_order") or 1))

        if team_type == "doubles":
            names = [
                p["display_name"] for p in sorted_p
                if p["display_name"] and p["display_name"] != "__NON_PERSON__"
            ]
            display = " / ".join(names) if names else "?"
        else:
            first = sorted_p[0] if sorted_p else {}
            display = first.get("display_name", "")
            if display == "__NON_PERSON__":
                display = ""

        if display:
            placements_by_disc[(ek, dk)].append((place, display))

    # Sort each list by place
    for key in placements_by_disc:
        placements_by_disc[key].sort(key=lambda x: x[0])

    return events_by_key, disciplines, dict(discs_per_event), dict(placements_by_disc)


# ── Player stats (recomputed from post-1997 subset only) ──────────────────────

def compute_player_stats(filtered_events, participants_raw):
    """
    Returns {person_id: {events, wins, podiums, placements, year_first, year_last}}.
    Counts are derived exclusively from the post-1997 mirror subset.
    Each (event_key, discipline_key, placement) participation counts as 1.
    """
    event_keys   = {r["event_key"] for r in filtered_events}
    year_by_key  = {r["event_key"]: int(r["year"]) for r in filtered_events}

    # Deduplicate on (person_id, event_key, disc_key, placement) so doubles
    # team members aren't double-counted when both participant rows are present.
    seen: set = set()
    stats: dict = defaultdict(lambda: {
        "event_set": set(), "p1": 0, "p2": 0, "p3": 0, "placements": 0, "years": set()
    })

    for r in participants_raw:
        if r["event_key"] not in event_keys:
            continue
        pid = r.get("person_id", "")
        if not pid or pid == "__NON_PERSON__":
            continue
        if not r.get("display_name"):
            continue

        key = (pid, r["event_key"], r["discipline_key"], r["placement"])
        if key in seen:
            continue
        seen.add(key)

        place = int(r.get("placement") or 0)
        year  = year_by_key.get(r["event_key"], 0)
        s     = stats[pid]
        s["event_set"].add(r["event_key"])
        s["placements"] += 1
        s["years"].add(year)
        if place == 1: s["p1"] += 1
        if place == 2: s["p2"] += 1
        if place == 3: s["p3"] += 1

    return {
        pid: {
            "events":     len(s["event_set"]),
            "wins":       s["p1"],
            "p2":         s["p2"],
            "p3":         s["p3"],
            "podiums":    s["p1"] + s["p2"] + s["p3"],
            "placements": s["placements"],
            "year_first": min(s["years"]) if s["years"] else 0,
            "year_last":  max(s["years"]) if s["years"] else 0,
        }
        for pid, s in stats.items()
    }


# ── Date formatting ────────────────────────────────────────────────────────────

def _fmt_date(start: str, end: str) -> str:
    def _parse(d):
        for fmt in ("%Y-%m-%d", "%Y-%m", "%Y"):
            try:
                return datetime.strptime(d, fmt)
            except ValueError:
                pass
        return None

    sd = _parse(start) if start else None
    if not sd:
        return start or ""
    s_str = sd.strftime("%d %b %Y").lstrip("0")

    ed = _parse(end) if end else None
    if ed and end != start:
        e_str = ed.strftime("%d %b %Y").lstrip("0")
        return f"{s_str}–{e_str}"
    return s_str


# ── README sheet ───────────────────────────────────────────────────────────────

def build_readme(wb, filtered_events):
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 72

    worlds_n = sum(1 for e in filtered_events if e.get("event_type") == "worlds")
    max_year = max(int(e["year"]) for e in filtered_events)

    rows = [
        # (label, value, row_type)  row_type: "header", "kv", "blank"
        ("FOOTBAG RESULTS — POST-1997 MIRROR DATASET", None, "header"),
        (None, None, "blank"),
        ("Dataset scope", "Mirror-only (Footbag.org HTML mirror, year ≥ 1997)", "kv"),
        ("Confidence level", "HIGH — all data derives from the authoritative online mirror", "kv"),
        ("Year range", f"{MIN_YEAR}–{max_year}", "kv"),
        ("Total events included", len(filtered_events), "kv"),
        ("World Championships events", worlds_n, "kv"),
        (None, None, "blank"),
        ("EXCLUSIONS", None, "header"),
        (None, None, "blank"),
        ("Pre-1997 data",
         "Excluded. Events before 1997 are not included in this dataset.",
         "kv"),
        ("FBW-derived data",
         "Excluded. Footbag World (FBW) magazine extractions are confined to "
         "pre-1997 years in this archive and are fully absent from this dataset.",
         "kv"),
        ("Legacy reconstruction",
         "Excluded. No manually inferred, estimated, or reconstructed results "
         "are present in this variant.",
         "kv"),
        (None, None, "blank"),
        ("IMPORTANT", None, "header"),
        (None, None, "blank"),
        ("Scope statement",
         "This dataset excludes all pre-1997 and FBW-derived results.",
         "kv"),
        ("Interpretation",
         "Statistics represent observed results within available mirror data — "
         "NOT definitive all-time records.",
         "kv"),
        ("Reliability principle",
         "This dataset prioritizes reliability over completeness.",
         "kv"),
    ]

    for row_i, (label, value, rtype) in enumerate(rows, 1):
        if rtype == "header":
            _hdr(ws, row_i, 1, label, ncols=2)
            ws.row_dimensions[row_i].height = 20
        elif rtype == "blank":
            ws.row_dimensions[row_i].height = 6
        else:
            _w(ws, row_i, 1, label, font=FONT_BOLD_10, align=ALIGN_L)
            _w(ws, row_i, 2, value, font=FONT_NORM_10, align=ALIGN_L)
            h = 15 + 15 * max(0, len(str(value or "")) // 72)
            ws.row_dimensions[row_i].height = h


# ── DATA NOTES sheet ───────────────────────────────────────────────────────────

def build_data_notes(wb):
    ws = wb.create_sheet("DATA NOTES")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 80

    sections = [
        ("A. SCOPE DEFINITION", [
            ("Dataset",
             "Post-1997 mirror-only subset of the Footbag canonical archive. "
             "High-confidence, reproducible, and suitable for public release."),
            ("Source",
             "All data derives from the Footbag.org HTML mirror — the authoritative "
             "online record of competitive results from approximately 1997 onward."),
            ("Confidence",
             "data_confidence = 'HIGH (mirror)' applies to all records in this dataset. "
             "Mirror records are treated as primary source; no external validation required."),
        ]),
        ("B. EXCLUSIONS", [
            ("FBW data excluded",
             "All data from Footbag World (FBW) magazine scans has been removed. "
             "FBW data is incomplete, inconsistently formatted, difficult to verify, "
             "and confined to the pre-internet era (pre-1997). Removing it improves "
             "dataset coherence."),
            ("Pre-1997 excluded",
             "Events before 1997 are not present. The pre-internet era has fragmentary "
             "coverage that would distort participation and achievement statistics if included."),
            ("No inference",
             "No results have been inferred, extrapolated, or manually reconstructed. "
             "If a result is absent from the mirror, it is absent here."),
        ]),
        ("C. DATA LIMITATIONS", [
            ("Formatting inconsistencies",
             "Mirror data contains formatting inconsistencies, occasional spelling "
             "errors in player names, and duplicate or fragmented event entries. "
             "These are minimised through canonicalisation but not fully eliminated."),
            ("Incomplete coverage",
             "Some events were never uploaded to Footbag.org. Some disciplines within "
             "events are missing (organiser did not post full results). "
             "Absence of a result does NOT mean the result did not happen."),
            ("Identity ambiguity",
             "Players may appear under multiple name variants. Not all aliases are "
             "resolved. Identity matching is conservative — when uncertain, the "
             "original source name is preserved rather than guessed."),
            ("Division name variation",
             "Division names are not standardised across events or years. The same "
             "discipline may appear under different labels (e.g., 'Open Freestyle' vs "
             "'Pro Freestyle'). No normalisation beyond the canonical pipeline is applied."),
            ("Temporal bias",
             "Earlier mirror years (late 1990s) are less complete than modern years. "
             "Participation metrics must not be compared across eras as if coverage "
             "were uniform. The dataset reflects internet-era recording practices."),
            ("This dataset prioritizes reliability over completeness.",
             "A result present here is very likely correct. "
             "Total-count statistics should be understood as lower bounds."),
        ]),
        ("D. INTERPRETATION GUIDANCE", [
            ("Statistics",
             "All statistics should be interpreted as: "
             "'Observed results within available mirror data (1997–present).'"),
            ("Not all-time records",
             "Do NOT treat placement counts, win tallies, or participation figures "
             "as definitive all-time records. Pre-1997 champions are not represented."),
            ("World Championships",
             "Worlds events are identified deterministically: event names containing "
             "'World Footbag' or 'IFPA World' are classified as Worlds. "
             "Ambiguous events are not tagged. Detection is conservative."),
            ("Event type normalisation",
             "event_type stores the event category only: 'worlds', 'mixed', 'social', "
             "or blank. Discipline-level types (net, freestyle, golf) have been removed "
             "from event_type and remain only in the discipline records."),
        ]),
    ]

    r = 1
    for section_title, items in sections:
        _hdr(ws, r, 1, section_title, ncols=2)
        ws.row_dimensions[r].height = 20
        r += 1
        for label, text in items:
            _w(ws, r, 1, label, font=FONT_BOLD_10, align=ALIGN_L)
            _w(ws, r, 2, text,  font=FONT_NORM_10, align=ALIGN_L)
            h = 15 + 15 * max(0, len(text) // 80)
            ws.row_dimensions[r].height = h
            r += 1
        ws.row_dimensions[r].height = 8
        r += 1   # blank gap between sections


# ── STATISTICS sheet ───────────────────────────────────────────────────────────

# Style constants aligned with full workbook
_STAT_FONT_TITLE   = Font(bold=True, size=13, name="Calibri")
_STAT_FONT_SECTION = Font(bold=True, size=12, name="Calibri")
_STAT_FONT_HEADER  = Font(bold=True, size=11, name="Calibri")
_STAT_FONT_DATA    = Font(size=11, name="Calibri")
_STAT_FILL_SECTION = PatternFill("solid", fgColor="E8F0FE")
_STAT_FILL_HEADER  = PatternFill("solid", fgColor="D9D9D9")
_STAT_ALN_L        = Alignment(horizontal="left",  vertical="top", wrap_text=False)
_STAT_ALN_R        = Alignment(horizontal="right", vertical="top", wrap_text=False)


def _stat_section(ws, row_num: int, text: str) -> int:
    _w(ws, row_num, 1, text, font=_STAT_FONT_SECTION, fill=_STAT_FILL_SECTION, align=_STAT_ALN_L)
    return row_num + 1


def _stat_hrow(ws, row_num: int, *headers) -> int:
    for col, h in enumerate(headers, 1):
        _w(ws, row_num, col, h, font=_STAT_FONT_HEADER, fill=_STAT_FILL_HEADER, align=_STAT_ALN_L)
    return row_num + 1


def _stat_drow(ws, row_num: int, *values) -> int:
    for col, v in enumerate(values, 1):
        align = _STAT_ALN_R if isinstance(v, (int, float)) else _STAT_ALN_L
        _w(ws, row_num, col, v, font=_STAT_FONT_DATA, align=align)
    return row_num + 1


def build_statistics(wb, player_stats, persons_raw, filtered_events):
    ws = wb.create_sheet("STATISTICS")
    ws.sheet_view.showGridLines = False

    persons_by_id = {p["person_id"]: p for p in persons_raw}

    # Build per-person podium and career data
    TOP_N = 25

    podium_rows = []
    wins_rows   = []
    events_rows = []
    career_rows = []

    for pid, s in player_stats.items():
        person = persons_by_id.get(pid, {})
        name   = person.get("person_name", pid)
        p1, p2, p3 = s["wins"], s["p2"], s["p3"]
        total  = p1 + p2 + p3
        if total > 0:
            podium_rows.append((name, p1, p2, p3, total))
        if p1 > 0:
            wins_rows.append((name, p1))
        events_rows.append((name, s["events"]))
        yf, yl = s["year_first"], s["year_last"]
        if yf and yl:
            career_rows.append((name, yf, yl, yl - yf))

    podium_rows.sort(key=lambda x: (-x[4], x[0].lower()))
    wins_rows.sort(key=lambda x: (-x[1], x[0].lower()))
    events_rows.sort(key=lambda x: (-x[1], x[0].lower()))
    career_rows.sort(key=lambda x: (-x[3], x[0].lower()))

    events_by_year = defaultdict(set)
    for e in filtered_events:
        events_by_year[int(e["year"])].add(e["event_key"])

    row_num = 1
    _w(ws, row_num, 1, "STATISTICS", font=_STAT_FONT_TITLE, align=_STAT_ALN_L)
    row_num += 2

    # ── 1. Most Career Podiums ────────────────────────────────────────────────
    row_num = _stat_section(ws, row_num, "MOST CAREER PODIUMS")
    row_num += 1
    row_num = _stat_hrow(ws, row_num, "Player", "1st", "2nd", "3rd", "Total Podiums")
    for canon, p1, p2, p3, total in podium_rows[:TOP_N]:
        row_num = _stat_drow(ws, row_num, canon, p1, p2, p3, total)
    row_num += 2

    # ── 2. Most Event Wins ────────────────────────────────────────────────────
    row_num = _stat_section(ws, row_num, "MOST EVENT WINS")
    row_num += 1
    row_num = _stat_hrow(ws, row_num, "Player", "Wins")
    for canon, wins in wins_rows[:TOP_N]:
        row_num = _stat_drow(ws, row_num, canon, wins)
    row_num += 2

    # ── 3. Most Events Competed ───────────────────────────────────────────────
    row_num = _stat_section(ws, row_num, "MOST EVENTS COMPETED")
    row_num += 1
    row_num = _stat_hrow(ws, row_num, "Player", "Events Competed")
    for canon, count in events_rows[:TOP_N]:
        row_num = _stat_drow(ws, row_num, canon, count)
    row_num += 2

    # ── 4. Longest Competitive Careers ───────────────────────────────────────
    row_num = _stat_section(ws, row_num, "LONGEST COMPETITIVE CAREERS")
    row_num += 1
    row_num = _stat_hrow(ws, row_num,
                         "Player", "First Event Year", "Last Event Year", "Career Span (Years)")
    for canon, yf, yl, span in career_rows[:TOP_N]:
        row_num = _stat_drow(ws, row_num, canon, yf, yl, span)
    row_num += 2

    # ── 5. Events by Year ─────────────────────────────────────────────────────
    row_num = _stat_section(ws, row_num, "EVENTS BY YEAR")
    row_num += 1
    row_num = _stat_hrow(ws, row_num, "Year", "Events")
    for year in sorted(events_by_year):
        row_num = _stat_drow(ws, row_num, year, len(events_by_year[year]))

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    ws.freeze_panes = "A2"
    print("  STATISTICS sheet written")


# ── PLAYER SUMMARY sheet ───────────────────────────────────────────────────────

def build_player_summary(wb, persons_raw, player_stats):
    ws = wb.create_sheet("PLAYER SUMMARY")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Columns: Player | Member ID | BAP Nickname | Wins | Podiums | Placements | Events
    COL_HEADERS = [
        ("Player",       32),
        ("Country",      18),
        ("Member ID",    13),
        ("BAP Nickname", 18),
        ("Wins",          7),
        ("Podiums",       9),
        ("Placements",   12),
        ("Events",        8),
    ]
    for col_i, (hdr, width) in enumerate(COL_HEADERS, 1):
        _w(ws, 1, col_i, hdr, font=FONT_HDR, fill=FILL_DARK_HDR, align=ALIGN_C)
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[1].height = 22

    persons_by_id = {p["person_id"]: p for p in persons_raw}

    # Sort alphabetically; only players with at least 1 post-1997 placement
    ordered = sorted(
        [(pid, s) for pid, s in player_stats.items() if s["placements"] > 0],
        key=lambda x: (persons_by_id.get(x[0], {}).get("person_name") or "").lower()
    )

    for row_i, (pid, s) in enumerate(ordered, 2):
        person      = persons_by_id.get(pid, {})
        name        = person.get("person_name", "")
        country     = person.get("country", "")
        member_id   = person.get("member_id", "") or ""
        bap_nick    = person.get("bap_nickname", "") or ""
        stripe      = FILL_STRIPE if row_i % 2 == 0 else None

        vals = [name, country, member_id, bap_nick,
                s["wins"], s["podiums"], s["placements"], s["events"]]
        for col_i, val in enumerate(vals, 1):
            _w(ws, row_i, col_i, val, font=FONT_NORM_9,
               fill=stripe, align=ALIGN_L_NW)
        ws.row_dimensions[row_i].height = 13


# ── CONSECUTIVE RECORDS placeholder ───────────────────────────────────────────

def build_consecutive_records(wb):
    ws = wb.create_sheet("CONSECUTIVE RECORDS")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    _hdr(ws, 1, 1, "CONSECUTIVE RECORDS — NOT INCLUDED IN THIS VARIANT", ncols=2)
    _w(ws, 2, 1, "Note", font=FONT_BOLD_10, align=ALIGN_L)
    _w(ws, 2, 2,
       "Consecutive records data is not derived solely from the Footbag.org mirror. "
       "This sheet is omitted from the post-1997 mirror-only variant to maintain scope "
       "purity. See the full community workbook for consecutive records.",
       font=FONT_NORM_10, align=ALIGN_L)
    ws.row_dimensions[2].height = 42


# ── FREESTYLE INSIGHTS sheet ───────────────────────────────────────────────────

def build_freestyle_insights(wb):
    """Build FREESTYLE INSIGHTS sheet — same structure as full workbook version.

    Loads from out/noise_aggregates/ CSVs.  All trick-sequence data is derived
    from post-1997 mirror events (Sick3 / sequence-scoring), so these values
    naturally reflect the post-1997 dataset.
    """
    from openpyxl.styles import Border, Side

    trick_freq  = _load(NOISE_AGG / "trick_frequency.csv")
    transitions = _load(NOISE_AGG / "trick_transition_network.csv")
    seq_diff    = _load(NOISE_AGG / "sequence_difficulty_conservative.csv")
    complexity  = _load(NOISE_AGG / "chain_complexity_by_year.csv")
    diversity   = _load(NOISE_AGG / "player_diversity_profiles.csv")
    trick_node  = _load(NOISE_AGG / "trick_node_metrics.csv")

    data_missing = not any([trick_freq, transitions, seq_diff])

    ws = wb.create_sheet("FREESTYLE INSIGHTS")

    if data_missing:
        ws.column_dimensions["A"].width = 80
        msg = ws.cell(row=2, column=1)
        msg.value = (
            "Freestyle analytics not available. "
            "Run tools/09_compute_difficulty_analytics.py, "
            "10_compute_extended_analytics.py, and "
            "11_build_transition_network.py to generate the required data, "
            "then rebuild this workbook."
        )
        msg.font      = Font(italic=True, size=11, color="888888")
        msg.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 40
        print("  FREESTYLE INSIGHTS: placeholder written (analytics CSVs not found)")
        return

    # ── Local style helpers (matching full workbook) ──────────────────────────
    _thin      = Side(style="thin")
    _border    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _hfill     = PatternFill("solid", fgColor="D9D9D9")
    _no_wrap   = Alignment(horizontal="left",  vertical="top", wrap_text=False)
    _no_wrap_r = Alignment(horizontal="right", vertical="top", wrap_text=False)

    def _c(row: int, col: int, value=None, *, header: bool = False) -> None:
        cell = ws.cell(row=row, column=col)
        cell.value  = value
        cell.border = _border
        if header:
            cell.font      = Font(bold=True, size=11)
            cell.fill      = _hfill
            cell.alignment = _no_wrap
        else:
            cell.font      = Font(size=11)
            cell.alignment = _no_wrap_r if isinstance(value, (int, float)) else _no_wrap

    def _title(row: int, text: str) -> int:
        ws.cell(row=row, column=1).value = text
        ws.cell(row=row, column=1).font  = Font(bold=True, size=12)
        return row + 1

    def _fhdr(row: int, *col_header_pairs) -> int:
        for col, h in col_header_pairs:
            _c(row, col, h, header=True)
        return row + 1

    def _narrative(row: int, text: str, *, italic: bool = False) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1)
        cell.value     = text
        cell.font      = Font(italic=italic, size=10, color="333333")
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        lines = max(2, len(text) // 90 + 1)
        ws.row_dimensions[row].height = max(28, lines * 14)
        return row + 1

    def _fsection(row: int, text: str, *, color: str = "1F3864") -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1)
        cell.value     = text
        cell.font      = Font(bold=True, size=12, color=color)
        cell.alignment = Alignment(horizontal="left", vertical="top")
        return row + 1

    # Column layout
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 50

    _MODIFIERS = {
        "pixie", "ducking", "spinning", "atomic", "symposium",
        "stepping", "gyro", "barraging", "blazing", "tapping", "paradox",
    }

    def _addv(trick: str, adds_raw: str):
        if trick in _MODIFIERS:
            return "modifier"
        try:
            return int(float(adds_raw)) if adds_raw else None
        except (ValueError, TypeError):
            return None

    node_by_trick = {r["trick"]: r for r in trick_node}

    row = 1

    # ── Scope note ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    scope_cell = ws.cell(row=row, column=1)
    scope_cell.value = (
        "All insights are derived from the post-1997 mirror-only dataset. "
        "Trick-sequence analytics are sourced from events that reported Sick3 "
        "or sequence-scoring results, all of which are post-1997 mirror events."
    )
    scope_cell.font      = Font(italic=True, size=10, color="333333")
    scope_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws.row_dimensions[row].height = 28
    row += 2

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 1: Most Used Freestyle Tricks
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Used Freestyle Tricks")
    row = _fhdr(row, (1, "#"), (2, "Trick"), (3, "ADD"),
                (4, "Mentions"), (5, "Players"), (6, "Events"))
    freq_sorted = sorted(trick_freq,
                         key=lambda r: _int(r, "total_mentions") or 0, reverse=True)
    for rank, r in enumerate(freq_sorted[:25], 1):
        trick = r.get("trick_canon", "")
        _c(row, 1, rank)
        _c(row, 2, trick)
        _c(row, 3, _addv(trick, r.get("adds", "")))
        _c(row, 4, _int(r, "total_mentions"))
        _c(row, 5, _int(r, "n_players"))
        _c(row, 6, _int(r, "n_events"))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 2: Most Influential Connector Tricks
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Influential Connector Tricks")
    row = _fhdr(row, (1, "#"), (2, "Trick"), (3, "ADD"),
                (4, "Connections"), (5, "Players"), (6, "Events"))
    enriched = []
    for r in trick_freq:
        trick = r.get("trick_canon", "")
        nd  = node_by_trick.get(trick, {})
        deg = _int(nd, "degree") or 0
        enriched.append((deg, r))
    enriched.sort(key=lambda x: x[0], reverse=True)
    for rank, (deg, r) in enumerate(enriched[:15], 1):
        trick = r.get("trick_canon", "")
        _c(row, 1, rank)
        _c(row, 2, trick)
        _c(row, 3, _addv(trick, r.get("adds", "")))
        _c(row, 4, deg)
        _c(row, 5, _int(r, "n_players"))
        _c(row, 6, _int(r, "n_events"))
        row += 1
    row += 1
    row = _narrative(row,
        "From a network perspective, freestyle sequences exhibit a clear directional structure. "
        "Blurry whirl functions as the primary launch node, initiating high-difficulty sequences, "
        "while whirl serves as the dominant attractor, acting as the most common resolution point. "
        "This creates a highly asymmetric flow pattern in which sequences tend to begin with "
        "high-complexity rotational entries and resolve into more stable, clipper-based terminations.")
    row = _narrative(row,
        "The most common two-trick structure — blurry whirl \u2192 whirl — represents an optimal "
        "difficulty architecture, combining a high-ADD entry (5 ADD) with a stable resolution "
        "(3 ADD), balancing risk and control.")
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 3: Most Common Trick Transitions
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Common Trick Transitions")
    row = _fhdr(row, (1, "#"), (2, "Transition"), (3, "Count"), (4, "Players"))
    trans_sorted = sorted(transitions,
                          key=lambda r: _int(r, "count") or 0, reverse=True)
    for rank, r in enumerate(trans_sorted[:20], 1):
        ta = r.get("trick_a", "")
        tb = r.get("trick_b", "")
        _c(row, 1, rank)
        _c(row, 2, f"{ta} \u2192 {tb}")
        _c(row, 3, _int(r, "count"))
        _c(row, 4, _int(r, "n_players"))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 4: Hardest Documented Sequences
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Hardest Documented Sequences")
    row = _fhdr(row, (1, "#"), (2, "Player"), (3, "Year"),
                (4, "ADD"), (5, "Length"), (6, "Sequence"))
    scored_seqs = [
        r for r in seq_diff
        if r.get("sequence_add") and r.get("person_canon", "").strip()
        and r["person_canon"].strip() not in ("", "__NON_PERSON__")
    ]
    scored_seqs.sort(key=lambda r: float(r.get("sequence_add") or 0), reverse=True)
    for rank, r in enumerate(scored_seqs[:10], 1):
        _c(row, 1, rank)
        _c(row, 2, r.get("person_canon"))
        _c(row, 3, _int(r, "year"))
        _c(row, 4, _floatv(r, "sequence_add"))
        _c(row, 5, _int(r, "normalized_length"))
        _c(row, 6, r.get("tricks_normalized", "").replace(">", " > "))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 5: Most Diverse Players
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Diverse Players")
    row = _fhdr(row, (1, "#"), (2, "Player"), (3, "Unique Tricks"), (4, "Years Active"))
    for rank, r in enumerate(diversity[:15], 1):
        y1 = _int(r, "year_first")
        y2 = _int(r, "year_last")
        _c(row, 1, rank)
        _c(row, 2, r.get("person_canon"))
        _c(row, 3, _int(r, "unique_tricks"))
        _c(row, 4, f"{y1}\u2013{y2}" if y1 and y2 else "")
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 6: Evolution of Difficulty
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Evolution of Difficulty")
    row = _fhdr(row, (2, "Era"), (3, "Chains"), (4, "Avg ADD"))
    complexity_by_year: dict = {}
    for r in complexity:
        y = _int(r, "year")
        if y:
            complexity_by_year[y] = r
    for label, y1, y2 in [
        ("2001\u20132003", 2001, 2003), ("2004\u20132006", 2004, 2006),
        ("2007\u20132009", 2007, 2009), ("2010\u20132015", 2010, 2015),
        ("2016\u20132025", 2016, 2025),
    ]:
        era_rows = [complexity_by_year[y] for y in range(y1, y2 + 1)
                    if y in complexity_by_year]
        if not era_rows:
            continue
        total_chains = sum(_int(r, "n_chains") or 0 for r in era_rows)
        weighted_sum = sum(
            (float(r.get("avg_avg_add") or 0)) * (_int(r, "n_chains") or 0)
            for r in era_rows
        )
        avg_add = round(weighted_sum / total_chains, 2) if total_chains else None
        _c(row, 2, label)
        _c(row, 3, total_chains)
        _c(row, 4, avg_add)
        row += 1
    row += 1
    row = _narrative(row,
        "This plateau suggests that freestyle did not continue to increase in raw technical "
        "difficulty after the mid-2000s. Instead, progress shifted toward consistency, execution "
        "quality, and the number of players capable of reaching the established ceiling, indicating "
        "a transition from technical expansion to competitive depth.")
    row = _narrative(row,
        "In this mature phase, innovation occurs primarily through recombination of existing "
        "components, rather than the introduction of fundamentally new trick structures.")
    row += 1
    row = _fsection(row, "European Dominance")
    row = _narrative(row,
        "The concentration of both podium finishes and high-difficulty sequence data among "
        "European players indicates that the competitive center of freestyle shifted geographically "
        "during this period. While early innovation was driven largely by North American players, "
        "the post-2005 era is characterized by European dominance in both performance and "
        "participation density.")
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 7: ADD Composition Examples
    # ═══════════════════════════════════════════════════════════════════════════
    row = _fsection(row, "ADD System")
    row = _narrative(row,
        "Modifiers represent additional body mechanics layered onto base tricks — including "
        "rotations (spinning, blurry), dexterities, and positional constraints (ducking, "
        "symposium, paradox, atomic). These increase not only nominal ADD value but also the "
        "timing precision, spatial coordination, and execution risk required within a single "
        "set cycle. Difficulty therefore scales not linearly, but through the interaction of "
        "multiple simultaneous constraints on body motion and control.")
    row = _narrative(row,
        "Some informal modifiers (e.g., quantum) have been proposed within the community but "
        "were never standardized within the ADD system. As such, they are excluded from this "
        "analysis to maintain consistency across the dataset.",
        italic=True)
    row += 1
    row = _title(row, "ADD Composition Examples")
    row = _fhdr(row, (2, "Trick"), (3, "ADD"), (4, "Notes"))
    for trick, add, note in [
        ("whirl",         3,           "Most-connected trick in the network"),
        ("blurry whirl",  5,           "Rotational base + blurry modifier (+2)"),
        ("blurriest",     6,           "Maximum documented base ADD"),
        ("ripwalk",       4,           "High-frequency transition trick"),
        ("ducking whirl", 4,           "Modifier stack on rotational base"),
        ("pixie",         "modifier",  "Standalone difficulty modifier, no fixed ADD"),
    ]:
        _c(row, 2, trick)
        _c(row, 3, add)
        _c(row, 4, note)
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION: Limits of Freestyle Difficulty
    # ═══════════════════════════════════════════════════════════════════════════
    row = _fsection(row, "Limits of Freestyle Difficulty")
    row = _narrative(row,
        "Despite the theoretical openness of the ADD system, the dataset shows no sustained "
        "increase in single-trick difficulty beyond 6 ADD. This suggests a practical ceiling "
        "imposed by human biomechanics rather than scoring rules.")
    row += 1
    for bullet in [
        "finite airtime within a single set",
        "constraints on rotational speed and body positioning",
        "increasing coordination complexity with stacked modifiers",
        "the requirement for controlled stall completion",
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1)
        cell.value     = "\u2022  " + bullet
        cell.font      = Font(size=10, color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="top", indent=2)
        row += 1
    row += 1
    row = _narrative(row,
        "While higher ADD values (7+) may be theoretically possible, they appear to be extremely "
        "rare and not reproducible in competitive conditions. The observed plateau therefore "
        "reflects a physical boundary on achievable complexity.")
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # CONCLUSION
    # ═══════════════════════════════════════════════════════════════════════════
    row = _fsection(row, "Conclusion")
    row = _narrative(row,
        "Freestyle footbag evolved through two distinct phases: an early period of rapid "
        "innovation in which the core vocabulary was established, followed by a mature phase "
        "in which that vocabulary was fully exploited. The stabilization of difficulty, combined "
        "with increasing competitive depth and a geographic shift toward Europe, indicates that "
        "the sport has reached a state of structural completeness, where progress is defined not "
        "by new elements, but by the refinement and recombination of existing ones.")

    ws.freeze_panes = "A2"
    print(f"  FREESTYLE INSIGHTS sheet written ({row - 1} rows)")


# ── EVENT INDEX sheet ──────────────────────────────────────────────────────────

def build_event_index(wb, filtered_events, discs_per_event, quarantine_keys: set):
    """Flat table of all included events. Quarantined rows highlighted red.
    Event Name cells are hyperlinked to the corresponding year-sheet column."""
    ws = wb.create_sheet("EVENT INDEX")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    _LINK_FONT = Font(underline="single", color="1F3864", size=9, name="Calibri")

    COL_HEADERS = [
        ("Event ID",     22),
        ("Year",          7),
        ("Event Name",   44),
        ("Event Type",   12),
        ("City",         18),
        ("Region",       14),
        ("Country",      12),
        ("Disciplines",  12),
        ("Notes",        40),
    ]
    for ci, (h, w) in enumerate(COL_HEADERS, 1):
        _w(ws, 1, ci, h, font=FONT_HDR, fill=FILL_DARK_HDR, align=ALIGN_C)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20

    events_sorted = sorted(filtered_events,
                            key=lambda e: (int(e["year"]), e.get("event_name", "")))

    for ri, ev in enumerate(events_sorted, 2):
        ek         = ev["event_key"]
        is_quar    = ek in quarantine_keys
        # Quarantined rows get red fill; otherwise alternate stripe
        row_fill   = FILL_QUARANTINE if is_quar else (FILL_STRIPE if ri % 2 == 0 else None)

        disc_list  = discs_per_event.get(ek, [])
        name_text  = ev.get("event_name", "")

        vals = [
            ek,
            int(ev["year"]),
            name_text,                          # col 3 — will get hyperlink
            ev.get("event_type", ""),
            ev.get("city", ""),
            ev.get("region", ""),
            ev.get("country", ""),
            len(disc_list),
            ev.get("notes", ""),
        ]
        for ci, val in enumerate(vals, 1):
            font = FONT_NORM_9
            if is_quar and ci == 3:
                font = Font(size=9, name="Calibri", color="990000")
            _w(ws, ri, ci, val, font=font, fill=row_fill, align=ALIGN_L_NW)

        # Hyperlink: Event Name cell → year-sheet column via DefinedName anchor
        anchor     = _anchor_name(ek)
        name_cell  = ws.cell(row=ri, column=3)
        name_cell.hyperlink = f"#{anchor}"
        name_cell.font      = Font(
            underline="single",
            color="990000" if is_quar else "1F3864",
            size=9, name="Calibri",
        )

        ws.row_dimensions[ri].height = 13


# ── YEAR SHEETS ────────────────────────────────────────────────────────────────

# Fixed row positions (same for every event column in a year sheet)
_R_NAME  = 1   # event name
_R_ID    = 2   # event_key / Event ID
_R_LOC   = 3   # location
_R_DATE  = 4   # dates
_R_TYPE  = 5   # event type (blank unless Worlds)
_R_GAP   = 6   # visual separator
_R_DATA  = 7   # placements start here


def _write_event_col(ws, col: int, event: dict,
                     discs_per_event: dict, placements_by_disc: dict,
                     is_quarantined: bool = False) -> int:
    """
    Write one event into column `col`.
    Returns the last row written (for callers that need to know extent).
    """
    ek        = event["event_key"]
    is_worlds = event.get("event_type") == "worlds"

    # Location string
    loc_parts = [p for p in (event.get("city",""), event.get("region",""),
                              event.get("country","")) if p]
    location  = ", ".join(loc_parts)

    # Header styling: quarantined > worlds > normal
    if is_quarantined:
        name_fill = FILL_QUARANTINE
        name_font = FONT_QUARANTINE_HDR
    elif is_worlds:
        name_fill = FILL_WORLDS_HDR
        name_font = FONT_WORLDS_COL
    else:
        name_fill = FILL_SECTION
        name_font = FONT_SECT_COL

    _w(ws, _R_NAME, col, event.get("event_name", ""), font=name_font, fill=name_fill, align=ALIGN_L)
    id_fill = FILL_QUARANTINE if is_quarantined else None
    _w(ws, _R_ID,   col, ek, font=FONT_NORM_9, fill=id_fill, align=ALIGN_L_NW)
    _w(ws, _R_LOC,  col, location,
       font=FONT_ITALIC_9, align=ALIGN_L)
    _w(ws, _R_DATE, col,
       _fmt_date(event.get("start_date",""), event.get("end_date","")),
       font=FONT_NORM_9, align=ALIGN_L_NW)
    if is_worlds:
        _w(ws, _R_TYPE, col, "WORLDS",
           font=FONT_WORLDS_TAG, fill=FILL_WORLDS_HDR, align=ALIGN_L_NW)

    # Placements
    row         = _R_DATA
    discs       = discs_per_event.get(ek, [])
    last_cat    = None

    for disc in discs:
        dk        = disc["discipline_key"]
        cat       = disc.get("discipline_category", "")
        disc_name = disc.get("discipline_name", dk)
        places    = placements_by_disc.get((ek, dk), [])

        if not places:
            continue  # skip disciplines with no results

        # Category separator header (NET / FREESTYLE / GOLF / …)
        if cat != last_cat:
            _w(ws, row, col, (cat.upper() if cat else "OTHER"),
               font=FONT_CAT_HDR, fill=FILL_CAT, align=ALIGN_L_NW)
            row      += 1
            last_cat  = cat

        # Discipline name header
        _w(ws, row, col, disc_name,
           font=FONT_DISC_HDR, fill=FILL_DISC, align=ALIGN_L)
        row += 1

        # Detect ties (same place number appearing more than once)
        place_counts = defaultdict(int)
        for place, _ in places:
            place_counts[place] += 1

        for place, display in places:
            medal     = MEDALS.get(place, "")
            tie_mark  = "T" if place_counts[place] > 1 else " "
            if medal:
                text  = f"{medal} {place:>3}{tie_mark} {display}"
            else:
                text  = f"    {place:>3}{tie_mark} {display}"

            if   place == 1: fill = FILL_GOLD
            elif place == 2: fill = FILL_SILVER
            elif place == 3: fill = FILL_BRONZE
            else:            fill = None

            _w(ws, row, col, text,
               font=FONT_BOLD_9 if place <= 3 else FONT_NORM_9,
               fill=fill, align=ALIGN_L)
            row += 1

    return row - 1


def build_year_sheets(wb, filtered_events, discs_per_event, placements_by_disc,
                      quarantine_keys: set):
    """Build one sheet per year. Registers DefinedName anchors for EVENT INDEX hyperlinks."""
    by_year = defaultdict(list)
    for e in filtered_events:
        by_year[int(e["year"])].append(e)

    for year in sorted(by_year):
        evs = sorted(
            by_year[year],
            key=lambda e: (e.get("start_date") or str(e["year"]), e.get("event_name", ""))
        )
        ws = wb.create_sheet(str(year))
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B1"

        # Column A: row labels
        lbl_fill = _fill("EEEEEE")
        lbl_font = _font(bold=True, size=9, color="555555")
        for row_n, label in [
            (_R_NAME, "Event"),    (_R_ID, "Event ID"),
            (_R_LOC,  "Location"), (_R_DATE, "Date"), (_R_TYPE, "Type"),
        ]:
            _w(ws, row_n, 1, label, font=lbl_font, fill=lbl_fill, align=ALIGN_L_NW)
        ws.column_dimensions["A"].width = 11

        # Row heights (fixed rows)
        ws.row_dimensions[_R_NAME].height = 30
        ws.row_dimensions[_R_ID].height   = 12
        ws.row_dimensions[_R_LOC].height  = 18
        ws.row_dimensions[_R_DATE].height = 14
        ws.row_dimensions[_R_TYPE].height = 14

        # Write each event into its own column; register DefinedName anchor
        sheet_name = str(year)
        for col_i, ev in enumerate(evs, 2):
            ek           = ev["event_key"]
            is_quar      = ek in quarantine_keys
            _write_event_col(ws, col_i, ev, discs_per_event, placements_by_disc,
                             is_quarantined=is_quar)
            # Register workbook-level anchor for EVENT INDEX hyperlinks
            col_letter   = get_column_letter(col_i)
            anchor       = _anchor_name(ek)
            safe_sheet   = sheet_name.replace("'", "''")
            dn = DefinedName(
                name=anchor,
                attr_text=f"'{safe_sheet}'!${col_letter}${_R_NAME}",
            )
            try:
                wb.defined_names[anchor] = dn
            except Exception:
                wb.defined_names.add(dn)
            # Column width: based on event name length, capped
            name_len = len(ev.get("event_name", ""))
            ws.column_dimensions[col_letter].width = max(22, min(52, name_len // 2 + 8))


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("Loading canonical CSVs…")
    events_raw       = _load(CANONICAL / "events.csv")
    disciplines_raw  = _load(CANONICAL / "event_disciplines.csv")
    results_raw      = _load(CANONICAL / "event_results.csv")       # not needed directly
    participants_raw = _load(CANONICAL / "event_result_participants.csv")
    persons_raw      = _load(CANONICAL / "persons.csv")

    print("Filtering to post-1997 mirror data…")
    filtered = filter_events(events_raw)
    excluded = len(events_raw) - len(filtered)
    worlds_n = sum(1 for e in filtered if e.get("event_type") == "worlds")
    print(f"  Included: {len(filtered)}  |  Excluded: {excluded}  |  Worlds: {worlds_n}")

    # Validate: no pre-1997 events, no wrong event_types
    assert all(int(e["year"]) >= MIN_YEAR for e in filtered), "Pre-1997 event slipped through"
    bad_types = {e["event_type"] for e in filtered if e["event_type"] in REMOVE_EVENT_TYPES}
    assert not bad_types, f"Invalid event_type values remain: {bad_types}"
    print("  Validation: OK")

    print("Building lookup structures…")
    events_by_key, disciplines, discs_per_event, placements_by_disc = build_lookups(
        filtered, disciplines_raw, participants_raw
    )
    print(f"  Disciplines: {len(disciplines)}")

    print("Computing player stats (post-1997 subset only)…")
    player_stats = compute_player_stats(filtered, participants_raw)
    print(f"  Players with post-1997 data: {len(player_stats)}")

    print("Loading quarantine list…")
    quarantine_raw     = _load(QUARANTINE_CSV)
    quarantine_eids    = {r.get("event_id", "").strip() for r in quarantine_raw
                          if r.get("event_id", "").strip()}
    # Map numeric event_id → event_key using legacy_event_id column in events.csv
    quarantine_keys    = {
        e["event_key"]
        for e in events_raw
        if e.get("legacy_event_id", "").strip() in quarantine_eids
    }
    print(f"  Quarantined event keys: {len(quarantine_keys)}")

    print("Building workbook…")
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet order: README, DATA NOTES, STATISTICS, PLAYER SUMMARY,
    #              FREESTYLE INSIGHTS, CONSECUTIVE RECORDS, year sheets, EVENT INDEX
    # EVENT INDEX must come AFTER year sheets so hyperlink anchors (DefinedNames) exist.
    build_readme(wb, filtered)
    build_data_notes(wb)
    build_statistics(wb, player_stats, persons_raw, filtered)
    build_player_summary(wb, persons_raw, player_stats)
    build_freestyle_insights(wb)
    build_consecutive_records(wb)
    build_year_sheets(wb, filtered, discs_per_event, placements_by_disc, quarantine_keys)
    build_event_index(wb, filtered, discs_per_event, quarantine_keys)

    print(f"Saving → {OUT_XLSX}")
    wb.save(OUT_XLSX)

    # ── Summary ──────────────────────────────────────────────────────────────
    print("\n────────────────────────────────────────────")
    print("SUMMARY")
    print(f"  Events included:       {len(filtered)}")
    print(f"  Events excluded:       {excluded}  (pre-1997 / all mirror source)")
    print(f"  World Championships:   {worlds_n}")
    print(f"  Year range:            {min(int(e['year']) for e in filtered)}–"
          f"{max(int(e['year']) for e in filtered)}")
    print(f"  Disciplines included:  {len(disciplines)}")
    print(f"  Players tracked:       {len(player_stats)}")
    print(f"  Output:                {OUT_XLSX}")
    print("────────────────────────────────────────────")
    print("Key differences vs full community workbook:")
    print(f"  – {excluded} pre-1997 events removed")
    print(f"  – event_type 'net/freestyle/golf' blanked; only 'worlds'/'mixed'/'social' kept")
    print(f"  – player stats recomputed from subset (not inherited from full PBP lock)")
    print(f"  – PLAYER SUMMARY: Member ID + BAP Nickname added; First/Last Year removed")
    print(f"  – FREESTYLE INSIGHTS sheet added (trick analytics, post-1997 only)")
    print(f"  – EVENT INDEX sheet added (flat event table with Event ID)")
    print(f"  – Event ID row added to each event column in year sheets")
    print(f"  – Worlds events highlighted with amber fill in year sheets")


if __name__ == "__main__":
    main()
