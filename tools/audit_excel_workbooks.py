#!/usr/bin/env python3
"""Audit two Excel workbooks before merging."""

import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FULL_PATH = "/home/james/projects/FOOTBAG_DATA/Footbag_Results_Community.xlsx"
V3_PATH   = "/home/james/projects/FOOTBAG_DATA/Footbag_Results_Community_v3.xlsx"

def sheet_summary(wb, name):
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    rows = ws.max_row
    cols = ws.max_column
    # Sample first few non-empty cells
    samples = []
    for r in ws.iter_rows(min_row=1, max_row=min(5, rows)):
        for c in r:
            if c.value is not None:
                samples.append(str(c.value)[:60])
            if len(samples) >= 10:
                break
        if len(samples) >= 10:
            break
    return {"rows": rows, "cols": cols, "samples": samples}

def compare_year_sheet(ws_full, ws_v3):
    """Compare two year sheets cell by cell. Returns (n_same, n_diff, diffs_sample)."""
    n_same = 0
    n_diff = 0
    diffs = []
    max_row = max(ws_full.max_row, ws_v3.max_row)
    max_col = max(ws_full.max_column, ws_v3.max_column)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v_full = ws_full.cell(r, c).value
            v_v3   = ws_v3.cell(r, c).value
            if v_full == v_v3:
                n_same += 1
            else:
                n_diff += 1
                if len(diffs) < 5:
                    diffs.append(f"  R{r}C{c}: full={repr(v_full)!s:.40s} | v3={repr(v_v3)!s:.40s}")
    return n_same, n_diff, diffs

def main():
    print("=" * 70)
    print("AUDIT: Footbag Excel Workbooks")
    print("=" * 70)

    print(f"\nLoading {FULL_PATH} ...")
    wb_full = load_workbook(FULL_PATH, data_only=True)
    print(f"  Sheets ({len(wb_full.sheetnames)}): {wb_full.sheetnames}")

    print(f"\nLoading {V3_PATH} ...")
    wb_v3 = load_workbook(V3_PATH, data_only=True)
    print(f"  Sheets ({len(wb_v3.sheetnames)}): {wb_v3.sheetnames}")

    # --- Year sheets comparison ---
    year_sheets_full = [s for s in wb_full.sheetnames if s.isdigit() and 1975 <= int(s) <= 2026]
    year_sheets_v3   = [s for s in wb_v3.sheetnames   if s.isdigit() and 1975 <= int(s) <= 2026]
    year_sheets_full.sort()
    year_sheets_v3.sort()

    print(f"\n--- YEAR SHEETS ---")
    print(f"  Full workbook years: {year_sheets_full}")
    print(f"  V3 workbook years:   {year_sheets_v3}")

    only_full = [y for y in year_sheets_full if y not in year_sheets_v3]
    only_v3   = [y for y in year_sheets_v3   if y not in year_sheets_full]
    in_both   = [y for y in year_sheets_full  if y in year_sheets_v3]

    if only_full:
        print(f"  Only in full: {only_full}")
    if only_v3:
        print(f"  Only in v3:   {only_v3}")

    print(f"\n  Comparing {len(in_both)} year sheets present in both...")
    all_identical = True
    for yr in in_both:
        ws_f = wb_full[yr]
        ws_v = wb_v3[yr]
        n_same, n_diff, diffs = compare_year_sheet(ws_f, ws_v)
        status = "IDENTICAL" if n_diff == 0 else f"DIFFERS ({n_diff} cells)"
        print(f"    {yr}: {status}  (rows_full={ws_f.max_row}, rows_v3={ws_v.max_row})")
        if n_diff > 0:
            all_identical = False
            for d in diffs:
                print(d)

    if all_identical:
        print("  RESULT: ALL year sheets are IDENTICAL between both workbooks.")
    else:
        print("  RESULT: Year sheets DIFFER — full workbook version will be used.")

    # 2026 check
    print(f"\n--- 2026 SHEET ---")
    if "2026" in wb_full.sheetnames:
        ws26 = wb_full["2026"]
        # Count non-empty cells
        non_empty = sum(1 for row in ws26.iter_rows() for c in row if c.value is not None)
        print(f"  Full workbook: 2026 PRESENT — {ws26.max_row} rows, {non_empty} non-empty cells")
    else:
        print("  Full workbook: 2026 NOT present")
    if "2026" in wb_v3.sheetnames:
        ws26v = wb_v3["2026"]
        non_empty = sum(1 for row in ws26v.iter_rows() for c in row if c.value is not None)
        print(f"  V3 workbook:   2026 PRESENT — {ws26v.max_row} rows, {non_empty} non-empty cells")
    else:
        print("  V3 workbook:   2026 NOT present")

    # --- Preliminary / front sheets ---
    non_year_full = [s for s in wb_full.sheetnames if not (s.isdigit() and 1975 <= int(s) <= 2026)]
    non_year_v3   = [s for s in wb_v3.sheetnames   if not (s.isdigit() and 1975 <= int(s) <= 2026)]

    print(f"\n--- FRONT/PRELIMINARY SHEETS ---")
    print(f"\n  FULL workbook front sheets: {non_year_full}")
    for name in non_year_full:
        info = sheet_summary(wb_full, name)
        if info:
            print(f"    '{name}': {info['rows']} rows x {info['cols']} cols")
            if info['samples']:
                print(f"      sample: {info['samples'][:4]}")

    print(f"\n  V3 workbook front sheets: {non_year_v3}")
    for name in non_year_v3:
        info = sheet_summary(wb_v3, name)
        if info:
            print(f"    '{name}': {info['rows']} rows x {info['cols']} cols")
            if info['samples']:
                print(f"      sample: {info['samples'][:4]}")

    # --- Index / Event Index row count ---
    print(f"\n--- INDEX / EVENT INDEX ROW COUNTS ---")
    for candidate in ["Index", "EVENT INDEX", "index", "Event Index"]:
        if candidate in wb_full.sheetnames:
            ws = wb_full[candidate]
            print(f"  Full workbook '{candidate}': {ws.max_row} rows x {ws.max_column} cols")
            # Print header row
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
            print(f"    Headers: {headers}")
        if candidate in wb_v3.sheetnames:
            ws = wb_v3[candidate]
            print(f"  V3 workbook '{candidate}': {ws.max_row} rows x {ws.max_column} cols")
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
            print(f"    Headers: {headers}")

    # --- V3 sheet detailed content ---
    print(f"\n--- V3 FRONT SHEETS DETAILED CONTENT ---")
    for name in non_year_v3:
        ws = wb_v3[name]
        print(f"\n  === {name} ({ws.max_row}r x {ws.max_column}c) ===")
        for r in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row)):
            row_vals = [str(c.value)[:50] if c.value is not None else "" for c in r]
            if any(v for v in row_vals):
                print(f"    Row {r[0].row}: {row_vals}")

    print("\n" + "=" * 70)
    print("AUDIT COMPLETE")
    print("=" * 70)

if __name__ == "__main__":
    main()
