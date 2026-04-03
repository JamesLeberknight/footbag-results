#!/usr/bin/env python3
"""
tools/build_workbook_v17.py

Community workbook v17 — final email-distribution polish from v16.

Changes vs v16 (presentation only, zero data changes):
  - Sheet order: README → EVENT INDEX → STATISTICS → ERA LEADERS →
                 PLAYER STATS → EXCLUDED EVENTS → year sheets
  - README: freeze_panes B2 (adds column-A freeze, consistent with all other sheets)
  - Version bump v16 → v17
"""

import csv
import os
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

csv.field_size_limit(sys.maxsize)

ROOT = Path(__file__).resolve().parents[1]

# ── Paths ──────────────────────────────────────────────────────────────────────
CA             = ROOT / "out" / "canonical_all"
RP_EVENTS_CSV  = ROOT / "out" / "release_publication" / "events.csv"
QUARANTINE_CSV = ROOT / "inputs" / "review_quarantine_events.csv"
OUTPUT_PATH    = ROOT / "Footbag_Results_Community_v17.xlsx"

VERSION        = "v17"
UPDATED        = date.today().isoformat()   # e.g. "2026-03-31"

# ── Styles ─────────────────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

FILL_NONE    = PatternFill(fill_type=None)
FILL_HEADER  = _fill("D9D9D9")
FILL_SECTION = _fill("1F3864")
FILL_WORLDS  = _fill("17375E")         # darker navy — worlds section dividers
FILL_WORLDS_ROW = _fill("C9DCF0")     # light blue — worlds event rows in EVENT INDEX
FILL_TITLE   = _fill("2E75B6")
FILL_NET     = _fill("E2EFDA")
FILL_FREE    = _fill("FFF2CC")
FILL_GOLF    = _fill("FCE4D6")
FILL_OTHER   = _fill("F2F2F2")
FILL_EVENT_A = _fill("DEEAF1")
FILL_EVENT_B = _fill("FFFFFF")
FILL_HOF     = _fill("FFF2CC")
FILL_BAP     = _fill("E2EFDA")
FILL_BOTH    = _fill("EAD1F5")
FILL_QUAR    = _fill("FFE0E0")
FILL_WARN    = _fill("FFF2CC")
WHITE        = _fill("FFFFFF")

FONT_TITLE   = Font(bold=True,  size=13, color="FFFFFF")
FONT_SECTION = Font(bold=True,  size=11, color="FFFFFF")
FONT_HEADER  = Font(bold=True,  size=11)
FONT_DATA    = Font(size=11)
FONT_SMALL   = Font(size=9,    color="808080")
FONT_NOTE    = Font(size=9,    italic=True, color="606060")
FONT_LINK    = Font(size=11,   color="0563C1", underline="single")
FONT_UNKNOWN = Font(size=10,   italic=True, color="A0A0A0")

ALIGN_L   = Alignment(horizontal="left",   vertical="top", wrap_text=False)
ALIGN_R   = Alignment(horizontal="right",  vertical="top")
ALIGN_C   = Alignment(horizontal="center", vertical="top")
ALIGN_LW  = Alignment(horizontal="left",   vertical="top", wrap_text=True)

# ── Sentinels (never included in stats / leaderboards) ─────────────────────────
_SKIP_PID   = {"", "__NON_PERSON__"}
_SKIP_DNAME = {"__NON_PERSON__", "__UNKNOWN_PARTNER__", "[UNKNOWN PARTNER]", "[UNKNOWN]", ""}
_UNKNOWN_DISPLAY = "[Unknown]"

# ── Helpers ─────────────────────────────────────────────────────────────────────

def _w(ws, row: int, col: int, value=None, *,
       font=None, fill=None, align=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font   is not None: cell.font          = font
    if fill   is not None: cell.fill          = fill
    if align  is not None: cell.alignment     = align
    if number_format:      cell.number_format = number_format
    return cell


def _title_row(ws, row: int, text: str, ncols: int = 8) -> int:
    _w(ws, row, 1, text, font=FONT_TITLE, fill=FILL_TITLE, align=ALIGN_L)
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = FILL_TITLE
    ws.row_dimensions[row].height = 22
    return row + 1


def _section_row(ws, row: int, text: str, ncols: int = 8,
                 fill=None, font=None) -> int:
    f = fill or FILL_SECTION
    fn = font or FONT_SECTION
    _w(ws, row, 1, text, font=fn, fill=f, align=ALIGN_L)
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = f
    ws.row_dimensions[row].height = 18
    return row + 1


def _worlds_section_row(ws, row: int, text: str, ncols: int = 8) -> int:
    return _section_row(ws, row, text, ncols=ncols, fill=FILL_WORLDS)


def _hrow(ws, row: int, *headers) -> int:
    for col, h in enumerate(headers, 1):
        _w(ws, row, col, h, font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_L)
    return row + 1


def _drow(ws, row: int, *values, fill=None) -> int:
    for col, v in enumerate(values, 1):
        align = ALIGN_R if isinstance(v, (int, float)) else ALIGN_L
        cell  = _w(ws, row, col, v, font=FONT_DATA, align=align)
        if fill:
            cell.fill = fill
    return row + 1


def _note_row(ws, row: int, text: str, ncols: int = 5) -> int:
    _w(ws, row, 1, text, font=FONT_NOTE, align=ALIGN_L)
    ws.row_dimensions[row].height = 13
    return row + 1


def _load(name: str) -> list[dict]:
    path = CA / name
    if not path.exists():
        print(f"  [WARN] missing {path}")
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _cat_fill(cat: str) -> PatternFill:
    c = (cat or "").lower()
    if c == "net":       return FILL_NET
    if c == "freestyle": return FILL_FREE
    if c in ("golf", "sideline"): return FILL_GOLF
    return FILL_OTHER


# ── Canonical ordering for year/event sheet layout ─────────────────────────────

_CAT_ORDER: dict[str, int] = {
    "OVERALL":     0,
    "CONSECUTIVE": 1,
    "NET":         2,
    "FREESTYLE":   3,
    "GOLF":        4,
    "DISTANCE":    5,
    "ACCURACY":    6,
    "SIDELINE":    7,
    "UNKNOWN":     8,
}

_DIV_ORDER: dict[str, dict[str, int]] = {
    "OVERALL": {
        "Men's Overall":              0,
        "Open Overall":               1,
        "Women's Overall":            2,
        "Freestyle Overall":          3,
        "Intermediate Overall":       4,
        "Novice Overall":             5,
    },
    "CONSECUTIVE": {
        "Open Singles Consecutive":       0,
        "Open Doubles Consecutive":       1,
        "Women's Singles Consecutive":    2,
        "Women's Doubles Consecutive":    3,
        "Intermediate Singles Consecutive": 4,
        "Singles Consecutive":            5,
        "Doubles Consecutive":            6,
    },
    "NET": {
        "Open Singles Net":           0,
        "Open Doubles Net":           1,
        "Open Mixed Doubles Net":     2,
        "Mixed Doubles Net":          3,
        "Women's Singles Net":        4,
        "Women's Doubles Net":        5,
        "Intermediate Singles Net":   6,
        "Intermediate Doubles Net":   7,
    },
    "FREESTYLE": {
        "Open Singles Freestyle":         0,
        "Open Singles Routines":          1,
        "Women's Singles Freestyle":      2,
        "Women's Singles Routines":       3,
        "Intermediate Singles Freestyle": 4,
        "Open Doubles Freestyle":         5,
        "Open Team Freestyle":            6,
        "Mixed Doubles Freestyle":        7,
    },
    "GOLF": {
        "Open Golf":                  0,
        "Open Singles Golf":          1,
        "Open Doubles Golf":          2,
        "Women's Golf":               3,
        "Intermediate Golf":          4,
        "Novice Golf":                5,
    },
}


def _disc_sort_key(disc_name: str, disc_rec: dict) -> tuple:
    """Sort key: canonical category order → canonical division order → alpha fallback."""
    cat = disc_rec.get("category_canonical") or ""
    div = disc_rec.get("division_canonical") or disc_name
    cat_idx = _CAT_ORDER.get(cat, 99)
    div_idx = _DIV_ORDER.get(cat, {}).get(div, 999)
    return (cat_idx, div_idx, div.lower(), disc_name.lower())


def _dedup_slot(slot_rows: list) -> list:
    """Deduplicate participant rows by person_id, preserving order.
    Rows with no person_id (unknowns) are kept once via object identity."""
    seen: set = set()
    result: list = []
    for row in slot_rows:
        pid = row.get("person_id", "").strip()
        key = pid if pid else id(row)
        if key not in seen:
            seen.add(key)
            result.append(row)
    return result


# ── Ranking helper ─────────────────────────────────────────────────────────────

def _add_ranks(rows: list, sort_key_idx: int) -> list:
    """
    Prepend a rank string to each row. Tied values get T-N notation.
    Input rows are assumed already sorted descending by sort_key_idx.
    Returns list of tuples: (rank_str, *original_row_values).
    """
    if not rows:
        return []
    # Count occurrences of each value to detect ties
    from collections import Counter as C
    val_count: dict = C(r[sort_key_idx] for r in rows)

    result = []
    rank = 1
    prev_val = None
    run_start = 1
    for i, row in enumerate(rows, 1):
        val = row[sort_key_idx]
        if val != prev_val:
            run_start = rank
        rank_str = f"T-{run_start}" if val_count[val] > 1 else str(run_start)
        result.append((rank_str,) + tuple(row))
        if val != prev_val:
            prev_val = val
        # Advance rank counter only when value changes on next row
        if i < len(rows) and rows[i][sort_key_idx] != val:
            rank = i + 1

    return result


# ── Data loading ───────────────────────────────────────────────────────────────

def load_all():
    print("Loading canonical_all CSVs…")

    raw_events  = _load("events.csv")
    raw_discs   = _load("event_disciplines.csv")
    raw_results = _load("event_result_participants.csv")
    raw_persons = _load("persons.csv")

    events:  dict[str, dict] = {r["event_id"]: r for r in raw_events}
    discs:   dict[tuple, dict] = {}
    for r in raw_discs:
        discs[(r["event_id"], r["discipline"])] = r
    persons: dict[str, dict] = {r["person_id"]: r for r in raw_persons}

    quarantine: set[str] = set()
    if QUARANTINE_CSV.exists():
        with open(QUARANTINE_CSV, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                eid = r.get("event_id", "").strip()
                if eid:
                    quarantine.add(eid)

    print(f"  Events: {len(events)}   Disciplines: {len(discs)}")
    print(f"  Participants: {len(raw_results)}   Persons: {len(persons)}")
    print(f"  Quarantined events: {len(quarantine)}")
    return events, discs, raw_results, persons, quarantine


# ── Statistics engine ──────────────────────────────────────────────────────────

def compute_stats(raw_results, events, discs, persons):
    stats: dict[str, dict] = defaultdict(lambda: {
        "wins": 0, "p1": 0, "p2": 0, "p3": 0, "podiums": 0,
        "events": set(), "years": set(),
        "worlds_wins": 0, "worlds_podiums": 0, "worlds_events": set(),
        "cat_wins":    defaultdict(int),
        "cat_podiums": defaultdict(int),
    })

    worlds_types = {"WFA_WORLD_CHAMPIONSHIPS", "WORLD_CHAMPIONSHIPS",
                    "IFAB_WORLD_CHAMPIONSHIPS", "NHSA_NATIONALS"}

    for row in raw_results:
        pid = row.get("person_id", "").strip()
        if not pid or pid in _SKIP_PID:
            continue
        dname = row.get("display_name", "").strip()
        if dname in _SKIP_DNAME:
            continue
        porder = row.get("participant_order", "1").strip()

        eid   = row.get("event_id", "").strip()
        ev    = events.get(eid, {})
        year  = ev.get("year", "").strip()
        etype = ev.get("event_type", "").strip()
        disc  = discs.get((eid, row.get("discipline", "").strip()), {})
        cat   = disc.get("discipline_category", "").lower().strip()

        try:
            place = int(row.get("placement", "0") or 0)
        except ValueError:
            place = 0

        s = stats[pid]
        if eid:
            s["events"].add(eid)
        if year:
            try:
                s["years"].add(int(year))
            except ValueError:
                pass

        if porder == "1":
            if place == 1:
                s["wins"] += 1
                s["p1"]   += 1
                s["cat_wins"][cat] += 1
            if 1 <= place <= 3:
                s["podiums"] += 1
                s["cat_podiums"][cat] += 1
            if place == 2: s["p2"] += 1
            if place == 3: s["p3"] += 1

            if etype in worlds_types:
                if eid:
                    s["worlds_events"].add(eid)
                if place == 1:
                    s["worlds_wins"] += 1
                if 1 <= place <= 3:
                    s["worlds_podiums"] += 1

    result = {}
    for pid, s in stats.items():
        if pid not in persons:
            continue
        result[pid] = {
            "wins":           s["wins"],
            "p1": s["p1"], "p2": s["p2"], "p3": s["p3"],
            "podiums":        s["podiums"],
            "events":         len(s["events"]),
            "year_first":     min(s["years"]) if s["years"] else None,
            "year_last":      max(s["years"])  if s["years"] else None,
            "worlds_wins":    s["worlds_wins"],
            "worlds_podiums": s["worlds_podiums"],
            "worlds_events":  len(s["worlds_events"]),
            "cat_wins":       dict(s["cat_wins"]),
            "cat_podiums":    dict(s["cat_podiums"]),
        }
    return result


# ── README sheet ──────────────────────────────────────────────────────────────

def build_readme(wb: Workbook, events: dict, persons: dict, n_parts: int) -> None:
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 90
    ws.freeze_panes = "B2"

    n_years   = len(set(e.get("year","") for e in events.values() if e.get("year")))
    n_events  = len(events)
    n_persons = sum(1 for p in persons.values()
                    if p.get("person_id") and p.get("person_canon"))
    n_with_results = sum(1 for e in events.values() if e.get("status","") != "no_results")

    row = _title_row(ws, 1,
        f"FOOTBAG COMPETITION RESULTS — COMMUNITY DATASET  "
        f"({VERSION} · Updated {UPDATED})", ncols=1)
    row += 1

    sections = [
        ("OVERVIEW", [
            ("Coverage",
             f"{n_with_results} events with results across {n_years} years · "
             f"{n_persons} identified competitors · {n_parts:,} placement records"),
            ("Years",    "1980 – present (all years included, no year suppression)"),
            ("Sources",  "Post-1997: Footbag.org HTML archive (mirror-derived, highest authority).  "
                         "Pre-1997: Footbag World magazine scans + oldresults.txt + "
                         "expert corrections from authoritative human sources."),
            ("Identity", "All statistics aggregated by person_id (UUID).  "
                         "No display-name matching — each person counted once regardless of "
                         "name variants, nicknames, or married-name changes.  "
                         "Canonical names sourced from persons.csv."),
            ("Divisions","Fully normalized (no Sgls/Dbls abbreviations).  "
                         "Categories: freestyle · net · golf · sideline."),
            ("Version",  f"{VERSION} — {UPDATED}"),
        ]),
        ("SHEETS", [
            ("README",          "This sheet — dataset overview, notes, known issues."),
            ("STATISTICS",      "Career leaderboards: top 25 per category with rank and tie notation."),
            ("ERA LEADERS",     "Decade leaderboards (1980s – 2020s): top 10 podiums and wins per era."),
            ("PLAYER STATS",    "One row per identified competitor — career summary."),
            ("EVENT INDEX",     "All events with metadata, coverage level, and links to year sheets."),
            ("EXCLUDED EVENTS", "Events in the archive without usable results (no results published "
                                "or data under review)."),
            ("<year>",          "One sheet per year (1980 – 2026).  Each column is one event; "
                                "rows show placements by division."),
        ]),
        ("COVERAGE LIMITATIONS", [
            ("Pre-1997",        "Data reconstructed from magazine scans and text archives.  "
                                "Coverage is incomplete: some events have only partial results, "
                                "and some early events are not represented at all.  "
                                "Coverage level noted in EVENT INDEX column 'Coverage'."),
            ("Post-1997",       "Mirror-derived results are the highest-authority source.  "
                                "Most post-1997 events are complete.  Some events have partial "
                                "coverage where the source page was incomplete at time of archiving."),
            ("NHSA vs WFA",     "1980–1985 featured distinct NHSA and WFA championships; "
                                "both are represented as separate events where data permits."),
            ("Sparse events",   "Events flagged SPARSE in EVENT INDEX have fewer than 3 divisions "
                                "or fewer than 10 placements — coverage may be incomplete."),
        ]),
        ("IDENTITY MODEL", [
            ("person_id",       "Every competitor is assigned a UUID (person_id).  "
                                "Statistics count each person_id once per placement slot."),
            ("Unknowns",        "Unresolved participants display as [Unknown] in year sheets.  "
                                "~750 genuine unknowns exist in the dataset (historical records "
                                "with no name or unresolvable identities).  "
                                "These are excluded from all leaderboards and statistics."),
            ("Sentinels",       "Team slots with a known partner but unknown second member "
                                "are marked [UNKNOWN PARTNER].  Also excluded from statistics."),
            ("Name variants",   "Nickname and married-name variants are resolved to a single "
                                "canonical name via the persons table.  No duplicates."),
        ]),
        ("STATISTICS NOTES", [
            ("Worlds types",    "Worlds leaderboards include: WORLD_CHAMPIONSHIPS, "
                                "WFA_WORLD_CHAMPIONSHIPS, IFAB_WORLD_CHAMPIONSHIPS, NHSA_NATIONALS."),
            ("Counting method", "Wins and podiums counted once per placement slot "
                                "(participant_order=1).  Team events: both members credited "
                                "equally; the slot is counted once in aggregate tables."),
            ("Top 25",          "Leaderboard tables show top 25.  Full data available "
                                "in PLAYER STATS sheet and canonical CSV files."),
            ("Ties",            "Equal-ranked positions use T-N notation (e.g. T-5).  "
                                "After a tie of N, the next rank skips N-1 positions."),
        ]),
        ("KNOWN ISSUES", [
            ("Pre-1997 gaps",   "Not all pre-1997 events are fully captured.  "
                                "Some divisions and placements below 3rd may be missing."),
            ("1984/1985 events","1984_worlds and 1984_wfa_nationals appear to represent "
                                "the same championship from two source records; "
                                "similarly for 1985_worlds and 1985_wfa_nationals.  "
                                "These are currently kept as separate events pending "
                                "authoritative confirmation.  Stats may double-count "
                                "placements at these events."),
            ("Unresolved IDs",  "~750 participants have no resolved person_id.  "
                                "These do not affect statistics (excluded by design) "
                                "but represent genuine historical uncertainty."),
            ("Source conflicts", "Where FBW and OLD_RESULTS sources disagree, "
                                 "OLD_RESULTS (authoritative human records) takes precedence.  "
                                 "Known cross-event contamination in 1982–1983 NHSA/WFA data "
                                 "has been cleaned; other early events may have similar issues."),
            ("Quarantined",     "9 events are quarantined (complex or uncertain data).  "
                                "They appear in year sheets with ⛔ marker but are "
                                "included in statistics with the caveat that results may be incomplete."),
        ]),
    ]

    for section_name, items in sections:
        row = _section_row(ws, row, section_name, ncols=1)
        for label, text in items:
            _w(ws, row, 1,
               f"▸ {label}:  {text}" if label else text,
               font=FONT_DATA, align=ALIGN_LW)
            ws.row_dimensions[row].height = max(16, min(60, len(text) // 6))
            row += 1
        row += 1

    ws.sheet_view.showGridLines = False
    print("  README done")


# ── DATA NOTES sheet ─────────────────────────────────────────────────────────

def build_data_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("DATA NOTES")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80
    ws.sheet_view.showGridLines = False

    _H  = Font(bold=True, size=12)
    _SH = Font(bold=True, size=10)
    _B  = Font(bold=True, size=10)
    _N  = Font(size=10)
    _SM = Font(size=9, italic=True, color="606060")

    NOTES: list[tuple] = [
        # (type, label, text)
        ("title",  "DATA NOTES",        ""),
        ("spacer", "", ""),

        ("heading", "1. Worlds Classification", ""),
        ("note",
         "Early tournaments (1980–1983)",
         "The NHSA and WFA held competitions widely recognized as de-facto World "
         "Championships starting with the 1980 NHSA event. These are classified as "
         "RETROACTIVE_WORLD_CHAMPIONSHIPS in the dataset. Beginning in 1984, the "
         "WFA formalized the World Footbag Championships in Golden, CO; these and "
         "all subsequent events are classified as OFFICIAL_WORLD_CHAMPIONSHIPS."),
        ("note",
         "NHSA / WFA / IFAB distinctions",
         "Original organizational names (NHSA, WFA, IFAB) are preserved in "
         "event_name and event_type. The worlds_classification field provides a "
         "unified classification across all governing bodies."),
        ("spacer", "", ""),

        ("heading", "2. Net Ruleset Variants (Ultra / Advanced)", ""),
        ("note",
         "Ultra and Advanced",
         "In the 1980s, Footbag Net was contested under two rule variants: "
         "'Ultra' (highest-level open competition) and 'Advanced' (intermediate-level). "
         "These are NOT separate divisions — they are rule variants of Net. "
         "In this dataset, Ultra and Advanced divisions are canonicalized to their "
         "standard Net equivalents (e.g., Open Singles Net, Open Doubles Net) and "
         "the variant is recorded in the 'ruleset' field (ULTRA / ADVANCED). "
         "The original division name is always preserved in 'division_raw'."),
        ("spacer", "", ""),

        ("heading", "3. Early Event Locations (1983–1984)", ""),
        ("note",
         "Year-level location assignment",
         "For 1983 and 1984, all major tournaments are assigned a single host "
         "location per year based on consolidated historical sources: "
         "Portland, Oregon (1983) and Boulder, Colorado (1984). "
         "This ensures consistency across overlapping NHSA/WFA/World Championship "
         "classifications for those years where the same host city served all events."),
        ("spacer", "", ""),

        ("heading", "4. Coverage Policy", ""),
        ("note",
         "SPARSE events excluded",
         "Events with fewer than 10 placements AND fewer than 2 disciplines are "
         "classified as SPARSE and excluded from the EVENT INDEX and year sheets. "
         "They remain in the full canonical dataset and are listed in the "
         "'QC - EXCLUDED EVENTS' sheet for reference."),
        ("note",
         "Coverage levels",
         "FULL = ≥20 placements and ≥3 disciplines (or multi-source confirmed). "
         "PARTIAL = ≥10 placements or ≥2 disciplines. "
         "SPARSE = any results but below PARTIAL threshold. "
         "NO RESULTS = no placement data available."),
        ("spacer", "", ""),

        ("heading", "5. Known Limitations", ""),
        ("note",
         "Pre-1990 data",
         "Results before 1990 are reconstructed from magazine scans (Footbag World) "
         "and secondary sources. Coverage is partial; dates and locations may be "
         "approximate. Original source references are preserved in the notes fields."),
        ("note",
         "Unresolved identities",
         "Some competitor names could not be matched to a canonical person record "
         "(single names, initials, ambiguous spelling). These appear with blank "
         "person_id in the dataset and are excluded from player statistics."),
        ("note",
         "Limited source availability pre-1990",
         "The historical record for events before 1990 is incomplete. Some events "
         "have only top-3 or top-5 results; others have only the winner recorded. "
         "Ongoing recovery efforts continue to improve early coverage."),
    ]

    row = 1
    for kind, label, text in NOTES:
        if kind == "title":
            cell = ws.cell(row, 1, label)
            cell.font = Font(bold=True, size=14)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
        elif kind == "spacer":
            row += 1
        elif kind == "heading":
            ws.cell(row, 1, label).font = _H
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row, 1).fill = PatternFill("solid", fgColor="E8EEF4")
            row += 1
        elif kind == "note":
            lc = ws.cell(row, 1, label)
            lc.font = _B
            lc.alignment = Alignment(vertical="top", wrap_text=True)
            tc = ws.cell(row, 2, text)
            tc.font = _N
            tc.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = max(15, len(text) // 4)
            row += 1

    print("  DATA NOTES done")


# ── STATISTICS sheet ──────────────────────────────────────────────────────────

def build_statistics(wb: Workbook, stats: dict, persons: dict) -> None:
    ws = wb.create_sheet("STATISTICS")
    ws.column_dimensions["A"].width = 6   # Rank
    ws.column_dimensions["B"].width = 34  # Player
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "B2"

    row = _title_row(ws, 1, "STATISTICS — ALL EVENTS ALL DIVISIONS", ncols=6)

    def canon(pid):
        return persons.get(pid, {}).get("person_canon", pid)

    TOP = 25

    def table(title, headers, raw_data, start_row,
              sort_key_idx=-1, is_worlds=False):
        """Write a ranked stats table. sort_key_idx indexes into raw_data tuples."""
        ncols = len(headers)
        if is_worlds:
            r = _worlds_section_row(ws, start_row, title, ncols=ncols)
        else:
            r = _section_row(ws, start_row, title, ncols=ncols)

        # Add note about top-N limit
        _note_row(ws, r, f"  Top {TOP} shown · full data in PLAYER STATS sheet")
        r += 1

        r = _hrow(ws, r, *headers)
        ranked = _add_ranks(raw_data[:TOP], sort_key_idx)
        for i, row_vals in enumerate(ranked):
            fill = FILL_NET if i % 2 == 0 else FILL_EVENT_B
            r = _drow(ws, r, *row_vals, fill=fill)
        return r + 1

    # ── General Stats ─────────────────────────────────────────────────────────
    row = _section_row(ws, row, "GENERAL STATISTICS", ncols=6)
    row += 1

    # 1 — Career Podiums
    podium_data = sorted(
        [(canon(pid), s["p1"], s["p2"], s["p3"], s["podiums"])
         for pid, s in stats.items() if s["podiums"] > 0],
        key=lambda x: (-x[4], x[0].lower())
    )
    row = table("MOST CAREER PODIUMS (ALL DIVISIONS)",
                ["Rank", "Player", "1st", "2nd", "3rd", "Total"],
                podium_data, row, sort_key_idx=4)

    # 2 — Wins
    wins_data = sorted(
        [(canon(pid), s["wins"]) for pid, s in stats.items() if s["wins"] > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST CAREER WINS (ALL DIVISIONS)",
                ["Rank", "Player", "Wins"],
                wins_data, row, sort_key_idx=1)

    # 3 — Freestyle Podiums
    free_pod = sorted(
        [(canon(pid), s["cat_podiums"].get("freestyle", 0))
         for pid, s in stats.items() if s["cat_podiums"].get("freestyle", 0) > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST FREESTYLE PODIUMS",
                ["Rank", "Player", "Podiums"],
                free_pod, row, sort_key_idx=1)

    # 4 — Net Podiums
    net_pod = sorted(
        [(canon(pid), s["cat_podiums"].get("net", 0))
         for pid, s in stats.items() if s["cat_podiums"].get("net", 0) > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST NET PODIUMS",
                ["Rank", "Player", "Podiums"],
                net_pod, row, sort_key_idx=1)

    # 5 — Events Competed
    events_data = sorted(
        [(canon(pid), s["events"]) for pid, s in stats.items() if s["events"] > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST EVENTS COMPETED",
                ["Rank", "Player", "Events"],
                events_data, row, sort_key_idx=1)

    # 6 — Career Spans
    career_data = sorted(
        [(canon(pid), s["year_first"], s["year_last"],
          (s["year_last"] - s["year_first"]) if s["year_first"] and s["year_last"] else 0)
         for pid, s in stats.items()
         if s["year_first"] and s["year_last"] and s["year_last"] > s["year_first"]],
        key=lambda x: (-x[3], x[0].lower())
    )
    row = table("LONGEST COMPETITIVE CAREERS",
                ["Rank", "Player", "First Year", "Last Year", "Span (yrs)"],
                career_data, row, sort_key_idx=3)

    # ── World Championships ────────────────────────────────────────────────────
    row += 1
    row = _worlds_section_row(ws, row,
        "═══  WORLD CHAMPIONSHIPS  ═══  "
        "(WORLD_CHAMPIONSHIPS · WFA_WORLD_CHAMPIONSHIPS · "
        "IFAB_WORLD_CHAMPIONSHIPS · NHSA_NATIONALS)",
        ncols=6)
    row += 1

    # 7 — Worlds Podiums
    worlds_pod = sorted(
        [(canon(pid), s["worlds_wins"], s["worlds_podiums"], s["worlds_events"])
         for pid, s in stats.items() if s["worlds_podiums"] > 0],
        key=lambda x: (-x[2], x[0].lower())
    )
    row = table("WORLDS PODIUMS",
                ["Rank", "Player", "Worlds Wins", "Worlds Podiums", "Worlds Events"],
                worlds_pod, row, sort_key_idx=2, is_worlds=True)

    # 8 — Worlds Wins
    worlds_wins = sorted(
        [(canon(pid), s["worlds_wins"])
         for pid, s in stats.items() if s["worlds_wins"] > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("WORLDS WINS",
                ["Rank", "Player", "Wins"],
                worlds_wins, row, sort_key_idx=1, is_worlds=True)

    print("  STATISTICS done")


# ── ERA LEADERS sheet ──────────────────────────────────────────────────────────

def build_era_leaders(wb: Workbook, raw_results, events, discs, persons) -> None:
    ws = wb.create_sheet("ERA LEADERS")
    ws.column_dimensions["A"].width = 6   # Rank
    ws.column_dimensions["B"].width = 34  # Player
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "B2"

    row = _title_row(ws, 1, "ERA LEADERS — BY DECADE", ncols=6)

    _ERAS = [
        ("1980s", 1980, 1989),
        ("1990s", 1990, 1999),
        ("2000s", 2000, 2009),
        ("2010s", 2010, 2019),
        ("2020s", 2020, 2029),
    ]
    TOP = 10

    def canon(pid):
        return persons.get(pid, {}).get("person_canon", pid)

    for era_name, yr_lo, yr_hi in _ERAS:
        era_podiums: dict[str, list] = defaultdict(lambda: [0, 0, 0])
        era_events:  dict[str, set]  = defaultdict(set)

        for row_r in raw_results:
            pid = row_r.get("person_id", "").strip()
            if not pid or pid in _SKIP_PID:
                continue
            if row_r.get("display_name", "").strip() in _SKIP_DNAME:
                continue
            if row_r.get("participant_order", "1").strip() != "1":
                continue

            eid = row_r.get("event_id", "").strip()
            ev  = events.get(eid, {})
            try:
                yr = int(ev.get("year", "0") or 0)
            except ValueError:
                yr = 0
            if not (yr_lo <= yr <= yr_hi):
                continue

            try:
                place = int(row_r.get("placement", "0") or 0)
            except ValueError:
                place = 0

            if 1 <= place <= 3:
                era_podiums[pid][place - 1] += 1
            if eid:
                era_events[pid].add(eid)

        row = _section_row(ws, row, f"ERA: {era_name}  ({yr_lo}–{yr_hi})", ncols=6)

        # Top Podiums
        _w(ws, row, 1, "Top Podiums", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_L)
        for c in range(2, 7):
            ws.cell(row=row, column=c).fill = FILL_HEADER
        row += 1
        row = _hrow(ws, row, "Rank", "Player", "1st", "2nd", "3rd", "Total")

        pod_rows = sorted(
            [(canon(pid), p[0], p[1], p[2], sum(p))
             for pid, p in era_podiums.items() if pid in persons and sum(p) > 0],
            key=lambda x: (-x[4], x[0].lower())
        )
        ranked_pod = _add_ranks(pod_rows[:TOP], sort_key_idx=4)
        for i, rv in enumerate(ranked_pod):
            fill = FILL_NET if i % 2 == 0 else FILL_EVENT_B
            row = _drow(ws, row, *rv, fill=fill)

        row += 1

        # Top Wins
        _w(ws, row, 1, "Top Wins", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_L)
        for c in range(2, 4):
            ws.cell(row=row, column=c).fill = FILL_HEADER
        row += 1
        row = _hrow(ws, row, "Rank", "Player", "Wins")

        win_rows = sorted(
            [(canon(pid), p[0]) for pid, p in era_podiums.items()
             if pid in persons and p[0] > 0],
            key=lambda x: (-x[1], x[0].lower())
        )
        ranked_win = _add_ranks(win_rows[:TOP], sort_key_idx=1)
        for i, rv in enumerate(ranked_win):
            fill = FILL_NET if i % 2 == 0 else FILL_EVENT_B
            row = _drow(ws, row, *rv, fill=fill)

        row += 2

    print("  ERA LEADERS done")


# ── PLAYER STATS sheet ─────────────────────────────────────────────────────────

def build_player_stats(wb: Workbook, stats: dict, persons: dict) -> None:
    ws = wb.create_sheet("PLAYER STATS")

    headers = ["Player", "Nickname", "Country", "Events", "Wins", "Podiums", "BAP", "FBHOF"]
    widths  = [34, 22, 18, 9, 9, 9, 6, 6]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"

    row = _title_row(ws, 1, "PLAYER STATS — ALL IDENTIFIED COMPETITORS", ncols=len(headers))
    row = _hrow(ws, row, *headers)

    rows_out = []
    for pid, p in sorted(persons.items(),
                         key=lambda kv: kv[1].get("person_canon", "").lower()):
        if not p.get("person_canon"):
            continue
        s = stats.get(pid, {})
        bap  = "Y" if p.get("bap_member",  "N") == "Y" else ""
        hof  = "Y" if p.get("fbhof_member","N") == "Y" else ""
        fill = FILL_BOTH if bap and hof else (FILL_HOF if hof else (FILL_BAP if bap else None))
        rows_out.append((
            p["person_canon"],
            p.get("bap_nickname", ""),
            p.get("country", ""),
            s.get("events",   0),
            s.get("wins",     0),
            s.get("podiums",  0),
            bap, hof,
            fill,
        ))

    for rec in rows_out:
        *vals, fill = rec
        r = row
        for col, v in enumerate(vals, 1):
            cell = _w(ws, r, col, v, font=FONT_DATA,
                      align=(ALIGN_R if isinstance(v, int) else ALIGN_L))
            if fill:
                cell.fill = fill
        row += 1

    print(f"  PLAYER STATS: {len(rows_out)} rows")


# ── EVENT INDEX sheet ──────────────────────────────────────────────────────────

def _coverage_level(ev: dict, n_discs: int, n_plc: int, is_quar: bool) -> str:
    """Map event metadata to a human-readable coverage level."""
    if is_quar:
        return "QUARANTINED"
    if ev.get("status","") == "no_results":
        return "NO RESULTS"
    vs = ev.get("validation_status","")
    src = ev.get("data_source","").upper()
    if vs == "CONFIRMED_MULTI_SOURCE":
        return "FULL"
    if vs == "VERIFIED":
        return "FULL"
    if n_plc >= 20 and n_discs >= 3:
        return "FULL"
    if n_plc >= 10 or n_discs >= 2:
        return "PARTIAL"
    if n_plc > 0:
        return "SPARSE"
    return "NO RESULTS"


def build_event_index(wb: Workbook, events: dict, discs: dict,
                      raw_results, quarantine: set,
                      event_col_map: dict, pub_eids: set) -> None:
    ws = wb.create_sheet("EVENT INDEX")

    headers = ["Year", "Event Name", "Location", "Event Type",
               "Disciplines", "Placements", "Coverage", "Source", "Notes"]
    widths  = [6, 52, 30, 26, 12, 11, 12, 14, 20]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"   # freeze row 1 AND column A (Year)

    row = _title_row(ws, 1, "EVENT INDEX — EVENTS WITH RESULTS", ncols=len(headers))
    row = _hrow(ws, row, *headers)

    placements_per_event: dict[str, int] = Counter(
        r["event_id"] for r in raw_results
        if r.get("person_id", "").strip() not in _SKIP_PID
    )
    discs_per_event: dict[str, int] = Counter(k[0] for k in discs)

    sorted_events = sorted(events.values(),
                           key=lambda e: (e.get("year",""),
                                          e.get("start_date",""),
                                          e.get("event_name","")))

    _COV_FILL = {
        "FULL":        None,
        "PARTIAL":     FILL_WARN,
        "SPARSE":      FILL_OTHER,
        "QUARANTINED": FILL_QUAR,
        "NO RESULTS":  FILL_OTHER,
    }

    shown = 0
    for ev in sorted_events:
        eid     = ev["event_id"]
        status  = ev.get("status", "")
        is_quar = eid in quarantine
        n_discs = discs_per_event.get(eid, 0)
        n_plc   = placements_per_event.get(eid, 0)
        src     = ev.get("source_types", ev.get("data_source", ""))
        cov     = _coverage_level(ev, n_discs, n_plc, is_quar)

        # Skip events not in publication whitelist (SPARSE / NO RESULTS excluded)
        if eid not in pub_eids:
            continue

        notes   = "⛔ under review" if is_quar else ""

        wc = ev.get("worlds_classification", "")
        is_worlds = bool(wc) or ev.get("event_type", "").lower() == "worlds"
        if is_worlds:
            row_fill = FILL_WORLDS_ROW
        else:
            row_fill = _COV_FILL.get(cov)

        # Build event name cell — with hyperlink to year sheet if available
        name_val = ev.get("event_name", eid)
        year_str = ev.get("year", "")
        sheet_ref = event_col_map.get(eid)

        for col_idx, v in enumerate([
            year_str,
            name_val,
            ev.get("location", ""),
            ev.get("event_type",""),
            n_discs or "",
            n_plc   or "",
            cov,
            src,
            notes,
        ], 1):
            font = FONT_DATA
            # Hyperlink style on event name column if year sheet exists
            if col_idx == 2 and sheet_ref:
                font = FONT_LINK
            cell = _w(ws, row, col_idx, v, font=font, align=ALIGN_L)
            if row_fill:
                cell.fill = row_fill
            # Apply hyperlink to event name cell
            if col_idx == 2 and sheet_ref:
                yr_sheet, col_letter = sheet_ref
                safe = yr_sheet.replace("'", "''")
                try:
                    cell.hyperlink = f"#{safe}!{col_letter}1"
                    cell.style = "Hyperlink"
                except Exception:
                    cell.font = FONT_LINK

        row += 1
        shown += 1

    # Coverage level legend at bottom
    row += 1
    _w(ws, row, 1, "Coverage levels:", font=FONT_HEADER, align=ALIGN_L)
    _w(ws, row, 2, "Events with SPARSE or NO RESULTS are in QC - EXCLUDED EVENTS sheet.", font=FONT_NOTE, align=ALIGN_L)
    row += 1
    for level, desc, fill_key in [
        ("FULL",        "≥20 placements and ≥3 divisions, or multi-source confirmed", "FULL"),
        ("PARTIAL",     "Some results available but not fully complete",               "PARTIAL"),
        ("QUARANTINED", "Data under review — results may be uncertain",                "QUARANTINED"),
        ("🌐 WORLDS",   "World Footbag Championship event",                            "_WORLDS"),
    ]:
        f = _COV_FILL.get(fill_key) or (FILL_WORLDS_ROW if fill_key == "_WORLDS" else FILL_NONE)
        _w(ws, row, 1, f"  {level}", font=FONT_DATA, fill=f, align=ALIGN_L)
        _w(ws, row, 2, desc, font=FONT_DATA, align=ALIGN_L)
        row += 1

    print(f"  EVENT INDEX: {shown} events (with results)")


# ── EXCLUDED EVENTS sheet ──────────────────────────────────────────────────────

def build_excluded_events(wb: Workbook, events: dict, quarantine: set,
                          discs: dict, raw_results) -> None:
    ws = wb.create_sheet("QC - EXCLUDED EVENTS")

    headers = ["Year", "Event ID", "Event Name", "Location", "Coverage", "Discs", "Placements", "Reason"]
    widths  = [6, 28, 52, 30, 13, 7, 11, 36]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"

    row = _title_row(ws, 1, "QC — EVENTS EXCLUDED FROM EVENT INDEX", ncols=len(headers))
    row = _hrow(ws, row, *headers)

    placements_per_event: dict[str, int] = Counter(
        r["event_id"] for r in raw_results
        if r.get("person_id", "").strip() not in _SKIP_PID
    )
    discs_per_event: dict[str, int] = Counter(k[0] for k in discs)

    _COV_FILL = {
        "SPARSE":      FILL_OTHER,
        "QUARANTINED": FILL_QUAR,
        "NO RESULTS":  FILL_OTHER,
    }

    excluded = []
    for eid, ev in events.items():
        status  = ev.get("status", "")
        is_quar = eid in quarantine
        n_discs = discs_per_event.get(eid, 0)
        n_plc   = placements_per_event.get(eid, 0)
        cov     = _coverage_level(ev, n_discs, n_plc, is_quar)

        if cov == "NO RESULTS":
            reason = "no results in database"
        elif cov == "SPARSE":
            reason = f"sparse coverage ({n_plc} placements, {n_discs} discipline(s))"
        elif is_quar:
            reason = "under review (complex or uncertain data)"
        else:
            continue

        excluded.append((ev.get("year",""), eid, ev.get("event_name",""),
                         ev.get("location",""), cov, n_discs or "", n_plc or "", reason))

    excluded.sort(key=lambda x: (x[0], x[2].lower()))
    for rec in excluded:
        cov_val  = rec[4]
        row_fill = _COV_FILL.get(cov_val)
        for col, v in enumerate(rec, 1):
            cell = _w(ws, row, col, v, font=FONT_DATA, align=ALIGN_L)
            if row_fill:
                cell.fill = row_fill
        row += 1

    print(f"  QC - EXCLUDED EVENTS: {len(excluded)} entries")


# ── Year sheets ────────────────────────────────────────────────────────────────

def _team_display(slot_rows: list[dict]) -> tuple[str, bool]:
    """
    Combine participant rows into a display string.
    Returns (display_str, is_unknown) where is_unknown=True if all names are unknown.
    """
    by_order: dict[int, str] = {}
    for r in slot_rows:
        try:
            order = int(r.get("participant_order", "1") or 1)
        except ValueError:
            order = 1
        name = r.get("display_name", "").strip()
        if name and name not in _SKIP_DNAME:
            by_order[order] = name
        elif not by_order.get(order):
            by_order[order] = _UNKNOWN_DISPLAY

    if not by_order:
        return (_UNKNOWN_DISPLAY, True)

    parts = [by_order.get(1, _UNKNOWN_DISPLAY)]
    if 2 in by_order:
        parts.append(by_order[2])

    parts = [p for p in parts if p]
    disp  = " / ".join(parts) if parts else _UNKNOWN_DISPLAY
    is_unk = all(p == _UNKNOWN_DISPLAY for p in parts)
    return (disp, is_unk)


def build_year_sheets(wb: Workbook, raw_results, events: dict,
                      discs: dict, persons: dict,
                      quarantine: set, pub_eids: set) -> dict:
    """Build one sheet per year. Returns event_col_map = {event_id: (sheet_name, col_letter)}.

    Placements are keyed by division_canonical so all raw disc_names that share the same
    canonical are merged into one row per event.  Participants are deduplicated by person_id.
    """

    # plcmt_by_event[eid][div_can][place] = merged list of participant rows
    plcmt_by_event: dict[str, dict] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )
    # div_can_rec[(eid, div_can)] = first disc record seen (for sort key + fill colour)
    div_can_rec: dict[tuple, dict] = {}
    # merge accounting
    _raw_discs_per_can: dict[tuple, set] = defaultdict(set)

    for r in raw_results:
        eid    = r["event_id"]
        disc   = r["discipline"]
        rec    = discs.get((eid, disc), {})
        div_can = rec.get("division_canonical") or disc
        try:
            place = int(r.get("placement", "0") or 0)
        except ValueError:
            place = 0
        plcmt_by_event[eid][div_can][place].append(r)
        _raw_discs_per_can[(eid, div_can)].add(disc)
        if (eid, div_can) not in div_can_rec:
            div_can_rec[(eid, div_can)] = rec

    # Report merge stats
    _dup_groups = sum(1 for v in _raw_discs_per_can.values() if len(v) > 1)
    _rows_merged = sum(len(v) - 1 for v in _raw_discs_per_can.values() if len(v) > 1)
    if _dup_groups:
        print(f"  Canonical merge: {_dup_groups} duplicate groups, "
              f"{_rows_merged} raw discipline rows consolidated")

    year_to_eids: dict[str, list] = defaultdict(list)
    for eid, ev in events.items():
        yr = ev.get("year", "").strip()
        if yr and yr.isdigit():
            # Use publication whitelist (same filter as release_publication):
            # includes FULL, PARTIAL, QUARANTINED; excludes SPARSE and NO RESULTS.
            if eid in pub_eids:
                year_to_eids[yr].append(eid)

    event_col_map: dict[str, tuple] = {}

    _R_NAME  = 1
    _R_LOC   = 2
    _R_DATE  = 3
    _R_TYPE  = 4
    _R_EID   = 5
    _R_DATA  = 7

    _YF_META = Font(bold=True, size=11)
    _YF_EID  = Font(size=8, color="808080")
    _YF_DIV  = Font(bold=True, size=10)
    _YF_PLC  = Font(size=10)
    _YF_UNK  = Font(size=10, italic=True, color="A0A0A0")
    _YF_QUAR = Font(bold=True, size=11, color="CC0000")

    _COL_A_W = 14
    _COL_MIN = 22
    _COL_MAX = 52
    _TOP_PLCS = 10

    for yr in sorted(year_to_eids, key=int):
        eids = sorted(
            year_to_eids[yr],
            key=lambda eid: (events[eid].get("start_date","") or "",
                             events[eid].get("event_name",""))
        )
        ws = wb.create_sheet(title=yr)
        ws.column_dimensions["A"].width = _COL_A_W

        label_map = {
            _R_NAME: "Event",
            _R_LOC:  "Location",
            _R_DATE: "Date",
            _R_TYPE: "Type",
            _R_EID:  "Event ID",
        }
        for r_idx, label in label_map.items():
            _w(ws, r_idx, 1, label,
               font=Font(size=9, color="606060"),
               align=Alignment(horizontal="right", vertical="top"))

        # Pre-compute discipline row layout.
        # Keys are division_canonical names — one row per canonical per year sheet.
        # Sort globally so NET < FREESTYLE < GOLF regardless of event order.
        div_sort_keys_yr: dict[str, tuple] = {}
        for eid in eids:
            for div_can in plcmt_by_event.get(eid, {}):
                if div_can not in div_sort_keys_yr:
                    rec = div_can_rec.get((eid, div_can), {})
                    div_sort_keys_yr[div_can] = _disc_sort_key(div_can, rec)

        all_discs_this_year: list[str] = sorted(div_sort_keys_yr, key=lambda d: div_sort_keys_yr[d])

        disc_row_start: dict[str, int] = {}
        cur_row = _R_DATA
        for div_can in all_discs_this_year:
            disc_row_start[div_can] = cur_row
            cur_row += 1 + _TOP_PLCS

        for div_can, dr in disc_row_start.items():
            _w(ws, dr, 1, div_can[:22],
               font=Font(size=8, italic=True, color="808080"),
               align=Alignment(horizontal="right", vertical="top"))
            for pi in range(1, _TOP_PLCS + 1):
                _w(ws, dr + pi, 1, f"  p{pi}",
                   font=Font(size=8, color="A0A0A0"),
                   align=Alignment(horizontal="right", vertical="top"))

        for col_offset, eid in enumerate(eids, start=2):
            ev         = events[eid]
            is_quar    = eid in quarantine
            placements = plcmt_by_event.get(eid, {})

            col_letter = get_column_letter(col_offset)
            event_col_map[eid] = (yr, col_letter)

            ev_name = ev.get("event_name", eid)
            if is_quar:
                ev_name = "⛔ " + ev_name

            _w(ws, _R_NAME, col_offset, ev_name,
               font=(_YF_QUAR if is_quar else _YF_META), align=ALIGN_LW)

            loc = ev.get("location") or \
                  f"{ev.get('city','')}, {ev.get('country','')}".strip(", ")
            _w(ws, _R_LOC,  col_offset, loc,                    font=_YF_PLC, align=ALIGN_L)
            _w(ws, _R_DATE, col_offset, ev.get("start_date",""), font=_YF_PLC, align=ALIGN_L)
            _w(ws, _R_TYPE, col_offset, ev.get("event_type",""), font=_YF_PLC, align=ALIGN_L)
            _w(ws, _R_EID,  col_offset, eid,                     font=_YF_EID, align=ALIGN_L)

            # Named anchor for EVENT INDEX hyperlinks
            anchor_name = f"event_{eid}"
            safe_yr = yr.replace("'", "''")
            try:
                dn = DefinedName(
                    name=anchor_name,
                    attr_text=f"'{safe_yr}'!${col_letter}$1",
                )
                wb.defined_names[anchor_name] = dn
            except Exception:
                pass

            max_name_len = len(ev_name)

            for div_can, disc_placements in placements.items():
                if div_can not in disc_row_start:
                    continue
                dr    = disc_row_start[div_can]
                rec   = div_can_rec.get((eid, div_can), {})
                cat   = rec.get("category_canonical", "") or rec.get("discipline_category", "")
                dfill = _cat_fill(cat)

                _w(ws, dr, col_offset, div_can,
                   font=_YF_DIV, fill=dfill, align=ALIGN_L)
                max_name_len = max(max_name_len, len(div_can))

                for pi in range(1, _TOP_PLCS + 1):
                    slot_rows = _dedup_slot(disc_placements.get(pi, []))
                    if slot_rows:
                        disp, is_unk = _team_display(slot_rows)
                    else:
                        disp, is_unk = ("", False)
                    font = _YF_UNK if is_unk and disp else _YF_PLC
                    _w(ws, dr + pi, col_offset, disp, font=font, align=ALIGN_L)
                    max_name_len = max(max_name_len, len(disp))

            ws.column_dimensions[col_letter].width = min(
                max(_COL_MIN, max_name_len + 2), _COL_MAX
            )

        ws.row_dimensions[_R_NAME].height = 32
        ws.row_dimensions[_R_LOC].height  = 14
        ws.row_dimensions[_R_DATE].height = 14
        ws.row_dimensions[_R_TYPE].height = 14
        ws.row_dimensions[_R_EID].height  = 11

        ws.freeze_panes = "B6"   # freeze rows 1-5 AND column A

        n_q = sum(1 for e in eids if e in quarantine)
        print(f"  {yr}: {len(eids)} events"
              + (f" ({n_q} quarantined)" if n_q else ""))

    return event_col_map


# ── Validation ─────────────────────────────────────────────────────────────────

def validate_workbook(wb: Workbook) -> dict:
    issues = []
    hidden_sheets = hidden_cols = hidden_rows = 0

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            hidden_sheets += 1
            issues.append(f"HIDDEN SHEET: {ws.title}")

        cd = ws.column_dimensions.get("A")
        if cd and (cd.hidden or (cd.width is not None and cd.width <= 0)):
            hidden_cols += 1
            issues.append(f"HIDDEN/ZERO COL A: {ws.title}")

        for cd in ws.column_dimensions.values():
            if cd.hidden:
                hidden_cols += 1
                issues.append(f"HIDDEN COL: {ws.title}")

        for rd in ws.row_dimensions.values():
            if rd.hidden:
                hidden_rows += 1
                issues.append(f"HIDDEN ROW {rd.index}: {ws.title}")

    return {
        "hidden_sheets": hidden_sheets,
        "hidden_cols":   hidden_cols,
        "hidden_rows":   hidden_rows,
        "issues":        issues,
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"\nbuild_workbook_v17.py  ({VERSION} · {UPDATED})")
    print(f"  Source:  {CA}")
    print(f"  Output:  {OUTPUT_PATH}\n")

    events, discs, raw_results, persons, quarantine = load_all()

    # Load publication whitelist — same filter applied by build_canonical_enrichment.py
    # (FULL + PARTIAL + QUARANTINED only; SPARSE and NO RESULTS excluded)
    pub_eids: set = set()
    if RP_EVENTS_CSV.exists():
        with open(RP_EVENTS_CSV, newline="", encoding="utf-8") as _f:
            pub_eids = {row["event_id"] for row in csv.DictReader(_f)}
        print(f"  Publication whitelist: {len(pub_eids)} events from release_publication/")
    else:
        print("  WARNING: release_publication/events.csv not found — year sheets unfiltered")
        pub_eids = set(events.keys())

    print("Computing statistics…")
    stats = compute_stats(raw_results, events, discs, persons)
    print(f"  Persons with stats: {len(stats)}")

    wb = Workbook()
    wb.remove(wb.active)

    print("\nBuilding sheets…")

    build_readme(wb, events, persons, len(raw_results))
    build_data_notes(wb)
    build_statistics(wb, stats, persons)
    build_era_leaders(wb, raw_results, events, discs, persons)
    build_player_stats(wb, stats, persons)

    event_col_map = build_year_sheets(
        wb, raw_results, events, discs, persons, quarantine, pub_eids)

    build_event_index(wb, events, discs, raw_results, quarantine, event_col_map, pub_eids)
    build_excluded_events(wb, events, quarantine, discs, raw_results)

    # Reorder front-matter sheets before year sheets
    desired_front = ["README", "EVENT INDEX", "STATISTICS", "ERA LEADERS",
                     "PLAYER STATS", "EXCLUDED EVENTS"]
    final_order = desired_front + [s for s in wb.sheetnames if s not in desired_front]
    for i, name in enumerate(final_order):
        if name in wb.sheetnames:
            current = wb.sheetnames.index(name)
            if current != i:
                wb.move_sheet(name, offset=i - current)

    print("\nValidating…")
    v = validate_workbook(wb)
    print(f"  Hidden sheets:  {v['hidden_sheets']}")
    print(f"  Hidden columns: {v['hidden_cols']}")
    print(f"  Hidden rows:    {v['hidden_rows']}")
    if v["issues"]:
        for iss in v["issues"][:10]:
            print(f"  ⚠  {iss}")
    else:
        print("  ✓ No hidden structure found")

    print(f"\nSaving → {OUTPUT_PATH.name}…")
    wb.save(str(OUTPUT_PATH))
    print(f"  Done.  ({OUTPUT_PATH.stat().st_size / 1_048_576:.1f} MB)")

    print("\n── Summary ──────────────────────────────────────────────")
    year_sheets = [s for s in wb.sheetnames if s.isdigit()]
    print(f"  Sheets total:       {len(wb.sheetnames)}")
    print(f"  Front-matter:       {len(desired_front)}")
    print(f"  Year sheets:        {len(year_sheets)}"
          + (f"  ({year_sheets[0]}–{year_sheets[-1]})" if year_sheets else ""))
    print(f"  Events covered:     {len(event_col_map)}")
    print(f"  Persons in stats:   {len(stats)}")
    print(f"  Persons in PLAYER STATS: {len(persons)}")
    print(f"  Hidden structure:   {v['hidden_sheets'] + v['hidden_cols'] + v['hidden_rows']}")
    print(f"  Version:            {VERSION}")
    print(f"  Updated:            {UPDATED}")


if __name__ == "__main__":
    main()
