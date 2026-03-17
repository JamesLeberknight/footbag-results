#!/usr/bin/env python3
"""
tools/64_patch_pbp_v67_to_v68.py

Build Placements_ByPerson_v68.csv by appending magazine placements to PBP v67.

Reads:
  inputs/identity_lock/Placements_ByPerson_v67.csv    — existing lock file
  out/review/magazine_pbp_review.xlsx                  — completed review workbook
  out/stage2_canonical_events.csv                      — event metadata

Writes:
  inputs/identity_lock/Placements_ByPerson_v68.csv

Resolution logic (per row in review sheet):
  1. If decision_person_id is filled → use it (overrides auto)
  2. Else if auto_person_id is filled → use it
  3. Else → person_id = "__NON_PERSON__", person_unresolved = 1

For doubles: two player rows sharing the same place get linked via team_person_key
(a deterministic UUID built from event_id + division + place + sorted player names).

Usage:
  .venv/bin/python tools/64_patch_pbp_v67_to_v68.py [--dry-run]
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

csv.field_size_limit(10_000_000)

ROOT   = Path(__file__).resolve().parents[1]
INPUTS = ROOT / "inputs" / "identity_lock"
OUT    = ROOT / "out"
REVIEW = OUT / "review"

V67_CSV      = INPUTS / "Placements_ByPerson_v67.csv"
REVIEW_XLSX  = REVIEW / "magazine_pbp_review.xlsx"
STAGE2_CSV   = OUT / "stage2_canonical_events.csv"
PT_CSV       = INPUTS / "Persons_Truth_Final_v42.csv"
V68_CSV      = INPUTS / "Placements_ByPerson_v68.csv"


def norm(s: str) -> str:
    s = unicodedata.normalize("NFD", s.lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip()


def make_team_key(event_id: str, div: str, place: str, p1: str, p2: str) -> str:
    key = f"{event_id}|{div}|{place}|{'|'.join(sorted([p1, p2]))}"
    return hashlib.md5(key.encode()).hexdigest()[:12]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    # ── Load PT for canon lookup ───────────────────────────────────────────
    pid_to_canon: dict[str, str] = {}
    with open(PT_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            pid_to_canon[row["effective_person_id"].strip()] = row["person_canon"].strip()
    print(f"PT loaded: {len(pid_to_canon):,} persons")

    # ── Load stage2 event metadata ─────────────────────────────────────────
    event_meta: dict[str, dict] = {}
    with open(STAGE2_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            eid = row["event_id"]
            event_meta[eid] = {
                "year":       row.get("year", ""),
                "event_name": row.get("event_name", ""),
            }

    # ── Load review workbook ───────────────────────────────────────────────
    if not REVIEW_XLSX.exists():
        print(f"ERROR: {REVIEW_XLSX} not found. Run tools/63_magazine_pbp_review.py first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(REVIEW_XLSX, read_only=True, data_only=True)
    ws = wb["Placements"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    review_rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[col["event_id"]]:
            continue
        review_rows.append({
            "event_id":          str(row[col["event_id"]]).strip(),
            "division_canon":    str(row[col["division_canon"]] or "").strip(),
            "division_category": str(row[col["division_category"]] or "").strip(),
            "division_raw":      str(row[col.get("division_raw", -1)] or "").strip() if col.get("division_raw") is not None else "",
            "place":             str(row[col["place"]] or "").strip(),
            "competitor_type":   str(row[col["competitor_type"]] or "player").strip(),
            "player_name":       str(row[col["player_name"]] or "").strip(),
            "auto_person_id":    str(row[col["auto_person_id"]] or "").strip(),
            "auto_person_canon": str(row[col["auto_person_canon"]] or "").strip(),
            "decision_person_id":str(row[col["decision_person_id"]] or "").strip(),
        })

    wb.close()
    print(f"Review rows loaded: {len(review_rows)}")

    # ── Resolve person_id for each row ─────────────────────────────────────
    new_pbp_rows: list[dict] = []

    # Group by (event_id, division_canon, place) to handle doubles teams
    from collections import defaultdict
    place_groups: dict[tuple, list[dict]] = defaultdict(list)
    for r in review_rows:
        key = (r["event_id"], r["division_canon"], r["place"])
        place_groups[key].append(r)

    stats = {"resolved": 0, "non_person": 0, "unresolved": 0}

    for (eid, div, place), group in sorted(place_groups.items()):
        meta = event_meta.get(eid, {})
        year = meta.get("year", "")
        is_doubles = group[0]["competitor_type"] == "team" and len(group) >= 2

        if is_doubles and len(group) >= 2:
            p1_name = group[0]["player_name"]
            p2_name = group[1]["player_name"]
            team_key = make_team_key(eid, div, place, p1_name, p2_name)
            team_display = f"{p1_name} / {p2_name}"
        else:
            team_key     = ""
            team_display = ""

        for r in group:
            # Resolve: decision overrides auto
            dec = r["decision_person_id"]
            auto = r["auto_person_id"]

            if dec and dec != "__NON_PERSON__":
                person_id = dec
                person_canon = pid_to_canon.get(dec, r["auto_person_canon"])
                person_unresolved = ""
                stats["resolved"] += 1
            elif dec == "__NON_PERSON__" or auto == "__NON_PERSON__":
                person_id = "__NON_PERSON__"
                person_canon = r["player_name"]
                person_unresolved = ""
                stats["non_person"] += 1
            elif auto and auto != "__NON_PERSON__":
                person_id = auto
                person_canon = r["auto_person_canon"] or pid_to_canon.get(auto, "")
                person_unresolved = ""
                stats["resolved"] += 1
            else:
                # Unresolved — include with person_unresolved flag
                person_id = ""
                person_canon = r["player_name"]
                person_unresolved = "1"
                stats["unresolved"] += 1

            new_pbp_rows.append({
                "event_id":         eid,
                "year":             year,
                "division_canon":   div,
                "division_category":r["division_category"],
                "place":            place,
                "competitor_type":  r["competitor_type"],
                "person_id":        person_id,
                "team_person_key":  team_key,
                "person_canon":     person_canon,
                "team_display_name":team_display if is_doubles else "",
                "coverage_flag":    "sparse",   # pre-mirror era
                "person_unresolved":person_unresolved,
                "norm":             norm(person_canon),
                "division_raw":     r.get("division_raw", div),
            })

    print(f"New PBP rows: {len(new_pbp_rows)}")
    print(f"  resolved:   {stats['resolved']}")
    print(f"  unresolved: {stats['unresolved']} (person_unresolved=1)")
    print(f"  non-person: {stats['non_person']} (__NON_PERSON__)")

    if args.dry_run:
        print("[dry-run] Not writing file.")
        return

    # ── Read v67 and append ────────────────────────────────────────────────
    v67_rows: list[dict] = []
    v67_fieldnames: list[str] = []
    with open(V67_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        v67_fieldnames = reader.fieldnames or []
        for row in reader:
            v67_rows.append(row)

    print(f"PBP v67 rows: {len(v67_rows):,}")

    # Ensure no duplicate magazine event_ids already present in v67
    existing_eids = {r["event_id"] for r in v67_rows}
    mag_eids      = {r["event_id"] for r in new_pbp_rows}
    overlap       = existing_eids & mag_eids
    if overlap:
        print(f"WARNING: {len(overlap)} magazine event_ids already in v67 — skipping those rows")
        new_pbp_rows = [r for r in new_pbp_rows if r["event_id"] not in overlap]

    combined = v67_rows + new_pbp_rows
    print(f"PBP v68 total rows: {len(combined):,} (+{len(new_pbp_rows)} magazine)")

    with open(V68_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=v67_fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(combined)

    print(f"Wrote: {V68_CSV}")
    print()
    print("Next steps:")
    print("  1. Update run_pipeline.sh to reference Placements_ByPerson_v68.csv")
    print("  2. Re-run: ./run_pipeline.sh release")
    print("  3. Update CLAUDE.md canonical state table")


if __name__ == "__main__":
    main()
