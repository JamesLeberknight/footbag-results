"""
tools/50_build_review_packet.py
================================
Build the Footbag Event Review Packet workbook and HTML dashboard.

Outputs:
  out/review/Footbag_Event_Review_Packet.xlsx
  out/review/event_review_dashboard.html
"""

import csv
import json
import re
import sys
from pathlib import Path
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

csv.field_size_limit(10**7)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).parent.parent
OUT_DIR = ROOT / "out" / "review"
OUT_DIR.mkdir(parents=True, exist_ok=True)

STAGE2_CSV        = ROOT / "out" / "stage2_canonical_events.csv"
PBP_CSV           = ROOT / "inputs" / "identity_lock" / "Placements_ByPerson_v59.csv"
KNOWN_ISSUES_CSV  = ROOT / "overrides" / "known_issues.csv"
OVERRIDES_JSONL   = ROOT / "overrides" / "events_overrides.jsonl"
CANONICAL_EVENTS  = ROOT / "out" / "canonical" / "events.csv"
PIPELINE_02       = ROOT / "pipeline" / "02_canonicalize_results.py"
QUARANTINE_CSV    = ROOT / "inputs" / "review_quarantine_events.csv"

WORKBOOK_PATH    = OUT_DIR / "Footbag_Event_Review_Packet.xlsx"
DASHBOARD_PATH   = OUT_DIR / "event_review_dashboard.html"

# ---------------------------------------------------------------------------
# Hard-coded recent changes
# ---------------------------------------------------------------------------
RECENTLY_MODIFIED_IDS = {
    '1235653935', '1241011525', '859923755',
    '1366240051', '1177512537', '1069791565', '937854594',
}

RECENT_FIXES = {
    '937854594':  {'fix_version': 'v2.6.0', 'fix_type': 'RESULTS_FILE_OVERRIDE + PBP division fix', 'files_changed': '02_canonicalize_results.py, legacy_data/event_results/937854594.txt, PBP v56'},
    '1069791565': {'fix_version': 'v2.7.0', 'fix_type': 'PBP division name cleanup (? → -)', 'files_changed': 'PBP v57 (21 rows)'},
    '1235653935': {'fix_version': 'v2.8.0', 'fix_type': 'PT entries removed (3 concat team names)', 'files_changed': 'PT v39→v41, PBP v57→v59'},
    '1241011525': {'fix_version': 'v2.8.0', 'fix_type': 'PT entry removed (concat team name)', 'files_changed': 'PT v39→v41, PBP v57→v59'},
    '859923755':  {'fix_version': 'v2.8.0', 'fix_type': 'PT entry removed (concat team name)', 'files_changed': 'PT v39→v41, PBP v57→v59'},
    '1366240051': {'fix_version': 'v2.8.0', 'fix_type': 'PT entry removed (concat team name)', 'files_changed': 'PT v39→v41, PBP v57→v59'},
    '1177512537': {'fix_version': 'v2.7.0', 'fix_type': 'PT entry removed (Grischa team concat)', 'files_changed': 'PT v38→v39, PBP v56→v57'},
}

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
FILL_HIGH        = PatternFill("solid", fgColor="FF4444")
FILL_MEDIUM      = PatternFill("solid", fgColor="FF8800")
FILL_LOW         = PatternFill("solid", fgColor="FFDD00")
FILL_CLEAN       = PatternFill("solid", fgColor="44BB44")
FILL_QUARANTINED = PatternFill("solid", fgColor="AAAAAA")
FILL_HEADER      = PatternFill("solid", fgColor="2E4057")

FONT_WHITE  = Font(color="FFFFFF", bold=True)
FONT_BLACK  = Font(color="000000", bold=True)
FONT_HEADER = Font(color="FFFFFF", bold=True)

# ---------------------------------------------------------------------------
# Step 7 — extract RESULTS_FILE_OVERRIDES keys
# ---------------------------------------------------------------------------
def extract_results_file_overrides():
    if not PIPELINE_02.exists():
        print(f"WARNING: {PIPELINE_02} not found; RESULTS_FILE_OVERRIDES will be empty.")
        return set()
    content = PIPELINE_02.read_text(encoding="utf-8", errors="replace")
    matches = re.findall(r'"(\d{9,10})":\s*\{', content)
    return set(matches)

# ---------------------------------------------------------------------------
# Step 8 — parse events_overrides.jsonl
# ---------------------------------------------------------------------------
def load_events_overrides():
    overrides = defaultdict(list)
    if not OVERRIDES_JSONL.exists():
        return overrides
    with open(OVERRIDES_JSONL, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                eid = str(obj.get("event_id", ""))
                if eid:
                    overrides[eid].append(obj)
            except json.JSONDecodeError:
                pass
    return overrides

# ---------------------------------------------------------------------------
# Load stage2
# ---------------------------------------------------------------------------
def load_stage2():
    rows = []
    with open(STAGE2_CSV, encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows

# ---------------------------------------------------------------------------
# Load PBP
# ---------------------------------------------------------------------------
def load_pbp():
    """Returns dict: event_id -> list of row dicts."""
    pbp = defaultdict(list)
    with open(PBP_CSV, encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pbp[str(row['event_id'])].append(row)
    return pbp

# ---------------------------------------------------------------------------
# Load quarantine list
# ---------------------------------------------------------------------------
def load_quarantine():
    """Returns dict: event_id -> row dict (reason, review_stage, notes)."""
    quarantine = {}
    if not QUARANTINE_CSV.exists():
        return quarantine
    with open(QUARANTINE_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            eid = str(row['event_id'])
            quarantine[eid] = row
    return quarantine

# ---------------------------------------------------------------------------
# Load known_issues
# ---------------------------------------------------------------------------
def load_known_issues():
    issues = {}
    with open(KNOWN_ISSUES_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            eid = str(row['event_id'])
            if eid not in issues:
                issues[eid] = row
            # keep highest severity if duplicates
    return issues

# ---------------------------------------------------------------------------
# Load canonical events
# ---------------------------------------------------------------------------
def load_canonical_events():
    canon = {}
    with open(CANONICAL_EVENTS, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            legacy_id = str(row.get('legacy_event_id', ''))
            if legacy_id:
                canon[legacy_id] = row
    return canon

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def parse_placements_json(raw):
    if not raw or raw.strip() in ('', '[]'):
        return []
    try:
        return json.loads(raw)
    except Exception:
        return []

def has_encoding_artifact(text):
    if not text:
        return False
    t = str(text)
    # non-ASCII chars — check if they look suspicious (control chars, replacement char, etc.)
    suspicious = re.search(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f\x80-\x9f\ufffd]', t)
    if suspicious:
        return True
    # consecutive non-Latin characters in a Latin context
    non_ascii = [c for c in t if ord(c) > 127]
    if len(non_ascii) > 0:
        # Check if they look like real diacritics vs garbage
        # Heuristic: if >30% non-ASCII in a short string, might be corrupted
        if len(t) < 20 and len(non_ascii) / len(t) > 0.3:
            return True
    return False

# ---------------------------------------------------------------------------
# Reconciliation helpers (Rules 1, 3, 4)
# ---------------------------------------------------------------------------
CONTEST_DIV_RE = re.compile(
    r'\b(sick\s*3|shred\s*30|circle\s+contest|request\s+contest|battle\s+contest)\b',
    re.IGNORECASE
)

def normalize_div_name(name: str) -> str:
    """Normalize division name tokens for fuzzy comparison (Rule 1)."""
    s = name.lower().strip()
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def compute_normalized_div_diff(divs_raw, divs_can):
    """
    Returns (only_in_raw, only_in_can) after normalizing both name sets.
    Empty sets = no real mismatch after normalization.
    """
    nr = {normalize_div_name(d) for d in divs_raw if d}
    nc = {normalize_div_name(d) for d in divs_can if d}
    return nr - nc, nc - nr

def has_contest_format(placements) -> bool:
    """True if any division is a contest format (Sick 3, Shred 30, etc.). Rule 4."""
    for p in placements:
        div = p.get('division_raw') or p.get('division', '') or p.get('division_canon', '') or ''
        if CONTEST_DIV_RE.search(div):
            return True
    return False

def count_exact_duplicates(placements) -> int:
    """Count exact duplicate placement rows (same div, place, player1). Rule 3."""
    seen = set()
    dupes = 0
    for p in placements:
        key = (
            p.get('division_raw') or p.get('division', ''),
            p.get('place'),
            p.get('player1_name', ''),
        )
        if key in seen:
            dupes += 1
        else:
            seen.add(key)
    return dupes

def render_excerpt(placements, max_divs=3, max_places=3):
    """Render a short text excerpt from placements_json list."""
    if not placements:
        return "(no results)"
    by_div = defaultdict(list)
    for p in placements:
        div = p.get('division_raw') or p.get('division', '') or p.get('division_canon', '')
        by_div[div].append(p)
    lines = []
    truncated = False
    for i, (div, entries) in enumerate(by_div.items()):
        if i >= max_divs:
            truncated = True
            break
        n = len(entries)
        lines.append(f"Division: {div} ({n} placements)")
        shown = 0
        for e in sorted(entries, key=lambda x: x.get('place', 99)):
            if shown >= max_places:
                break
            place = e.get('place', '?')
            p1 = e.get('player1_name', '')
            p2 = e.get('player2_name', '')
            name = f"{p1} & {p2}" if p2 else p1
            lines.append(f"  {place}. {name}")
            shown += 1
        if n > max_places:
            lines.append(f"  ...")
    if truncated:
        lines.append("...")
    return "\n".join(lines)

def render_pbp_excerpt(pbp_rows, max_divs=3, max_places=3):
    """Render a short text excerpt from PBP rows."""
    if not pbp_rows:
        return "(no canonical results)"
    by_div = defaultdict(list)
    for r in pbp_rows:
        div = r.get('division_canon', '')
        by_div[div].append(r)
    lines = []
    truncated = False
    for i, (div, entries) in enumerate(by_div.items()):
        if i >= max_divs:
            truncated = True
            break
        n = len(entries)
        lines.append(f"Division: {div} ({n} placements)")
        shown = 0
        for e in sorted(entries, key=lambda x: int(x.get('place', 99)) if str(x.get('place', 99)).isdigit() else 99):
            if shown >= max_places:
                break
            place = e.get('place', '?')
            team = e.get('team_display_name', '')
            person = e.get('person_canon', '')
            name = team if team else person
            lines.append(f"  {place}. {name}")
            shown += 1
        if n > max_places:
            lines.append(f"  ...")
    if truncated:
        lines.append("...")
    return "\n".join(lines)

def render_full(placements, max_chars=500):
    if not placements:
        return "(no results)"
    by_div = defaultdict(list)
    for p in placements:
        div = p.get('division_raw') or p.get('division', '') or p.get('division_canon', '')
        by_div[div].append(p)
    lines = []
    for div, entries in by_div.items():
        lines.append(f"[{div}]")
        for e in sorted(entries, key=lambda x: x.get('place', 99)):
            place = e.get('place', '?')
            p1 = e.get('player1_name', '')
            p2 = e.get('player2_name', '')
            name = f"{p1} & {p2}" if p2 else p1
            lines.append(f"  {place}. {name}")
    text = "\n".join(lines)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n..."
    return text

def render_pbp_full(pbp_rows, max_chars=500):
    if not pbp_rows:
        return "(no canonical results)"
    by_div = defaultdict(list)
    for r in pbp_rows:
        div = r.get('division_canon', '')
        by_div[div].append(r)
    lines = []
    for div, entries in by_div.items():
        lines.append(f"[{div}]")
        for e in sorted(entries, key=lambda x: int(x.get('place', 99)) if str(x.get('place', 99)).isdigit() else 99):
            place = e.get('place', '?')
            team = e.get('team_display_name', '')
            person = e.get('person_canon', '')
            name = team if team else person
            lines.append(f"  {place}. {name}")
    text = "\n".join(lines)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n..."
    return text

def compute_diff_notes(row):
    div_raw = row['division_count_raw']
    div_can = row['division_count_canonical']
    plc_raw = row['placement_count_raw']
    plc_can = row['placement_count_canonical']
    placements = row['_placements']
    pbp_rows   = row['_pbp_rows']

    notes = []
    if div_raw != div_can and div_raw > 0 and div_can > 0:
        raw_divs = {p.get('division_raw') or p.get('division', '') for p in placements}
        can_divs = {r.get('division_canon', '') for r in pbp_rows}
        only_raw = raw_divs - can_divs
        only_can = can_divs - raw_divs
        if only_raw:
            notes.append(f"In raw only: {', '.join(sorted(only_raw)[:5])}")
        if only_can:
            notes.append(f"In canonical only: {', '.join(sorted(only_can)[:5])}")
    if abs(plc_raw - plc_can) > 3:
        notes.append(f"Placement delta: raw={plc_raw} canonical={plc_can} (diff={plc_raw - plc_can:+d})")
    if not notes:
        if plc_raw == plc_can and div_raw == div_can:
            notes.append("counts match")
        else:
            notes.append(f"raw={plc_raw} placements, {div_raw} divs; canonical={plc_can} placements, {div_can} divs")
    return "\n".join(notes)

# ---------------------------------------------------------------------------
# openpyxl helpers
# ---------------------------------------------------------------------------
def style_header_row(ws, row_num=1):
    for cell in ws[row_num]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_column_widths(ws, widths_dict):
    """widths_dict: {col_letter_or_name: width}"""
    for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        header = col[0].value
        letter = get_column_letter(i)
        if header in widths_dict:
            ws.column_dimensions[letter].width = widths_dict[header]

def write_sheet(wb, sheet_name, data_rows, columns, col_widths=None,
                heat_col=None, wrap_cols=None, freeze=True):
    """Generic sheet writer."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(columns)
    style_header_row(ws)

    for row in data_rows:
        ws.append([row.get(c, '') for c in columns])

    if freeze:
        ws.freeze_panes = ws['A2']

    ws.auto_filter.ref = ws.dimensions

    # Heat label colouring (cell only)
    if heat_col and heat_col in columns:
        col_idx = columns.index(heat_col) + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val == 'HIGH':
                cell.fill = FILL_HIGH
                cell.font = Font(color="FFFFFF", bold=True)
            elif val == 'MEDIUM':
                cell.fill = FILL_MEDIUM
                cell.font = Font(color="FFFFFF", bold=True)
            elif val == 'LOW':
                cell.fill = FILL_LOW
                cell.font = Font(color="000000", bold=True)
            elif val == 'CLEAN':
                cell.fill = FILL_CLEAN
                cell.font = Font(color="000000", bold=True)
            elif val == 'QUARANTINED':
                cell.fill = FILL_QUARANTINED
                cell.font = Font(color="FFFFFF", bold=False)

    # Wrap text for designated columns
    if wrap_cols:
        for wc in wrap_cols:
            if wc in columns:
                ci = columns.index(wc) + 1
                for ri in range(2, ws.max_row + 1):
                    ws.cell(row=ri, column=ci).alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    if col_widths:
        for col_name, width in col_widths.items():
            if col_name in columns:
                ci = columns.index(col_name) + 1
                ws.column_dimensions[get_column_letter(ci)].width = width

    return ws

# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------
def main():
    print("Loading data sources...")

    stage2_rows     = load_stage2()
    pbp_by_event    = load_pbp()
    known_issues    = load_known_issues()
    events_overrides= load_events_overrides()
    canonical_events= load_canonical_events()
    rfo_ids         = extract_results_file_overrides()
    quarantine      = load_quarantine()

    print(f"  stage2: {len(stage2_rows)} events")
    print(f"  PBP: {sum(len(v) for v in pbp_by_event.values())} rows across {len(pbp_by_event)} events")
    print(f"  known_issues: {len(known_issues)} events")
    print(f"  events_overrides: {len(events_overrides)} events")
    print(f"  quarantine: {len(quarantine)} events")
    print(f"  canonical_events: {len(canonical_events)} events")
    print(f"  RESULTS_FILE_OVERRIDES: {len(rfo_ids)} entries")

    # -----------------------------------------------------------------------
    # Step 1 — gather per-event data
    # -----------------------------------------------------------------------
    events = []

    for s2 in stage2_rows:
        eid = str(s2['event_id'])
        year_raw = s2.get('year', '')
        try:
            year = int(year_raw) if year_raw else 0
        except ValueError:
            year = 0

        event_name   = s2.get('event_name', '')
        location_raw = s2.get('location', '')
        host_club    = s2.get('host_club', '')

        # Parse placements_json
        placements = parse_placements_json(s2.get('placements_json', ''))

        # Raw counts
        raw_results_present    = 1 if placements else 0
        placement_count_raw    = len(placements)
        divisions_raw          = {p.get('division_raw') or p.get('division', '') for p in placements}
        division_count_raw     = len(divisions_raw)

        # PBP (canonical) counts
        pbp_rows               = pbp_by_event.get(eid, [])
        canonical_results_present = 1 if pbp_rows else 0
        placement_count_canonical = len(pbp_rows)
        divisions_canonical    = {r.get('division_canon', '') for r in pbp_rows}
        division_count_canonical  = len(divisions_canonical)

        # Flags
        known_issue_flag = 1 if eid in known_issues else 0
        known_issue_type = known_issues[eid]['severity'] if eid in known_issues else ''
        known_issue_note = known_issues[eid]['note'] if eid in known_issues else ''

        in_jsonl  = eid in events_overrides
        in_rfo    = eid in rfo_ids
        override_flag = 1 if (in_jsonl or in_rfo) else 0

        recently_modified_flag = 1 if eid in RECENTLY_MODIFIED_IDS else 0

        # canonical event metadata
        canon = canonical_events.get(eid, {})
        start_date = canon.get('start_date', '')
        end_date   = canon.get('end_date', '')
        city       = canon.get('city', '')
        region     = canon.get('region', '')
        country    = canon.get('country', '')
        event_type = canon.get('event_type', s2.get('event_type', ''))

        # Step 2 — metadata anomaly
        anomaly = 0
        if not location_raw or location_raw.strip() == '':
            anomaly = 1
        elif '?' in str(event_name) or '?' in str(location_raw):
            anomaly = 1
        elif has_encoding_artifact(location_raw):
            anomaly = 1
        if not host_club or host_club.strip() == '':
            anomaly = 1

        # ── Reconciliation flags (Rules 1-6) ──────────────────────────────
        # Rule 1: RFO divergence — expected by design (Rule 1 + STEP 1)
        rfo_divergence_expected = in_rfo and canonical_results_present

        # Rule 1: Division name normalization
        only_raw_norm, only_can_norm = compute_normalized_div_diff(
            divisions_raw, divisions_canonical
        )
        div_mismatch_after_norm = bool(only_raw_norm or only_can_norm)

        # PBP stale: stage2 has MORE divisions than PBP, not RFO, parser already fixed
        pbp_stale_flag = (
            division_count_raw > division_count_canonical > 0
            and not in_rfo
            and not known_issue_flag
        )

        # S2 < PBP: PBP has MORE divisions than stage2 — investigate source
        s2_less_flag = (
            division_count_canonical > division_count_raw > 0
            and not in_rfo
            and not known_issue_flag
        )

        # Rule 4: Contest format — legitimately non-standard placements
        contest_format_flag = has_contest_format(placements)

        # Rule 3: Exact duplicate count
        dup_count = count_exact_duplicates(placements)

        # Real division mismatch: not RFO, not stale, not normalized-match
        div_mismatch_real = (
            division_count_raw > 0
            and division_count_canonical > 0
            and div_mismatch_after_norm
            and not rfo_divergence_expected
            and not pbp_stale_flag
            and not s2_less_flag
        )

        # Real placement mismatch: not RFO, not stale PBP, not contest, not explained by dedup
        plc_diff = abs(placement_count_raw - placement_count_canonical)
        plc_mismatch_real = (
            placement_count_raw > 0
            and plc_diff > 3
            and not rfo_divergence_expected
            and not pbp_stale_flag    # stale PBP naturally has fewer placements too
            and not contest_format_flag
            and not (dup_count > 0 and abs(plc_diff - dup_count) <= 3)
        )

        # Step 3 — heat score (Rule 6: only penalize real mismatches)
        heat = 0
        if known_issue_flag:
            heat += 4
        if override_flag:
            heat += 3
        if div_mismatch_real:
            heat += 3
        if plc_mismatch_real:
            heat += 3
        if recently_modified_flag:
            heat += 2
        if anomaly:
            heat += 1

        if heat >= 7:
            heat_label = 'HIGH'
            heat_rank  = 1
        elif heat >= 4:
            heat_label = 'MEDIUM'
            heat_rank  = 2
        elif heat >= 1:
            heat_label = 'LOW'
            heat_rank  = 3
        else:
            heat_label = 'CLEAN'
            heat_rank  = 4

        # Quarantine override: suppress HIGH events that are documented and deferred
        quarantine_row = quarantine.get(eid)
        if quarantine_row:
            heat_label = 'QUARANTINED'
            heat_rank  = 5   # sort after CLEAN
            heat       = 0

        # Step 4 — priority tier
        SPECIAL_YEARS = {1997, 1998, 1999, 2000, 2001, 2002, 2003, 2024, 2025, 2026}
        is_tier1 = (
            known_issue_flag
            or override_flag
            or recently_modified_flag
            or (division_count_raw > 0 and division_count_canonical > 0 and division_count_raw != division_count_canonical)
            or (placement_count_raw > 0 and abs(placement_count_raw - placement_count_canonical) > 3)
        )
        is_tier2_natural = (year in SPECIAL_YEARS) or (event_type in ('worlds', 'mixed'))

        # HIGH/MEDIUM are at least Tier2
        if is_tier1:
            priority_tier = 'Tier1'
        elif is_tier2_natural or heat_label in ('HIGH', 'MEDIUM'):
            priority_tier = 'Tier2'
        else:
            priority_tier = 'Tier3'

        # Step 5 — diff summary
        if known_issue_flag:
            diff_summary = 'known issue'
        elif rfo_divergence_expected:
            diff_summary = 'override ok'
        elif pbp_stale_flag:
            diff_summary = 'pbp stale'
        elif s2_less_flag:
            diff_summary = 'investigate: pbp>stage2'
        elif not div_mismatch_after_norm and division_count_raw != division_count_canonical:
            diff_summary = 'normalized match'
        elif contest_format_flag and not div_mismatch_real:
            diff_summary = 'contest format'
        elif div_mismatch_real:
            diff_summary = 'division mismatch'
        elif plc_mismatch_real:
            diff_summary = 'placement mismatch'
        elif override_flag:
            diff_summary = 'override used'
        elif recently_modified_flag:
            diff_summary = 'recently modified'
        elif anomaly:
            diff_summary = 'metadata only'
        else:
            diff_summary = 'clean'

        events.append({
            'event_id': eid,
            'year': year if year else '',
            'event_name': event_name,
            'start_date': start_date,
            'end_date': end_date,
            'city': city,
            'region': region,
            'country': country,
            'location_raw': location_raw,
            'host_club': host_club,
            'priority_tier': priority_tier,
            'known_issue_flag': known_issue_flag,
            'known_issue_type': known_issue_type,
            '_known_issue_note': known_issue_note,
            'override_flag': override_flag,
            'recently_modified_flag': recently_modified_flag,
            'reviewer_flag': 0,
            'member_id_enrichment_pending': 0,
            'raw_results_present': raw_results_present,
            'canonical_results_present': canonical_results_present,
            'division_count_raw': division_count_raw,
            'division_count_canonical': division_count_canonical,
            'placement_count_raw': placement_count_raw,
            'placement_count_canonical': placement_count_canonical,
            'diff_summary': diff_summary,
            'review_heat_score': heat,
            'review_heat_label': heat_label,
            'review_priority_rank': heat_rank,
            'review_status': '',
            'review_note': '',
            'action_needed': '',
            'action_target': '',
            'final_disposition': '',
            'metadata_anomaly': anomaly,
            'quarantine_flag': 1 if quarantine_row else 0,
            'quarantine_reason': quarantine_row['reason'] if quarantine_row else '',
            'quarantine_notes': quarantine_row['notes'] if quarantine_row else '',
            'rfo_divergence_expected': int(rfo_divergence_expected),
            'pbp_stale_flag': int(pbp_stale_flag),
            'contest_format_flag': int(contest_format_flag),
            's2_less_flag': int(s2_less_flag),
            'div_mismatch_real': int(div_mismatch_real),
            'plc_mismatch_real': int(plc_mismatch_real),
            'event_type': event_type,
            '_placements': placements,
            '_pbp_rows': pbp_rows,
            '_in_jsonl': in_jsonl,
            '_in_rfo': in_rfo,
            '_events_override_rows': events_overrides.get(eid, []),
        })

    # Sort by review_priority_rank, year, event_name
    events.sort(key=lambda r: (r['review_priority_rank'], r['year'] or 0, r['event_name']))

    # Step 6 — text excerpts (lazily compute for Tier1/Tier2)
    for ev in events:
        placements = ev['_placements']
        pbp_rows   = ev['_pbp_rows']
        ev['raw_truth_excerpt']      = render_excerpt(placements)
        ev['canonical_excerpt']      = render_pbp_excerpt(pbp_rows)
        ev['raw_truth_full']         = render_full(placements)
        ev['canonical_render_full']  = render_pbp_full(pbp_rows)
        ev['diff_notes']             = compute_diff_notes(ev)

    # -----------------------------------------------------------------------
    # Print summary stats
    # -----------------------------------------------------------------------
    heat_counts = {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'CLEAN': 0, 'QUARANTINED': 0}
    tier_counts = {'Tier1': 0, 'Tier2': 0, 'Tier3': 0}
    for ev in events:
        heat_counts[ev['review_heat_label']] += 1
        tier_counts[ev['priority_tier']] += 1

    print(f"\nHeat distribution: {heat_counts}")
    print(f"Priority tiers: {tier_counts}")

    # -----------------------------------------------------------------------
    # Build workbook
    # -----------------------------------------------------------------------
    print("\nBuilding workbook...")
    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ------- Sheet: Queue -------
    queue_cols = [
        'event_id', 'year', 'event_name', 'start_date', 'end_date',
        'city', 'region', 'country', 'location_raw', 'host_club',
        'priority_tier', 'known_issue_flag', 'known_issue_type', 'override_flag',
        'recently_modified_flag', 'reviewer_flag', 'member_id_enrichment_pending',
        'raw_results_present', 'canonical_results_present',
        'division_count_raw', 'division_count_canonical',
        'placement_count_raw', 'placement_count_canonical',
        'diff_summary', 'review_heat_score', 'review_heat_label', 'review_priority_rank',
        'rfo_divergence_expected', 'pbp_stale_flag', 'contest_format_flag',
        's2_less_flag', 'div_mismatch_real', 'plc_mismatch_real',
        'review_status', 'review_note', 'action_needed', 'action_target', 'final_disposition',
    ]
    queue_widths = {
        'event_id': 14, 'year': 6, 'event_name': 35, 'priority_tier': 8,
        'review_heat_label': 12, 'diff_summary': 22, 'review_status': 14,
        'review_note': 30, 'action_needed': 20,
    }
    write_sheet(wb, 'Queue', events, queue_cols, col_widths=queue_widths, heat_col='review_heat_label')
    print(f"  Queue: {len(events)} rows")

    # ------- Sheet: Tier1_Review -------
    tier1_events = [ev for ev in events if ev['priority_tier'] == 'Tier1']

    def notable_differences(ev):
        parts = []
        div_raw = ev['division_count_raw']
        div_can = ev['division_count_canonical']
        plc_raw = ev['placement_count_raw']
        plc_can = ev['placement_count_canonical']
        if div_raw > 0 and div_can > 0 and div_raw != div_can:
            parts.append(f"Div count: raw={div_raw} vs canonical={div_can}")
        if abs(plc_raw - plc_can) > 3:
            parts.append(f"Place count: raw={plc_raw} vs canonical={plc_can}")
        if ev['known_issue_flag']:
            note = ev['_known_issue_note'][:80] if ev['_known_issue_note'] else ''
            parts.append(f"Known issue: {note}")
        if ev['recently_modified_flag']:
            fix = RECENT_FIXES.get(ev['event_id'], {})
            ver = fix.get('fix_version', 'v2.x.x')
            parts.append(f"Recently modified ({ver})")
        return "; ".join(parts)

    tier1_enriched = []
    for ev in tier1_events:
        row = dict(ev)
        row['notable_differences'] = notable_differences(ev)
        row['current_known_issue_note'] = ev['_known_issue_note']
        tier1_enriched.append(row)

    tier1_cols = [
        'event_id', 'year', 'event_name', 'host_club',
        'known_issue_type', 'diff_summary', 'review_heat_score', 'review_heat_label',
        'raw_truth_excerpt', 'canonical_excerpt',
        'notable_differences', 'current_known_issue_note',
        'review_status', 'review_note', 'action_needed',
    ]
    tier1_widths = {
        'event_id': 14, 'year': 6, 'event_name': 35, 'host_club': 20,
        'raw_truth_excerpt': 40, 'canonical_excerpt': 40,
        'notable_differences': 50, 'current_known_issue_note': 50,
        'review_status': 14, 'review_note': 30, 'action_needed': 20,
    }
    write_sheet(wb, 'Tier1_Review', tier1_enriched, tier1_cols,
                col_widths=tier1_widths, heat_col='review_heat_label',
                wrap_cols=['raw_truth_excerpt', 'canonical_excerpt', 'notable_differences', 'current_known_issue_note'])
    print(f"  Tier1_Review: {len(tier1_enriched)} rows")

    # ------- Sheet: KnownIssues -------
    SEVERITY_CLASS = {'severe': 'DATA_LOSS', 'moderate': 'INTEGRITY', 'minor': 'QUALITY'}
    SEVERITY_NEXT  = {
        'severe':   'Investigate source; consider RESULTS_FILE_OVERRIDE',
        'moderate': 'Review duplicates; confirm or document as pool-play artifact',
        'minor':    'Confirm acceptable; mark resolved or defer',
    }

    # Build event lookup
    ev_lookup = {ev['event_id']: ev for ev in events}

    ki_rows = []
    with open(KNOWN_ISSUES_CSV, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            eid  = str(row['event_id'])
            sev  = row.get('severity', '')
            note = row.get('note', '')
            ev   = ev_lookup.get(eid, {})
            ki_rows.append({
                'event_id':               eid,
                'event_name':             ev.get('event_name', ''),
                'issue_class':            SEVERITY_CLASS.get(sev, sev),
                'severity':               sev,
                'issue_note':             note,
                'host_club':              ev.get('host_club', ''),
                'raw_truth_excerpt':      ev.get('raw_truth_excerpt', ''),
                'canonical_excerpt':      ev.get('canonical_excerpt', ''),
                'recommended_next_step':  SEVERITY_NEXT.get(sev, ''),
                'review_status':          '',
                'review_note':            '',
            })

    ki_cols = [
        'event_id', 'event_name', 'issue_class', 'severity', 'issue_note', 'host_club',
        'raw_truth_excerpt', 'canonical_excerpt', 'recommended_next_step',
        'review_status', 'review_note',
    ]
    ki_widths = {
        'event_id': 14, 'event_name': 35, 'issue_class': 14, 'severity': 10,
        'issue_note': 50, 'raw_truth_excerpt': 40, 'canonical_excerpt': 40,
        'recommended_next_step': 50, 'review_note': 30,
    }
    write_sheet(wb, 'KnownIssues', ki_rows, ki_cols, col_widths=ki_widths,
                wrap_cols=['issue_note', 'raw_truth_excerpt', 'canonical_excerpt', 'recommended_next_step'])
    print(f"  KnownIssues: {len(ki_rows)} rows")

    # ------- Sheet: Overrides -------
    ov_rows = []
    seen_eid_type = set()
    for ev in events:
        eid = ev['event_id']
        ov_list = ev['_events_override_rows']
        for obj in ov_list:
            ov_type = []
            if 'location' in obj:
                ov_type.append('location')
            if 'host_club' in obj:
                ov_type.append('host_club')
            reason = obj.get('reason', obj.get('source', ''))
            key = (eid, 'jsonl', reason)
            if key not in seen_eid_type:
                seen_eid_type.add(key)
                ov_rows.append({
                    'event_id':           eid,
                    'event_name':         ev['event_name'],
                    'year':               ev['year'],
                    'host_club':          ev['host_club'],
                    'override_type':      ', '.join(ov_type) if ov_type else 'metadata',
                    'override_source_file': 'events_overrides.jsonl',
                    'reason':             reason,
                    'review_status':      '',
                    'review_note':        '',
                })
        if ev['_in_rfo']:
            key = (eid, 'rfo')
            if key not in seen_eid_type:
                seen_eid_type.add(key)
                ov_rows.append({
                    'event_id':           eid,
                    'event_name':         ev['event_name'],
                    'year':               ev['year'],
                    'host_club':          ev['host_club'],
                    'override_type':      'results_file',
                    'override_source_file': '02_canonicalize_results.py',
                    'reason':             'Parser results override',
                    'review_status':      '',
                    'review_note':        '',
                })

    ov_cols = [
        'event_id', 'event_name', 'year', 'host_club',
        'override_type', 'override_source_file', 'reason',
        'review_status', 'review_note',
    ]
    ov_widths = {
        'event_id': 14, 'event_name': 35, 'year': 6, 'host_club': 20,
        'override_type': 16, 'override_source_file': 28, 'reason': 45,
        'review_note': 30,
    }
    write_sheet(wb, 'Overrides', ov_rows, ov_cols, col_widths=ov_widths,
                wrap_cols=['reason'])
    print(f"  Overrides: {len(ov_rows)} rows")

    # ------- Sheet: RecentFixes -------
    rf_rows = []
    for eid_str, fix in RECENT_FIXES.items():
        ev = ev_lookup.get(eid_str, {})
        rf_rows.append({
            'event_id':     eid_str,
            'event_name':   ev.get('event_name', ''),
            'year':         ev.get('year', ''),
            'host_club':    ev.get('host_club', ''),
            'fix_version':  fix['fix_version'],
            'fix_type':     fix['fix_type'],
            'files_changed': fix['files_changed'],
            'review_status': '',
            'review_note':  '',
        })
    rf_cols = [
        'event_id', 'event_name', 'year', 'host_club',
        'fix_version', 'fix_type', 'files_changed',
        'review_status', 'review_note',
    ]
    rf_widths = {
        'event_id': 14, 'event_name': 35, 'year': 6, 'host_club': 20,
        'fix_version': 10, 'fix_type': 45, 'files_changed': 55,
        'review_note': 30,
    }
    write_sheet(wb, 'RecentFixes', rf_rows, rf_cols, col_widths=rf_widths,
                wrap_cols=['fix_type', 'files_changed'])
    print(f"  RecentFixes: {len(rf_rows)} rows")

    # ------- Sheet: PerEvent_Text -------
    pet_events = [ev for ev in events if ev['priority_tier'] in ('Tier1', 'Tier2')]
    pet_rows   = []
    for ev in pet_events:
        pet_rows.append({
            'event_id':              ev['event_id'],
            'event_name':            ev['event_name'],
            'year':                  ev['year'],
            'host_club':             ev['host_club'],
            'raw_truth_full':        ev['raw_truth_full'],
            'canonical_render_full': ev['canonical_render_full'],
            'diff_notes':            ev['diff_notes'],
        })
    pet_cols = [
        'event_id', 'event_name', 'year', 'host_club',
        'raw_truth_full', 'canonical_render_full', 'diff_notes',
    ]
    pet_widths = {
        'event_id': 14, 'event_name': 35, 'year': 6, 'host_club': 20,
        'raw_truth_full': 50, 'canonical_render_full': 50, 'diff_notes': 30,
    }
    ws_pet = write_sheet(wb, 'PerEvent_Text', pet_rows, pet_cols,
                         col_widths=pet_widths,
                         wrap_cols=['raw_truth_full', 'canonical_render_full', 'diff_notes'])
    # Set row height hint
    for ri in range(2, ws_pet.max_row + 1):
        ws_pet.row_dimensions[ri].height = 80
    print(f"  PerEvent_Text: {len(pet_rows)} rows")

    # ------- Sheet: Instructions -------
    if 'Instructions' in wb.sheetnames:
        del wb['Instructions']
    ws_inst = wb.create_sheet('Instructions')

    def add_header(ws, row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=13, color="2E4057")
        return row + 1

    def add_text(ws, row, text, bold=False, indent=0):
        prefix = '    ' * indent
        cell = ws.cell(row=row, column=1, value=prefix + text)
        if bold:
            cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        return row + 1

    r = 1
    r = add_header(ws_inst, r, "FOOTBAG EVENT REVIEW PACKET — Instructions")
    r = add_text(ws_inst, r, "This workbook is for human review of event data quality. It is read-only with respect to canonical identity.")
    r += 1

    r = add_header(ws_inst, r, "Priority Tier Definitions")
    r = add_text(ws_inst, r, "Tier1: Events with known issues, overrides, recent modifications, or mismatched division/placement counts.", indent=1)
    r = add_text(ws_inst, r, "Tier2: Special-year events (1997–2003, 2024–2026) or Worlds/Mixed type events. Also HIGH/MEDIUM heat events.", indent=1)
    r = add_text(ws_inst, r, "Tier3: All other events with no detected anomalies.", indent=1)
    r += 1

    r = add_header(ws_inst, r, "review_status Values")
    for val, desc in [
        ('(blank)', 'Unreviewed'),
        ('ACCEPTED', 'Data is correct and acceptable as-is'),
        ('NEEDS_ACTION', 'A fix is required; fill action_needed and action_target'),
        ('DEFERRED', 'Defer to a future sprint; document reason in review_note'),
        ('DUPLICATE', 'Flag as duplicate of another event'),
    ]:
        r = add_text(ws_inst, r, f"  {val}: {desc}", indent=1)
    r += 1

    r = add_header(ws_inst, r, "action_needed Values")
    for val in ['NONE', 'PARSER_FIX', 'PBP_PATCH', 'RESULTS_FILE_OVERRIDE',
                'IDENTITY_FIX', 'METADATA_FIX', 'DOCUMENT_ONLY']:
        r = add_text(ws_inst, r, f"  {val}", indent=1)
    r += 1

    r = add_header(ws_inst, r, "Heat Score Breakdown")
    for pts, desc in [
        ('+4', 'known_issue_flag — event in known_issues.csv'),
        ('+3', 'override_flag — event has an override (results file or metadata)'),
        ('+3', 'div_mismatch_real — real division count mismatch (after normalization, not RFO, not stale PBP)'),
        ('+3', 'plc_mismatch_real — real placement count mismatch (not RFO, not contest format, not dedup-explained)'),
        ('+2', 'recently_modified_flag — touched in v2.6.0–v2.8.0'),
        ('+1', 'metadata_anomaly — missing location, host_club, or encoding issues'),
    ]:
        r = add_text(ws_inst, r, f"  {pts}: {desc}", indent=1)
    r = add_text(ws_inst, r, "Labels: HIGH ≥7, MEDIUM 4–6, LOW 1–3, CLEAN 0", indent=1)
    r += 1
    r = add_header(ws_inst, r, "Reconciliation Flags (suppress false HIGH)")
    for flag, desc in [
        ('rfo_divergence_expected', 'RFO event — stage2 vs PBP divergence is expected; div/plc mismatch suppressed'),
        ('pbp_stale_flag', 'Parser has been fixed but PBP was generated from older parse; div mismatch suppressed'),
        ('contest_format_flag', 'Division is Sick 3/Shred 30/Circle/Request/Battle — non-standard placements allowed'),
        ('s2_less_flag', 'PBP has MORE divisions than stage2 — investigate source divergence'),
        ('div_mismatch_real', '1 = genuine division mismatch after reconciliation'),
        ('plc_mismatch_real', '1 = genuine placement mismatch after reconciliation'),
    ]:
        r = add_text(ws_inst, r, f"  {flag}: {desc}", indent=1)
    r += 1

    r = add_header(ws_inst, r, "Fastest Review Workflow (5 Steps)")
    steps = [
        "1. Open Queue sheet. Sort by review_priority_rank, then year.",
        "2. For each HIGH/MEDIUM event: open Tier1_Review tab, check raw_truth_excerpt vs canonical_excerpt.",
        "3. If counts match and data looks correct: set review_status=ACCEPTED.",
        "4. If a fix is needed: set review_status=NEEDS_ACTION, fill action_needed.",
        "5. Use KnownIssues tab to track all severe/moderate issues systematically.",
    ]
    for step in steps:
        r = add_text(ws_inst, r, step, indent=1)

    ws_inst.column_dimensions['A'].width = 90
    ws_inst.freeze_panes = None
    print("  Instructions: static")

    # ------- Sheet: Quarantined -------
    quarantine_rows_for_sheet = []
    for ev in events:
        if ev.get('quarantine_flag'):
            quarantine_rows_for_sheet.append({
                'event_id':          ev['event_id'],
                'year':              ev['year'],
                'event_name':        ev['event_name'],
                'quarantine_reason': ev['quarantine_reason'],
                'diff_summary':      ev['diff_summary'],
                'known_issue_flag':  ev['known_issue_flag'],
                'quarantine_notes':  ev['quarantine_notes'],
                'review_status':     '',
                'review_note':       '',
            })
    q_cols = [
        'event_id', 'year', 'event_name', 'quarantine_reason',
        'diff_summary', 'known_issue_flag',
        'quarantine_notes', 'review_status', 'review_note',
    ]
    q_widths = {
        'event_id': 14, 'year': 6, 'event_name': 40, 'quarantine_reason': 28,
        'diff_summary': 22, 'quarantine_notes': 80, 'review_note': 30,
    }
    write_sheet(wb, 'Quarantined', quarantine_rows_for_sheet, q_cols,
                col_widths=q_widths,
                wrap_cols=['quarantine_notes', 'review_note'])
    print(f"  Quarantined: {len(quarantine_rows_for_sheet)} rows")

    # -----------------------------------------------------------------------
    # Save workbook
    # -----------------------------------------------------------------------
    print(f"\nSaving workbook to {WORKBOOK_PATH}...")
    wb.save(WORKBOOK_PATH)
    print("  Saved.")

    # -----------------------------------------------------------------------
    # Export adjudication subsets
    # -----------------------------------------------------------------------
    def _add_focus_rank(row):
        """Compute review_focus_rank for adjudication prioritization."""
        label = row.get('review_heat_label', '')
        div_mis = row.get('div_mismatch_real', 0)
        plc_mis = row.get('plc_mismatch_real', 0)
        if label == 'HIGH' and div_mis:
            return 1
        if label == 'HIGH' and plc_mis:
            return 2
        if label == 'HIGH' and row.get('override_flag', 0):
            return 3
        if label == 'HIGH' and row.get('known_issue_flag', 0):
            return 4
        if label == 'MEDIUM' and div_mis:
            return 5
        if label == 'MEDIUM' and plc_mis:
            return 6
        if label == 'HIGH':
            return 7
        if label == 'MEDIUM':
            return 8
        return 9

    def _count_mismatch_flag(row):
        return 1 if (row.get('div_mismatch_real', 0) or row.get('plc_mismatch_real', 0)) else 0

    def _medium_pattern_family(ev):
        """Derive pattern_family for MEDIUM events from diff_summary and flags."""
        summary = ev.get('diff_summary', '')
        if summary == 'override ok':
            return 'OVERRIDE_OK'
        if summary == 'pbp stale':
            return 'PBP_STALE'
        if summary == 'investigate: pbp>stage2':
            return 'INVESTIGATE_SOURCE'
        if summary in ('division mismatch', 'placement mismatch'):
            return 'DATA_MISMATCH'
        if summary == 'known issue':
            return 'KNOWN_ISSUE_MINOR'
        if summary == 'override used':
            return 'METADATA_OVERRIDE'
        if summary == 'recently modified':
            return 'RECENTLY_MODIFIED'
        if summary == 'metadata only':
            return 'METADATA_ONLY'
        if summary == 'contest format':
            return 'CONTEST_FORMAT'
        if summary == 'normalized match':
            return 'NORMALIZED_MATCH'
        return 'OTHER'

    adj_cols = queue_cols + ['review_focus_rank', 'count_mismatch_flag']

    # Queue_HIGH_ONLY: non-quarantined HIGH events only
    high_only = [ev for ev in events if ev['review_heat_label'] == 'HIGH']
    for ev in high_only:
        ev['review_focus_rank'] = _add_focus_rank(ev)
        ev['count_mismatch_flag'] = _count_mismatch_flag(ev)
    high_only.sort(key=lambda r: (r['review_focus_rank'], -r['review_heat_score']))
    wb_high = openpyxl.Workbook()
    ws_h = wb_high.active
    ws_h.title = 'Queue_HIGH_ONLY'
    ws_h.append(adj_cols)
    for ev in high_only:
        ws_h.append([ev.get(c, '') for c in adj_cols])
    HIGH_ONLY_PATH = OUT_DIR / "Queue_HIGH_ONLY.xlsx"
    wb_high.save(HIGH_ONLY_PATH)
    print(f"  Queue_HIGH_ONLY: {len(high_only)} rows → {HIGH_ONLY_PATH}")

    # Queue_HIGH_MEDIUM_MISMATCH: non-quarantined HIGH+MEDIUM with real data mismatch
    medium_mismatch = [
        ev for ev in events
        if ev['review_heat_label'] in ('HIGH', 'MEDIUM')
        and (ev.get('div_mismatch_real', 0) or ev.get('plc_mismatch_real', 0))
    ]
    for ev in medium_mismatch:
        if 'review_focus_rank' not in ev:
            ev['review_focus_rank'] = _add_focus_rank(ev)
            ev['count_mismatch_flag'] = _count_mismatch_flag(ev)
    medium_mismatch.sort(key=lambda r: (r.get('review_focus_rank', 9), -r['review_heat_score']))
    wb_mm = openpyxl.Workbook()
    ws_mm = wb_mm.active
    ws_mm.title = 'Queue_HIGH_MEDIUM_MISMATCH'
    ws_mm.append(adj_cols)
    for ev in medium_mismatch:
        ws_mm.append([ev.get(c, '') for c in adj_cols])
    MM_PATH = OUT_DIR / "Queue_HIGH_MEDIUM_MISMATCH.xlsx"
    wb_mm.save(MM_PATH)
    print(f"  Queue_HIGH_MEDIUM_MISMATCH: {len(medium_mismatch)} rows → {MM_PATH}")

    # Queue_MEDIUM_REVIEW: MEDIUM events grouped by pattern family (active review target)
    medium_review_cols = adj_cols + ['pattern_family']
    medium_events = [ev for ev in events if ev['review_heat_label'] == 'MEDIUM']
    for ev in medium_events:
        ev['pattern_family'] = _medium_pattern_family(ev)
        if 'review_focus_rank' not in ev:
            ev['review_focus_rank'] = _add_focus_rank(ev)
            ev['count_mismatch_flag'] = _count_mismatch_flag(ev)
    # Sort by pattern_family then year for batch review
    medium_events.sort(key=lambda r: (r['pattern_family'], r['year'] or 0))
    wb_med = openpyxl.Workbook()
    ws_med = wb_med.active
    ws_med.title = 'Queue_MEDIUM_REVIEW'
    ws_med.append(medium_review_cols)
    for ev in medium_events:
        ws_med.append([ev.get(c, '') for c in medium_review_cols])
    MED_PATH = OUT_DIR / "Queue_MEDIUM_REVIEW.xlsx"
    wb_med.save(MED_PATH)
    # Print pattern family breakdown
    pf_counts = defaultdict(int)
    for ev in medium_events:
        pf_counts[ev['pattern_family']] += 1
    print(f"  Queue_MEDIUM_REVIEW: {len(medium_events)} rows → {MED_PATH}")
    for pf, cnt in sorted(pf_counts.items()):
        print(f"    {pf}: {cnt}")

    # -----------------------------------------------------------------------
    # HTML Dashboard
    # -----------------------------------------------------------------------
    print(f"Building HTML dashboard...")

    high_events = [ev for ev in events if ev['review_heat_label'] == 'HIGH']

    # Year distribution
    year_heat = defaultdict(lambda: {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'CLEAN': 0, 'QUARANTINED': 0})
    for ev in events:
        yr = ev['year'] if ev['year'] else 'Unknown'
        year_heat[yr][ev['review_heat_label']] += 1

    years_sorted = sorted([k for k in year_heat.keys() if k != 'Unknown'], key=lambda x: int(x))
    if 'Unknown' in year_heat:
        years_sorted.append('Unknown')

    # Build year chart rows
    year_rows = []
    for yr in years_sorted:
        h = year_heat[yr]
        total = sum(h.values())
        bar_cells = []
        for label, color in [('HIGH', '#FF4444'), ('MEDIUM', '#FF8800'), ('LOW', '#FFDD00'), ('CLEAN', '#44BB44'), ('QUARANTINED', '#AAAAAA')]:
            if h[label] > 0:
                bar_cells.append(f'<td style="background:{color};width:{h[label]*8}px;min-width:4px;" title="{label}: {h[label]}">&nbsp;</td>')
        year_rows.append(f'<tr><td style="padding:2px 6px;font-size:12px">{yr}</td><td style="padding:2px 6px;font-size:12px">{total}</td><td>{"".join(bar_cells)}</td></tr>')

    # High events table
    high_rows = []
    for ev in high_events:
        name = ev['event_name'][:60]
        high_rows.append(
            f'<tr>'
            f'<td style="padding:3px 8px;font-size:12px">{ev["event_id"]}</td>'
            f'<td style="padding:3px 8px;font-size:12px">{ev["year"]}</td>'
            f'<td style="padding:3px 8px;font-size:12px">{name}</td>'
            f'<td style="padding:3px 8px;font-size:12px">{ev["review_heat_score"]}</td>'
            f'<td style="padding:3px 8px;font-size:12px">{ev["diff_summary"]}</td>'
            f'<td style="padding:3px 8px;font-size:12px">{ev["priority_tier"]}</td>'
            f'</tr>'
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Footbag Event Review Dashboard</title>
<style>
  body {{ font-family: Arial, sans-serif; background: #f4f4f4; margin: 20px; color: #222; }}
  h1 {{ color: #2E4057; }}
  h2 {{ color: #2E4057; margin-top: 32px; }}
  .summary-table {{ border-collapse: collapse; margin-bottom: 20px; }}
  .summary-table td, .summary-table th {{ border: 1px solid #ccc; padding: 6px 14px; text-align: center; }}
  .summary-table th {{ background: #2E4057; color: white; }}
  .high {{ background: #FF4444; color: white; font-weight: bold; }}
  .medium {{ background: #FF8800; color: white; font-weight: bold; }}
  .low {{ background: #FFDD00; color: black; font-weight: bold; }}
  .clean {{ background: #44BB44; color: white; font-weight: bold; }}
  .tier1 {{ background: #FF6666; color: white; }}
  .tier2 {{ background: #FFBB66; color: black; }}
  .tier3 {{ background: #AADDAA; color: black; }}
  table.data {{ border-collapse: collapse; width: 100%; }}
  table.data td, table.data th {{ border: 1px solid #ddd; padding: 4px 8px; font-size: 12px; }}
  table.data th {{ background: #2E4057; color: white; }}
  table.data tr:nth-child(even) {{ background: #f9f9f9; }}
</style>
</head>
<body>
<h1>Footbag Event Review Dashboard</h1>
<p>Generated 2026-03-11 | Total events: {len(events)}</p>

<h2>Summary Statistics</h2>
<table class="summary-table">
  <tr>
    <th>Total Events</th>
    <th class="high">HIGH</th>
    <th class="medium">MEDIUM</th>
    <th class="low">LOW</th>
    <th class="clean">CLEAN</th>
    <th style="background:#AAAAAA;color:white">QUARANTINED</th>
    <th class="tier1">Tier1</th>
    <th class="tier2">Tier2</th>
    <th class="tier3">Tier3</th>
  </tr>
  <tr>
    <td><strong>{len(events)}</strong></td>
    <td class="high">{heat_counts['HIGH']}</td>
    <td class="medium">{heat_counts['MEDIUM']}</td>
    <td class="low">{heat_counts['LOW']}</td>
    <td class="clean">{heat_counts['CLEAN']}</td>
    <td style="background:#AAAAAA;color:white;font-weight:bold">{heat_counts['QUARANTINED']}</td>
    <td class="tier1">{tier_counts['Tier1']}</td>
    <td class="tier2">{tier_counts['Tier2']}</td>
    <td class="tier3">{tier_counts['Tier3']}</td>
  </tr>
</table>

<h2>Heat Distribution by Year</h2>
<table style="border-collapse:collapse">
  <tr>
    <th style="padding:2px 6px;background:#2E4057;color:white;font-size:12px">Year</th>
    <th style="padding:2px 6px;background:#2E4057;color:white;font-size:12px">Count</th>
    <th style="padding:2px 6px;background:#2E4057;color:white;font-size:12px">Heat (HIGH/MEDIUM/LOW/CLEAN)</th>
  </tr>
  {''.join(year_rows)}
</table>

<h2>HIGH Priority Events ({len(high_events)})</h2>
<table class="data">
  <tr>
    <th>Event ID</th>
    <th>Year</th>
    <th>Event Name</th>
    <th>Heat Score</th>
    <th>Diff Summary</th>
    <th>Priority Tier</th>
  </tr>
  {''.join(high_rows) if high_rows else '<tr><td colspan="6">None</td></tr>'}
</table>

</body>
</html>
"""

    DASHBOARD_PATH.write_text(html, encoding='utf-8')
    print(f"  Saved HTML dashboard.")

    # Final summary
    print("\n=== DONE ===")
    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Dashboard: {DASHBOARD_PATH}")
    print(f"\nSheet row counts:")
    print(f"  Queue:         {len(events)}")
    print(f"  Tier1_Review:  {len(tier1_enriched)}")
    print(f"  KnownIssues:   {len(ki_rows)}")
    print(f"  Overrides:     {len(ov_rows)}")
    print(f"  RecentFixes:   {len(rf_rows)}")
    print(f"  PerEvent_Text: {len(pet_rows)}")
    print(f"\nHeat distribution:")
    for label in ('HIGH', 'MEDIUM', 'LOW', 'CLEAN', 'QUARANTINED'):
        print(f"  {label}: {heat_counts[label]}")

if __name__ == '__main__':
    main()
