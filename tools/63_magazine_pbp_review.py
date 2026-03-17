#!/usr/bin/env python3
"""
tools/63_magazine_pbp_review.py

Generate a review workbook for magazine placements to support PBP v68 creation.

Outputs out/review/magazine_pbp_review.xlsx with three sheets:
  Placements  — one row per (event, division, place, player) with auto-resolved
                person_ids and fuzzy-match candidates for unresolved names
  PT_Ref      — full Persons_Truth for lookup during review
  Instructions — column guide

Workflow:
  1. Run this tool → open out/review/magazine_pbp_review.xlsx
  2. Fill in 'decision_person_id' for unresolved rows (copy from PT_Ref)
  3. Use '__NON_PERSON__' for handles/non-real-persons (e.g. 'Big Al')
  4. Leave blank to keep auto-resolved entries unchanged
  5. Run tools/64_patch_pbp_v67_to_v68.py to generate PBP v68

Usage:
  .venv/bin/python tools/63_magazine_pbp_review.py
"""
from __future__ import annotations

import csv
import difflib
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

csv.field_size_limit(10_000_000)

ROOT    = Path(__file__).resolve().parents[1]
OUT     = ROOT / "out"
REVIEW  = OUT / "review"
INPUTS  = ROOT / "inputs" / "identity_lock"

STAGE2_CSV   = OUT / "stage2_canonical_events.csv"
PT_CSV       = INPUTS / "Persons_Truth_Final_v42.csv"
ALIASES_CSV  = ROOT / "overrides" / "person_aliases.csv"
OUT_XLSX     = REVIEW / "magazine_pbp_review.xlsx"

# ── Known manual overrides confirmed by the curator ──────────────────────────
# Format: norm(magazine_name) → (person_id_or_sentinel, person_canon, note)
# Use person_id="" with a canon to flag as known-unresolved (not yet in PT).
MANUAL_OVERRIDES: dict[str, tuple[str, str, str]] = {
    # Typo in magazine source → existing PT entry
    "bruce guettlich":  ("886e5b99-bd46-5707-a490-a2e938b09a35", "Bruce Guettich",  "magazine typo"),
    "garry griggs":     ("8b8c45bf-a622-55cd-bd2d-67f3b395873e", "Gary Griggs",      "Garry/Gary variant"),
    "john lind":        ("a6e260ae-74e5-5ef5-8b20-d7c2fbd65b83", "Jon Lind",         "John/Jon variant"),
    # Handle → real name (not yet in PT; will appear as person_unresolved=1)
    "big al":           ("", "Alan Cook", "handle for Alan Cook; add to PT in future v2.0.0"),
    # Same person appears under full name too — keep consistent canon
    "alan cook":        ("", "Alan Cook", "pre-mirror era competitor; add to PT in future v2.0.0"),
    # Maiden/married name variant → existing PT entry
    "jody badger":      ("e333a5e6-3d76-513c-a9a4-f34c09c31b96", "Jody Welch",    "Jody Badger = Jody Welch (maiden name)"),
    "tricia sullivan":  ("26349aa8-a1ff-5e6a-bff5-f93a89d20c68", "Tricia George", "Tricia Sullivan = Tricia George (maiden name)"),
    # Location artifact leaked into results — not a person
    "london england":   ("__NON_PERSON__", "__NON_PERSON__", "location annotation, not a person"),
}

REVIEW.mkdir(parents=True, exist_ok=True)


# ── Normalise name for fuzzy matching ─────────────────────────────────────────

def norm_name(s: str) -> str:
    """Lowercase, strip accents, collapse whitespace."""
    s = unicodedata.normalize("NFD", s.lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip()


def first_last(name: str) -> tuple[str, str]:
    parts = name.strip().split()
    if len(parts) >= 2:
        return parts[0].lower(), parts[-1].lower()
    return parts[0].lower() if parts else "", ""


# ── Load PT ───────────────────────────────────────────────────────────────────

pt_rows: list[dict] = []
token_to_pid: dict[str, str]  = {}   # stage2 UUID → person_id
token_to_canon: dict[str, str] = {}   # stage2 UUID → person_canon
name_to_pid: dict[str, str]   = {}   # norm(name) → person_id
name_to_canon: dict[str, str]  = {}   # norm(name) → person_canon

with open(PT_CSV, newline="", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        pid   = row["effective_person_id"].strip()
        canon = row["person_canon"].strip()
        pt_rows.append({"person_id": pid, "person_canon": canon, "notes": row.get("notes", "")})

        # Token map: player_ids_seen is pipe-separated list of stage2 UUIDs
        for tok in (row.get("player_ids_seen") or "").split("|"):
            tok = tok.strip()
            if tok:
                token_to_pid[tok]   = pid
                token_to_canon[tok] = canon

        # Name map: person_canon + all player_names_seen
        all_names = [canon] + [
            n.strip()
            for n in (row.get("player_names_seen") or "").split("|")
            if n.strip()
        ]
        for n in all_names:
            k = norm_name(n)
            if k and k not in name_to_pid:
                name_to_pid[k]   = pid
                name_to_canon[k] = canon

print(f"PT loaded: {len(pt_rows):,} persons, {len(token_to_pid):,} tokens, {len(name_to_pid):,} name variants")

# ── Augment name map from aliases CSV ─────────────────────────────────────────
_alias_loaded = 0
if ALIASES_CSV.exists():
    with open(ALIASES_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            alias  = (row.get("alias") or "").strip()
            pid    = (row.get("person_id") or "").strip()
            canon  = (row.get("person_canon") or "").strip()
            if alias and pid and canon:
                k = norm_name(alias)
                if k and k not in name_to_pid:
                    name_to_pid[k]   = pid
                    name_to_canon[k] = canon
                    _alias_loaded += 1
    print(f"Aliases loaded: {_alias_loaded:,} new name variants from {ALIASES_CSV.name}")

# Build sorted list of (norm_name, canon, person_id) for fuzzy search
_pt_sorted = sorted(name_to_pid.items())   # (norm_name, pid)
_pt_norm_list   = [k for k, _ in _pt_sorted]
_pt_pid_list    = [v for _, v in _pt_sorted]
_pt_canon_list  = [name_to_canon[k] for k, _ in _pt_sorted]


def resolve(token: str, raw_name: str) -> tuple[str, str, str]:
    """Return (person_id, person_canon, source_note) or ("", "", "") if unresolved.

    person_id may be "" even on a successful match when the match is a known
    unresolved identity (e.g. Big Al → Alan Cook, not yet in PT).  In that case
    person_canon is set so the PBP row uses the correct display name.
    source_note indicates where the match came from.
    """
    # 1. Manual override (curator-confirmed)
    k = norm_name(raw_name)
    if k in MANUAL_OVERRIDES:
        pid, canon, note = MANUAL_OVERRIDES[k]
        return pid, canon, f"manual: {note}"
    # 2. Stage2 token → PT
    if token and token in token_to_pid:
        return token_to_pid[token], token_to_canon[token], "token"
    # 3. Name → PT (includes aliases CSV augmentation)
    if k in name_to_pid:
        return name_to_pid[k], name_to_canon[k], "name"
    return "", "", ""


def fuzzy_candidates(raw_name: str, n: int = 5) -> list[tuple[str, str, float]]:
    """Return top-n (canon, person_id, score) fuzzy matches from PT."""
    k = norm_name(raw_name)
    # Exact first/last name match gets priority
    first, last = first_last(raw_name)
    results: list[tuple[float, str, str]] = []
    for pt_norm, pid, canon in zip(_pt_norm_list, _pt_pid_list, _pt_canon_list):
        score = difflib.SequenceMatcher(None, k, pt_norm).ratio()
        # Boost if last name matches exactly
        pt_first, pt_last = first_last(canon)
        if last and pt_last == last:
            score = max(score, 0.70)
        if first and pt_first == first:
            score += 0.05
        results.append((score, canon, pid))
    results.sort(reverse=True)
    return [(c, p, s) for s, c, p in results[:n] if s > 0.35]


# ── Load stage2 magazine placements ───────────────────────────────────────────

rows: list[dict] = []

with open(STAGE2_CSV, newline="", encoding="utf-8") as f:
    for ev in csv.DictReader(f):
        eid = ev["event_id"]
        if len(eid) != 7:          # magazine IDs are exactly 7 digits
            continue
        year       = ev.get("year", "")
        event_name = ev.get("event_name", "")
        pj_raw     = ev.get("placements_json") or "[]"
        try:
            placements = json.loads(pj_raw)
        except Exception:
            placements = []

        for p in placements:
            div_canon  = p.get("division_canon", "")
            div_cat    = p.get("division_category", "")
            div_raw    = p.get("division_raw", "")
            place      = str(p.get("place", ""))
            comp_type  = p.get("competitor_type", "player")
            p1_name    = (p.get("player1_name") or "").strip()
            p2_name    = (p.get("player2_name") or "").strip()
            p1_token   = (p.get("player1_id") or "").strip()
            p2_token   = (p.get("player2_id") or "").strip()

            players = [(p1_name, p1_token)]
            if p2_name:
                players.append((p2_name, p2_token))

            for player_name, player_token in players:
                if not player_name:
                    continue

                auto_pid, auto_canon, resolve_note = resolve(player_token, player_name)
                # __NON_PERSON__: location/noise, treat as resolved (no review needed)
                is_non_person = (auto_pid == "__NON_PERSON__")
                # "known-unresolved": manual override supplied a canon name but no PT ID yet
                is_known_unresolved = (not auto_pid and bool(auto_canon) and not is_non_person)
                is_unresolved = not auto_pid and not auto_canon and not is_non_person

                cands_str = ""
                if is_known_unresolved:
                    cands_str = f"[known alias] {resolve_note}"
                elif is_unresolved:
                    cands = fuzzy_candidates(player_name)
                    cands_str = " | ".join(
                        f"{c} ({s:.0%})" for c, p_id, s in cands
                    )

                rows.append({
                    "event_id":           eid,
                    "year":               year,
                    "event_name":         event_name,
                    "division_canon":     div_canon,
                    "division_category":  div_cat,
                    "division_raw":       div_raw,
                    "place":              place,
                    "competitor_type":    comp_type,
                    "player_name":        player_name,
                    "player_token":       player_token,
                    # Auto-resolved
                    "auto_person_id":     auto_pid,
                    "auto_person_canon":  auto_canon,
                    # User fills in:
                    "decision_person_id": "",
                    "decision_notes":     "",
                    # Fuzzy candidates (for unresolved only)
                    "fuzzy_candidates":   cands_str,
                    "_unresolved":        is_unresolved,
                    "_known_unresolved":  is_known_unresolved,
                    "_non_person":        is_non_person,
                })

n_total           = len(rows)
n_unresolved      = sum(1 for r in rows if r["_unresolved"])
n_known_unresolved= sum(1 for r in rows if r["_known_unresolved"])
n_non_person      = sum(1 for r in rows if r["_non_person"])
n_resolved        = n_total - n_unresolved - n_known_unresolved - n_non_person
print(f"Magazine placements: {n_total} total")
print(f"  {n_resolved} auto-resolved (PT match)")
print(f"  {n_known_unresolved} known-unresolved (manual alias, not yet in PT)")
print(f"  {n_non_person} non-person (__NON_PERSON__ — location/noise artifacts)")
print(f"  {n_unresolved} truly unresolved (need review)")


# ── Write workbook ────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── Sheet 1: Placements ────────────────────────────────────────────────────

ws = wb.active
ws.title = "Placements"

HEADER = [
    "event_id", "year", "event_name", "division_canon", "division_category",
    "place", "competitor_type", "player_name",
    "auto_person_id", "auto_person_canon",
    "decision_person_id", "decision_notes",
    "fuzzy_candidates",
]

HDR_STYLE   = Font(bold=True)
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")   # resolved (PT match)
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")   # known-unresolved (manual alias, not in PT yet)
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")   # truly unresolved
PINK_FILL   = PatternFill("solid", fgColor="FFC7CE")   # header

for col_idx, col in enumerate(HEADER, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center")

for r_idx, r in enumerate(rows, start=2):
    is_unresolved       = r["_unresolved"]
    is_known_unresolved = r["_known_unresolved"]
    row_fill = YELLOW_FILL if is_unresolved else (ORANGE_FILL if is_known_unresolved else GREEN_FILL)
    values = [
        r["event_id"], r["year"], r["event_name"], r["division_canon"],
        r["division_category"], r["place"], r["competitor_type"], r["player_name"],
        r["auto_person_id"], r["auto_person_canon"],
        r["decision_person_id"], r["decision_notes"],
        r["fuzzy_candidates"],
    ]
    for c_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        # Highlight unresolved rows; shade decision columns
        if c_idx in (9, 10):   # auto columns
            if not is_unresolved and not is_known_unresolved:
                cell.fill = GREEN_FILL
            elif is_known_unresolved:
                cell.fill = ORANGE_FILL
        if c_idx in (11, 12):  # decision columns
            if is_unresolved:
                cell.fill = YELLOW_FILL
            elif is_known_unresolved:
                cell.fill = ORANGE_FILL
            else:
                cell.fill = PatternFill("solid", fgColor="EBF3FB")
        if c_idx == 13:        # fuzzy_candidates
            cell.alignment = Alignment(wrap_text=True)

# Column widths
col_widths = [12, 6, 40, 32, 16, 6, 14, 28, 38, 28, 38, 20, 55]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ── Sheet 2: PT Reference ─────────────────────────────────────────────────

ws2 = wb.create_sheet("PT_Ref")
pt_cols = ["person_id", "person_canon", "notes"]
for c_idx, col in enumerate(pt_cols, start=1):
    cell = ws2.cell(row=1, column=c_idx, value=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="7030A0")

for r_idx, r in enumerate(sorted(pt_rows, key=lambda x: x["person_canon"]), start=2):
    ws2.cell(row=r_idx, column=1, value=r["person_id"])
    ws2.cell(row=r_idx, column=2, value=r["person_canon"])
    ws2.cell(row=r_idx, column=3, value=r["notes"])

ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 40
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

# ── Sheet 3: Instructions ─────────────────────────────────────────────────

ws3 = wb.create_sheet("Instructions")
instructions = [
    ["Column", "Description"],
    ["event_id", "Magazine synthetic event ID (7 digits, 99xxxxx)"],
    ["year", "Event year (after auto-correction for retrospective articles)"],
    ["event_name", "Canonical event name from stage2"],
    ["division_canon", "Canonical division name (Division: prefix stripped)"],
    ["division_category", "freestyle / net / golf / unknown"],
    ["place", "Finishing placement (integer)"],
    ["competitor_type", "player (singles) or team (doubles)"],
    ["player_name", "Player name as it appeared in the magazine source"],
    ["auto_person_id", "Person ID auto-resolved from PT (via token or name match)"],
    ["auto_person_canon", "Canonical name from PT for the auto-resolved ID"],
    ["decision_person_id", "FILL IN: PT person_id to assign (or __NON_PERSON__ for handles/non-persons)"],
    ["decision_notes", "Optional notes on your decision"],
    ["fuzzy_candidates", "Top PT name matches for unresolved players (format: Name (score%) | ...)"],
    ["", ""],
    ["HOW TO USE", ""],
    ["1", "Green rows = auto-resolved. Verify the auto_person_canon looks right."],
    ["2", "Yellow rows = unresolved. Check fuzzy_candidates column for suggestions."],
    ["3", "For unresolved: copy person_id from PT_Ref sheet into decision_person_id."],
    ["4", "Use __NON_PERSON__ for handles like 'Big Al', place-holders, non-real-persons."],
    ["5", "Leave decision_person_id blank to accept the auto-resolved value (green rows)."],
    ["6", "When done, run: .venv/bin/python tools/64_patch_pbp_v67_to_v68.py"],
    ["", ""],
    ["NOTES", ""],
    ["•", "Doubles: each player appears as a separate row with the same place."],
    ["•", "Kenny Shults → look for 'Kenneth Shults' in PT_Ref (auto-resolved via token)."],
    ["•", "The patch script will produce PBP v68 with all magazine rows appended."],
    ["•", "Magazine placements get coverage_flag='sparse' (pre-mirror era, partial sources)."],
]
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 70
for r_idx, (col_a, col_b) in enumerate(instructions, start=1):
    ws3.cell(row=r_idx, column=1, value=col_a)
    ws3.cell(row=r_idx, column=2, value=col_b)
    if col_a in ("Column", "HOW TO USE", "NOTES"):
        ws3.cell(row=r_idx, column=1).font = Font(bold=True)
        ws3.cell(row=r_idx, column=2).font = Font(bold=True)

wb.save(OUT_XLSX)
print(f"Wrote: {OUT_XLSX}")
print(f"  {n_resolved} auto-resolved (green) — PT match via token or name")
print(f"  {n_known_unresolved} known-unresolved (orange) — manual alias, person_canon set, not yet in PT")
print(f"  {n_non_person} non-person (green/__NON_PERSON__) — location artifacts")
print(f"  {n_unresolved} truly unresolved (yellow) — need your decision")
print()
print("Next step: open the workbook, fill in decision_person_id for yellow rows,")
print("then run: .venv/bin/python tools/64_patch_pbp_v67_to_v68.py")
