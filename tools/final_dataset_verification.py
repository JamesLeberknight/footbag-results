#!/usr/bin/env python3
"""
final_dataset_verification.py
Part 2: Source vs Canonical Comparison

Compares stage2 placements (source) against Placements_Flat.csv (canonical)
to identify missing divisions, count mismatches, and placement max mismatches.

Outputs to out/final_verification/
"""

import csv
import json
import os
import sys
import re
from collections import defaultdict
from difflib import SequenceMatcher

csv.field_size_limit(sys.maxsize)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE_DIR, "out", "final_verification")
os.makedirs(OUT_DIR, exist_ok=True)

STAGE2_CSV = os.path.join(BASE_DIR, "out", "stage2_canonical_events.csv")
PF_CSV = os.path.join(BASE_DIR, "out", "Placements_Flat.csv")
QUARANTINE_CSV = os.path.join(BASE_DIR, "inputs", "review_quarantine_events.csv")
COVERAGE_CSV = os.path.join(BASE_DIR, "out", "Coverage_ByEventDivision.csv")

# ── helpers ─────────────────────────────────────────────────────────────────

def normalize_div(name: str) -> str:
    """Normalize division name for fuzzy comparison."""
    n = name.lower().strip()
    n = re.sub(r'[:\-–—/]', ' ', n)
    n = re.sub(r'\s+', ' ', n)
    n = n.strip()
    return n


def div_similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, normalize_div(a), normalize_div(b)).ratio()


def find_best_match(div: str, candidates: list, threshold: float = 0.75):
    """Return (best_match, score) or (None, 0.0) if below threshold."""
    best = None
    best_score = 0.0
    for c in candidates:
        s = div_similarity(div, c)
        if s > best_score:
            best_score = s
            best = c
    if best_score >= threshold:
        return best, best_score
    return None, best_score


# ── load data ────────────────────────────────────────────────────────────────

print("Loading quarantine list...")
quarantine_ids = set()
with open(QUARANTINE_CSV, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        quarantine_ids.add(row["event_id"])
print(f"  {len(quarantine_ids)} quarantined events")

print("Loading stage2 canonical events...")
# source_data: {event_id: {division: {"count": N, "max_place": M, "event_name": ..., "year": ...}}}
source_data = {}
event_meta = {}  # event_id -> {name, year}

with open(STAGE2_CSV, encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        eid = row["event_id"]
        year = row.get("year", "")
        name = row.get("event_name", "")
        event_meta[eid] = {"name": name, "year": year}

        placements_raw = row.get("placements_json", "")
        if not placements_raw:
            continue
        try:
            placements = json.loads(placements_raw)
        except Exception:
            continue

        if not isinstance(placements, list):
            continue

        div_data = defaultdict(list)
        for p in placements:
            div = p.get("division_canon") or p.get("division_raw", "")
            place = p.get("place", 0)
            div_data[div].append(place)

        source_data[eid] = {}
        for div, places in div_data.items():
            source_data[eid][div] = {
                "count": len(places),
                "max_place": max(places) if places else 0,
            }

print(f"  {len(source_data)} events with placements in stage2")

print("Loading Placements_Flat (canonical)...")
# canonical_data: {event_id: {division: {"count": N, "max_place": M}}}
canonical_data = defaultdict(lambda: defaultdict(list))

with open(PF_CSV, encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        eid = row["event_id"]
        div = row.get("division_canon", "")
        try:
            place = int(row.get("place", 0))
        except ValueError:
            place = 0
        canonical_data[eid][div].append(place)

# Convert to summary
canonical_summary = {}
for eid, divs in canonical_data.items():
    canonical_summary[eid] = {}
    for div, places in divs.items():
        canonical_summary[eid][div] = {
            "count": len(places),
            "max_place": max(places) if places else 0,
        }

print(f"  {len(canonical_summary)} events with placements in PF")

# ── load SOURCE_PARTIAL status from EVENT INDEX (workbook) via Coverage file ─
# We'll mark events that are SOURCE_PARTIAL from the Coverage_ByEventDivision
# (doesn't have data_status) — instead we rely on event_meta patterns and known-issues
# Check if there's a separate source_status file
source_partial_events = set()
metadata_only_events = set()

# Try to load from the existing workbook EVENT INDEX (has Data Status col)
import openpyxl
wb_path = os.path.join(BASE_DIR, "Footbag_Results_Community_FINAL.xlsx")
if os.path.exists(wb_path):
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    if "EVENT INDEX" in wb.sheetnames:
        ws = wb["EVENT INDEX"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            eid_col = headers.index("Event ID")
            status_col = headers.index("Data Status")
            for row in ws.iter_rows(min_row=2, values_only=True):
                eid = str(row[eid_col]) if row[eid_col] else ""
                status = str(row[status_col]) if row[status_col] else ""
                if "SOURCE_PARTIAL" in status:
                    source_partial_events.add(eid)
                elif "METADATA_ONLY" in status:
                    metadata_only_events.add(eid)
        except (ValueError, IndexError):
            pass
    wb.close()

print(f"  {len(source_partial_events)} SOURCE_PARTIAL events, {len(metadata_only_events)} METADATA_ONLY events")

# ── comparison ───────────────────────────────────────────────────────────────

missing_divisions = []      # div in source but not in canonical (non-quarantined)
count_mismatches = []       # count differs between source and canonical
max_mismatches = []         # max_place differs

stats = {
    "total_source_event_divs": 0,
    "quarantined_skipped": 0,
    "metadata_only_skipped": 0,
    "source_partial_flagged": 0,
    "exact_match": 0,
    "drift_match": 0,
    "missing_blocker": 0,
    "missing_drift": 0,
    "count_mismatch": 0,
    "max_mismatch": 0,
}

print("Comparing source vs canonical...")

for eid, source_divs in sorted(source_data.items()):
    if eid in quarantine_ids:
        stats["quarantined_skipped"] += len(source_divs)
        continue
    if eid in metadata_only_events:
        stats["metadata_only_skipped"] += len(source_divs)
        continue

    is_partial = eid in source_partial_events
    meta = event_meta.get(eid, {})

    canonical_divs = canonical_summary.get(eid, {})
    canonical_div_names = list(canonical_divs.keys())

    for div, src_info in source_divs.items():
        stats["total_source_event_divs"] += 1
        if is_partial:
            stats["source_partial_flagged"] += 1

        # Check exact match
        if div in canonical_divs:
            stats["exact_match"] += 1
            can_info = canonical_divs[div]
            # Count mismatch check
            if src_info["count"] != can_info["count"]:
                stats["count_mismatch"] += 1
                count_mismatches.append({
                    "event_id": eid,
                    "year": meta.get("year", ""),
                    "event_name": meta.get("name", ""),
                    "division": div,
                    "source_count": src_info["count"],
                    "canonical_count": can_info["count"],
                    "source_max_place": src_info["max_place"],
                    "canonical_max_place": can_info["max_place"],
                    "is_source_partial": is_partial,
                })
            if src_info["max_place"] != can_info["max_place"]:
                stats["max_mismatch"] += 1
                max_mismatches.append({
                    "event_id": eid,
                    "year": meta.get("year", ""),
                    "event_name": meta.get("name", ""),
                    "division": div,
                    "source_max_place": src_info["max_place"],
                    "canonical_max_place": can_info["max_place"],
                    "source_count": src_info["count"],
                    "canonical_count": can_info["count"],
                    "is_source_partial": is_partial,
                })
        else:
            # Try fuzzy match
            best_match, score = find_best_match(div, canonical_div_names)
            if best_match is not None:
                stats["drift_match"] += 1
                # Check counts with the drift match
                can_info = canonical_divs[best_match]
                if src_info["count"] != can_info["count"]:
                    stats["count_mismatch"] += 1
                    count_mismatches.append({
                        "event_id": eid,
                        "year": meta.get("year", ""),
                        "event_name": meta.get("name", ""),
                        "division": div,
                        "canonical_div_matched": best_match,
                        "match_score": round(score, 3),
                        "source_count": src_info["count"],
                        "canonical_count": can_info["count"],
                        "source_max_place": src_info["max_place"],
                        "canonical_max_place": can_info["max_place"],
                        "is_source_partial": is_partial,
                    })
                if src_info["max_place"] != can_info["max_place"]:
                    stats["max_mismatch"] += 1
                    max_mismatches.append({
                        "event_id": eid,
                        "year": meta.get("year", ""),
                        "event_name": meta.get("name", ""),
                        "division": div,
                        "canonical_div_matched": best_match,
                        "match_score": round(score, 3),
                        "source_max_place": src_info["max_place"],
                        "canonical_max_place": can_info["max_place"],
                        "source_count": src_info["count"],
                        "canonical_count": can_info["count"],
                        "is_source_partial": is_partial,
                    })
            else:
                # Genuinely missing
                blocker_type = "DRIFT" if canonical_div_names else "BLOCKER"
                if not canonical_div_names:
                    # No canonical divisions at all for this event
                    blocker_type = "BLOCKER"
                else:
                    blocker_type = "BLOCKER" if not is_partial else "SOURCE_PARTIAL_MISSING"

                if blocker_type == "BLOCKER":
                    stats["missing_blocker"] += 1
                else:
                    stats["missing_drift"] += 1

                missing_divisions.append({
                    "event_id": eid,
                    "year": meta.get("year", ""),
                    "event_name": meta.get("name", ""),
                    "division": div,
                    "source_count": src_info["count"],
                    "source_max_place": src_info["max_place"],
                    "best_fuzzy_match": best_match or "",
                    "best_fuzzy_score": round(score, 3) if score else 0,
                    "canonical_div_count": len(canonical_div_names),
                    "is_source_partial": is_partial,
                    "blocker_type": blocker_type,
                })

# ── write reports ─────────────────────────────────────────────────────────────

print("Writing reports...")

# missing_divisions_report.csv
missing_path = os.path.join(OUT_DIR, "missing_divisions_report.csv")
with open(missing_path, "w", newline="", encoding="utf-8") as f:
    if missing_divisions:
        writer = csv.DictWriter(f, fieldnames=list(missing_divisions[0].keys()))
        writer.writeheader()
        writer.writerows(missing_divisions)
    else:
        f.write("# No missing divisions found\n")

# division_count_mismatch.csv
count_path = os.path.join(OUT_DIR, "division_count_mismatch.csv")
with open(count_path, "w", newline="", encoding="utf-8") as f:
    if count_mismatches:
        # Add canonical_div_matched col for all rows
        for r in count_mismatches:
            r.setdefault("canonical_div_matched", r.get("division", ""))
            r.setdefault("match_score", 1.0)
        writer = csv.DictWriter(f, fieldnames=list(count_mismatches[0].keys()))
        writer.writeheader()
        writer.writerows(count_mismatches)
    else:
        f.write("# No count mismatches found\n")

# placement_max_mismatch.csv
max_path = os.path.join(OUT_DIR, "placement_max_mismatch.csv")
with open(max_path, "w", newline="", encoding="utf-8") as f:
    if max_mismatches:
        for r in max_mismatches:
            r.setdefault("canonical_div_matched", r.get("division", ""))
            r.setdefault("match_score", 1.0)
        writer = csv.DictWriter(f, fieldnames=list(max_mismatches[0].keys()))
        writer.writeheader()
        writer.writerows(max_mismatches)
    else:
        f.write("# No max_place mismatches found\n")

# final_validation_report.md
report_path = os.path.join(OUT_DIR, "final_validation_report.md")
blocker_rows = [r for r in missing_divisions if r["blocker_type"] == "BLOCKER"]
partial_missing = [r for r in missing_divisions if r["blocker_type"] == "SOURCE_PARTIAL_MISSING"]

with open(report_path, "w", encoding="utf-8") as f:
    f.write("# Final Dataset Verification Report\n\n")
    f.write("## Summary Statistics\n\n")
    f.write(f"| Metric | Count |\n")
    f.write(f"|--------|-------|\n")
    f.write(f"| Total source event-divisions evaluated | {stats['total_source_event_divs']:,} |\n")
    f.write(f"| Quarantined (skipped) | {stats['quarantined_skipped']:,} |\n")
    f.write(f"| METADATA_ONLY (skipped) | {stats['metadata_only_skipped']:,} |\n")
    f.write(f"| SOURCE_PARTIAL (flagged separately) | {stats['source_partial_flagged']:,} |\n")
    f.write(f"| Exact division match | {stats['exact_match']:,} |\n")
    f.write(f"| Fuzzy/drift match | {stats['drift_match']:,} |\n")
    f.write(f"| Missing — BLOCKER (no match) | {stats['missing_blocker']:,} |\n")
    f.write(f"| Missing — SOURCE_PARTIAL | {len(partial_missing):,} |\n")
    f.write(f"| Count mismatches | {stats['count_mismatch']:,} |\n")
    f.write(f"| Max-place mismatches | {stats['max_mismatch']:,} |\n")
    f.write(f"\n")

    f.write("## BLOCKER Divisions (genuine gaps)\n\n")
    if not blocker_rows:
        f.write("**PASS — No genuine BLOCKER divisions found.**\n\n")
    else:
        f.write(f"**{len(blocker_rows)} BLOCKER(s) found:**\n\n")
        f.write("| Event ID | Year | Event Name | Division | Source Count | Source Max |\n")
        f.write("|----------|------|------------|----------|--------------|------------|\n")
        for r in blocker_rows:
            f.write(f"| {r['event_id']} | {r['year']} | {r['event_name'][:50]} | {r['division'][:40]} | {r['source_count']} | {r['source_max_place']} |\n")
        f.write("\n")

    f.write("## Count Mismatches (source vs canonical)\n\n")
    if not count_mismatches:
        f.write("**PASS — No count mismatches found.**\n\n")
    else:
        f.write(f"**{len(count_mismatches)} count mismatch(es):**\n\n")
        f.write("| Event ID | Year | Division | Src Count | Can Count | Δ | SOURCE_PARTIAL? |\n")
        f.write("|----------|------|----------|-----------|-----------|---|----------------|\n")
        for r in sorted(count_mismatches, key=lambda x: abs(x['source_count'] - x['canonical_count']), reverse=True)[:30]:
            delta = r['source_count'] - r['canonical_count']
            f.write(f"| {r['event_id']} | {r['year']} | {r['division'][:35]} | {r['source_count']} | {r['canonical_count']} | {delta:+d} | {'Y' if r['is_source_partial'] else 'N'} |\n")
        if len(count_mismatches) > 30:
            f.write(f"\n*... and {len(count_mismatches)-30} more (see division_count_mismatch.csv)*\n")
        f.write("\n")

    f.write("## Max-Place Mismatches (source vs canonical)\n\n")
    if not max_mismatches:
        f.write("**PASS — No max-place mismatches found.**\n\n")
    else:
        f.write(f"**{len(max_mismatches)} max-place mismatch(es):**\n\n")
        f.write("| Event ID | Year | Division | Src Max | Can Max | SOURCE_PARTIAL? |\n")
        f.write("|----------|------|----------|---------|---------|----------------|\n")
        for r in sorted(max_mismatches, key=lambda x: abs(x['source_max_place'] - x['canonical_max_place']), reverse=True)[:30]:
            f.write(f"| {r['event_id']} | {r['year']} | {r['division'][:35]} | {r['source_max_place']} | {r['canonical_max_place']} | {'Y' if r['is_source_partial'] else 'N'} |\n")
        if len(max_mismatches) > 30:
            f.write(f"\n*... and {len(max_mismatches)-30} more (see placement_max_mismatch.csv)*\n")
        f.write("\n")

    f.write("## Conclusion\n\n")
    if stats["missing_blocker"] == 0:
        f.write("**SOURCE_COVERAGE: PASS** — No genuine BLOCKER divisions. All source data is accounted for in canonical PF (either matched, quarantined, or documented as SOURCE_PARTIAL).\n")
    else:
        f.write(f"**SOURCE_COVERAGE: FAIL** — {stats['missing_blocker']} genuine BLOCKER division(s) found. See BLOCKER section above.\n")

print(f"\n=== RESULTS ===")
print(f"Total source event-divisions: {stats['total_source_event_divs']:,}")
print(f"Quarantined (skipped):        {stats['quarantined_skipped']:,}")
print(f"Metadata only (skipped):      {stats['metadata_only_skipped']:,}")
print(f"Exact match:                  {stats['exact_match']:,}")
print(f"Drift/fuzzy match:            {stats['drift_match']:,}")
print(f"Missing BLOCKER:              {stats['missing_blocker']:,}")
print(f"Missing SOURCE_PARTIAL:       {len(partial_missing):,}")
print(f"Count mismatches:             {stats['count_mismatch']:,}")
print(f"Max-place mismatches:         {stats['max_mismatch']:,}")
print(f"\nReports written to: {OUT_DIR}")
if stats["missing_blocker"] == 0:
    print("VERDICT: SOURCE_COVERAGE PASS — no genuine blockers")
else:
    print(f"VERDICT: SOURCE_COVERAGE FAIL — {stats['missing_blocker']} blockers")
    print("Blocker events:")
    for r in blocker_rows[:10]:
        print(f"  {r['event_id']} {r['year']} {r['event_name'][:40]} | div={r['division']!r}")
