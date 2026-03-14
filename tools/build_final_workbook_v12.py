#!/usr/bin/env python3
"""
build_final_workbook_v12.py

Changes from v11:
  - PLAYER SUMMARY  rebuilt with honors + freestyle enrichment columns
  - HONORS          new sheet  (FBHOF section + BAP section)
  - FREESTYLE INSIGHTS  new sheet  (trick stats, difficulty, transitions)
  - Year sheets     column widths standardised (data/formatting untouched)

New PLAYER SUMMARY columns (appended after existing core columns):
  years_active, honors_bap, honors_fbhof,
  freestyle_chains, max_sequence_add, distinct_tricks, most_common_trick

Derivation:
  years_active      ← min/max year in Placements_Flat per person
  honors_bap        ← bap_data_updated.csv  →  "BAP #N (YYYY)"
  honors_fbhof      ← fbhof_data_updated.csv →  "FBHOF YYYY" or "FBHOF"
  freestyle_chains  ← player_difficulty_profiles.csv  chains_total
  max_sequence_add  ← player_difficulty_profiles.csv  max_sequence_add
  distinct_tricks   ← player_difficulty_profiles.csv  n_distinct_tricks
  most_common_trick ← player_diversity_profiles.csv   top_tricks (first)
"""

import copy
import csv
import os
import sys
import unicodedata
from collections import Counter, defaultdict

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

csv.field_size_limit(sys.maxsize)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── Paths ─────────────────────────────────────────────────────────────────────

SOURCE_V11       = os.path.join(BASE_DIR, "Footbag_Results_Community_FINAL_v11.xlsx")
OUTPUT_PATH      = os.path.join(BASE_DIR, "Footbag_Results_Community_FINAL_v12.xlsx")

PF_CSV           = os.path.join(BASE_DIR, "out", "Placements_Flat.csv")
PT_CSV           = os.path.join(BASE_DIR, "out", "Persons_Truth.csv")
EVENTS_CSV       = os.path.join(BASE_DIR, "out", "canonical", "events_normalized.csv")
QUARANTINE_CSV   = os.path.join(BASE_DIR, "inputs", "review_quarantine_events.csv")

BAP_CSV          = os.path.join(BASE_DIR, "inputs", "bap_data_updated.csv")
FBHOF_CSV        = os.path.join(BASE_DIR, "inputs", "fbhof_data_updated.csv")

DIFFICULTY_CSV   = os.path.join(BASE_DIR, "out", "noise_aggregates", "player_difficulty_profiles.csv")
DIVERSITY_CSV    = os.path.join(BASE_DIR, "out", "noise_aggregates", "player_diversity_profiles.csv")
TRICK_FREQ_CSV   = os.path.join(BASE_DIR, "out", "noise_aggregates", "trick_frequency.csv")
TRANSITIONS_CSV  = os.path.join(BASE_DIR, "out", "noise_aggregates", "trick_transition_network.csv")
DIFF_YEAR_CSV    = os.path.join(BASE_DIR, "out", "noise_aggregates", "difficulty_by_year.csv")
INNOVATION_CSV   = os.path.join(BASE_DIR, "out", "noise_aggregates", "trick_innovation_timeline.csv")
TRICK_NODE_CSV        = os.path.join(BASE_DIR, "out", "noise_aggregates", "trick_node_metrics.csv")
SEQ_DIFF_CSV          = os.path.join(BASE_DIR, "out", "noise_aggregates", "sequence_difficulty_conservative.csv")
CHAIN_COMPLEXITY_CSV  = os.path.join(BASE_DIR, "out", "noise_aggregates", "chain_complexity_by_year.csv")


# ── Styles ────────────────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

FILL_NONE    = PatternFill(fill_type=None)
FILL_HEADER  = _fill("D9D9D9")
FILL_SECTION = _fill("E8F0FE")   # light blue for section titles
FILL_HOF     = _fill("FFF2CC")   # pale gold for FBHOF rows
FILL_BAP     = _fill("E2EFDA")   # pale green for BAP rows
FILL_BOTH    = _fill("EAD1F5")   # light purple for both

FONT_TITLE   = Font(bold=True, size=13)
FONT_SECTION = Font(bold=True, size=12)
FONT_HEADER  = Font(bold=True, size=11)
FONT_DATA    = Font(size=11)

ALIGN_LEFT   = Alignment(horizontal="left",  vertical="top", wrap_text=False)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="top")


# ── Name normalisation helpers ────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Lower-case, strip accents, collapse spaces."""
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode()
    return " ".join(s.lower().split())


# Manual overrides: honor-CSV name  →  person_canon (exact string)
# Extend this dict whenever a new mismatch is found.
_HONOR_OVERRIDES: dict[str, str] = {
    "ken shults":               "Kenneth Shults",
    "kenny shults":             "Kenneth Shults",
    "vasek klouda":             "Václav Klouda",
    "vaclav (vasek) klouda":    "Václav Klouda",
    "tina aberli":              "Tina Aeberli",
    "eli piltz":                "Eliott Piltz Galán",
    "eliott piltz galan":       "Eliott Piltz Galán",
    "evanne lamarch":           "Evanne Lemarche",
    "evanne lamarche":          "Evanne Lemarche",
    "arek dzudzinski":          "Arkadiusz Dudzinski",
    "martin cote":              "Martin Côté",
    "sebastien duchesne":       "Sébastien Duchesne",
    "sebastien duschesne":      "Sébastien Duchesne",
    "jonathan schneider":       "Jonathan Schneider",
    "lon smith":                "Lon Skyler Smith",
    "lon skyler smith":         "Lon Skyler Smith",
    "aleksi airinen":           "Aleksi Airinen",
    "lauri airinen":            "Lauri Airinen",
    "ales zelinka":             "Aleš Zelinka",
    "jere vainikka":            "Jere Väinikkä",
    "jukka peltola":            "Jukka Peltola",
    "tuomas karki":             "Tuomas Kärki",
    "tuukka antikainen":        "Tuukka Antikainen",
    "rafal kaleta":             "Rafał Kaleta",
    "pawel nowak":              "Paweł Nowak",
    "jakub mosciszewski":       "Jakub Mościszewski",
    "dominik simku":            "Dominik Šimků",
    "honza weber":              "Jan Weber",
    "carol wedemeyer":          "Carol Wedemeyer",
    "scott-mag hughes":         "Scott-Mag Hughes",
    "cheryl aubin hughes":      "Cheryl Aubin Hughes",
    "heather squires thomas":   "Heather Squires Thomas",
    "lisa mcdaniel jones":      "Lisa McDaniel Jones",
    "lori jean conover":        "Lori Jean Conover",
    "jody badger welch":        "Jody Badger Welch",
    "genevieve bousquet":       "Geneviève Bousquet",
    "becca english":            "Becca English",
    "becca english-ross":       "Becca English",
    "pt lovern":                "P.T. Lovern",
    "p.t. lovern":              "P.T. Lovern",
    "kendall kic":              "Kendall KIC",
    "taishi ishida":            "Taishi Ishida",
    "wiktor debski":            "Wiktor Dębski",
    "wiktor d\u0119bski":       "Wiktor Dębski",
    "florian gotze":            "Florian Götze",
    "grischa tellenbach":       "Grischa Tellenbach",
    "chantelle laurent":        "Chantelle Laurent",
}


def match_honor_name(raw_name: str, canon_by_norm: dict[str, str]) -> str | None:
    """
    Try to resolve an honor-CSV name to a person_canon string.
    1. Check _HONOR_OVERRIDES (normalised key)
    2. Exact normalised lookup in Persons_Truth
    Returns person_canon or None.
    """
    key = _norm(raw_name)
    if key in _HONOR_OVERRIDES:
        return _HONOR_OVERRIDES[key]
    if key in canon_by_norm:
        return canon_by_norm[key]
    return None


# ── Data loading helpers ───────────────────────────────────────────────────────

def load_csv(path: str) -> list[dict]:
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_persons_truth() -> tuple[list[dict], dict[str, str]]:
    """Returns (rows, norm→person_canon mapping)."""
    rows = load_csv(PT_CSV)
    norm_map: dict[str, str] = {}
    for r in rows:
        pc = r.get("person_canon", "").strip()
        if pc:
            norm_map[_norm(pc)] = pc
            # Also index the norm_key field if present
            nk = r.get("norm_key", "").strip()
            if nk:
                norm_map[nk] = pc
    return rows, norm_map


def is_real_person(r: dict) -> bool:
    pid = r.get("effective_person_id", "") or r.get("person_id", "")
    pc  = r.get("person_canon", "")
    if pid == "__NON_PERSON__" or pc == "__NON_PERSON__":
        return False
    excl = r.get("exclusion_reason", "") or ""
    if "non_person" in excl.lower():
        return False
    return bool(pc)


# ── Placement aggregations ────────────────────────────────────────────────────

def build_placement_stats(
    pf_rows: list[dict],
) -> dict[str, dict]:
    """
    Returns {person_id: {events, wins, podiums, placements, year_first, year_last}}
    Counts are across ALL division categories.
    """
    stats: dict[str, dict] = defaultdict(lambda: {
        "events": set(), "wins": 0, "podiums": 0,
        "placements": 0, "years": set(),
    })

    for row in pf_rows:
        pid = row.get("person_id", "")
        if not pid or pid == "__NON_PERSON__":
            continue
        eid  = row.get("event_id", "")
        year = row.get("year", "")
        try:
            place = int(row.get("place", 0) or 0)
        except ValueError:
            place = 0

        s = stats[pid]
        if eid:
            s["events"].add(eid)
        if year:
            try:
                s["years"].add(int(year))
            except ValueError:
                pass
        s["placements"] += 1
        if place == 1:
            s["wins"] += 1
        if 1 <= place <= 3:
            s["podiums"] += 1

    return {
        pid: {
            "events":     len(d["events"]),
            "wins":       d["wins"],
            "podiums":    d["podiums"],
            "placements": d["placements"],
            "year_first": min(d["years"]) if d["years"] else None,
            "year_last":  max(d["years"]) if d["years"] else None,
        }
        for pid, d in stats.items()
    }


def years_active_str(s: dict) -> str:
    yf = s.get("year_first")
    yl = s.get("year_last")
    if yf and yl:
        return str(yf) if yf == yl else f"{yf}\u2013{yl}"
    return ""


# ── Honors loading ────────────────────────────────────────────────────────────

def load_bap(canon_by_norm: dict[str, str]) -> dict[str, dict]:
    """
    Returns {person_canon: {bap_number, year_inducted, nickname}}
    BAP members are numbered 1..N in chronological order by year_inducted.
    """
    rows = load_csv(BAP_CSV)
    result: dict[str, dict] = {}
    for i, row in enumerate(rows, 1):
        raw  = row.get("name", "").strip()
        year = row.get("year_inducted", "").strip()
        nick = row.get("nickname", "").strip()
        pc   = match_honor_name(raw, canon_by_norm)
        if pc:
            result[pc] = {
                "bap_number":    i,
                "year_inducted": year,
                "nickname":      nick,
                "raw_name":      raw,
            }
        else:
            print(f"  [WARN] BAP name unmatched: {raw!r}")
    return result


def load_fbhof(canon_by_norm: dict[str, str]) -> dict[str, dict]:
    """
    Returns {person_canon: {year_inducted}}
    year_inducted may be "unknown".
    """
    rows = load_csv(FBHOF_CSV)
    result: dict[str, dict] = {}
    for row in rows:
        raw  = row.get("name", "").strip()
        year = row.get("year_inducted", "").strip()
        pc   = match_honor_name(raw, canon_by_norm)
        if pc:
            result[pc] = {"year_inducted": year, "raw_name": raw}
        else:
            print(f"  [WARN] FBHOF name unmatched: {raw!r}")
    return result


def bap_label(info: dict) -> str:
    n    = info["bap_number"]
    year = info["year_inducted"]
    return f"BAP #{n} ({year})" if year else f"BAP #{n}"


def fbhof_label(info: dict) -> str:
    year = info.get("year_inducted", "")
    return f"FBHOF {year}" if year and year != "unknown" else "FBHOF"


# ── Freestyle analytics ───────────────────────────────────────────────────────

def load_difficulty_profiles() -> dict[str, dict]:
    """Returns {person_id: row_dict}"""
    rows = load_csv(DIFFICULTY_CSV)
    return {r["person_id"]: r for r in rows if r.get("person_id")}


def load_diversity_profiles() -> dict[str, str]:
    """Returns {person_id: most_common_trick (first of top_tricks pipe-list)}"""
    rows = load_csv(DIVERSITY_CSV)
    result: dict[str, str] = {}
    for r in rows:
        pid  = r.get("person_id", "")
        tops = r.get("top_tricks", "")
        if pid and tops:
            result[pid] = tops.split(" | ")[0].strip()
    return result


# ── Sheet helpers ─────────────────────────────────────────────────────────────

def _w(ws, row: int, col: int, value=None, *,
       font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font  is not None: cell.font      = font
    if fill  is not None: cell.fill      = fill
    if align is not None: cell.alignment = align
    return cell


def _hrow(ws, row: int, *headers) -> int:
    for col, h in enumerate(headers, 1):
        _w(ws, row, col, h, font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT)
    return row + 1


def _drow(ws, row: int, *values) -> int:
    for col, v in enumerate(values, 1):
        align = ALIGN_RIGHT if isinstance(v, (int, float)) else ALIGN_LEFT
        _w(ws, row, col, v, font=FONT_DATA, align=align)
    return row + 1


def _section(ws, row: int, text: str) -> int:
    _w(ws, row, 1, text, font=FONT_SECTION, fill=FILL_SECTION, align=ALIGN_LEFT)
    return row + 1


# ── README sheet ──────────────────────────────────────────────────────────────

def build_readme(wb: Workbook) -> None:
    if "README" in wb.sheetnames:
        del wb["README"]
    idx = 0
    ws = wb.create_sheet("README", idx)

    row = 1
    _w(ws, row, 1, "Footbag Historical Results — Community Workbook",
       font=FONT_TITLE, align=ALIGN_LEFT)
    row += 2

    _w(ws, row, 1, "About This Workbook", font=FONT_SECTION, align=ALIGN_LEFT)
    row += 1
    for line in [
        "This workbook contains historical footbag competition results spanning 1980 to the present.",
        "Results are sourced from Footbag.org and supplementary archival records.",
        "Player identities are human-verified. Unresolved names are preserved as-is from the source.",
        "21 events are quarantined due to parsing ambiguity and excluded from statistics.",
    ]:
        _w(ws, row, 1, line, font=FONT_DATA, align=ALIGN_LEFT)
        row += 1
    row += 1

    _w(ws, row, 1, "Sheet Guide", font=FONT_SECTION, align=ALIGN_LEFT)
    row += 1
    _w(ws, row, 1, "Sheet", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT)
    _w(ws, row, 2, "Contents", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT)
    row += 1
    for name, desc in [
        ("README",              "This sheet — dataset overview and sheet guide"),
        ("DATA NOTES",          "Source quality notes, known limitations, quarantined events"),
        ("STATISTICS",          "Career podiums, event wins, events competed, career spans, events by year"),
        ("EVENT INDEX",         "One row per event — year, name, location, discipline counts"),
        ("PLAYER SUMMARY",      "Competition history per player — wins, podiums, placements, events"),
        ("CONSECUTIVE RECORDS", "Documented consecutives world records"),
        ("FREESTYLE INSIGHTS",  "Trick-sequence analytics — difficulty by year, backbone tricks, transitions, innovation timeline"),
        ("1980 – 2026",         "One sheet per year — all placement results for that year (including unresolved entries)"),
    ]:
        _w(ws, row, 1, name, font=FONT_DATA, align=ALIGN_LEFT)
        _w(ws, row, 2, desc, font=FONT_DATA, align=ALIGN_LEFT)
        row += 1
    row += 1

    _w(ws, row, 1, "Coverage Notes", font=FONT_SECTION, align=ALIGN_LEFT)
    row += 1
    for note in [
        "774 events documented, 1980–2026.",
        "3,441 canonically identified players.",
        "28,511 identity-locked placements.",
        "Coverage is comprehensive from 1997 onward (primary Footbag.org mirror).",
        "Pre-1997 data is partial — top-3 only for most divisions in 1980–1986 and 1990–1991.",
        "Years 1987–1989 and 1992–1996 have no coverage.",
        "FREESTYLE INSIGHTS draws from events that reported trick sequences; coverage is a subset of all events.",
    ]:
        _w(ws, row, 1, "•  " + note, font=FONT_DATA, align=ALIGN_LEFT)
        row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70
    print("  README sheet written")


# ── STATISTICS sheet ───────────────────────────────────────────────────────────

def build_statistics(wb: Workbook, pf_rows: list[dict], pt_rows: list[dict]) -> None:
    if "STATISTICS" in wb.sheetnames:
        del wb["STATISTICS"]
    # Insert after DATA NOTES
    if "DATA NOTES" in wb.sheetnames:
        idx = wb.sheetnames.index("DATA NOTES") + 1
    else:
        idx = 1
    ws = wb.create_sheet("STATISTICS", idx)

    # Build lookup tables
    valid_pids: set[str] = {
        r.get("effective_person_id", "") for r in pt_rows if is_real_person(r)
    }
    pid_to_canon: dict[str, str] = {
        r.get("effective_person_id", ""): r.get("person_canon", "")
        for r in pt_rows if r.get("effective_person_id")
    }

    # Aggregate from placements
    from collections import defaultdict
    podiums:       dict[str, dict] = defaultdict(lambda: {1: 0, 2: 0, 3: 0})
    events_by_pid: dict[str, set]  = defaultdict(set)
    years_by_pid:  dict[str, set]  = defaultdict(set)
    events_by_year: dict[int, set] = defaultdict(set)

    for row in pf_rows:
        pid = row.get("person_id", "")
        if not pid or pid not in valid_pids:
            continue
        eid  = row.get("event_id", "")
        year = row.get("year", "")
        try:
            place = int(row.get("place", 0) or 0)
        except (ValueError, TypeError):
            place = 0

        if eid:
            events_by_pid[pid].add(eid)
        if year:
            try:
                y = int(year)
                years_by_pid[pid].add(y)
                if eid:
                    events_by_year[y].add(eid)
            except (ValueError, TypeError):
                pass
        if 1 <= place <= 3:
            podiums[pid][place] += 1

    # ── Table helpers ─────────────────────────────────────────────────────────
    row_num = 1
    _w(ws, row_num, 1, "STATISTICS", font=FONT_TITLE, align=ALIGN_LEFT)
    row_num += 2

    TOP_N = 25

    # ── 1. Most Career Podiums ────────────────────────────────────────────────
    row_num = _section(ws, row_num, "MOST CAREER PODIUMS")
    row_num += 1
    row_num = _hrow(ws, row_num, "Player", "1st", "2nd", "3rd", "Total Podiums")
    podium_rows = [
        (pid_to_canon.get(pid, pid),
         d[1], d[2], d[3], d[1] + d[2] + d[3])
        for pid, d in podiums.items()
        if d[1] + d[2] + d[3] > 0
    ]
    podium_rows.sort(key=lambda x: (-x[4], x[0].lower()))
    for canon, p1, p2, p3, total in podium_rows[:TOP_N]:
        row_num = _drow(ws, row_num, canon, p1, p2, p3, total)
    row_num += 2

    # ── 2. Most Event Wins ────────────────────────────────────────────────────
    row_num = _section(ws, row_num, "MOST EVENT WINS")
    row_num += 1
    row_num = _hrow(ws, row_num, "Player", "Wins")
    wins_rows = [
        (pid_to_canon.get(pid, pid), d[1])
        for pid, d in podiums.items()
        if d[1] > 0
    ]
    wins_rows.sort(key=lambda x: (-x[1], x[0].lower()))
    for canon, wins in wins_rows[:TOP_N]:
        row_num = _drow(ws, row_num, canon, wins)
    row_num += 2

    # ── 3. Most Events Competed ───────────────────────────────────────────────
    row_num = _section(ws, row_num, "MOST EVENTS COMPETED")
    row_num += 1
    row_num = _hrow(ws, row_num, "Player", "Events Competed")
    events_rows = [
        (pid_to_canon.get(pid, pid), len(eids))
        for pid, eids in events_by_pid.items()
        if pid in valid_pids
    ]
    events_rows.sort(key=lambda x: (-x[1], x[0].lower()))
    for canon, count in events_rows[:TOP_N]:
        row_num = _drow(ws, row_num, canon, count)
    row_num += 2

    # ── 4. Longest Competitive Careers ───────────────────────────────────────
    row_num = _section(ws, row_num, "LONGEST COMPETITIVE CAREERS")
    row_num += 1
    row_num = _hrow(ws, row_num,
                    "Player", "First Event Year", "Last Event Year", "Career Span (Years)")
    career_rows = []
    for pid, years in years_by_pid.items():
        if pid not in valid_pids or not years:
            continue
        yf, yl = min(years), max(years)
        span = yl - yf
        career_rows.append((pid_to_canon.get(pid, pid), yf, yl, span))
    career_rows.sort(key=lambda x: (-x[3], x[0].lower()))
    for canon, yf, yl, span in career_rows[:TOP_N]:
        row_num = _drow(ws, row_num, canon, yf, yl, span)
    row_num += 2

    # ── 5. Events by Year ─────────────────────────────────────────────────────
    row_num = _section(ws, row_num, "EVENTS BY YEAR")
    row_num += 1
    row_num = _hrow(ws, row_num, "Year", "Events")
    for year in sorted(events_by_year):
        row_num = _drow(ws, row_num, year, len(events_by_year[year]))

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    ws.freeze_panes = "A2"
    print("  STATISTICS sheet written")


# ── PLAYER SUMMARY ────────────────────────────────────────────────────────────

PLAYER_SUMMARY_HEADERS = [
    "Player",
    "Wins", "Podiums", "Placements", "Events",
    "Years Active",
    "BAP", "FBHOF",
    "Freestyle Chains", "Max ADD", "Distinct Tricks", "Top Trick",
    "Legacy ID",
]

PLAYER_COL_WIDTHS = [32, 5, 7, 10, 6, 12, 14, 12, 16, 8, 15, 20, 10]


def build_player_summary(wb: Workbook,
                         pt_rows: list[dict],
                         placement_stats: dict[str, dict],
                         bap_map: dict[str, dict],
                         fbhof_map: dict[str, dict],
                         difficulty_map: dict[str, dict],
                         diversity_map: dict[str, str]) -> None:
    if "PLAYER SUMMARY" in wb.sheetnames:
        idx = wb.sheetnames.index("PLAYER SUMMARY")
        del wb["PLAYER SUMMARY"]
        ws = wb.create_sheet("PLAYER SUMMARY", idx)
    else:
        ws = wb.create_sheet("PLAYER SUMMARY")

    row = 1
    row = _hrow(ws, row, *PLAYER_SUMMARY_HEADERS)

    # Build rows: one per real person who has at least one placement
    persons = [r for r in pt_rows if is_real_person(r)]
    # Sort: most placements first, then alphabetically
    def _sort_key(r):
        pid   = r.get("effective_person_id", "")
        stats = placement_stats.get(pid, {})
        return (-stats.get("placements", 0), r.get("person_canon", "").lower())

    persons.sort(key=_sort_key)

    n_written = 0
    for r in persons:
        pid   = r.get("effective_person_id", "")
        pc    = r.get("person_canon", "")
        lid   = r.get("legacyid", "") or None

        stats = placement_stats.get(pid, {})
        if not stats:
            continue  # skip persons with no placements

        ya     = years_active_str(stats)
        bap    = bap_label(bap_map[pc])    if pc in bap_map    else None
        fbhof  = fbhof_label(fbhof_map[pc]) if pc in fbhof_map else None

        diff = difficulty_map.get(pid, {})
        chains = diff.get("chains_total")
        max_add = diff.get("max_sequence_add")
        n_tricks = diff.get("n_distinct_tricks")
        top_trick = diversity_map.get(pid)

        # Format numerics
        chains   = int(chains)   if chains   and str(chains)   != "nan" else None
        max_add  = float(max_add) if max_add  and str(max_add)  != "nan" else None
        n_tricks = int(n_tricks) if n_tricks and str(n_tricks) != "nan" else None

        # Highlight BAP/FBHOF rows
        if pc in bap_map and pc in fbhof_map:
            row_fill = FILL_BOTH
        elif pc in fbhof_map:
            row_fill = FILL_HOF
        elif pc in bap_map:
            row_fill = FILL_BAP
        else:
            row_fill = None

        values = [
            pc,
            stats.get("wins"),
            stats.get("podiums"),
            stats.get("placements"),
            stats.get("events"),
            ya or None,
            bap,
            fbhof,
            chains,
            max_add,
            n_tricks,
            top_trick,
            lid,
        ]
        for col, v in enumerate(values, 1):
            align = ALIGN_RIGHT if isinstance(v, (int, float)) else ALIGN_LEFT
            cell = _w(ws, row, col, v, font=FONT_DATA, align=align)
            if row_fill:
                cell.fill = row_fill
        row += 1
        n_written += 1

    # Column widths
    for col, width in enumerate(PLAYER_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PLAYER_SUMMARY_HEADERS))}1"

    print(f"  PLAYER SUMMARY: {n_written} rows written")


# ── HONORS sheet ──────────────────────────────────────────────────────────────

def build_honors_sheet(wb: Workbook,
                       pt_rows: list[dict],
                       placement_stats: dict[str, dict],
                       bap_map: dict[str, dict],
                       fbhof_map: dict[str, dict]) -> None:
    if "HONORS" in wb.sheetnames:
        del wb["HONORS"]

    # Insert after CONSECUTIVE RECORDS (or PLAYER SUMMARY if not found)
    insert_after = "CONSECUTIVE RECORDS"
    if insert_after in wb.sheetnames:
        idx = wb.sheetnames.index(insert_after) + 1
    else:
        idx = wb.sheetnames.index("PLAYER SUMMARY") + 1
    ws = wb.create_sheet("HONORS", idx)

    # Build lookup: person_canon → placement stats
    pc_to_pid = {
        r.get("person_canon", ""): r.get("effective_person_id", "")
        for r in pt_rows if r.get("person_canon")
    }

    def ya_for_canon(pc: str) -> str:
        pid   = pc_to_pid.get(pc, "")
        stats = placement_stats.get(pid, {})
        return years_active_str(stats)

    row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    _w(ws, row, 1, "COMMUNITY HONORS", font=FONT_TITLE, align=ALIGN_LEFT)
    row += 2

    # ── FBHOF Section ─────────────────────────────────────────────────────────
    row = _section(ws, row, "FOOTBAG HALL OF FAME")
    row += 1
    row = _hrow(ws, row, "Player", "Year Inducted", "Years Active")

    # Sort: numeric years first (ascending), "unknown" at end, then alpha
    def _fbhof_sort(item):
        pc, info = item
        y = info.get("year_inducted", "")
        try:
            return (0, int(y), pc.lower())
        except (ValueError, TypeError):
            return (1, 0, pc.lower())

    for pc, info in sorted(fbhof_map.items(), key=_fbhof_sort):
        year = info.get("year_inducted", "")
        ya   = ya_for_canon(pc)
        cell = ws.cell(row=row, column=1)
        cell.value     = pc
        cell.font      = FONT_DATA
        cell.fill      = FILL_HOF
        cell.alignment = ALIGN_LEFT
        for col, v in enumerate([year if year != "unknown" else "?", ya], 2):
            c = ws.cell(row=row, column=col)
            c.value     = v
            c.font      = FONT_DATA
            c.fill      = FILL_HOF
            c.alignment = ALIGN_LEFT
        row += 1

    row += 2

    # ── BAP Section ───────────────────────────────────────────────────────────
    row = _section(ws, row, "BIG ADD POSSE")
    row += 1
    row = _hrow(ws, row, "Player", "BAP #", "Year Inducted", "Nickname", "Years Active")

    # Sort by BAP number (chronological)
    for pc, info in sorted(bap_map.items(), key=lambda x: x[1]["bap_number"]):
        ya   = ya_for_canon(pc)
        bnum = info["bap_number"]
        year = info["year_inducted"]
        nick = info["nickname"]
        cell = ws.cell(row=row, column=1)
        cell.value     = pc
        cell.font      = FONT_DATA
        cell.fill      = FILL_BAP
        cell.alignment = ALIGN_LEFT
        for col, v in enumerate([bnum, year, nick or None, ya], 2):
            c = ws.cell(row=row, column=col)
            c.value     = v
            c.font      = FONT_DATA
            c.fill      = FILL_BAP
            c.alignment = ALIGN_RIGHT if isinstance(v, int) else ALIGN_LEFT
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12

    ws.freeze_panes = "A2"

    n_hof = len(fbhof_map)
    n_bap = len(bap_map)
    print(f"  HONORS sheet: {n_hof} FBHOF + {n_bap} BAP entries")


# ── FREESTYLE INSIGHTS sheet ─────────────────────────────────────────────────

def build_freestyle_insights(wb: Workbook) -> None:
    """Build FREESTYLE INSIGHTS sheet — vertical stacked tables, compact formatting."""
    from openpyxl.styles import Border, Side

    trick_freq  = load_csv(TRICK_FREQ_CSV)
    transitions = load_csv(TRANSITIONS_CSV)
    seq_diff    = load_csv(SEQ_DIFF_CSV)
    complexity  = load_csv(CHAIN_COMPLEXITY_CSV)
    diversity   = load_csv(DIVERSITY_CSV)
    trick_node  = load_csv(TRICK_NODE_CSV)

    if not any([trick_freq, transitions, seq_diff]):
        print("  FREESTYLE INSIGHTS: skipped (no analytics CSVs found)")
        return

    if "FREESTYLE INSIGHTS" in wb.sheetnames:
        del wb["FREESTYLE INSIGHTS"]

    for anchor in ("CONSECUTIVE RECORDS", "PLAYER SUMMARY"):
        if anchor in wb.sheetnames:
            idx = wb.sheetnames.index(anchor) + 1
            break
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet("FREESTYLE INSIGHTS", idx)

    # ── Styles ────────────────────────────────────────────────────────────────
    _thin    = Side(style="thin")
    _border  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _hfill   = _fill("D9D9D9")
    _no_wrap = Alignment(horizontal="left",  vertical="top", wrap_text=False)
    _no_wrap_r = Alignment(horizontal="right", vertical="top", wrap_text=False)

    def _c(row: int, col: int, value=None, *, header: bool = False) -> None:
        """Write a single bordered, non-wrapping cell."""
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = _border
        if header:
            cell.font  = Font(bold=True, size=11)
            cell.fill  = _hfill
            cell.alignment = _no_wrap
        else:
            cell.font  = Font(size=11)
            cell.alignment = (
                _no_wrap_r if isinstance(value, (int, float)) else _no_wrap
            )

    def _title(row: int, text: str) -> int:
        ws.cell(row=row, column=1).value = text
        ws.cell(row=row, column=1).font  = Font(bold=True, size=12)
        return row + 1

    def _hdr(row: int, *col_header_pairs) -> int:
        """Write header cells. col_header_pairs: (col, label), ..."""
        for col, h in col_header_pairs:
            _c(row, col, h, header=True)
        return row + 1

    # Column layout (shared across all stacked tables):
    #  A(1)=5   rank / era
    #  B(2)=28  trick / player / transition / era label
    #  C(3)=8   ADD / year / count / unique tricks / chains
    #  D(4)=12  mentions / connections / avg add / years active
    #  E(5)=10  players / length
    #  F(6)=50  events / sequence / notes
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 50

    _MODIFIERS = {
        "pixie", "ducking", "spinning", "atomic", "symposium",
        "stepping", "gyro", "barraging", "blazing", "tapping", "paradox",
    }

    def _addv(trick: str, adds_raw: str) -> object:
        if trick in _MODIFIERS:
            return "modifier"
        try:
            return int(float(adds_raw)) if adds_raw else None
        except (ValueError, TypeError):
            return None

    node_by_trick = {r["trick"]: r for r in trick_node}

    row = 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 1: Most Used Freestyle Tricks
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Used Freestyle Tricks")
    row = _hdr(row, (1, "#"), (2, "Trick"), (3, "ADD"),
               (4, "Mentions"), (5, "Players"), (6, "Events"))
    freq_sorted = sorted(trick_freq,
                         key=lambda r: _int(r, "total_mentions") or 0, reverse=True)
    for rank, r in enumerate(freq_sorted[:25], 1):
        trick = r.get("trick_canon", "")
        _c(row, 1, rank)
        _c(row, 2, trick)
        _c(row, 3, _addv(trick, r.get("adds", "")))
        _c(row, 4, _int(r, "total_mentions"))
        _c(row, 5, _int(r, "n_players"))
        _c(row, 6, _int(r, "n_events"))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 2: Most Influential Connector Tricks
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Influential Connector Tricks")
    row = _hdr(row, (1, "#"), (2, "Trick"), (3, "ADD"),
               (4, "Connections"), (5, "Players"), (6, "Events"))
    enriched = []
    for r in trick_freq:
        trick = r.get("trick_canon", "")
        nd  = node_by_trick.get(trick, {})
        deg = _int(nd, "degree") or 0
        enriched.append((deg, r))
    enriched.sort(key=lambda x: x[0], reverse=True)
    for rank, (deg, r) in enumerate(enriched[:15], 1):
        trick = r.get("trick_canon", "")
        _c(row, 1, rank)
        _c(row, 2, trick)
        _c(row, 3, _addv(trick, r.get("adds", "")))
        _c(row, 4, deg)
        _c(row, 5, _int(r, "n_players"))
        _c(row, 6, _int(r, "n_events"))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 3: Most Common Trick Transitions
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Common Trick Transitions")
    row = _hdr(row, (1, "#"), (2, "Transition"), (3, "Count"), (4, "Players"))
    trans_sorted = sorted(transitions,
                          key=lambda r: _int(r, "count") or 0, reverse=True)
    for rank, r in enumerate(trans_sorted[:20], 1):
        ta = r.get("trick_a", "")
        tb = r.get("trick_b", "")
        _c(row, 1, rank)
        _c(row, 2, f"{ta} → {tb}")
        _c(row, 3, _int(r, "count"))
        _c(row, 4, _int(r, "n_players"))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 4: Hardest Documented Sequences
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Hardest Documented Sequences")
    row = _hdr(row, (1, "#"), (2, "Player"), (3, "Year"),
               (4, "ADD"), (5, "Length"), (6, "Sequence"))
    scored_seqs = [
        r for r in seq_diff
        if r.get("sequence_add") and r.get("person_canon", "").strip()
        and r["person_canon"].strip() not in ("", "__NON_PERSON__")
    ]
    scored_seqs.sort(key=lambda r: float(r.get("sequence_add") or 0), reverse=True)
    for rank, r in enumerate(scored_seqs[:10], 1):
        _c(row, 1, rank)
        _c(row, 2, r.get("person_canon"))
        _c(row, 3, _int(r, "year"))
        _c(row, 4, _floatv(r, "sequence_add"))
        _c(row, 5, _int(r, "normalized_length"))
        _c(row, 6, r.get("tricks_normalized", "").replace(">", " > "))
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 5: Most Diverse Players
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Most Diverse Players")
    row = _hdr(row, (1, "#"), (2, "Player"), (3, "Unique Tricks"), (4, "Years Active"))
    for rank, r in enumerate(diversity[:15], 1):
        y1 = _int(r, "year_first")
        y2 = _int(r, "year_last")
        _c(row, 1, rank)
        _c(row, 2, r.get("person_canon"))
        _c(row, 3, _int(r, "unique_tricks"))
        _c(row, 4, f"{y1}–{y2}" if y1 and y2 else "")
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 6: Evolution of Difficulty
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "Evolution of Difficulty")
    row = _hdr(row, (2, "Era"), (3, "Chains"), (4, "Avg ADD"))
    complexity_by_year: dict[int, dict] = {}
    for r in complexity:
        y = _int(r, "year")
        if y:
            complexity_by_year[y] = r
    for label, y1, y2 in [
        ("2001–2003", 2001, 2003), ("2004–2006", 2004, 2006),
        ("2007–2009", 2007, 2009), ("2010–2015", 2010, 2015),
        ("2016–2025", 2016, 2025),
    ]:
        era_rows = [complexity_by_year[y] for y in range(y1, y2 + 1)
                    if y in complexity_by_year]
        if not era_rows:
            continue
        total_chains = sum(_int(r, "n_chains") or 0 for r in era_rows)
        weighted_sum = sum(
            (float(r.get("avg_avg_add") or 0)) * (_int(r, "n_chains") or 0)
            for r in era_rows
        )
        avg_add = round(weighted_sum / total_chains, 2) if total_chains else None
        _c(row, 2, label)
        _c(row, 3, total_chains)
        _c(row, 4, avg_add)
        row += 1
    row += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLE 7: ADD Composition Examples
    # ═══════════════════════════════════════════════════════════════════════════
    row = _title(row, "ADD Composition Examples")
    row = _hdr(row, (2, "Trick"), (3, "ADD"), (4, "Notes"))
    for trick, add, note in [
        ("whirl",         3,           "Most-connected trick in the network"),
        ("blurry whirl",  5,           "Rotational base + blurry modifier (+2)"),
        ("blurriest",     6,           "Maximum documented base ADD"),
        ("ripwalk",       4,           "High-frequency transition trick"),
        ("ducking whirl", 4,           "Modifier stack on rotational base"),
        ("pixie",         "modifier",  "Standalone difficulty modifier, no fixed ADD"),
    ]:
        _c(row, 2, trick)
        _c(row, 3, add)
        _c(row, 4, note)
        row += 1

    ws.freeze_panes = "A2"
    print(f"  FREESTYLE INSIGHTS sheet written ({row - 1} rows)")


def _int(r: dict, key: str) -> int | None:
    v = r.get(key)
    if v is None or v == "" or str(v).lower() == "nan":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _floatv(r: dict, key: str) -> float | None:
    v = r.get(key)
    if v is None or v == "" or str(v).lower() == "nan":
        return None
    try:
        return round(float(v), 3)
    except (ValueError, TypeError):
        return None


# ── Sheet copy utilities (unchanged from v11) ─────────────────────────────────

def copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        for attr in ("font", "fill", "alignment", "border", "number_format"):
            try:
                val = getattr(src_cell, attr)
                setattr(dst_cell, attr,
                        copy.copy(val) if attr != "number_format" else val)
            except Exception:
                pass
    if src_cell.hyperlink:
        try:
            dst_cell.hyperlink = copy.copy(src_cell.hyperlink)
        except Exception:
            pass


def copy_sheet_to(src_ws, dst_ws, *, standardise_year_cols: bool = False):
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width  = col_dim.width or 8.43
        dst_ws.column_dimensions[col_letter].hidden = col_dim.hidden
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height or 15
    for row in src_ws.iter_rows():
        for cell in row:
            copy_cell(cell, dst_ws.cell(row=cell.row, column=cell.column))
    for merge_range in src_ws.merged_cells.ranges:
        try:
            dst_ws.merge_cells(str(merge_range))
        except Exception:
            pass
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref

    if standardise_year_cols:
        # Standardise column widths for year sheets only
        # A=place(4), B=player(34), C=team(18), D=division/notes(22)
        _YEAR_WIDTHS = {"A": 5, "B": 34, "C": 18, "D": 22}
        for col_letter, width in _YEAR_WIDTHS.items():
            if col_letter in dst_ws.column_dimensions:
                dst_ws.column_dimensions[col_letter].width = width


# ── Main ──────────────────────────────────────────────────────────────────────

FRONT_SHEETS_COPY = {
    "DATA NOTES", "EVENT INDEX", "CONSECUTIVE RECORDS",
}

# Sheets rebuilt or newly created (not copied from v11)
SHEETS_REBUILT = {"README", "STATISTICS", "FREESTYLE INSIGHTS"}


def main():

    print(f"\nOpening source workbook: {os.path.basename(SOURCE_V11)}")
    src_wb = openpyxl.load_workbook(SOURCE_V11)
    print(f"  Sheets: {src_wb.sheetnames[:6]} ...")

    print("\nLoading canonical data...")
    pt_rows, canon_by_norm = load_persons_truth()
    pf_rows = load_csv(PF_CSV)
    print(f"  Persons Truth: {len(pt_rows)} rows")
    print(f"  Placements Flat: {len(pf_rows)} rows")

    print("\nCreating output workbook...")
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    # Build README first (position 0)
    print("\nBuilding README sheet...")
    build_readme(out_wb)

    # Copy DATA NOTES from v11
    for sheet_name in ("DATA NOTES",):
        if sheet_name in src_wb.sheetnames:
            print(f"  Copying {sheet_name}...", end="", flush=True)
            dst_ws = out_wb.create_sheet(sheet_name)
            copy_sheet_to(src_wb[sheet_name], dst_ws)
            print(" done")

    # Build STATISTICS (after DATA NOTES)
    print("\nBuilding STATISTICS sheet...")
    build_statistics(out_wb, pf_rows, pt_rows)

    # Copy EVENT INDEX, PLAYER SUMMARY, CONSECUTIVE RECORDS from v11
    for sheet_name in ("EVENT INDEX", "PLAYER SUMMARY", "CONSECUTIVE RECORDS"):
        if sheet_name in src_wb.sheetnames:
            print(f"  Copying {sheet_name}...", end="", flush=True)
            dst_ws = out_wb.create_sheet(sheet_name)
            copy_sheet_to(src_wb[sheet_name], dst_ws)
            print(" done")

    # Build FREESTYLE INSIGHTS after CONSECUTIVE RECORDS
    print("\nBuilding FREESTYLE INSIGHTS sheet...")
    build_freestyle_insights(out_wb)

    # Copy all year sheets from v11
    for sheet_name in src_wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        print(f"  Copying {sheet_name}...", end="", flush=True)
        dst_ws = out_wb.create_sheet(sheet_name)
        copy_sheet_to(src_wb[sheet_name], dst_ws, standardise_year_cols=True)
        print(" done")

    # Final sheet order check
    sheets = out_wb.sheetnames
    year_sheets = [s for s in sheets if s.isdigit()]
    front = [s for s in sheets if not s.isdigit()]
    print(f"\nFront sheets: {front}")
    print(f"Year sheets: {year_sheets[0]}–{year_sheets[-1]} ({len(year_sheets)} sheets)")

    print(f"\nSaving to: {OUTPUT_PATH}")
    out_wb.save(OUTPUT_PATH)
    size_mb = os.path.getsize(OUTPUT_PATH) / (1024 * 1024)
    print(f"Saved: {size_mb:.1f} MB")
    print("Done.")


def _insert_before(wb: Workbook, before_sheet: str, new_sheet_name: str) -> None:
    """Create new_sheet_name positioned immediately before before_sheet."""
    if new_sheet_name in wb.sheetnames:
        return  # already exists
    if before_sheet in wb.sheetnames:
        idx = wb.sheetnames.index(before_sheet)
        wb.create_sheet(new_sheet_name, idx)
    else:
        wb.create_sheet(new_sheet_name)


if __name__ == "__main__":
    main()
