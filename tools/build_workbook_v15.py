#!/usr/bin/env python3
"""
tools/build_workbook_v15.py

Community workbook v15 — full rebuild from out/canonical_all/ CSVs.

Design principles (v15):
  - Source: out/canonical_all/*.csv  (merged pre-1997 + post-1997, fully normalized)
  - ALL aggregation uses person_id — no display_name matching
  - ALL names displayed via persons.csv person_canon
  - ALL years 1980-present included; no year suppression
  - ZERO hidden rows, columns, or sheets anywhere
  - Column A always visible on every sheet
  - Division names from canonical_all (already normalized, no Sgls/Dbls abbreviations)
  - BAP / FBHOF data read directly from canonical_all/persons.csv

Sheets produced (in order):
  README
  STATISTICS
  ERA LEADERS
  PLAYER STATS
  EVENT INDEX
  EXCLUDED EVENTS
  <year>  (one per year present in data, e.g. 1980 … 2026)
"""

import csv
import os
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

csv.field_size_limit(sys.maxsize)

ROOT = Path(__file__).resolve().parents[1]

# ── Paths ──────────────────────────────────────────────────────────────────────
CA            = ROOT / "out" / "canonical_all"
QUARANTINE_CSV = ROOT / "inputs" / "review_quarantine_events.csv"
OUTPUT_PATH   = ROOT / "Footbag_Results_Community_v15.xlsx"

# ── Styles ─────────────────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

FILL_NONE    = PatternFill(fill_type=None)
FILL_HEADER  = _fill("D9D9D9")         # grey header
FILL_SECTION = _fill("1F3864")         # dark navy section dividers
FILL_TITLE   = _fill("2E75B6")         # blue title bands
FILL_NET     = _fill("E2EFDA")         # pale green — net
FILL_FREE    = _fill("FFF2CC")         # pale gold — freestyle
FILL_GOLF    = _fill("FCE4D6")         # pale orange — golf / sideline
FILL_OTHER   = _fill("F2F2F2")         # light grey — other / unknown
FILL_EVENT_A = _fill("DEEAF1")         # even event bands
FILL_EVENT_B = _fill("FFFFFF")         # odd event bands
FILL_HOF     = _fill("FFF2CC")         # FBHOF rows
FILL_BAP     = _fill("E2EFDA")         # BAP rows
FILL_BOTH    = _fill("EAD1F5")         # BAP + FBHOF
FILL_QUAR    = _fill("FFE0E0")         # quarantined

WHITE        = _fill("FFFFFF")

FONT_TITLE   = Font(bold=True,  size=13, color="FFFFFF")
FONT_SECTION = Font(bold=True,  size=11, color="FFFFFF")
FONT_HEADER  = Font(bold=True,  size=11)
FONT_DATA    = Font(size=11)
FONT_SMALL   = Font(size=9,    color="808080")
FONT_LINK    = Font(size=11,   color="0563C1", underline="single")

ALIGN_L   = Alignment(horizontal="left",   vertical="top", wrap_text=False)
ALIGN_R   = Alignment(horizontal="right",  vertical="top")
ALIGN_C   = Alignment(horizontal="center", vertical="top")
ALIGN_LW  = Alignment(horizontal="left",   vertical="top", wrap_text=True)

# ── Sentinels (never included in stats / leaderboards) ─────────────────────────
_SKIP_PID = {"", "__NON_PERSON__"}
_SKIP_DNAME = {"__NON_PERSON__", "__UNKNOWN_PARTNER__", "[UNKNOWN PARTNER]", "[UNKNOWN]", ""}

# ── Helpers ─────────────────────────────────────────────────────────────────────

def _w(ws, row: int, col: int, value=None, *,
       font=None, fill=None, align=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font   is not None: cell.font          = font
    if fill   is not None: cell.fill          = fill
    if align  is not None: cell.alignment     = align
    if number_format:      cell.number_format = number_format
    return cell


def _title_row(ws, row: int, text: str, ncols: int = 8) -> int:
    _w(ws, row, 1, text, font=FONT_TITLE, fill=FILL_TITLE, align=ALIGN_L)
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = FILL_TITLE
    ws.row_dimensions[row].height = 22
    return row + 1


def _section_row(ws, row: int, text: str, ncols: int = 8) -> int:
    _w(ws, row, 1, text, font=FONT_SECTION, fill=FILL_SECTION, align=ALIGN_L)
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION
    ws.row_dimensions[row].height = 18
    return row + 1


def _hrow(ws, row: int, *headers) -> int:
    for col, h in enumerate(headers, 1):
        _w(ws, row, col, h, font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_L)
    return row + 1


def _drow(ws, row: int, *values, fill=None) -> int:
    for col, v in enumerate(values, 1):
        align = ALIGN_R if isinstance(v, (int, float)) else ALIGN_L
        cell  = _w(ws, row, col, v, font=FONT_DATA, align=align)
        if fill:
            cell.fill = fill
    return row + 1


def _load(name: str) -> list[dict]:
    path = CA / name
    if not path.exists():
        print(f"  [WARN] missing {path}")
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _cat_fill(cat: str) -> PatternFill:
    c = (cat or "").lower()
    if c == "net":       return FILL_NET
    if c == "freestyle": return FILL_FREE
    if c in ("golf", "sideline"): return FILL_GOLF
    return FILL_OTHER


# ── Data loading ───────────────────────────────────────────────────────────────

def load_all():
    """Load all canonical_all CSVs and return structured data."""
    print("Loading canonical_all CSVs…")

    raw_events  = _load("events.csv")
    raw_discs   = _load("event_disciplines.csv")
    raw_results = _load("event_result_participants.csv")
    raw_persons = _load("persons.csv")

    # Events dict
    events: dict[str, dict] = {r["event_id"]: r for r in raw_events}

    # Disciplines dict: (event_id, discipline) → row
    discs: dict[tuple, dict] = {}
    for r in raw_discs:
        discs[(r["event_id"], r["discipline"])] = r

    # Persons dict: person_id → row
    persons: dict[str, dict] = {r["person_id"]: r for r in raw_persons}

    # Quarantine set
    quarantine: set[str] = set()
    if QUARANTINE_CSV.exists():
        with open(QUARANTINE_CSV, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                eid = r.get("event_id", "").strip()
                if eid:
                    quarantine.add(eid)

    print(f"  Events: {len(events)}   Disciplines: {len(discs)}")
    print(f"  Participants: {len(raw_results)}   Persons: {len(persons)}")
    print(f"  Quarantined events: {len(quarantine)}")
    return events, discs, raw_results, persons, quarantine


# ── Statistics engine ──────────────────────────────────────────────────────────

def compute_stats(raw_results, events, discs, persons):
    """
    Aggregate by person_id only. Returns:
      stats[pid] = {wins, podiums, p1, p2, p3, events_set, years_set,
                    worlds_wins, worlds_podiums, worlds_events}
      event_cat_stats[pid][cat] = {wins, podiums}
    """
    stats:           dict[str, dict] = defaultdict(lambda: {
        "wins": 0, "p1": 0, "p2": 0, "p3": 0, "podiums": 0,
        "events": set(), "years": set(),
        "worlds_wins": 0, "worlds_podiums": 0, "worlds_events": set(),
        "cat_wins":    defaultdict(int),
        "cat_podiums": defaultdict(int),
    })

    worlds_types = {"WFA_WORLD_CHAMPIONSHIPS", "WORLD_CHAMPIONSHIPS",
                    "IFAB_WORLD_CHAMPIONSHIPS", "NHSA_NATIONALS"}

    for row in raw_results:
        pid = row.get("person_id", "").strip()
        if not pid or pid in _SKIP_PID:
            continue
        dname = row.get("display_name", "").strip()
        if dname in _SKIP_DNAME:
            continue
        # Only count participant_order=1 for win/podium tracking
        # (each placement slot counts once, not once per team member)
        porder = row.get("participant_order", "1").strip()

        eid   = row.get("event_id", "").strip()
        ev    = events.get(eid, {})
        year  = ev.get("year", "").strip()
        etype = ev.get("event_type", "").strip()
        disc  = discs.get((eid, row.get("discipline", "").strip()), {})
        cat   = disc.get("discipline_category", "").lower().strip()

        try:
            place = int(row.get("placement", "0") or 0)
        except ValueError:
            place = 0

        s = stats[pid]
        if eid:
            s["events"].add(eid)
        if year:
            try:
                s["years"].add(int(year))
            except ValueError:
                pass

        # Win/podium only counted once per placement slot (participant_order=1)
        if porder == "1":
            if place == 1:
                s["wins"] += 1
                s["p1"]   += 1
                s["cat_wins"][cat] += 1
            if 1 <= place <= 3:
                s["podiums"] += 1
                s["cat_podiums"][cat] += 1
            if place == 2: s["p2"] += 1
            if place == 3: s["p3"] += 1

            if etype in worlds_types:
                if eid:
                    s["worlds_events"].add(eid)
                if place == 1:
                    s["worlds_wins"] += 1
                if 1 <= place <= 3:
                    s["worlds_podiums"] += 1

    # Flatten sets to counts
    result = {}
    for pid, s in stats.items():
        if pid not in persons:
            continue  # skip persons not in master list
        result[pid] = {
            "wins":           s["wins"],
            "p1": s["p1"], "p2": s["p2"], "p3": s["p3"],
            "podiums":        s["podiums"],
            "events":         len(s["events"]),
            "year_first":     min(s["years"]) if s["years"] else None,
            "year_last":      max(s["years"])  if s["years"] else None,
            "worlds_wins":    s["worlds_wins"],
            "worlds_podiums": s["worlds_podiums"],
            "worlds_events":  len(s["worlds_events"]),
            "cat_wins":       dict(s["cat_wins"]),
            "cat_podiums":    dict(s["cat_podiums"]),
        }
    return result


# ── README sheet ──────────────────────────────────────────────────────────────

def build_readme(wb: Workbook, events: dict, persons: dict, n_parts: int) -> None:
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 90

    n_years   = len(set(e.get("year","") for e in events.values() if e.get("year")))
    n_events  = len(events)
    n_persons = sum(1 for p in persons.values()
                    if p.get("person_id") and p.get("person_canon"))

    row = _title_row(ws, 1, "FOOTBAG COMPETITION RESULTS — COMMUNITY DATASET", ncols=1)
    row += 1

    lines = [
        ("Coverage",
         f"{n_events} events · {n_years} years · {n_persons} identified competitors · "
         f"{n_parts:,} placement records"),
        ("Years",       "1980 – present (all years included; no year suppression)"),
        ("Sources",     "Post-1997: Footbag.org HTML archive.  "
                        "Pre-1997: Footbag World magazine scans + oldresults.txt."),
        ("Identity",    "All statistics aggregated by person_id. "
                        "Canonical names from persons.csv only."),
        ("Divisions",   "Fully normalized (no abbreviations). "
                        "Categories: freestyle · net · golf · sideline."),
        ("Visibility",  "No hidden rows, columns, or sheets anywhere."),
        ("", ""),
        ("SHEETS", ""),
        ("README",          "This sheet — dataset overview."),
        ("STATISTICS",      "Career leaderboards: podiums, wins, events competed, career spans."),
        ("ERA LEADERS",     "Decade-by-decade leaderboards (1980s – 2020s)."),
        ("PLAYER STATS",    "One row per identified competitor with career summary."),
        ("EVENT INDEX",     "All events with metadata, linked to year sheets."),
        ("EXCLUDED EVENTS", "Events in archive without usable results (no_results / quarantined)."),
        ("<year>",          "One sheet per year (1980 – 2026). "
                            "Each row is one placement. Grouped by event → division → place."),
        ("", ""),
        ("NOTES", ""),
        ("Unknown participants",
         "Rows with unresolved identity ([UNKNOWN PARTNER], __UNKNOWN_PARTNER__) "
         "appear in year sheets but are excluded from all leaderboards."),
        ("Pre-1997 data",
         "Reconstructed from magazine scans. Some events have partial coverage. "
         "Coverage quality noted in EVENT INDEX."),
        ("Quarantined events",
         "Included in year sheets with ⛔ marker. Results may be incomplete or uncertain."),
    ]

    for label, text in lines:
        if not label and not text:
            row += 1
            continue
        if label in ("SHEETS", "NOTES"):
            row = _section_row(ws, row, label, ncols=1)
            continue
        _w(ws, row, 1, f"{'  ' if label == '<year>' else ''}"
           f"{'▸ ' + label + ':  ' if label else ''}{text}",
           font=FONT_DATA, align=ALIGN_LW)
        ws.row_dimensions[row].height = 16
        row += 1

    ws.sheet_view.showGridLines = True
    print("  README done")


# ── STATISTICS sheet ──────────────────────────────────────────────────────────

def build_statistics(wb: Workbook, stats: dict, persons: dict) -> None:
    ws = wb.create_sheet("STATISTICS")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A2"

    row = _title_row(ws, 1, "STATISTICS — ALL EVENTS ALL DIVISIONS", ncols=5)

    def canon(pid):
        return persons.get(pid, {}).get("person_canon", pid)

    TOP = 30

    # Helper: write a ranked table
    def table(title, headers, data_rows, start_row):
        r = _section_row(ws, start_row, title, ncols=len(headers))
        r = _hrow(ws, r, *headers)
        for i, vals in enumerate(data_rows[:TOP]):
            fill = FILL_NET if i % 2 == 0 else FILL_EVENT_B
            r = _drow(ws, r, *vals, fill=fill)
        return r + 1

    # 1 — Career Podiums
    podium_data = sorted(
        [(canon(pid), s["p1"], s["p2"], s["p3"], s["podiums"])
         for pid, s in stats.items() if s["podiums"] > 0],
        key=lambda x: (-x[4], x[0].lower())
    )
    row = table("MOST CAREER PODIUMS (ALL DIVISIONS)",
                ["Player", "1st", "2nd", "3rd", "Total"],
                podium_data, row)

    # 2 — Wins
    wins_data = sorted(
        [(canon(pid), s["wins"]) for pid, s in stats.items() if s["wins"] > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST CAREER WINS (ALL DIVISIONS)",
                ["Player", "Wins"], wins_data, row)

    # 3 — Freestyle Podiums
    free_pod = sorted(
        [(canon(pid), s["cat_podiums"].get("freestyle", 0))
         for pid, s in stats.items() if s["cat_podiums"].get("freestyle", 0) > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST FREESTYLE PODIUMS",
                ["Player", "Podiums"], free_pod, row)

    # 4 — Net Podiums
    net_pod = sorted(
        [(canon(pid), s["cat_podiums"].get("net", 0))
         for pid, s in stats.items() if s["cat_podiums"].get("net", 0) > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST NET PODIUMS",
                ["Player", "Podiums"], net_pod, row)

    # 5 — Worlds Podiums
    worlds_pod = sorted(
        [(canon(pid), s["worlds_wins"], s["worlds_podiums"], s["worlds_events"])
         for pid, s in stats.items() if s["worlds_podiums"] > 0],
        key=lambda x: (-x[2], x[0].lower())
    )
    row = table("WORLDS PODIUMS (WORLD CHAMPIONSHIPS ONLY)",
                ["Player", "Worlds Wins", "Worlds Podiums", "Worlds Events"],
                worlds_pod, row)

    # 6 — Worlds Wins
    worlds_wins = sorted(
        [(canon(pid), s["worlds_wins"])
         for pid, s in stats.items() if s["worlds_wins"] > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("WORLDS WINS",
                ["Player", "Wins"], worlds_wins, row)

    # 7 — Events Competed
    events_data = sorted(
        [(canon(pid), s["events"]) for pid, s in stats.items() if s["events"] > 0],
        key=lambda x: (-x[1], x[0].lower())
    )
    row = table("MOST EVENTS COMPETED",
                ["Player", "Events"], events_data, row)

    # 8 — Career Spans
    career_data = sorted(
        [(canon(pid), s["year_first"], s["year_last"],
          (s["year_last"] - s["year_first"]) if s["year_first"] and s["year_last"] else 0)
         for pid, s in stats.items()
         if s["year_first"] and s["year_last"] and s["year_last"] > s["year_first"]],
        key=lambda x: (-x[3], x[0].lower())
    )
    row = table("LONGEST COMPETITIVE CAREERS",
                ["Player", "First Year", "Last Year", "Span (yrs)"],
                career_data, row)

    print("  STATISTICS done")


# ── ERA LEADERS sheet ──────────────────────────────────────────────────────────

def build_era_leaders(wb: Workbook, raw_results, events, discs, persons) -> None:
    ws = wb.create_sheet("ERA LEADERS")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A2"

    row = _title_row(ws, 1, "ERA LEADERS — BY DECADE", ncols=5)

    _ERAS = [
        ("1980s", 1980, 1989),
        ("1990s", 1990, 1999),
        ("2000s", 2000, 2009),
        ("2010s", 2010, 2019),
        ("2020s", 2020, 2029),
    ]
    TOP = 10

    def canon(pid):
        return persons.get(pid, {}).get("person_canon", pid)

    for era_name, yr_lo, yr_hi in _ERAS:
        # Gather placements in this era
        era_podiums:  dict[str, list] = defaultdict(lambda: [0, 0, 0])  # pid → [p1,p2,p3]
        era_events:   dict[str, set]  = defaultdict(set)

        for row_r in raw_results:
            pid = row_r.get("person_id", "").strip()
            if not pid or pid in _SKIP_PID:
                continue
            if row_r.get("display_name", "").strip() in _SKIP_DNAME:
                continue
            if row_r.get("participant_order", "1").strip() != "1":
                continue  # count placement once per slot

            eid  = row_r.get("event_id", "").strip()
            ev   = events.get(eid, {})
            try:
                yr = int(ev.get("year", "0") or 0)
            except ValueError:
                yr = 0
            if not (yr_lo <= yr <= yr_hi):
                continue

            try:
                place = int(row_r.get("placement", "0") or 0)
            except ValueError:
                place = 0

            if 1 <= place <= 3:
                era_podiums[pid][place - 1] += 1
            if eid:
                era_events[pid].add(eid)

        row = _section_row(ws, row, f"ERA: {era_name}  ({yr_lo}–{yr_hi})", ncols=5)

        # Top Podiums
        _w(ws, row, 1, "Top Podiums", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_L)
        row += 1
        row = _hrow(ws, row, "Player", "1st", "2nd", "3rd", "Total")
        pod_rows = sorted(
            [(canon(pid), p[0], p[1], p[2], sum(p))
             for pid, p in era_podiums.items() if pid in persons and sum(p) > 0],
            key=lambda x: (-x[4], x[0].lower())
        )
        for i, (name, p1, p2, p3, tot) in enumerate(pod_rows[:TOP]):
            fill = FILL_NET if i % 2 == 0 else FILL_EVENT_B
            row = _drow(ws, row, name, p1, p2, p3, tot, fill=fill)

        row += 1

        # Top Wins
        _w(ws, row, 1, "Top Wins", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_L)
        row += 1
        row = _hrow(ws, row, "Player", "Wins")
        win_rows = sorted(
            [(canon(pid), p[0]) for pid, p in era_podiums.items()
             if pid in persons and p[0] > 0],
            key=lambda x: (-x[1], x[0].lower())
        )
        for i, (name, wins) in enumerate(win_rows[:TOP]):
            fill = FILL_NET if i % 2 == 0 else FILL_EVENT_B
            row = _drow(ws, row, name, wins, fill=fill)

        row += 2

    print("  ERA LEADERS done")


# ── PLAYER STATS sheet ─────────────────────────────────────────────────────────

def build_player_stats(wb: Workbook, stats: dict, persons: dict) -> None:
    ws = wb.create_sheet("PLAYER STATS")

    headers = ["Player", "Country", "First Year", "Last Year",
               "Events", "Wins", "Podiums", "BAP", "FBHOF"]
    widths  = [34, 18, 11, 11, 9, 9, 9, 6, 6]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    row = _title_row(ws, 1, "PLAYER STATS — ALL IDENTIFIED COMPETITORS", ncols=len(headers))
    row = _hrow(ws, row, *headers)

    # Sort by person_canon
    rows_out = []
    for pid, p in sorted(persons.items(),
                         key=lambda kv: kv[1].get("person_canon", "").lower()):
        if not p.get("person_canon"):
            continue
        s = stats.get(pid, {})
        bap  = "Y" if p.get("bap_member",  "N") == "Y" else ""
        hof  = "Y" if p.get("fbhof_member","N") == "Y" else ""
        fill = FILL_BOTH if bap and hof else (FILL_HOF if hof else (FILL_BAP if bap else None))
        rows_out.append((
            p["person_canon"],
            p.get("country", ""),
            s.get("year_first", p.get("first_year", "")),
            s.get("year_last",  p.get("last_year",  "")),
            s.get("events",   0),
            s.get("wins",     0),
            s.get("podiums",  0),
            bap, hof,
            fill,
        ))

    for rec in rows_out:
        *vals, fill = rec
        r = row
        for col, v in enumerate(vals, 1):
            cell = _w(ws, r, col, v, font=FONT_DATA,
                      align=(ALIGN_R if isinstance(v, int) else ALIGN_L))
            if fill:
                cell.fill = fill
        row += 1

    print(f"  PLAYER STATS: {len(rows_out)} rows")


# ── EVENT INDEX sheet ──────────────────────────────────────────────────────────

def build_event_index(wb: Workbook, events: dict, discs: dict,
                      raw_results, quarantine: set,
                      event_col_map: dict) -> None:
    ws = wb.create_sheet("EVENT INDEX")

    headers = ["Year", "Event Name", "Location", "Event Type",
               "Disciplines", "Placements", "Source", "Coverage", "Notes"]
    widths  = [6, 52, 30, 26, 12, 11, 10, 10, 20]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    row = _title_row(ws, 1, "EVENT INDEX — ALL DOCUMENTED EVENTS", ncols=len(headers))
    row = _hrow(ws, row, *headers)

    # Count placements per event
    placements_per_event: dict[str, int] = Counter(
        r["event_id"] for r in raw_results
        if r.get("person_id", "").strip() not in _SKIP_PID
        and r.get("display_name", "").strip() not in _SKIP_DNAME
    )
    # Count disciplines per event
    discs_per_event: dict[str, int] = Counter(k[0] for k in discs)

    sorted_events = sorted(events.values(),
                           key=lambda e: (e.get("year",""), e.get("start_date",""), e.get("event_name","")))

    for ev in sorted_events:
        eid    = ev["event_id"]
        status = ev.get("status", "")
        is_quar = eid in quarantine
        n_discs = discs_per_event.get(eid, 0)
        n_plc   = placements_per_event.get(eid, 0)
        src     = ev.get("source_types", ev.get("data_source", ""))
        cov     = ev.get("validation_status", "")
        notes   = "⛔ quarantined" if is_quar else (
                  "no results"    if status == "no_results" else "")

        fill = FILL_QUAR if is_quar else (
               FILL_OTHER if status == "no_results" else None)

        # Hyperlink to year sheet if we have a defined name anchor
        sheet_ref = event_col_map.get(eid)
        name_val  = ev.get("event_name", eid)
        if sheet_ref:
            sheet_name, _ = sheet_ref
            # Use plain text (hyperlinks via defined names are fragile)
            pass

        for col, v in enumerate([
            ev.get("year",""),
            name_val,
            ev.get("location", ""),
            ev.get("event_type",""),
            n_discs or "",
            n_plc   or "",
            src, cov, notes,
        ], 1):
            cell = _w(ws, row, col, v, font=FONT_DATA, align=ALIGN_L)
            if fill:
                cell.fill = fill
        row += 1

    print(f"  EVENT INDEX: {len(sorted_events)} events")


# ── EXCLUDED EVENTS sheet ──────────────────────────────────────────────────────

def build_excluded_events(wb: Workbook, events: dict, quarantine: set) -> None:
    ws = wb.create_sheet("EXCLUDED EVENTS")

    headers = ["Year", "Event ID", "Event Name", "Location", "Reason"]
    widths  = [6, 28, 52, 30, 26]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    row = _title_row(ws, 1, "EXCLUDED EVENTS — NOT IN MAIN DATASET", ncols=len(headers))
    row = _hrow(ws, row, *headers)

    excluded = []
    for eid, ev in events.items():
        status = ev.get("status", "")
        if status == "no_results":
            reason = "no results published"
        elif eid in quarantine:
            reason = "quarantined (complex/uncertain data)"
        else:
            continue
        excluded.append((ev.get("year",""), eid, ev.get("event_name",""),
                         ev.get("location",""), reason))

    excluded.sort(key=lambda x: (x[0], x[2].lower()))
    for rec in excluded:
        for col, v in enumerate(rec, 1):
            _w(ws, row, col, v, font=FONT_DATA, align=ALIGN_L)
        row += 1

    print(f"  EXCLUDED EVENTS: {len(excluded)} entries")


# ── Year sheets ────────────────────────────────────────────────────────────────

def _team_display(slot_rows: list[dict]) -> str:
    """Combine participant_order=1 and =2 into 'Name1 / Name2'."""
    by_order: dict[int, str] = {}
    for r in slot_rows:
        try:
            order = int(r.get("participant_order", "1") or 1)
        except ValueError:
            order = 1
        name = r.get("display_name", "").strip()
        if name and name not in _SKIP_DNAME:
            by_order[order] = name
    if not by_order:
        return ""
    parts = [by_order.get(1, "?"), by_order.get(2, "")]
    parts = [p for p in parts if p and p != "?"]
    return " / ".join(parts) if len(parts) > 1 else (parts[0] if parts else "")


def build_year_sheets(wb: Workbook, raw_results, events: dict,
                      discs: dict, persons: dict,
                      quarantine: set) -> dict:
    """
    Build one sheet per year. Returns event_col_map = {event_id: (sheet_name, col)}.

    Year sheet column layout:
      Col A:  Row label (Event Name / Location / Date / Type / EID)
      Col B+: One column per event in that year

    Placement rows (below fixed metadata rows):
      One cell-block per discipline: header row (discipline name) + p1, p2, p3 rows.
      Different events in the same year may have different placements; rows are NOT
      shared across event columns — each column is self-contained vertically.
      Because of this we use an independent row-per-slot design within each column:
        • We pre-compute the global row layout for the year (max disciplines across events)
        • Each event column writes its divisions in the shared row slots; gaps are left blank.
    """
    # ── Group results by event and discipline ──────────────────────────────────
    # plcmt_by_event[eid][discipline] = sorted list of (placement, [participant rows])
    plcmt_by_event: dict[str, dict[str, dict[int, list]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )
    for r in raw_results:
        eid  = r["event_id"]
        disc = r["discipline"]
        try:
            place = int(r.get("placement", "0") or 0)
        except ValueError:
            place = 0
        plcmt_by_event[eid][disc][place].append(r)

    # ── Group events by year ───────────────────────────────────────────────────
    year_to_eids: dict[str, list] = defaultdict(list)
    for eid, ev in events.items():
        yr = ev.get("year", "").strip()
        if yr and yr.isdigit():
            # Only include events that have data OR are in quarantine
            if eid in plcmt_by_event or eid in quarantine:
                year_to_eids[yr].append(eid)

    event_col_map: dict[str, tuple] = {}

    # ── Fixed row offsets (shared across all event columns in a sheet) ─────────
    _R_NAME   = 1
    _R_LOC    = 2
    _R_DATE   = 3
    _R_TYPE   = 4
    _R_EID    = 5
    _R_DATA   = 7   # first row of placement data

    # ── Style constants for year sheets ───────────────────────────────────────
    _YF_META  = Font(bold=True, size=11)
    _YF_EID   = Font(size=8, color="808080")
    _YF_DIV   = Font(bold=True, size=10)
    _YF_PLC   = Font(size=10)
    _YF_QUAR  = Font(bold=True, size=11, color="CC0000")

    _COL_A_W  = 14     # column A (label column) width
    _COL_MIN  = 22     # min event column width
    _COL_MAX  = 52     # max event column width
    _TOP_PLCS = 3      # show top 3 placements per division

    for yr in sorted(year_to_eids, key=int):
        eids = sorted(
            year_to_eids[yr],
            key=lambda eid: (events[eid].get("start_date","") or "",
                             events[eid].get("event_name",""))
        )
        ws = wb.create_sheet(title=yr)

        # Column A — row label column
        ws.column_dimensions["A"].width = _COL_A_W
        label_map = {
            _R_NAME: "Event",
            _R_LOC:  "Location",
            _R_DATE: "Date",
            _R_TYPE: "Type",
            _R_EID:  "Event ID",
        }
        for r_idx, label in label_map.items():
            _w(ws, r_idx, 1, label,
               font=Font(size=9, color="606060"),
               align=Alignment(horizontal="right", vertical="top"))

        # ── Pre-compute the common discipline row layout for this year ─────────
        # We want each discipline to occupy the same rows across all event columns.
        # Strategy: collect all distinct disciplines across all events in this year,
        # assign each a fixed block of rows in a canonical order.
        # Discipline sort order: by discipline_category (net first) then name.
        all_discs_this_year: list[str] = []
        seen_discs: set[str] = set()
        for eid in eids:
            for disc_name in sorted(plcmt_by_event.get(eid, {}).keys(),
                                    key=lambda d: (
                                        discs.get((eid, d), {}).get("discipline_category","zzz"),
                                        d.lower()
                                    )):
                if disc_name not in seen_discs:
                    all_discs_this_year.append(disc_name)
                    seen_discs.add(disc_name)

        # Build row map: discipline → starting row (header), then _TOP_PLCS rows
        disc_row_start: dict[str, int] = {}
        cur_row = _R_DATA
        for disc_name in all_discs_this_year:
            disc_row_start[disc_name] = cur_row
            cur_row += 1 + _TOP_PLCS  # header + placements

        # Write row labels for discipline/placement rows in col A
        for disc_name, dr in disc_row_start.items():
            _w(ws, dr, 1, disc_name[:18],
               font=Font(size=8, italic=True, color="808080"),
               align=Alignment(horizontal="right", vertical="top"))
            for pi in range(1, _TOP_PLCS + 1):
                _w(ws, dr + pi, 1, f"  p{pi}",
                   font=Font(size=8, color="A0A0A0"),
                   align=Alignment(horizontal="right", vertical="top"))

        # ── Write event columns ────────────────────────────────────────────────
        for col_offset, eid in enumerate(eids, start=2):
            ev       = events[eid]
            is_quar  = eid in quarantine
            placements = plcmt_by_event.get(eid, {})

            col_letter = get_column_letter(col_offset)
            event_col_map[eid] = (yr, col_letter)

            # Event name
            ev_name = ev.get("event_name", eid)
            if is_quar:
                ev_name = "⛔ " + ev_name
            name_cell = _w(ws, _R_NAME, col_offset, ev_name,
                           font=(_YF_QUAR if is_quar else _YF_META),
                           align=ALIGN_LW)

            # Location
            loc = ev.get("location") or f"{ev.get('city','')}, {ev.get('country','')}".strip(", ")
            _w(ws, _R_LOC, col_offset, loc, font=_YF_PLC, align=ALIGN_L)

            # Date
            date_str = ev.get("start_date", "")
            _w(ws, _R_DATE, col_offset, date_str, font=_YF_PLC, align=ALIGN_L)

            # Type
            etype = ev.get("event_type", "")
            _w(ws, _R_TYPE, col_offset, etype, font=_YF_PLC, align=ALIGN_L)

            # Event ID (small)
            _w(ws, _R_EID, col_offset, eid, font=_YF_EID, align=ALIGN_L)

            # Define a named anchor for EVENT INDEX hyperlinking
            anchor_name = f"event_{eid}"
            safe_yr = yr.replace("'", "''")
            try:
                dn = DefinedName(
                    name=anchor_name,
                    attr_text=f"'{safe_yr}'!${col_letter}$1",
                )
                wb.defined_names[anchor_name] = dn
            except Exception:
                pass

            # ── Write discipline blocks ────────────────────────────────────────
            max_name_len = len(ev_name)

            for disc_name, disc_placements in placements.items():
                if disc_name not in disc_row_start:
                    continue  # shouldn't happen, but guard
                dr   = disc_row_start[disc_name]
                cat  = discs.get((eid, disc_name), {}).get("discipline_category", "")
                dfill = _cat_fill(cat)

                # Discipline header
                _w(ws, dr, col_offset, disc_name,
                   font=_YF_DIV, fill=dfill, align=ALIGN_L)
                max_name_len = max(max_name_len, len(disc_name))

                # Placements p1 .. p_TOP_PLCS
                for pi in range(1, _TOP_PLCS + 1):
                    slot_rows = disc_placements.get(pi, [])
                    disp = _team_display(slot_rows)
                    _w(ws, dr + pi, col_offset, disp,
                       font=_YF_PLC, align=ALIGN_L)
                    max_name_len = max(max_name_len, len(disp))

            # Auto column width
            ws.column_dimensions[col_letter].width = min(
                max(_COL_MIN, max_name_len + 2), _COL_MAX
            )

        # Fixed row heights
        ws.row_dimensions[_R_NAME].height = 32
        ws.row_dimensions[_R_LOC].height  = 14
        ws.row_dimensions[_R_DATE].height = 14
        ws.row_dimensions[_R_TYPE].height = 14
        ws.row_dimensions[_R_EID].height  = 11

        ws.freeze_panes = "B6"

        n_q = sum(1 for e in eids if e in quarantine)
        print(f"  {yr}: {len(eids)} events"
              + (f" ({n_q} quarantined)" if n_q else ""))

    return event_col_map


# ── Validation ─────────────────────────────────────────────────────────────────

def validate_workbook(wb: Workbook) -> dict:
    issues = []
    hidden_sheets = hidden_cols = hidden_rows = 0

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            hidden_sheets += 1
            issues.append(f"HIDDEN SHEET: {ws.title}")

        # Check column A hidden or zero-width
        cd = ws.column_dimensions.get("A")
        if cd and (cd.hidden or (cd.width is not None and cd.width <= 0)):
            hidden_cols += 1
            issues.append(f"HIDDEN/ZERO COL A: {ws.title}")

        for cd in ws.column_dimensions.values():
            if cd.hidden:
                hidden_cols += 1
                issues.append(f"HIDDEN COL {cd}: {ws.title}")

        for rd in ws.row_dimensions.values():
            if rd.hidden:
                hidden_rows += 1
                issues.append(f"HIDDEN ROW {rd.index}: {ws.title}")

    return {
        "hidden_sheets": hidden_sheets,
        "hidden_cols":   hidden_cols,
        "hidden_rows":   hidden_rows,
        "issues":        issues,
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"\nbuild_workbook_v15.py")
    print(f"  Source:  {CA}")
    print(f"  Output:  {OUTPUT_PATH}\n")

    # 1. Load data
    events, discs, raw_results, persons, quarantine = load_all()

    # 2. Compute statistics
    print("Computing statistics…")
    stats = compute_stats(raw_results, events, discs, persons)
    print(f"  Persons with stats: {len(stats)}")

    # 3. Build workbook
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    print("\nBuilding sheets…")

    build_readme(wb, events, persons, len(raw_results))
    build_statistics(wb, stats, persons)
    build_era_leaders(wb, raw_results, events, discs, persons)
    build_player_stats(wb, stats, persons)

    # Year sheets (returns event_col_map for EVENT INDEX)
    event_col_map = build_year_sheets(wb, raw_results, events, discs, persons, quarantine)

    # Front sheets that reference event_col_map
    build_event_index(wb, events, discs, raw_results, quarantine, event_col_map)
    build_excluded_events(wb, events, quarantine)

    # Reorder sheets so front matter comes before year sheets
    desired_front = ["README", "STATISTICS", "ERA LEADERS", "PLAYER STATS",
                     "EVENT INDEX", "EXCLUDED EVENTS"]
    final_order = desired_front + [s for s in wb.sheetnames if s not in desired_front]
    for i, name in enumerate(final_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    # 4. Validate
    print("\nValidating…")
    v = validate_workbook(wb)
    print(f"  Hidden sheets:  {v['hidden_sheets']}")
    print(f"  Hidden columns: {v['hidden_cols']}")
    print(f"  Hidden rows:    {v['hidden_rows']}")
    if v["issues"]:
        for iss in v["issues"][:10]:
            print(f"  ⚠  {iss}")
    else:
        print("  ✓ No hidden structure found")

    # 5. Save
    print(f"\nSaving → {OUTPUT_PATH.name}…")
    wb.save(str(OUTPUT_PATH))
    print(f"  Done.  ({OUTPUT_PATH.stat().st_size / 1_048_576:.1f} MB)")

    # 6. Summary
    print("\n── Summary ──────────────────────────────────────────────")
    year_sheets = [s for s in wb.sheetnames if s.isdigit()]
    print(f"  Sheets total:       {len(wb.sheetnames)}")
    print(f"  Front-matter:       {len(desired_front)}")
    print(f"  Year sheets:        {len(year_sheets)}"
          + (f"  ({year_sheets[0]}–{year_sheets[-1]})" if year_sheets else ""))
    print(f"  Events covered:     {len(event_col_map)}")
    print(f"  Persons in stats:   {len(stats)}")
    print(f"  Persons in PLAYER STATS: {len(persons)}")
    print(f"  Hidden structure:   {v['hidden_sheets'] + v['hidden_cols'] + v['hidden_rows']}")


if __name__ == "__main__":
    main()
