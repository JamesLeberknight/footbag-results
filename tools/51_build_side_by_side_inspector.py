#!/usr/bin/env python3
"""
Tool 51: Build Side-by-Side Event Inspector HTML
Generates out/review/event_side_by_side_inspector.html
showing 44 HIGH-heat events with raw mirror vs canonical placements.
"""

import csv
import json
import os
import re
import sys
from collections import defaultdict

import openpyxl

csv.field_size_limit(10**7)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = "/home/james/projects/FOOTBAG_DATA"
QUEUE_XLSX = f"{BASE}/out/review/Queue_HIGH_ONLY.xlsx"
STAGE2_CSV = f"{BASE}/out/stage2_canonical_events.csv"
PBP_CSV = f"{BASE}/inputs/identity_lock/Placements_ByPerson_v59.csv"
KNOWN_ISSUES_CSV = f"{BASE}/overrides/known_issues.csv"
EVENTS_OVERRIDES_JSONL = f"{BASE}/overrides/events_overrides.jsonl"
CANONICALIZE_PY = f"{BASE}/pipeline/02_canonicalize_results.py"
OUT_DIR = f"{BASE}/out/review"
OUT_HTML = f"{OUT_DIR}/event_side_by_side_inspector.html"


# ---------------------------------------------------------------------------
# HTML entity escape helper
# ---------------------------------------------------------------------------
def esc(s):
    if not s:
        return ""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


# ---------------------------------------------------------------------------
# 1. Load HIGH event list from xlsx
# ---------------------------------------------------------------------------
def load_high_events():
    wb = openpyxl.load_workbook(QUEUE_XLSX, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        events.append(d)
    wb.close()
    return events


# ---------------------------------------------------------------------------
# 2. Load raw placements from stage2_canonical_events.csv
# ---------------------------------------------------------------------------
def load_raw_placements(event_ids):
    """Returns dict: event_id -> list of placement dicts"""
    raw = {}
    with open(STAGE2_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            eid = row["event_id"]
            if eid in event_ids:
                try:
                    placements = json.loads(row["placements_json"] or "[]")
                except Exception:
                    placements = []
                raw[eid] = placements
    return raw


# ---------------------------------------------------------------------------
# 3. Load canonical placements from Placements_ByPerson_v59.csv
# ---------------------------------------------------------------------------
def load_canonical_placements(event_ids):
    """Returns dict: event_id -> list of canonical placement rows"""
    canon = defaultdict(list)
    with open(PBP_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            eid = row["event_id"]
            if eid in event_ids:
                canon[eid].append(row)
    return dict(canon)


# ---------------------------------------------------------------------------
# 4. Load known issues
# ---------------------------------------------------------------------------
def load_known_issues():
    issues = {}
    with open(KNOWN_ISSUES_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            issues[str(row["event_id"])] = {
                "severity": row.get("severity", ""),
                "note": row.get("note", ""),
            }
    return issues


# ---------------------------------------------------------------------------
# 5. Load events overrides
# ---------------------------------------------------------------------------
def load_events_overrides():
    overrides = defaultdict(list)
    with open(EVENTS_OVERRIDES_JSONL, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                d = json.loads(line)
                eid = str(d.get("event_id", ""))
                if eid:
                    overrides[eid].append(d)
            except Exception:
                pass
    return dict(overrides)


# ---------------------------------------------------------------------------
# 6. Load RESULTS_FILE_OVERRIDES event IDs from canonicalize script
# ---------------------------------------------------------------------------
def load_results_file_override_ids():
    ids = set()
    try:
        with open(CANONICALIZE_PY, encoding="utf-8") as f:
            content = f.read()
        ids = set(re.findall(r'"(\d{8,11})":\s*\{', content))
    except Exception as e:
        print(f"  Warning: could not read {CANONICALIZE_PY}: {e}", file=sys.stderr)
    return ids


# ---------------------------------------------------------------------------
# 7. Build text + HTML for raw placements panel
# ---------------------------------------------------------------------------
def build_raw_panel(placements, canon_divisions_set):
    """Returns (plain_text, html_str)"""
    # Group by division
    by_div = defaultdict(list)
    for p in placements:
        div = p.get("division", "") or "Unknown Division"
        by_div[div].append(p)

    # Sort divisions by first occurrence place
    div_order = list(by_div.keys())

    lines_text = []
    lines_html = []

    for div in div_order:
        entries = sorted(by_div[div], key=lambda x: (int(x.get("place", 999)) if str(x.get("place", "999")).isdigit() else 999))
        n = len(entries)
        header = f"{div} ({n} placement{'s' if n != 1 else ''})"

        # Determine highlight class
        if div in canon_divisions_set:
            cls = "div-header shared"
        else:
            cls = "div-header raw-only"

        lines_text.append(header)
        lines_html.append(f'<span class="{cls}">{esc(header)}</span>')

        for p in entries:
            place = p.get("place", "?")
            p1 = p.get("player1_name", "") or ""
            p2 = p.get("player2_name", "") or ""
            if p2:
                name = f"{p1} / {p2}"
            else:
                name = p1 or "[unknown]"
            line = f"  {place}. {name}"
            lines_text.append(line)
            lines_html.append(esc(line))

    plain = "\n".join(lines_text)
    html = "\n".join(lines_html)
    return plain, html


# ---------------------------------------------------------------------------
# 8. Build text + HTML for canonical placements panel
# ---------------------------------------------------------------------------
def build_canon_panel(pbp_rows, raw_divisions_set):
    """Returns (plain_text, html_str)"""
    by_div = defaultdict(list)
    for row in pbp_rows:
        div = row.get("division_canon", "") or "Unknown Division"
        by_div[div].append(row)

    div_order = list(by_div.keys())

    lines_text = []
    lines_html = []

    for div in div_order:
        entries = sorted(
            by_div[div],
            key=lambda x: (int(x.get("place", 999)) if str(x.get("place", "999")).lstrip("-").isdigit() else 999),
        )
        n = len(entries)
        header = f"{div} ({n} placement{'s' if n != 1 else ''})"

        if div in raw_divisions_set:
            cls = "div-header shared"
        else:
            cls = "div-header canon-only"

        lines_text.append(header)
        lines_html.append(f'<span class="{cls}">{esc(header)}</span>')

        # Deduplicate by place+person for display (teams appear once per person)
        seen_team_keys = set()
        for row in entries:
            place = row.get("place", "?")
            comp_type = row.get("competitor_type", "player")
            team_name = row.get("team_display_name", "") or ""
            person_canon = row.get("person_canon", "") or ""

            if comp_type == "team" and team_name:
                # Use team key to deduplicate
                key = (div, str(place), team_name)
                if key in seen_team_keys:
                    continue
                seen_team_keys.add(key)
                name = team_name
            else:
                if person_canon == "__NON_PERSON__":
                    name = "[unresolved]"
                else:
                    name = person_canon or "[unknown]"

            line = f"  {place}. {name}"
            lines_text.append(line)
            lines_html.append(esc(line))

    plain = "\n".join(lines_text)
    html = "\n".join(lines_html)
    return plain, html


# ---------------------------------------------------------------------------
# 9. Compute primary_review_reason
# ---------------------------------------------------------------------------
def compute_primary_review_reason(ev, known_issues, events_overrides):
    eid = str(ev.get("event_id", ""))
    div_raw = ev.get("division_count_raw") or 0
    div_canon = ev.get("division_count_canonical") or 0
    pl_raw = ev.get("placement_count_raw") or 0
    pl_canon = ev.get("placement_count_canonical") or 0
    diff_summary = (ev.get("diff_summary") or "").lower()

    has_div_mismatch = (div_raw > 0 or div_canon > 0) and (div_raw != div_canon)
    has_pl_mismatch = abs(pl_raw - pl_canon) > 3
    has_known_issue = eid in known_issues
    has_override = bool(ev.get("override_flag"))
    has_recently_modified = bool(ev.get("recently_modified_flag"))
    has_text_cleanup = "text cleanup" in diff_summary

    if has_div_mismatch:
        return "division_mismatch"
    if has_pl_mismatch:
        return "placement_mismatch"
    if has_known_issue:
        return "known_issue"
    if has_override:
        return "override_used"
    if has_recently_modified:
        return "recently_modified"
    if has_text_cleanup:
        return "text_cleanup_candidate"
    # Check if there's truly no interesting data difference
    return "metadata_only"


# ---------------------------------------------------------------------------
# 10. Compute groups
# ---------------------------------------------------------------------------
def compute_groups(ev, primary_reason):
    groups = []
    if primary_reason == "division_mismatch":
        groups.append("division_mismatch")
    if primary_reason == "placement_mismatch":
        groups.append("placement_mismatch")
    if ev.get("known_issue_flag"):
        groups.append("known_issue")
    if ev.get("override_flag"):
        groups.append("override_used")
    if ev.get("recently_modified_flag"):
        groups.append("recently_modified")
    if str(ev.get("count_mismatch_flag", "") or "").upper() == "YES":
        groups.append("count_mismatch")
    if primary_reason == "metadata_only":
        groups.append("metadata_only")
    return groups


# ---------------------------------------------------------------------------
# 11. Compute diff_notes
# ---------------------------------------------------------------------------
def compute_diff_notes(ev, raw_divs, canon_divs, known_issues):
    eid = str(ev.get("event_id", ""))
    div_raw = ev.get("division_count_raw") or 0
    div_canon = ev.get("division_count_canonical") or 0
    pl_raw = ev.get("placement_count_raw") or 0
    pl_canon = ev.get("placement_count_canonical") or 0

    notes = []

    if div_raw != div_canon and (div_raw > 0 or div_canon > 0):
        raw_only = sorted(raw_divs - canon_divs)
        canon_only = sorted(canon_divs - raw_divs)
        parts = []
        if raw_only:
            parts.append(f"Raw-only divs: [{', '.join(raw_only)}]")
        if canon_only:
            parts.append(f"Canon-only divs: [{', '.join(canon_only)}]")
        if parts:
            notes.append(" | ".join(parts))
        else:
            notes.append(f"Division count mismatch: raw={div_raw} canon={div_canon}")

    if abs(pl_raw - pl_canon) > 3:
        diff = pl_raw - pl_canon
        sign = "+" if diff > 0 else ""
        notes.append(f"Placement delta: raw={pl_raw} vs canonical={pl_canon} (diff={sign}{diff})")

    if eid in known_issues:
        ki = known_issues[eid]
        notes.append(f"Known issue [{ki['severity']}]: {ki['note']}")

    if not notes:
        notes.append(f"Counts match (div={div_raw}, placements={pl_raw})")

    return " | ".join(notes)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Loading HIGH event list...")
    high_events = load_high_events()
    event_ids = {str(ev["event_id"]) for ev in high_events}
    print(f"  {len(high_events)} HIGH events loaded")

    print("Loading known issues...")
    known_issues = load_known_issues()

    print("Loading events overrides...")
    events_overrides = load_events_overrides()

    print("Loading RESULTS_FILE_OVERRIDES IDs...")
    results_override_ids = load_results_file_override_ids()
    print(f"  {len(results_override_ids)} RESULTS_FILE_OVERRIDES entries found")

    print("Loading raw placements from stage2...")
    raw_placements = load_raw_placements(event_ids)
    print(f"  Loaded for {len(raw_placements)} events")

    print("Loading canonical placements from PBP v59...")
    canon_placements = load_canonical_placements(event_ids)
    print(f"  Loaded for {len(canon_placements)} events")

    print("Building event data...")
    event_data = []
    empty_raw = []
    empty_canon = []

    for ev in high_events:
        eid = str(ev.get("event_id", ""))

        # Raw placements
        raw_list = raw_placements.get(eid, [])
        raw_divs = set(p.get("division", "") or "" for p in raw_list)
        raw_divs.discard("")

        # Canon placements
        canon_list = canon_placements.get(eid, [])
        canon_divs = set(r.get("division_canon", "") or "" for r in canon_list)
        canon_divs.discard("")

        # Build panels
        raw_text, raw_html = build_raw_panel(raw_list, canon_divs)
        canon_text, canon_html = build_canon_panel(canon_list, raw_divs)

        if not raw_text.strip():
            empty_raw.append(eid)
        if not canon_text.strip():
            empty_canon.append(eid)

        # Compute derived fields
        primary_reason = compute_primary_review_reason(ev, known_issues, events_overrides)
        groups = compute_groups(ev, primary_reason)
        diff_notes = compute_diff_notes(ev, raw_divs, canon_divs, known_issues)
        known_issue_note = known_issues.get(eid, {}).get("note", "")

        obj = {
            "event_id": eid,
            "year": ev.get("year"),
            "event_name": ev.get("event_name") or "",
            "start_date": str(ev.get("start_date") or ""),
            "end_date": str(ev.get("end_date") or ""),
            "city": ev.get("city") or "",
            "region": ev.get("region") or "",
            "country": ev.get("country") or "",
            "location_raw": ev.get("location_raw") or "",
            "host_club": ev.get("host_club") or "",
            "review_heat_score": ev.get("review_heat_score") or 0,
            "review_heat_label": ev.get("review_heat_label") or "",
            "diff_summary": ev.get("diff_summary") or "",
            "known_issue_flag": 1 if ev.get("known_issue_flag") else 0,
            "known_issue_type": ev.get("known_issue_type") or "",
            "override_flag": 1 if ev.get("override_flag") else 0,
            "recently_modified_flag": 1 if ev.get("recently_modified_flag") else 0,
            "reviewer_flag": 1 if ev.get("reviewer_flag") else 0,
            "division_count_raw": ev.get("division_count_raw") or 0,
            "division_count_canonical": ev.get("division_count_canonical") or 0,
            "placement_count_raw": ev.get("placement_count_raw") or 0,
            "placement_count_canonical": ev.get("placement_count_canonical") or 0,
            "raw_text": raw_text,
            "raw_html": raw_html,
            "canonical_text": canon_text,
            "canonical_html": canon_html,
            "diff_notes": diff_notes,
            "known_issue_note": known_issue_note,
            "primary_review_reason": primary_reason,
            "review_focus_rank": ev.get("review_focus_rank") or 999,
            "groups": groups,
            "count_mismatch_flag": str(ev.get("count_mismatch_flag") or ""),
        }
        event_data.append(obj)

    # Sort: review_focus_rank asc, review_heat_score desc
    event_data.sort(key=lambda x: (x["review_focus_rank"], -x["review_heat_score"]))

    # Group breakdown
    group_counts = defaultdict(int)
    for ev in event_data:
        for g in ev["groups"]:
            group_counts[g] += 1

    print(f"\nGroup breakdown:")
    for g, c in sorted(group_counts.items()):
        print(f"  {g}: {c}")

    print(f"\nEmpty raw panels: {len(empty_raw)}")
    if empty_raw:
        print(f"  event_ids: {empty_raw}")
    print(f"Empty canonical panels: {len(empty_canon)}")
    if empty_canon:
        print(f"  event_ids: {empty_canon}")

    # Build HTML
    print("\nBuilding HTML...")
    os.makedirs(OUT_DIR, exist_ok=True)
    html = build_html(event_data, group_counts)
    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    size = os.path.getsize(OUT_HTML)
    print(f"\nOutput: {OUT_HTML}")
    print(f"File size: {size:,} bytes ({size/1024:.1f} KB)")
    print(f"Events embedded: {len(event_data)}")

    return empty_raw, empty_canon, group_counts, len(event_data)


# ---------------------------------------------------------------------------
# HTML builder
# ---------------------------------------------------------------------------
def build_html(event_data, group_counts):
    events_json = json.dumps(event_data, ensure_ascii=False, indent=None)

    all_groups = [
        ("division_mismatch", "Div Mismatch"),
        ("placement_mismatch", "Placement Mismatch"),
        ("known_issue", "Known Issue"),
        ("override_used", "Override Used"),
        ("recently_modified", "Recently Modified"),
        ("count_mismatch", "Count Mismatch"),
        ("metadata_only", "Metadata Only"),
    ]

    group_buttons_html = ""
    for gkey, glabel in all_groups:
        cnt = group_counts.get(gkey, 0)
        group_buttons_html += f'<button class="group-btn" data-group="{gkey}" onclick="filterGroup(this)">{glabel} ({cnt})</button>\n'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Footbag Event Side-by-Side Inspector</title>
<style>
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

body {{
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
  font-size: 14px;
  background: #f8f8f8;
  color: #222;
}}

/* ---- NAV BAR ---- */
#navbar {{
  position: sticky;
  top: 0;
  z-index: 100;
  background: #fff;
  border-bottom: 2px solid #ccc;
  padding: 8px 12px;
}}

.nav-row {{
  display: flex;
  align-items: center;
  gap: 8px;
  flex-wrap: wrap;
  margin-bottom: 6px;
}}

.nav-counter {{
  font-weight: bold;
  font-size: 15px;
  min-width: 80px;
}}

button.nav-btn {{
  padding: 5px 14px;
  cursor: pointer;
  background: #e0e0e0;
  border: 1px solid #aaa;
  border-radius: 4px;
  font-size: 13px;
}}
button.nav-btn:hover {{ background: #ccc; }}

select#jumpSelect {{
  padding: 4px 8px;
  font-size: 13px;
  max-width: 360px;
}}

.group-filters {{
  display: flex;
  gap: 6px;
  flex-wrap: wrap;
  align-items: center;
}}

.group-filters label {{
  font-size: 12px;
  font-weight: bold;
  color: #555;
  margin-right: 4px;
}}

button.group-btn {{
  padding: 3px 10px;
  cursor: pointer;
  background: #f0f0f0;
  border: 1px solid #bbb;
  border-radius: 12px;
  font-size: 12px;
}}
button.group-btn:hover {{ background: #ddd; }}
button.group-btn.active {{ background: #4a90d9; color: #fff; border-color: #2b6cb0; }}
button#btnAll {{
  padding: 3px 10px;
  cursor: pointer;
  background: #4a90d9;
  color: #fff;
  border: 1px solid #2b6cb0;
  border-radius: 12px;
  font-size: 12px;
}}
button#btnAll:hover {{ background: #2b6cb0; }}
button#btnAll.inactive {{ background: #f0f0f0; color: #333; border-color: #bbb; }}

/* ---- EVENT HEADER ---- */
#eventHeader {{
  background: #fff;
  border: 1px solid #ddd;
  border-radius: 4px;
  margin: 8px 12px;
  padding: 10px 14px;
  display: grid;
  grid-template-columns: 1fr auto;
  gap: 12px;
}}

.event-meta h2 {{
  font-size: 16px;
  margin-bottom: 4px;
}}
.event-meta .meta-row {{
  font-size: 12px;
  color: #555;
  margin-bottom: 3px;
}}

.badge {{
  display: inline-block;
  padding: 2px 8px;
  border-radius: 10px;
  font-size: 11px;
  font-weight: bold;
  margin-right: 4px;
}}
.badge-HIGH {{ background: #f5c6cb; color: #721c24; }}
.badge-MEDIUM {{ background: #ffeeba; color: #856404; }}
.badge-LOW {{ background: #d4edda; color: #155724; }}
.badge-division_mismatch {{ background: #f8d7da; color: #721c24; }}
.badge-placement_mismatch {{ background: #fff3cd; color: #856404; }}
.badge-known_issue {{ background: #f5c6cb; color: #721c24; }}
.badge-override_used {{ background: #cce5ff; color: #004085; }}
.badge-recently_modified {{ background: #d4edda; color: #155724; }}
.badge-text_cleanup_candidate {{ background: #e2e3e5; color: #383d41; }}
.badge-metadata_only {{ background: #e2e3e5; color: #383d41; }}
.badge-other {{ background: #e2e3e5; color: #383d41; }}

.count-ok {{ color: #155724; }}
.count-mismatch {{ color: #721c24; font-weight: bold; }}

/* ---- REVIEW STUB ---- */
.review-stub {{
  min-width: 220px;
  background: #f9f9f9;
  border: 1px solid #ddd;
  border-radius: 4px;
  padding: 10px;
  font-size: 13px;
}}
.review-stub label {{
  display: block;
  font-size: 11px;
  color: #555;
  margin-bottom: 2px;
  margin-top: 8px;
}}
.review-stub label:first-child {{ margin-top: 0; }}
.review-stub select, .review-stub textarea {{
  width: 100%;
  font-size: 12px;
  padding: 3px 5px;
  border: 1px solid #ccc;
  border-radius: 3px;
}}
.review-stub textarea {{
  height: 60px;
  resize: vertical;
}}

.export-btn {{
  margin-top: 8px;
  padding: 5px 10px;
  background: #28a745;
  color: #fff;
  border: none;
  border-radius: 3px;
  cursor: pointer;
  font-size: 12px;
  width: 100%;
}}
.export-btn:hover {{ background: #218838; }}

/* ---- DIFF BOX ---- */
#diffBox {{
  margin: 0 12px 8px;
  padding: 8px 12px;
  border-radius: 4px;
  font-size: 12px;
  border: 2px solid #ccc;
  background: #fafafa;
}}
#diffBox.red-border {{ border-color: #dc3545; background: #fff5f5; }}
#diffBox.orange-border {{ border-color: #fd7e14; background: #fff8f0; }}
#diffBox.gray-border {{ border-color: #adb5bd; background: #fafafa; }}

/* ---- PANELS ---- */
#panels {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 12px;
  margin: 0 12px 12px;
}}

.panel-wrap {{
  background: #fff;
  border: 1px solid #ccc;
  border-radius: 4px;
  overflow: hidden;
}}

.panel-title {{
  background: #e9ecef;
  padding: 6px 10px;
  font-size: 12px;
  font-weight: bold;
  border-bottom: 1px solid #ccc;
}}

.panel-content {{
  padding: 10px;
  overflow-y: auto;
  max-height: 55vh;
}}

.panel-content pre {{
  font-family: 'Courier New', 'Consolas', monospace;
  font-size: 12px;
  white-space: pre-wrap;
  word-break: break-word;
  line-height: 1.5;
}}

.div-header {{
  display: block;
  padding: 1px 3px;
  border-radius: 2px;
  font-weight: bold;
}}
.div-header.raw-only {{ background: #fff3cd; }}
.div-header.canon-only {{ background: #cce5ff; }}
.div-header.shared {{ background: #d4edda; }}

.empty-panel {{
  color: #999;
  font-style: italic;
  padding: 20px;
  text-align: center;
}}

/* ---- LEGEND ---- */
.legend {{
  margin: 4px 12px 8px;
  font-size: 11px;
  color: #555;
  display: flex;
  gap: 12px;
  flex-wrap: wrap;
}}
.legend-item {{
  display: flex;
  align-items: center;
  gap: 4px;
}}
.legend-swatch {{
  display: inline-block;
  width: 14px;
  height: 14px;
  border-radius: 2px;
  border: 1px solid #bbb;
}}
</style>
</head>
<body>

<div id="navbar">
  <div class="nav-row">
    <button class="nav-btn" onclick="prevEvent()">&#8592; Prev</button>
    <span class="nav-counter" id="navCounter">1 / 44</span>
    <button class="nav-btn" onclick="nextEvent()">Next &#8594;</button>
    <span style="margin: 0 4px; color:#aaa;">|</span>
    <span style="font-size:12px; color:#555;">Jump to:</span>
    <select id="jumpSelect" onchange="jumpTo(this.value)">
    </select>
    <span style="margin: 0 4px; color:#aaa;">|</span>
    <button class="nav-btn" onclick="exportStubs()" style="background:#28a745;color:#fff;border-color:#28a745;">Export Stubs CSV</button>
  </div>
  <div class="group-filters">
    <label>Filter:</label>
    <button id="btnAll" onclick="filterAll()">All (__TOTAL__)</button>
    {group_buttons_html}
  </div>
</div>

<div class="legend">
  <span class="legend-item"><span class="legend-swatch" style="background:#d4edda;"></span> Division in both</span>
  <span class="legend-item"><span class="legend-swatch" style="background:#fff3cd;"></span> Raw-only division</span>
  <span class="legend-item"><span class="legend-swatch" style="background:#cce5ff;"></span> Canon-only division</span>
</div>

<div id="eventHeader">
  <div class="event-meta" id="eventMeta"></div>
  <div class="review-stub" id="reviewStub"></div>
</div>

<div id="diffBox" class="gray-border"></div>

<div id="panels">
  <div class="panel-wrap">
    <div class="panel-title" id="rawTitle">RAW (stage2)</div>
    <div class="panel-content" id="rawPanel"></div>
  </div>
  <div class="panel-wrap">
    <div class="panel-title" id="canonTitle">CANONICAL (PBP v59)</div>
    <div class="panel-content" id="canonPanel"></div>
  </div>
</div>

<script>
const events = {events_json};

// Review stubs storage: event_id -> {{status, action, note}}
const stubs = {{}};

let filtered = events.slice();
let currentIdx = 0;
let activeGroup = null;

function getStub(eid) {{
  if (!stubs[eid]) {{
    stubs[eid] = {{ status: 'UNREVIEWED', action: 'NONE', note: '' }};
  }}
  return stubs[eid];
}}

function saveStub(eid) {{
  const s = document.getElementById('stubStatus');
  const a = document.getElementById('stubAction');
  const n = document.getElementById('stubNote');
  if (s && a && n) {{
    stubs[eid] = {{ status: s.value, action: a.value, note: n.value }};
  }}
}}

function prevEvent() {{
  if (filtered.length === 0) return;
  const eid = filtered[currentIdx] && filtered[currentIdx].event_id;
  if (eid) saveStub(eid);
  currentIdx = (currentIdx - 1 + filtered.length) % filtered.length;
  showEvent(currentIdx);
}}

function nextEvent() {{
  if (filtered.length === 0) return;
  const eid = filtered[currentIdx] && filtered[currentIdx].event_id;
  if (eid) saveStub(eid);
  currentIdx = (currentIdx + 1) % filtered.length;
  showEvent(currentIdx);
}}

function jumpTo(val) {{
  const idx = parseInt(val);
  if (isNaN(idx)) return;
  const eid = filtered[currentIdx] && filtered[currentIdx].event_id;
  if (eid) saveStub(eid);
  currentIdx = idx;
  showEvent(currentIdx);
}}

function filterGroup(btn) {{
  const eid = filtered[currentIdx] && filtered[currentIdx].event_id;
  if (eid) saveStub(eid);

  const group = btn.getAttribute('data-group');
  document.querySelectorAll('.group-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById('btnAll').classList.add('inactive');

  activeGroup = group;
  filtered = events.filter(e => e.groups.includes(group));
  currentIdx = 0;
  rebuildJump();
  updateCounter();
  if (filtered.length > 0) showEvent(0);
  else clearPanels();
}}

function filterAll() {{
  const eid = filtered[currentIdx] && filtered[currentIdx].event_id;
  if (eid) saveStub(eid);

  document.querySelectorAll('.group-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('btnAll').classList.remove('inactive');
  activeGroup = null;
  filtered = events.slice();
  currentIdx = 0;
  rebuildJump();
  updateCounter();
  if (filtered.length > 0) showEvent(0);
}}

function rebuildJump() {{
  const sel = document.getElementById('jumpSelect');
  sel.innerHTML = '';
  filtered.forEach((ev, i) => {{
    const opt = document.createElement('option');
    opt.value = i;
    opt.textContent = `[${{ev.year || '?'}}] ${{ev.event_id}} — ${{ev.event_name.substring(0, 50)}}`;
    sel.appendChild(opt);
  }});
}}

function updateCounter() {{
  document.getElementById('navCounter').textContent = `${{currentIdx + 1}} / ${{filtered.length}}`;
  document.getElementById('jumpSelect').value = currentIdx;
}}

function clearPanels() {{
  document.getElementById('eventMeta').innerHTML = '<p style="color:#999">No events in this filter.</p>';
  document.getElementById('reviewStub').innerHTML = '';
  document.getElementById('diffBox').innerHTML = '';
  document.getElementById('rawPanel').innerHTML = '<div class="empty-panel">No events.</div>';
  document.getElementById('canonPanel').innerHTML = '<div class="empty-panel">No events.</div>';
  document.getElementById('navCounter').textContent = '0 / 0';
}}

function showEvent(idx) {{
  currentIdx = idx;
  updateCounter();

  const ev = filtered[idx];
  if (!ev) {{ clearPanels(); return; }}

  // Header meta
  const divRaw = ev.division_count_raw;
  const divCanon = ev.division_count_canonical;
  const plRaw = ev.placement_count_raw;
  const plCanon = ev.placement_count_canonical;
  const divClass = divRaw !== divCanon ? 'count-mismatch' : 'count-ok';
  const plClass = Math.abs(plRaw - plCanon) > 3 ? 'count-mismatch' : 'count-ok';

  let flags = '';
  if (ev.known_issue_flag) flags += '<span class="badge badge-known_issue">Known Issue</span>';
  if (ev.override_flag) flags += '<span class="badge badge-override_used">Override</span>';
  if (ev.recently_modified_flag) flags += '<span class="badge badge-recently_modified">Recently Modified</span>';
  if (ev.reviewer_flag) flags += '<span class="badge" style="background:#e2d9f3;color:#432874">Reviewer</span>';
  if (ev.count_mismatch_flag === 'YES') flags += '<span class="badge count-mismatch">Count Mismatch</span>';

  const reasonLabel = ev.primary_review_reason.replace(/_/g, ' ');
  const reasonBadge = `<span class="badge badge-${{ev.primary_review_reason}}">${{reasonLabel}}</span>`;

  const locParts = [ev.city, ev.region, ev.country].filter(Boolean);
  const loc = locParts.join(', ');
  const dates = ev.start_date !== ev.end_date && ev.end_date
    ? `${{ev.start_date}} – ${{ev.end_date}}`
    : ev.start_date;

  document.getElementById('eventMeta').innerHTML = `
    <h2>[${{ev.year || '?'}}] ${{escHtml(ev.event_name)}}</h2>
    <div class="meta-row">
      <strong>ID:</strong> ${{ev.event_id}}
      &nbsp;|&nbsp; <strong>Heat:</strong> <span class="badge badge-${{ev.review_heat_label}}">${{ev.review_heat_label}} (${{ev.review_heat_score}})</span>
      &nbsp;${{reasonBadge}}
      ${{flags}}
    </div>
    <div class="meta-row">
      <strong>Location:</strong> ${{escHtml(loc) || '—'}}
      ${{ev.host_club ? '&nbsp;|&nbsp; <strong>Club:</strong> ' + escHtml(ev.host_club) : ''}}
      &nbsp;|&nbsp; <strong>Dates:</strong> ${{escHtml(dates) || '—'}}
    </div>
    <div class="meta-row">
      <strong>Divisions:</strong>
      <span class="${{divClass}}">raw=${{divRaw}} / canon=${{divCanon}}</span>
      &nbsp;&nbsp;
      <strong>Placements:</strong>
      <span class="${{plClass}}">raw=${{plRaw}} / canon=${{plCanon}}</span>
    </div>
    <div class="meta-row"><strong>Diff summary:</strong> ${{escHtml(ev.diff_summary)}}</div>
  `;

  // Review stub
  const stub = getStub(ev.event_id);
  document.getElementById('reviewStub').innerHTML = `
    <label>Status</label>
    <select id="stubStatus">
      ${{['UNREVIEWED','ACCEPT','MINOR_TEXT_FIX','STRUCTURAL_FIX','SOURCE_PARTIAL','DEFER']
        .map(o => `<option${{o === stub.status ? ' selected' : ''}}>${{o}}</option>`).join('')}}
    </select>
    <label>Action</label>
    <select id="stubAction">
      ${{['NONE','PARSER_FIX','PBP_PATCH','RESULTS_FILE_OVERRIDE','IDENTITY_FIX','METADATA_FIX','DOCUMENT_ONLY']
        .map(o => `<option${{o === stub.action ? ' selected' : ''}}>${{o}}</option>`).join('')}}
    </select>
    <label>Note</label>
    <textarea id="stubNote" placeholder="Review notes...">${{escHtml(stub.note)}}</textarea>
    <button class="export-btn" onclick="exportStubs()">Export Stubs CSV</button>
  `;

  // Diff box
  let diffClass = 'gray-border';
  if (divRaw !== divCanon) diffClass = 'red-border';
  else if (Math.abs(plRaw - plCanon) > 3) diffClass = 'orange-border';
  const diffBox = document.getElementById('diffBox');
  diffBox.className = `${{diffClass}}`;
  diffBox.innerHTML = `<strong>Diff:</strong> ${{escHtml(ev.diff_notes)}}`;

  // Panels
  document.getElementById('rawTitle').textContent = `RAW (stage2) — ${{plRaw}} placements`;
  document.getElementById('canonTitle').textContent = `CANONICAL (PBP v59) — ${{plCanon}} placements`;

  const rawPanel = document.getElementById('rawPanel');
  const canonPanel = document.getElementById('canonPanel');

  if (ev.raw_html && ev.raw_html.trim()) {{
    rawPanel.innerHTML = `<pre>${{ev.raw_html}}</pre>`;
  }} else {{
    rawPanel.innerHTML = '<div class="empty-panel">No raw placements available.</div>';
  }}

  if (ev.canonical_html && ev.canonical_html.trim()) {{
    canonPanel.innerHTML = `<pre>${{ev.canonical_html}}</pre>`;
  }} else {{
    canonPanel.innerHTML = '<div class="empty-panel">No canonical placements available.</div>';
  }}
}}

function escHtml(s) {{
  if (!s) return '';
  return String(s)
    .replace(/&/g, '&amp;')
    .replace(/</g, '&lt;')
    .replace(/>/g, '&gt;')
    .replace(/"/g, '&quot;');
}}

function exportStubs() {{
  // Save current stub first
  const ev = filtered[currentIdx];
  if (ev) saveStub(ev.event_id);

  const rows = [['event_id','year','event_name','review_status','action_needed','review_note','primary_review_reason']];
  events.forEach(e => {{
    const s = getStub(e.event_id);
    rows.push([e.event_id, e.year, e.event_name, s.status, s.action, s.note, e.primary_review_reason]);
  }});

  const csv = rows.map(r => r.map(v => `"${{String(v || '').replace(/"/g, '""')}}"` ).join(',')).join('\\n');
  const blob = new Blob([csv], {{ type: 'text/csv;charset=utf-8;' }});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url;
  a.download = 'event_review_stubs.csv';
  a.click();
  URL.revokeObjectURL(url);
}}

// Init
rebuildJump();
updateCounter();
if (filtered.length > 0) showEvent(0);
</script>
</body>
</html>
"""

    # Replace placeholder
    html = html.replace("__TOTAL__", str(len(event_data)))
    return html


if __name__ == "__main__":
    empty_raw, empty_canon, group_counts, n_events = main()
