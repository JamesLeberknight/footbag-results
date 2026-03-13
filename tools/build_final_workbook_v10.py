#!/usr/bin/env python3
"""
build_final_workbook_v10.py
Part 8: Build Final Community Workbook v10

Changes from v9:
- Year sheets rebuilt from PBP data (out/Placements_ByPerson.csv) so that
  ALL placements are visible, including unresolved (person_unresolved=1) and
  __NON_PERSON__ doubles entries with a team_display_name.
  Previously these rows were silently filtered out by 04B; they now appear
  as-is, giving the complete historical record.
- ALL-CAPS text normalization applied to every text cell in every year sheet
  and to front-matter sheets when copying (division names, event type labels,
  category headers like "NET", "FREESTYLE" are intentional abbreviations and
  are left unchanged).  Truly ALL-CAPS person names / event names are
  title-cased.
- clean_display_name() applied to all person_canon values before formatting.

Approach for year sheets: APPROACH A (rebuilt from PBP, not copied from
Community.xlsx), with the same formatting logic as 04B but without the
__NON_PERSON__ / unresolved filter.

Front matter (README, DATA NOTES, STATISTICS, EVENT INDEX, PLAYER SUMMARY,
CONSECUTIVE RECORDS) are copied unchanged from v8.

Produces: Footbag_Results_Community_FINAL_v9.xlsx

Sheet order:
1. README
2. DATA NOTES
3. STATISTICS
4. EVENT INDEX
5. PLAYER SUMMARY
6. CONSECUTIVE RECORDS
7. Year sheets: 1980, 1981, ... 2025, 2026
"""

import copy
import csv
import json
import os
import re
import sys
import unicodedata
from collections import OrderedDict, defaultdict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

csv.field_size_limit(sys.maxsize)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SOURCE_V9_PATH  = os.path.join(BASE_DIR, "Footbag_Results_Community_FINAL_v9.xlsx")
OUTPUT_PATH     = os.path.join(BASE_DIR, "Footbag_Results_Community_FINAL_v10.xlsx")
PBP_CSV         = os.path.join(BASE_DIR, "out", "Placements_Flat.csv")
STAGE2_EVENTS   = os.path.join(BASE_DIR, "out", "stage2_canonical_events.csv")
QUARANTINE_CSV  = os.path.join(BASE_DIR, "inputs", "review_quarantine_events.csv")
EVENTS_NORM_CSV = os.path.join(BASE_DIR, "out", "canonical", "events_normalized.csv")

FRONT_SHEETS = ["README", "DATA NOTES", "STATISTICS",
                "EVENT INDEX", "PLAYER SUMMARY", "CONSECUTIVE RECORDS"]

# ── Palette & styles (mirrors 04B) ────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

def _font(**kw) -> Font:
    return Font(**kw)

def _border_top(color="BBBBBB") -> Border:
    return Border(top=Side(style="thin", color=color))

FILL_BANNER        = _fill("1F4E79")
FILL_BANNER_WORLDS = _fill("7D4607")
FILL_META          = _fill("EBF3FB")
FILL_PLAYERS       = _fill("F5F5F5")
FILL_DIV           = _fill("E2E2E2")
FILL_GOLD          = _fill("FFF3CC")
FILL_SILVER        = _fill("F0F0F0")
FILL_BRONZE        = _fill("FDEBD0")
FILL_WHITE         = _fill("FFFFFF")
FILL_CAT           = _fill("D0D0D0")
FILL_ROW_LABEL     = _fill("F0F4F8")
FILL_STATUS_WARN   = _fill("FFF9C4")

FONT_BANNER  = Font(bold=True,   size=12, color="FFFFFF")
FONT_META    = Font(             size=9,  color="1F4E79")
FONT_HOST    = Font(italic=True, size=9,  color="444444")
FONT_PLAYERS = Font(             size=9,  color="888888")
FONT_DIV     = Font(bold=True,   size=9)
FONT_PODIUM  = Font(bold=True,   size=9)
FONT_PLACE   = Font(             size=9)
FONT_ROW_LBL = Font(             size=8,  color="888888")
FONT_CAT     = Font(bold=True,   size=8,  color="444444")
FONT_STATUS_WARN = Font(bold=True, size=8, color="996600")

ALIGN_WRAP   = Alignment(wrap_text=True, vertical="top")
ALIGN_TOP    = Alignment(vertical="top")
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="top")

MEDALS = {1: "🥇", 2: "🥈", 3: "🥉"}

COL_W_MIN   = 24
COL_W_LABEL = 12

# Row numbers in year sheets
_R_NAME    = 1
_R_LOC     = 2
_R_HOST    = 3
_R_DATE    = 4
_R_PLAYERS = 5
_R_EVTYPE  = 6
_R_EID     = 7
_R_STATUS  = 8
_R_DATA    = 9

_ROW_LABELS = {
    _R_NAME:    "Event",
    _R_LOC:     "Location",
    _R_HOST:    "Host Club",
    _R_DATE:    "Date",
    _R_PLAYERS: "Players",
    _R_EVTYPE:  "Event Type",
    _R_EID:     "Event ID",
    _R_STATUS:  "Status",
}

CAT_ORDER  = ["net", "freestyle", "golf", "sideline", "unknown"]
CAT_LABELS = {
    "net":       "Net",
    "freestyle": "Freestyle",
    "golf":      "Golf",
    "sideline":  "Sideline",
    "unknown":   "Other",
}

# ── Name / display helpers ─────────────────────────────────────────────────────

_RE_MOJI_QUOTE   = re.compile(r"Ï(.+?)Ó")
_RE_REPL_UPPER   = re.compile(r"\ufffd([A-Z])")
_RE_TEAM_REPL_UP = re.compile(r"\ufffd([A-Z])")
_RE_APOS_CORRUPT = re.compile(r"\b(\w+)\?[Ss]\b")
_RE_STAR         = re.compile(r"^\*+\s*")
_RE_BBU          = re.compile(r"\[/?U\]", re.I)
_RE_TRAIL_DASH   = re.compile(r"\s*-\s*$")
_RE_QUESTION_SEP = re.compile(r"\s+\?\s+")
_RE_EVENT_ISO2   = re.compile(r"[¿\u00bf](?=[a-zA-Z])")
_RE_ANNOTATION   = re.compile(r"\s*\(([^)]+)\)\s*$")
_RE_CTRL         = re.compile(r"[\x00-\x1f\x7f\x80-\x9f]")

_TRANSLIT = str.maketrans("łŁøØðÐđĐ", "lloodddd")

_WEEKDAYS  = {"monday","tuesday","wednesday","thursday","friday","saturday","sunday"}
_COUNTRIES = {
    "czech republic","germany","france","poland","switzerland","venezuela",
    "colombia","slovakia","austria","sweden","finland","usa","canada",
    "australia","spain","russia","brazil","argentina","mexico","netherlands",
    "belgium","norway","denmark","hungary","ukraine","czech",
}

_NAME_CORRECTIONS: dict = {
    "Alexandre B\ufffdlanger":           "Alexandre Bélanger",
    "Andr\ufffd Lemaire":                "André Lemaire",
    "Carlos M\ufffdRquez":               "Carlos Marquez",
    "Chris L\ufffdW":                    "Chris Löw",
    "Filip W\ufffdJcik":                 "Filip Wojcik",
    "Fran\ufffdois Leh":                 "François Leh",
    "Fran\ufffdois Pelletier":           "Francois Pelletier",
    "Genevi\ufffdve Bousquet":           "Genevieve Bousquet",
    "Gosia D\ufffdBska":                 "Gosia Debska",
    "Heike K\ufffdLler":                 "Heike Köller",
    "Jean-Francois B\ufffdLanger":       "Jean François Bélanger",
    "Jean-Fran\ufffdois Lemieux":        "Jean-Francois Lemieux",
    "Kinga Gw\ufffd\u017add\u017c":      "Kinga Gwozdz",
    "Klemenz L\ufffdNgauer":             "Klemenz Längauer",
    "Krzysztof Sob\ufffdTka":            "Krzysztof Sobótka",
    "L\ufffdA L'Esp\ufffdRance":         "Léa Lespérance",
    "Marcin Staro\ufffd":               "Marcin Staron",
    "Martin C\ufffdT\ufffd":            "Martin Cote",
    "Martin Sl\ufffdDek":               "Martin Sladek",
    "Mał\ufffdGorzata D\u0119B\ufffdSka":"Malgorzata Debska",
    "Mał\ufffdGorzata Ol\u0119Dzka":    "Malgorzata Oledzka",
    "Micha\ufffd R\ufffdG":             "Micha Rog",
    "Oskari Forst\ufffdN":              "Oskari Forstén",
    "Petteri Pet\ufffdInen":            "Petteri Petäinen",
    "Piia Tantarim\ufffdKi":            "Piia Tantarimäki",
    "Rados\ufffdAw Turek":              "Rados Turek",
    "Robin P\ufffdChel":                "Robin Puchel",
    "S\ufffdBastien Duchesne":          "Sebastien Duchesne",
    "S\ufffdBastien Maillet":           "Sébastien Maillet",
    "St\ufffdPhane Tailleur":           "Stéphane Tailleur",
    "Tuomas K\ufffdRki":                "Tuomas Karki",
    "Ulrike H\ufffd\ufffdLer":          "Ulrike Häßler",
    "Wiktor D\ufffdBski":               "Wiktor Debski",
    "\ufffdUkasz Domin":                "Lukasz Domin",
    "Florian G\ufffdTze":               "Florian Goetze",
    "J. B\ufffdHm":                     "Jule Böhm",
    "Renato Z\ufffdLli":                "Renatto Zülli",
    "Thomas F\ufffdRster":              "Thomas Forster",
    "Olivier Berthiaume-Berg\ufffdE":   "Olivier B.-Bergé",
    "Tomá\u00b9 Tu\u00e8ek":           "Tomáš Tuček",
    "Ale\u00b9 Pelko":                 "Aleš Pelko",
    "Paweł \u00a6cierski":             "Paweł Ścierski",
    "Rafał Kaleta":                     "Rafał Kaleta",
    "Kinga Gwó\u00bcd\u00bc":          "Kinga Gwóźdź",
    "Robin P\u00b8chel":               "Robin Puchel",
}


def clean_display_name(name: str) -> str:
    """Clean a player/field name for display.
    - Strip leading/trailing whitespace
    - Remove U+FFFD replacement characters
    - Remove C0/C1 control characters (0-31 except space, 128-159)
    - Collapse multiple spaces to single space
    """
    if not name:
        return ""
    s = str(name).strip()
    s = s.replace("\ufffd", "")
    s = _RE_CTRL.sub("", s)
    s = re.sub(r"  +", " ", s)
    return s.strip()


# Short CAPS tokens that are intentional abbreviations — leave unchanged.
_CAPS_EXEMPT = {
    "OK", "USA", "UK",
}

def fix_caps(value: str) -> str:
    """Normalize ALL-CAPS strings to title case.
    - Only applies if ALL alphabetic characters are uppercase AND len > 3
    - Leaves short strings (<=3) alone (abbreviations like 'NET' already exempt)
    - Does NOT touch mixed-case strings
    - Exempt set covers intentional category labels
    """
    if not isinstance(value, str):
        return value
    if not value:
        return value
    if value in _CAPS_EXEMPT:
        return value
    alpha = [c for c in value if c.isalpha()]
    if not alpha:
        return value
    if len(alpha) <= 3:
        return value
    if all(c.isupper() for c in alpha):
        return value.title()
    return value


def _fix_name_encoding(p: str) -> str:
    """Apply encoding corrections to a single team member name."""
    if p in _NAME_CORRECTIONS:
        return _NAME_CORRECTIONS[p]
    if "Ï" in p or "Ó" in p:
        p = _RE_MOJI_QUOTE.sub(lambda m: f'"{m.group(1)}"', p)
    if "\ufffd" in p:
        p = _RE_TEAM_REPL_UP.sub(lambda m: m.group(1).lower(), p)
        p = p.replace("\ufffd", "")
    return p


def _strip_annotation_tail(p: str) -> str:
    """Remove trailing parenthetical annotations (not short codes like '(BC)')."""
    while True:
        m = _RE_ANNOTATION.search(p)
        if not m:
            break
        content = m.group(1)
        if len(content) > 4 and " " in content:
            p = p[: m.start()].rstrip()
        else:
            break
    return p


def _clean_team_display(s: str) -> str:
    """Clean team display names: remove noise tokens, capitalize, strip annotation tails."""
    s = (s or "").strip()
    if "/" not in s:
        return s
    parts = [p.strip() for p in s.split("/", 1)]
    cleaned = []
    for p in parts:
        if p.lower() in _WEEKDAYS or p.lower() in _COUNTRIES:
            cleaned.append("[?]")
        else:
            p = _fix_name_encoding(p)
            p = _strip_annotation_tail(p)
            if p and p[0].islower():
                p = p[0].upper() + p[1:]
            cleaned.append(p)
    return " / ".join(cleaned)


def _clean_div(s: str) -> str:
    """Strip workbook-visible markup and encoding artifacts from division names."""
    s = (s or "").strip()
    s = _RE_STAR.sub("", s)
    s = _RE_BBU.sub("", s)
    s = _RE_TRAIL_DASH.sub("", s)
    s = _RE_QUESTION_SEP.sub(" - ", s)
    s = _RE_APOS_CORRUPT.sub(lambda m: m.group(1) + "'s", s)
    s = _RE_REPL_UPPER.sub(lambda m: m.group(1).lower(), s)
    s = s.replace("\ufffd", "")
    return s.strip()


def _display_name(s: str) -> str:
    """Return s ready for workbook display.
    - Fixes encoding artifacts
    - Title-cases ALL-CAPS strings
    """
    s = clean_display_name(s or "")
    if not s:
        return s
    s = _RE_MOJI_QUOTE.sub(lambda m: f'"{m.group(1)}"', s)
    alpha = [c for c in s if c.isalpha()]
    if alpha and all(c.isupper() for c in alpha):
        return s.title()
    return s


def _count_participants(ep: dict) -> int:
    """Count actual participants; doubles entries (containing ' / ') count as 2."""
    return sum(2 if " / " in disp else 1 for v in ep.values() for (_, disp, _) in v)


# ── Data loading ──────────────────────────────────────────────────────────────

def load_quarantine_ids() -> set:
    ids = set()
    with open(QUARANTINE_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ids.add(str(row["event_id"]))
    return ids


def load_normalized_locations() -> dict:
    """
    Returns dict legacy_event_id → canonical location string for year-sheet display.
    Applies the same presentation rule as EVENT INDEX:
      US/Canada → "City, State/Province, Country"
      All others → "City, Country"
    Skips region for non-US/CA to avoid duplicates like "Stara Zagora, Stara Zagora, Bulgaria".
    """
    locs = {}
    try:
        with open(EVENTS_NORM_CSV, newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                eid = str(row.get("legacy_event_id") or "").strip()
                if not eid:
                    continue
                city    = (row.get("city") or "").strip()
                region  = (row.get("region") or "").strip()
                country = (row.get("country") or "").strip()
                if not city and not country:
                    continue
                if country in ("United States", "Canada"):
                    parts = [p for p in [city, region, country] if p]
                else:
                    parts = [p for p in [city, country] if p]
                locs[eid] = ", ".join(parts)
    except FileNotFoundError:
        pass
    return locs


def load_stage2_events(quarantine_ids: set, norm_locs: dict) -> dict:
    """
    Returns dict event_id → {event_id, year, event_name, date, location,
                              host_club, event_type, div_order}
    """
    events = {}
    with open(STAGE2_EVENTS, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.DictReader(fh):
            eid = str(row["event_id"]).strip()
            if eid in quarantine_ids:
                continue
            try:
                placements = json.loads(row.get("placements_json") or "[]")
            except (json.JSONDecodeError, TypeError):
                placements = []
            seen, div_order = set(), []
            for p in placements:
                dc = (p.get("division_canon") or "").strip()
                if dc and dc not in seen:
                    div_order.append(dc)
                    seen.add(dc)
            event_name = fix_caps(clean_display_name((row.get("event_name") or "").strip()))
            host_club  = fix_caps(clean_display_name((row.get("host_club") or "").strip()))
            # Prefer normalized location; fall back to stage2 location string
            location = norm_locs.get(eid) or (row.get("location") or "").strip()
            events[eid] = {
                "event_id":   eid,
                "year":       int(row.get("year") or 0),
                "event_name": event_name,
                "date":       (row.get("date") or "").strip(),
                "location":   location,
                "host_club":  host_club,
                "event_type": (row.get("event_type") or "").strip(),
                "div_order":  div_order,
            }
    return events


def load_pbp(quarantine_ids: set) -> dict:
    """
    Returns dict event_id → OrderedDict{division_canon: [(place_int, display, cat)]}.

    KEY DIFFERENCE FROM 04B: includes unresolved AND __NON_PERSON__ entries.
    For __NON_PERSON__ doubles: use team_display_name.
    For __NON_PERSON__ singles: show the raw person_canon string ("__NON_PERSON__"
        is replaced with a placeholder showing raw player_name via norm column).
    For unresolved: show person_canon as-is (it's the raw name, typically
        first-name-only or abbreviated).
    """
    # Group PBP rows by event_id
    rows_by_event: dict = defaultdict(list)
    with open(PBP_CSV, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.DictReader(fh):
            eid = str(row.get("event_id", "")).strip()
            if eid and eid not in quarantine_ids:
                rows_by_event[eid].append(row)

    result: dict = {}
    for eid, rows in rows_by_event.items():
        div_placements: dict = {}

        # Group by division_canon
        by_div: dict = defaultdict(list)
        for row in rows:
            dc = _clean_div((row.get("division_canon") or "").rstrip(":").strip())
            by_div[dc].append(row)

        for div_canon, drows in by_div.items():
            # Sort by place, then team entries before solo entries (so team display
            # is recorded first and the redundant solo entry is deduped away),
            # then team_person_key, then person_canon.
            def sort_key(r):
                try:
                    pl = int(float(r.get("place") or 0))
                except (ValueError, TypeError):
                    pl = 9999
                tpk = r.get("team_person_key", "")
                is_solo = 1 if not tpk else 0  # 0 = team row first
                return (pl, is_solo, tpk, r.get("person_canon", ""))

            drows_sorted = sorted(drows, key=sort_key)

            entries = []
            seen_teams: set = set()
            seen_place_name: set = set()   # dedup (place, display)
            seen_place_person: set = set() # dedup (place, person_canon) — prevents same person appearing solo AND in a team
            # Determine if this is a true team division
            # Use ALL rows (not just visible ones — since we show everything now)
            team_count = sum(1 for r in drows_sorted
                             if (r.get("competitor_type") or "player").lower() == "team")
            is_team_division = len(drows_sorted) > 0 and (team_count / len(drows_sorted)) > 0.5

            for row in drows_sorted:
                person       = (row.get("person_canon") or "").strip()
                comp         = (row.get("competitor_type") or "player").lower()
                tpk          = (row.get("team_person_key") or "").strip()
                cat          = (row.get("division_category") or "").strip()
                team_display = _clean_team_display((row.get("team_display_name") or "").strip())
                unresolved   = (row.get("person_unresolved") or "").strip()
                is_unresolved = unresolved in ("1", "true", "True")
                is_non_person = person == "__NON_PERSON__"

                try:
                    place_int = int(float(row.get("place") or 0))
                except (ValueError, TypeError):
                    continue
                if place_int <= 0:
                    continue

                if comp == "team" and tpk:
                    # Doubles: deduplicate by team_person_key
                    if tpk in seen_teams:
                        continue
                    seen_teams.add(tpk)
                    if team_display and not team_display.rstrip().endswith("/ ?"):
                        display = team_display
                    elif is_non_person and not team_display:
                        # No usable display name — skip
                        continue
                    else:
                        # Build from members
                        members = [r.get("person_canon","") for r in drows_sorted
                                   if r.get("team_person_key","") == tpk
                                   and r.get("person_canon","") not in ("","__NON_PERSON__")]
                        display = " / ".join(_display_name(m) for m in members if m)
                        if not display:
                            if team_display:
                                display = team_display
                            else:
                                continue
                elif comp == "team" and team_display:
                    # team_person_key missing — dedup on display string
                    if team_display in seen_teams:
                        continue
                    seen_teams.add(team_display)
                    if team_display.rstrip().endswith("/ ?"):
                        continue
                    display = team_display
                elif is_non_person and not is_unresolved:
                    # Singleton __NON_PERSON__ in singles division — no display name
                    continue
                elif is_unresolved:
                    # Show raw name (unresolved means a real person not yet in PT)
                    display = _display_name(person) if person else "[unknown]"
                else:
                    display = _display_name(person)

                key = (place_int, display)
                if key in seen_place_name:
                    continue
                seen_place_name.add(key)
                # Also skip if this person already appears at this place in a team entry
                person_key = (place_int, _display_name(person))
                if person_key in seen_place_person:
                    continue
                # Record all person names that appear in this display (team: both members)
                for name_part in display.split(" / "):
                    seen_place_person.add((place_int, name_part.strip()))
                entries.append((place_int, display, cat))

            if entries:
                div_placements[div_canon] = entries

        result[eid] = div_placements

    return result


def merge_div_order(pbp_placements: dict, stage2_events: dict) -> dict:
    """
    For each event, return an OrderedDict with divisions in stage2 div_order
    first, then any additional divisions from pbp.
    """
    result: dict = {}
    for eid, raw_divs in pbp_placements.items():
        div_order = stage2_events.get(eid, {}).get("div_order", [])
        # Clean div_order keys
        clean_order = [_clean_div(dc.rstrip(":").strip()) for dc in div_order]
        ordered: OrderedDict = OrderedDict()
        for dc in clean_order:
            if dc in raw_divs:
                ordered[dc] = raw_divs[dc]
        for dc, dp in raw_divs.items():
            if dc not in ordered:
                ordered[dc] = dp
        result[eid] = ordered

    return result


# ── Sheet copy utilities (for front matter) ───────────────────────────────────

def copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        for attr in ("font", "fill", "alignment", "border", "number_format"):
            try:
                val = getattr(src_cell, attr)
                setattr(dst_cell, attr,
                        copy.copy(val) if attr != "number_format" else val)
            except Exception:
                pass
    if src_cell.hyperlink:
        try:
            dst_cell.hyperlink = copy.copy(src_cell.hyperlink)
        except Exception:
            pass


def copy_sheet_to(src_ws, dst_ws):
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width  = col_dim.width or 8.43
        dst_ws.column_dimensions[col_letter].hidden = col_dim.hidden
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height or 15
    for row in src_ws.iter_rows():
        for cell in row:
            copy_cell(cell, dst_ws.cell(row=cell.row, column=cell.column))
    for merge_range in src_ws.merged_cells.ranges:
        try:
            dst_ws.merge_cells(str(merge_range))
        except Exception:
            pass
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref


# ── Year sheet construction ────────────────────────────────────────────────────

def _c(ws, row: int, col: int, value=None, *,
       font=None, fill=None, align=None, border=None):
    """Write a cell with optional formatting."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font   is not None: cell.font      = font
    if fill   is not None: cell.fill      = fill
    if align  is not None: cell.alignment = align
    if border is not None: cell.border    = border


def _date_sort_key(date_str: str, eid: str) -> tuple:
    """Return (month_number, day, eid) for chronological sorting within a year."""
    MONTHS = {m: i for i, m in enumerate([
        "january","february","march","april","may","june",
        "july","august","september","october","november","december"], 1)}
    m = re.match(r"([A-Za-z]+)\s+(\d+)", (date_str or "").strip())
    if m:
        mon = MONTHS.get(m.group(1).lower(), 13)
        day = int(m.group(2))
        return (mon, day, eid)
    return (13, 0, eid)


def build_year_sheet(wb: Workbook, year: int, eids: list,
                     events: dict, event_placements: dict) -> None:
    """
    Build one year sheet with:
    - Column A: row labels
    - Columns B onward: one event per column
    - freeze_panes = "B1"
    - Auto column width (min COL_W_MIN)
    """
    ws = wb.create_sheet(title=str(year))

    # Column A: row-label column
    ws.column_dimensions["A"].width = COL_W_LABEL
    for row_num, label in _ROW_LABELS.items():
        _c(ws, row_num, 1, label,
           font=FONT_ROW_LBL, fill=FILL_ROW_LABEL, align=ALIGN_RIGHT)

    sorted_eids = sorted(
        eids,
        key=lambda eid: _date_sort_key(events[eid].get("date", ""), eid),
    )

    col_max_widths: dict = {}

    for col_offset, eid in enumerate(sorted_eids, start=2):
        ev         = events[eid]
        placements = event_placements.get(eid, OrderedDict())
        n_players  = _count_participants(placements)
        max_content = max(len(ev.get("event_name", "")), 24)

        def _write(r, val, font, fill, align=ALIGN_TOP):
            nonlocal max_content
            if val:
                max_content = max(max_content, len(str(val)))
            _c(ws, r, col_offset, val, font=font, fill=fill, align=align)

        is_worlds   = ev.get("event_type", "") == "worlds"
        banner_fill = FILL_BANNER_WORLDS if is_worlds else FILL_BANNER

        # Fix ISO-8859-2 artifact in event name (¿ → ż)
        _evt_name = _RE_EVENT_ISO2.sub("\u017c", ev["event_name"])

        _write(_R_NAME,    _evt_name,                              FONT_BANNER,  banner_fill, ALIGN_WRAP)
        _write(_R_LOC,     ev["location"] or "—",                  FONT_META,    FILL_META)
        _write(_R_HOST,    ev["host_club"] or "Not recorded",      FONT_HOST,    FILL_META)
        _write(_R_DATE,    ev["date"]      or "Not recorded",      FONT_META,    FILL_META)
        _write(_R_PLAYERS, f"Players: {n_players}",                FONT_PLAYERS, FILL_PLAYERS)
        _write(_R_EVTYPE,  ev.get("event_type") or "Not recorded", FONT_META,    FILL_META)
        _write(_R_EID,     ev.get("event_id") or "",               FONT_ROW_LBL, FILL_META)

        row = _R_DATA

        # Group divisions by category, preserving order within category
        cat_to_divs: dict = {}
        for div_name, entries in placements.items():
            if not entries:
                continue
            cat = (entries[0][2] or "unknown")
            if cat not in cat_to_divs:
                cat_to_divs[cat] = []
            cat_to_divs[cat].append((div_name, entries))

        for cat in CAT_ORDER:
            if cat not in cat_to_divs:
                continue
            cat_label = CAT_LABELS.get(cat, "Other")
            divs_in_cat = cat_to_divs[cat]
            # Suppress category header when there is exactly one division and its
            # display name is the same as the category label (e.g. "Freestyle" / "Freestyle")
            single_div_matches = (
                len(divs_in_cat) == 1
                and fix_caps(_clean_div(divs_in_cat[0][0])).lower() == cat_label.lower()
            )
            if not single_div_matches:
                _c(ws, row, col_offset, cat_label, font=FONT_CAT, fill=FILL_CAT, align=ALIGN_TOP)
                max_content = max(max_content, len(cat_label) + 2)
                row += 1

            for div_name, entries in divs_in_cat:
                div_name = _clean_div(div_name)
                # Apply fix_caps to division name — fixes "OPEN SINGLES NET" etc.
                # but leaves short abbreviations alone
                div_name_display = fix_caps(div_name)
                _c(ws, row, col_offset, div_name_display,
                   font=FONT_DIV, fill=FILL_DIV, border=_border_top(), align=ALIGN_TOP)
                max_content = max(max_content, len(div_name_display) + 2)
                row += 1

                for place_int, display, _ in entries:
                    medal = MEDALS.get(place_int, "")

                    # Apply fix_caps to the player/team name portion
                    # Apply fix_caps per name segment so mixed-case team strings
                    # like "Pavel Hejra / PETR FUCIK" get each part fixed independently
                    display_fixed = " / ".join(
                        fix_caps(part) for part in display.split(" / ")
                    )

                    parts = []
                    if medal:
                        parts.append(medal)
                    parts.append(f"{place_int:>3} ")
                    parts.append(display_fixed)
                    text = " ".join(parts) if medal else "".join(parts)

                    if place_int == 1:
                        fill, font = FILL_GOLD,   FONT_PODIUM
                    elif place_int == 2:
                        fill, font = FILL_SILVER, FONT_PODIUM
                    elif place_int == 3:
                        fill, font = FILL_BRONZE, FONT_PODIUM
                    else:
                        fill, font = FILL_WHITE,  FONT_PLACE

                    _c(ws, row, col_offset, text, font=font, fill=fill, align=ALIGN_TOP)
                    max_content = max(max_content, len(text) + 2)
                    row += 1

                row += 1  # blank between divisions

        col_max_widths[col_offset] = max_content

    # Set column widths
    for col_i, max_w in col_max_widths.items():
        col_letter = get_column_letter(col_i)
        ws.column_dimensions[col_letter].width = max(COL_W_MIN, min(max_w + 2, 60))

    # Row heights
    ws.row_dimensions[_R_NAME].height   = 30
    ws.row_dimensions[_R_STATUS].height = 15
    ws.freeze_panes = "B1"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading quarantine list...")
    quarantine_ids = load_quarantine_ids()
    print(f"  {len(quarantine_ids)} quarantined events")

    print("Loading normalized locations...")
    norm_locs = load_normalized_locations()
    print(f"  {len(norm_locs)} normalized location entries")

    print("Loading stage2 events...")
    stage2_events = load_stage2_events(quarantine_ids, norm_locs)
    print(f"  {len(stage2_events)} events loaded")

    print("Loading Placements_ByPerson (PBP) — including unresolved + __NON_PERSON__...")
    raw_pbp = load_pbp(quarantine_ids)
    print(f"  {len(raw_pbp)} events with placements")

    print("Merging division order from stage2...")
    event_placements = merge_div_order(raw_pbp, stage2_events)

    # Count total visible placements (all entries including unresolved)
    total_placements = sum(
        len(entries)
        for ep in event_placements.values()
        for entries in ep.values()
    )
    print(f"  Total placement entries (including unresolved): {total_placements:,}")

    # Group events by year
    events_by_year: dict = defaultdict(list)
    for eid, ev in stage2_events.items():
        year = ev.get("year", 0)
        if year and eid in event_placements:
            events_by_year[year].append(eid)
    years = sorted(events_by_year.keys())
    print(f"  Years: {years[0]}–{years[-1]} ({len(years)} years)")

    print(f"\nOpening source workbook (v9) for front matter: {SOURCE_V9_PATH}")
    src_wb = openpyxl.load_workbook(SOURCE_V9_PATH)
    print(f"  Sheets in v9: {src_wb.sheetnames[:8]} ...")

    print("\nCreating output workbook...")
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    # Status code strings that must not be title-cased
    FRONT_CAPS_EXEMPT = {
        "QUARANTINED", "OK", "ACCEPT", "DEFER",
        "DATASET OVERVIEW", "EVENTS BY YEAR",
        "DISCIPLINE HISTORY", "GEOGRAPHIC DISTRIBUTION",
        "README", "DATA NOTES", "STATISTICS",
        "EVENT INDEX", "PLAYER SUMMARY", "CONSECUTIVE RECORDS",
        "SOURCE_PARTIAL", "KNOWN_ISSUE",
    }

    def _fix_front_cell_caps(value):
        """Apply fix_caps to a front-sheet cell, exempting known status codes."""
        if not isinstance(value, str):
            return value
        if value in FRONT_CAPS_EXEMPT:
            return value
        # Only fix strings where ALL alpha chars are uppercase and len > 4
        alpha = [c for c in value if c.isalpha()]
        if not alpha or len(alpha) <= 3:
            return value
        if all(c.isupper() for c in alpha):
            return value.title()
        return value

    # ── 1–6: Copy front matter from v8, applying caps fix to EVENT INDEX ──────
    for sheet_name in FRONT_SHEETS:
        if sheet_name in src_wb.sheetnames:
            print(f"  Copying {sheet_name}...")
            dst_ws = out_wb.create_sheet(sheet_name)
            copy_sheet_to(src_wb[sheet_name], dst_ws)
            # Apply caps normalization to EVENT INDEX data cells
            if sheet_name == "EVENT INDEX":
                fixed_count = 0
                for row in dst_ws.iter_rows():
                    for cell in row:
                        new_val = _fix_front_cell_caps(cell.value)
                        if new_val != cell.value:
                            cell.value = new_val
                            fixed_count += 1
                if fixed_count:
                    print(f"    Fixed {fixed_count} ALL-CAPS cells in EVENT INDEX")
        else:
            print(f"  WARNING: {sheet_name} not found in v8 source")

    # ── 7+: Build year sheets from PBP data ───────────────────────────────────
    print(f"\nBuilding {len(years)} year sheets from PBP data...")
    for year in years:
        eids = events_by_year[year]
        build_year_sheet(out_wb, year, eids, stage2_events, event_placements)
        print(f"  {year}: {len(eids)} events", end="", flush=True)
        # Spot-check: Bulgaria 2025
        if year == 2025:
            bg_eid = "1739036206"
            if bg_eid in event_placements:
                bg_ep = event_placements[bg_eid]
                sn = bg_ep.get("Singles Net", [])
                print(f"  [Bulgaria Singles Net: {len(sn)} entries]", end="")
        print()

    # ── Validation ────────────────────────────────────────────────────────────
    print("\nValidating output workbook...")
    sheet_names = out_wb.sheetnames
    print(f"  Total sheets: {len(sheet_names)}")

    expected_front = FRONT_SHEETS
    actual_front   = list(sheet_names[:6])
    if actual_front == expected_front:
        print("  Front sheet order: CORRECT")
    else:
        print(f"  WARNING: Front sheet order mismatch")
        print(f"    Expected: {expected_front}")
        print(f"    Got:      {actual_front}")

    year_sheets = [s for s in sheet_names if s.isdigit()]
    print(f"  Year sheets: {year_sheets[0]}–{year_sheets[-1]} ({len(year_sheets)} sheets)")

    # Spot-check: Bulgaria 2025 Singles Net
    ws2025 = out_wb["2025"] if "2025" in out_wb.sheetnames else None
    if ws2025:
        # Find Bulgaria column
        bg_col = None
        for row in ws2025.iter_rows(values_only=True):
            for j, v in enumerate(row):
                if str(v) == "1739036206":
                    bg_col = j
                    break
            if bg_col is not None:
                break
        if bg_col is not None:
            print(f"\n  Bulgaria 2025 at column {bg_col + 1}:")
            for i, row in enumerate(ws2025.iter_rows(values_only=True)):
                if bg_col < len(row) and row[bg_col] is not None:
                    print(f"    Row {i+1}: {str(row[bg_col])[:60]}")
        else:
            print("  WARNING: Bulgaria 2025 column not found in 2025 sheet")

    # ── Save ──────────────────────────────────────────────────────────────────
    print(f"\nSaving to: {OUTPUT_PATH}")
    out_wb.save(OUTPUT_PATH)
    size_bytes = os.path.getsize(OUTPUT_PATH)
    size_mb    = size_bytes / (1024 * 1024)
    print(f"Saved: {size_mb:.1f} MB ({size_bytes:,} bytes)")
    print("\nDone.")


if __name__ == "__main__":
    main()
