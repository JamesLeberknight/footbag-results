#!/usr/bin/env python3
"""
04B_create_community_excel.py

Produces Footbag_Results_Community.xlsx — a reader-friendly Footbag Historical
Almanac targeted at the footbag community.

Read-only transformation only.  No identity changes, no canonical mutations.

Inputs  (from out/):
    stage2_canonical_events.csv  — event metadata + division source order
    index.csv                    — clean display metadata (name, date, location)
    Placements_Flat.csv          — identity-resolved placements
    Placements_ByPerson.csv      — for leaderboard computation
    Persons_Truth.csv            — for honours matching

Inputs  (from inputs/):
    bap_data.csv                 — Big Add Posse inductees
    fbhof_data.csv               — Footbag Hall of Fame inductees

Output:
    Footbag_Results_Community.xlsx
"""

import csv
import json
import sys
import unicodedata
from collections import defaultdict, OrderedDict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

csv.field_size_limit(10_000_000)

REPO      = Path(__file__).resolve().parent.parent
OUT_DIR   = REPO / "out"
INPUT_DIR = REPO / "inputs"
XLSX      = REPO / "Footbag_Results_Community.xlsx"


# ── Palette & styles ──────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

def _font(**kw) -> Font:
    return Font(**kw)

def _border_top(color="BBBBBB") -> Border:
    return Border(top=Side(style="thin", color=color))

def _border_bottom(color="BBBBBB") -> Border:
    return Border(bottom=Side(style="thin", color=color))

FILL_BANNER   = _fill("1F4E79")   # dark navy  — event banner
FILL_META     = _fill("EBF3FB")   # pale blue  — location / host / date
FILL_PLAYERS  = _fill("F5F5F5")   # near-white — players count
FILL_DIV      = _fill("E2E2E2")   # light grey — division header
FILL_GOLD     = _fill("FFF3CC")   # soft gold  — 1st place
FILL_SILVER   = _fill("F0F0F0")   # near-white — 2nd place
FILL_BRONZE   = _fill("FDEBD0")   # pale orange— 3rd place
FILL_WHITE    = _fill("FFFFFF")
FILL_HDR      = _fill("1F4E79")   # sheet header row
FILL_HON_BAP  = _fill("FFF8E1")   # BAP honour row tint
FILL_HON_FBHOF= _fill("E8F5E9")   # FBHOF honour row tint
FILL_HON_BOTH = _fill("F3E5F5")   # both honours row tint
FILL_ROW_LABEL= _fill("F0F4F8")   # column-A row labels in year sheets

FONT_BANNER   = Font(bold=True,   size=12, color="FFFFFF")
FONT_META     = Font(             size=9,  color="1F4E79")
FONT_HOST     = Font(italic=True, size=9,  color="444444")
FONT_PLAYERS  = Font(             size=9,  color="888888")
FONT_DIV      = Font(bold=True,   size=9)
FONT_PODIUM   = Font(bold=True,   size=9)
FONT_PLACE    = Font(             size=9)
FONT_TITLE    = Font(bold=True,   size=16)
FONT_SECTION  = Font(bold=True,   size=12)
FONT_SUBHEAD  = Font(bold=True,   size=10)
FONT_NORMAL   = Font(             size=10)
FONT_SMALL    = Font(             size=9,  color="555555")
FONT_HDR      = Font(bold=True,   size=10, color="FFFFFF")
FONT_LINK     = Font(             size=10, color="0563C1", underline="single")
FONT_ITALIC   = Font(italic=True, size=10)
FONT_ROW_LBL  = Font(             size=8,  color="888888")

ALIGN_WRAP    = Alignment(wrap_text=True, vertical="top")
ALIGN_TOP     = Alignment(vertical="top")
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT   = Alignment(horizontal="right",  vertical="top")

MEDALS = {1: "🥇", 2: "🥈", 3: "🥉"}

COL_W_MIN     = 24   # minimum column width for event columns
COL_W_LABEL   = 12   # column A (row-label) width in year sheets

# Honour symbols
SYM_BAP   = "★"
SYM_FBHOF = "☆"

# Known name-variant pairs between BAP and FBHOF files
_BAP_TO_FBHOF_ALIASES = {
    "Ken Shults":      "Kenny Shults",
    "Vasek Klouda":    "Václav Klouda",
    "Sebastien Lefay": "Sébastien Lefay",
}
_FBHOF_TO_BAP_ALIASES = {v: k for k, v in _BAP_TO_FBHOF_ALIASES.items()}

# Honours display names → person_canon in Persons_Truth
# (for cases where normalisation alone doesn't find the match)
_HONOURS_TO_PT = {
    # BAP file variants (name as it appears in BAP file → person_canon in Persons_Truth)
    "Sebastien Duschesne":      "Sebastien Duchesne",
    "Arek Dzudzinski":          "Arek Dudzinski",
    "Rene Ruhr":                "Rene Ruehr",
    "Nick Polini":              "Nick Pollini",
    "Ken Shults":               "Kenneth Shults",
    "Kenny Shults":             "Kenneth Shults",
    "Dave Holton":              "David Holton",
    "Gordon Scott Bevier":      "Scott Bevier",
    "Bryan Fournier":           "Brian Fournier",
    "Johnny Murphy":            "Jonathan Murphy",
    "Phillip Morrison":         "Philip Morrison",
    "Tina Aberli":              "Tina Aeberli",
    "Vasek Klouda":             "Vaclav Klouda",
    "Vaclav (Vasek) Klouda":    "Vaclav Klouda",
    "Václav Klouda":            "Vaclav Klouda",
    "Olav Piwowar":             "Olaf Piwowar",
    "Jindra Smola":             "Jindrich Smola",
    "Rafał Kaleta":             "Rafal Kaleta",
    "Jani Markkanen":           "Jani Sakari Markkanen",
    "Jakob Wagner Revstein":    "Jakob Wagner",
    # FBHOF file variants
    "Becca English":            "Becca English Ross",
    "Chris Siebert":            "Christopher Michael Siebert",
    "David Leberknight":        "Dave Leberknight",
    "Jim Caveney":              "Jimmy Caveney",
    "Lon Skyler Smith":         "Skyler Lon Smith",
    "Maude Landreville":        "Maude Laudreville",
    "PT Lovern":                "Paul Lovern",
    "Sam Conlon":               "Samantha Conlon",
    "Vasek Klouda":             "Vaclav Klouda",
}


# ── Name helpers ──────────────────────────────────────────────────────────────

_TRANSLIT = str.maketrans("łŁøØðÐđĐ", "lloodddd")

def _display_name(s: str) -> str:
    """Return s with title-casing applied if stored in ALL-CAPS (Latin American event data)."""
    s = (s or "").strip()
    if not s:
        return s
    alpha = [c for c in s if c.isalpha()]
    if alpha and all(c.isupper() for c in alpha):
        return s.title()
    return s


def _norm_name(s: str) -> str:
    """Lowercase + strip diacritics for fuzzy matching.
    Handles Polish ł, Norwegian ø, Icelandic ð, etc. that don't NFD-decompose."""
    s = s.translate(_TRANSLIT)
    nfd = unicodedata.normalize("NFD", s.lower().strip())
    return "".join(c for c in nfd if unicodedata.category(c) != "Mn")


def _to_int(v) -> int:
    try:
        return int(float(v or 0))
    except (ValueError, TypeError):
        return 0


def _split_location(loc: str):
    """'City, State, Country' → (city_region, country)"""
    parts = [p.strip() for p in loc.split(",")]
    if len(parts) >= 2:
        return ", ".join(parts[:-1]), parts[-1]
    return loc, ""


# ── Data loading ──────────────────────────────────────────────────────────────

def load_index() -> dict:
    """Load index.csv → dict event_id → {name, date, location, host_club}."""
    path = OUT_DIR / "index.csv"
    meta = {}
    try:
        df = pd.read_csv(path, dtype=str, encoding="latin-1").fillna("")
        for _, row in df.iterrows():
            eid = str(row.get("event_id", "")).strip()
            if not eid:
                continue
            loc = row.get("Location", "").strip()
            city, country = _split_location(loc)
            meta[eid] = {
                "event_name": row.get("Tournament Name", "").strip(),
                "date":       row.get("Date", "").strip(),
                "location":   loc,
                "city":       city,
                "country":    country,
                "host_club":  row.get("Host Club", "").strip(),
                "year":       _to_int(row.get("year", "")),
                "event_type": row.get("Event Type", "").strip(),
            }
    except Exception as exc:
        print(f"  WARN: could not load index.csv: {exc}", file=sys.stderr)
    return meta


def load_stage2_events() -> dict:
    """
    Load stage2_canonical_events.csv.
    Returns dict event_id → {year, event_name, date, location, host_club,
                              div_order: [division_canon ...]}
    """
    path = OUT_DIR / "stage2_canonical_events.csv"
    events = {}
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.DictReader(fh):
            eid = row["event_id"].strip()
            try:
                placements = json.loads(row.get("placements_json") or "[]")
            except (json.JSONDecodeError, TypeError):
                placements = []

            seen, div_order = set(), []
            for p in placements:
                dc = (p.get("division_canon") or "").strip()
                if dc and dc not in seen:
                    div_order.append(dc)
                    seen.add(dc)

            loc = (row.get("location") or "").strip()
            city, country = _split_location(loc)
            events[eid] = {
                "event_id":   eid,
                "year":       _to_int(row.get("year")),
                "event_name": (row.get("event_name") or "").strip(),
                "date":       (row.get("date") or "").strip(),
                "location":   loc,
                "city":       city,
                "country":    country,
                "host_club":  (row.get("host_club") or "").strip(),
                "div_order":  div_order,
            }
    return events


def load_placements_flat() -> pd.DataFrame:
    return pd.read_csv(
        OUT_DIR / "Placements_Flat.csv", dtype=str, encoding="utf-8",
    ).fillna("")


def load_placements_by_person() -> pd.DataFrame:
    return pd.read_csv(
        OUT_DIR / "Placements_ByPerson.csv", dtype=str, encoding="utf-8",
    ).fillna("")


def load_persons_truth() -> pd.DataFrame:
    df = pd.read_csv(
        OUT_DIR / "Persons_Truth.csv", dtype=str, encoding="utf-8",
    ).fillna("")
    internal = {"effective_person_id", "player_ids_seen", "player_names_seen",
                "alias_statuses", "norm_key", "last_token",
                "person_canon_clean", "person_canon_clean_reason"}
    return df.drop(columns=[c for c in df.columns if c in internal])


def load_honours(pt_df: pd.DataFrame) -> dict:
    """
    Load BAP and FBHOF CSVs, unify name variants, match against person_canon.

    Returns dict:
        person_canon → {
            'bap':        bool,
            'fbhof':      bool,
            'nickname':   str,
            'bap_year':   int,
            'fbhof_year': int,
            'symbol':     str,   # e.g. '★', '☆', '★☆'
        }

    Also returns (bap_rows, fbhof_rows) as raw lists for the Honours sheet.
    """
    # Build normalised-name → person_canon reverse lookup
    canon_list = pt_df["person_canon"].dropna().tolist() if not pt_df.empty else []
    norm_to_canon: dict = {}
    for pc in canon_list:
        norm_to_canon[_norm_name(pc)] = pc

    def _match_canon(name: str) -> str:
        """Return matched person_canon or '' if not found."""
        import re as _re
        # Hard-coded alias table
        if name in _HONOURS_TO_PT:
            return _HONOURS_TO_PT[name]
        # Exact
        if name in norm_to_canon.values():
            return name
        # Normalised
        n = _norm_name(name)
        if n in norm_to_canon:
            return norm_to_canon[n]
        # Strip parenthetical variants: "Vaclav (Vasek) Klouda" → "Vaclav Klouda"
        stripped = _re.sub(r"\s*\([^)]*\)", "", name).strip()
        if stripped != name:
            if stripped in _HONOURS_TO_PT:
                return _HONOURS_TO_PT[stripped]
            n2 = _norm_name(stripped)
            if n2 in norm_to_canon:
                return norm_to_canon[n2]
        return ""

    honours: dict = {}   # person_canon → info dict

    def _ensure(canon: str) -> dict:
        if canon not in honours:
            honours[canon] = {
                "bap": False, "fbhof": False,
                "nickname": "", "bap_year": 0, "fbhof_year": 0, "symbol": "",
            }
        return honours[canon]

    bap_rows  = []
    fbhof_rows = []

    # ── Load BAP ──────────────────────────────────────────────────────────────
    try:
        bap_df = pd.read_csv(INPUT_DIR / "bap_data_updated.csv", dtype=str).fillna("")
        for _, row in bap_df.iterrows():
            name = row["name"].strip()
            if not name:
                continue
            nick  = row.get("nickname", "").strip()
            year  = _to_int(row.get("year_inducted", ""))
            canon = _match_canon(name)
            bap_rows.append({"name": name, "nickname": nick, "year": year, "canon": canon})
            if canon:
                h = _ensure(canon)
                h["bap"]      = True
                h["nickname"] = h["nickname"] or nick
                h["bap_year"] = year
    except Exception as exc:
        print(f"  WARN: could not load bap_data.csv: {exc}", file=sys.stderr)

    # ── Load FBHOF ────────────────────────────────────────────────────────────
    try:
        fbhof_df = pd.read_csv(INPUT_DIR / "fbhof_data_updated.csv", dtype=str).fillna("")
        for _, row in fbhof_df.iterrows():
            name = row["name"].strip()
            if not name:
                continue
            raw_year = row.get("year_inducted", "")
            year = 0 if str(raw_year).strip().lower() == "unknown" else _to_int(raw_year)
            # Map FBHOF name to BAP canonical name if known
            bap_name = _FBHOF_TO_BAP_ALIASES.get(name, name)
            canon    = _match_canon(bap_name) or _match_canon(name)
            fbhof_rows.append({"name": name, "year": year, "canon": canon})
            if canon:
                h = _ensure(canon)
                h["fbhof"]      = True
                h["fbhof_year"] = year
    except Exception as exc:
        print(f"  WARN: could not load fbhof_data.csv: {exc}", file=sys.stderr)

    # Compute symbol strings
    for h in honours.values():
        h["symbol"] = (SYM_BAP if h["bap"] else "") + (SYM_FBHOF if h["fbhof"] else "")

    matched   = sum(1 for r in bap_rows if r["canon"])
    unmatched = [r["name"] for r in bap_rows if not r["canon"]]
    print(f"  Honours: {matched}/{len(bap_rows)} BAP names matched to person_canon")
    if unmatched:
        print(f"  Honours unmatched: {unmatched}", file=sys.stderr)

    return honours, bap_rows, fbhof_rows


# ── Placement data for year sheets ────────────────────────────────────────────

def build_event_placements(pf: pd.DataFrame, events: dict) -> dict:
    """
    Returns dict: event_id → OrderedDict{division_canon: [(place_int, display, cat)]}.
    Divisions are in source order (from events[eid]['div_order']).
    Doubles are deduplicated by team_person_key; team_display_name used.
    __NON_PERSON__ and unresolved rows are excluded.
    """
    result = {}

    for eid, edf in pf.groupby("event_id"):
        if eid not in events:
            continue

        div_order     = events[eid]["div_order"]
        div_placements: dict = {}

        for div_canon, ddf in edf.groupby("division_canon"):
            ddf = ddf.copy()
            ddf["_place"] = pd.to_numeric(ddf["place"], errors="coerce")
            ddf = ddf.sort_values(["_place", "team_person_key", "person_canon"],
                                  na_position="last")

            entries    = []
            seen_teams: set = set()

            for _, row in ddf.iterrows():
                person = (row.get("person_canon") or "").strip()
                if not person or person == "__NON_PERSON__":
                    continue
                if (row.get("person_unresolved") or "").lower() == "true":
                    continue

                try:
                    place_int = int(float(row["place"]))
                except (ValueError, TypeError):
                    continue

                comp = (row.get("competitor_type") or "player").lower()
                tpk  = (row.get("team_person_key") or "").strip()
                cat  = (row.get("division_category") or "").strip()

                if comp == "team" and tpk:
                    if tpk in seen_teams:
                        continue
                    seen_teams.add(tpk)
                    display = (row.get("team_display_name") or "").strip()
                    if not display:
                        members = ddf[ddf["team_person_key"] == tpk]["person_canon"].tolist()
                        display = " / ".join(_display_name(m) for m in members if m)
                else:
                    display = _display_name(person)
                    # Solo entry in a doubles division — partner not recorded
                    if "double" in div_canon.lower():
                        display = f"{display} / ?"

                entries.append((place_int, display, cat))

            if entries:
                div_placements[div_canon] = entries

        # Reorder: source order first, then any unseen divisions
        ordered: OrderedDict = OrderedDict()
        for dc in div_order:
            if dc in div_placements:
                ordered[dc] = div_placements[dc]
        for dc, dp in div_placements.items():
            if dc not in ordered:
                ordered[dc] = dp

        result[eid] = ordered

    return result


# ── Leaderboard computation ───────────────────────────────────────────────────

def compute_leaderboards(pbp: pd.DataFrame) -> pd.DataFrame:
    """Compute wins / podiums / placements / events / career_span per person."""
    df = pbp.copy()
    df = df[df["person_unresolved"].str.lower() != "true"]
    df = df[df["person_canon"].str.strip() != ""]
    df = df[df["person_canon"].str.strip() != "__NON_PERSON__"]
    df["_place"] = pd.to_numeric(df["place"], errors="coerce")
    df["_year"]  = pd.to_numeric(df["year"],  errors="coerce")

    wins     = df[df["_place"] == 1].groupby("person_canon").size().rename("wins")
    podiums  = df[df["_place"] <= 3].groupby("person_canon").size().rename("podiums")
    total    = df.groupby("person_canon").size().rename("placements")
    events   = df.groupby("person_canon")["event_id"].nunique().rename("events")
    first_yr = df.groupby("person_canon")["_year"].min().rename("first_year")
    last_yr  = df.groupby("person_canon")["_year"].max().rename("last_year")

    stats = pd.concat([wins, podiums, total, events, first_yr, last_yr], axis=1).fillna(0)
    stats["wins"]        = stats["wins"].astype(int)
    stats["podiums"]     = stats["podiums"].astype(int)
    stats["placements"]  = stats["placements"].astype(int)
    stats["events"]      = stats["events"].astype(int)
    stats["first_year"]  = stats["first_year"].astype(int)
    stats["last_year"]   = stats["last_year"].astype(int)
    stats["career_span"] = stats["last_year"] - stats["first_year"]
    return stats.reset_index()


def compute_leaderboards_by_cat(pbp: pd.DataFrame) -> dict:
    """Wins per person per division_category."""
    df = pbp.copy()
    df = df[df["person_unresolved"].str.lower() != "true"]
    df = df[~df["person_canon"].str.strip().isin(["", "__NON_PERSON__"])]
    df["_place"] = pd.to_numeric(df["place"], errors="coerce")
    wins = df[df["_place"] == 1]
    by_cat = {}
    for cat, cdf in wins.groupby("division_category"):
        by_cat[cat] = (
            cdf.groupby("person_canon").size()
               .rename("wins")
               .sort_values(ascending=False)
               .reset_index()
        )
    return by_cat


# ── Cell helper ───────────────────────────────────────────────────────────────

def _c(ws, row: int, col: int, value=None, *,
       font=None, fill=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    return cell


def _get_symbol(display: str, honours: dict) -> str:
    """Return honour symbol string for a display name (person or team)."""
    if display in honours:
        return honours[display]["symbol"]
    if "/" in display:
        symbols = set()
        for m in display.split("/"):
            m = m.strip()
            if m in honours:
                if honours[m]["bap"]:   symbols.add(SYM_BAP)
                if honours[m]["fbhof"]: symbols.add(SYM_FBHOF)
        if symbols:
            return (SYM_BAP if SYM_BAP in symbols else "") + \
                   (SYM_FBHOF if SYM_FBHOF in symbols else "")
    return ""


# ── ReadMe sheet ──────────────────────────────────────────────────────────────

def build_readme(wb: Workbook, events: dict, pf: pd.DataFrame,
                 honours: dict):
    ws = wb.create_sheet("ReadMe")
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22

    n_events     = len(events)
    n_placements = len(pf[~pf["person_canon"].isin(["", "__NON_PERSON__"])])
    years        = sorted({ev["year"] for ev in events.values() if ev["year"]})
    yr_range     = f"{years[0]}–{years[-1]}" if years else "?"

    _c(ws, 1, 1, "Footbag Historical Results Archive", font=FONT_TITLE)

    sections = [
        (3,  "Coverage",   yr_range),
        (5,  "Events",     f"{n_events:,}"),
        (7,  "Placements", f"{n_placements:,}"),
    ]
    for row, label, value in sections:
        _c(ws, row, 1, label, font=FONT_SUBHEAD)
        _c(ws, row, 2, value, font=FONT_NORMAL)

    _c(ws, 9,  1, "Note", font=FONT_SUBHEAD)
    _c(ws, 10, 1, "Results data is incomplete for early years (pre-1997).", font=FONT_NORMAL)
    _c(ws, 11, 1, "Coverage improves significantly from 1997 onward.", font=FONT_SMALL)

    _c(ws, 13, 1, "Sources", font=FONT_SUBHEAD)
    _c(ws, 14, 1, "footbag.org archive",          font=FONT_NORMAL)
    _c(ws, 15, 1, "historical tournament records", font=FONT_NORMAL)

    _c(ws, 17, 1, "Compiled by",        font=FONT_SUBHEAD)
    _c(ws, 18, 1, "James Leberknight",  font=FONT_NORMAL)
    _c(ws, 19, 1, str(datetime.now().year), font=FONT_SMALL)

    _c(ws, 21, 1, "Honours", font=FONT_SUBHEAD)
    _c(ws, 22, 1, "BAP and Hall of Fame inductees are listed on the Honours sheet.",
       font=FONT_NORMAL)


# ── Honours sheet ─────────────────────────────────────────────────────────────

def build_honours_sheet(wb: Workbook, honours: dict,
                        bap_rows: list, fbhof_rows: list,
                        stats: pd.DataFrame):
    ws = wb.create_sheet("Honours")
    ws.freeze_panes = "A3"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 8

    # Title
    ws.merge_cells("A1:H1")
    _c(ws, 1, 1, "Footbag Honours — Big Add Posse & Hall of Fame",
       font=FONT_TITLE, align=ALIGN_CENTER)

    # Sub-header
    ws.merge_cells("A2:H2")
    _c(ws, 2, 1,
       f"{SYM_BAP} Big Add Posse  |  {SYM_FBHOF} Footbag Hall of Fame",
       font=FONT_SMALL, align=ALIGN_CENTER)

    # Column headers
    hdrs = ["Year", "Name", SYM_BAP, SYM_FBHOF, "Nickname", "Wins", "Podiums", "Events"]
    for c, h in enumerate(hdrs, start=1):
        _c(ws, 3, c, h, font=FONT_HDR, fill=FILL_HDR, align=ALIGN_CENTER)

    # Build unified inductee list, keyed by BAP canonical name
    # (FBHOF-only entries use FBHOF name if no BAP match)
    bap_set  = {r["name"] for r in bap_rows}
    fbhof_set = {r["name"] for r in fbhof_rows}

    inductees = []  # list of dicts: {year, name, bap, fbhof, nickname, canon}

    # Collect all BAP entries first
    for r in bap_rows:
        fbhof_year = 0
        # Check if this person is also in FBHOF (via alias or exact)
        fbhof_name = _BAP_TO_FBHOF_ALIASES.get(r["name"], r["name"])
        for fr in fbhof_rows:
            if fr["name"] == fbhof_name or fr["name"] == r["name"]:
                fbhof_year = fr["year"]
                break
        inductees.append({
            "year":       r["year"],
            "name":       r["name"],
            "bap":        True,
            "fbhof":      fbhof_year > 0,
            "fbhof_year": fbhof_year,
            "nickname":   r["nickname"],
            "canon":      r["canon"],
        })

    # FBHOF-only entries (not in BAP)
    bap_canonical_names = {r["name"] for r in bap_rows}
    bap_aliases         = set(_BAP_TO_FBHOF_ALIASES.values())
    for fr in fbhof_rows:
        if fr["name"] not in bap_canonical_names and fr["name"] not in bap_aliases:
            inductees.append({
                "year":       fr["year"],
                "name":       fr["name"],
                "bap":        False,
                "fbhof":      True,
                "fbhof_year": fr["year"],
                "nickname":   "",
                "canon":      fr["canon"],
            })

    # Sort by induction year
    inductees.sort(key=lambda x: x["year"])

    # Build stats lookup  canon → plain dict (avoid Series truth-value issues)
    stats_map: dict = {}
    if not stats.empty:
        for _, row in stats.iterrows():
            stats_map[row["person_canon"]] = {
                "wins":    int(row.get("wins", 0)    or 0),
                "podiums": int(row.get("podiums", 0) or 0),
                "events":  int(row.get("events", 0)  or 0),
            }

    row_idx = 4
    for ind in inductees:
        canon  = ind["canon"]
        s      = stats_map.get(canon)
        wins   = s["wins"]    if s else 0
        pods   = s["podiums"] if s else 0
        evts   = s["events"]  if s else 0
        symbol = (SYM_BAP if ind["bap"] else "") + (SYM_FBHOF if ind["fbhof"] else "")

        if ind["bap"] and ind["fbhof"]:
            fill = FILL_HON_BOTH
        elif ind["bap"]:
            fill = FILL_HON_BAP
        else:
            fill = FILL_HON_FBHOF

        ws.cell(row=row_idx, column=1, value=ind["year"] or None)
        ws.cell(row=row_idx, column=2, value=_display_name(ind["name"]))
        ws.cell(row=row_idx, column=3, value=SYM_BAP   if ind["bap"]   else "")
        ws.cell(row=row_idx, column=4, value=SYM_FBHOF if ind["fbhof"] else "")
        ws.cell(row=row_idx, column=5, value=ind["nickname"] or "")
        ws.cell(row=row_idx, column=6, value=wins  or None)
        ws.cell(row=row_idx, column=7, value=pods  or None)
        ws.cell(row=row_idx, column=8, value=evts  or None)

        for c in range(1, 9):
            ws.cell(row=row_idx, column=c).fill = fill
            ws.cell(row=row_idx, column=c).alignment = ALIGN_CENTER \
                if c in (1, 3, 4, 6, 7, 8) else ALIGN_TOP

        row_idx += 1

    # Legend below table
    row_idx += 1
    _c(ws, row_idx, 1, "Legend", font=FONT_SUBHEAD)
    _c(ws, row_idx + 1, 1, f"{SYM_BAP} = Big Add Posse inductee",          font=FONT_SMALL)
    _c(ws, row_idx + 2, 1, f"{SYM_FBHOF} = Footbag Hall of Fame inductee", font=FONT_SMALL)
    _c(ws, row_idx + 3, 1, "BAP data includes nickname column where available.", font=FONT_SMALL)
    fill_legend = [FILL_HON_BAP, FILL_HON_FBHOF, FILL_HON_BOTH]
    legend_labels = ["BAP only", "FBHOF only", "Both"]
    for i, (f, lbl) in enumerate(zip(fill_legend, legend_labels)):
        _c(ws, row_idx + 5 + i, 1, lbl, font=FONT_SMALL, fill=f)


# ── Summary sheet ─────────────────────────────────────────────────────────────

def build_summary(wb: Workbook, events: dict, event_placements: dict,
                  stats: pd.DataFrame, pbp: pd.DataFrame,
                  honours: dict):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width =  4
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10

    ws.merge_cells("A1:F1")
    _c(ws, 1, 1, "Footbag Historical Results Archive",
       font=FONT_TITLE, align=ALIGN_CENTER)

    # ── Dataset overview ──────────────────────────────────────────────────────
    _c(ws, 3, 1, "Dataset Overview", font=FONT_SECTION)

    n_events     = len(events)
    n_placements = sum(len(v) for ep in event_placements.values() for v in ep.values())
    n_players    = len(stats) if not stats.empty else "?"
    years        = sorted({ev["year"] for ev in events.values() if ev["year"]})
    yr_range     = f"{years[0]}–{years[-1]}" if years else "?"

    for r, (label, value) in enumerate([
        ("Events",         f"{n_events:,}"),
        ("Years covered",  yr_range),
        ("Placements",     f"{n_placements:,}"),
        ("Unique players", f"{n_players:,}" if isinstance(n_players, int) else n_players),
    ], start=4):
        _c(ws, r, 1, label, font=FONT_SUBHEAD)
        _c(ws, r, 2, value, font=FONT_NORMAL)

    # ── Honours section ───────────────────────────────────────────────────────
    _c(ws, 9, 1, "Honours", font=FONT_SECTION)

    n_bap   = sum(1 for h in honours.values() if h["bap"])
    n_fbhof = sum(1 for h in honours.values() if h["fbhof"])
    n_both  = sum(1 for h in honours.values() if h["bap"] and h["fbhof"])

    for r, (label, value) in enumerate([
        ("Big Add Posse",  f"{n_bap} members in dataset"),
        ("Hall of Fame",   f"{n_fbhof} members in dataset"),
        ("Both Honours",   f"{n_both} players"),
    ], start=10):
        _c(ws, r, 1, label, font=FONT_SUBHEAD)
        _c(ws, r, 2, value, font=FONT_NORMAL)

    # ── Leaderboards ─────────────────────────────────────────────────────────
    lb_row = 15
    _c(ws, lb_row, 1, "Leaderboards", font=FONT_SECTION)
    lb_row += 1

    def _write_lb(ws, start_row, col, title, df_in, val_col, val_label, n=10):
        _c(ws, start_row, col,   title,     font=FONT_SUBHEAD)
        _c(ws, start_row, col+1, val_label, font=FONT_SUBHEAD)
        r = start_row + 1
        try:
            top = (df_in[["person_canon", val_col]].copy()
                   .assign(**{val_col: pd.to_numeric(df_in[val_col], errors="coerce")})
                   .dropna(subset=[val_col])
                   .nlargest(n, val_col))
            for _, row in top.iterrows():
                ws.cell(row=r, column=col,   value=_display_name(row["person_canon"]))
                ws.cell(row=r, column=col+1, value=int(row[val_col]))
                r += 1
        except Exception:
            ws.cell(row=r, column=col, value="(unavailable)")
        return r + 1

    if not stats.empty:
        lb_row = _write_lb(ws, lb_row, 1, "Most Wins",        stats, "wins",       "Wins")
        lb_row = _write_lb(ws, lb_row, 1, "Most Podiums",     stats, "podiums",    "Podiums")
        lb_row = _write_lb(ws, lb_row, 1, "Most Appearances", stats, "events",     "Events")
        _write_lb(ws, lb_row, 1,           "Longest Careers", stats, "career_span", "Years")


# ── Records sheet ─────────────────────────────────────────────────────────────

def build_records(wb: Workbook, stats: pd.DataFrame, cat_stats: dict,
                  events: dict, event_placements: dict,
                  honours: dict):
    ws = wb.create_sheet("Records")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width =  4
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width =  4

    _c(ws, 1, 1, "Records", font=FONT_TITLE)

    def _write_lb(ws, row, col, title, df_in, name_col, val_col, val_label, n=15):
        _c(ws, row, col,   title,     font=FONT_SUBHEAD, border=_border_top())
        _c(ws, row, col+1, val_label, font=FONT_SUBHEAD, border=_border_top())
        r = row + 1
        try:
            top = (df_in[[name_col, val_col]].copy()
                   .assign(**{val_col: pd.to_numeric(df_in[val_col], errors="coerce")})
                   .dropna(subset=[val_col])
                   .nlargest(n, val_col))
            for _, row_data in top.iterrows():
                ws.cell(row=r, column=col,   value=_display_name(str(row_data[name_col])))
                ws.cell(row=r, column=col+1, value=int(row_data[val_col]))
                r += 1
        except Exception:
            ws.cell(row=r, column=col, value="(unavailable)")
        return r + 1

    # ── Column A/B: all-time records ──────────────────────────────────────────
    r = 3
    if not stats.empty:
        r = _write_lb(ws, r, 1, "Most Wins — All Time",    stats, "person_canon", "wins",        "Wins")
        r = _write_lb(ws, r, 1, "Most Podium Finishes",    stats, "person_canon", "podiums",     "Podiums")
        r = _write_lb(ws, r, 1, "Most Events Competed",    stats, "person_canon", "events",      "Events")
        r = _write_lb(ws, r, 1, "Longest Careers (years)", stats, "person_canon", "career_span", "Span")

    # ── Column D/E: by division category ─────────────────────────────────────
    cat_labels = {"freestyle": "Freestyle Wins", "net": "Net Wins", "golf": "Golf Wins"}
    rc = 3
    for cat, label in cat_labels.items():
        if cat in cat_stats and not cat_stats[cat].empty:
            rc = _write_lb(ws, rc, 4, label, cat_stats[cat], "person_canon", "wins", "Wins")

    # Largest events
    event_sizes = []
    for eid, ep in event_placements.items():
        n = sum(len(v) for v in ep.values())
        ev = events.get(eid, {})
        event_sizes.append((ev.get("event_name", eid), ev.get("year", 0), n))
    top_events = sorted(event_sizes, key=lambda x: x[2], reverse=True)[:15]

    rc2 = rc + 1
    _c(ws, rc2, 4, "Largest Events", font=FONT_SUBHEAD, border=_border_top())
    _c(ws, rc2, 5, "Players",        font=FONT_SUBHEAD, border=_border_top())
    rc2 += 1
    for name, year, n in top_events:
        ws.cell(row=rc2, column=4, value=f"{name} ({year})")
        ws.cell(row=rc2, column=5, value=n)
        rc2 += 1



# ── Index sheet ───────────────────────────────────────────────────────────────

def build_index_real(wb: Workbook, events: dict, event_placements: dict,
                     event_col_map: dict, insert_at: int):
    """Build the Index sheet with hyperlinks and insert at the correct position."""
    ws = wb.create_sheet("Index")
    wb.move_sheet("Index", offset=-(len(wb.sheetnames) - 1 - insert_at))

    ws.freeze_panes = "A2"

    hdrs   = ["Year", "Event", "City / Region", "Country", "Host Club",
              "Divisions", "Players"]
    widths = [7, 48, 32, 12, 30, 11, 10]
    for c, (h, w) in enumerate(zip(hdrs, widths), start=1):
        _c(ws, 1, c, h, font=FONT_HDR, fill=FILL_HDR, align=ALIGN_CENTER)
        ws.column_dimensions[get_column_letter(c)].width = w

    all_eids = sorted(
        events.keys(),
        key=lambda eid: (events[eid]["year"], events[eid].get("date", eid)),
    )

    for row_idx, eid in enumerate(all_eids, start=2):
        ev  = events[eid]
        ep  = event_placements.get(eid, {})
        n_p = sum(len(v) for v in ep.values())
        n_d = len(ep)

        ws.cell(row=row_idx, column=1, value=ev["year"] or "?")
        ws.cell(row=row_idx, column=3, value=ev.get("city", ev.get("location", "")))
        ws.cell(row=row_idx, column=4, value=ev.get("country", ""))
        ws.cell(row=row_idx, column=5, value=ev.get("host_club", ""))
        ws.cell(row=row_idx, column=6, value=n_d)
        ws.cell(row=row_idx, column=7, value=n_p)

        cell = ws.cell(row=row_idx, column=2, value=ev["event_name"])
        if eid in event_col_map:
            sheet_name, col_letter = event_col_map[eid]
            safe = sheet_name.replace("'", "''")
            cell.hyperlink = f"#'{safe}'!{col_letter}1"
            cell.font = FONT_LINK
        else:
            cell.font = FONT_NORMAL

        if row_idx % 2 == 0:
            for c in range(1, 8):
                cell_obj = ws.cell(row=row_idx, column=c)
                if cell_obj.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell_obj.fill = _fill("F7F9FC")


# ── Player Stats sheet ────────────────────────────────────────────────────────

def build_player_stats(wb: Workbook, stats: pd.DataFrame, honours: dict,
                       persons_df: pd.DataFrame | None = None):
    """
    One row per resolved player showing career statistics.
    Filterable so any player can look themselves up.
    """
    ws = wb.create_sheet("Player Stats")
    ws.freeze_panes = "A2"

    hdrs   = ["Player", "Wins", "Podiums", "Placements", "Events",
              "First Year", "Last Year", "Career (yrs)", "Legacy ID"]
    widths = [32, 8, 8, 12, 8, 12, 12, 12, 10]

    for c, (h, w) in enumerate(zip(hdrs, widths), start=1):
        _c(ws, 1, c, h, font=FONT_HDR, fill=FILL_HDR)
        ws.column_dimensions[get_column_letter(c)].width = w

    if stats.empty:
        return

    # Build legacyid lookup from Persons_Truth
    legacyid_map: dict = {}
    if persons_df is not None and "legacyid" in persons_df.columns:
        for _, pr in persons_df.iterrows():
            lid = pr.get("legacyid", "")
            if lid:
                legacyid_map[pr["person_canon"]] = lid

    df = stats.sort_values("person_canon").reset_index(drop=True)

    for r_idx, row in df.iterrows():
        pc        = row["person_canon"]
        excel_row = r_idx + 2

        ws.cell(row=excel_row, column=1, value=_display_name(pc))
        ws.cell(row=excel_row, column=2, value=int(row["wins"]))
        ws.cell(row=excel_row, column=3, value=int(row["podiums"]))
        ws.cell(row=excel_row, column=4, value=int(row["placements"]))
        ws.cell(row=excel_row, column=5, value=int(row["events"]))
        ws.cell(row=excel_row, column=6, value=int(row["first_year"]) or None)
        ws.cell(row=excel_row, column=7, value=int(row["last_year"])  or None)
        ws.cell(row=excel_row, column=8, value=int(row["career_span"]) or None)
        lid = legacyid_map.get(pc, "")
        ws.cell(row=excel_row, column=9, value=int(lid) if lid else None)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{len(df) + 1}"


# ── Player Results sheet ──────────────────────────────────────────────────────

def build_player_results(wb: Workbook, pf: pd.DataFrame, events: dict):
    ws = wb.create_sheet("Player Results")
    ws.freeze_panes = "A2"

    hdrs   = ["Year", "Event", "Location", "Division", "Category",
              "Place", "Player", "Partner"]
    widths = [7, 48, 32, 28, 12, 7, 28, 28]

    for c, (h, w) in enumerate(zip(hdrs, widths), start=1):
        _c(ws, 1, c, h, font=FONT_HDR, fill=FILL_HDR)
        ws.column_dimensions[get_column_letter(c)].width = w

    df = pf.copy()
    df = df[df["person_unresolved"].str.lower() != "true"]
    df = df[~df["person_canon"].isin(["", "__NON_PERSON__"])]

    df["_place"] = pd.to_numeric(df["place"], errors="coerce")
    df["_year"]  = pd.to_numeric(df["year"],  errors="coerce")
    df = df.sort_values(["_year", "event_id", "division_canon", "_place",
                          "team_person_key", "person_canon"],
                        na_position="last")

    # Pre-build partner lookup for doubles
    seen_teams: dict = {}
    for _, row in df[df["competitor_type"] == "team"].iterrows():
        tpk = (row.get("team_person_key") or "").strip()
        if not tpk:
            continue
        grp = df[
            (df["event_id"]       == row["event_id"]) &
            (df["division_canon"] == row["division_canon"]) &
            (df["place"]          == row["place"]) &
            (df["team_person_key"]== tpk) &
            (df["person_canon"]   != row["person_canon"])
        ]["person_canon"].tolist()
        seen_teams[(row["event_id"], row["division_canon"], row["place"],
                    tpk, row["person_canon"])] = " / ".join(grp) if grp else ""

    row_idx = 2
    for _, row in df.iterrows():
        eid    = row["event_id"]
        ev     = events.get(eid, {})
        person = (row.get("person_canon") or "").strip()
        tpk    = (row.get("team_person_key") or "").strip()

        partner = ""
        if (row.get("competitor_type") or "").lower() == "team" and tpk:
            partner = seen_teams.get(
                (eid, row["division_canon"], row["place"], tpk, person), ""
            )
            if not partner:
                td = (row.get("team_display_name") or "").strip()
                if td and person in td:
                    partner = td.replace(person, "").strip(" /")
                elif td:
                    partner = td

        try:
            place_val = int(float(row["place"]))
        except (ValueError, TypeError):
            place_val = row["place"]

        ws.cell(row=row_idx, column=1, value=ev.get("year") or _to_int(row.get("year")))
        ws.cell(row=row_idx, column=2, value=ev.get("event_name") or eid)
        ws.cell(row=row_idx, column=3, value=ev.get("location", ""))
        ws.cell(row=row_idx, column=4, value=row.get("division_canon", ""))
        ws.cell(row=row_idx, column=5, value=row.get("division_category", ""))
        ws.cell(row=row_idx, column=6, value=place_val)
        ws.cell(row=row_idx, column=7, value=_display_name(person))
        ws.cell(row=row_idx, column=8, value=partner)
        row_idx += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{row_idx - 1}"


# ── Year sheets ───────────────────────────────────────────────────────────────

# Fixed row positions for the event header block
# (columns B onward; column A holds row labels)
_R_NAME    = 1   # Event name
_R_LOC     = 2   # Location
_R_HOST    = 3   # Host club  (italic per spec)
_R_DATE    = 4   # Date
_R_PLAYERS = 5   # Players count
_R_EVTYPE  = 6   # Event type
_R_BLANK   = 7   # Spacer
_R_DATA    = 8   # First division / placement row

_ROW_LABELS = {
    _R_NAME:    "Event",
    _R_LOC:     "Location",
    _R_HOST:    "Host Club",
    _R_DATE:    "Date",
    _R_PLAYERS: "Players",
    _R_EVTYPE:  "Event Type",
}


# Category display constants for grouped division headers in year sheets
CAT_ORDER  = ["net", "freestyle", "golf", "sideline", "unknown"]
CAT_LABELS = {"net": "NET", "freestyle": "FREESTYLE", "golf": "GOLF",
              "sideline": "SIDELINE", "unknown": "OTHER"}
FILL_CAT   = _fill("D0D0D0")
FONT_CAT   = Font(bold=True, size=8, color="444444")


def _write_event_col(ws, col: int, ev: dict, placements: OrderedDict,
                     honours: dict) -> tuple:
    """
    Write one event into column `col` (1-based, already offset for label col).
    Returns (last_row_written, max_content_length).
    Divisions are grouped by category (NET / FREESTYLE / GOLF / SIDELINE / OTHER).
    """
    n_players   = sum(len(v) for v in placements.values())
    max_content = max(len(ev.get("event_name", "")), 24)

    def _write(r, val, font, fill, align=ALIGN_TOP):
        nonlocal max_content
        if val:
            max_content = max(max_content, len(str(val)))
        _c(ws, r, col, val, font=font, fill=fill, align=align)

    _write(_R_NAME,    ev["event_name"],                FONT_BANNER,  FILL_BANNER, ALIGN_WRAP)
    _write(_R_LOC,     ev["location"] or "—",           FONT_META,    FILL_META)
    _write(_R_HOST,    ev["host_club"] or "",            FONT_HOST,    FILL_META)
    _write(_R_DATE,    ev["date"] or "",                 FONT_META,    FILL_META)
    _write(_R_PLAYERS, f"Players: {n_players}",         FONT_PLAYERS, FILL_PLAYERS)
    _write(_R_EVTYPE,  ev.get("event_type") or "",      FONT_META,    FILL_META)

    row = _R_DATA

    # Group divisions by category while preserving source order within each category
    cat_to_divs: dict = {}
    for div_name, entries in placements.items():
        if not entries:
            continue
        cat = (entries[0][2] or "unknown")
        if cat not in cat_to_divs:
            cat_to_divs[cat] = []
        cat_to_divs[cat].append((div_name, entries))

    for cat in CAT_ORDER:
        if cat not in cat_to_divs:
            continue
        # Category header row (only if this category has ≥1 division)
        cat_label = CAT_LABELS.get(cat, "OTHER")
        _c(ws, row, col, cat_label, font=FONT_CAT, fill=FILL_CAT, align=ALIGN_TOP)
        max_content = max(max_content, len(cat_label) + 2)
        row += 1

        for div_name, entries in cat_to_divs[cat]:
            # Division header: bold, light-grey, top border
            _c(ws, row, col, div_name,
               font=FONT_DIV, fill=FILL_DIV, border=_border_top(), align=ALIGN_TOP)
            max_content = max(max_content, len(div_name) + 2)
            row += 1

            for place_int, display, _ in entries:
                medal = MEDALS.get(place_int, "")

                # Build display text:  🥇 1  Name
                parts = []
                if medal:
                    parts.append(medal)
                parts.append(f"{place_int:>3} ")
                parts.append(display)
                text = " ".join(parts) if medal else "".join(parts)

                if place_int == 1:
                    fill, font = FILL_GOLD,   FONT_PODIUM
                elif place_int == 2:
                    fill, font = FILL_SILVER, FONT_PODIUM
                elif place_int == 3:
                    fill, font = FILL_BRONZE, FONT_PODIUM
                else:
                    fill, font = FILL_WHITE,  FONT_PLACE

                _c(ws, row, col, text, font=font, fill=fill, align=ALIGN_TOP)
                max_content = max(max_content, len(text) + 2)
                row += 1

            row += 1   # blank row between divisions

    return row - 1, max_content


def build_year_sheet(wb: Workbook, year: int, eids: list,
                     events: dict, event_placements: dict,
                     honours: dict) -> dict:
    """
    Build one year sheet with:
    - Column A: row labels (Event, Location, Host Club, Date, Players)
    - Columns B onward: one event per column
    - freeze_panes = "B1"  (column A stays visible when scrolling right)
    - Auto column width (min COL_W_MIN)

    Returns dict event_id → column_letter (B, C, D, …)
    """
    ws = wb.create_sheet(title=str(year))

    # ── Column A: row-label column ────────────────────────────────────────────
    ws.column_dimensions["A"].width = COL_W_LABEL
    for row_num, label in _ROW_LABELS.items():
        _c(ws, row_num, 1, label,
           font=FONT_ROW_LBL, fill=FILL_ROW_LABEL, align=ALIGN_RIGHT)

    sorted_eids = sorted(
        eids,
        key=lambda eid: (events[eid].get("date", ""), eid),
    )

    event_col_map: dict = {}
    col_max_widths: dict = {}

    for col_offset, eid in enumerate(sorted_eids, start=2):   # B=2, C=3, …
        ev         = events[eid]
        placements = event_placements.get(eid, OrderedDict())
        last_row, max_w = _write_event_col(ws, col_offset, ev, placements, honours)
        event_col_map[eid]       = get_column_letter(col_offset)
        col_max_widths[col_offset] = max_w

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[_R_NAME].height    = 36
    ws.row_dimensions[_R_LOC].height     = 15
    ws.row_dimensions[_R_HOST].height    = 15
    ws.row_dimensions[_R_DATE].height    = 15
    ws.row_dimensions[_R_PLAYERS].height = 15
    ws.row_dimensions[_R_EVTYPE].height  = 15

    # ── Auto-width per event column (min COL_W_MIN, cap at 60) ───────────────
    for col_idx, max_w in col_max_widths.items():
        ltr = get_column_letter(col_idx)
        ws.column_dimensions[ltr].width = max(min(max_w + 4, 60), COL_W_MIN)

    # ── Freeze: column A always visible when scrolling right ─────────────────
    ws.freeze_panes = "B1"

    return event_col_map


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading data…")
    index_meta = load_index()
    s2_events  = load_stage2_events()
    pf         = load_placements_flat()
    pbp        = load_placements_by_person()
    persons_df = load_persons_truth()

    # Merge index metadata (better display strings) into s2 events.
    # NOTE: location/city/country are NOT overwritten — stage2_canonical_events.csv
    # has canonical "City, Country" form (e.g. "Quebec, Canada"), while index.csv
    # has raw verbose strings (e.g. "La Ronde and Plaine des Jeux Montreal, Quebec, Canada").
    for eid, s2 in s2_events.items():
        if eid in index_meta:
            im = index_meta[eid]
            if im["event_name"]:  s2["event_name"]  = im["event_name"]
            if im["date"]:        s2["date"]         = im["date"]
            if im["host_club"]:   s2["host_club"]    = im["host_club"]
            if im.get("event_type"): s2["event_type"] = im["event_type"]

    print("Loading honours (BAP / FBHOF)…")
    honours, bap_rows, fbhof_rows = load_honours(persons_df)
    print(f"  {sum(1 for h in honours.values() if h['bap'])} BAP  "
          f"| {sum(1 for h in honours.values() if h['fbhof'])} FBHOF  "
          f"| {sum(1 for h in honours.values() if h['bap'] and h['fbhof'])} both")

    print("Building event placements…")
    event_placements = build_event_placements(pf, s2_events)

    print("Computing leaderboards…")
    stats     = compute_leaderboards(pbp)
    cat_stats = compute_leaderboards_by_cat(pbp)

    # Group events by year (only years with placements get a sheet)
    year_to_eids: dict = defaultdict(list)
    for eid in s2_events:
        yr = s2_events[eid]["year"]
        if yr and eid in event_placements and event_placements[eid]:
            year_to_eids[yr].append(eid)

    print("Creating workbook…")
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet order: ReadMe, Summary, Records, Honours, [Index placeholder],
    #              Players, Player Results, year sheets
    build_readme(wb, s2_events, pf, honours)
    build_summary(wb, s2_events, event_placements, stats, pbp, honours)
    build_records(wb, stats, cat_stats, s2_events, event_placements, honours)
    build_honours_sheet(wb, honours, bap_rows, fbhof_rows, stats)

    # Index placeholder — correct content added after year sheets are built
    idx_placeholder = wb.create_sheet("Index")

    build_player_stats(wb, stats, honours, persons_df=persons_df)
    build_player_results(wb, pf, s2_events)

    # ── Year sheets ───────────────────────────────────────────────────────────
    all_event_col_map: dict = {}   # event_id → (sheet_title, col_letter)

    sorted_years = sorted(year_to_eids.keys())
    print(f"Building {len(sorted_years)} year sheets…")
    for year in sorted_years:
        col_map = build_year_sheet(
            wb, year, year_to_eids[year], s2_events, event_placements, honours
        )
        for eid, col_letter in col_map.items():
            all_event_col_map[eid] = (str(year), col_letter)

    # ── Rebuild Index now that year positions are known ───────────────────────
    # Sheet order after year sheets: [..., Index placeholder at position 4]
    # We remove placeholder and insert real Index at position 4
    # (ReadMe=0, Summary=1, Records=2, Honours=3, Index=4, …)
    wb.remove(idx_placeholder)
    build_index_real(wb, s2_events, event_placements, all_event_col_map,
                     insert_at=4)

    # ── Save ──────────────────────────────────────────────────────────────────
    print(f"Saving {XLSX}…")
    wb.save(XLSX)
    n_placements = sum(len(v) for ep in event_placements.values() for v in ep.values())
    print(f"Done.  Events: {len(s2_events)}, Placements: {n_placements:,}, "
          f"Year sheets: {len(sorted_years)}")


if __name__ == "__main__":
    main()
