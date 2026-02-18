#!/usr/bin/env python3
"""
05_build_persons_review.py — Build Persons Truth Review Workbook

Produces an Excel workbook designed for human review of person data quality.
Sheets:
  1. Persons_All        — Every person from Persons_Truth_Full + Excluded, enriched
                           with placement counts, years active, division categories.
  2. Ambiguous_Aliases   — Aliases that map to 2+ person_ids (need human disambiguation).
  3. Canon_Conflicts     — Person_ids with 2+ different person_canon values.
  4. Duplicate_Display   — Quarantined entries (two-people concatenations).
  5. Rejected_Placements — Placements that couldn't be matched to a person_id.
  6. Coverage_Gaps       — Persons in placements but missing from Persons_Truth.
"""

import csv
import sys
import os
from collections import defaultdict, Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

csv.field_size_limit(10_000_000)

OUT_DIR = Path("out")
OVERRIDES_DIR = Path("overrides")
OUTPUT_FILE = OUT_DIR / "Persons_Truth_Review.xlsx"


# ── Styling ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
OK_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_header(ws):
    """Apply header styling to first row."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws, max_width=50):
    """Auto-fit column widths with a cap."""
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            val = str(cell.value or "")
            max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3


def write_sheet(ws, headers, rows, row_fill_fn=None):
    """Write headers + data rows to a worksheet."""
    ws.append(headers)
    for i, row in enumerate(rows):
        ws.append(row)
        if row_fill_fn:
            fill = row_fill_fn(i, row, headers)
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill
    style_header(ws)
    auto_width(ws)


# ── Data Loading ─────────────────────────────────────────────────────────────

def load_csv(path):
    """Load a CSV file into a list of dicts."""
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_placement_stats():
    """Build per-person_id stats from Placements_Flat."""
    stats = defaultdict(lambda: {
        "count": 0, "years": set(), "div_cats": set(),
        "events": set(), "names_seen": set(), "wins": 0, "podiums": 0,
    })
    with open(OUT_DIR / "Placements_Flat.csv", newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            year = row.get("year", "")
            div_cat = row.get("division_category", "")
            event_id = row.get("event_id", "")
            place_str = row.get("place", "")
            try:
                place = int(place_str)
            except (ValueError, TypeError):
                place = None

            for prefix in ("player1", "player2"):
                pid = row.get(f"{prefix}_person_id", "")
                if not pid:
                    continue
                s = stats[pid]
                s["count"] += 1
                if year:
                    s["years"].add(year)
                if div_cat:
                    s["div_cats"].add(div_cat)
                if event_id:
                    s["events"].add(event_id)
                name = row.get(f"{prefix}_person_canon", "")
                if name:
                    s["names_seen"].add(name)
                if place == 1:
                    s["wins"] += 1
                if place is not None and place <= 3:
                    s["podiums"] += 1
    return stats


def load_aliases():
    """Load person_aliases.csv."""
    return load_csv(OVERRIDES_DIR / "person_aliases.csv")


# ── Sheet Builders ───────────────────────────────────────────────────────────

def build_persons_all(wb, placement_stats):
    """Sheet 1: Every person, enriched with placement stats and quality flags."""
    ws = wb.create_sheet("Persons_All")

    # Load all persons from Full + Excluded
    full = load_csv(OUT_DIR / "Persons_Truth_Full.csv")
    excluded = load_csv(OUT_DIR / "Persons_Truth_Excluded.csv")

    # Tag source file
    for row in full:
        row["_file"] = "truth_full"
    for row in excluded:
        row["_file"] = "excluded"

    # Combine (Full already includes presentable + candidates;
    # Excluded has not_presentable + quarantine)
    # Deduplicate by effective_person_id (Excluded overlaps with Full for quarantine)
    seen_ids = set()
    all_persons = []
    for row in full:
        pid = row["effective_person_id"]
        if pid not in seen_ids:
            seen_ids.add(pid)
            all_persons.append(row)
    for row in excluded:
        pid = row["effective_person_id"]
        if pid not in seen_ids:
            seen_ids.add(pid)
            all_persons.append(row)

    headers = [
        "effective_person_id", "person_canon", "person_canon_clean",
        "status", "source",
        "placement_count", "event_count", "win_count", "podium_count",
        "year_first", "year_last", "years_active",
        "div_categories", "names_seen_count",
        "player_ids_seen", "player_names_seen",
        "aliases", "alias_statuses",
        "clean_reason", "quality_flags",
    ]

    rows = []
    for p in all_persons:
        pid = p["effective_person_id"]
        ps = placement_stats.get(pid, {})
        years = sorted(ps.get("years", set()))
        div_cats = sorted(ps.get("div_cats", set()))

        # Determine status
        exclude_reason = p.get("exclude_reason", "")
        quarantine_reason = p.get("quarantine_reason", "")
        if exclude_reason and exclude_reason != "two_people_quarantine":
            status = f"excluded:{exclude_reason}"
        elif quarantine_reason:
            status = f"quarantine:{quarantine_reason}"
        elif p["_file"] == "truth_full":
            status = "presentable"
        else:
            status = "excluded"

        # Quality flags
        flags = []
        canon = p.get("person_canon", "")
        canon_clean = p.get("person_canon_clean", "")
        if canon != canon_clean and canon_clean:
            flags.append("canon_cleaned")
        if canon and canon == canon.upper() and len(canon) > 2:
            flags.append("ALL_CAPS")
        if canon and canon[-1] in ".-,;:":
            flags.append("trailing_punct")
        if len(canon.split()) == 1 and len(canon) > 1:
            flags.append("single_word")
        if ps.get("count", 0) == 0:
            flags.append("no_placements")
        names_seen_count = len(ps.get("names_seen", set()))
        if names_seen_count > 3:
            flags.append(f"many_names({names_seen_count})")

        rows.append([
            pid,
            canon,
            canon_clean,
            status,
            p.get("source", ""),
            ps.get("count", 0),
            len(ps.get("events", set())),
            ps.get("wins", 0),
            ps.get("podiums", 0),
            years[0] if years else "",
            years[-1] if years else "",
            len(years),
            " | ".join(div_cats) if div_cats else "",
            names_seen_count,
            p.get("player_ids_seen", ""),
            p.get("player_names_seen", ""),
            p.get("aliases", ""),
            p.get("alias_statuses", ""),
            p.get("person_canon_clean_reason", ""),
            " | ".join(flags) if flags else "",
        ])

    # Sort: presentable first, then by placement count desc
    status_order = {"presentable": 0}
    rows.sort(key=lambda r: (status_order.get(r[3], 1), -r[5]))

    def row_fill(i, row, hdrs):
        status = row[3]
        if status == "presentable":
            return None
        if "quarantine" in status:
            return WARN_FILL
        if "excluded" in status:
            return ERROR_FILL
        return None

    write_sheet(ws, headers, rows, row_fill_fn=row_fill)
    print(f"  Persons_All: {len(rows)} rows")
    return len(rows)


def build_ambiguous_aliases(wb):
    """Sheet 2: Aliases that map to 2+ person_ids."""
    ws = wb.create_sheet("Ambiguous_Aliases")
    aliases = load_aliases()

    # Group by alias → set of (person_id, person_canon)
    alias_map = defaultdict(list)
    for row in aliases:
        if row["status"].lower() == "verified":
            alias_map[row["alias"]].append(row)

    # Load placement stats for context
    placement_stats = load_placement_stats()

    headers = [
        "alias", "person_id", "person_canon", "status",
        "placement_count", "event_count", "year_first", "year_last",
        "decision", "notes",
    ]

    rows = []
    ambiguous = {a: entries for a, entries in alias_map.items()
                 if len(set(e["person_id"] for e in entries)) > 1}

    for alias in sorted(ambiguous.keys()):
        entries = ambiguous[alias]
        unique_pids = sorted(set(e["person_id"] for e in entries))
        for pid in unique_pids:
            canons = set(e["person_canon"] for e in entries if e["person_id"] == pid)
            ps = placement_stats.get(pid, {})
            years = sorted(ps.get("years", set()))
            rows.append([
                alias,
                pid,
                " / ".join(sorted(canons)),
                "verified",
                ps.get("count", 0),
                len(ps.get("events", set())),
                years[0] if years else "",
                years[-1] if years else "",
                "",  # decision (for human)
                "",  # notes (for human)
            ])

    write_sheet(ws, headers, rows)
    print(f"  Ambiguous_Aliases: {len(rows)} rows ({len(ambiguous)} distinct aliases)")
    return len(rows)


def build_canon_conflicts(wb):
    """Sheet 3: Person_ids with 2+ different person_canon values."""
    ws = wb.create_sheet("Canon_Conflicts")
    aliases = load_aliases()
    placement_stats = load_placement_stats()

    # Group by person_id → set of person_canon
    pid_canons = defaultdict(lambda: {"canons": set(), "aliases": set()})
    for row in aliases:
        if row["status"].lower() == "verified":
            pid_canons[row["person_id"]]["canons"].add(row["person_canon"])
            pid_canons[row["person_id"]]["aliases"].add(row["alias"])

    conflicts = {pid: data for pid, data in pid_canons.items()
                 if len(data["canons"]) > 1}

    headers = [
        "person_id", "canon_variant_1", "canon_variant_2", "canon_variant_3",
        "all_aliases",
        "placement_count", "event_count", "year_first", "year_last",
        "chosen_canon", "notes",
    ]

    rows = []
    for pid in sorted(conflicts.keys()):
        data = conflicts[pid]
        canons = sorted(data["canons"])
        ps = placement_stats.get(pid, {})
        years = sorted(ps.get("years", set()))
        rows.append([
            pid,
            canons[0] if len(canons) > 0 else "",
            canons[1] if len(canons) > 1 else "",
            canons[2] if len(canons) > 2 else "",
            " | ".join(sorted(data["aliases"])),
            ps.get("count", 0),
            len(ps.get("events", set())),
            years[0] if years else "",
            years[-1] if years else "",
            "",  # chosen_canon (for human)
            "",  # notes (for human)
        ])

    rows.sort(key=lambda r: -r[5])  # Most placements first

    write_sheet(ws, headers, rows)
    print(f"  Canon_Conflicts: {len(rows)} rows")
    return len(rows)


def build_duplicate_display(wb):
    """Sheet 4: Quarantined entries (two-people concatenations)."""
    ws = wb.create_sheet("Duplicate_Display")
    quarantine = load_csv(OUT_DIR / "Persons_Truth_Quarantine_TwoPeople.csv")

    headers = [
        "person_canon", "effective_person_id",
        "quarantine_reason", "quarantine_evidence",
        "player_names_seen", "source",
        "action", "notes",
    ]

    rows = []
    for q in quarantine:
        rows.append([
            q.get("person_canon", ""),
            q.get("effective_person_id", ""),
            q.get("quarantine_reason", ""),
            q.get("quarantine_evidence", ""),
            q.get("player_names_seen", ""),
            q.get("source", ""),
            "",  # action (for human)
            "",  # notes (for human)
        ])

    write_sheet(ws, headers, rows)
    print(f"  Duplicate_Display: {len(rows)} rows")
    return len(rows)


def build_rejected_placements(wb):
    """Sheet 5: Placements that couldn't be matched to a person_id."""
    ws = wb.create_sheet("Rejected_Placements")
    rejected = load_csv(OUT_DIR / "Placements_ByPerson_Rejected.csv")

    headers = [
        "reject_reason", "event_id", "year", "division_canon",
        "division_category", "competitor_type", "place",
        "player1_name_raw", "player1_name_clean", "player1_name",
        "player2_name_raw", "player2_name",
        "player1_id", "player2_id",
    ]

    rows = []
    for r in rejected:
        rows.append([
            r.get("reject_reason", ""),
            r.get("event_id", ""),
            r.get("year", ""),
            r.get("division_canon", ""),
            r.get("division_category", ""),
            r.get("competitor_type", ""),
            r.get("place", ""),
            r.get("player1_name_raw", ""),
            r.get("player1_name_clean", ""),
            r.get("player1_name", ""),
            r.get("player2_name_raw", ""),
            r.get("player2_name", ""),
            r.get("player1_id", ""),
            r.get("player2_id", ""),
        ])

    # Sort by reject_reason, then year
    rows.sort(key=lambda r: (r[0], r[2], r[1]))

    write_sheet(ws, headers, rows)
    print(f"  Rejected_Placements: {len(rows)} rows")
    return len(rows)


def build_coverage_gaps(wb, placement_stats):
    """Sheet 6: Person_ids in placements but missing from Persons_Truth."""
    ws = wb.create_sheet("Coverage_Gaps")

    # Load all known person_ids from Persons_Truth_Full
    full = load_csv(OUT_DIR / "Persons_Truth_Full.csv")
    excluded = load_csv(OUT_DIR / "Persons_Truth_Excluded.csv")
    known_pids = set()
    for row in full:
        known_pids.add(row["effective_person_id"])
    for row in excluded:
        known_pids.add(row["effective_person_id"])

    # Find person_ids in placements but not in truth
    gaps = {}
    for pid, stats in placement_stats.items():
        if pid not in known_pids:
            gaps[pid] = stats

    headers = [
        "person_id", "canon_name_observed",
        "placement_count", "event_count",
        "year_first", "year_last",
        "div_categories", "all_names_seen",
        "action", "notes",
    ]

    rows = []
    for pid in sorted(gaps.keys(), key=lambda p: -gaps[p]["count"]):
        s = gaps[pid]
        years = sorted(s["years"])
        names = sorted(s["names_seen"])
        rows.append([
            pid,
            names[0] if names else "",
            s["count"],
            len(s["events"]),
            years[0] if years else "",
            years[-1] if years else "",
            " | ".join(sorted(s["div_cats"])),
            " | ".join(names),
            "",  # action (for human)
            "",  # notes (for human)
        ])

    write_sheet(ws, headers, rows)
    print(f"  Coverage_Gaps: {len(rows)} rows")
    return len(rows)


def build_summary(wb, sheet_stats):
    """Add a Summary sheet at the beginning."""
    ws = wb.create_sheet("Summary", 0)

    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)

    ws["A1"] = "Persons Truth Review Workbook"
    ws["A1"].font = title_font
    ws["A3"] = "Sheet"
    ws["B3"] = "Rows"
    ws["C3"] = "Purpose"
    ws["A3"].font = subtitle_font
    ws["B3"].font = subtitle_font
    ws["C3"].font = subtitle_font

    sheet_info = [
        ("Persons_All", sheet_stats.get("Persons_All", 0),
         "Every person (presentable + excluded + quarantined) with placement stats and quality flags. Pivot-table ready."),
        ("Ambiguous_Aliases", sheet_stats.get("Ambiguous_Aliases", 0),
         "Aliases that map to 2+ different person_ids. Fill 'decision' column to resolve."),
        ("Canon_Conflicts", sheet_stats.get("Canon_Conflicts", 0),
         "Person_ids with 2+ different person_canon spellings. Fill 'chosen_canon' to resolve."),
        ("Duplicate_Display", sheet_stats.get("Duplicate_Display", 0),
         "Quarantined entries detected as two-person concatenations (unsplit teams)."),
        ("Rejected_Placements", sheet_stats.get("Rejected_Placements", 0),
         "Placements dropped because player couldn't be matched to a person_id."),
        ("Coverage_Gaps", sheet_stats.get("Coverage_Gaps", 0),
         "Person_ids found in placements but missing from Persons_Truth tables."),
    ]

    for i, (name, count, purpose) in enumerate(sheet_info, 4):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = count
        ws[f"C{i}"] = purpose

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 80

    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("Building Persons Truth Review Workbook...")
    print(f"Loading placement stats...")
    placement_stats = load_placement_stats()
    print(f"  {len(placement_stats)} unique person_ids with placements")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheet_stats = {}

    print("Building sheets...")
    sheet_stats["Persons_All"] = build_persons_all(wb, placement_stats)
    sheet_stats["Ambiguous_Aliases"] = build_ambiguous_aliases(wb)
    sheet_stats["Canon_Conflicts"] = build_canon_conflicts(wb)
    sheet_stats["Duplicate_Display"] = build_duplicate_display(wb)
    sheet_stats["Rejected_Placements"] = build_rejected_placements(wb)
    sheet_stats["Coverage_Gaps"] = build_coverage_gaps(wb, placement_stats)

    build_summary(wb, sheet_stats)

    print(f"\nWriting: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("Done!")


if __name__ == "__main__":
    main()
