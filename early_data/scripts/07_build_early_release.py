#!/usr/bin/env python3
"""
07_build_early_release.py — Pre-1997 early-data release package.

Applies conservative identity resolution policy, builds canonical CSVs,
and produces a standalone Excel workbook for the pre-1997 recovery dataset.

DOES NOT modify any post-1997 output files.
"""

import csv
import hashlib
import re
import uuid
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
ED = REPO_ROOT / "early_data"

# Inputs
PLACEMENTS_FBW  = ED / "placements"    / "placements_flat.csv"
PLACEMENTS_OR   = ED / "old_results"   / "old_results_placements_flat.csv"
EVENTS_FBW      = ED / "event_blocks"  / "event_blocks.csv"
EVENTS_OR       = ED / "old_results"   / "old_results_event_blocks.csv"
CANONICAL_EV    = ED / "canonical"     / "canonical_events.csv"
EID_MAPPING     = ED / "canonical"     / "event_id_mapping.csv"
EV_COMPARISON   = ED / "canonical"     / "event_source_comparison.csv"
CANDIDATES      = ED / "identity"      / "person_match_candidates.csv"
PERSONS_TRUTH   = REPO_ROOT / "inputs" / "Persons_Truth.csv"

# Outputs — identity review
IDENTITY_DIR    = ED / "identity"
OUT_AUTOACC     = IDENTITY_DIR / "person_aliases_autoaccepted.csv"
OUT_REVIEW      = IDENTITY_DIR / "person_aliases_needs_review.csv"
OUT_NEWPLAYERS  = IDENTITY_DIR / "new_early_players.csv"
OUT_NOISE       = IDENTITY_DIR / "unresolved_noise.csv"

# Outputs — canonical CSVs
CANON_DIR       = ED / "canonical"
OUT_EVENTS      = CANON_DIR / "events_pre1997.csv"
OUT_DISCIPLINES = CANON_DIR / "event_disciplines_pre1997.csv"
OUT_RESULTS     = CANON_DIR / "event_results_pre1997.csv"
OUT_PARTICIPANTS= CANON_DIR / "event_result_participants_pre1997.csv"
OUT_PERSONS     = CANON_DIR / "persons_pre1997.csv"
OUT_PALIASES    = CANON_DIR / "person_aliases_pre1997.csv"

# Output — spreadsheet
OUT_DIR         = ED / "out"
OUT_XLSX        = OUT_DIR / "footbag_results_pre1997_recovery.xlsx"


# ---------------------------------------------------------------------------
# PART 1 — IDENTITY RESOLUTION POLICY TABLES
# ---------------------------------------------------------------------------
# All decisions below are EXPLICIT and documented. No heuristic auto-resolution.

# Safe auto-accepted aliases: obvious 1-char spelling variants or common
# nickname shortenings where no ambiguity is possible.
# Format: raw_name -> (matched_person_id_prefix, matched_person_canon, reason)
AUTO_ACCEPT: dict[str, tuple] = {
    "Billy Hayne":    ("92b0ee3b", "Bill Hayne",    "Billy/Bill standard nickname pair"),
    "Fred Kipley":    ("8acf4e97", "Fred Kippley",  "Single missing 'p' — OCR/transcription error"),
    "Misty Helme":    ("e85a6af9", "Misty Helms",   "Missing terminal 's' — transcription error"),
    "Max Smith Jr.":  ("95fb4def", "Max Smith",     "Identical person, suffix 'Jr.' only"),
    "Gary Laut":      ("66a5ee0b", "Gary Lautt",    "Single missing 't' — transcription error"),
}

# Names requiring manual review: cannot be auto-resolved under the policy.
# Format: raw_name -> (candidate_pid_prefix or "", candidate_canon, review_notes)
NEEDS_REVIEW: dict[str, tuple] = {
    "Jim Caveney": (
        "df329352", "Jimmy Caveney",
        "Nickname vs formal name. PT rename Jimmy→Jim was planned but not yet in PT. "
        "Confirm before accepting.",
    ),
    "Jody Grace": (
        "73728c19", "Judy Grace",
        "Different first name (Jody vs Judy). Could be same person or different. "
        "Needs external evidence.",
    ),
    "Jenny Davidson": (
        "3f48caf5", "Jenny Davison",
        "Single char surname difference (Davidson vs Davison). Could be same person. "
        "Verify via source image.",
    ),
    "Ken Shults": (
        "2a6a7c9e", "Kenneth Shults",
        "Ken is common short form of Kenneth but PT alias is 'Kenny Schults' (different "
        "spelling). No 'Ken Shults' alias registered.",
    ),
    "Kenny Shults": (
        "2a6a7c9e", "Kenneth Shults",
        "PT alias is 'Kenny Schults', not 'Kenny Shults'. Spelling difference in last name. "
        "Confirm before accepting.",
    ),
    "Karin Atogpian": (
        "", "Karen Atgopian (also unresolved)",
        "Both 'Karin Atogpian' and 'Karen Atgopian' are unresolved. Likely same person "
        "with variant spelling. Merge only after verifying against source.",
    ),
    "Lori Jean Tarr": (
        "77f0f32d", "Lori Jean Conover",
        "Different surname. Could be maiden/married name change. "
        "Needs external confirmation.",
    ),
    "Tim Fitzgerald": (
        "b54020bc", "Jim Fitzgerald",
        "Completely different first name. Likely source transcription error (Tim vs Jim). "
        "Verify against source image before accepting.",
    ),
    "Tobin Wigger": (
        "f38e2f8e", "Torbin Wigger",
        "Scandinavian name variant (Tobin vs Torbin). Single char, but these are proper "
        "Scandinavian names with distinct forms. Needs verification.",
    ),
    "Torben Wigger": (
        "f38e2f8e", "Torbin Wigger",
        "Scandinavian name variant (Torben vs Torbin). PT spells it 'Torbin'. "
        "Needs verification.",
    ),
    "Steve Fennell": (
        "fb605451", "Steve Femmel",
        "Different spelling (Fennell vs Femmel). Not a 1-char change — could be same "
        "person or different. Verify source.",
    ),
    "Ted Martens": (
        "3d1b24de", "Ted Martin",
        "Similar but different surname (Martens vs Martin). Could be transcription error. "
        "Needs verification.",
    ),
}

# Genuine new players: absent from PT, not noise, not ambiguous.
# These are pre-1997 athletes who likely never competed post-1997.
NEW_PLAYERS: list[str] = [
    "Colin Cowles",
    "Doug Little",
    "James Stallcup",
    "Jan-Olof Karlsson",
    "Karen Atgopian",
    "Karen Uppinghouse",
    "Ken Eldrick",
    "Kevin Gaunce",
    "Randy Nelson",
    "Randy Ross",
    "Sindre Madsen",
    "Steve Brown",
    "Ted Johnson",
    "Torgeir Rygh",
]

# Literal noise
NOISE: list[str] = ["unknown"]

# Namespace UUID for deterministic new early-player IDs
EARLY_NS = uuid.UUID("a1b2c3d4-e5f6-5890-abcd-ef1234567890")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def stable_id(s: str) -> str:
    return str(uuid.uuid5(EARLY_NS, norm(s)))


def make_result_id(canonical_event_id: str, division_raw: str, place: str) -> str:
    key = f"{canonical_event_id}|{division_raw}|{place}"
    return hashlib.sha1(key.encode()).hexdigest()[:10]


def write_csv(path: Path, fields: list, rows: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


# ---------------------------------------------------------------------------
# Load inputs
# ---------------------------------------------------------------------------

def load_inputs():
    pt          = list(csv.DictReader(open(PERSONS_TRUTH,  encoding="utf-8")))
    candidates  = list(csv.DictReader(open(CANDIDATES,     encoding="utf-8")))
    plc_fbw     = list(csv.DictReader(open(PLACEMENTS_FBW, encoding="utf-8")))
    plc_or      = list(csv.DictReader(open(PLACEMENTS_OR,  encoding="utf-8")))
    ev_fbw      = list(csv.DictReader(open(EVENTS_FBW,     encoding="utf-8")))
    ev_or       = list(csv.DictReader(open(EVENTS_OR,      encoding="utf-8")))
    canon_ev    = list(csv.DictReader(open(CANONICAL_EV,   encoding="utf-8")))
    eid_map     = list(csv.DictReader(open(EID_MAPPING,    encoding="utf-8")))
    comparison  = list(csv.DictReader(open(EV_COMPARISON,  encoding="utf-8")))
    return pt, candidates, plc_fbw, plc_or, ev_fbw, ev_or, canon_ev, eid_map, comparison


# ---------------------------------------------------------------------------
# PART 2 — Build identity review files
# ---------------------------------------------------------------------------

def build_identity_review(pt_rows, candidates):
    """Classify all unresolved names and produce the 4 review files."""

    # PT lookup by person_id prefix (8 chars)
    pt_by_pid = {}
    for r in pt_rows:
        pt_by_pid[r["effective_person_id"][:8]] = r
        pt_by_pid[r["effective_person_id"]] = r

    # Build auto-accepted rows
    aa_rows = []
    for raw_name, (pid_prefix, canon_name, reason) in AUTO_ACCEPT.items():
        pt_row = pt_by_pid.get(pid_prefix, {})
        full_pid = pt_row.get("effective_person_id", pid_prefix)
        aa_rows.append({
            "raw_name":           raw_name,
            "matched_person_id":  full_pid,
            "matched_person_name": canon_name,
            "reason":             reason,
            "confidence":         "HIGH",
        })

    # Build review rows
    rev_rows = []
    for raw_name, (pid_prefix, cand_name, notes) in NEEDS_REVIEW.items():
        pt_row = pt_by_pid.get(pid_prefix, {})
        full_pid = pt_row.get("effective_person_id", pid_prefix)
        rev_rows.append({
            "raw_name":              raw_name,
            "candidate_person_ids":  full_pid,
            "candidate_person_names": cand_name,
            "reason":                "near-miss — requires manual verification",
            "confidence":            "LOW",
            "review_notes":          notes,
        })

    # Build new early player rows
    # Count how many placements each name appears in
    plc_fbw = list(csv.DictReader(open(PLACEMENTS_FBW, encoding="utf-8")))
    plc_or  = list(csv.DictReader(open(PLACEMENTS_OR,  encoding="utf-8")))
    all_plc = plc_fbw + plc_or
    name_sources: dict = defaultdict(set)
    for row in all_plc:
        for field in ("player_raw", "team_raw"):
            raw = row.get(field, "")
            if raw and "/" in raw:
                for part in raw.split("/"):
                    name_sources[part.strip()].add(row.get("source_file", ""))
            elif raw:
                name_sources[raw.strip()].add(row.get("source_file", ""))

    np_rows = []
    for raw_name in NEW_PLAYERS:
        sources = name_sources.get(raw_name, set())
        np_rows.append({
            "raw_name":            raw_name,
            "proposed_person_name": raw_name,  # exact raw spelling initially
            "reason":              "not in PT; not obviously noise; pre-1997 only athlete",
            "source_count":        len(sources),
            "notes":               "; ".join(sorted(sources)) if sources else "",
        })

    # Build noise rows
    noise_rows = [
        {"raw_name": n, "reason": "literal noise / unknown value", "notes": ""}
        for n in NOISE
    ]

    return aa_rows, rev_rows, np_rows, noise_rows


# ---------------------------------------------------------------------------
# PART 3 — Build person resolution lookup
# ---------------------------------------------------------------------------

def build_resolution_lookup(pt_rows, aa_rows, np_rows):
    """
    Returns a dict: norm(raw_name) -> {
        person_id, person_canon, resolution_status
    }
    """
    lookup: dict = {}

    # Step 1: Load all EXACT matches from person_match_candidates.csv
    for row in csv.DictReader(open(CANDIDATES, encoding="utf-8")):
        if row["match_type"] in ("EXACT", "CASE", "ALIAS"):
            lookup[norm(row["raw_name"])] = {
                "person_id":         row["person_id"],
                "person_canon":      row["person_canon"],
                "resolution_status": "MATCHED",
            }

    # Step 2: Auto-accepted aliases → map to existing PT person
    pt_by_pid = {r["effective_person_id"]: r for r in pt_rows}
    pt_by_pid.update({r["effective_person_id"][:8]: r for r in pt_rows})

    for aa in aa_rows:
        pid = aa["matched_person_id"]
        pt_row = pt_by_pid.get(pid, {})
        full_pid = pt_row.get("effective_person_id", pid)
        lookup[norm(aa["raw_name"])] = {
            "person_id":         full_pid,
            "person_canon":      aa["matched_person_name"],
            "resolution_status": "AUTOACCEPTED",
        }

    # Step 3: New early players → assign new stable IDs
    for np in np_rows:
        raw = np["raw_name"]
        new_pid = stable_id(raw)
        lookup[norm(raw)] = {
            "person_id":         new_pid,
            "person_canon":      raw,
            "resolution_status": "NEW_PLAYER",
        }

    # Step 4: Review-needed names → mark but leave person_id blank
    for rn in NEEDS_REVIEW:
        if norm(rn) not in lookup:
            lookup[norm(rn)] = {
                "person_id":         "",
                "person_canon":      "",
                "resolution_status": "REVIEW_NEEDED",
            }

    # Step 5: Noise
    for n in NOISE:
        lookup[norm(n)] = {
            "person_id":         "",
            "person_canon":      "",
            "resolution_status": "NOISE",
        }

    return lookup


# ---------------------------------------------------------------------------
# PART 3b — Build persons_pre1997.csv and person_aliases_pre1997.csv
# ---------------------------------------------------------------------------

def build_persons_table(pt_rows, aa_rows, np_rows, resolution_lookup):
    """
    Build persons_pre1997.csv:
      - Only PT persons referenced by pre-1997 placements
      - Plus new early players
    """
    # Collect all person_ids referenced in pre-1997 data
    referenced_pids = set()
    for entry in resolution_lookup.values():
        if entry["person_id"]:
            referenced_pids.add(entry["person_id"])

    pt_by_pid = {r["effective_person_id"]: r for r in pt_rows}

    persons_rows = []
    alias_rows   = []

    # PT persons referenced by pre-1997 data
    for pid in sorted(referenced_pids):
        pt_row = pt_by_pid.get(pid)
        if pt_row:
            persons_rows.append({
                "person_id":     pid,
                "person_canon":  pt_row["person_canon"],
                "source_scope":  "POST1997",
                "person_status": "ESTABLISHED",
                "aliases":       pt_row.get("aliases", ""),
                "notes":         "Referenced in pre-1997 placement data",
            })

    # New early players
    for np in np_rows:
        raw  = np["raw_name"]
        pid  = stable_id(raw)
        persons_rows.append({
            "person_id":     pid,
            "person_canon":  raw,
            "source_scope":  "PRE1997_ONLY",
            "person_status": "NEW_EARLY_PLAYER",
            "aliases":       "",
            "notes":         np.get("notes", ""),
        })

    # person_aliases_pre1997.csv — all name→pid mappings
    for raw_name_norm, entry in sorted(resolution_lookup.items()):
        # Find the original raw name
        raw_name = raw_name_norm  # use norm'd as approximation
        alias_rows.append({
            "raw_name":        raw_name_norm,
            "person_id":       entry["person_id"],
            "alias_status":    entry["resolution_status"],
        })

    return persons_rows, alias_rows


# ---------------------------------------------------------------------------
# PART 4 — Build canonical CSVs
# ---------------------------------------------------------------------------

def build_canonical_csvs(plc_fbw, plc_or, ev_fbw, ev_or,
                          canon_ev, eid_map, resolution_lookup):
    """
    Build events/disciplines/results/participants canonical CSVs.
    """
    all_placements = [(r, "FBW")  for r in plc_fbw] + \
                     [(r, "OLD_RESULTS") for r in plc_or]

    # Build source_event_id → canonical_event_id map
    src_to_canon = {r["event_id"]: r["canonical_event_id"] for r in eid_map}

    # Pre-1997 canonical events (year < 1997)
    canon_events_pre97 = [r for r in canon_ev if r["year"] < "1997"]
    canon_pids = {r["canonical_event_id"] for r in canon_events_pre97}

    # ---- events_pre1997.csv ----
    events_rows = []
    for r in sorted(canon_events_pre97, key=lambda x: (x["year"], x["event_name_raw"])):
        events_rows.append({
            "canonical_event_id":    r["canonical_event_id"],
            "event_name":            r["event_name_raw"],
            "year":                  r["year"],
            "location":              r["location_raw"],
            "normalized_event_type": r["normalized_event_type"],
            "source_types":          r["source_types"],
            "num_sources":           r["num_sources"],
            "validation_status":     r["validation_status"],
            "confidence":            r["confidence"],
            "num_placements":        r["num_placements"],
        })

    # ---- Process placements → results + participants ----
    result_rows: list = []
    participant_rows: list = []
    disc_set: dict = {}  # (canonical_event_id, division_raw) → count

    for plc_row, stype in all_placements:
        src_eid = plc_row["event_id"]
        canon_eid = src_to_canon.get(src_eid)
        if canon_eid not in canon_pids:
            continue  # out of scope

        division_raw = plc_row["division_raw"]
        place        = plc_row["placement_num"]
        player_raw   = plc_row["player_raw"].strip()
        team_raw     = plc_row["team_raw"].strip()
        result_id    = make_result_id(canon_eid, division_raw, place)

        # Track discipline
        disc_key = (canon_eid, division_raw)
        disc_set[disc_key] = disc_set.get(disc_key, 0) + 1

        result_rows.append({
            "result_id":          result_id,
            "canonical_event_id": canon_eid,
            "division_raw":       division_raw,
            "place":              place,
            "player_raw":         player_raw,
            "team_raw":           team_raw,
            "source_event_id":    src_eid,
            "source_type":        stype,
        })

        # Expand participants
        if team_raw:
            raw_names = [p.strip() for p in team_raw.split("/") if p.strip()]
        elif "/" in player_raw:
            raw_names = [p.strip() for p in player_raw.split("/") if p.strip()]
        else:
            raw_names = [player_raw] if player_raw else []

        for raw_name in raw_names:
            entry = resolution_lookup.get(norm(raw_name), {
                "person_id": "", "person_canon": "", "resolution_status": "UNRESOLVED"
            })
            participant_rows.append({
                "result_id":          result_id,
                "canonical_event_id": canon_eid,
                "division_raw":       division_raw,
                "place":              place,
                "player_name_raw":    raw_name,
                "person_id":          entry["person_id"],
                "person_canon":       entry["person_canon"],
                "resolution_status":  entry["resolution_status"],
                "source_type":        stype,
            })

    # ---- event_disciplines_pre1997.csv ----
    discipline_rows = []
    for (canon_eid, div_raw), count in sorted(disc_set.items()):
        discipline_rows.append({
            "canonical_event_id": canon_eid,
            "division_raw":       div_raw,
            "total_placements":   count,
        })

    return events_rows, discipline_rows, result_rows, participant_rows


# ---------------------------------------------------------------------------
# PART 5 — Build Excel workbook
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUBHEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
EVENT_FILL = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)
BOLD_WHITE = Font(bold=True, color="FFFFFF")
BOLD_DARK = Font(bold=True, color="1F3864")


def _hdr(ws, row_idx: int, values: list, fill=None, font=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row_idx + 1


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_sheet_readme(wb):
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 80
    ws.sheet_view.showGridLines = False

    data = [
        ("DATASET",    "Footbag Results 1980–1996 — Historical Recovery"),
        ("VERSION",    "Pre-1997 Early Data Recovery Package"),
        ("STATUS",     "INCOMPLETE — UNDER ACTIVE RECONSTRUCTION"),
        ("",           ""),
        ("PURPOSE",    "This workbook contains reconstructed historical footbag competition "
                       "results from 1980 to 1996. It is a SEPARATE dataset from the published "
                       "post-1997 release and should be treated independently."),
        ("",           ""),
        ("SOURCES",    "FBW Magazine scans (Vols 2–14, pages extracted via AI)\n"
                       "IFAB Worlds History page (scanned)\n"
                       "OLD_RESULTS.txt (contributed text file, 1980–1986)"),
        ("",           ""),
        ("PROVENANCE", "All raw names and event titles are preserved exactly as found in "
                       "sources. No silent normalisation has been applied to results data."),
        ("",           ""),
        ("PERSONS",    "Person identity has been resolved conservatively against the "
                       "post-1997 Persons_Truth table where unambiguous. Uncertain cases "
                       "are preserved as unresolved or flagged for review."),
        ("",           ""),
        ("WARNING",    "Data is incomplete. Many years have partial coverage only. "
                       "Some results may be duplicate across sources. Cross-source "
                       "validation is noted in the VALIDATION SUMMARY sheet."),
    ]
    r = 1
    ws.cell(r, 1, "PRE-1997 RECOVERY DATASET").font = Font(bold=True, size=14)
    ws.merge_cells(f"A{r}:B{r}")
    r += 1
    ws.cell(r, 1, "footbag_results_pre1997_recovery.xlsx").font = Font(italic=True, color="595959")
    ws.merge_cells(f"A{r}:B{r}")
    r += 2
    for label, text in data:
        ws.cell(r, 1, label).font = BOLD
        ws.cell(r, 2, text).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(30, len(text) // 3) if text else 15
        r += 1


def build_sheet_data_notes(wb):
    ws = wb.create_sheet("DATA NOTES")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 75
    ws.sheet_view.showGridLines = False

    notes = [
        ("Completeness",
         "This is NOT a complete record of pre-1997 footbag events. Coverage depends "
         "entirely on what was printed in FBW Magazine and what text sources are available. "
         "Many events are missing entirely."),
        ("Source types",
         "FBW: Footbag World magazine scans, AI-extracted. Accuracy depends on scan "
         "quality and AI extraction.\n"
         "IFAB: IFAB Worlds History page (lists only World Championships).\n"
         "OLD_RESULTS: contributed text file with 1980-1986 championships results."),
        ("Cross-source validation",
         "Events confirmed by 2+ independent sources are marked CONFIRMED_MULTI_SOURCE. "
         "Events from a single source only are SINGLE_SOURCE. No CONFLICT cases found."),
        ("Identity resolution",
         "83 of 115 unique player names matched exactly to the post-1997 Persons_Truth. "
         "5 names auto-accepted as safe spelling variants. 12 names require manual review. "
         "14 names added as new pre-1997-only players. 1 literal 'unknown' preserved."),
        ("Team entries",
         "Team players are listed with '/' separator in player_raw. "
         "Individual identities within teams are resolved independently."),
        ("Year coverage",
         "1980–1986: Old Results text + FBW Magazine (well-covered for top events)\n"
         "1987–1992: FBW Magazine WFA World Championships\n"
         "1993–1996: FBW + IFAB IFAB World Championships\n"
         "Other events: sparse, varies by year"),
        ("Conflicts",
         "No location conflicts were detected across sources. "
         "Some division names differ across sources for the same event — this is expected "
         "and all raw division names are preserved."),
        ("Auditability",
         "All raw names, raw event titles, and source files are preserved in the canonical "
         "CSVs. This workbook is derived from those CSVs and should not be treated as a "
         "correction to them."),
    ]

    r = 1
    ws.cell(r, 1, "DATA NOTES").font = Font(bold=True, size=13)
    ws.merge_cells(f"A{r}:B{r}")
    r += 2
    for topic, body in notes:
        ws.cell(r, 1, topic).font = BOLD
        cell = ws.cell(r, 2, body)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(40, len(body) // 2)
        r += 1
        ws.cell(r, 1, "").fill = PatternFill("solid", fgColor="F2F2F2")
        r += 1


def build_sheet_event_index(wb, events_rows, comparison_rows):
    ws = wb.create_sheet("EVENT INDEX")
    _set_col_widths(ws, [14, 38, 6, 28, 12, 10, 22, 12])
    headers = ["canonical_event_id", "event_name", "year", "normalized_event_type",
               "source_types", "num_sources", "validation_status", "num_placements"]
    r = _hdr(ws, 1, headers, fill=HEADER_FILL, font=BOLD_WHITE)
    ws.freeze_panes = "A2"

    for ev in sorted(events_rows, key=lambda x: (x["year"], x["event_name"])):
        for col, fld in enumerate(headers, 1):
            ws.cell(r, col, ev.get(fld, ""))
        r += 1


def build_sheet_player_summary(wb, participant_rows, persons_rows):
    ws = wb.create_sheet("PLAYER SUMMARY")
    _set_col_widths(ws, [36, 12, 36, 14, 12])
    headers = ["player_name_raw", "appearances", "events (years)", "person_id", "resolution_status"]
    r = _hdr(ws, 1, headers, fill=HEADER_FILL, font=BOLD_WHITE)
    ws.freeze_panes = "A2"

    # Build summary
    summary: dict = {}
    for row in participant_rows:
        raw = row["player_name_raw"]
        if not raw:
            continue
        if raw not in summary:
            summary[raw] = {
                "appearances": 0,
                "events": set(),
                "years": set(),
                "person_id": row["person_id"],
                "resolution_status": row["resolution_status"],
            }
        summary[raw]["appearances"] += 1
        summary[raw]["events"].add(row["canonical_event_id"])

    for name in sorted(summary.keys()):
        s = summary[name]
        ev_list = ", ".join(sorted(s["events"])[:4])
        ws.cell(r, 1, name)
        ws.cell(r, 2, s["appearances"])
        ws.cell(r, 3, ev_list)
        ws.cell(r, 4, s["person_id"])
        ws.cell(r, 5, s["resolution_status"])
        r += 1


def build_sheet_validation_summary(wb, comparison_rows):
    ws = wb.create_sheet("VALIDATION SUMMARY")
    _set_col_widths(ws, [14, 28, 6, 24, 24, 12, 50])
    headers = ["group_id", "normalized_event_type", "year",
               "validation_status", "source_types", "num_sources", "sources"]
    r = _hdr(ws, 1, headers, fill=HEADER_FILL, font=BOLD_WHITE)
    ws.freeze_panes = "A2"

    counts: dict = defaultdict(int)
    for row in sorted(comparison_rows, key=lambda x: (x["year"], x["normalized_event_type"])):
        yr = int(row["year"]) if row["year"] else 9999
        if yr >= 1997:
            continue
        for col, fld in enumerate(headers, 1):
            ws.cell(r, col, row.get(fld, ""))
        counts[row["validation_status"]] += 1
        r += 1

    r += 2
    ws.cell(r, 1, "SUMMARY").font = BOLD
    r += 1
    for status in ["CONFIRMED_MULTI_SOURCE", "SINGLE_SOURCE", "CONFLICT"]:
        ws.cell(r, 1, status)
        ws.cell(r, 2, counts.get(status, 0))
        r += 1


def build_year_sheets(wb, events_rows, result_rows, participant_rows, canon_ev):
    """One sheet per year, showing all events/divisions/results."""

    # Index structures
    canon_ev_by_id = {r["canonical_event_id"]: r for r in canon_ev}
    canon_ev_map = {r["canonical_event_id"]: r for r in events_rows}

    # Group results by year → canonical_event_id → division_raw → place
    by_year: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for prow in participant_rows:
        ceid  = prow["canonical_event_id"]
        ev    = canon_ev_map.get(ceid, {})
        year  = ev.get("year", "")
        if not year or int(year) >= 1997:
            continue
        div   = prow["division_raw"] or "(no division)"
        place = prow["place"]
        by_year[year][ceid][div].append(prow)

    for year in sorted(by_year.keys()):
        ws = wb.create_sheet(str(year))
        _set_col_widths(ws, [14, 32, 26, 6, 34, 36, 14])
        r = 1

        # Year header
        ws.cell(r, 1, f"  {year} Results").font = Font(bold=True, size=13)
        ws.merge_cells(f"A{r}:G{r}")
        ws.cell(r, 1).fill = HEADER_FILL
        ws.cell(r, 1).font = BOLD_WHITE
        r += 2

        for ceid in sorted(by_year[year].keys()):
            ev = canon_ev_map.get(ceid, {})
            ev_name = ev.get("event_name", ceid)
            ev_loc  = ev.get("location", "")
            ev_src  = ev.get("source_types", "")
            ev_val  = ev.get("validation_status", "")

            # Event name row
            header_text = ev_name
            if ev_loc:
                header_text += f"  —  {ev_loc}"
            ws.cell(r, 1, header_text).font = BOLD_DARK
            ws.merge_cells(f"A{r}:D{r}")
            ws.cell(r, 5, f"Source: {ev_src}").font = Font(italic=True, color="595959", size=9)
            ws.cell(r, 6, ev_val).font = Font(italic=True, color="595959", size=9)
            ws.cell(r, 7, ceid[:8]).font = Font(italic=True, color="595959", size=9)
            for col in range(1, 8):
                ws.cell(r, col).fill = EVENT_FILL
            r += 1

            # Column headers for this event
            r = _hdr(ws, r,
                     ["event_id", "division", "source_type", "place",
                      "player_raw", "person_canon", "resolution"],
                     fill=SUBHEADER_FILL, font=BOLD_DARK)

            divs = by_year[year][ceid]
            for div in sorted(divs.keys()):
                prows = sorted(divs[div],
                               key=lambda x: (int(x["place"]) if x["place"].isdigit() else 999,
                                              x["source_type"]))
                for prow in prows:
                    ws.cell(r, 1, ceid[:8])
                    ws.cell(r, 2, div if div != "(no division)" else "")
                    ws.cell(r, 3, prow["source_type"])
                    ws.cell(r, 4, prow["place"])
                    ws.cell(r, 5, prow["player_name_raw"])
                    ws.cell(r, 6, prow["person_canon"])
                    status = prow["resolution_status"]
                    ws.cell(r, 7, status)
                    if status in ("REVIEW_NEEDED", "UNRESOLVED"):
                        ws.cell(r, 7).font = Font(color="C00000")
                    elif status in ("NEW_PLAYER", "AUTOACCEPTED"):
                        ws.cell(r, 7).font = Font(color="7030A0")
                    r += 1
            r += 1  # blank row between events

        ws.freeze_panes = "A3"


def build_workbook(events_rows, discipline_rows, result_rows, participant_rows,
                   persons_rows, canon_ev, comparison_rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    build_sheet_readme(wb)
    build_sheet_data_notes(wb)
    build_sheet_event_index(wb, events_rows, comparison_rows)
    build_sheet_player_summary(wb, participant_rows, persons_rows)
    build_year_sheets(wb, events_rows, result_rows, participant_rows, canon_ev)
    build_sheet_validation_summary(wb, comparison_rows)

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

AUTOACC_FIELDS  = ["raw_name", "matched_person_id", "matched_person_name", "reason", "confidence"]
REVIEW_FIELDS   = ["raw_name", "candidate_person_ids", "candidate_person_names",
                   "reason", "confidence", "review_notes"]
NEWPLAYER_FIELDS= ["raw_name", "proposed_person_name", "reason", "source_count", "notes"]
NOISE_FIELDS    = ["raw_name", "reason", "notes"]
PERSONS_FIELDS  = ["person_id", "person_canon", "source_scope", "person_status", "aliases", "notes"]
PALIASES_FIELDS = ["raw_name", "person_id", "alias_status"]
EVENTS_FIELDS   = ["canonical_event_id", "event_name", "year", "location",
                   "normalized_event_type", "source_types", "num_sources",
                   "validation_status", "confidence", "num_placements"]
DISC_FIELDS     = ["canonical_event_id", "division_raw", "total_placements"]
RESULTS_FIELDS  = ["result_id", "canonical_event_id", "division_raw", "place",
                   "player_raw", "team_raw", "source_event_id", "source_type"]
PARTS_FIELDS    = ["result_id", "canonical_event_id", "division_raw", "place",
                   "player_name_raw", "person_id", "person_canon",
                   "resolution_status", "source_type"]


def main():
    print("=== 07_build_early_release.py ===\n")

    print("Loading inputs…")
    (pt_rows, candidates, plc_fbw, plc_or,
     ev_fbw, ev_or, canon_ev, eid_map, comparison) = load_inputs()
    print(f"  PT: {len(pt_rows)} persons")
    print(f"  Placements: {len(plc_fbw)} FBW + {len(plc_or)} OLD_RESULTS")
    print(f"  Canonical events: {len(canon_ev)}")

    # --- Part 2: identity review files ---
    print("\nApplying identity resolution policy…")
    aa_rows, rev_rows, np_rows, noise_rows = build_identity_review(pt_rows, candidates)
    write_csv(OUT_AUTOACC,  AUTOACC_FIELDS,   aa_rows)
    write_csv(OUT_REVIEW,   REVIEW_FIELDS,    rev_rows)
    write_csv(OUT_NEWPLAYERS, NEWPLAYER_FIELDS, np_rows)
    write_csv(OUT_NOISE,    NOISE_FIELDS,     noise_rows)
    print(f"  AUTO-ACCEPT:    {len(aa_rows)} aliases")
    print(f"  REVIEW NEEDED:  {len(rev_rows)} aliases")
    print(f"  NEW PLAYERS:    {len(np_rows)}")
    print(f"  NOISE:          {len(noise_rows)}")

    # --- Part 3: resolution lookup + persons table ---
    print("\nBuilding resolution lookup and persons table…")
    resolution_lookup = build_resolution_lookup(pt_rows, aa_rows, np_rows)
    persons_rows, alias_rows = build_persons_table(pt_rows, aa_rows, np_rows, resolution_lookup)
    write_csv(OUT_PERSONS,  PERSONS_FIELDS,  persons_rows)
    write_csv(OUT_PALIASES, PALIASES_FIELDS, alias_rows)
    established = sum(1 for r in persons_rows if r["person_status"] == "ESTABLISHED")
    new_players = sum(1 for r in persons_rows if r["person_status"] == "NEW_EARLY_PLAYER")
    print(f"  persons_pre1997.csv: {len(persons_rows)} rows "
          f"({established} established, {new_players} new early players)")

    # --- Part 4: canonical CSVs ---
    print("\nBuilding canonical CSVs…")
    events_rows, discipline_rows, result_rows, participant_rows = build_canonical_csvs(
        plc_fbw, plc_or, ev_fbw, ev_or, canon_ev, eid_map, resolution_lookup
    )
    write_csv(OUT_EVENTS,      EVENTS_FIELDS, events_rows)
    write_csv(OUT_DISCIPLINES, DISC_FIELDS,   discipline_rows)
    write_csv(OUT_RESULTS,     RESULTS_FIELDS,result_rows)
    write_csv(OUT_PARTICIPANTS,PARTS_FIELDS,  participant_rows)
    print(f"  events_pre1997.csv:              {len(events_rows)} events")
    print(f"  event_disciplines_pre1997.csv:   {len(discipline_rows)} disciplines")
    print(f"  event_results_pre1997.csv:       {len(result_rows)} placements")
    print(f"  event_result_participants_pre1997.csv: {len(participant_rows)} participants")

    # Resolution status breakdown in participants
    status_counts: dict = defaultdict(int)
    for r in participant_rows:
        status_counts[r["resolution_status"]] += 1
    print("  Participant resolution breakdown:")
    for s in ["MATCHED", "AUTOACCEPTED", "NEW_PLAYER", "REVIEW_NEEDED", "UNRESOLVED", "NOISE"]:
        print(f"    {s:15s}: {status_counts.get(s, 0)}")

    # --- Part 5: Excel workbook ---
    print("\nBuilding Excel workbook…")
    wb = build_workbook(events_rows, discipline_rows, result_rows, participant_rows,
                        persons_rows, canon_ev, comparison)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    sheets = [ws.title for ws in wb.worksheets]
    print(f"  Sheets ({len(sheets)}): {', '.join(sheets)}")
    print(f"  Saved: {OUT_XLSX.name}")

    # --- Final summary ---
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"\nIdentity resolution:")
    print(f"  Auto-accepted aliases:  {len(aa_rows)}")
    for r in aa_rows:
        print(f"    {r['raw_name']!r:22s} → {r['matched_person_name']}  [{r['reason'][:45]}]")
    print(f"\n  Review-needed ({len(rev_rows)}):")
    for r in rev_rows:
        print(f"    {r['raw_name']!r:22s} ↔ {r['candidate_person_names'][:30]}")
    print(f"\n  New early players ({len(np_rows)}):")
    for r in np_rows:
        print(f"    {r['proposed_person_name']}")
    print(f"\n  Noise preserved ({len(noise_rows)}):  {[r['raw_name'] for r in noise_rows]}")
    print(f"\nIntentionally left unresolved: the {len(rev_rows)} REVIEW_NEEDED names above.")
    print("These require manual source verification before any PT promotion.")
    print("\nDone.")


if __name__ == "__main__":
    main()
