"""
11_finalize_pre1997.py — Produce stable PRE-1997 v1.0 release artifacts

Reads from early_data/canonical/ (PRE1997 V2 state) and:

1. Applies expert-confirmed identity corrections
2. Verifies strict year < 1997 separation
3. Writes early_data/final_pre1997/ (release-ready canonical CSVs)
4. Writes identity_review_log.csv
5. Writes validation_summary.txt
6. Writes footbag_results_pre1997_v1.xlsx

Expert-confirmed identity fixes applied here
(in addition to V2 automated/reviewed aliases):

  Billy Hayne     → Bill Hayne        (already AUTOACCEPTED — promote to EXPERT_CONFIRMED)
  Fred Kipley     → Fred Kippley      (already AUTOACCEPTED — promote to EXPERT_CONFIRMED)
  Misty Helme     → Misty Helms       (already AUTOACCEPTED — promote to EXPERT_CONFIRMED)
  Tobin Wigger    → Torben Wigger     (already ACCEPTED — promote + update display name)
  Torbin Wigger   → Torben Wigger     (new: "Torbin" variant → canonical "Torben" in pre-1997)
  Tim Fitzgerald  → Jim Fitzgerald    (already ACCEPTED — promote to EXPERT_CONFIRMED)

Wigger note: The PT has "Torbin Wigger" as the post-1997 canonical. Expert confirms
the historically correct spelling is "Torben". The pre-1997 dataset uses "Torben Wigger"
as the canonical display while preserving the same person_id (f38e2f8e) that links
to the post-1997 PT person. This does NOT modify the post-1997 PT.
"""

import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT       = Path(__file__).resolve().parents[2]
EARLY      = ROOT / "early_data"
CANONICAL  = EARLY / "canonical"
FINAL      = EARLY / "final_pre1997"
IDENTITY   = EARLY / "identity"
REVIEW     = EARLY / "review"

FINAL.mkdir(exist_ok=True)

TODAY = date.today().isoformat()
VERSION = "v1.0"

# ── Expert-confirmed identity mappings ────────────────────────────────────────
# raw_name_lower → (person_id, canonical_display, previous_status)
EXPERT_CONFIRMED = {
    "billy hayne":    ("92b0ee3b-efaa-545a-b07e-30ab6d8ebeb0",  "Bill Hayne",       "AUTOACCEPTED"),
    "fred kipley":    ("8acf4e97-e1e8-50d7-8fd4-3a1d8ff4e731",  "Fred Kippley",     "AUTOACCEPTED"),
    "misty helme":    ("e85a6af9-3ac6-550d-8b35-045ec4886d06",  "Misty Helms",      "AUTOACCEPTED"),
    "tobin wigger":   ("f38e2f8e-cba6-5fff-8b20-90791fd0d794",  "Torben Wigger",    "ACCEPTED"),
    "torben wigger":  ("f38e2f8e-cba6-5fff-8b20-90791fd0d794",  "Torben Wigger",    "ACCEPTED"),
    "tim fitzgerald": ("b54020bc-1a1a-5d23-89e1-34617b3514fa",  "Jim Fitzgerald",   "ACCEPTED"),
}

# The PT entry for Wigger is "Torbin Wigger" — in the pre-1997 context we override
# the display to "Torben Wigger" for this specific person_id.
WIGGER_PERSON_ID = "f38e2f8e-cba6-5fff-8b20-90791fd0d794"
WIGGER_DISPLAY   = "Torben Wigger"   # expert-confirmed correct historical spelling


# ── helpers ───────────────────────────────────────────────────────────────────

def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def write_csv(path, rows, fieldnames=None):
    if fieldnames is None and rows:
        fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames or [])
        w.writeheader()
        w.writerows(rows)
    print(f"  Wrote {len(rows):4d} rows → {path.relative_to(ROOT)}")

def _norm_div(s: str) -> str:
    s = s.lower().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _div_category(s: str) -> str:
    n = s.lower()
    if any(k in n for k in ("freestyle", "frstyl", "shred", "sick", "circle",
                             "routine", "battle", "combo")):
        return "freestyle"
    if "golf" in n:
        return "golf"
    if "sideline" in n:
        return "sideline"
    return "net"


# ── PART 1 — Apply expert-confirmed identity corrections ──────────────────────

def apply_expert_corrections(aliases, persons, participants):
    """
    Promote alias statuses to EXPERT_CONFIRMED.
    Update person_canon for Wigger entries to 'Torben Wigger'.
    Returns (aliases, persons, participants, review_log).
    """
    review_log = []

    # Update aliases table
    for alias_row in aliases:
        raw = alias_row["raw_name"].lower().strip()
        if raw in EXPERT_CONFIRMED:
            pid, canon, prev_status = EXPERT_CONFIRMED[raw]
            old_status = alias_row["alias_status"]
            if old_status != "EXPERT_CONFIRMED":
                alias_row["alias_status"] = "EXPERT_CONFIRMED"
                alias_row["person_id"] = pid
                review_log.append({
                    "raw_name": alias_row["raw_name"],
                    "action": "PROMOTE_TO_EXPERT_CONFIRMED",
                    "previous_status": old_status,
                    "new_status": "EXPERT_CONFIRMED",
                    "person_id": pid,
                    "canonical_display": canon,
                    "note": "",
                })

    # Update persons table: Torbin Wigger → Torben Wigger display in pre-1997 context
    for person_row in persons:
        if person_row["person_id"] == WIGGER_PERSON_ID:
            old_canon = person_row["person_canon"]
            if old_canon != WIGGER_DISPLAY:
                person_row["person_canon"] = WIGGER_DISPLAY
                review_log.append({
                    "raw_name": "(persons_table)",
                    "action": "UPDATE_PERSON_CANON",
                    "previous_status": old_canon,
                    "new_status": WIGGER_DISPLAY,
                    "person_id": WIGGER_PERSON_ID,
                    "canonical_display": WIGGER_DISPLAY,
                    "note": "PT has 'Torbin'; expert confirms correct historical spelling is 'Torben'",
                })

    # Update participants: any row with person_id == WIGGER_PERSON_ID → update person_canon
    for p in participants:
        if p["person_id"] == WIGGER_PERSON_ID and p["person_canon"] != WIGGER_DISPLAY:
            p["person_canon"] = WIGGER_DISPLAY

    return aliases, persons, participants, review_log


# ── PART 2 — Year separation verification ────────────────────────────────────

def verify_year_separation(events, results, participants):
    """Confirm no post-1997 data in the pre-1997 outputs. Returns list of violations."""
    violations = []
    event_years = {e["canonical_event_id"]: int(e["year"]) for e in events}
    for e in events:
        yr = int(e["year"])
        if yr >= 1997:
            violations.append(f"EVENT year≥1997: {e['canonical_event_id']} {e['year']} {e['event_name']}")
    for r in results:
        eid = r["canonical_event_id"]
        yr = event_years.get(eid)
        if yr and yr >= 1997:
            violations.append(f"RESULT from post-1997 event: {eid} result_id={r['result_id']}")
    return violations


# ── PART 3 — Write final_pre1997/ canonical CSVs ─────────────────────────────

def write_final_csvs(events, results, participants, persons, disciplines, aliases):
    """Write all canonical CSVs to final_pre1997/ directory."""

    # events_pre1997.csv — add version annotation
    for e in events:
        e["dataset_version"] = VERSION
    write_csv(FINAL / "events_pre1997.csv", events)

    # event_results_pre1997.csv
    write_csv(FINAL / "event_results_pre1997.csv", results)

    # event_result_participants_pre1997.csv
    write_csv(FINAL / "event_result_participants_pre1997.csv", participants)

    # persons_pre1997.csv
    write_csv(FINAL / "persons_pre1997.csv", persons)

    # event_disciplines_pre1997.csv
    write_csv(FINAL / "event_disciplines_pre1997.csv", disciplines)

    # person_aliases_pre1997.csv — includes raw_name, person_id, alias_status
    write_csv(FINAL / "person_aliases_pre1997.csv", aliases)


# ── PART 4 — Preserve review/unresolved files ─────────────────────────────────

def copy_review_files():
    """Copy identity review files into final_pre1997/ unchanged."""
    import shutil
    for src_dir, fname in [
        (IDENTITY, "unresolved_names.csv"),
        (IDENTITY, "person_aliases_needs_review.csv"),
        (REVIEW,   "person_alias_resolution.csv"),
        (REVIEW,   "event_group_resolution.csv"),
    ]:
        src = src_dir / fname
        if src.exists():
            shutil.copy2(src, FINAL / fname)
            print(f"  Copied  → {(FINAL / fname).relative_to(ROOT)}")
        else:
            print(f"  WARN: {src} not found, skipping")


# ── PART 5 — Validation summary ───────────────────────────────────────────────

def write_validation_summary(events, results, participants, persons, aliases):
    val_counts    = Counter(e["validation_status"] for e in events)
    status_counts = Counter(p["resolution_status"] for p in participants)
    alias_counts  = Counter(a["alias_status"] for a in aliases)
    scope_counts  = Counter(p["source_scope"] for p in persons)

    lines = [
        f"PRE-1997 Historical Footbag Dataset — Validation Summary",
        f"Dataset version: {VERSION}",
        f"Generated: {TODAY}",
        "",
        "=" * 56,
        "EVENT COVERAGE",
        "=" * 56,
        f"  Canonical events (year < 1997): {len(events)}",
        f"  Years covered:                  {min(e['year'] for e in events)}–{max(e['year'] for e in events)}",
        f"  Source placement rows:          {len(results)}",
        f"  Participant rows (expanded):    {len(participants)}",
        "",
        "CROSS-SOURCE VALIDATION",
        "-" * 40,
    ]
    for status in ("CONFIRMED_MULTI_SOURCE", "SINGLE_SOURCE", "CONFLICT"):
        lines.append(f"  {status:30s} {val_counts.get(status, 0)}")

    lines += [
        "",
        "IDENTITY RESOLUTION",
        "-" * 40,
    ]
    for status in ("MATCHED", "AUTOACCEPTED", "ACCEPTED",
                   "EXPERT_CONFIRMED", "NEW_PLAYER", "REVIEW_NEEDED", "UNRESOLVED", "NOISE"):
        cnt = status_counts.get(status, 0)
        if cnt:
            lines.append(f"  {status:25s} {cnt}")

    lines += [
        "",
        "ALIAS STATUS BREAKDOWN",
        "-" * 40,
    ]
    for status, cnt in sorted(alias_counts.items()):
        lines.append(f"  {status:25s} {cnt}")

    lines += [
        "",
        "PERSONS TABLE",
        "-" * 40,
        f"  Total persons referenced:  {len(persons)}",
    ]
    for scope, cnt in sorted(scope_counts.items()):
        lines.append(f"  {scope:25s} {cnt}")

    lines += [
        "",
        "YEAR DISTRIBUTION",
        "-" * 40,
    ]
    year_plc = Counter()
    from collections import defaultdict
    yr_events = defaultdict(list)
    for e in events:
        yr_events[e["year"]].append(e)
    for p in participants:
        yr = next((e["year"] for e in events
                   if e["canonical_event_id"] == p["canonical_event_id"]), "?")
        year_plc[yr] += 1
    for yr in sorted(yr_events.keys()):
        evts = yr_events[yr]
        lines.append(f"  {yr}  {len(evts):2d} events   {year_plc.get(yr, 0):4d} participants")

    lines += [
        "",
        "UNRESOLVED IDENTITIES",
        "-" * 40,
    ]
    unresolved = [p["player_name_raw"] for p in participants
                  if p["resolution_status"] in ("REVIEW_NEEDED", "UNRESOLVED")]
    uniq_unresolved = sorted(set(unresolved))
    if uniq_unresolved:
        for name in uniq_unresolved:
            lines.append(f"  {name}")
    else:
        lines.append("  (none)")

    lines += [
        "",
        "SEPARATION GUARANTEE",
        "-" * 40,
        "  All events in this dataset have year < 1997.",
        "  No post-1997 data is included.",
        "  The post-1997 published dataset is unchanged.",
        "",
        "SOURCES",
        "-" * 40,
        "  FBW Magazine (Vols 2-14) — AI extraction via Gemini",
        "  IFAB History Pages — AI extraction via Gemini",
        "  OLD_RESULTS.txt — Contributed historical text file",
    ]

    out_path = FINAL / "validation_summary.txt"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Wrote validation_summary.txt → {out_path.relative_to(ROOT)}")
    return lines


# ── PART 6 — Identity review log ─────────────────────────────────────────────

def write_identity_review_log(review_log):
    fields = ["raw_name", "action", "previous_status", "new_status",
              "person_id", "canonical_display", "note"]
    out_path = FINAL / "identity_review_log.csv"
    write_csv(out_path, review_log, fields)


# ── PART 7 — Release spreadsheet ─────────────────────────────────────────────

H1_FILL = PatternFill("solid", fgColor="1F4E79")
H1_FONT = Font(bold=True, color="FFFFFF", size=11)
H2_FILL = PatternFill("solid", fgColor="2E75B6")
H2_FONT = Font(bold=True, color="FFFFFF")
COL_FILL = PatternFill("solid", fgColor="D6E4F0")
COL_FONT = Font(bold=True)
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
GRN_FILL = PatternFill("solid", fgColor="C6EFCE")
YEL_FILL = PatternFill("solid", fgColor="FFEB9C")
ORG_FILL = PatternFill("solid", fgColor="FFD966")
GRY_FILL = PatternFill("solid", fgColor="F2F2F2")

VALID_COLORS = {
    "CONFIRMED_MULTI_SOURCE": GRN_FILL,
    "SINGLE_SOURCE":          YEL_FILL,
    "CONFLICT":               PatternFill("solid", fgColor="FFC7CE"),
}
STATUS_COLORS = {
    "MATCHED":          GRN_FILL,
    "AUTOACCEPTED":     GRN_FILL,
    "ACCEPTED":         GRN_FILL,
    "EXPERT_CONFIRMED": GRN_FILL,
    "NEW_PLAYER":       YEL_FILL,
    "REVIEW_NEEDED":    ORG_FILL,
    "UNRESOLVED":       ORG_FILL,
    "NOISE":            GRY_FILL,
}


def _col_header(ws, row, cols, fill=COL_FILL, font=COL_FONT):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws, min_w=8, max_w=45):
    for col in ws.columns:
        w = min_w
        for cell in col:
            if cell.value:
                w = max(w, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = w


def build_readme(wb):
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    rows = [
        ("Pre-1997 Footbag Competition Results",),
        ("Historical Reconstruction — v1.0",),
        ("",),
        ("SCOPE",),
        ("This workbook covers footbag competition results from 1980–1996.",),
        ("It is a historical reconstruction from limited sources and is INCOMPLETE.",),
        ("Coverage is best for major championships; many regional and club events are absent.",),
        ("",),
        ("SOURCES",),
        ("1. FBW Magazine (Volumes 2–14) — scanned pages, AI-extracted via Gemini.",),
        ("   Coverage: major events, primarily 1980–1996.",),
        ("2. IFAB History Pages — AI-extracted results from IFAB World Championships page.",),
        ("   Coverage: World Championships 1980–1996.",),
        ("3. OLD_RESULTS.txt — contributed historical text file.",),
        ("   Coverage: NHSA nationals, WFA nationals, World Championships 1982–1990.",),
        ("",),
        ("DATA PHILOSOPHY",),
        ("All source evidence is preserved. Where multiple sources cover the same event,",),
        ("both are retained. Conflicts are flagged, not auto-resolved.",),
        ("Raw player names are stored alongside resolved canonical identities.",),
        ("Unresolved identities are kept in the dataset, not silently dropped.",),
        ("",),
        ("IDENTITY STATUS",),
        ("MATCHED          — name matched exactly to the post-1997 Persons Truth database",),
        ("AUTOACCEPTED     — safe alias accepted automatically (obvious 1-char variant)",),
        ("ACCEPTED         — alias accepted by human review",),
        ("EXPERT_CONFIRMED — alias confirmed by domain expert",),
        ("NEW_PLAYER       — pre-1997 only player, not in post-1997 records",),
        ("REVIEW_NEEDED    — not yet resolved; flagged for future research",),
        ("UNRESOLVED       — no plausible match found",),
        ("",),
        ("VERSION",),
        (f"Dataset version: v1.0  |  Generated: {TODAY}",),
    ]
    ws.column_dimensions["A"].width = 80
    for i, (text,) in enumerate(rows, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif i == 2:
            cell.font = Font(bold=True, size=11, italic=True)
        elif text and text.isupper() and len(text) < 30:
            cell.font = Font(bold=True)


def build_data_notes(wb):
    ws = wb.create_sheet("DATA NOTES")
    ws.sheet_view.showGridLines = False
    rows = [
        ("Limitation", "Details"),
        ("Incomplete coverage", "Only major championships well-covered 1980-1996. Most regional/state/club events absent."),
        ("Source quality", "Magazine scans vary in legibility. AI extraction may have errors in names, divisions, or placements."),
        ("Division names", "Abbreviated or inconsistent in sources (e.g., 'Open Sgls Net' vs 'Open Singles Net'). Raw names preserved."),
        ("Identity resolution", "Names matched conservatively. Unresolved entries kept — do not assume they are different people."),
        ("Doubles teams", "Some sources list team entries; others list individuals. Team pairings reconstructed where possible."),
        ("Year coverage gap", "1993 has no confirmed results. 1990-1992 sparse. 1988 US Nationals: only 1 placement recovered."),
        ("1980 Worlds", "Only 1 placement recovered (source page very sparse)."),
        ("1981 Worlds", "Only 2 placements recovered."),
        ("1984 Euro Champs", "Only 1 placement recovered."),
        ("Multi-source events", "CONFIRMED_MULTI_SOURCE means 2+ independent sources agree. SINGLE_SOURCE means only one source found."),
        ("Merged events (1986-1989)", "WFA World Championships and World Footbag Championships confirmed as same event per domain expert."),
        ("Merged event (1994)", "IFAB World Championships (FBW source) and World Footbag Championships (IFAB source) confirmed same event."),
        ("Wigger name", "'Torben Wigger' is the expert-confirmed correct historical spelling. Post-1997 PT has 'Torbin Wigger'."),
    ]
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    _col_header(ws, 1, ["Limitation", "Details"])
    for i, (a, b) in enumerate(rows[1:], 2):
        ws.cell(i, 1, a).font = Font(bold=True)
        cell_b = ws.cell(i, 2, b)
        cell_b.alignment = Alignment(wrap_text=True)
        if i % 2 == 0:
            ws.cell(i, 1).fill = ALT_FILL
            cell_b.fill = ALT_FILL
    ws.row_dimensions[1].height = 18
    for row_idx in range(2, len(rows) + 1):
        ws.row_dimensions[row_idx].height = 30


def build_event_index(wb, events):
    ws = wb.create_sheet("EVENT INDEX")
    ws.freeze_panes = "A2"
    cols = ["ID (10-char)", "Year", "Event Name", "Location", "Type",
            "Sources", "Validation", "Placements", "Version"]
    _col_header(ws, 1, cols)
    ws.row_dimensions[1].height = 18
    for i, e in enumerate(sorted(events, key=lambda x: (x["year"], x["event_name"])), 2):
        row_vals = [
            e["canonical_event_id"],
            int(e["year"]),
            e["event_name"],
            e.get("location", ""),
            e["normalized_event_type"],
            e["source_types"],
            e["validation_status"],
            int(e["num_placements"]),
            e.get("dataset_version", VERSION),
        ]
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(i, c, val)
            if i % 2 == 0:
                cell.fill = ALT_FILL
            if c == 7:
                cell.fill = VALID_COLORS.get(val, PatternFill())
    _autowidth(ws)


def build_player_summary(wb, persons, participants):
    ws = wb.create_sheet("PLAYER SUMMARY")
    ws.freeze_panes = "A2"

    app_count = Counter()
    yr_set    = defaultdict(set)
    div_set   = defaultdict(set)
    event_yr  = {}

    # Build event→year map from participants
    for p in participants:
        pid = p["person_id"]
        if pid:
            app_count[pid] += 1
            div_set[pid].add(p["division_raw"][:20] if p["division_raw"] else "")

    cols = ["Person ID", "Canonical Name", "Scope", "Status",
            "Appearances", "Sample Division", "Raw Names Seen"]
    _col_header(ws, 1, cols)
    ws.row_dimensions[1].height = 18

    # Build raw-names-seen per person_id from aliases
    from collections import defaultdict as _dd
    raw_by_pid = _dd(list)
    # This will be populated from aliases table passed in via main
    # For now, build from participants
    raw_parts = _dd(set)
    for p in participants:
        if p["person_id"]:
            raw_parts[p["person_id"]].add(p["player_name_raw"])

    for i, p in enumerate(sorted(persons, key=lambda x: x["person_canon"]), 2):
        pid = p["person_id"]
        raw_names = "; ".join(sorted(raw_parts.get(pid, set())))
        divs = list(div_set.get(pid, set()))
        row_vals = [
            pid,
            p["person_canon"],
            p["source_scope"],
            p.get("person_status", "ACTIVE"),
            app_count.get(pid, 0),
            divs[0] if divs else "",
            raw_names[:60],
        ]
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(i, c, val)
            if i % 2 == 0:
                cell.fill = ALT_FILL
    _autowidth(ws)


def build_year_sheet(wb, year, events, results, participants):
    ws = wb.create_sheet(str(year))
    ws.sheet_view.showGridLines = False

    year_events = [e for e in events if e["year"] == str(year)]
    if not year_events:
        ws.cell(1, 1, f"No events recorded for {year}")
        return

    # Index results and participants
    parts_by_result = defaultdict(list)
    for p in participants:
        parts_by_result[p["result_id"]].append(p)

    current_row = 1
    for event in sorted(year_events, key=lambda e: e["event_name"]):
        eid = event["canonical_event_id"]

        # Event title
        cell = ws.cell(current_row, 1, f"{event['event_name']}")
        cell.fill = H1_FILL
        cell.font = H1_FONT
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=7)
        current_row += 1

        # Metadata bar
        meta_parts = []
        if event.get("location"):
            meta_parts.append(f"Location: {event['location']}")
        meta_parts.append(f"Sources: {event['source_types']}")
        meta_parts.append(f"ID: {eid}")
        meta_parts.append(event["validation_status"])
        cell = ws.cell(current_row, 1, "  |  ".join(meta_parts))
        cell.fill = H2_FILL
        cell.font = Font(color="FFFFFF", size=9, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=7)
        current_row += 1

        # Group results by division_raw
        event_results = [r for r in results if r["canonical_event_id"] == eid]
        divs: dict = defaultdict(list)
        for r in event_results:
            divs[r["division_raw"]].append(r)

        for div_raw, div_results in sorted(divs.items()):
            # Division header
            cell = ws.cell(current_row, 1, div_raw or "(unspecified division)")
            cell.fill = H2_FILL
            cell.font = H2_FONT
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=7)
            current_row += 1

            # Column headers
            for c, h in enumerate(["Place", "Player(s)", "Canon Name",
                                    "Status", "Source", "Source File", "Person ID"], 1):
                cell = ws.cell(current_row, c, h)
                cell.fill = GRY_FILL
                cell.font = Font(bold=True, size=9)
            current_row += 1

            for res in sorted(div_results,
                               key=lambda x: int(x["place"]) if x["place"].isdigit() else 999):
                pparts = parts_by_result.get(res["result_id"], [])
                players_raw = " / ".join(p["player_name_raw"] for p in pparts) or res.get("player_raw", "?")
                canon_names = " / ".join(p["person_canon"] for p in pparts if p["person_canon"])
                statuses    = " / ".join(p["resolution_status"] for p in pparts)
                primary_st  = pparts[0]["resolution_status"] if pparts else "UNRESOLVED"
                src_type    = res.get("source_type", "")
                src_file    = res.get("source_event_id", "")
                pid_disp    = " / ".join(p["person_id"] for p in pparts if p["person_id"])

                place_val = int(res["place"]) if res["place"].isdigit() else res["place"]
                fill = STATUS_COLORS.get(primary_st, PatternFill())
                for c, val in enumerate([place_val, players_raw, canon_names,
                                         statuses, src_type, src_file, pid_disp], 1):
                    cell = ws.cell(current_row, c, val)
                    cell.fill = fill
                    cell.font = Font(size=9)
                current_row += 1

        current_row += 1

    # Column widths
    for col_letter, width in [("A", 8), ("B", 30), ("C", 28),
                               ("D", 20), ("E", 14), ("F", 14), ("G", 38)]:
        ws.column_dimensions[col_letter].width = width


def build_validation_sheet(wb, events, participants, aliases):
    ws = wb.create_sheet("VALIDATION SUMMARY")
    ws.sheet_view.showGridLines = False

    val_counts    = Counter(e["validation_status"] for e in events)
    status_counts = Counter(p["resolution_status"] for p in participants)
    alias_counts  = Counter(a["alias_status"] for a in aliases)

    row = 1
    ws.cell(row, 1, "Event Validation").font = Font(bold=True, size=12)
    row += 2
    _col_header(ws, row, ["Status", "Count"])
    row += 1
    for status, cnt in sorted(val_counts.items()):
        ws.cell(row, 1, status).fill = VALID_COLORS.get(status, PatternFill())
        ws.cell(row, 2, cnt)
        row += 1

    row += 1
    ws.cell(row, 1, "Participant Identity Resolution").font = Font(bold=True, size=12)
    row += 2
    _col_header(ws, row, ["Status", "Count"])
    row += 1
    for status in ("MATCHED", "AUTOACCEPTED", "ACCEPTED", "EXPERT_CONFIRMED",
                   "NEW_PLAYER", "REVIEW_NEEDED", "UNRESOLVED", "NOISE"):
        cnt = status_counts.get(status, 0)
        if cnt:
            ws.cell(row, 1, status).fill = STATUS_COLORS.get(status, PatternFill())
            ws.cell(row, 2, cnt)
            row += 1

    row += 1
    ws.cell(row, 1, "Alias Status").font = Font(bold=True, size=12)
    row += 2
    _col_header(ws, row, ["Status", "Count"])
    row += 1
    for status, cnt in sorted(alias_counts.items()):
        ws.cell(row, 1, status)
        ws.cell(row, 2, cnt)
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12


def build_workbook(events, results, participants, persons, aliases):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_readme(wb)
    build_data_notes(wb)
    build_event_index(wb, events)
    build_player_summary(wb, persons, participants)
    for year in range(1980, 1997):
        build_year_sheet(wb, year, events, results, participants)
    build_validation_sheet(wb, events, participants, aliases)

    out_path = EARLY / "out" / "footbag_results_pre1997_v1.xlsx"
    wb.save(out_path)
    print(f"  Saved workbook → {out_path.relative_to(ROOT)}")
    return out_path


# ── PART 8 — Consistency check ────────────────────────────────────────────────

def consistency_check(events, results, participants, persons, aliases):
    """Run basic integrity checks and return (warnings, errors)."""
    warnings, errors = [], []
    event_ids  = {e["canonical_event_id"] for e in events}
    person_ids = {p["person_id"] for p in persons}

    # Year separation
    for e in events:
        if int(e["year"]) >= 1997:
            errors.append(f"POST-1997 event in pre-1997 dataset: {e['canonical_event_id']} {e['year']}")

    # Orphan results
    orphan_results = {r["canonical_event_id"] for r in results if r["canonical_event_id"] not in event_ids}
    for eid in orphan_results:
        errors.append(f"Result with unknown canonical_event_id: {eid}")

    # Orphan participants
    orphan_parts = {p["canonical_event_id"] for p in participants if p["canonical_event_id"] not in event_ids}
    for eid in orphan_parts:
        errors.append(f"Participant with unknown canonical_event_id: {eid}")

    # Orphan person_ids in participants (allow empty = unresolved)
    for p in participants:
        pid = p.get("person_id", "")
        if pid and pid not in person_ids:
            errors.append(f"Participant has unknown person_id: {pid} (name: {p['player_name_raw']})")

    # Alias → person_id consistency
    alias_pid_set = {a["person_id"] for a in aliases if a["person_id"]}
    orphan_alias_pids = alias_pid_set - person_ids
    for pid in sorted(orphan_alias_pids):
        warnings.append(f"Alias points to person_id not in persons table: {pid}")

    # Duplicate event IDs
    dup_events = [eid for eid, cnt in Counter(e["canonical_event_id"] for e in events).items() if cnt > 1]
    for eid in dup_events:
        errors.append(f"Duplicate canonical_event_id: {eid}")

    return warnings, errors


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    print(f"=== 11_finalize_pre1997.py — PRE-1997 {VERSION} ===\n")

    # Load V2 canonical state
    events       = read_csv(CANONICAL / "events_pre1997.csv")
    results      = read_csv(CANONICAL / "event_results_pre1997.csv")
    participants = read_csv(CANONICAL / "event_result_participants_pre1997.csv")
    persons      = read_csv(CANONICAL / "persons_pre1997.csv")
    disciplines  = read_csv(CANONICAL / "event_disciplines_pre1997.csv")
    aliases      = read_csv(CANONICAL / "person_aliases_pre1997.csv")

    print(f"Loaded (V2 baseline):")
    print(f"  Events:       {len(events)}")
    print(f"  Results:      {len(results)}")
    print(f"  Participants: {len(participants)}")
    print(f"  Persons:      {len(persons)}")
    print(f"  Aliases:      {len(aliases)}")

    # --- Part 1: Expert corrections ---
    print("\n--- PART 1: Expert-confirmed identity corrections ---")
    aliases, persons, participants, review_log = apply_expert_corrections(
        aliases, persons, participants)
    print(f"  Changes applied: {len(review_log)}")
    for entry in review_log:
        print(f"    [{entry['action']}] {entry['raw_name']} → {entry['canonical_display']}")
        if entry["note"]:
            print(f"      Note: {entry['note']}")

    # --- Part 2: Year separation check ---
    print("\n--- PART 2: Year separation verification ---")
    violations = verify_year_separation(events, results, participants)
    if violations:
        for v in violations:
            print(f"  ERROR: {v}")
    else:
        print(f"  OK — all {len(events)} events have year < 1997")

    # --- Part 3: Write final CSVs ---
    print("\n--- PART 3: Write final_pre1997/ canonical CSVs ---")
    write_final_csvs(events, results, participants, persons, disciplines, aliases)

    # --- Part 4: Copy review/unresolved files ---
    print("\n--- PART 4: Copy review/unresolved files ---")
    copy_review_files()

    # --- Part 5: Validation summary ---
    print("\n--- PART 5: Validation summary ---")
    write_validation_summary(events, results, participants, persons, aliases)

    # --- Part 5b: Identity review log ---
    write_identity_review_log(review_log)

    # --- Part 6: Workbook ---
    print("\n--- PART 6: Build footbag_results_pre1997_v1.xlsx ---")
    build_workbook(events, results, participants, persons, aliases)

    # --- Part 7: Consistency check ---
    print("\n--- PART 7: Consistency check ---")
    warnings, errors = consistency_check(events, results, participants, persons, aliases)
    for w in warnings:
        print(f"  WARN: {w}")
    for e in errors:
        print(f"  ERROR: {e}")
    if not warnings and not errors:
        print("  All checks passed.")

    # --- Final summary ---
    val_counts    = Counter(e["validation_status"] for e in events)
    status_counts = Counter(p["resolution_status"] for p in participants)
    print(f"\n{'='*56}")
    print(f"PRE-1997 {VERSION} FINALIZATION COMPLETE")
    print(f"{'='*56}")
    print(f"\nEvents:             {len(events)}")
    print(f"  CONFIRMED_MULTI_SOURCE: {val_counts.get('CONFIRMED_MULTI_SOURCE', 0)}")
    print(f"  SINGLE_SOURCE:          {val_counts.get('SINGLE_SOURCE', 0)}")
    print(f"  CONFLICT:               {val_counts.get('CONFLICT', 0)}")
    print(f"Placements:         {len(results)}")
    print(f"Participants:       {len(participants)}")
    print(f"  Resolved:         {sum(status_counts.get(s, 0) for s in ('MATCHED','AUTOACCEPTED','ACCEPTED','EXPERT_CONFIRMED','NEW_PLAYER'))}")
    print(f"  Unresolved:       {sum(status_counts.get(s, 0) for s in ('REVIEW_NEEDED','UNRESOLVED'))}")
    print(f"Persons:            {len(persons)}")
    print(f"  POST1997 refs:    {sum(1 for p in persons if p['source_scope']=='POST1997')}")
    print(f"  PRE1997_ONLY:     {sum(1 for p in persons if p['source_scope']=='PRE1997_ONLY')}")
    print(f"\nOutput directory:   early_data/final_pre1997/")
    print(f"Spreadsheet:        early_data/out/footbag_results_pre1997_v1.xlsx")
    if errors:
        print(f"\n⚠  {len(errors)} consistency error(s) found — review above")
    if warnings:
        print(f"⚠  {len(warnings)} warning(s) — review above")
    print(f"\nDone.")


if __name__ == "__main__":
    main()
