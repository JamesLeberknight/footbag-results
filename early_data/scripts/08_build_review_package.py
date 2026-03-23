#!/usr/bin/env python3
"""
08_build_review_package.py — Build human review files for person aliases
and event group resolution.

Reads:
  early_data/identity/person_aliases_needs_review.csv
  early_data/canonical/event_source_comparison.csv
  early_data/canonical/canonical_events.csv
  early_data/canonical/event_result_participants_pre1997.csv
  early_data/canonical/event_disciplines_pre1997.csv
  early_data/canonical/events_pre1997.csv

Produces:
  early_data/review/person_alias_resolution.csv
  early_data/review/review_aliases.xlsx
  early_data/review/event_group_resolution.csv
  early_data/review/review_event_groups.xlsx

Does NOT modify any canonical outputs.
"""

import csv
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
ED = REPO_ROOT / "early_data"

# Inputs
NEEDS_REVIEW_CSV = ED / "identity"  / "person_aliases_needs_review.csv"
EV_COMPARISON    = ED / "canonical" / "event_source_comparison.csv"
CANONICAL_EV     = ED / "canonical" / "canonical_events.csv"
PARTICIPANTS     = ED / "canonical" / "event_result_participants_pre1997.csv"
DISCIPLINES      = ED / "canonical" / "event_disciplines_pre1997.csv"
EVENTS_PRE97     = ED / "canonical" / "events_pre1997.csv"

# Outputs
REVIEW_DIR = ED / "review"
ALIAS_CSV  = REVIEW_DIR / "person_alias_resolution.csv"
ALIAS_XLSX = REVIEW_DIR / "review_aliases.xlsx"
EVENT_CSV  = REVIEW_DIR / "event_group_resolution.csv"
EVENT_XLSX = REVIEW_DIR / "review_event_groups.xlsx"

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

BLUE_FILL    = PatternFill("solid", fgColor="4472C4")
YELLOW_FILL  = PatternFill("solid", fgColor="FFD966")
GREEN_FILL   = PatternFill("solid", fgColor="E2EFDA")
ORANGE_FILL  = PatternFill("solid", fgColor="F4B942")
GREY_FILL    = PatternFill("solid", fgColor="F2F2F2")
LIGHT_FILL   = PatternFill("solid", fgColor="DEEAF1")

WHITE_BOLD   = Font(bold=True, color="FFFFFF")
DARK_BOLD    = Font(bold=True, color="1F3864")
RED_FONT     = Font(color="C00000")
ITALIC_GREY  = Font(italic=True, color="595959", size=9)

THIN_BORDER = Border(
    bottom=Side(style="thin", color="BFBFBF"),
)


def _hdr(ws, row: int, values: list, fill=None, font=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        if fill: c.fill = fill
        if font: c.font = font
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def _col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_csv(path: Path, fields: list, rows: list):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


# ---------------------------------------------------------------------------
# Part 1: Person alias review
# ---------------------------------------------------------------------------

# Context notes added per case — what to look for when reviewing source
REVIEW_CONTEXT = {
    "Jim Caveney": {
        "source_evidence": "Appears in OLD_RESULTS.txt section 1 (1983 NHSA, 1984 WFA) and section 2 (1983 NHSA, 1983 WFA, 1984, 1985). PT has 'Jimmy Caveney' — a rename to 'Jim Caveney' was planned (PT v44→v45) but may not have been applied to this PT version.",
        "suggested_action": "Check PT version. If rename was applied, ACCEPT. If not, verify source image before accepting.",
        "risk_if_wrong": "LOW — Jim and Jimmy Caveney almost certainly the same person.",
    },
    "Jody Grace": {
        "source_evidence": "Appears in FBW placements. PT has 'Judy Grace'. Could be OCR error (Jody/Judy) or genuinely different person.",
        "suggested_action": "Check source scan (FBW page image) for legibility of first name.",
        "risk_if_wrong": "MEDIUM — Jody and Judy are distinct names.",
    },
    "Jenny Davidson": {
        "source_evidence": "Appears in FBW placements (1983 Women's Doubles Net). PT has 'Jenny Davison' (single 'o' vs double 'o'). Single-char difference in Scandinavian/British surname.",
        "suggested_action": "Check source scan. Davidson and Davison are distinct surnames in some contexts.",
        "risk_if_wrong": "LOW — single char, same first name, same era, same division type.",
    },
    "Ken Shults": {
        "source_evidence": "Appears in OLD_RESULTS.txt multiple times. PT has 'Kenneth Shults' with alias 'Kenny Schults' — note different spelling of alias (Schults, not Shults). 'Ken' is not registered as an alias.",
        "suggested_action": "If Kenneth Shults is confirmed = Ken Shults, add 'Ken Shults' as alias. Verify spelling of surname in source.",
        "risk_if_wrong": "LOW — Ken is standard short form of Kenneth.",
    },
    "Kenny Shults": {
        "source_evidence": "Appears in OLD_RESULTS.txt. PT alias is 'Kenny Schults' (different spelling: Schults). This entry has 'Shults'. Could be a different spelling variant or same person.",
        "suggested_action": "Confirm which spelling is correct. If 'Shults' is correct, update PT alias from 'Kenny Schults' to 'Kenny Shults' and ACCEPT.",
        "risk_if_wrong": "LOW — but PT alias spelling conflict must be resolved first.",
    },
    "Karin Atogpian": {
        "source_evidence": "Appears in FBW placements. 'Karen Atgopian' also appears in FBW placements as a separate entry (both are unresolved). Possible alternate spelling of the same Scandinavian/Eastern European name.",
        "suggested_action": "Check whether 'Karin' and 'Karen' appear in the same event or different events. If same event, likely two different people. If different events, could be the same person.",
        "risk_if_wrong": "MEDIUM — Karin and Karen are distinct given names in some cultures.",
    },
    "Lori Jean Tarr": {
        "source_evidence": "Appears in FBW placements (1984–1985 events). PT has 'Lori Jean Conover'. Different surname — could be married vs maiden name, or completely different person.",
        "suggested_action": "Cannot accept without external evidence of name change. Search for 'Tarr' in source scans and external records.",
        "risk_if_wrong": "HIGH — different surnames. Must not auto-merge.",
    },
    "Tim Fitzgerald": {
        "source_evidence": "Appears in FBW placements. PT has 'Jim Fitzgerald'. Completely different first name — almost certainly an OCR error (T vs J are not visually similar, but 'Tim' and 'Jim' can be confused in handwriting/print).",
        "suggested_action": "Check source scan image for this placement. If scan shows 'Jim', ACCEPT. If genuinely 'Tim', CREATE_NEW.",
        "risk_if_wrong": "MEDIUM — Tim and Jim are distinct people, but OCR errors are common.",
    },
    "Tobin Wigger": {
        "source_evidence": "Appears in FBW placements. PT has 'Torbin Wigger'. Tobin/Torbin differ by one char (missing 'r'). These are Scandinavian name forms.",
        "suggested_action": "Check source scan for spelling. 'Torbin' is the established PT form.",
        "risk_if_wrong": "LOW — single char difference, same surname, same era.",
    },
    "Torben Wigger": {
        "source_evidence": "Appears in FBW placements. PT has 'Torbin Wigger'. Torben is a distinct Danish/Norwegian name from Torbin. Could be the same person or a relation.",
        "suggested_action": "Verify source scan. 'Torben' is a common Scandinavian name; 'Torbin' may be a transcription of the same person.",
        "risk_if_wrong": "LOW — same surname, similar era.",
    },
    "Steve Fennell": {
        "source_evidence": "Appears in FBW placements. PT has 'Steve Femmel'. The difference is 'nnell' vs 'mmel' — this is more than a single-char OCR error. Could be a distinct person or a badly transcribed name.",
        "suggested_action": "Check source scan for the exact spelling. If the source genuinely reads 'Fennell', consider CREATE_NEW rather than ACCEPT.",
        "risk_if_wrong": "MEDIUM — not a trivial spelling variant.",
    },
    "Ted Martens": {
        "source_evidence": "Appears in FBW placements. PT has 'Ted Martin'. Martens vs Martin differ in the ending. Could be OCR adding an 's', or genuinely different person.",
        "suggested_action": "Check source scan. Martens is a distinct surname (common in Dutch/Flemish contexts).",
        "risk_if_wrong": "MEDIUM — Martens and Martin are distinct surnames.",
    },
}


def build_alias_review(needs_review_rows, participants_rows):
    """
    Build person_alias_resolution.csv with context and blank DECISION column.
    """
    # Count appearances of each raw name in participants
    name_appearances: dict = defaultdict(lambda: {"count": 0, "events": set(),
                                                    "years": set(), "sources": set()})
    for p in participants_rows:
        raw = p["player_name_raw"]
        if p["resolution_status"] == "REVIEW_NEEDED" and raw:
            name_appearances[raw]["count"]  += 1
            name_appearances[raw]["events"].add(p["canonical_event_id"])
            name_appearances[raw]["sources"].add(p["source_type"])

    rows = []
    for nr in needs_review_rows:
        raw     = nr["raw_name"]
        cand_id = nr["candidate_person_ids"]
        cand_nm = nr["candidate_person_names"]
        ctx     = REVIEW_CONTEXT.get(raw, {})
        app     = name_appearances.get(raw, {"count": 0, "events": set(), "sources": set()})

        rows.append({
            "raw_name":              raw,
            "candidate_person_id":   cand_id,
            "candidate_person_name": cand_nm,
            "appearances_in_data":   app["count"],
            "events_seen":           "; ".join(sorted(app["events"])[:4]),
            "source_types_seen":     "; ".join(sorted(app["sources"])),
            "review_notes":          nr["review_notes"],
            "source_evidence":       ctx.get("source_evidence", ""),
            "suggested_action":      ctx.get("suggested_action", ""),
            "risk_if_wrong":         ctx.get("risk_if_wrong", ""),
            "DECISION":              "",   # human fills in: ACCEPT | CREATE_NEW | REJECT | DEFER
            "DECISION_NOTES":        "",
        })
    return rows


def build_alias_xlsx(wb, alias_rows):
    ws = wb.create_sheet("Alias Review")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [22, 38, 22, 8, 32, 16, 50, 70, 60, 12, 14, 40])

    r = 1
    # Title
    ws.cell(r, 1, "PERSON ALIAS REVIEW — 12 items requiring human decision").font = Font(bold=True, size=13)
    ws.merge_cells(f"A{r}:L{r}")
    ws.cell(r, 1).fill = BLUE_FILL
    ws.cell(r, 1).font = Font(bold=True, size=13, color="FFFFFF")
    r += 1

    # Instructions
    instr = ("For each row, verify the raw name against source images, then record a DECISION: "
             "ACCEPT (map to candidate) | CREATE_NEW (new early player) | REJECT (wrong candidate) | DEFER (needs more research)")
    ws.cell(r, 1, instr).font = ITALIC_GREY
    ws.merge_cells(f"A{r}:L{r}")
    ws.row_dimensions[r].height = 30
    r += 2

    headers = [
        "raw_name", "candidate_person_name", "candidate_person_id",
        "appearances", "events_seen", "sources",
        "review_notes", "source_evidence", "suggested_action", "risk",
        "DECISION ▼", "DECISION_NOTES",
    ]
    r = _hdr(ws, r, headers, fill=BLUE_FILL, font=WHITE_BOLD)
    ws.freeze_panes = f"A{r}"

    field_map = [
        "raw_name", "candidate_person_name", "candidate_person_id",
        "appearances_in_data", "events_seen", "source_types_seen",
        "review_notes", "source_evidence", "suggested_action", "risk_if_wrong",
        "DECISION", "DECISION_NOTES",
    ]

    for row_data in alias_rows:
        risk = row_data.get("risk_if_wrong", "")
        fill = ORANGE_FILL if "HIGH" in risk else (YELLOW_FILL if "MEDIUM" in risk else GREEN_FILL)
        for col, fld in enumerate(field_map, 1):
            c = ws.cell(r, col, row_data.get(fld, ""))
            c.alignment = Alignment(wrap_text=True, vertical="top")
        # Highlight DECISION column
        ws.cell(r, 11).fill = LIGHT_FILL
        ws.cell(r, 11).font = Font(bold=True)
        # Risk column color
        ws.cell(r, 10).fill = fill
        ws.row_dimensions[r].height = 60
        r += 1

    # Decision key
    r += 2
    ws.cell(r, 1, "DECISION KEY").font = DARK_BOLD
    r += 1
    for code, meaning in [
        ("ACCEPT",     "Map raw_name to candidate PT person (will be added as alias)"),
        ("CREATE_NEW", "Treat as a new pre-1997-only early player — do not link to candidate"),
        ("REJECT",     "The candidate is wrong; leave completely unresolved"),
        ("DEFER",      "Cannot decide yet; flag for future research"),
    ]:
        ws.cell(r, 1, code).font = Font(bold=True)
        ws.cell(r, 2, meaning)
        r += 1


# ---------------------------------------------------------------------------
# Part 2: Event group review
# ---------------------------------------------------------------------------

# Explicit review questions per group or type
GROUP_QUESTIONS = {
    # Same-year WORLD_CHAMPIONSHIPS vs WFA_WORLD_CHAMPIONSHIPS
    "1986": "WORLD_CHAMPIONSHIPS (FBW '1986 World Footbag Championships') vs WFA_WORLD_CHAMPIONSHIPS (FBW+OLD '1986 WFA World Championships') — are these the SAME event or parallel events?",
    "1987": "WORLD_CHAMPIONSHIPS (FBW '1987 World Footbag Championships', 3 placements) vs WFA_WORLD_CHAMPIONSHIPS (FBW '1987 WFA World Championships', 44 placements) — same event?",
    "1988": "WORLD_CHAMPIONSHIPS (FBW, 2 placements) vs WFA_WORLD_CHAMPIONSHIPS (44 placements). Also US_NATIONALS (1 placement) — complete?",
    "1989": "WORLD_CHAMPIONSHIPS (IFAB, 1 placement) vs WFA_WORLD_CHAMPIONSHIPS (44 placements) — same event, different source coverage?",
    "1994": "IFAB_WORLD_CHAMPIONSHIPS (FBW, 44 placements) vs WORLD_CHAMPIONSHIPS (IFAB, 1 placement) — almost certainly the same event. Should these share a canonical ID?",
}

# Same-year event pairs worth flagging
SAME_YEAR_PAIRS = {
    "1986": [("WFA_WORLD_CHAMPIONSHIPS", "WORLD_CHAMPIONSHIPS")],
    "1987": [("WFA_WORLD_CHAMPIONSHIPS", "WORLD_CHAMPIONSHIPS")],
    "1988": [("WFA_WORLD_CHAMPIONSHIPS", "WORLD_CHAMPIONSHIPS")],
    "1989": [("WFA_WORLD_CHAMPIONSHIPS", "WORLD_CHAMPIONSHIPS")],
    "1994": [("IFAB_WORLD_CHAMPIONSHIPS", "WORLD_CHAMPIONSHIPS")],
}


def build_event_review(comparison_rows, events_rows, disciplines_rows, participants_rows):
    """Build event_group_resolution.csv with context and blank DECISION column."""

    # Build lookups
    ev_by_id  = {r["canonical_event_id"]: r for r in events_rows}
    disc_count = defaultdict(int)
    for d in disciplines_rows:
        disc_count[d["canonical_event_id"]] += 1
    plc_count = defaultdict(int)
    for p in participants_rows:
        plc_count[p["canonical_event_id"]] += 1

    # Index same-year events
    by_year_type: dict = defaultdict(list)
    for e in events_rows:
        by_year_type[e["year"]].append((e["normalized_event_type"], e["canonical_event_id"]))

    rows = []
    for cmp in comparison_rows:
        ceid  = cmp["group_id"]
        year  = cmp["year"]
        ntype = cmp["normalized_event_type"]
        vstatus = cmp["validation_status"]
        sources = cmp["sources"]
        ev    = ev_by_id.get(ceid, {})

        # Detect same-year events that might be the same real-world event
        same_year_events = [
            f"{t} ({eid[:8]})"
            for t, eid in by_year_type.get(year, [])
            if eid != ceid
        ]
        potential_merge = "; ".join(same_year_events) if same_year_events else ""

        # Determine review question
        question = ""
        if vstatus == "CONFIRMED_MULTI_SOURCE":
            question = "Verify: are both sources describing the same real-world event?"
        elif vstatus == "SINGLE_SOURCE" and potential_merge:
            question = f"Same year has other events ({ntype} this group). Could any be the same event?"
        elif vstatus == "SINGLE_SOURCE":
            question = "Single source only. Are additional placements available elsewhere?"

        # Override with specific known questions
        year_pairs = SAME_YEAR_PAIRS.get(year, [])
        for t1, t2 in year_pairs:
            if ntype in (t1, t2):
                question = GROUP_QUESTIONS.get(year, question)
                break

        rows.append({
            "canonical_event_id":    ceid,
            "year":                  year,
            "normalized_event_type": ntype,
            "event_name":            ev.get("event_name", ""),
            "location":              ev.get("location", ""),
            "validation_status":     vstatus,
            "num_sources":           cmp["num_sources"],
            "source_types":          cmp["source_types"],
            "sources":               sources,
            "num_disciplines":       disc_count.get(ceid, 0),
            "num_placements":        plc_count.get(ceid, 0),
            "other_events_same_year": potential_merge,
            "review_question":       question,
            "DECISION":              "",  # CORRECT | MERGE_WITH:<id> | SPLIT | NEEDS_DATA | DEFER
            "DECISION_NOTES":        "",
        })

    return sorted(rows, key=lambda x: (x["year"], x["normalized_event_type"]))


def build_event_xlsx(wb, event_rows):
    ws = wb.create_sheet("Event Groups")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, [14, 6, 26, 38, 22, 22, 8, 14, 50, 8, 10, 44, 60, 16, 40])

    r = 1
    ws.cell(r, 1, "EVENT GROUP REVIEW — 37 canonical event groups").font = Font(bold=True, size=13)
    ws.merge_cells(f"A{r}:O{r}")
    ws.cell(r, 1).fill = BLUE_FILL
    ws.cell(r, 1).font = Font(bold=True, size=13, color="FFFFFF")
    r += 1

    instr = ("Review each group. CONFIRMED_MULTI_SOURCE groups need source verification. "
             "Pay special attention to same-year events that may be the same real-world event "
             "(e.g., WFA World Championships and World Footbag Championships in the same year).")
    ws.cell(r, 1, instr).font = ITALIC_GREY
    ws.merge_cells(f"A{r}:O{r}")
    ws.row_dimensions[r].height = 30
    r += 2

    headers = [
        "canonical_event_id", "year", "normalized_event_type", "event_name",
        "location", "validation_status", "num_src", "source_types", "sources",
        "divs", "plc", "other_events_same_year", "review_question",
        "DECISION ▼", "DECISION_NOTES",
    ]
    r = _hdr(ws, r, headers, fill=BLUE_FILL, font=WHITE_BOLD)
    ws.freeze_panes = f"A{r}"

    field_map = [
        "canonical_event_id", "year", "normalized_event_type", "event_name",
        "location", "validation_status", "num_sources", "source_types", "sources",
        "num_disciplines", "num_placements", "other_events_same_year",
        "review_question", "DECISION", "DECISION_NOTES",
    ]

    prev_year = None
    for row_data in event_rows:
        yr = row_data["year"]
        vstatus = row_data["validation_status"]
        has_merge_q = bool(row_data.get("other_events_same_year"))

        if yr != prev_year:
            # Year separator row
            ws.cell(r, 1, f"  {yr}").font = Font(bold=True, color="1F3864")
            ws.merge_cells(f"A{r}:O{r}")
            ws.cell(r, 1).fill = GREY_FILL
            ws.row_dimensions[r].height = 18
            r += 1
            prev_year = yr

        for col, fld in enumerate(field_map, 1):
            c = ws.cell(r, col, row_data.get(fld, ""))
            c.alignment = Alignment(wrap_text=True, vertical="top")

        # Color-code validation status
        if vstatus == "CONFIRMED_MULTI_SOURCE":
            ws.cell(r, 6).fill = GREEN_FILL
        elif has_merge_q:
            ws.cell(r, 6).fill = ORANGE_FILL

        # Highlight DECISION column
        ws.cell(r, 14).fill = LIGHT_FILL
        ws.cell(r, 14).font = Font(bold=True)
        ws.row_dimensions[r].height = 45
        r += 1

    # Decision key
    r += 2
    ws.cell(r, 1, "DECISION KEY").font = DARK_BOLD
    r += 1
    for code, meaning in [
        ("CORRECT",        "Group is correct as-is — no changes needed"),
        ("MERGE_WITH:<id>", "This group should be merged with the specified canonical_event_id"),
        ("SPLIT",          "This group contains events that should be separated"),
        ("NEEDS_DATA",     "Group is correct but additional placement data should be sourced"),
        ("DEFER",          "Cannot decide yet"),
    ]:
        ws.cell(r, 1, code).font = Font(bold=True)
        ws.cell(r, 2, meaning)
        r += 1

    # Highlight CONFIRMED_MULTI_SOURCE summary
    r += 2
    ws.cell(r, 1, "CONFIRMED_MULTI_SOURCE events (these need source-agreement verification):").font = DARK_BOLD
    r += 1
    for row_data in event_rows:
        if row_data["validation_status"] == "CONFIRMED_MULTI_SOURCE":
            ws.cell(r, 1, row_data["year"])
            ws.cell(r, 2, row_data["normalized_event_type"])
            ws.cell(r, 3, row_data["sources"])
            r += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=== 08_build_review_package.py ===\n")

    # Load
    needs_review  = list(csv.DictReader(open(NEEDS_REVIEW_CSV)))
    comparison    = list(csv.DictReader(open(EV_COMPARISON)))
    canon_events  = list(csv.DictReader(open(CANONICAL_EV)))
    participants  = list(csv.DictReader(open(PARTICIPANTS)))
    disciplines   = list(csv.DictReader(open(DISCIPLINES)))
    events_pre97  = list(csv.DictReader(open(EVENTS_PRE97)))

    print(f"Loaded: {len(needs_review)} review aliases, {len(comparison)} event groups")

    # Part 1: alias review
    alias_rows = build_alias_review(needs_review, participants)
    ALIAS_CSV_FIELDS = [
        "raw_name", "candidate_person_name", "candidate_person_id",
        "appearances_in_data", "events_seen", "source_types_seen",
        "review_notes", "source_evidence", "suggested_action", "risk_if_wrong",
        "DECISION", "DECISION_NOTES",
    ]
    write_csv(ALIAS_CSV, ALIAS_CSV_FIELDS, alias_rows)
    print(f"Wrote: {ALIAS_CSV.name}  ({len(alias_rows)} rows)")

    wb_alias = openpyxl.Workbook()
    wb_alias.remove(wb_alias.active)
    build_alias_xlsx(wb_alias, alias_rows)
    wb_alias.save(ALIAS_XLSX)
    print(f"Wrote: {ALIAS_XLSX.name}")

    # Part 2: event group review
    event_rows = build_event_review(comparison, events_pre97, disciplines, participants)
    EVENT_CSV_FIELDS = [
        "canonical_event_id", "year", "normalized_event_type", "event_name",
        "location", "validation_status", "num_sources", "source_types", "sources",
        "num_disciplines", "num_placements", "other_events_same_year",
        "review_question", "DECISION", "DECISION_NOTES",
    ]
    write_csv(EVENT_CSV, EVENT_CSV_FIELDS, event_rows)
    print(f"Wrote: {EVENT_CSV.name}  ({len(event_rows)} rows)")

    wb_ev = openpyxl.Workbook()
    wb_ev.remove(wb_ev.active)
    build_event_xlsx(wb_ev, event_rows)
    wb_ev.save(EVENT_XLSX)
    print(f"Wrote: {EVENT_XLSX.name}")

    # Summary
    print("\n" + "=" * 55)
    print("REVIEW PACKAGE SUMMARY")
    print("=" * 55)
    print(f"\nPerson alias review: {len(alias_rows)} items")
    by_risk = defaultdict(int)
    for r in alias_rows:
        by_risk[r.get("risk_if_wrong", "UNKNOWN")] += 1
    for risk, count in sorted(by_risk.items()):
        print(f"  {risk:8s}: {count}")

    print(f"\nEvent group review: {len(event_rows)} items")
    by_vstatus = defaultdict(int)
    has_merge_q = 0
    for r in event_rows:
        by_vstatus[r["validation_status"]] += 1
        if r.get("other_events_same_year"):
            has_merge_q += 1
    for s, c in sorted(by_vstatus.items()):
        print(f"  {s}: {c}")
    print(f"  Events with possible-merge questions: {has_merge_q}")

    print("\nDone.")


if __name__ == "__main__":
    main()
