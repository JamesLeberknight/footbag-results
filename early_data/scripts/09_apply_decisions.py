"""
09_apply_decisions.py — Apply human review decisions → PRE1997 V2

Decisions applied (verbally confirmed by user 2026-03-23):

PERSON ALIASES — all 12 ACCEPT:
  Jim Caveney       → Jimmy Caveney    (df329352)
  Jody Grace        → Judy Grace       (73728c19)
  Jenny Davidson    → Jenny Davison    (3f48caf5)
  Ken Shults        → Kenneth Shults   (2a6a7c9e)
  Kenny Shults      → Kenneth Shults   (2a6a7c9e)
  Karin Atogpian    → Karen Atgopian   (86de3135, PRE1997_ONLY)
  Lori Jean Tarr    → Lori Jean Conover (77f0f32d)
  Tim Fitzgerald    → Jim Fitzgerald   (b54020bc)
  Tobin Wigger      → Torbin Wigger    (f38e2f8e)
  Torben Wigger     → Torbin Wigger    (f38e2f8e)
  Steve Fennell     → Steve Femmel     (fb605451)
  Ted Martens       → Ted Martin       (3d1b24de)

EVENT MERGES — 5 same-year pairs (keep WFA/IFAB type, absorb WORLD_CHAMPIONSHIPS):
  1986: keep b4b6a194e2 (WFA_WORLD), absorb dea8d6f019 (WORLD)
  1987: keep 8667de3590 (WFA_WORLD), absorb 8a1c282011 (WORLD)
  1988: keep d272be24f7 (WFA_WORLD), absorb 1398fe01e4 (WORLD)
  1989: keep 334c632f6e (WFA_WORLD), absorb 6ce90a8f5d (WORLD)
  1994: keep b9d35c4646 (IFAB_WORLD), absorb 9bced09b82 (WORLD)
"""

import csv
import os
import sys
import shutil
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
CANONICAL = ROOT / "canonical"
INPUTS = ROOT.parent / "inputs"
OUT = ROOT / "out"

PT_FILE = INPUTS / "Persons_Truth.csv"

# ── decisions ──────────────────────────────────────────────────────────────

# raw_name_lower → (person_id, person_canon, resolution_status)
# resolution_status: MATCHED for PT persons, NEW_PLAYER for PRE1997_ONLY
ALIAS_ACCEPTS = {
    "jim caveney":     ("df329352-6f3b-5e98-b23b-1af6737d100b", "Jimmy Caveney",     "MATCHED"),
    "jody grace":      ("73728c19-e412-5006-aa48-1d2685a77f7e", "Judy Grace",         "MATCHED"),
    "jenny davidson":  ("3f48caf5-14c9-59f1-8255-1db7ef2fd049", "Jenny Davison",      "MATCHED"),
    "ken shults":      ("2a6a7c9e-1d8a-4f9a-a8f5-6f3a3c1e9b0f", "Kenneth Shults",    "MATCHED"),
    "kenny shults":    ("2a6a7c9e-1d8a-4f9a-a8f5-6f3a3c1e9b0f", "Kenneth Shults",    "MATCHED"),
    "karin atogpian":  ("86de3135-51c7-5ad5-9144-112f984acd3f", "Karen Atgopian",     "NEW_PLAYER"),
    "lori jean tarr":  ("77f0f32d-98bf-5742-8228-4e06ac07bd9d", "Lori Jean Conover",  "MATCHED"),
    "tim fitzgerald":  ("b54020bc-1a1a-5d23-89e1-34617b3514fa", "Jim Fitzgerald",     "MATCHED"),
    "tobin wigger":    ("f38e2f8e-cba6-5fff-8b20-90791fd0d794", "Torbin Wigger",       "MATCHED"),
    "torben wigger":   ("f38e2f8e-cba6-5fff-8b20-90791fd0d794", "Torbin Wigger",       "MATCHED"),
    "steve fennell":   ("fb605451-fa41-5e6d-ad21-1b541b3fc82b", "Steve Femmel",        "MATCHED"),
    "ted martens":     ("3d1b24de-41bd-56eb-86b8-8afb919a63b5", "Ted Martin",          "MATCHED"),
}

# (keep_id, absorb_id)
EVENT_MERGES = [
    ("b4b6a194e2", "dea8d6f019"),  # 1986 WFA_WORLD ← WORLD
    ("8667de3590", "8a1c282011"),  # 1987 WFA_WORLD ← WORLD
    ("d272be24f7", "1398fe01e4"),  # 1988 WFA_WORLD ← WORLD
    ("334c632f6e", "6ce90a8f5d"),  # 1989 WFA_WORLD ← WORLD
    ("b9d35c4646", "9bced09b82"),  # 1994 IFAB_WORLD ← WORLD
]

VERSION = "PRE1997_V2"
TODAY = date.today().isoformat()


# ── helpers ────────────────────────────────────────────────────────────────

def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def write_csv(path, rows, fieldnames=None):
    if not rows:
        return
    if fieldnames is None:
        fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print(f"  Wrote {len(rows):4d} rows → {path.relative_to(ROOT.parent)}")

def load_pt():
    """Load Persons_Truth → dict effective_person_id → row."""
    rows = read_csv(PT_FILE)
    return {r["effective_person_id"]: r for r in rows}


# ── PART 1 — person aliases ────────────────────────────────────────────────

def apply_person_decisions(persons, aliases, participants, pt):
    """
    Returns updated (persons, aliases, participants) plus stats dict.
    """
    stats = {
        "aliases_added": [],
        "participants_resolved": 0,
        "persons_added": [],
    }

    # PT persons not yet in persons_pre1997 that we need to add
    existing_person_ids = {p["person_id"] for p in persons}

    for raw_lower, (pid, canon, status) in ALIAS_ACCEPTS.items():
        # 1. Update existing alias entry (they exist as REVIEW_NEEDED stubs)
        updated = False
        for alias_row in aliases:
            if alias_row["raw_name"].lower() == raw_lower:
                alias_row["person_id"] = pid
                alias_row["alias_status"] = "ACCEPTED"
                updated = True
                break
        if not updated:
            # Add new row if not present
            aliases.append({
                "raw_name": raw_lower,
                "person_id": pid,
                "alias_status": "ACCEPTED",
            })
        stats["aliases_added"].append(raw_lower)

        # 2. Add person to persons_pre1997 if not present
        if pid not in existing_person_ids:
            # Pull from PT for POST1997 scope persons
            if status == "MATCHED" and pid in pt:
                pt_row = pt[pid]
                persons.append({
                    "person_id": pid,
                    "person_canon": pt_row.get("person_canon") or canon,
                    "source_scope": "POST1997",
                    "person_status": "ACTIVE",
                    "aliases": raw_lower,
                    "notes": f"Added by {VERSION} alias accept",
                })
            elif status == "NEW_PLAYER":
                # Should already exist (Karin→Karen, Karen is PRE1997_ONLY)
                pass
            existing_person_ids.add(pid)
            stats["persons_added"].append(canon)

        # 3. Update participants: REVIEW_NEEDED rows with this raw name
        for row in participants:
            if (row["player_name_raw"].lower() == raw_lower
                    and row["resolution_status"] == "REVIEW_NEEDED"):
                row["person_id"] = pid
                row["person_canon"] = canon
                row["resolution_status"] = status
                stats["participants_resolved"] += 1

    return persons, aliases, participants, stats


# ── PART 2 — event merges ──────────────────────────────────────────────────

def apply_event_merges(events, id_mapping, results, participants, disciplines,
                       canonical_events, event_groups, source_comparison):
    """
    Returns all updated tables plus merge_log list.
    """
    merge_log = []
    # Build remap: absorb_id → keep_id
    remap = {absorb: keep for keep, absorb in EVENT_MERGES}

    # Index events by canonical_event_id
    event_idx = {e["canonical_event_id"]: e for e in events}

    for keep_id, absorb_id in EVENT_MERGES:
        if keep_id not in event_idx or absorb_id not in event_idx:
            print(f"  WARN: merge {keep_id}←{absorb_id} — one or both IDs missing, skipping")
            continue

        keeper = event_idx[keep_id]
        absorbed = event_idx[absorb_id]

        # Merge source_types
        keep_sources = set(keeper["source_types"].split("|"))
        absorb_sources = set(absorbed["source_types"].split("|"))
        merged_sources = sorted(keep_sources | absorb_sources)
        keeper["source_types"] = "|".join(merged_sources)
        keeper["num_sources"] = str(len(merged_sources))
        keeper["validation_status"] = "CONFIRMED_MULTI_SOURCE"

        # Merge location (take first non-empty)
        if not keeper.get("location") and absorbed.get("location"):
            keeper["location"] = absorbed["location"]

        # Recount placements after we remap results rows
        merge_log.append({
            "keep_id": keep_id,
            "absorb_id": absorb_id,
            "keep_name": keeper["event_name"],
            "absorb_name": absorbed["event_name"],
            "keep_type": keeper["normalized_event_type"],
            "absorb_type": absorbed["normalized_event_type"],
            "year": keeper["year"],
        })

    # Remove absorbed events from events list
    events = [e for e in events if e["canonical_event_id"] not in remap]

    # Remap all tables
    for row in id_mapping:
        if row["canonical_event_id"] in remap:
            row["canonical_event_id"] = remap[row["canonical_event_id"]]

    for row in results:
        if row["canonical_event_id"] in remap:
            row["canonical_event_id"] = remap[row["canonical_event_id"]]

    for row in participants:
        if row["canonical_event_id"] in remap:
            row["canonical_event_id"] = remap[row["canonical_event_id"]]

    for row in disciplines:
        if row["canonical_event_id"] in remap:
            row["canonical_event_id"] = remap[row["canonical_event_id"]]

    # Deduplicate disciplines (same canonical_event_id + division_raw)
    seen_disc = set()
    deduped_disc = []
    for row in disciplines:
        key = (row["canonical_event_id"], row["division_raw"])
        if key not in seen_disc:
            seen_disc.add(key)
            deduped_disc.append(row)
    disciplines = deduped_disc

    # Remap canonical_events (uses canonical_event_id field)
    canonical_events = [e for e in canonical_events
                        if e["canonical_event_id"] not in remap]
    # Update num_placements on surviving canonical_events
    # (canonical_events has different fields from events_pre1997 — update only shared ones)
    from collections import Counter as _Counter
    plc_counts2 = _Counter(r["canonical_event_id"] for r in results)
    for e in canonical_events:
        e["num_placements"] = str(plc_counts2.get(e["canonical_event_id"], 0))
    # Sync validation_status from updated events
    updated_events_idx = {ev["canonical_event_id"]: ev for ev in events}
    for e in canonical_events:
        eid = e["canonical_event_id"]
        if eid in updated_events_idx:
            e["validation_status"] = updated_events_idx[eid]["validation_status"]
            e["num_sources"] = updated_events_idx[eid]["num_sources"]
            e["source_types"] = updated_events_idx[eid]["source_types"]

    # Remap event_groups (uses group_id field, not canonical_event_id)
    for row in event_groups:
        if row.get("group_id") in remap:
            row["group_id"] = remap[row["group_id"]]

    # Remap source_comparison (uses group_id field)
    source_comparison = [r for r in source_comparison
                         if r["group_id"] not in remap]

    # Update num_placements in events
    from collections import Counter
    plc_counts = Counter(r["canonical_event_id"] for r in results)
    for e in events:
        e["num_placements"] = str(plc_counts.get(e["canonical_event_id"], 0))
    for e in canonical_events:
        e["num_placements"] = str(plc_counts.get(e["canonical_event_id"], 0))

    return (events, id_mapping, results, participants, disciplines,
            canonical_events, event_groups, source_comparison, merge_log)


# ── PART 3 — rebuild spreadsheet ──────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
ORANGE_FILL = PatternFill("solid", fgColor="FFCC99")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")

STATUS_COLORS = {
    "MATCHED": GREEN_FILL,
    "AUTOACCEPTED": GREEN_FILL,
    "ACCEPTED": GREEN_FILL,
    "NEW_PLAYER": YELLOW_FILL,
    "REVIEW_NEEDED": ORANGE_FILL,
    "UNRESOLVED": ORANGE_FILL,
    "NOISE": GRAY_FILL,
}
VALIDATION_COLORS = {
    "CONFIRMED_MULTI_SOURCE": GREEN_FILL,
    "SINGLE_SOURCE": YELLOW_FILL,
    "CONFLICT": RED_FILL,
}


def _header(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autowidth(ws, min_w=8, max_w=40):
    for col in ws.columns:
        w = min_w
        for cell in col:
            if cell.value:
                w = max(w, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = w


def build_readme(wb, version, today):
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    data = [
        ("Pre-1997 Footbag Historical Results", None),
        (f"Dataset version: {version}", None),
        (f"Generated: {today}", None),
        ("", None),
        ("CONTENTS", None),
        ("Sheet", "Description"),
        ("README", "This page"),
        ("DATA NOTES", "Coverage gaps and data quality notes"),
        ("EVENT INDEX", "All 37 canonical event groups"),
        ("PLAYER SUMMARY", "All persons appearing in pre-1997 results"),
        ("1980 – 1996", "17 year sheets with full placement data"),
        ("VALIDATION SUMMARY", "Cross-source validation statistics"),
    ]
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50
    for i, (a, b) in enumerate(data, 1):
        wa = ws.cell(row=i, column=1, value=a)
        if b is not None:
            ws.cell(row=i, column=2, value=b)
        if i == 1:
            wa.font = Font(bold=True, size=14)
        elif a in ("CONTENTS", "Sheet"):
            wa.font = Font(bold=True)


def build_data_notes(wb):
    ws = wb.create_sheet("DATA NOTES")
    ws.sheet_view.showGridLines = False
    notes = [
        "Coverage gaps and data quality notes",
        "",
        "SOURCES",
        "FBW Magazine Vols 2-14 (scans, AI extraction via Gemini)",
        "IFAB Worlds History page (AI extraction)",
        "OLD_RESULTS.txt (contributed text file)",
        "",
        "KNOWN GAPS",
        "1980-1981: World Footbag Championships has only 1-2 placements (IFAB history page limited)",
        "1982-1985: Regional/state/club events not in scanned sources",
        "1986-1992: Non-WFA events not covered (only WFA World Championships)",
        "1988: US Nationals has only 1 placement (scan quality or page missing)",
        "1984: European Championships has only 1 placement (same)",
        "",
        "IDENTITY NOTES",
        "MATCHED: Name matched exactly to post-1997 Persons Truth database",
        "ACCEPTED: Alias accepted by human reviewer (spelling variant or nickname)",
        "AUTOACCEPTED: Safe alias accepted automatically (obvious 1-char error or nickname)",
        "NEW_PLAYER: Pre-1997 only player — not in post-1997 records",
        "REVIEW_NEEDED: Unresolved — requires further verification",
        "UNRESOLVED: No plausible match found",
        "",
        "EVENT NOTES",
        "WFA World Championships (1986-1989): confirmed same event as 'World Footbag Championships'",
        "1994 IFAB World Championships: confirmed same event as '1994 World Footbag Championships'",
        "Event groups with CONFIRMED_MULTI_SOURCE have results verified from 2+ independent sources",
    ]
    ws.column_dimensions["A"].width = 80
    for i, note in enumerate(notes, 1):
        cell = ws.cell(row=i, column=1, value=note)
        if note and not note.startswith(" ") and note.isupper():
            cell.font = Font(bold=True)


def build_event_index(wb, events):
    ws = wb.create_sheet("EVENT INDEX")
    ws.freeze_panes = "A2"
    cols = ["ID", "Year", "Event Name", "Location", "Type",
            "Sources", "Validation", "Placements"]
    _header(ws, 1, cols)
    ws.row_dimensions[1].height = 20
    for i, e in enumerate(sorted(events, key=lambda x: (x["year"], x["event_name"])), 2):
        row_data = [
            e["canonical_event_id"],
            int(e["year"]),
            e["event_name"],
            e.get("location", ""),
            e["normalized_event_type"],
            e["source_types"],
            e["validation_status"],
            int(e["num_placements"]),
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=c, value=val)
            if c == 7:  # validation
                cell.fill = VALIDATION_COLORS.get(val, PatternFill())
            if i % 2 == 0 and not cell.fill.fgColor.rgb.endswith("000000"):
                pass  # keep status color
    _autowidth(ws)


def build_player_summary(wb, persons, participants):
    ws = wb.create_sheet("PLAYER SUMMARY")
    ws.freeze_panes = "A2"

    # Count appearances per person
    from collections import Counter, defaultdict
    app_count = Counter()
    year_set = defaultdict(set)
    div_set = defaultdict(set)
    for p in participants:
        pid = p["person_id"]
        if pid:
            app_count[pid] += 1
            year_set[pid].add(p["canonical_event_id"][:4] if len(p["canonical_event_id"]) >= 4 else "?")
            div_set[pid].add(p["division_raw"][:30] if p["division_raw"] else "")

    cols = ["Person ID", "Name", "Scope", "Status", "Appearances",
            "Earliest Year", "Sample Division"]
    _header(ws, 1, cols)
    ws.row_dimensions[1].height = 20

    sorted_persons = sorted(persons, key=lambda p: p["person_canon"])
    for i, p in enumerate(sorted_persons, 2):
        pid = p["person_id"]
        years = sorted(year_set[pid])
        row_data = [
            pid,
            p["person_canon"],
            p["source_scope"],
            p.get("person_status", ""),
            app_count.get(pid, 0),
            years[0] if years else "",
            list(div_set[pid])[0] if div_set[pid] else "",
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=c, value=val)
            if i % 2 == 0:
                cell.fill = ALT_FILL
    _autowidth(ws)


def build_year_sheet(wb, year, events, results, participants):
    ws = wb.create_sheet(str(year))
    ws.sheet_view.showGridLines = False

    year_events = [e for e in events if e["year"] == str(year)]
    if not year_events:
        ws.cell(1, 1, f"No events for {year}")
        return

    # Index results and participants by result_id
    result_map = {r["result_id"]: r for r in results}
    parts_by_result = {}
    for p in participants:
        parts_by_result.setdefault(p["result_id"], []).append(p)

    current_row = 1
    for event in sorted(year_events, key=lambda e: e["event_name"]):
        eid = event["canonical_event_id"]

        # Event header
        cell = ws.cell(current_row, 1, event["event_name"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=6
        )
        current_row += 1

        # Event metadata
        meta = f"Location: {event.get('location','unknown')} | Sources: {event['source_types']} | {event['validation_status']}"
        cell = ws.cell(current_row, 1, meta)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(color="FFFFFF", italic=True)
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=6
        )
        current_row += 1

        # Group results by division
        event_results = [r for r in results if r["canonical_event_id"] == eid]
        divs = {}
        for r in event_results:
            divs.setdefault(r["division_raw"], []).append(r)

        for div_name, div_results in sorted(divs.items()):
            # Division header
            cell = ws.cell(current_row, 1, div_name)
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=6
            )
            current_row += 1

            # Column headers
            col_heads = ["Place", "Player(s)", "Person Canon", "Status", "Source", "Person ID"]
            for c, h in enumerate(col_heads, 1):
                cell = ws.cell(current_row, c, h)
                cell.fill = GRAY_FILL
                cell.font = Font(bold=True, size=9)
            current_row += 1

            for res in sorted(div_results, key=lambda x: int(x["place"]) if x["place"].strip().isdigit() else 999):
                rid = res["result_id"]
                pparts = parts_by_result.get(rid, [])

                players_raw = " / ".join(p["player_name_raw"] for p in pparts) if pparts else res.get("player_raw", "")
                canon_names = " / ".join(p["person_canon"] for p in pparts if p["person_canon"]) if pparts else ""
                statuses = " / ".join(p["resolution_status"] for p in pparts) if pparts else "?"
                primary_status = pparts[0]["resolution_status"] if pparts else "UNRESOLVED"
                source = res.get("source_type", "")
                pid_display = " / ".join(p["person_id"] for p in pparts if p["person_id"]) if pparts else ""

                place_val = int(res["place"]) if res["place"].strip().isdigit() else res["place"]
                row_vals = [place_val, players_raw, canon_names,
                            statuses, source, pid_display]
                fill = STATUS_COLORS.get(primary_status, PatternFill())
                for c, val in enumerate(row_vals, 1):
                    cell = ws.cell(current_row, c, val)
                    cell.fill = fill
                    cell.font = Font(size=9)
                current_row += 1

        current_row += 1  # blank row between events

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 38


def build_validation_summary(wb, events, results, participants):
    ws = wb.create_sheet("VALIDATION SUMMARY")
    ws.sheet_view.showGridLines = False

    from collections import Counter
    val_counts = Counter(e["validation_status"] for e in events)
    res_counts = Counter(p["resolution_status"] for p in participants)

    ws.cell(1, 1, "Cross-Source Validation").font = Font(bold=True, size=12)
    row = 3
    for status, count in sorted(val_counts.items()):
        ws.cell(row, 1, status)
        ws.cell(row, 2, count)
        ws.cell(row, 1).fill = VALIDATION_COLORS.get(status, PatternFill())
        row += 1

    row += 1
    ws.cell(row, 1, "Identity Resolution").font = Font(bold=True, size=12)
    row += 2
    for status, count in sorted(res_counts.items()):
        ws.cell(row, 1, status)
        ws.cell(row, 2, count)
        ws.cell(row, 1).fill = STATUS_COLORS.get(status, PatternFill())
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12


def build_workbook(events, results, participants, persons, version, today):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_readme(wb, version, today)
    build_data_notes(wb)
    build_event_index(wb, events)
    build_player_summary(wb, persons, participants)
    for year in range(1980, 1997):
        build_year_sheet(wb, year, events, results, participants)
    build_validation_summary(wb, events, results, participants)

    out_path = OUT / f"footbag_results_pre1997_recovery_{version}.xlsx"
    wb.save(out_path)
    print(f"  Saved workbook → {out_path.relative_to(ROOT.parent)}")
    return out_path


# ── main ───────────────────────────────────────────────────────────────────

def main():
    print(f"=== 09_apply_decisions.py → {VERSION} ===\n")

    # Load all canonical inputs
    events        = read_csv(CANONICAL / "events_pre1997.csv")
    id_mapping    = read_csv(CANONICAL / "event_id_mapping.csv")
    results       = read_csv(CANONICAL / "event_results_pre1997.csv")
    participants  = read_csv(CANONICAL / "event_result_participants_pre1997.csv")
    disciplines   = read_csv(CANONICAL / "event_disciplines_pre1997.csv")
    persons       = read_csv(CANONICAL / "persons_pre1997.csv")
    aliases       = read_csv(CANONICAL / "person_aliases_pre1997.csv")
    can_events    = read_csv(CANONICAL / "canonical_events.csv")
    event_groups  = read_csv(CANONICAL / "event_groups.csv")
    src_compare   = read_csv(CANONICAL / "event_source_comparison.csv")
    pt            = load_pt()

    print(f"Loaded (V1 baseline):")
    print(f"  Events:       {len(events)}")
    print(f"  Results:      {len(results)}")
    print(f"  Participants: {len(participants)}")
    print(f"  Persons:      {len(persons)}")
    print(f"  Aliases:      {len(aliases)}")
    print(f"  Disciplines:  {len(disciplines)}")

    # Baseline REVIEW_NEEDED
    rn_before = sum(1 for p in participants if p["resolution_status"] == "REVIEW_NEEDED")

    # ── Part 1: person decisions ──
    print("\n--- PART 1: Apply person alias decisions ---")
    persons, aliases, participants, person_stats = apply_person_decisions(
        persons, aliases, participants, pt
    )
    print(f"  Aliases added:          {len(person_stats['aliases_added'])}")
    for a in person_stats["aliases_added"]:
        print(f"    + {a}")
    print(f"  Persons added to table: {len(person_stats['persons_added'])}")
    for p in person_stats["persons_added"]:
        print(f"    + {p}")
    print(f"  Participants resolved:  {person_stats['participants_resolved']}")

    # ── Part 2: event merges ──
    print("\n--- PART 2: Apply event merge decisions ---")
    (events, id_mapping, results, participants, disciplines,
     can_events, event_groups, src_compare, merge_log) = apply_event_merges(
        events, id_mapping, results, participants, disciplines,
        can_events, event_groups, src_compare
    )
    print(f"  Merges applied: {len(merge_log)}")
    for m in merge_log:
        print(f"    {m['year']}: [{m['absorb_type']}] \"{m['absorb_name']}\" → [{m['keep_type']}] \"{m['keep_name']}\"")

    # ── Write V2 canonical CSVs ──
    print("\n--- PART 3: Write PRE1997_V2 canonical CSVs ---")

    write_csv(CANONICAL / "events_pre1997.csv",       events)
    write_csv(CANONICAL / "event_id_mapping.csv",     id_mapping)
    write_csv(CANONICAL / "event_results_pre1997.csv", results)
    write_csv(CANONICAL / "event_result_participants_pre1997.csv", participants)
    write_csv(CANONICAL / "event_disciplines_pre1997.csv", disciplines)
    write_csv(CANONICAL / "persons_pre1997.csv",      persons)
    write_csv(CANONICAL / "person_aliases_pre1997.csv", aliases)
    write_csv(CANONICAL / "canonical_events.csv",     can_events)
    write_csv(CANONICAL / "event_groups.csv",         event_groups)
    write_csv(CANONICAL / "event_source_comparison.csv", src_compare)

    # ── Build workbook ──
    print("\n--- PART 4: Build PRE1997_V2 spreadsheet ---")
    build_workbook(events, results, participants, persons, VERSION, TODAY)

    # ── Final summary ──
    rn_after = sum(1 for p in participants if p["resolution_status"] == "REVIEW_NEEDED")
    resolved_now = sum(1 for p in participants
                       if p["resolution_status"] in ("MATCHED", "AUTOACCEPTED", "ACCEPTED", "NEW_PLAYER"))

    print(f"\n{'='*56}")
    print(f"PRE1997_V2 SUMMARY")
    print(f"{'='*56}")
    print(f"\nPerson decisions:")
    print(f"  Aliases accepted:        12")
    print(f"  REVIEW_NEEDED before:    {rn_before}")
    print(f"  REVIEW_NEEDED after:     {rn_after}")
    print(f"  Participants resolved:   {person_stats['participants_resolved']}")
    print(f"  New persons in table:    {len(person_stats['persons_added'])}")

    print(f"\nEvent merges:")
    for m in merge_log:
        print(f"  {m['year']}: {m['absorb_name']} absorbed into {m['keep_name']}")

    print(f"\nFinal counts:")
    print(f"  Events:       {len(events)}")
    print(f"  Results:      {len(results)}")
    print(f"  Participants: {len(participants)}")
    print(f"    MATCHED/ACCEPTED:  {resolved_now}")
    print(f"    REVIEW_NEEDED:     {rn_after}")
    print(f"    UNRESOLVED:        {sum(1 for p in participants if p['resolution_status']=='UNRESOLVED')}")
    print(f"    NOISE:             {sum(1 for p in participants if p['resolution_status']=='NOISE')}")
    print(f"  Persons:      {len(persons)}")
    print(f"  Aliases:      {len(aliases)}")
    print(f"  Disciplines:  {len(disciplines)}")
    print(f"\nValidation:")
    from collections import Counter
    for status, cnt in sorted(Counter(e["validation_status"] for e in events).items()):
        print(f"  {status}: {cnt}")
    print(f"\nDone.")


if __name__ == "__main__":
    main()
